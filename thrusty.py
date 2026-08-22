"""
Thrusty — Python/tkinter port of Forden's MATLAB GUIDE application.

Layout mirrors the original MATLAB GUIDE application:
  Left panel  : booster type, units, launch site (decimal °), target (decimal °),
                cutoff time, run buttons, range/apogee results
  Right panel : 4-up matplotlib plots (altitude, speed, trajectory, ground track)
  Bottom bar  : status line
"""

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import tkinter.font as tkfont
import copy
import threading
import numpy as np
import sys
import os
import json
import types
from pathlib import Path

sys.path.insert(0, os.path.dirname(__file__))

import matplotlib
try:
    matplotlib.use("TkAgg")
except ImportError:
    # Headless import (test collection with pyplot already on a non-GUI
    # backend): keep whatever backend is active.  The app itself needs a
    # display and will fail later at Tk() init, which callers handle.
    pass
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure

import matplotlib.ticker

import booster_models as mm
from booster_models import (BOOSTER_DB, get_booster,
                           booster_to_dict, booster_from_dict,
                           extract_flight_plan, apply_flight_plan, save_flight_plan,
                           total_burn_time, tumbling_cylinder_beta,
                           NOSE_SHAPES, NOSE_SHAPE_LABELS,
                           GRAIN_LABELS, grain_fill_factor, _GRAIN_FILL_RANGE,
                           ROParams, ro_from_dict, ro_to_dict, effective_ro,
                           extract_reentry_plan, apply_reentry_plan,
                           load_reentry_plan, save_reentry_plan)
from trajectory import (integrate_trajectory, maximize_range, aim_booster,
                        plan_orbital_insertion, MaxRangeCancelled,
                        wheelon_burnout_angle)
from coordinates import range_between
from slv_performance import schilling_performance
import mass_estimator as mest
import heating
import thresholds

# ---------------------------------------------------------------------------
# Country border map data (Natural Earth 110m, bundled GeoJSON)
# ---------------------------------------------------------------------------

_BORDERS_CACHE = None   # loaded once on first draw, then reused

def _open_file(path: str) -> None:
    """Open a file with the default viewer, cross-platform."""
    import subprocess
    try:
        if sys.platform == 'darwin':
            subprocess.Popen(['open', path])
        elif sys.platform == 'win32':
            os.startfile(path)
        else:
            subprocess.Popen(['xdg-open', path])
    except Exception:
        pass


def _load_borders():
    """Return the bundled Natural Earth country GeoJSON (lazy, cached).
    Prefers the 50m tier (finer coastlines on every map); falls back to
    the 110m file if that's all a checkout carries."""
    global _BORDERS_CACHE
    if _BORDERS_CACHE is None:
        d = Path(__file__).parent / "data"
        for name in ("ne_50m_countries.geojson",
                     "ne_110m_countries.geojson"):
            p = d / name
            if p.exists():
                _BORDERS_CACHE = json.loads(p.read_text())
                break
    return _BORDERS_CACHE


def _draw_borders(ax, center_lon):
    """Overlay country borders on *ax*, re-centred on *center_lon* degrees."""
    data = _load_borders()
    if data is None:
        return

    def _plot_ring(ring):
        # Shift every vertex into [−180, +180] relative to center_lon
        xs = [((pt[0] - center_lon + 180.0) % 360.0) - 180.0 for pt in ring]
        ys = [pt[1] for pt in ring]
        # Split the ring wherever it still crosses ±180° in centred space
        seg_x, seg_y = [[]], [[]]
        for i in range(len(xs)):
            if i > 0 and abs(xs[i] - xs[i - 1]) > 180:
                seg_x.append([])
                seg_y.append([])
            seg_x[-1].append(xs[i])
            seg_y[-1].append(ys[i])
        for sx, sy in zip(seg_x, seg_y):
            if len(sx) > 1:
                ax.plot(sx, sy, color='#777777', linewidth=0.4,
                        solid_capstyle='round', zorder=1)

    for feature in data.get('features', []):
        geom  = feature.get('geometry') or {}
        gtype = geom.get('type', '')
        coords = geom.get('coordinates', [])
        if gtype == 'Polygon':
            for ring in coords:
                _plot_ring(ring)
        elif gtype == 'MultiPolygon':
            for polygon in coords:
                for ring in polygon:
                    _plot_ring(ring)


# ---------------------------------------------------------------------------
# Custom booster persistence
# ---------------------------------------------------------------------------

# Sentinel string inserted into the booster combobox between non-Forden and
# Forden entries.  It is never a valid booster name.

# Blunted-cone hypersonic Cd — physics lives in booster_models (headless-
# testable); the "Estimate Object β" dialog is presentation over it.
from booster_models import (cd_blunted_cone_newtonian as _cd_blunted_cone_newtonian,
                            cd_cone_hypersonic as _cd_cone_hypersonic,
                            cd_biconic_hypersonic as _cd_biconic_hypersonic,
                            biconic_angles as _biconic_angles,
                            wing_geometry as _wing_geometry,
                            lifting_body_sweep as _lifting_body_sweep,
                            CONE_CF_TURBULENT as _CONE_CF)


# Names that ship with the program and cannot be deleted
_PACKAGED_NAMES: set[str] = set(BOOSTER_DB.keys())
# Packaged boosters the user has overridden with custom edits
_OVERRIDDEN_PACKAGED: set[str] = set()
# Where user-created boosters are saved.  Writers emit the new custom_boosters
# path; the loader still reads the legacy custom_missiles file (see below).
_CUSTOM_PATH        = Path.home() / ".gui_missile_flyout" / "custom_boosters.json"
_CUSTOM_PATH_LEGACY = Path.home() / ".gui_missile_flyout" / "custom_missiles.json"
_TRAJ_PATH        = Path.home() / ".gui_missile_flyout" / "trajectory_profiles.json"
# Which named flight plan is active per booster (booster name -> plan name).
_ACTIVE_PLANS_PATH = Path.home() / ".gui_missile_flyout" / "active_flight_plans.json"
_ACTIVE_REENTRY_PLANS_PATH = Path.home() / ".gui_missile_flyout" / "active_reentry_plans.json"
# ── Export folder layout (visible under ~/Documents for Finder access) ───
_THRUSTY_ROOT     = Path.home() / "Documents" / "Thrusty"
_DIR_BOOSTERS     = _THRUSTY_ROOT / "boosters"
_DIR_BOOSTERS_LEGACY = _THRUSTY_ROOT / "boosters"   # old export dir, still browsable


def _boosters_dir():
    """Default folder for booster file dialogs.  Prefer the new boosters/ dir,
    but fall back to the legacy boosters/ dir if that's where the user's files
    already are (boosters/ not yet created)."""
    if not _DIR_BOOSTERS.exists() and _DIR_BOOSTERS_LEGACY.exists():
        return _DIR_BOOSTERS_LEGACY
    return _ensure_dir(_DIR_BOOSTERS)
_RO_LIBRARY_PATH  = _THRUSTY_ROOT / "ro_library"
# Writable user flight-plan library.  Booster hardware is stored without a
# flight plan; when the user saves a booster we drop a timing-only flight plan
# here so subsystem-deployment timing (shroud/strap-on/grid-fin) survives a
# reload.  Registered with booster_models so get_booster merges it over the
# bundled plan.  (Separate from the bundled flight_plans/ next to the module.)
_FLIGHT_PLAN_LIBRARY_PATH = _THRUSTY_ROOT / "flight_plans"
mm.USER_FLIGHT_PLAN_DIRS = [str(_FLIGHT_PLAN_LIBRARY_PATH)]
# Reentry objects are the down-leg equivalent: hardware-only files plus a
# separate reentry plan.  User plans live here and override the bundled ones.
_REENTRY_PLAN_LIBRARY_PATH = _THRUSTY_ROOT / "reentry_plans"
mm.USER_REENTRY_PLAN_DIRS = [str(_REENTRY_PLAN_LIBRARY_PATH)]
# Canonical RVs that ship with the code, next to this file (e.g. SWERVE, AHW).
# These are always available; the writable user library above overrides them.
_BUNDLED_RO_LIBRARY_PATH = Path(__file__).resolve().parent / "ro_library"
# Back-compat: reentry objects used to live in rv_library/*.ro.json.  We still
# read those (old locally-saved files) but only ever write the new .ro.json form.
_LEGACY_RO_LIBRARY_PATH = _THRUSTY_ROOT / "rv_library"
_DIR_GUIDANCE     = _THRUSTY_ROOT / "guidance"
_DIR_SITES        = _THRUSTY_ROOT / "sites"
_DIR_TRAJECTORIES = _THRUSTY_ROOT / "trajectories"
_DIR_EVENTS       = _THRUSTY_ROOT / "events"
_DIR_MAPS         = _THRUSTY_ROOT / "maps"
_DIR_PLOTS        = _THRUSTY_ROOT / "plots"
_DIR_SCENARIOS    = _THRUSTY_ROOT / "scenarios"
_DIR_BLENDER      = _THRUSTY_ROOT / "blender_models"


def _safe_name(s: str, maxlen: int = 40) -> str:
    """Sanitize a string for use as a filename component.
    Collapses whitespace, strips characters that confuse file pickers
    or extension parsing, and truncates to a readable length."""
    import re as _re
    s = _re.sub(r'\s+', '_', (s or '').strip())
    s = _re.sub(r'[^\w\-]', '-', s)
    return s[:maxlen] or "untitled"


def _ensure_dir(d: Path) -> Path:
    """Create directory if missing; return the path."""
    d.mkdir(parents=True, exist_ok=True)
    return d


def _parse_deploy_schedule(text: str):
    """Parse a grid-fin deploy schedule 't:n, t:n' into [[t_s, n], ...].

    Each entry is 'deploy_time_seconds : number_of_fins'.  Whitespace is
    ignored; an empty string means all fins are deployed from t=0 (-> []).
    Raises ValueError on malformed input."""
    text = (text or "").strip()
    if not text:
        return []
    out = []
    for chunk in text.replace(";", ",").split(","):
        chunk = chunk.strip()
        if not chunk:
            continue
        if ":" not in chunk:
            raise ValueError(f"deploy entry '{chunk}' must be 'time:count'")
        t_s, n_s = chunk.split(":", 1)
        out.append([float(t_s.strip()), int(float(n_s.strip()))])
    return out


def _format_deploy_schedule(sched) -> str:
    """Format [[t_s, n], ...] back to 't:n, t:n' for the entry field."""
    parts = []
    for entry in (sched or []):
        try:
            t_s, n = float(entry[0]), int(entry[1])
        except (TypeError, ValueError, IndexError):
            continue
        t_str = f"{t_s:g}"
        parts.append(f"{t_str}:{n}")
    return ", ".join(parts)



def _load_traj_profiles() -> dict:
    """Return saved trajectory profiles keyed by booster name."""
    if not _TRAJ_PATH.exists():
        return {}
    try:
        return json.loads(_TRAJ_PATH.read_text())
    except Exception:
        return {}


def _save_traj_profiles(profiles: dict) -> None:
    _TRAJ_PATH.parent.mkdir(parents=True, exist_ok=True)
    _TRAJ_PATH.write_text(json.dumps(profiles, indent=2))


def _load_custom_boosters():
    """Read the saved custom boosters and register them in BOOSTER_DB.

    Reads custom_boosters.json (current) or falls back to the legacy
    custom_missiles.json, so boosters saved before the rename still load."""
    path = (_CUSTOM_PATH if _CUSTOM_PATH.exists()
            else _CUSTOM_PATH_LEGACY if _CUSTOM_PATH_LEGACY.exists() else None)
    if path is None:
        return
    try:
        data = json.loads(path.read_text())
        for name, d in data.items():
            p = booster_from_dict(d)
            BOOSTER_DB[name] = lambda _p=p: _p
            if name in _PACKAGED_NAMES:
                _OVERRIDDEN_PACKAGED.add(name)
    except Exception as exc:
        print(f"Warning: could not load custom boosters: {exc}")


def _save_custom_boosters():
    """Write all non-packaged and overridden-packaged boosters to custom_boosters.json.

    Boosters are stored hardware-only -- guidance/flight-plan fields are not
    embedded.  The flight plan is persisted separately (traj_profiles.json in
    the GUI, .flightplan.json on export), so it is never lost."""
    _CUSTOM_PATH.parent.mkdir(parents=True, exist_ok=True)
    data = {}
    for name in BOOSTER_DB:
        if name not in _PACKAGED_NAMES or name in _OVERRIDDEN_PACKAGED:
            data[name] = booster_to_dict(BOOSTER_DB[name](), include_flight_plan=False)
    _CUSTOM_PATH.write_text(json.dumps(data, indent=2))


# ---------------------------------------------------------------------------
# RV library — RVs are first-class objects, decoupled from any specific
# booster.  Each .ro.json file in ro_library/ becomes an entry in RO_DB,
# keyed by RV name.  The main control panel exposes the library via a
# combobox; at run time the selected RV is injected into the booster.
# ---------------------------------------------------------------------------

RO_DB: dict = {}   # name -> callable returning a fresh ROParams


def _load_ro_library():
    """Register every .ro.json into RO_DB.

    RVs are loaded from two places: the **bundled** ro_library/ shipped next to
    thrusty.py (the canonical RVs that travel with the code — SWERVE, AHW, …),
    and the user's writable ~/Documents/Thrusty/ro_library (saved or edited
    RVs).  Bundled RVs load first; a user RV of the same name overrides the
    bundled one, so a fresh checkout always exposes the shipped RVs while the
    user's own edits still win."""
    _ensure_dir(_RO_LIBRARY_PATH)
    RO_DB.clear()
    dirs, seen = [], set()
    # Legacy dir last so a new-form file of the same name wins on override.
    for d in (_BUNDLED_RO_LIBRARY_PATH, _LEGACY_RO_LIBRARY_PATH, _RO_LIBRARY_PATH):
        rp = str(d.resolve())
        if rp not in seen:          # skip if bundled == user (e.g. run in place)
            seen.add(rp); dirs.append(d)
    for d in dirs:
        if not d.exists():
            continue
        # Accept both the new .ro.json and legacy .ro.json extensions.
        files = sorted(list(d.glob("*.ro.json")) + list(d.glob("*.ro.json")))
        for fp in files:
            try:
                ro = ro_from_dict(json.loads(fp.read_text()))
                # Reentry-object files are hardware-only; merge the reentry plan
                # (glide mode, turns, dives, separation) on top so RO_DB holds
                # ready-to-fly objects, exactly as get_booster does for boosters.
                _rp = load_reentry_plan(ro.name, extra_dirs=mm.USER_REENTRY_PLAN_DIRS)
                if _rp is not None:
                    ro = apply_reentry_plan(ro, _rp)
                key = ro.name or fp.stem.replace(".ro", "").replace(".ro", "")
                RO_DB[key] = lambda _r=ro: _r
            except Exception as exc:
                print(f"Warning: could not load Reentry object '{fp.name}': {exc}")


def _ro_plan_sibling(ro_path) -> Path:
    """Companion .reentryplan.json path paired to a .ro.json file by stem."""
    p = Path(ro_path)
    for suf in ('.ro.json', '.json'):
        if p.name.endswith(suf):
            return p.with_name(p.name[:-len(suf)] + '.reentryplan.json')
    return p.with_name(p.stem + '.reentryplan.json')


def _save_ro_to_library(ro) -> Path:
    """Write a hardware-only <safe_name>.ro.json, persist its reentry plan, and
    register the ready-to-fly object in RO_DB.

    Object files are hardware-only; the reentry plan (glide mode, turns, dives,
    separation) is written to the user reentry-plan library, where
    _load_ro_library merges it back on top by name.  This keeps GUI edits from
    re-embedding the plan into the object file."""
    _ensure_dir(_RO_LIBRARY_PATH)
    safe = _safe_name(ro.name) or "RO"
    fp = _RO_LIBRARY_PATH / f"{safe}.ro.json"
    fp.write_text(json.dumps(ro_to_dict(ro, include_reentry_plan=False), indent=2))
    save_reentry_plan(ro.name, extract_reentry_plan(ro), _REENTRY_PLAN_LIBRARY_PATH)
    RO_DB[ro.name] = lambda _r=ro: _r
    return fp


def _extract_ros_from_boosters():
    """One-time migration: copy every booster-embedded RV into ro_library/
    if not already present.  Guarded by a marker file so the user can
    delete library entries without having them re-extracted on next launch."""
    marker = _RO_LIBRARY_PATH / ".migrated"
    if marker.exists():
        return
    for name in list(BOOSTER_DB.keys()):
        try:
            p = BOOSTER_DB[name]()
        except Exception:
            continue
        ero = effective_ro(p)
        if ero is None or not ero.name or ero.name in RO_DB:
            continue
        try:
            _save_ro_to_library(ero)
        except Exception as exc:
            print(f"Warning: could not extract Reentry object '{ero.name}' from '{name}': {exc}")
    try:
        marker.touch()
    except Exception:
        pass


def _migrate_terminal_dive_default():
    """One-time migration: rewrite glider_terminal_alt_km 30.0 -> 0.0 in user
    reentry-plan files.  30 km was the old ride-along default and -- since no
    reentry-plan editor exists yet -- can only have been written by earlier
    migrations/saves, never chosen by the user.  Under the new semantics
    0 = glide to impact (dive on the target trigger only).  Marker-gated so a
    user who later hand-sets 30.0 keeps it."""
    marker = _REENTRY_PLAN_LIBRARY_PATH / ".dive_default_0"
    if marker.exists():
        return
    if _REENTRY_PLAN_LIBRARY_PATH.is_dir():
        for fp in _REENTRY_PLAN_LIBRARY_PATH.glob("*.reentryplan.json"):
            try:
                d = json.loads(fp.read_text())
                if float(d.get('glider_terminal_alt_km', 0.0)) == 30.0:
                    d['glider_terminal_alt_km'] = 0.0
                    fp.write_text(json.dumps(d, indent=2))
            except Exception as exc:
                print(f"Warning: could not migrate reentry plan '{fp.name}': {exc}")
    # Legacy user boosters may still embed a reentry object carrying the old
    # ride-along 30 km; rewrite those embeds too.
    try:
        if _CUSTOM_PATH.exists():
            allb = json.loads(_CUSTOM_PATH.read_text())
            changed = False
            for bd in allb.values():
                ro_d = bd.get('ro') if isinstance(bd, dict) else None
                if (isinstance(ro_d, dict)
                        and float(ro_d.get('glider_terminal_alt_km', 0.0)) == 30.0):
                    ro_d['glider_terminal_alt_km'] = 0.0
                    changed = True
            if changed:
                _CUSTOM_PATH.write_text(json.dumps(allb, indent=2))
    except Exception as exc:
        print(f"Warning: could not migrate custom boosters: {exc}")
    try:
        _REENTRY_PLAN_LIBRARY_PATH.mkdir(parents=True, exist_ok=True)
        marker.touch()
    except Exception:
        pass


def _migrate_analytic_family():
    """One-time migration for the family-identity split
    (REENTRY_FAMILY_DESIGN.md): a user plan that names a CLOSED-FORM ANALYTIC
    law (Tracy/Acton) but carries numerical-family capabilities — a non-trivial
    bank schedule or an armed dive-at-target — was, under the old silent
    fallback, actually being flown on the numerical EOM.  That fallback is
    deleted, so rewrite such plans to the numerical family's equilibrium law
    (dynamic_equilibrium_glide), which keeps their banking/dive-target
    functional and preserves the equilibrium-glide intent.  Marker-gated."""
    marker = _REENTRY_PLAN_LIBRARY_PATH / ".family_v1"
    if marker.exists():
        return
    if _REENTRY_PLAN_LIBRARY_PATH.is_dir():
        for fp in _REENTRY_PLAN_LIBRARY_PATH.glob("*.reentryplan.json"):
            try:
                d = json.loads(fp.read_text())
                if mm.glide_family(d.get('glider_guidance')) != 'analytic':
                    continue
                _banks = any(
                    (b and len(b) == 3 and float(b[0]) < float(b[1])
                     and float(b[2]) != 0.0)
                    for b in (d.get('glider_bank_schedule') or []))
                _dt = float(d.get('glider_dive_target_radius_km') or 0.0) > 0.0
                if _banks or _dt:
                    d['glider_guidance'] = 'dynamic_equilibrium_glide'
                    fp.write_text(json.dumps(d, indent=2))
                    print(f"Reentry plan '{fp.name}': analytic law with "
                          f"banking/dive-target — migrated to the numerical "
                          f"family (dynamic_equilibrium_glide).")
            except Exception as exc:
                print(f"Warning: could not migrate reentry plan '{fp.name}': {exc}")
    try:
        _REENTRY_PLAN_LIBRARY_PATH.mkdir(parents=True, exist_ok=True)
        marker.touch()
    except Exception:
        pass


_SITE_SEPARATOR = "──────────────────────────────"


def _bind_typeahead(cb):
    """
    Prefix-typeahead via a Toplevel autocomplete popup.

    As the user types, a popup appears directly below the combobox listing
    every item whose name begins with the typed prefix (case-insensitive).
    Separator entries (starting with '─') are excluded.  Works on macOS,
    Linux, and Windows without relying on the native dropdown widget.

    Commit paths
    ────────────
    • Click an item in the popup  → select + fire <<ComboboxSelected>>
    • Enter / Tab                 → best-prefix match + fire the event
    • ↓ arrow                     → move keyboard focus into the popup list
    • Escape                      → dismiss popup, leave field unchanged
    • FocusOut (click elsewhere)  → silently snap to best match
    """
    _all   = list(cb['values'])
    _popup = [None]   # Toplevel reference (reused, not recreated)
    _lb    = [None]   # Listbox inside the popup

    cb.config(state='normal')

    def _is_sep(v): return v.startswith('─')

    def _best(prefix):
        p = prefix.lower()
        return next((v for v in _all if not _is_sep(v) and v.lower().startswith(p)), None)

    def _matches(prefix):
        p = prefix.lower()
        return [v for v in _all if not _is_sep(v) and v.lower().startswith(p)]

    # ── popup lifecycle ───────────────────────────────────────────────────

    def _dismiss(event=None):
        if _popup[0] and _popup[0].winfo_exists():
            _popup[0].withdraw()

    def _show(items):
        if not _popup[0] or not _popup[0].winfo_exists():
            pop = tk.Toplevel(cb)
            pop.wm_overrideredirect(True)
            pop.attributes('-topmost', True)
            lb = tk.Listbox(pop, selectmode=tk.SINGLE,
                            exportselection=False, activestyle='dotbox')
            sb = ttk.Scrollbar(pop, orient=tk.VERTICAL, command=lb.yview)
            lb.config(yscrollcommand=sb.set)
            lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            sb.pack(side=tk.RIGHT, fill=tk.Y)
            lb.bind('<ButtonRelease-1>', _pick)
            lb.bind('<Return>', _pick)
            lb.bind('<Escape>', lambda e: (_dismiss(), cb.focus_set()))
            _popup[0] = pop
            _lb[0]    = lb

        lb = _lb[0]
        lb.delete(0, tk.END)
        for item in items:
            lb.insert(tk.END, item)

        n = min(10, len(items))
        lb.config(height=n)
        x = cb.winfo_rootx()
        y = cb.winfo_rooty() + cb.winfo_height()
        w = max(cb.winfo_width(), 180)
        _popup[0].geometry(f'{w}x{n * 20}+{x}+{y}')
        _popup[0].deiconify()
        _popup[0].lift()

    # ── selection ─────────────────────────────────────────────────────────

    def _pick(event=None):
        lb  = _lb[0]
        sel = lb.curselection()
        if not sel and event:
            sel = (lb.nearest(event.y),)
        if sel:
            value = lb.get(sel[0])
            _dismiss()
            cb.focus_set()
            cb.set(value)
            cb['values'] = _all
            cb.event_generate('<<ComboboxSelected>>')

    def _commit_fire(event=None):
        _dismiss()
        cb['values'] = _all
        m = _best(cb.get())
        if m:
            cb.set(m)
            cb.event_generate('<<ComboboxSelected>>')

    def _commit_silent_later(event=None):
        # Delay so a listbox click can register before we snap.
        cb.after(150, _do_commit_silent)

    def _do_commit_silent():
        # Wrapped in a broad try/except so that any TclError raised while
        # the combobox's parent window is in a modal-dialog state on macOS
        # (Python 3.11) does not propagate to tkinter's CallWrapper.__call__
        # and trigger a spurious UnboundLocalError crash there.
        try:
            try:
                focused = cb.focus_get()
            except Exception:
                focused = None
            if focused is _lb[0]:
                return   # user is navigating the popup — let _pick handle it
            _dismiss()
            cb['values'] = _all
            m = _best(cb.get())
            if m:
                cb.set(m)
        except Exception:
            pass

    def _on_selected(event=None):
        cb['values'] = _all
        _dismiss()

    # ── key handler ───────────────────────────────────────────────────────

    def _on_key(event=None):
        keysym = event.keysym if event else ''
        if keysym == 'Escape':
            _dismiss()
            return
        if keysym in ('Return', 'KP_Enter', 'Tab'):
            _commit_fire()
            return
        if keysym == 'Down':
            if _lb[0] and _popup[0] and _popup[0].winfo_exists():
                _lb[0].focus_set()
                if not _lb[0].curselection():
                    _lb[0].selection_set(0)
            return
        typed = cb.get()
        if not typed:
            _dismiss()
            return
        hits = _matches(typed)
        if hits:
            _show(hits)
        else:
            _dismiss()

    cb.bind('<KeyRelease>', _on_key)
    cb.bind('<FocusOut>',   _commit_silent_later)
    cb.bind('<<ComboboxSelected>>', _on_selected, add='+')


# Bundled sites (read-only) come from launch_sites.json in the source tree.
# User-added sites are stored separately so the bundled file stays clean.
_USER_SITES_PATH = Path.home() / ".gui_missile_flyout" / "user_sites.json"
_BUNDLED_SITE_NAMES: set = set()   # populated by _load_launch_sites()


def _load_user_sites() -> list:
    """Return list of user-defined site dicts, or [] on error/missing."""
    if not _USER_SITES_PATH.exists():
        return []
    try:
        return json.loads(_USER_SITES_PATH.read_text())
    except Exception as exc:
        print(f"Warning: could not load user_sites.json: {exc}")
        return []


def _save_user_sites(sites: list) -> None:
    _USER_SITES_PATH.parent.mkdir(parents=True, exist_ok=True)
    _USER_SITES_PATH.write_text(json.dumps(sites, indent=2))


def _load_launch_sites():
    """Return (combobox_values, name→site_dict) from bundled + user sites."""
    global _BUNDLED_SITE_NAMES
    path = Path(__file__).parent / "launch_sites.json"
    bundled = []
    if path.exists():
        try:
            bundled = json.loads(path.read_text())
        except Exception as exc:
            print(f"Warning: could not load launch_sites.json: {exc}")
    _BUNDLED_SITE_NAMES = {s["name"] for s in bundled}
    all_sites = bundled + _load_user_sites()
    by_country = {}
    for s in all_sites:
        by_country.setdefault(s["country"], []).append(s)
    values, site_map = [], {}
    for country in sorted(by_country):
        values.append(f"── {country} ──")
        for s in sorted(by_country[country], key=lambda x: x["name"]):
            values.append(s["name"])
            site_map[s["name"]] = s
    return values, site_map


# ---------------------------------------------------------------------------
# Reusable labelled entry helper
# ---------------------------------------------------------------------------

def _entry_row(parent, label, row, default, unit="", width=10):
    """Grid a Label + Entry + unit-label; return the StringVar."""
    ttk.Label(parent, text=label).grid(row=row, column=0,
                                       sticky=tk.W, padx=(6, 2), pady=2)
    var = tk.StringVar(value=default)
    inner = ttk.Frame(parent)
    inner.grid(row=row, column=1, sticky=tk.W, padx=(0, 6), pady=2)
    ttk.Entry(inner, textvariable=var, width=width).pack(side=tk.LEFT)
    if unit:
        ttk.Label(inner, text=unit).pack(side=tk.LEFT, padx=(2, 0))
    return var


# ---------------------------------------------------------------------------
# Stage sub-frame used inside BoosterDialog
# ---------------------------------------------------------------------------

class _StageFrame(ttk.LabelFrame):
    """Entry widgets for one rocket stage."""

    # Default thrust derived from default prop/isp/burn:
    # T = 230 × 9.80665 × (5000−1500) / 70 ≈ 112.9 kN
    _DEFAULTS = dict(fueled="5000", dry="1500", dia="0.88",
                     length="12.0", thrust_kn="112.9", isp="230",
                     nozzle_area="0", n_nozzles="1", nozzle_each="0",
                     coast="0")

    _G0 = 9.80665  # m/s²

    def __init__(self, parent, label, stage_num=1, defaults=None):
        super().__init__(parent, text=label)
        self._stage_num = stage_num
        d = {**self._DEFAULTS, **(defaults or {})}
        self._fueled      = _entry_row(self, "Fueled mass (kg):",    0, d["fueled"],      "kg")
        self._dry         = _entry_row(self, "Dry mass (kg):",       1, d["dry"],         "kg")
        self._dia         = _entry_row(self, "Diameter (m):",        2, d["dia"],         "m")
        self._length      = _entry_row(self, "Length (m):",          3, d["length"],      "m")
        # Thrust row (row 4) with Suggest button
        self._thrust_lbl = ttk.Label(self, text="Thrust (kN):")
        self._thrust_lbl.grid(row=4, column=0, sticky=tk.W, padx=(6, 2), pady=2)
        self._thrust_kn = tk.StringVar(value=d["thrust_kn"])
        _thr_inner = ttk.Frame(self)
        _thr_inner.grid(row=4, column=1, sticky=tk.W, padx=(0, 6), pady=2)
        ttk.Entry(_thr_inner, textvariable=self._thrust_kn, width=10).pack(side=tk.LEFT)
        ttk.Label(_thr_inner, text="kN").pack(side=tk.LEFT, padx=(2, 6))
        if self._stage_num == 1:
            ttk.Button(_thr_inner, text="Estimate…",
                       command=self._suggest_thrust).pack(side=tk.LEFT)
        # Isp (row 5) — user-entered for liquid; computed (readonly) for solid.
        ttk.Label(self, text="Isp (vacuum, s):").grid(
            row=5, column=0, sticky=tk.W, padx=(6, 2), pady=2)
        self._isp = tk.StringVar(value=d["isp"])
        _isp_inner = ttk.Frame(self)
        _isp_inner.grid(row=5, column=1, sticky=tk.W, padx=(0, 6), pady=2)
        self._isp_entry = ttk.Entry(_isp_inner, textvariable=self._isp, width=10)
        self._isp_entry.pack(side=tk.LEFT)
        self._isp_hint_lbl = ttk.Label(_isp_inner, text="s", foreground="gray50")
        self._isp_hint_lbl.pack(side=tk.LEFT, padx=(2, 0))
        # Nozzle exit area — TOTAL (physics + Estimate) on row 6, plus the
        # geometry inputs (count × per-nozzle) that FEED the total and the
        # 3-D export's base disks.  Per-nozzle drives the total when set;
        # Estimate fills the total directly and clears per-nozzle.
        ttk.Label(self, text="Nozzle exit area (m²):").grid(
            row=6, column=0, sticky=tk.W, padx=(6, 2), pady=2)
        self._nozzle_area = tk.StringVar(value=d["nozzle_area"])
        _noz_inner = ttk.Frame(self)
        _noz_inner.grid(row=6, column=1, sticky=tk.W, padx=(0, 6), pady=2)
        ttk.Entry(_noz_inner, textvariable=self._nozzle_area, width=10).pack(side=tk.LEFT)
        ttk.Label(_noz_inner, text="m² total").pack(side=tk.LEFT, padx=(2, 6))
        if self._stage_num == 1:
            ttk.Button(_noz_inner, text="Estimate…",
                       command=self._suggest_nozzle_area).pack(side=tk.LEFT)
        ttk.Label(self, text="  nozzles × each:").grid(
            row=8, column=0, sticky=tk.W, padx=(6, 2), pady=2)
        _noz_geo = ttk.Frame(self)
        _noz_geo.grid(row=8, column=1, sticky=tk.W, padx=(0, 6), pady=2)
        self._n_nozzles = tk.StringVar(value=str(d.get("n_nozzles", 1)))
        ttk.Entry(_noz_geo, textvariable=self._n_nozzles, width=4).pack(side=tk.LEFT)
        ttk.Label(_noz_geo, text="×").pack(side=tk.LEFT, padx=3)
        self._nozzle_each = tk.StringVar(value=str(d.get("nozzle_each", 0)))
        ttk.Entry(_noz_geo, textvariable=self._nozzle_each, width=9).pack(side=tk.LEFT)
        ttk.Label(_noz_geo, text="m² each  (→ total)").pack(side=tk.LEFT, padx=(2, 0))

        def _sum_nozzles(*_):
            # per-nozzle × count -> total, only when per-nozzle is set (>0),
            # so a directly-typed / estimated total is never clobbered.
            try:
                each = float(self._nozzle_each.get() or 0.0)
                n = int(float(self._n_nozzles.get() or 1))
            except (ValueError, tk.TclError):
                return
            if each > 0 and n > 0:
                self._nozzle_area.set(f"{each * n:.4f}")
        self._n_nozzles.trace_add("write", _sum_nozzles)
        self._nozzle_each.trace_add("write", _sum_nozzles)

        # Burn time (row 7) — readonly/computed for liquid; user-entered for solid.
        ttk.Label(self, text="Burn time (s):").grid(
            row=7, column=0, sticky=tk.W, padx=(6, 2), pady=2)
        self._burn_var = tk.StringVar()
        _burn_inner = ttk.Frame(self)
        _burn_inner.grid(row=7, column=1, sticky=tk.W, padx=(0, 6), pady=2)
        self._burn_entry = ttk.Entry(_burn_inner, textvariable=self._burn_var,
                                     width=10, state="readonly")
        self._burn_entry.pack(side=tk.LEFT)
        self._burn_hint_lbl = ttk.Label(_burn_inner, text="s  (computed)",
                                        foreground="gray50")
        self._burn_hint_lbl.pack(side=tk.LEFT, padx=(2, 0))

        # Propellant type selector (row 9)
        _prop_row = ttk.Frame(self)
        _prop_row.grid(row=9, column=0, columnspan=2,
                       sticky=tk.W, padx=(6, 2), pady=(2, 4))
        ttk.Label(_prop_row, text="Propellant:").pack(side=tk.LEFT)
        self._propellant_var = tk.StringVar(value="Liquid")
        self._propellant_cb = ttk.Combobox(
            _prop_row, textvariable=self._propellant_var,
            values=["Liquid", "Solid"], state="readonly", width=10)
        self._propellant_cb.pack(side=tk.LEFT, padx=(4, 0))
        self._propellant_cb.bind("<<ComboboxSelected>>",
                                 lambda _e: self._on_propellant_changed())

        # ── Solid grain profile block (row 10) — hidden until solid is checked ──
        _GRAIN_KEYS   = list(GRAIN_LABELS.keys())
        _GRAIN_LABELS = [GRAIN_LABELS[k] for k in _GRAIN_KEYS]
        self._grain_keys = _GRAIN_KEYS

        self._solid_frame = ttk.LabelFrame(self, text="Grain profile")
        self._solid_frame.grid(row=10, column=0, columnspan=2,
                               sticky=tk.EW, padx=4, pady=(0, 4))
        self._solid_frame.columnconfigure(1, weight=1)
        self._solid_frame.grid_remove()

        # Row 0: grain type selector
        ttk.Label(self._solid_frame, text="Grain type:").grid(
            row=0, column=0, sticky=tk.W, padx=(6, 2), pady=2)
        self._grain_var = tk.StringVar(value=GRAIN_LABELS["star"])
        self._grain_cb = ttk.Combobox(self._solid_frame, textvariable=self._grain_var,
                                      values=_GRAIN_LABELS, state="readonly", width=28)
        self._grain_cb.current(3)   # star
        self._grain_cb.grid(row=0, column=1, sticky=tk.W, padx=(0, 6), pady=2)
        self._grain_cb.bind("<<ComboboxSelected>>", self._on_grain_changed)

        # Row 1: thrust specification toggle (peak vs average)
        ttk.Label(self._solid_frame, text="Specify:").grid(
            row=1, column=0, sticky=tk.W, padx=(6, 2), pady=2)
        _tmode_f = ttk.Frame(self._solid_frame)
        _tmode_f.grid(row=1, column=1, sticky=tk.W, padx=(0, 6), pady=2)
        self._thrust_mode_var = tk.StringVar(value="average")
        ttk.Radiobutton(_tmode_f, text="Peak thrust",
                        variable=self._thrust_mode_var, value="peak",
                        command=self._on_thrust_mode_changed).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(_tmode_f, text="Average thrust",
                        variable=self._thrust_mode_var, value="average",
                        command=self._on_thrust_mode_changed).pack(side=tk.LEFT)

        # Row 2: computed alternate thrust
        ttk.Label(self._solid_frame, text="Computed:").grid(
            row=2, column=0, sticky=tk.W, padx=(6, 2), pady=2)
        self._alt_thrust_lbl = ttk.Label(self._solid_frame, text="—",
                                         foreground="navy")
        self._alt_thrust_lbl.grid(row=2, column=1, sticky=tk.W, padx=(0, 6), pady=2)

        # Row 3: boost-phase fraction (two-phase grains only)
        self._boost_frac_lbl = ttk.Label(self._solid_frame, text="Boost phase:")
        self._boost_frac_lbl.grid(row=3, column=0, sticky=tk.W, padx=(6, 2), pady=2)
        _boost_f = ttk.Frame(self._solid_frame)
        _boost_f.grid(row=3, column=1, sticky=tk.W, padx=(0, 6), pady=2)
        self._boost_frac_var = tk.StringVar(value="35")
        ttk.Entry(_boost_f, textvariable=self._boost_frac_var,
                  width=6).pack(side=tk.LEFT)
        ttk.Label(_boost_f, text="% of burn time").pack(side=tk.LEFT, padx=(4, 0))
        self._boost_frac_inner = _boost_f
        self._boost_frac_lbl.grid_remove()
        self._boost_frac_inner.grid_remove()

        # Row 4: custom CSV profile
        ttk.Label(self._solid_frame, text="Custom profile:").grid(
            row=4, column=0, sticky=tk.W, padx=(6, 2), pady=2)
        _csv_f = ttk.Frame(self._solid_frame)
        _csv_f.grid(row=4, column=1, sticky=tk.W, padx=(0, 6), pady=2)
        self._profile_path_var = tk.StringVar(value="")
        ttk.Label(_csv_f, textvariable=self._profile_path_var,
                  foreground="gray50", width=22,
                  anchor=tk.W).pack(side=tk.LEFT)
        ttk.Button(_csv_f, text="Browse…",
                   command=self._browse_profile).pack(side=tk.LEFT, padx=(6, 0))

        # Row 5: fill-factor warning
        self._fill_warn_lbl = ttk.Label(self._solid_frame, text="",
                                        foreground="darkorange", wraplength=340)
        self._fill_warn_lbl.grid(row=5, column=0, columnspan=2,
                                 sticky=tk.W, padx=(6, 2), pady=(2, 4))

        # ── Conical stage (row 11) ─────────────────────────────────────────
        # A tapered stage body: frustum from Diameter (base) to Top ⌀.
        _con_f = ttk.Frame(self)
        _con_f.grid(row=11, column=0, columnspan=2, sticky=tk.W, padx=6, pady=(4, 0))
        self._conical_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(_con_f, text="Conical stage", variable=self._conical_var,
                        command=self._on_conical).pack(side=tk.LEFT)
        ttk.Label(_con_f, text="Top ⌀ (m):").pack(side=tk.LEFT, padx=(10, 2))
        self._top_dia_var = tk.StringVar(value="0")
        self._top_dia_entry = ttk.Entry(_con_f, textvariable=self._top_dia_var, width=8)
        self._top_dia_entry.pack(side=tk.LEFT)

        # ── Interstage on top (row 12) ─────────────────────────────────────
        # Adapter sitting on this stage.  Diameters are DERIVED (this stage's
        # top → the next stage's base); only length, mass, jettison are free.
        _is_f = ttk.Frame(self)
        _is_f.grid(row=12, column=0, columnspan=2, sticky=tk.W, padx=6, pady=(2, 4))
        self._interstage_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(_is_f, text="Interstage on top",
                        variable=self._interstage_var,
                        command=self._on_interstage).pack(side=tk.LEFT)
        ttk.Label(_is_f, text="L (m):").pack(side=tk.LEFT, padx=(10, 2))
        self._is_len_var = tk.StringVar(value="0")
        self._is_len_entry = ttk.Entry(_is_f, textvariable=self._is_len_var, width=6)
        self._is_len_entry.pack(side=tk.LEFT)
        ttk.Label(_is_f, text="mass (kg):").pack(side=tk.LEFT, padx=(8, 2))
        self._is_mass_var = tk.StringVar(value="0")
        self._is_mass_entry = ttk.Entry(_is_f, textvariable=self._is_mass_var, width=8)
        self._is_mass_entry.pack(side=tk.LEFT)
        ttk.Label(_is_f, text="jettison (s):").pack(side=tk.LEFT, padx=(8, 2))
        self._is_jett_var = tk.StringVar(value="")
        self._is_jett_entry = ttk.Entry(_is_f, textvariable=self._is_jett_var, width=7)
        self._is_jett_entry.pack(side=tk.LEFT)
        ttk.Label(_is_f, text="(blank = stage sep.)",
                  foreground="gray50").pack(side=tk.LEFT, padx=(4, 0))
        self._on_conical()
        self._on_interstage()

        # Recompute burn whenever any of the four driving fields change
        for _v in (self._fueled, self._dry, self._thrust_kn, self._isp):
            _v.trace_add("write", self._recompute_burn)
        self._recompute_burn()

    def _on_conical(self):
        st = "normal" if self._conical_var.get() else "disabled"
        self._top_dia_entry.config(state=st)

    def _on_interstage(self):
        st = "normal" if self._interstage_var.get() else "disabled"
        for e in (self._is_len_entry, self._is_mass_entry, self._is_jett_entry):
            e.config(state=st)

    def _recompute_burn(self, *_):
        """Compute burn = Isp × g₀ × prop / avg_thrust.

        For liquids, avg_thrust = entered thrust.  For solids, avg_thrust is
        peak_thrust × fill_factor when peak is specified, or the entered
        value directly when average is specified.
        """
        # Solid-only display side effects (alt-thrust label, fill warning)
        self._update_solid_display()
        try:
            prop      = float(self._fueled.get()) - float(self._dry.get())
            thrust_kn = float(self._thrust_kn.get())
            isp       = float(self._isp.get())
            if prop <= 0 or thrust_kn <= 0 or isp <= 0:
                raise ValueError
            if getattr(self, '_propellant_var', None) and self._propellant_var.get() == "Solid":
                key  = self._get_grain_key()
                fill = grain_fill_factor(key) if key else 1.0
                thrust_avg_kn = (thrust_kn * fill
                                 if self._thrust_mode_var.get() == "peak"
                                 else thrust_kn)
            else:
                thrust_avg_kn = thrust_kn
            thrust_n = thrust_avg_kn * 1000.0
            if thrust_n <= 0:
                raise ValueError
            self._burn_var.set(f"{isp * self._G0 * prop / thrust_n:.1f}")
        except (ValueError, ZeroDivisionError):
            self._burn_var.set("—")

    def _on_propellant_changed(self):
        """Show/hide grain frame; relabel thrust field for peak/avg context."""
        is_solid = self._propellant_var.get() == "Solid"
        if is_solid:
            self._solid_frame.grid()
            self._thrust_lbl.config(text=self._thrust_label_text())
        else:
            self._solid_frame.grid_remove()
            self._thrust_lbl.config(text="Thrust (kN):")
        # burn_time and Isp behaviour is identical for liquid and solid:
        # burn_time is always computed (readonly), Isp is always user-input.
        self._burn_entry.config(state="readonly")
        self._burn_hint_lbl.config(text="s  (computed)")
        self._isp_entry.config(state="normal")
        self._isp_hint_lbl.config(text="s")
        self._recompute_burn()

    def _thrust_label_text(self):
        mode = getattr(self, '_thrust_mode_var', None)
        if mode and mode.get() == "average":
            return "Avg thrust (kN):"
        return "Peak thrust (kN):"

    def _on_thrust_mode_changed(self):
        self._thrust_lbl.config(text=self._thrust_label_text())
        self._recompute_burn()

    def _on_grain_changed(self, *_):
        key = self._get_grain_key()
        two_phase = key in ("multi_fin", "dual_composition")
        if two_phase:
            self._boost_frac_lbl.grid()
            self._boost_frac_inner.grid()
        else:
            self._boost_frac_lbl.grid_remove()
            self._boost_frac_inner.grid_remove()
        self._recompute_burn()

    def _get_grain_key(self):
        label = self._grain_var.get()
        for k, v in GRAIN_LABELS.items():
            if v == label:
                return k
        return ""

    def _update_solid_display(self, *_):
        """Update the alternate-thrust label and fill-factor warning (solid mode only)."""
        if not (hasattr(self, '_grain_cb') and self._propellant_var.get() == "Solid"):
            return
        key = self._get_grain_key()
        fill = grain_fill_factor(key) if key else 1.0
        try:
            thrust_entered = float(self._thrust_kn.get())
            mode = self._thrust_mode_var.get()
            if mode == "peak":
                alt_kn = thrust_entered * fill
                self._alt_thrust_lbl.config(
                    text=f"{alt_kn:.1f} kN  (average, fill factor {fill:.3f})")
            else:
                if fill > 0:
                    alt_kn = thrust_entered / fill
                    self._alt_thrust_lbl.config(
                        text=f"{alt_kn:.1f} kN  (peak, fill factor {fill:.3f})")
                else:
                    self._alt_thrust_lbl.config(text="—")
        except (ValueError, ZeroDivisionError):
            self._alt_thrust_lbl.config(text="—")
        # Fill-factor warning
        if key and key in _GRAIN_FILL_RANGE:
            lo, hi = _GRAIN_FILL_RANGE[key]
            if not (lo <= fill <= hi):
                self._fill_warn_lbl.config(
                    text=f"\u26a0 Fill factor {fill:.3f} outside typical range "
                         f"{lo}–{hi} for {GRAIN_LABELS.get(key, key)}")
            else:
                self._fill_warn_lbl.config(text="")
        else:
            self._fill_warn_lbl.config(text="")

    def _browse_profile(self):
        """Let user pick a CSV thrust-profile file; show preview plot."""
        import tkinter.filedialog as fd
        import csv as _csv
        path = fd.askopenfilename(
            parent=self,
            title="Select thrust profile CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
        if not path:
            return
        # Parse two-column CSV: t_frac, F_frac
        pairs = []
        try:
            with open(path, newline='') as f:
                reader = _csv.reader(f)
                for row in reader:
                    row = [c.strip() for c in row if c.strip()]
                    if len(row) >= 2:
                        try:
                            pairs.append((float(row[0]), float(row[1])))
                        except ValueError:
                            pass   # skip header rows
        except OSError as exc:
            tk.messagebox.showerror("Cannot open file", str(exc), parent=self)
            return
        if len(pairs) < 2:
            tk.messagebox.showerror(
                "Invalid profile",
                "File must contain at least two rows with columns: t_frac, F_frac",
                parent=self)
            return
        import os
        self._profile_path_var.set(os.path.basename(path))
        self._profile_data = pairs   # stored for get()
        self._show_profile_preview(pairs, path)

    def _show_profile_preview(self, pairs, path):
        """Pop up a small Matplotlib preview of the thrust profile."""
        try:
            import matplotlib
            matplotlib.use("TkAgg")
            import matplotlib.pyplot as _plt
        except ImportError:
            tk.messagebox.showinfo(
                "No preview",
                "matplotlib is not installed; profile loaded without preview.",
                parent=self)
            return
        fig, ax = _plt.subplots(figsize=(5, 3), tight_layout=True)
        ts = [p[0] for p in pairs]
        fs = [p[1] for p in pairs]
        ax.step(ts, fs, where='post', color='steelblue', linewidth=1.5)
        ax.set_xlabel("t / burn time")
        ax.set_ylabel("F / F_peak")
        ax.set_title(f"Thrust profile: {path}")
        ax.set_xlim(0, 1)
        ax.set_ylim(0, None)
        ax.grid(True, alpha=0.3)
        _plt.show()

    @staticmethod
    def _iter_entries(widget):
        """Yield all ttk.Entry descendants of widget."""
        for child in widget.winfo_children():
            if isinstance(child, ttk.Entry):
                yield child
            else:
                yield from _StageFrame._iter_entries(child)

    def _suggest_nozzle_area(self):
        """Estimate Ae = (g₀ / p₀) × ṁ × Isp_vac × performance_factor"""
        try:
            isp_vac = float(self._isp.get())
            prop    = float(self._fueled.get()) - float(self._dry.get())
            burn    = float(self._burn_var.get())
            if prop <= 0 or burn <= 0 or isp_vac <= 0:
                raise ValueError
        except (ValueError, TypeError):
            tk.messagebox.showerror(
                "Cannot estimate",
                "Please enter valid fueled mass, dry mass, Isp, and thrust first.",
                parent=self)
            return

        mdot = prop / burn

        dlg = tk.Toplevel(self)
        dlg.title("Estimate Nozzle Exit Area")
        dlg.resizable(False, False)
        dlg.grab_set()

        ttk.Label(dlg, text="Isp_vac (s):").grid(
            row=0, column=0, sticky=tk.W, padx=(10, 4), pady=(10, 2))
        isp_var = tk.StringVar(value=str(isp_vac))
        ttk.Entry(dlg, textvariable=isp_var, width=10).grid(
            row=0, column=1, sticky=tk.W, padx=(0, 10), pady=(10, 2))

        ttk.Label(dlg, text="Performance factor:").grid(
            row=1, column=0, sticky=tk.W, padx=(10, 4), pady=2)
        pf_var = tk.StringVar(value="0.10")
        ttk.Entry(dlg, textvariable=pf_var, width=10).grid(
            row=1, column=1, sticky=tk.W, padx=(0, 10), pady=2)

        result_var = tk.StringVar(value="")
        ttk.Label(dlg, textvariable=result_var, foreground="navy").grid(
            row=2, column=0, columnspan=2, padx=10, pady=(6, 4))

        def _compute(*_):
            try:
                isp  = float(isp_var.get())
                pf   = float(pf_var.get())
                if isp <= 0 or pf <= 0:
                    raise ValueError
                ae = (self._G0 / 101325.0) * mdot * isp * pf
                result_var.set(f"Ae ≈ {ae:.4f} m²")
                return ae
            except (ValueError, TypeError):
                result_var.set("Enter valid Isp and performance factor.")
                return None

        isp_var.trace_add("write", lambda *_: _compute())
        pf_var .trace_add("write", lambda *_: _compute())

        btn_row = ttk.Frame(dlg)
        btn_row.grid(row=3, column=0, columnspan=2, pady=(4, 10))

        def _accept():
            ae = _compute()
            if ae is not None:
                # Estimate produces a TOTAL — clear per-nozzle so its
                # coupling trace can't overwrite it; the export then
                # derives each = total / n_nozzles.
                if hasattr(self, "_nozzle_each"):
                    self._nozzle_each.set("0")
                self._nozzle_area.set(f"{ae:.4f}")
                dlg.destroy()

        ttk.Button(btn_row, text="Accept", command=_accept).pack(
            side=tk.LEFT, padx=6)
        ttk.Button(btn_row, text="Cancel",
                   command=dlg.destroy).pack(side=tk.LEFT, padx=6)

        # Centre over parent
        dlg.update_idletasks()
        px = self.winfo_rootx() + (self.winfo_width()  - dlg.winfo_reqwidth())  // 2
        py = self.winfo_rooty() + (self.winfo_height() - dlg.winfo_reqheight()) // 2
        dlg.geometry(f"+{px}+{py}")

    def _suggest_thrust(self):
        """Estimate thrust from observed rocket acceleration during boost."""
        import math
        G0 = 9.80665

        dlg = tk.Toplevel(self)
        dlg.title("Estimate Thrust")
        dlg.resizable(False, False)
        dlg.grab_set()

        # ── Input fields ──────────────────────────────────────────────
        frm = ttk.Frame(dlg, padding=12)
        frm.pack(fill=tk.X)
        frm.columnconfigure(1, weight=1)

        def _lbl(row, text):
            ttk.Label(frm, text=text).grid(
                row=row, column=0, sticky=tk.W, padx=(0, 8), pady=3)

        _lbl(0, "Fueled mass (kg):")
        mass_var = tk.StringVar(value=self._fueled.get())
        ttk.Entry(frm, textvariable=mass_var, width=10).grid(
            row=0, column=1, sticky=tk.W)

        _lbl(1, "Vertical acceleration (m/s², upward +):")
        av_var = tk.StringVar(value="")
        ttk.Entry(frm, textvariable=av_var, width=10).grid(
            row=1, column=1, sticky=tk.W)

        _lbl(2, "Horizontal acceleration (m/s²):")
        ah_inner = ttk.Frame(frm)
        ah_inner.grid(row=2, column=1, sticky=tk.W)
        ah_var = tk.StringVar(value="0")
        ttk.Entry(ah_inner, textvariable=ah_var, width=10).pack(side=tk.LEFT)
        ttk.Label(ah_inner, text="  (0 for vertical flight)",
                  foreground="gray50").pack(side=tk.LEFT)

        ttk.Separator(dlg, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=12)

        # ── Result display ────────────────────────────────────────────
        res_frm = ttk.Frame(dlg, padding=(12, 8))
        res_frm.pack(fill=tk.X)
        res_frm.columnconfigure(1, weight=1)
        ttk.Label(res_frm, text="Estimated thrust:").grid(
            row=0, column=0, sticky=tk.W, padx=(0, 8), pady=2)
        thrust_lbl = ttk.Label(res_frm, text="—",
                               font=("", 11, "bold"), foreground="navy")
        thrust_lbl.grid(row=0, column=1, sticky=tk.W)
        note_lbl = ttk.Label(res_frm, text="", foreground="gray50", wraplength=360)
        note_lbl.grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=(2, 0))

        _thrust_result = [None]

        def _compute(*_):
            try:
                mass = float(mass_var.get())
                if mass <= 0:
                    raise ValueError
            except (ValueError, TypeError):
                thrust_lbl.config(text="—")
                note_lbl.config(text="invalid mass")
                _thrust_result[0] = None
                return
            try:
                av  = float(av_var.get())
                ah  = float(ah_var.get())
                f_n = mass * math.sqrt(ah**2 + (av + G0)**2)
                if f_n <= 0:
                    raise ValueError
            except (ValueError, TypeError):
                thrust_lbl.config(text="—")
                note_lbl.config(text="")
                _thrust_result[0] = None
                return
            note = (f"T = m·√(a_h²+(a_v+g)²)  =  {mass:.0f}·"
                    f"√({ah:.2f}²+{av+G0:.3f}²)  =  {f_n/1000:.2f} kN")
            thrust_lbl.config(text=f"{f_n/1000:,.1f} kN")
            note_lbl.config(text=note)
            _thrust_result[0] = f_n / 1000.0

        for _v in (mass_var, av_var, ah_var):
            _v.trace_add("write", _compute)

        # ── Buttons ───────────────────────────────────────────────────
        ttk.Separator(dlg, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=12)
        btn_frm = ttk.Frame(dlg, padding=(12, 8))
        btn_frm.pack(fill=tk.X)

        def _use():
            if _thrust_result[0] is not None:
                self._thrust_kn.set(f"{_thrust_result[0]:.1f}")
            dlg.destroy()

        ttk.Button(btn_frm, text="Use this value", command=_use).pack(side=tk.LEFT)
        ttk.Button(btn_frm, text="Cancel",
                   command=dlg.destroy).pack(side=tk.LEFT, padx=6)

    def set_readonly(self, readonly: bool):
        """Set all editable entry fields to readonly (for Forden reference boosters)."""
        state = "readonly" if readonly else "normal"
        self._propellant_cb.config(state="disabled" if readonly else "readonly")
        self._grain_cb.config(state="disabled" if readonly else "readonly")
        for entry in self._iter_entries(self):
            # Burn entry: readonly unless solid motor is enabled (managed by _on_propellant_changed)
            if entry is self._burn_entry:
                if not readonly and self._propellant_var.get() == "Solid":
                    entry.config(state="normal")
                else:
                    entry.config(state="readonly")
            else:
                entry.config(state=state)

    def get(self):
        burn_str = self._burn_var.get()
        if burn_str == "—":
            raise ValueError("Burn time could not be computed — check thrust, Isp, and masses.")
        _LABELS = {
            "fueled": "Fueled Mass", "dry": "Dry Mass", "dia": "Diameter",
            "length": "Length", "thrust_kn": "Thrust", "isp": "Isp",
            "nozzle_area": "Nozzle Exit Area",
        }
        result = {}
        for k, v in [
            ("fueled",      self._fueled),      ("dry",         self._dry),
            ("dia",         self._dia),         ("length",      self._length),
            ("thrust_kn",   self._thrust_kn),   ("isp",         self._isp),
            ("nozzle_area", self._nozzle_area),
        ]:
            try:
                result[k] = float(v.get())
            except ValueError:
                raise ValueError(
                    f"{_LABELS.get(k, k)}: expected a number, got {v.get()!r:.40s}"
                )
        try:
            result["burn"] = float(burn_str)
        except ValueError:
            raise ValueError(f"Burn time: expected a number, got {burn_str!r:.40s}")
        # Nozzle geometry (count + per-nozzle) — optional; per-nozzle already
        # summed into nozzle_area by the coupling trace.  Tolerant parse:
        # blanks/junk fall back to 1 nozzle / 0 per-nozzle.
        try:
            result["n_nozzles"] = max(1, int(float(self._n_nozzles.get() or 1)))
        except (ValueError, tk.TclError):
            result["n_nozzles"] = 1
        try:
            result["nozzle_each"] = max(0.0, float(self._nozzle_each.get() or 0.0))
        except (ValueError, tk.TclError):
            result["nozzle_each"] = 0.0
        result["coast"] = 0.0   # coast is now set in the advanced pitch panel
        result["solid_motor"] = (self._propellant_var.get() == "Solid")

        # Solid-motor grain fields
        result["grain_type"]    = ""
        result["thrust_peak_N"] = 0.0
        result["thrust_profile"] = []
        if result["solid_motor"]:
            grain_key = self._get_grain_key()
            result["grain_type"] = grain_key
            fill = grain_fill_factor(grain_key) if grain_key else 1.0
            try:
                thrust_entered_n = float(self._thrust_kn.get()) * 1000.0
                if self._thrust_mode_var.get() == "peak":
                    result["thrust_peak_N"] = thrust_entered_n
                    # thrust_kn already holds peak; override to avg for prop-mass consistency
                    result["thrust_kn"] = thrust_entered_n * fill / 1000.0
                else:
                    result["thrust_peak_N"] = (thrust_entered_n / fill
                                               if fill > 0 else thrust_entered_n)
                    # thrust_kn holds avg — leave as is
            except ValueError:
                pass
            result["thrust_profile"] = getattr(self, '_profile_data', [])

        # Conical stage + interstage (per-stage structure)
        result["conical"] = bool(self._conical_var.get())
        try:
            result["top_diameter_m"] = (max(0.0, float(self._top_dia_var.get()))
                                        if result["conical"] else 0.0)
        except ValueError:
            raise ValueError("Conical top diameter must be a number.")
        result["has_interstage"] = bool(self._interstage_var.get())
        if result["has_interstage"]:
            try:
                result["interstage_length_m"] = max(0.0, float(self._is_len_var.get()))
                result["interstage_mass_kg"]  = max(0.0, float(self._is_mass_var.get()))
            except ValueError:
                raise ValueError("Interstage length and mass must be numbers.")
            _jt = self._is_jett_var.get().strip()
            try:
                result["interstage_jettison_s"] = float(_jt) if _jt else None
            except ValueError:
                raise ValueError("Interstage jettison time must be a number (or blank).")
        else:
            result["interstage_length_m"] = 0.0
            result["interstage_mass_kg"]  = 0.0
            result["interstage_jettison_s"] = None
        return result

    def populate(self, d):
        # Back-calculate thrust_kn from stored burn/isp/prop so the round-trip
        # is exact: T = Isp × g₀ × prop / burn
        prop = d["fueled"] - d["dry"]
        burn = d["burn"]
        thrust_kn = (d["isp"] * self._G0 * prop / burn / 1000.0
                     if burn > 0 and prop > 0 else 0.0)

        self._fueled      .set(str(d["fueled"]))
        self._dry         .set(str(d["dry"]))
        self._dia         .set(str(d["dia"]))
        self._length      .set(str(d["length"]))
        self._thrust_kn   .set(f"{thrust_kn:.1f}")
        self._isp         .set(str(d["isp"]))
        self._nozzle_area .set(str(d.get("nozzle_area", 0)))
        # per-nozzle first so its coupling trace doesn't clobber the total
        self._nozzle_each .set(str(d.get("nozzle_each", 0)))
        self._n_nozzles   .set(str(d.get("n_nozzles", 1)))
        self._nozzle_area .set(str(d.get("nozzle_area", 0)))
        # _burn_var is updated automatically by traces on Isp/thrust/masses
        self._propellant_var.set("Solid" if d.get("solid_motor", False) else "Liquid")

        # Grain profile fields
        grain_key = d.get("grain_type", "")
        if grain_key and grain_key in GRAIN_LABELS:
            self._grain_var.set(GRAIN_LABELS[grain_key])
        else:
            self._grain_var.set(GRAIN_LABELS.get("star", ""))

        thrust_peak_N = float(d.get("thrust_peak_N", 0.0))
        if thrust_peak_N > 0.0:
            self._thrust_mode_var.set("peak")
            self._thrust_kn.set(f"{thrust_peak_N / 1000.0:.1f}")
        else:
            self._thrust_mode_var.set("average")
            # thrust_kn is already set above from Isp/prop/burn (average thrust)

        profile = d.get("thrust_profile", [])
        self._profile_data = [tuple(p) for p in profile] if profile else []
        if self._profile_data:
            self._profile_path_var.set(f"<{len(self._profile_data)} pts>")
        else:
            self._profile_path_var.set("")

        # Conical stage + interstage
        self._conical_var.set(bool(d.get("conical", False)))
        self._top_dia_var.set(str(d.get("top_diameter_m", 0.0) or 0.0))
        self._interstage_var.set(bool(d.get("has_interstage", False)))
        self._is_len_var.set(str(d.get("interstage_length_m", 0.0) or 0.0))
        self._is_mass_var.set(str(d.get("interstage_mass_kg", 0.0) or 0.0))
        _jt = d.get("interstage_jettison_s", None)
        self._is_jett_var.set("" if _jt is None else f"{_jt:g}")
        self._on_conical()
        self._on_interstage()

        # Trigger UI state update — burn time will be recomputed from
        # Isp / prop / (peak × fill_factor or average) thrust by _recompute_burn.
        if self._propellant_var.get() == "Solid":
            self._on_propellant_changed()
            self._on_grain_changed()
            self._on_thrust_mode_changed()


# ---------------------------------------------------------------------------
# New / Edit booster dialog
# ---------------------------------------------------------------------------

class BoosterDialog(tk.Toplevel):
    """Modal dialog for creating or editing a custom booster."""

    def __init__(self, parent, on_save, existing_name=None):
        super().__init__(parent)
        self._on_save = on_save
        self._existing_name = existing_name
        self._readonly_mode = (existing_name is not None
                               and existing_name.endswith(" (Forden)"))
        if self._readonly_mode:
            self.title("View Booster — Forden Reference")
        elif existing_name:
            self.title("Edit Booster")
        else:
            self.title("New Booster")
        self.resizable(False, True)
        self.grab_set()               # modal
        self._build(existing_name)
        # Centre over parent; cap height to 90 % of screen so dialog is scrollable
        self.update_idletasks()
        max_h = int(parent.winfo_screenheight() * 0.90)
        nat_h = self.winfo_reqheight()
        dlg_h = min(nat_h, max_h)
        dlg_w = self.winfo_reqwidth()
        px = parent.winfo_rootx() + (parent.winfo_width()  - dlg_w) // 2
        py = parent.winfo_rooty() + (parent.winfo_height() - dlg_h) // 2
        self.geometry(f"{dlg_w}x{dlg_h}+{px}+{py}")

    # ------------------------------------------------------------------
    def _build(self, existing_name):
        pad = dict(padx=8, pady=4)

        # Name row
        nf = ttk.Frame(self)
        nf.pack(fill=tk.X, **pad)
        ttk.Label(nf, text="Booster name:").pack(side=tk.LEFT)
        self._name_var = tk.StringVar(value=existing_name or "My Booster")
        self._name_entry = ttk.Entry(nf, textvariable=self._name_var, width=24)
        self._name_entry.pack(side=tk.LEFT, padx=(6, 16))
        ttk.Label(nf, text="Stages:").pack(side=tk.LEFT)
        self._n_stages_var = tk.StringVar(value="1")
        self._stages_cb = ttk.Combobox(nf, textvariable=self._n_stages_var,
                                       values=["1", "2", "3", "4"],
                                       state="readonly", width=3)
        self._stages_cb.pack(side=tk.LEFT, padx=(4, 0))
        self._stages_cb.bind("<<ComboboxSelected>>",
                             lambda _: self._update_stage_frames())
        ttk.Label(nf, text="  Boosters:").pack(side=tk.LEFT)
        self._n_boosters_var = tk.StringVar(value="0")
        self._n_boosters_spin = ttk.Spinbox(
            nf, textvariable=self._n_boosters_var, from_=0, to=9,
            width=2, command=self._update_booster_frame)
        self._n_boosters_spin.pack(side=tk.LEFT, padx=(4, 0))
        self._n_boosters_var.trace_add("write",
                                       lambda *_: self._update_booster_frame())
        ttk.Button(nf, text="Measure from image…",
                   command=self._measure_from_image).pack(side=tk.LEFT, padx=(16, 0))

        # Scrollable body: canvas + scrollbar sandwiched between the name row
        # and the Save/Cancel buttons so buttons are always visible.
        scroll_outer = ttk.Frame(self)
        scroll_outer.pack(fill=tk.BOTH, expand=True)

        self._scroll_canvas = tk.Canvas(
            scroll_outer, borderwidth=0, highlightthickness=0, height=500)
        vsb = ttk.Scrollbar(scroll_outer, orient="vertical",
                            command=self._scroll_canvas.yview)
        self._scroll_canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self._scroll_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Inner frame — all middle content lives here
        body = ttk.Frame(self._scroll_canvas)
        body_id = self._scroll_canvas.create_window(
            (0, 0), window=body, anchor="nw")

        def _on_body_configure(_event):
            self._scroll_canvas.configure(
                scrollregion=self._scroll_canvas.bbox("all"))

        def _on_canvas_configure(event):
            self._scroll_canvas.itemconfig(body_id, width=event.width)

        body.bind("<Configure>", _on_body_configure)
        self._scroll_canvas.bind("<Configure>", _on_canvas_configure)

        # Mousewheel scrolling (Mac/Windows uses <MouseWheel>; Linux Button-4/5)
        def _on_mousewheel(event):
            self._scroll_canvas.yview_scroll(
                int(-1 * (event.delta / 120)), "units")

        def _on_mousewheel_linux_up(_event):
            self._scroll_canvas.yview_scroll(-1, "units")

        def _on_mousewheel_linux_down(_event):
            self._scroll_canvas.yview_scroll(1, "units")

        self._scroll_canvas.bind("<MouseWheel>", _on_mousewheel)
        self._scroll_canvas.bind("<Button-4>",   _on_mousewheel_linux_up)
        self._scroll_canvas.bind("<Button-5>",   _on_mousewheel_linux_down)
        body.bind("<MouseWheel>", _on_mousewheel)
        body.bind("<Button-4>",   _on_mousewheel_linux_up)
        body.bind("<Button-5>",   _on_mousewheel_linux_down)

        # ── Front End panel ──────────────────────────────────────────────────
        pl = ttk.LabelFrame(body, text="Front End")
        pl.pack(fill=tk.X, padx=8, pady=4)
        pl.columnconfigure(1, weight=1)

        _ns_labels = list(NOSE_SHAPE_LABELS.values())

        def _fe_entry(parent, label, row, default, unit, pady=2):
            """Helper: label in col 0, entry+unit label in col 1."""
            ttk.Label(parent, text=label).grid(
                row=row, column=0, sticky=tk.W, padx=(6, 2), pady=pady)
            _inner = ttk.Frame(parent)
            _inner.grid(row=row, column=1, sticky=tk.W, padx=(0, 6), pady=pady)
            _var = tk.StringVar(value=default)
            _ent = ttk.Entry(_inner, textvariable=_var, width=10)
            _ent.pack(side=tk.LEFT)
            ttk.Label(_inner, text=unit).pack(side=tk.LEFT, padx=(2, 0))
            return _var, _ent

        # The booster owns only its own front-end hardware: the bus (PBV,
        # carried as dead mass for now) and the fairing.  The loadout — which
        # reentry object and how many — is a run-level choice made in the
        # sidebar; boost carries bus + N × object mass, composed at run time.
        ttk.Label(pl,
                  text="Loadout (reentry object × N) is chosen in the sidebar "
                       "at run time; boost carries bus + N × object mass.",
                  foreground="gray", wraplength=340).grid(
            row=0, column=0, columnspan=2, sticky=tk.W, padx=(6, 2), pady=(4, 2))

        # ── Row 1: Has PBV toggle ─────────────────────────────────────────────
        self._has_pbv_var = tk.BooleanVar(value=False)
        self._has_pbv_check = ttk.Checkbutton(
            pl, text="Has PBV (post-boost bus)",
            variable=self._has_pbv_var,
            command=self._update_pbv_state)
        self._has_pbv_check.grid(
            row=1, column=0, columnspan=2, sticky=tk.W, padx=(6, 2), pady=(4, 0))

        # ── Row 2: PBV sub-section (hidden until checkbox ticked) ────────────
        self._pbv_section = ttk.Frame(pl)
        self._pbv_section.grid(row=2, column=0, columnspan=2,
                               sticky=tk.EW, padx=(16, 0))
        self._pbv_section.columnconfigure(1, weight=1)
        self._pbv_section.grid_remove()

        self._pbv_mass_var, self._pbv_mass_entry = _fe_entry(
            self._pbv_section, "PBV mass (kg):", 0, "0", "kg")
        self._pbv_diameter_var, self._pbv_diameter_entry = _fe_entry(
            self._pbv_section, "PBV diameter (m):", 1, "0", "m")
        self._pbv_length_var, self._pbv_length_entry = _fe_entry(
            self._pbv_section, "PBV length (m):", 2, "0", "m")

        # Legacy alias: bus_var → pbv_mass_var
        self._bus_var = self._pbv_mass_var

        # ── Row 6: Has Fairing toggle ─────────────────────────────────────────
        self._shroud_var = tk.BooleanVar(value=False)
        self._shroud_check = ttk.Checkbutton(
            pl, text="Has Fairing",
            variable=self._shroud_var,
            command=self._update_shroud_state)
        self._shroud_check.grid(
            row=6, column=0, columnspan=2, sticky=tk.W, padx=(6, 2), pady=(4, 0))

        # ── Row 7: Shroud section (hidden until checkbox ticked) ─────────────
        self._shroud_section = ttk.Frame(pl)
        self._shroud_section.grid(row=7, column=0, columnspan=2,
                                  sticky=tk.EW, padx=(16, 0), pady=(0, 4))
        self._shroud_section.columnconfigure(1, weight=1)
        self._shroud_section.grid_remove()

        self._shroud_mass_var, self._shroud_mass_entry = _fe_entry(
            self._shroud_section, "Mass (kg):", 0, "0", "kg")
        ttk.Label(self._shroud_section, text="Shape:").grid(
            row=1, column=0, sticky=tk.W, padx=(6, 2), pady=2)
        self._shroud_nose_shape_var = tk.StringVar(value=NOSE_SHAPE_LABELS["cone"])
        self._shroud_nose_shape_cb = ttk.Combobox(
            self._shroud_section, textvariable=self._shroud_nose_shape_var,
            values=_ns_labels, state="readonly", width=18)
        self._shroud_nose_shape_cb.grid(row=1, column=1, sticky=tk.W, padx=(0, 6), pady=2)
        self._shroud_diameter_var, self._shroud_diameter_entry = _fe_entry(
            self._shroud_section, "Diameter (m):", 2, "0", "m")
        self._shroud_length_var, self._shroud_length_entry = _fe_entry(
            self._shroud_section, "Total fairing length (m):", 3, "0", "m")
        self._shroud_nose_length_var, self._shroud_nose_length_entry = _fe_entry(
            self._shroud_section, "Nose segment length (m):", 4, "0", "m")
        # Jettison altitude is a flight-plan choice, not hardware — it lives in
        # the main window's "Fairing jettison" panel, not this booster editor.

        # ── Aerospike (drag-reduction probe attached to shroud) ─────────────
        # Effect applies only while shroud is attached; it stops at jettison.
        self._aerospike_var = tk.BooleanVar(value=False)
        self._aerospike_check = ttk.Checkbutton(
            self._shroud_section, text="Aerospike",
            variable=self._aerospike_var,
            command=self._update_aerospike_state)
        self._aerospike_check.grid(row=6, column=0, columnspan=2,
                                   sticky=tk.W, padx=(6, 2), pady=(4, 0))

        self._aerospike_section = ttk.Frame(self._shroud_section)
        self._aerospike_section.grid(row=7, column=0, columnspan=2,
                                     sticky=tk.EW, padx=(16, 0))
        self._aerospike_section.columnconfigure(1, weight=1)
        self._aerospike_section.grid_remove()
        self._aerospike_LD_var, self._aerospike_LD_entry = _fe_entry(
            self._aerospike_section, "Spike length (L/D):", 0, "1.5", "")
        self._aerospike_dD_var, self._aerospike_dD_entry = _fe_entry(
            self._aerospike_section, "Aerodisk diameter (d/D):", 1, "0.0", "",
            pady=(2, 4))

        # ── Non-separating vehicle (body reenters) ──────────────────────────
        # A whole-vehicle property: the last stage IS the reentry body
        # (Hwasong-11 / Iskander / Scud / KN-23 class).  When ticked, the
        # sidebar locks the reentry-plan separation to "body" so the separation
        # choice lives with the missile, not the flight plan (the reentry mode
        # still defaults to ballistic and stays switchable).
        ttk.Separator(pl, orient="horizontal").grid(
            row=8, column=0, columnspan=2, sticky=tk.EW, padx=6, pady=(6, 2))
        self._body_reenters_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            pl, text="Non-separating — the body reenters (Scud / KN-23 class)",
            variable=self._body_reenters_var).grid(
            row=9, column=0, columnspan=2, sticky=tk.W, padx=(6, 2), pady=(0, 4))
        ttk.Label(pl, foreground="gray", wraplength=340,
                  text="Locks the reentry plan to non-separating; the last "
                       "stage is the reentry body.").grid(
            row=10, column=0, columnspan=2, sticky=tk.W, padx=(24, 2),
            pady=(0, 4))

        # ── Fins ─────────────────────────────────────────────────────────
        ff = ttk.LabelFrame(body, text="Fins")
        ff.pack(fill=tk.X, padx=8, pady=4)
        ff.columnconfigure(1, weight=1)

        self._fins_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(ff, text="Has fins",
                        variable=self._fins_var,
                        command=self._update_fins_state).grid(
            row=0, column=0, columnspan=2, sticky=tk.W,
            padx=(6, 2), pady=(4, 0))

        self._fins_section = ttk.Frame(ff)
        self._fins_section.grid(row=1, column=0, columnspan=2,
                                sticky=tk.EW, padx=(16, 0))
        self._fins_section.columnconfigure(1, weight=1)
        self._fins_section.grid_remove()

        def _ff_entry(row, label, default, unit, pady=2):
            ttk.Label(self._fins_section, text=label).grid(
                row=row, column=0, sticky=tk.W, padx=(6, 2), pady=pady)
            _inner = ttk.Frame(self._fins_section)
            _inner.grid(row=row, column=1, sticky=tk.W, padx=(0, 6), pady=pady)
            _v = tk.StringVar(value=default)
            ttk.Entry(_inner, textvariable=_v, width=8).pack(side=tk.LEFT)
            ttk.Label(_inner, text=unit).pack(side=tk.LEFT, padx=(2, 0))
            return _v

        self._fin_n_var        = _ff_entry(0, "Number of fins:", "4", "")
        self._fin_span_var     = _ff_entry(1, "Span (exposed, m):", "0", "m")
        self._fin_root_var     = _ff_entry(2, "Root chord (m):", "0", "m")
        self._fin_tip_var      = _ff_entry(3, "Tip chord (m):", "0", "m")
        self._fin_thick_var    = _ff_entry(4, "Thickness (m):", "0", "m")
        self._fin_sweep_var    = _ff_entry(5, "L.E. sweep (°):", "0", "°", pady=(2, 4))

        # ── Grid (lattice) fins ──────────────────────────────────────────
        self._gridfins_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(ff, text="Has grid (lattice) fins",
                        variable=self._gridfins_var,
                        command=self._update_gridfins_state).grid(
            row=2, column=0, columnspan=2, sticky=tk.W,
            padx=(6, 2), pady=(4, 0))

        self._gridfins_section = ttk.Frame(ff)
        self._gridfins_section.grid(row=3, column=0, columnspan=2,
                                    sticky=tk.EW, padx=(16, 0))
        self._gridfins_section.columnconfigure(1, weight=1)
        self._gridfins_section.grid_remove()

        def _gf_entry(row, label, default, unit, pady=2):
            ttk.Label(self._gridfins_section, text=label).grid(
                row=row, column=0, sticky=tk.W, padx=(6, 2), pady=pady)
            _inner = ttk.Frame(self._gridfins_section)
            _inner.grid(row=row, column=1, sticky=tk.W, padx=(0, 6), pady=pady)
            _v = tk.StringVar(value=default)
            ttk.Entry(_inner, textvariable=_v, width=12).pack(side=tk.LEFT)
            ttk.Label(_inner, text=unit).pack(side=tk.LEFT, padx=(2, 0))
            return _v

        self._gfin_n_var      = _gf_entry(0, "Number of grid fins:", "0", "")
        self._gfin_width_var  = _gf_entry(1, "Frame width (m):", "0", "m")
        self._gfin_height_var = _gf_entry(2, "Frame height (m):", "0", "m")
        self._gfin_chord_var  = _gf_entry(3, "Chord / lattice depth (m):", "0", "m")
        # Solidity row carries a "Calculate σ…" helper (σ from web/pitch).
        ttk.Label(self._gridfins_section, text="Solidity σ (0–1):").grid(
            row=4, column=0, sticky=tk.W, padx=(6, 2), pady=2)
        _sol_inner = ttk.Frame(self._gridfins_section)
        _sol_inner.grid(row=4, column=1, sticky=tk.W, padx=(0, 6), pady=2)
        self._gfin_solidity_var = tk.StringVar(value="0")
        ttk.Entry(_sol_inner, textvariable=self._gfin_solidity_var,
                  width=12).pack(side=tk.LEFT)
        ttk.Button(_sol_inner, text="Calculate σ…",
                   command=self._calc_gridfin_solidity).pack(side=tk.LEFT, padx=(6, 0))
        self._gfin_edge_var   = _gf_entry(5, "Edge factor (0.6–1.0):", "1.0", "",
                                          pady=(2, 4))
        # Grid-fin DEPLOY SCHEDULE is a flight-plan choice (when the fins open),
        # not hardware — it lives in the Flight Plan editor, not here.
        ttk.Label(self._gridfins_section,
                  text="Deploy schedule is set in the Flight Plan editor.",
                  foreground="#666").grid(row=6, column=0, columnspan=2,
                                          sticky=tk.W, padx=(6, 6), pady=(0, 4))

        # Guidance mode is flight-plan data, edited in the sidebar / Flight
        # Plan editor -- not here.  The hidden variable only round-trips the
        # in-memory value so _collect keeps working; the plan wins at run time.
        self._guidance_var = tk.StringVar(value="pitch_program")

        # ── Strap-on Boosters panel ─────────────────────────────────────────
        self._booster_frame = ttk.LabelFrame(body, text="Strap-on Boosters")
        self._booster_frame.columnconfigure(1, weight=1)
        # (packed/unpacked dynamically by _update_booster_frame)

        def _be_entry(row, label, default, unit, pady=2):
            ttk.Label(self._booster_frame, text=label).grid(
                row=row, column=0, sticky=tk.W, padx=(6, 2), pady=pady)
            _inner = ttk.Frame(self._booster_frame)
            _inner.grid(row=row, column=1, sticky=tk.W, padx=(0, 6), pady=pady)
            _var = tk.StringVar(value=default)
            ttk.Entry(_inner, textvariable=_var, width=10).pack(side=tk.LEFT)
            ttk.Label(_inner, text=unit).pack(side=tk.LEFT, padx=(2, 0))
            return _var

        # Strap-on JETTISON TIME (booster drop) is a flight-plan choice and
        # lives in the Flight Plan editor; only the motor/mass hardware is here.
        self._b_thrust_var      = _be_entry(0, "Thrust per booster (kN):", "500",  "kN")
        self._b_burn_var        = _be_entry(1, "Burn time (s):",            "60",   "s")
        self._b_core_delay_var  = _be_entry(2, "Core ignition delay (s):", "0",    "s")
        self._b_inert_var       = _be_entry(3, "Inert mass per booster (kg):", "2000", "kg")
        self._b_prop_var        = _be_entry(4, "Propellant per booster (kg):", "10000","kg")
        self._b_isp_var         = _be_entry(5, "Isp (vacuum, s):",          "270",  "s")
        self._b_nozzle_var      = _be_entry(6, "Nozzle exit area (m²):",    "0",    "m²")
        self._b_diam_var        = _be_entry(7, "Diameter (m):",              "1.2",  "m")
        self._b_length_var      = _be_entry(8, "Length (m):",               "0",    "m  (0 = 2×dia)")
        self._b_cd_var          = _be_entry(9, "Cd (drag coeff):",          "0.20", "",
                                            pady=(2, 6))
        ttk.Label(self._booster_frame, text="Cd guide: 0.10 ogive · 0.20 cone · 0.40 hemi · 1.0 flat",
                  foreground="gray50").grid(
            row=11, column=0, columnspan=2, sticky=tk.W, padx=(6, 6), pady=(0, 4))

        # Stage frames (1 always visible; 2-4 toggled).
        # A dedicated container ensures dynamically-packed stages always appear
        # between the payload row and the buttons (not after the buttons).
        self._stages_container = ttk.Frame(body)
        self._stages_container.pack(fill=tk.X)
        self._stage_frames = [_StageFrame(self._stages_container, f"Stage {i+1}",
                                           stage_num=i+1)
                               for i in range(4)]
        self._stage_frames[0].pack(fill=tk.X, **pad)  # Stage 1 always shown

        # Notes / provenance — free text saved with the booster (parity with
        # the RO editor's box): the image tool's provenance stamp appends
        # here, and BoosterParams.notes already round-trips through the JSON.
        notes_frm = ttk.LabelFrame(body, text="Notes / provenance")
        notes_frm.pack(fill=tk.X, **pad)
        self._notes_text = tk.Text(notes_frm, width=44, height=3, wrap=tk.WORD)
        self._notes_text.pack(fill=tk.X, padx=4, pady=4)

        # Buttons — outside the scroll area so always visible
        bf = ttk.Frame(self)
        bf.pack(fill=tk.X, padx=8, pady=(4, 8))
        ttk.Button(bf, text="Cancel", command=self.destroy).pack(
            side=tk.RIGHT, padx=(4, 0))
        self._save_btn = ttk.Button(bf, text="Save Booster", command=self._save)
        self._save_btn.pack(side=tk.RIGHT)
        self._save_as_btn = ttk.Button(bf, text="Save as New Booster",
                                       command=self._save_as_new)
        self._save_as_btn.pack(side=tk.RIGHT, padx=(0, 8))

        # Pre-fill if editing an existing booster
        if existing_name:
            self._prefill(existing_name)
            if self._readonly_mode:
                self._apply_readonly()

    # ------------------------------------------------------------------
    def _apply_readonly(self):
        """Lock all input fields for Forden reference boosters."""
        for sf in self._stage_frames:
            sf.set_readonly(True)
        self._name_entry.config(state="readonly")
        self._stages_cb.config(state="disabled")
        # Front End (PBV)
        self._has_pbv_check.config(state="disabled")
        self._pbv_mass_entry.config(state="disabled")
        self._pbv_diameter_entry.config(state="disabled")
        self._pbv_length_entry.config(state="disabled")
        # Shroud section
        self._shroud_check.config(state="disabled")
        self._shroud_mass_entry.config(state="disabled")
        self._shroud_nose_shape_cb.config(state="disabled")
        self._shroud_diameter_entry.config(state="disabled")
        self._shroud_length_entry.config(state="disabled")
        self._shroud_nose_length_entry.config(state="disabled")
        # Fins section
        self._fins_var.set(False)
        self._update_fins_state()
        # Booster section
        self._n_boosters_spin.config(state="disabled")
        for _bv in (self._b_thrust_var, self._b_burn_var, self._b_core_delay_var,
                    self._b_inert_var, self._b_prop_var, self._b_isp_var,
                    self._b_nozzle_var, self._b_diam_var, self._b_length_var,
                    self._b_cd_var):
            for _w in self._booster_frame.winfo_children():
                if isinstance(_w, ttk.Frame):
                    for _c in _w.winfo_children():
                        if isinstance(_c, ttk.Entry):
                            try:
                                _c.config(state="disabled")
                            except tk.TclError:
                                pass
        self._notes_text.config(state="disabled")
        self._save_btn.pack_forget()
        self._save_as_btn.pack_forget()

    # ------------------------------------------------------------------
    def _update_pbv_state(self):
        """Show/hide PBV sub-section."""
        if self._has_pbv_var.get():
            self._pbv_section.grid()
        else:
            self._pbv_section.grid_remove()

    def _update_shroud_state(self):
        """Show/hide the shroud sub-section."""
        if self._shroud_var.get():
            self._shroud_section.grid()
        else:
            self._shroud_section.grid_remove()

    def _update_aerospike_state(self):
        """Show/hide the aerospike L/D and d/D entries."""
        if self._aerospike_var.get():
            self._aerospike_section.grid()
        else:
            self._aerospike_section.grid_remove()

    def _update_fins_state(self):
        """Show/hide the fin geometry entries."""
        if self._fins_var.get():
            self._fins_section.grid()
        else:
            self._fins_section.grid_remove()

    def _update_gridfins_state(self):
        """Show/hide the grid-fin geometry entries."""
        if self._gridfins_var.get():
            self._gridfins_section.grid()
        else:
            self._gridfins_section.grid_remove()

    def _calc_gridfin_solidity(self):
        """Open a dialog (both inputs explicit) to compute solidity
        σ = 1 − ((p−t)/p)² from the lattice web (wall) thickness t and cell
        pitch p, then fill the σ field.  Styled like Estimate Nozzle Exit Area."""
        dlg = tk.Toplevel(self)
        dlg.title("Calculate σ (solidity)")
        dlg.resizable(False, False)
        dlg.grab_set()

        ttk.Label(dlg, text="Web (wall) thickness  t  (mm):").grid(
            row=0, column=0, sticky=tk.W, padx=(10, 4), pady=(10, 2))
        t_var = tk.StringVar(value="")
        ttk.Entry(dlg, textvariable=t_var, width=12).grid(
            row=0, column=1, sticky=tk.W, padx=(0, 10), pady=(10, 2))

        ttk.Label(dlg, text="Cell pitch  p  (centre-to-centre, mm):").grid(
            row=1, column=0, sticky=tk.W, padx=(10, 4), pady=2)
        p_var = tk.StringVar(value="")
        ttk.Entry(dlg, textvariable=p_var, width=12).grid(
            row=1, column=1, sticky=tk.W, padx=(0, 10), pady=2)

        ttk.Label(dlg, text="σ = 1 − ((p − t) / p)²", foreground="#666").grid(
            row=2, column=0, columnspan=2, sticky=tk.W, padx=10, pady=(8, 0))
        ttk.Label(dlg, text="(σ is a ratio — any consistent unit for t and p works)",
                  foreground="#888").grid(
            row=3, column=0, columnspan=2, sticky=tk.W, padx=10, pady=(0, 2))
        ttk.Label(dlg,
                  text="real fins (σ from published geometry):  open ≈ 0.04–0.06 "
                       "(US MICOM; RU patent)  ·  typical ≈ 0.12 (AA-12-class)  ·  "
                       "dense ≈ 0.22 only with an atypically thick web",
                  foreground="#555").grid(
            row=4, column=0, columnspan=2, sticky=tk.W, padx=10, pady=(0, 2))

        result_var = tk.StringVar(value="Enter t and p.")
        ttk.Label(dlg, textvariable=result_var, foreground="navy").grid(
            row=5, column=0, columnspan=2, padx=10, pady=(4, 4))

        def _compute(*_):
            try:
                t = float(t_var.get())
                p = float(p_var.get())
            except (ValueError, TypeError):
                result_var.set("Enter valid t and p.")
                return None
            if p <= 0.0:
                result_var.set("Cell pitch p must be greater than 0.")
                return None
            if t < 0.0 or t > p:
                result_var.set("Need 0 ≤ t ≤ p.")
                return None
            sigma = mm.grid_fin_solidity(t, p)
            result_var.set(f"σ ≈ {sigma:.3f}")
            return sigma

        t_var.trace_add("write", lambda *_: _compute())
        p_var.trace_add("write", lambda *_: _compute())

        btn_row = ttk.Frame(dlg)
        btn_row.grid(row=6, column=0, columnspan=2, pady=(4, 10))

        def _accept():
            sigma = _compute()
            if sigma is not None:
                self._gfin_solidity_var.set(f"{sigma:.3f}")
                dlg.destroy()

        ttk.Button(btn_row, text="Accept", command=_accept).pack(
            side=tk.LEFT, padx=6)
        ttk.Button(btn_row, text="Cancel",
                   command=dlg.destroy).pack(side=tk.LEFT, padx=6)

        # Centre over parent (matches _suggest_nozzle_area)
        dlg.update_idletasks()
        px = self.winfo_rootx() + (self.winfo_width()  - dlg.winfo_reqwidth())  // 2
        py = self.winfo_rooty() + (self.winfo_height() - dlg.winfo_reqheight()) // 2
        dlg.geometry(f"+{px}+{py}")

    # ------------------------------------------------------------------
    def _update_booster_frame(self, *_):
        """Show or hide the booster parameter panel based on booster count."""
        try:
            n = int(self._n_boosters_var.get())
        except (ValueError, tk.TclError):
            n = 0
        if n > 0:
            self._booster_frame.pack(fill=tk.X, padx=8, pady=4,
                                     before=self._stages_container)
        else:
            self._booster_frame.pack_forget()

    # ------------------------------------------------------------------
    def _measure_from_image(self):
        """Open the shared image-measure dialog for the booster.  Prompts are
        generated from the topology the editor ALREADY declares (stage count,
        fairing on/off, fins on/off + count, strap-on count) — the editor IS
        the topology declaration; the tool reads it."""
        import image_measure as im
        def _i(var, default=0):
            try:
                return int(float(var.get()))
            except (ValueError, tk.TclError, AttributeError):
                return default
        _nst = _i(self._n_stages_var, 1)
        inter = [i + 1 for i, fr in enumerate(self._stage_frames[:_nst])
                 if bool(getattr(fr, "_interstage_var", None)
                         and fr._interstage_var.get())]
        conical = [i + 1 for i, fr in enumerate(self._stage_frames[:_nst])
                   if bool(getattr(fr, "_conical_var", None)
                           and fr._conical_var.get())]
        prompts = im.booster_prompts(
            n_stages=_nst,
            has_fairing=bool(self._shroud_var.get()),
            has_fins=bool(self._fins_var.get()),
            n_fins=_i(self._fin_n_var, 0),
            n_strapons=_i(self._n_boosters_var, 0),
            interstage_stages=inter, conical_stages=conical) \
            + im.booster_angle_prompts(has_fins=bool(self._fins_var.get()))

        # The editor's LOADED geometry rides along so the diagrams draw
        # THIS vehicle's proportions (empty stage 1 → representative art).
        def _fv(var):
            try:
                return float(var.get())
            except (ValueError, tk.TclError, AttributeError):
                return 0.0
        stages = []
        for i, fr in enumerate(self._stage_frames[:_nst], start=1):
            stages.append(dict(
                length_m=_fv(fr._length), diameter_m=_fv(fr._dia),
                top_diameter_m=(_fv(getattr(fr, "_top_dia_var", None))
                                if i in conical else 0.0),
                interstage=(i in inter),
                interstage_len_m=_fv(getattr(fr, "_is_len_var", None))))
        dims = dict(
            stages=stages,
            fairing=(dict(length_m=_fv(self._shroud_length_var),
                          diameter_m=_fv(self._shroud_diameter_var),
                          nose_len_m=_fv(self._shroud_nose_length_var))
                     if self._shroud_var.get() else None),
            fins=(dict(root_m=_fv(self._fin_root_var),
                       span_m=_fv(self._fin_span_var),
                       tip_m=_fv(self._fin_tip_var))
                  if self._fins_var.get() else None),
            strapon=(dict(length_m=_fv(self._b_length_var),
                          diameter_m=_fv(self._b_diam_var))
                     if _i(self._n_boosters_var, 0) > 0 else None),
            nozzle=dict(dia_m=self._nozzle_dia_loaded(1)))
        _open_image_measure_dialog(
            self, "Measure from image — booster", prompts,
            self._apply_image_measurements,
            current_fn=self._current_image_values,
            ctx=dict(dims=dims))

    def _nozzle_dia_loaded(self, stage=1):
        """Stage `stage`'s loaded per-nozzle exit DIAMETER in metres (0 if
        unset), implied by the stored per-nozzle area — or total ÷ count
        when only the total is filled.  Feeds the measure tool's diagram
        proportions and the R8 delta preview (both speak metres)."""
        frames = getattr(self, "_stage_frames", [])
        if not 1 <= stage <= len(frames):
            return 0.0
        fr = frames[stage - 1]
        def _v(name, default=0.0):
            try:
                return float(getattr(fr, name).get() or default)
            except (ValueError, tk.TclError, AttributeError):
                return default
        each = _v("_nozzle_each")
        if each <= 0.0:
            tot, n = _v("_nozzle_area"), _v("_n_nozzles", 1.0)
            if tot > 0.0 and n >= 1.0:
                each = tot / int(n)
        return 2.0 * np.sqrt(each / np.pi) if each > 0.0 else 0.0

    def _img_field_var(self, field):
        """The editor StringVar an image-tool field writes to, or None for a
        field this editor cannot store (the check-only cross-checks).  ONE
        map shared by apply and the R8 delta preview, so they can never
        disagree about what is writable.
        UNITS CONTRACT: the tool writes METRES.  Every var mapped here must
        be a metre-labelled field backed by a *_m model attribute (audited
        2026-08-01) — except fin_sweep, DEGREES, written only by the 3-click
        angle prompt (never through the metre CONVENTIONS).  Other non-metre
        editor fields (masses kg, motor web mm, jettison km) must never be
        added without a conversion at the boundary.  stageN_nozzle_dia is
        therefore deliberately NOT mapped here: the stored field is an AREA
        (m²), so apply converts π(d/2)² at the boundary and the preview
        derives the implied diameter (_nozzle_dia_loaded)."""
        if field.startswith("stage"):
            # "stage2_len" / "stage3_dia" / "stage1_interstage_len" → the
            # StringVar on that stage frame.
            try:
                n = int(field[5])
            except (ValueError, IndexError):
                return None
            frames = getattr(self, "_stage_frames", [])
            if not (1 <= n <= len(frames)):
                return None
            fr = frames[n - 1]
            if field.endswith("_nozzle_dia"):
                return None                      # area field: apply converts
            if field.endswith("_interstage_len"):
                return getattr(fr, "_is_len_var", None)
            if field.endswith("_top_dia"):
                return getattr(fr, "_top_dia_var", None)
            return fr._length if field.endswith("_len") else fr._dia
        return {
            "fairing_len": getattr(self, "_shroud_length_var", None),
            "fairing_dia": getattr(self, "_shroud_diameter_var", None),
            "fairing_nose_len": getattr(self, "_shroud_nose_length_var", None),
            "fin_span": getattr(self, "_fin_span_var", None),
            "fin_root": getattr(self, "_fin_root_var", None),
            "fin_tip": getattr(self, "_fin_tip_var", None),
            "strapon_dia": getattr(self, "_b_diam_var", None),
            "strapon_len": getattr(self, "_b_length_var", None),
            "fin_sweep": getattr(self, "_fin_sweep_var", None),
        }.get(field)

    def _current_image_values(self, fields):
        """{field: current editor value} for the R8 delta preview — only the
        fields this editor can actually write; None = blank/unparseable."""
        out = {}
        for f in fields:
            if f.startswith("stage") and f.endswith("_nozzle_dia"):
                # Stored as per-nozzle AREA; the preview compares diameters.
                try:
                    d = self._nozzle_dia_loaded(int(f[5:].split("_", 1)[0]))
                except ValueError:
                    d = 0.0
                out[f] = d if d > 0.0 else None
                continue
            var = self._img_field_var(f)
            if var is None:
                continue
            try:
                out[f] = float(var.get())
            except (ValueError, tk.TclError):
                out[f] = None
        return out

    def _apply_image_measurements(self, accepted, measurements, scale):
        """Write accepted booster measurements into the editor's fields.  The
        image tool measures ONE of a repeated feature (a fin, a strap-on); the
        model already stores count + one geometry, so the declared count is
        untouched and the single geometry is filled — measure-one-declare-count
        (design).  The provenance stamp is appended to the Notes box and
        saved with the booster (BoosterParams.notes) — parity with the RO
        editor."""
        for field, value in (accepted or {}).items():
            var = self._img_field_var(field)
            if var is not None:
                var.set(f"{float(value):.4g}")
        # Nozzles: the tool measured ONE exit DIAMETER (metres) per stage;
        # the editor stores per-nozzle exit AREA (m²), so the units-contract
        # conversion π(d/2)² happens HERE, at the boundary.  Writing the
        # per-nozzle field fires that stage frame's _sum_nozzles trace,
        # which refreshes its total = each × declared count (count
        # untouched — measure one, declare count).
        for field, value in (accepted or {}).items():
            if not (field.startswith("stage")
                    and field.endswith("_nozzle_dia")):
                continue
            try:
                fr = self._stage_frames[int(field[5:].split("_", 1)[0]) - 1]
                fr._nozzle_each.set(f"{np.pi * (0.5 * float(value)) ** 2:.4g}")
            except (ValueError, IndexError, tk.TclError, AttributeError):
                pass
        # Fin sweep is DERIVED from the fin's root/span/tip (already measured):
        # tan Λ = (root − tip)/span — a length ratio, no angle to get backwards.
        if any(f in (accepted or {}) for f in ("fin_root", "fin_span",
                                               "fin_tip")):
            import image_measure as _im
            try:
                root = float(self._fin_root_var.get() or 0.0)
                span = float(self._fin_span_var.get() or 0.0)
                tip = float(self._fin_tip_var.get() or 0.0)
                sw = _im.sweep_from_planform(root, span, tip)
                if sw is not None:
                    self._fin_sweep_var.set(f"{sw:.4g}")
            except (ValueError, tk.TclError, AttributeError):
                pass
        # A measured interstage length / conical top ⌀ keeps its stage's
        # section on and enabled (the value would otherwise sit in a disabled
        # entry) — measured-it-so-show-it, per stage.
        for field in (accepted or {}):
            if not field.startswith("stage"):
                continue
            try:
                fr = self._stage_frames[int(field[5]) - 1]
            except (ValueError, IndexError):
                continue
            if field.endswith("_interstage_len") \
                    and getattr(fr, "_interstage_var", None) is not None \
                    and not fr._interstage_var.get():
                fr._interstage_var.set(True)
                if hasattr(fr, "_on_interstage"):
                    fr._on_interstage()
            elif field.endswith("_top_dia") \
                    and getattr(fr, "_conical_var", None) is not None \
                    and not fr._conical_var.get():
                fr._conical_var.set(True)
                if hasattr(fr, "_on_conical"):
                    fr._on_conical()
        # Turning on the fairing/fins sections if their geometry was measured
        # keeps what-you-measured visible (the editor owns the count/on flags).
        if any(f.startswith("fairing") for f in (accepted or {})) and \
                hasattr(self, "_shroud_var"):
            self._shroud_var.set(True); self._update_shroud_state()
        if any(f.startswith("fin_") for f in (accepted or {})) and \
                hasattr(self, "_fins_var"):
            self._fins_var.set(True); self._update_fins_state()
        # Provenance stamp → the Notes box (durable, saved with the booster
        # via BoosterParams.notes) — parity with the RO editor.
        if measurements and scale is not None:
            import datetime
            import image_measure as _im
            stamp = _im.provenance_stamp(measurements, scale,
                                         datetime.date.today().isoformat())
            try:
                cur = self._notes_text.get("1.0", "end-1c")
                self._notes_text.insert(
                    "end", ("\n" if cur.strip() else "") + stamp)
            except (AttributeError, tk.TclError):
                pass

    # ------------------------------------------------------------------
    def _update_stage_frames(self):
        """Show the right number of stage frames and coast-time rows."""
        n = int(self._n_stages_var.get())
        pad = dict(padx=8, pady=4)
        for i, sf in enumerate(self._stage_frames):
            if i < n:
                sf.pack(fill=tk.X, **pad)
            else:
                sf.pack_forget()

    # ------------------------------------------------------------------
    def _prefill(self, name):
        """Populate all fields from an existing booster (custom or packaged)."""
        # get_booster (not the raw hardware) so the dialog shows the booster's
        # real deployment timing, which lives in the flight plan; otherwise a
        # re-edit would display defaults and clobber the saved timing on save.
        p = get_booster(name)

        if getattr(p, "notes", ""):
            self._notes_text.delete("1.0", "end")
            self._notes_text.insert("1.0", p.notes)

        payload      = p.payload_kg
        shroud_mass  = p.shroud_mass_kg

        # Walk the linked list to collect per-stage data.
        # mass_initial is cumulative (includes all upper stages); we recover
        # per-stage fueled mass by differencing adjacent mass_initial values
        # and stripping payload/shroud from the appropriate stages.
        # dry is always fueled - mass_propellant: mass_propellant is per-stage
        # and reliable even in boosters loaded from older JSON files, so this
        # avoids any dependency on the (potentially corrupt) mass_final field.
        stage_data = []
        node = p
        stage_idx = 0
        while node is not None:
            nxt      = node.stage2
            is_first = (stage_idx == 0)
            is_last  = (nxt is None)
            if is_last and is_first:
                # Single-stage booster
                fueled = node.mass_initial - payload - shroud_mass
            elif is_last:
                # Last of multiple stages: no shroud here (lives on stage 1)
                fueled = node.mass_initial - payload
            elif is_first:
                # First of multiple stages: subtract shroud and upper stack
                fueled = node.mass_initial - shroud_mass - nxt.mass_initial
            else:
                # Middle stage
                fueled = node.mass_initial - nxt.mass_initial
            # Per-stage dry mass is always: fueled - own propellant.
            # Do NOT use node.mass_final here — it may be a cumulative burnout
            # mass (includes upper stages) if the model was loaded from an
            # older serialised file.
            dry = fueled - node.mass_propellant
            stage_data.append({
                "fueled":        fueled,                   "dry":          dry,
                "dia":           node.diameter_m,          "length":       node.length_m,
                "burn":          node.burn_time_s,         "isp":          node.isp_s,
                "nozzle_area":   node.nozzle_exit_area_m2, "coast":        node.coast_time_s,
                "n_nozzles":     getattr(node, 'n_nozzles', 1),
                "nozzle_each":   getattr(node, 'nozzle_area_each_m2', 0.0),
                "solid_motor":   getattr(node, 'solid_motor', False),
                "grain_type":    getattr(node, 'grain_type', ''),
                "thrust_peak_N": getattr(node, 'thrust_peak_N', 0.0),
                "thrust_profile": list(getattr(node, 'thrust_profile', [])),
                "conical":        getattr(node, 'conical', False),
                "top_diameter_m": getattr(node, 'top_diameter_m', 0.0),
                "has_interstage": getattr(node, 'has_interstage', False),
                "interstage_length_m": getattr(node, 'interstage_length_m', 0.0),
                "interstage_mass_kg":  getattr(node, 'interstage_mass_kg', 0.0),
                "interstage_jettison_s": getattr(node, 'interstage_jettison_s', None),
            })
            node = nxt
            stage_idx += 1

        n = len(stage_data)
        self._n_stages_var.set(str(n))
        self._update_stage_frames()
        for i, sd in enumerate(stage_data):
            self._stage_frames[i].populate(sd)

        # PBV (bus) — the only front-end mass the booster owns; the loadout
        # (object × N) is composed at run time from the sidebar selection.
        has_pbv = p.bus_mass_kg > 0
        self._has_pbv_var.set(has_pbv)
        self._pbv_mass_var.set(f"{p.bus_mass_kg:.0f}" if has_pbv else "0")
        self._body_reenters_var.set(bool(getattr(p, 'body_reenters', False)))
        self._pbv_diameter_var.set(f"{getattr(p, 'pbv_diameter_m', 0.0):.2f}")
        self._pbv_length_var.set(f"{getattr(p, 'pbv_length_m', 0.0):.2f}")

        # Aerospike
        _aero_LD = float(getattr(p, 'aerospike_LD', 0.0) or 0.0)
        _aero_dD = float(getattr(p, 'aerospike_dD', 0.0) or 0.0)
        self._aerospike_var.set(_aero_LD > 0.0)
        self._aerospike_LD_var.set(f"{_aero_LD:.2f}" if _aero_LD > 0 else "1.5")
        self._aerospike_dD_var.set(f"{_aero_dD:.2f}")
        self._update_aerospike_state()

        # Fins
        _has_fins = bool(getattr(p, 'has_fins', False))
        self._fins_var.set(_has_fins)
        self._fin_n_var.set(str(int(getattr(p, 'n_fins', 4) or 4)))
        self._fin_span_var.set(f"{float(getattr(p, 'fin_span_m', 0.0)):.3f}")
        self._fin_root_var.set(f"{float(getattr(p, 'fin_root_chord_m', 0.0)):.3f}")
        self._fin_tip_var.set(f"{float(getattr(p, 'fin_tip_chord_m', 0.0)):.3f}")
        self._fin_thick_var.set(f"{float(getattr(p, 'fin_thickness_m', 0.0)):.4f}")
        self._fin_sweep_var.set(f"{float(getattr(p, 'fin_sweep_deg', 0.0)):.1f}")
        self._update_fins_state()

        # Grid (lattice) fins
        self._gridfins_var.set(bool(getattr(p, 'has_grid_fins', False)))
        self._gfin_n_var.set(str(int(getattr(p, 'n_grid_fins', 0) or 0)))
        self._gfin_width_var.set(f"{float(getattr(p, 'grid_fin_width_m', 0.0)):.3f}")
        self._gfin_height_var.set(f"{float(getattr(p, 'grid_fin_height_m', 0.0)):.3f}")
        self._gfin_chord_var.set(f"{float(getattr(p, 'grid_fin_chord_m', 0.0)):.3f}")
        self._gfin_solidity_var.set(f"{float(getattr(p, 'grid_fin_solidity', 0.0)):.3f}")
        self._gfin_edge_var.set(f"{float(getattr(p, 'grid_fin_edge_factor', 1.0) or 1.0):.2f}")
        self._update_gridfins_state()

        # Shroud
        has_shroud = shroud_mass > 0
        self._shroud_var.set(has_shroud)
        self._shroud_mass_var.set(f"{shroud_mass:.0f}")
        self._shroud_length_var.set(f"{p.shroud_length_m:.1f}")
        self._shroud_diameter_var.set(f"{p.shroud_diameter_m:.2f}")
        self._shroud_nose_shape_var.set(
            NOSE_SHAPE_LABELS.get(p.shroud_nose_shape, NOSE_SHAPE_LABELS["cone"]))
        self._shroud_nose_length_var.set(f"{p.shroud_nose_length_m:.2f}")

        # Strap-on boosters
        nb = getattr(p, 'n_boosters', 0)
        self._n_boosters_var.set(str(nb))
        if nb > 0:
            G0 = 9.80665
            b_prop = getattr(p, 'booster_prop_kg', 0.0)
            b_burn = getattr(p, 'booster_burn_time_s', 0.0)
            b_isp  = getattr(p, 'booster_isp_s', 0.0)
            b_thrust_kn = (b_isp * G0 * b_prop / b_burn / 1000.0
                           if b_burn > 0 and b_prop > 0 and b_isp > 0
                           else getattr(p, 'booster_thrust_n', 0.0) / 1000.0)
            self._b_thrust_var.set(f"{b_thrust_kn:.1f}")
            self._b_burn_var.set(f"{b_burn:.1f}")
            self._b_inert_var.set(f"{getattr(p, 'booster_inert_kg', 0.0):.0f}")
            self._b_prop_var.set(f"{b_prop:.0f}")
            self._b_isp_var.set(f"{b_isp:.1f}")
            self._b_nozzle_var.set(f"{getattr(p, 'booster_nozzle_area_m2', 0.0):.4f}")
            self._b_diam_var.set(f"{getattr(p, 'booster_diam_m', 0.0):.2f}")
            self._b_length_var.set(f"{getattr(p, 'booster_length_m', 0.0):.2f}")
            self._b_cd_var.set(f"{getattr(p, 'booster_cd', 0.20):.2f}")
            self._b_core_delay_var.set(f"{getattr(p, 'booster_core_delay_s', 0.0):.1f}")
        self._update_booster_frame()

        # Apply show/hide state for all sections
        self._update_pbv_state()
        self._update_shroud_state()

        self._name_var.set(name)
        self._guidance_var.set(p.guidance)

    # ------------------------------------------------------------------
    def _collect(self) -> 'BoosterParams':
        """Read and validate all fields; return a BoosterParams linked list."""
        from booster_models import BoosterParams, _FORDEN_MACH, _FORDEN_CD

        name = self._name_var.get().strip()
        if not name:
            raise ValueError("Booster name cannot be blank.")

        n = int(self._n_stages_var.get())

        # Front-end hardware the booster owns: the bus (PBV) only.  The
        # loadout — which reentry object and how many — is a run-level
        # sidebar choice composed onto the chain by compose_loadout(), so
        # the stage masses are built STACK-ONLY (payload = 0, never baked
        # in) and the booster no longer references the reentry object.
        try:
            bus_mass = float(self._pbv_mass_var.get()) if self._has_pbv_var.get() else 0.0
        except ValueError:
            raise ValueError("PBV mass must be a number.")
        payload  = 0.0
        num_ros  = 1
        ro_mass  = 0.0
        ro_separates = True   # build-era record: stack-only masses (mass_final = dry)

        # Front-end ascent shape now follows the run-level object (or the
        # fairing while attached) via _boost_front_geometry; the booster
        # keeps no hand-entered payload shape of its own.
        nose_shape         = ""
        payload_diameter_m = 0.0
        nose_length_m      = 0.0

        # Aerospike (drag-reduction probe)
        if self._aerospike_var.get():
            try:
                aerospike_LD = max(0.0, float(self._aerospike_LD_var.get()))
                aerospike_dD = max(0.0, float(self._aerospike_dD_var.get()))
            except ValueError:
                raise ValueError("Aerospike L/D and d/D must be numbers.")
        else:
            aerospike_LD = 0.0
            aerospike_dD = 0.0

        # Fins
        has_fins = bool(self._fins_var.get())
        if has_fins:
            try:
                n_fins         = max(1, int(float(self._fin_n_var.get())))
                fin_span_m     = max(0.0, float(self._fin_span_var.get()))
                fin_root_m     = max(0.0, float(self._fin_root_var.get()))
                fin_tip_m      = max(0.0, float(self._fin_tip_var.get()))
                fin_thick_m    = max(0.0, float(self._fin_thick_var.get()))
                fin_sweep_deg  = float(self._fin_sweep_var.get())
            except ValueError:
                raise ValueError("Fin dimensions must be numbers.")
        else:
            n_fins = 4; fin_span_m = 0.0; fin_root_m = 0.0
            fin_tip_m = 0.0; fin_thick_m = 0.0; fin_sweep_deg = 0.0

        # Grid (lattice) fins
        has_grid_fins = bool(self._gridfins_var.get())
        if has_grid_fins:
            try:
                n_grid_fins      = max(1, int(float(self._gfin_n_var.get())))
                gfin_width_m     = max(0.0, float(self._gfin_width_var.get()))
                gfin_height_m    = max(0.0, float(self._gfin_height_var.get()))
                gfin_chord_m     = max(0.0, float(self._gfin_chord_var.get()))
                gfin_solidity    = max(0.0, min(1.0, float(self._gfin_solidity_var.get())))
                gfin_edge_factor = float(self._gfin_edge_var.get())
            except ValueError:
                raise ValueError("Grid-fin dimensions must be numbers.")
        else:
            n_grid_fins = 0; gfin_width_m = 0.0; gfin_height_m = 0.0
            gfin_chord_m = 0.0; gfin_solidity = 0.0; gfin_edge_factor = 1.0

        # PBV geometry
        try:
            pbv_diameter_m = (float(self._pbv_diameter_var.get())
                              if self._has_pbv_var.get() else 0.0)
            pbv_length_m   = (float(self._pbv_length_var.get())
                              if self._has_pbv_var.get() else 0.0)
        except ValueError:
            raise ValueError("PBV diameter and length must be numbers.")

        # Shroud (hardware only; jettison timing is a flight-plan field)
        shroud_mass          = 0.0
        shroud_length_m      = 0.0
        shroud_diameter_m    = 0.0
        shroud_nose_shape    = ""
        shroud_nose_length_m = 0.0
        if self._shroud_var.get():
            try:
                shroud_mass          = float(self._shroud_mass_var.get())
                shroud_length_m      = float(self._shroud_length_var.get())
                shroud_diameter_m    = float(self._shroud_diameter_var.get())
                _snl = self._shroud_nose_length_var.get().strip()
                shroud_nose_length_m = float(_snl) if _snl and float(_snl) > 0 \
                                       else shroud_length_m
            except ValueError:
                raise ValueError("Fairing fields must be numbers.")
            _slabel = self._shroud_nose_shape_var.get()
            shroud_nose_shape = next(
                (k for k, v in NOSE_SHAPE_LABELS.items() if v == _slabel), "")

        # Read and validate all active stage frames
        stages = []
        for i in range(n):
            sd = self._stage_frames[i].get()
            if sd["fueled"] <= sd["dry"]:
                raise ValueError(
                    f"Stage {i+1}: fueled mass must exceed dry mass.")
            stages.append(sd)

        # Build the linked list from the last stage back to the first.
        # Shroud lives on the first (bottom) stage; payload is part of the
        # last (top) stage's mass until final burnout.
        node = None
        upper_mass = 0.0
        for idx, sd in enumerate(reversed(stages)):
            stage_num = n - idx
            is_last  = (idx == 0)        # last stage of booster (first in reversed loop)
            is_first = (idx == n - 1)    # first stage of booster (last in reversed loop)
            prop = sd["fueled"] - sd["dry"]
            if is_last and is_first:
                # Single-stage booster (stack-only: loadout composed at run)
                m0     = sd["fueled"] + shroud_mass
                mfinal = sd["dry"]
            elif is_last:
                # Last of multiple stages: shroud is on stage 1
                m0     = sd["fueled"]
                mfinal = sd["dry"]
            elif is_first:
                # First of multiple stages: add shroud here
                m0     = sd["fueled"] + shroud_mass + upper_mass
                mfinal = sd["dry"]
            else:
                # Middle stage
                m0     = sd["fueled"] + upper_mass
                mfinal = sd["dry"]
            upper_mass = m0
            node = BoosterParams(
                name=f"{name} Stage {stage_num}",
                mass_initial=m0,
                mass_propellant=prop,
                mass_final=mfinal,
                diameter_m=sd["dia"],  length_m=sd["length"],
                thrust_N=round(sd["thrust_kn"] * 1000.0),
                burn_time_s=sd["burn"], isp_s=sd["isp"],
                coast_time_s=sd["coast"] if not is_last else 0.0,
                nozzle_exit_area_m2=sd["nozzle_area"],
                n_nozzles=sd.get("n_nozzles", 1),
                nozzle_area_each_m2=sd.get("nozzle_each", 0.0),
                mach_table=list(_FORDEN_MACH), cd_table=list(_FORDEN_CD),
                stage2=node,
                solid_motor=bool(sd.get("solid_motor", False)),
                grain_type=sd.get("grain_type", ""),
                thrust_peak_N=float(sd.get("thrust_peak_N", 0.0)),
                thrust_profile=list(sd.get("thrust_profile", [])),
                conical=bool(sd.get("conical", False)),
                top_diameter_m=float(sd.get("top_diameter_m", 0.0)),
                has_interstage=bool(sd.get("has_interstage", False)),
                interstage_length_m=float(sd.get("interstage_length_m", 0.0)),
                interstage_mass_kg=float(sd.get("interstage_mass_kg", 0.0)),
                interstage_jettison_s=sd.get("interstage_jettison_s", None),
            )

        # Saved boosters no longer embed an RV — RV identity lives in the
        # library and is injected at run time from the sidebar selection.
        node.name                   = name
        node.guidance               = self._guidance_var.get()
        node.payload_kg             = payload
        node.ro                     = None
        node.bus_mass_kg            = bus_mass
        node.num_ros                = num_ros
        node.ro_mass_kg             = ro_mass
        node.ro_separates           = ro_separates
        node.body_reenters          = bool(self._body_reenters_var.get())
        node.nose_shape             = nose_shape
        node.nose_length_m          = nose_length_m
        node.payload_diameter_m     = payload_diameter_m
        node.pbv_diameter_m         = pbv_diameter_m
        node.pbv_length_m           = pbv_length_m
        node.shroud_mass_kg         = shroud_mass
        node.shroud_length_m        = shroud_length_m
        node.shroud_diameter_m      = shroud_diameter_m
        node.shroud_nose_shape      = shroud_nose_shape
        node.shroud_nose_length_m   = shroud_nose_length_m
        node.aerospike_LD           = aerospike_LD
        node.aerospike_dD           = aerospike_dD
        node.has_fins               = has_fins
        node.n_fins                 = n_fins
        node.fin_span_m             = fin_span_m
        node.fin_root_chord_m       = fin_root_m
        node.fin_tip_chord_m        = fin_tip_m
        node.fin_thickness_m        = fin_thick_m
        node.fin_sweep_deg          = fin_sweep_deg
        node.has_grid_fins          = has_grid_fins
        node.n_grid_fins            = n_grid_fins
        node.grid_fin_width_m       = gfin_width_m
        node.grid_fin_height_m      = gfin_height_m
        node.grid_fin_chord_m       = gfin_chord_m
        node.grid_fin_solidity      = gfin_solidity
        node.grid_fin_edge_factor   = gfin_edge_factor
        # grid_fin_deploy_schedule is flight-plan data (owned by the Flight Plan
        # editor); left at its default here so a booster save never clobbers it.
        node.notes = self._notes_text.get("1.0", "end-1c").strip()

        # Strap-on boosters
        try:
            _n_b = max(0, min(9, int(self._n_boosters_var.get())))
        except (ValueError, tk.TclError):
            _n_b = 0
        if _n_b > 0:
            try:
                _b_thrust_kn   = float(self._b_thrust_var.get())
                _b_burn        = float(self._b_burn_var.get())
                _b_core_delay  = float(self._b_core_delay_var.get())
                _b_inert       = float(self._b_inert_var.get())
                _b_prop        = float(self._b_prop_var.get())
                _b_isp         = float(self._b_isp_var.get())
                _b_nozzle      = float(self._b_nozzle_var.get())
                _b_diam        = float(self._b_diam_var.get())
                _b_length      = float(self._b_length_var.get())
                _b_cd          = float(self._b_cd_var.get())
            except ValueError as exc:
                raise ValueError(f"Booster field: {exc}") from exc
            if _b_burn <= 0:
                raise ValueError("Booster burn time must be > 0.")
            if _b_diam <= 0:
                raise ValueError("Booster diameter must be > 0.")
            node.n_boosters             = _n_b
            node.booster_thrust_n       = _b_thrust_kn * 1000.0
            node.booster_burn_time_s    = _b_burn
            node.booster_core_delay_s   = max(0.0, _b_core_delay)
            # booster_jettison_s is flight-plan data (owned by the Flight Plan
            # editor); left at its default here so a save never clobbers it.
            node.booster_inert_kg       = _b_inert
            node.booster_prop_kg        = _b_prop
            node.booster_isp_s          = _b_isp
            node.booster_nozzle_area_m2 = _b_nozzle
            node.booster_diam_m         = _b_diam
            node.booster_length_m       = _b_length
            node.booster_cd             = _b_cd

        return node

    # ------------------------------------------------------------------
    def _save(self):
        try:
            p = self._collect()
        except ValueError as e:
            messagebox.showerror("Invalid input", str(e), parent=self)
            return
        self._on_save(p)
        self.destroy()

    def _save_as_new(self):
        try:
            p = self._collect()
        except ValueError as e:
            messagebox.showerror("Invalid input", str(e), parent=self)
            return
        new_name = simpledialog.askstring(
            "Save as New Booster",
            "Enter a name for the new booster:",
            initialvalue=p.name,
            parent=self)
        if not new_name or not new_name.strip():
            return
        new_name = new_name.strip()
        if new_name in BOOSTER_DB:
            if not messagebox.askyesno(
                    "Overwrite?",
                    f"A booster named '{new_name}' already exists. Overwrite it?",
                    parent=self):
                return
        p.name = new_name
        self._on_save(p)
        self.destroy()


# ---------------------------------------------------------------------------
# RV editor dialog
# ---------------------------------------------------------------------------

# Integration-family display names.  The reentry-mode dropdown is FAMILY-
# SCOPED (it lists only the active plan's family — numerical EOM or closed-form
# analytic — so the family is the plan's identity and cannot be crossed from
# the strip; New Reentry Plan chooses the family).  See REENTRY_FAMILY_DESIGN.md.
_FAMILY_LABELS = {'numerical': "numerical (EOM)",
                  'analytic':  "closed-form analytic"}


def _open_image_measure_dialog(parent, title, prompts, apply_fn,
                               current_fn=None, ctx=None):
    """Shared image-dimensioning dialog (Phase A) for the RO and booster
    editors.  `prompts` is a list of {field,label,view,convention} (from
    image_measure); `apply_fn(accepted, measurements, scale)` writes the
    accepted {field:value_m} into the parent's fields and stamps its notes.

    The image PROPOSES; the human Accepts each value; Apply commits.  The
    image, scale, and clicks are session-only (no persistence) — only the
    accepted values and the stamp survive.  image_measure.py is the tested
    core; this is the shell (canvas + scale + prompt loop).
    """
    try:
        from PIL import Image, ImageTk
    except ImportError:
        messagebox.showerror(
            "Pillow required",
            "Measuring from an image needs the Pillow library.\n"
            "Install it with:  pip install pillow", parent=parent)
        return
    import image_measure as im

    dlg = tk.Toplevel(parent)
    dlg.title(title)
    dlg.grab_set()

    # ── Views (Phase B) ────────────────────────────────────────────────
    # Each named view (side / plan) carries its OWN image and its OWN scale —
    # two figures in a paper are almost never at the same resolution, so a
    # side-view m/px is meaningless on the plan view.  Single-view checklists
    # (booster, axisymmetric RO) get one "side" slot and no selector.
    views_needed = sorted({p["view"] for p in prompts}) or ["side"]
    multiview = len(views_needed) > 1
    _VIEW_LABELS = {"side": "Side view", "plan": "Plan (top) view"}

    def _blank_view():
        return dict(img=None, photo=None, zoom=1.0, scale=None,
                    anchor_total=None)

    views = {v: _blank_view() for v in views_needed}
    state = dict(cur=("side" if "side" in views else views_needed[0]),
                 mode="idle", clicks=[],       # clicks in ORIGINAL-image px
                 prompt=None, accepted={}, measurements=[],
                 overlay=True, annotations={})  # field → (view, p1, p2, label)

    def _cv():
        return views[state["cur"]]

    body = ttk.Frame(dlg, padding=8); body.pack(fill=tk.BOTH, expand=True)
    canvas = tk.Canvas(body, width=760, height=560, bg="#222",
                       highlightthickness=1, highlightbackground="#888")
    canvas.grid(row=0, column=0, rowspan=2, sticky="nsew")
    panel = ttk.Frame(body, padding=(10, 0, 0, 0))
    panel.grid(row=0, column=1, sticky="n")

    status = tk.StringVar(value="Load an image to begin.")
    ttk.Label(panel, textvariable=status, foreground="#367",
              wraplength=250, justify=tk.LEFT).pack(anchor=tk.W, pady=(0, 6))

    _MAXW, _MAXH = 760, 560

    def _fit_zoom(img):
        ow, oh = img.size
        return min(_MAXW / ow, _MAXH / oh, 1.0)

    def _render():
        """Redraw the current view at its zoom: image, the accepted-
        measurement overlay (audit of what was clicked, per view), and any
        in-progress click marks.  All geometry lives in original-image px;
        zoom is display-only, so measurements are zoom-independent."""
        canvas.delete("all")
        v = _cv()
        if v["img"] is None:
            return
        z = v["zoom"]
        ow, oh = v["img"].size
        disp = v["img"].resize((max(1, int(ow * z)), max(1, int(oh * z))))
        v["photo"] = ImageTk.PhotoImage(disp)
        canvas.create_image(0, 0, anchor="nw", image=v["photo"])
        canvas.configure(scrollregion=(0, 0, int(ow * z), int(oh * z)))
        if state["overlay"]:
            # A field being RE-measured hides its old accepted mark, so the
            # fresh clicks can't be confused with the mark they're replacing
            # (field report: Measure "doesn't kill the older measure" — the
            # VALUE stays recorded until the new reading is Accepted, but
            # the old overlay must get out of the way while re-clicking).
            redo = ((state.get("prompt") or {}).get("field")
                    if (state["mode"] == "measure"
                        or state.get("_pending") is not None) else None)
            for i, (_f, ann) in enumerate(state["annotations"].items()):
                if _f == redo:
                    continue
                vw, p1, p2, lab = ann[0], ann[1], ann[2], ann[3]
                vtx = ann[4] if len(ann) > 4 else None
                if vw != state["cur"]:
                    continue
                x0, y0, x1, y1 = p1[0] * z, p1[1] * z, p2[0] * z, p2[1] * z
                if vtx is None:                      # linear: one segment
                    canvas.create_line(x0, y0, x1, y1, fill="#3c6", width=1,
                                       tags="ann")
                    corners = ((x0, y0), (x1, y1))
                else:                                # angle: two rays + vertex
                    vx, vy = vtx[0] * z, vtx[1] * z
                    canvas.create_line(vx, vy, x0, y0, fill="#3c6", width=1,
                                       tags="ann")
                    canvas.create_line(vx, vy, x1, y1, fill="#3c6", width=1,
                                       tags="ann")
                    corners = ((vx, vy), (x0, y0), (x1, y1))
                for xx, yy in corners:
                    canvas.create_oval(xx - 2, yy - 2, xx + 2, yy + 2,
                                       outline="#3c6", tags="ann")
                # stagger the label rows so co-located segments (a chord and a
                # sweep measured at the same fin) stay readable
                canvas.create_text((x0 + x1) / 2,
                                   (y0 + y1) / 2 - 9 - 12 * (i % 3), text=lab,
                                   fill="#3c6", font=("TkDefaultFont", 8),
                                   tags="ann")
        for ix, iy in state["clicks"]:
            x, y = ix * z, iy * z
            canvas.create_oval(x - 3, y - 3, x + 3, y + 3, outline="#ff4",
                               width=2, tags="mark")
        pts = [(ix * z, iy * z) for ix, iy in state["clicks"]]
        if len(pts) == 2:
            canvas.create_line(*pts[0], *pts[1], fill="#ff4", width=1,
                               tags="mark")
        elif len(pts) == 3:                    # angle: vertex first, two rays
            canvas.create_line(*pts[0], *pts[1], fill="#ff4", width=1,
                               tags="mark")
            canvas.create_line(*pts[0], *pts[2], fill="#ff4", width=1,
                               tags="mark")

    def _sync_quantum():
        s = _cv()["scale"]
        vtag = f"{state['cur']} " if multiview else ""
        quantum.set(f"{vtag}scale: " + (s.quantum_str() if s else "—"))

    def _set_image(img):
        """Display a new source image IN THE CURRENT VIEW.  That view's scale
        always resets with it — metres-per-pixel belongs to the image it was
        anchored on; carrying it to a different picture would be silently
        wrong.  Other views' images and scales are untouched."""
        _cv().update(img=img, zoom=_fit_zoom(img), scale=None,
                     anchor_total=None)
        state.update(mode="idle", clicks=[])
        _sync_quantum()
        _render()
        canvas.xview_moveto(0); canvas.yview_moveto(0)
        ow, oh = img.size
        vtag = f" into the {state['cur'].upper()} view" if multiview else ""
        status.set(f"Loaded {ow}×{oh}px{vtag}.  Set the scale next "
                   "(click two points a known distance apart).")
        _refresh_closure()

    def _load_path(path):
        try:
            img = Image.open(path)
        except Exception as e:
            messagebox.showerror("Cannot open image", str(e), parent=dlg)
            return
        _set_image(img)

    def _load():
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            parent=dlg, title="Load vehicle image",
            filetypes=[("Images",
                        "*.png *.jpg *.jpeg *.webp *.gif *.bmp *.tif *.tiff"),
                       ("All files", "*.*")])
        if path:
            _load_path(path)

    def _paste(*_ev):
        """Clipboard image (⌘V / Ctrl-V) — Pillow's grabclipboard returns an
        image (screenshot), a list of file paths (copied files), or None."""
        try:
            from PIL import ImageGrab
            clip = ImageGrab.grabclipboard()
        except Exception as e:
            messagebox.showerror(
                "Paste", f"Clipboard image unavailable here: {e}", parent=dlg)
            return "break"
        if isinstance(clip, list):
            if clip:
                _load_path(clip[0])
        elif clip is not None:
            _set_image(clip)
        else:
            status.set("No image on the clipboard.")
        return "break"

    dlg.bind("<Command-v>", _paste)
    dlg.bind("<Control-v>", _paste)

    # Per-click prompts so the user always knows a click landed and how many
    # remain (an angle's THREE clicks were being confused with a 2-click line,
    # and a stale result read as the current one).
    _ANGLE_STEPS = ("click the vertex — the wing-root LE corner",
                    "vertex set ✓ — now click a point ALONG the leading edge",
                    "✓ — now click a point OUT spanwise (perpendicular to the "
                    "body axis, NOT down the body)")
    _LINE_STEPS = ("click the first point", "✓ — click the second point")
    _SCALE_STEPS = ("SCALE: click the first point",
                    "✓ — click the second point a known distance away")

    def _click_progress():
        n = len(state["clicks"])
        if state["mode"] == "scale":
            steps = _SCALE_STEPS
        elif (state.get("prompt") or {}).get("angle"):
            steps = _ANGLE_STEPS
        else:
            steps = _LINE_STEPS
        if n < len(steps):
            status.set(f"[{n}/{len(steps)}]  {steps[n]}")

    def _on_click(ev):
        v = _cv()
        if state["mode"] not in ("scale", "measure") or v["img"] is None:
            return
        z = v["zoom"]
        state["clicks"].append((canvas.canvasx(ev.x) / z,
                                canvas.canvasy(ev.y) / z))
        _render()
        # angles take THREE clicks (vertex + two rays); everything else two
        need = (3 if (state["mode"] == "measure"
                      and (state.get("prompt") or {}).get("angle")) else 2)
        if len(state["clicks"]) == need:
            (state["_finish_scale"] if state["mode"] == "scale"
             else state["_finish_measure"])()
        else:
            _click_progress()

    canvas.bind("<Button-1>", _on_click)

    # Zoom (wheel, about the cursor) + pan (right- or middle-drag).  Clicks
    # are stored in original-image px, so zoom never touches a measurement.
    def _zoom_at(factor, wx=None, wy=None):
        v = _cv()
        if v["img"] is None:
            return
        z0 = v["zoom"]
        z1 = min(8.0, max(0.05, z0 * float(factor)))
        if z1 == z0:
            return
        if wx is None:
            wx, wy = canvas.winfo_width() / 2, canvas.winfo_height() / 2
        ix, iy = canvas.canvasx(wx) / z0, canvas.canvasy(wy) / z0
        v["zoom"] = z1
        _render()
        ow, oh = v["img"].size
        canvas.xview_moveto(max(0.0, (ix * z1 - wx) / (ow * z1)))
        canvas.yview_moveto(max(0.0, (iy * z1 - wy) / (oh * z1)))

    def _fit(*_ev):
        v = _cv()
        if v["img"] is None:
            return
        v["zoom"] = _fit_zoom(v["img"])
        _render()
        canvas.xview_moveto(0); canvas.yview_moveto(0)

    # Wheel routing (Mac convention): plain scroll PANS — that's a trackpad's
    # two-finger drag — and ⌘/Ctrl-scroll ZOOMS about the cursor, the
    # Preview/maps idiom.  A trackpad delivers a rapid stream of small-delta
    # events plus a momentum tail after the fingers lift (some events with
    # delta 0, which the old ">0 else out" test read as ZOOM OUT), so a
    # per-event 1.15× slammed the zoom limits on one swipe; steps are now
    # normalized (Windows ±120/notch, Mac trackpad ±1-ish) and clamped, and
    # delta-0 events are dropped.  Pinch never reaches Tk at all — the
    # buttons/keys below are the guaranteed path.
    def _wheel_steps(delta):
        d = float(delta)
        if d == 0.0:
            return 0.0                       # momentum tail / stray event
        s = d / 120.0 if abs(d) >= 120.0 else d
        return max(-2.0, min(2.0, s))

    def _pan_by(dx_px, dy_px):
        if _cv()["img"] is None:
            return
        if dx_px:
            canvas.xview_scroll(int(dx_px), "units")
        if dy_px:
            canvas.yview_scroll(int(dy_px), "units")

    def _wheel_zoom(ev, cx=None, cy=None):
        s = _wheel_steps(getattr(ev, "delta", 0))
        if s:
            _zoom_at(1.15 ** s,
                     ev.x if cx is None else cx,
                     ev.y if cy is None else cy)
        return "break"

    def _wheel_pan(ev, horizontal=False):
        d = getattr(ev, "delta", 0)
        if not d:
            return "break"
        px = int(d if abs(d) >= 40 else d * 30)      # notches → pixels
        _pan_by(-px if horizontal else 0, 0 if horizontal else -px)
        return "break"

    # Aqua quirk: some macOS Tk builds deliver wheel events to the FOCUSED
    # widget, not the one under the pointer — and a canvas never takes
    # keyboard focus, so its wheel bindings simply never fire ("scroll does
    # nothing" on a Mac trackpad).  The canvas bindings below return "break"
    # when they DO fire, which stops the bindtag chain before this toplevel
    # fallback — so each event is handled exactly once: the fallback catches
    # only events that landed on some other widget, hit-tests the POINTER
    # against the canvas, and re-routes with canvas-relative coordinates.
    def _route_wheel(ev, kind):
        try:
            under = dlg.winfo_containing(ev.x_root, ev.y_root)
        except (KeyError, tk.TclError):
            under = None
        if under is not canvas:
            return None
        if kind == "zoom":
            return _wheel_zoom(ev, ev.x_root - canvas.winfo_rootx(),
                               ev.y_root - canvas.winfo_rooty())
        return _wheel_pan(ev, horizontal=(kind == "panx"))

    canvas.configure(xscrollincrement=1, yscrollincrement=1)
    canvas.bind("<MouseWheel>", _wheel_pan)
    canvas.bind("<Shift-MouseWheel>", lambda e: _wheel_pan(e, True))
    canvas.bind("<Control-MouseWheel>", _wheel_zoom)
    dlg.bind("<MouseWheel>", lambda e: _route_wheel(e, "pan"))
    dlg.bind("<Shift-MouseWheel>", lambda e: _route_wheel(e, "panx"))
    dlg.bind("<Control-MouseWheel>", lambda e: _route_wheel(e, "zoom"))
    try:
        canvas.bind("<Command-MouseWheel>", _wheel_zoom)   # Aqua only
        dlg.bind("<Command-MouseWheel>", lambda e: _route_wheel(e, "zoom"))
    except tk.TclError:
        pass
    # X11 reports the wheel as buttons 4/5 (no delta): same routing.
    canvas.bind("<Button-4>",
                lambda e: (_pan_by(0, -30), "break")[1])
    canvas.bind("<Button-5>",
                lambda e: (_pan_by(0, 30), "break")[1])
    canvas.bind("<Shift-Button-4>",
                lambda e: (_pan_by(-30, 0), "break")[1])
    canvas.bind("<Shift-Button-5>",
                lambda e: (_pan_by(30, 0), "break")[1])
    canvas.bind("<Control-Button-4>",
                lambda e: (_zoom_at(1.15, e.x, e.y), "break")[1])
    canvas.bind("<Control-Button-5>",
                lambda e: (_zoom_at(1 / 1.15, e.x, e.y), "break")[1])
    for press, drag in (("<ButtonPress-2>", "<B2-Motion>"),
                        ("<ButtonPress-3>", "<B3-Motion>")):
        canvas.bind(press, lambda e: canvas.scan_mark(e.x, e.y))
        canvas.bind(drag, lambda e: canvas.scan_dragto(e.x, e.y, gain=1))

    # Keyboard zoom + pan — the device-independent guarantee (trackpad
    # pinch never reaches Tk; wheel DELIVERY varies by platform and Tk
    # build): +/= in, − out, 0 = fit, arrow keys pan.  Skipped while a text
    # widget has focus so typing never zooms or pans.
    def _typing():
        try:
            w = dlg.focus_get()
        except (KeyError, tk.TclError):
            w = None
        return isinstance(w, (tk.Entry, ttk.Entry, tk.Text,
                              tk.Spinbox, ttk.Spinbox))

    def _key_zoom(factor):
        if _typing():
            return
        _fit() if factor is None else _zoom_at(factor)

    def _key_pan(dx, dy):
        if not _typing():
            _pan_by(dx, dy)
    for _seq, _f in (("<Key-plus>", 1.25), ("<Key-equal>", 1.25),
                     ("<Key-minus>", 1 / 1.25), ("<Key-0>", None)):
        dlg.bind(_seq, lambda _e, f=_f: _key_zoom(f))
    # Plain arrows can be captured by whichever widget has focus (Tk's own
    # focus-traversal / combobox bindings run first) — a modifier bypasses
    # those, so ⌘-arrows (Mac) / Ctrl-arrows are bound explicitly as the
    # always-works spelling; plain arrows still pan when nothing eats them.
    for _seq, _d in (("<Key-Left>", (-40, 0)), ("<Key-Right>", (40, 0)),
                     ("<Key-Up>", (0, -40)), ("<Key-Down>", (0, 40))):
        dlg.bind(_seq, lambda _e, d=_d: _key_pan(*d))
        dlg.bind("<Control-" + _seq[1:], lambda _e, d=_d: _key_pan(*d))
        try:
            dlg.bind("<Command-" + _seq[1:], lambda _e, d=_d: _key_pan(*d))
        except tk.TclError:
            pass

    def _clear_marks():
        state["clicks"] = []
        _render()

    quantum = tk.StringVar(value="scale: —")

    # View selector (multi-view checklists only: wedge needs side + plan).
    view_var = tk.StringVar(value=state["cur"])

    def _select_view(vname):
        if vname not in views:
            return
        state.update(cur=vname, mode="idle", clicks=[])
        view_var.set(vname)
        _sync_quantum()
        _render()
        v = _cv()
        if v["img"] is None:
            status.set(f"{_VIEW_LABELS.get(vname, vname)}: load its image "
                       "(each view has its own image AND its own scale).")
        elif v["scale"] is None:
            status.set(f"{_VIEW_LABELS.get(vname, vname)}: set its scale.")
        else:
            status.set(f"{_VIEW_LABELS.get(vname, vname)}: ready.")

    if multiview:
        vrow = ttk.Frame(panel); vrow.pack(anchor=tk.W, fill=tk.X, pady=(0, 4))
        ttk.Label(vrow, text="View:").pack(side=tk.LEFT)
        for vn in views_needed:
            ttk.Radiobutton(vrow, text=_VIEW_LABELS.get(vn, vn),
                            variable=view_var, value=vn,
                            command=lambda n=vn: _select_view(n)).pack(
                side=tk.LEFT, padx=(6, 0))

    def _begin_scale():
        if _cv()["img"] is None:
            return
        _clear_marks(); state["mode"] = "scale"
        status.set("SCALE: click two points a known distance apart.")

    def _finish_scale():
        from tkinter import simpledialog
        v = _cv()
        p1, p2 = state["clicks"]
        d = simpledialog.askfloat("Scale", "Real distance between the two "
                                  "points (metres):", parent=dlg, minvalue=1e-6)
        if not d:
            _clear_marks(); state["mode"] = "idle"; return
        note = simpledialog.askstring(
            "Scale provenance", "Where does this distance come from?\n"
            "(e.g. 'claimed overall length 10.2 m, source X')",
            parent=dlg) or ""
        try:
            v["scale"] = im.Scale(p1, p2, d, anchor_note=note)
        except ValueError as e:
            messagebox.showerror("Bad scale", str(e), parent=dlg)
            _clear_marks(); state["mode"] = "idle"; return
        # Anchor-as-total: when the anchor IS the overall length (the common
        # case), it doubles as the closure total for free — no need to
        # re-measure the same span through the check-only prompt.
        if _closure_applies:
            v["anchor_total"] = (
                float(d) if messagebox.askyesno(
                    "Scale anchor",
                    "Is this distance the vehicle's OVERALL length?\n"
                    "(if so it doubles as the total for the stage-length "
                    "cross-check)", parent=dlg)
                else None)
        _sync_quantum()
        status.set("Scale set.  Pick a dimension and click Measure.")
        _clear_marks(); state["mode"] = "idle"
        _refresh_prompts()
        _refresh_closure()
    state["_finish_scale"] = _finish_scale

    ttk.Button(panel, text="Load image…", command=_load).pack(
        anchor=tk.W, fill=tk.X)
    ttk.Button(panel, text="Paste image  (⌘V / Ctrl-V)", command=_paste).pack(
        anchor=tk.W, fill=tk.X, pady=(4, 0))
    ttk.Button(panel, text="Set scale…", command=_begin_scale).pack(
        anchor=tk.W, fill=tk.X, pady=(4, 0))

    zrow = ttk.Frame(panel); zrow.pack(anchor=tk.W, fill=tk.X, pady=(4, 0))
    ttk.Button(zrow, text="+", width=3,
               command=lambda: _zoom_at(1.25)).pack(side=tk.LEFT)
    ttk.Button(zrow, text="−", width=3,
               command=lambda: _zoom_at(1 / 1.25)).pack(side=tk.LEFT,
                                                        padx=(2, 0))
    ttk.Button(zrow, text="Fit", width=4, command=_fit).pack(side=tk.LEFT,
                                                             padx=(2, 0))
    overlay_var = tk.BooleanVar(value=True)

    def _toggle_overlay():
        state["overlay"] = bool(overlay_var.get())
        _render()
    ttk.Checkbutton(zrow, text="Overlay accepted", variable=overlay_var,
                    command=_toggle_overlay).pack(side=tk.LEFT, padx=(8, 0))
    ttk.Label(panel, text="pan: scroll · ⌘-arrows (Mac) or arrows · "
                          "right-drag\nzoom: ⌘/Ctrl-scroll · + / − · 0 = fit",
              foreground="#999", justify=tk.LEFT).pack(anchor=tk.W)
    ttk.Label(panel, textvariable=quantum, foreground="#555").pack(
        anchor=tk.W, pady=(2, 8))

    ttk.Label(panel, text="Dimension to measure:").pack(anchor=tk.W)
    prompt_var = tk.StringVar()
    prompt_combo = ttk.Combobox(panel, textvariable=prompt_var,
                                state="readonly", width=36)
    prompt_combo.pack(anchor=tk.W, fill=tk.X)
    prompt_hint = ttk.Label(panel, text="", foreground="#777",
                            wraplength=250, justify=tk.LEFT)
    prompt_hint.pack(anchor=tk.W, pady=(2, 4))

    # The little "what to click" diagram: a stylized outline with the exact
    # clicks numbered, redrawn whenever the selected dimension changes
    # (including checklist auto-advance).  Data lives in image_measure
    # (diagram_spec / DIAGRAM_BASES); this only scales and draws.
    _DIAG_W, _DIAG_H = 232, 160
    diag_canvas = tk.Canvas(panel, width=_DIAG_W, height=_DIAG_H, bg="white",
                            highlightthickness=1, highlightbackground="#ccc")
    diag_canvas.pack(anchor=tk.W, pady=(0, 4))

    def _draw_prompt_diagram(p):
        # Rendered through matplotlib/Agg (diagram_render, the schematic's
        # architecture) for antialiased, resolution-independent art — the
        # canvas just displays the resulting image.  Same pure data
        # (diagram_spec / DIAGRAM_BASES / DIAGRAM_STYLE) as before.
        diag_canvas.configure(bg=im.DIAGRAM_STYLE["bg"])
        diag_canvas.delete("all")
        diag_canvas._photo = None        # ref for the GC; None = blank strip
        spec = im.diagram_spec(p) if p else None
        if not spec:
            return
        # Proportion-aware art: when the editor already holds real
        # dimensions (ctx["dims"]), the diagram draws THOSE proportions —
        # uniform scale, measurement points follow the reshaped geometry.
        # A straight-"new" object (core dims unset → layout None) keeps
        # the representative default art below.
        layout = None
        if ctx and ctx.get("dims"):
            # LIVE dims: this session's accepted values overlay the ones
            # captured at open, so the art tracks what you measure (typing
            # 0 for a pointed tip redraws the fin pointed).  The tip-chord
            # prompts keep the visible-tip floor — their diagram must show
            # the edge they ask for.
            if spec["base"] == "ro_side":
                layout = im.ro_side_layout(
                    im.ro_dims_with_accepted(ctx["dims"], state["accepted"]),
                    _DIAG_W, _DIAG_H,
                    tip_floor=(p["field"] == "_wing_tip_derive"))
            elif spec["base"] == "booster":
                layout = im.booster_side_layout(
                    im.booster_dims_with_accepted(ctx["dims"],
                                                  state["accepted"]),
                    _DIAG_W, _DIAG_H,
                    tip_floor=(p["field"] == "fin_tip"))
        if layout and p["field"] in layout["pts"]:
            base_polys = layout["polys"]
            spec = dict(spec, pts=layout["pts"][p["field"]],
                        subject=layout["subjects"].get(p["field"]))
        # Shape-aware base art: the RO nose outline follows the DECLARED
        # profile (cone / ogive / Sears-Haack / parabola / blunt) so the
        # picture matches the shape; everything else uses the static art.
        elif spec["base"] == "ro_side" and ctx and ctx.get("nose_shape"):
            base_polys = im.ro_side_base(ctx["nose_shape"],
                                         bool(ctx.get("biconic")))
        else:
            base_polys = im.DIAGRAM_BASES[spec["base"]]
        try:
            import diagram_render
            from PIL import ImageTk
        except ImportError:              # no PIL: blank strip, never crash
            return
        img = diagram_render.render_prompt_diagram(spec, base_polys,
                                                   _DIAG_W, _DIAG_H)
        diag_canvas._photo = ImageTk.PhotoImage(img)
        diag_canvas.create_image(0, 0, anchor="nw", image=diag_canvas._photo)

    # R1 clocking (only shown for spans a ×-roll can foreshorten, e.g. fins).
    # Default in-plane → no silent inflation; the correction is offered here.
    # The control is built only when the declared topology HAS such a dimension,
    # so its mere presence means "this vehicle has a clocking-sensitive span".
    _clock_labels = [lab for lab, _ in im.CLOCKING_OPTIONS]
    _clock_key = dict(im.CLOCKING_OPTIONS)
    _has_clocking = any(p.get("clocking_sensitive") for p in prompts)
    clock_var = tk.StringVar(value=_clock_labels[0])
    clock_frame = None
    if _has_clocking:
        clock_frame = ttk.Frame(panel)
        ttk.Label(clock_frame, text="Fin/wing clocking (R1):").pack(anchor=tk.W)
        clock_combo = ttk.Combobox(clock_frame, textvariable=clock_var,
                                   state="readonly", width=36,
                                   values=_clock_labels)
        clock_combo.pack(anchor=tk.W, fill=tk.X)          # hidden until selected

    result_var = tk.StringVar(value="")
    ttk.Label(panel, textvariable=result_var, foreground="#2a7",
              wraplength=250, justify=tk.LEFT).pack(anchor=tk.W)

    _label_by_field = {}

    def _refresh_prompts():
        _label_by_field.clear()
        labels = []
        for p in prompts:
            disp = f"{p['field']}  —  {p['label']}"
            labels.append(disp)
            _label_by_field[disp] = p
        prompt_combo["values"] = labels
        if labels:
            prompt_var.set(labels[0]); _on_prompt()

    def _on_prompt(*_):
        p = _label_by_field.get(prompt_var.get())
        prompt_hint.config(text=(p["label"] + f"   [{p['view']} view]") if p else "")
        _draw_prompt_diagram(p)
        if clock_frame is not None:
            if p and p.get("clocking_sensitive"):
                clock_frame.pack(anchor=tk.W, fill=tk.X, pady=(0, 4),
                                 after=prompt_hint)
            else:
                clock_frame.pack_forget()
    prompt_combo.bind("<<ComboboxSelected>>", _on_prompt)

    def _advance_prompt():
        """Checklist behavior (the design's 'walked through a checklist'):
        after a value is recorded, the selection ADVANCES to the next
        unmeasured dimension instead of parking on the old one — a selection
        left behind is how a fin chord ends up recorded as the vehicle length
        (observed in use; the delta preview caught it, but the tool should
        not set the trap)."""
        labels = list(prompt_combo["values"] or ())
        if not labels:
            return
        try:
            start = labels.index(prompt_var.get()) + 1
        except ValueError:
            start = 0
        for disp in labels[start:] + labels[:start]:
            p = _label_by_field.get(disp)
            if p and p["field"] not in state["accepted"]:
                prompt_var.set(disp)
                _on_prompt()
                return

    def _begin_measure():
        p = _label_by_field.get(prompt_var.get())
        if not p:
            return
        # View gating (Phase B): a plan-view dimension can only be clicked on
        # the plan view — the old label-only warning is now a hard guard (a
        # span clicked off a side elevation would be pure garbage: the span
        # runs into the page).  Auto-switch when the right view is loaded.
        if multiview and p["view"] != state["cur"]:
            _select_view(p["view"])
        v = _cv()
        if v["img"] is None:
            status.set(f"This dimension needs the {p['view'].upper()} view — "
                       "load that image first (each view has its own scale).")
            return
        # Angles are anchor-free (R2): they need an image but NO scale.
        if v["scale"] is None and not p.get("angle"):
            status.set(f"Set the {p['view'].upper()} view's scale first.")
            return
        # A fresh measurement retires any un-accepted prior reading: clear the
        # pending value, disarm Accept, and wipe the result line so a stale
        # number can never be mistaken for — or accepted as — the new one.
        state["prompt"] = p; state["_pending"] = None
        acc_btn.config(state="disabled")
        # Re-measuring an already-recorded field: say so (the old value
        # survives until the new reading is Accepted), and set mode BEFORE
        # rendering so the old overlay mark hides for the redo.
        prev = state["accepted"].get(p["field"])
        result_var.set("" if prev is None else
                       f"re-measuring {p['field']} — recorded {prev:.4g} "
                       "stays until you Accept the new reading")
        state["mode"] = "measure"
        _clear_marks()
        _click_progress()

    def _finish_measure():
        p = state["prompt"]
        if p.get("angle"):
            vertex, a1, a2 = state["clicks"]
            m = im.AngleMeasurement(p["field"], vertex, a1, a2,
                                    complement=bool(p.get("complement")))
            state["_pending_pts"] = (vertex, a1, a2, state["cur"])
            _clear_marks(); state["mode"] = "idle"
            if m.refused:
                result_var.set("✗ " + "; ".join(m.flags)
                               + " — nothing recorded.")
                return
            conv = (f"  (LE↔root {m.raw_deg:.1f}° → Λ)"
                    if getattr(m, "complement", False) else "")
            result_var.set(f"{p['field']} = {m.value_deg:.1f}°{conv}  "
                           "(anchor-free)\nAccept to record, or re-measure.")
            state["_pending"] = m

            def _accept_angle():
                mm = state.get("_pending")
                if mm is None:
                    return
                state["accepted"][mm.field] = mm.value_deg
                state["measurements"] = [x for x in state["measurements"]
                                         if x.field != mm.field] + [mm]
                vtx, r1, r2, pview = state["_pending_pts"]
                state["annotations"][mm.field] = (
                    pview, r1, r2, f"{mm.field} = {mm.value_deg:.1f}°", vtx)
                result_var.set(f"✓ recorded {mm.field} = {mm.value_deg:.1f}°")
                state["_pending"] = None
                _refresh_angle_checks()
                _render()
                _advance_prompt()
                _note_advance(mm.field)
            acc_btn.config(command=_accept_angle, state="normal")
            return
        s = _cv()["scale"]
        p1, p2 = state["clicks"]
        click_m, span = s.measure(p1, p2)
        clocking = (_clock_key.get(clock_var.get(), "in_plane")
                    if p.get("clocking_sensitive") else "in_plane")
        m = im.Measurement(p["field"], click_m, span, scale=s,
                           view=p["view"], convention=p["convention"],
                           clocking=clocking)
        state["_pending_pts"] = (p1, p2, state["cur"])
        _clear_marks(); state["mode"] = "idle"
        if m.refused:
            result_var.set("✗ " + "; ".join(m.flags) + " — nothing recorded.")
            return
        flag = ("  (" + "; ".join(m.flags) + ")") if m.flags else ""
        result_var.set(f"{p['field']} = {m.value_m:.4g} m   [{m.quantum_str()}]"
                       f"{flag}\nAccept to record, or re-measure.")
        state["_pending"] = m

        def _accept():
            mm = state.get("_pending")
            if mm is None:
                return
            state["accepted"][mm.field] = mm.value_m
            state["measurements"] = [x for x in state["measurements"]
                                     if x.field != mm.field] + [mm]
            pp1, pp2, pview = state["_pending_pts"]
            state["annotations"][mm.field] = (
                pview, pp1, pp2, f"{mm.field} = {mm.value_m:.4g} m", None)
            result_var.set(f"✓ recorded {mm.field} = {mm.value_m:.4g} m")
            state["_pending"] = None
            _refresh_closure()
            _render()
            _advance_prompt()
            _note_advance(mm.field)
        acc_btn.config(command=_accept, state="normal")
    state["_finish_measure"] = _finish_measure

    def _note_advance(done_field):
        """After Accept the checklist AUTO-ADVANCES — pressing Measure again
        now measures the NEXT field.  Say so, or a redo silently lands on
        the wrong dimension (field report: the redo 'didn't kill the older
        measure' — it had recorded a fresh value under the next field).
        Also ALWAYS redraw the diagram: the live dims just changed, and
        when the checklist has nowhere left to advance no redraw would
        otherwise fire (field report: typing 0 for the pointed tip never
        updated the picture)."""
        _draw_prompt_diagram(_label_by_field.get(prompt_var.get()))
        nxt = _label_by_field.get(prompt_var.get())
        if nxt and nxt["field"] != done_field:
            result_var.set(result_var.get()
                           + f"\n→ selection advanced to {nxt['field']} — "
                             f"re-select {done_field} to redo it")

    def _type_value():
        """Manual entry for a dimension the user already knows precisely (a
        published diameter, the anchor length itself): the checklist never
        FORCES a click.  Needs no image and no scale — the value is typed in
        stored metres — and it is recorded as entered-by-hand so the
        provenance stamp cannot claim it was measured off the image."""
        from tkinter import simpledialog
        p = _label_by_field.get(prompt_var.get())
        if not p:
            return
        conv = im.CONVENTIONS.get(p.get("convention"))
        hint = f"\n({conv[0]})" if conv else ""
        unit = "degrees" if p.get("angle") else "metres"
        v = simpledialog.askfloat(
            "Enter known value",
            f"{p['label']}{hint}\n\nEnter the STORED value in {unit} — "
            "recorded as entered by hand, not measured:",
            parent=dlg, minvalue=0.0)
        if v is None:
            return
        he = im.HandEntry(p["field"], v)
        state["accepted"][he.field] = he.value_m
        state["measurements"] = [x for x in state["measurements"]
                                 if x.field != he.field] + [he]
        result_var.set(f"✓ entered {he.field} = {v:g} m (by hand, not measured)")
        _refresh_closure()
        _advance_prompt()
        _note_advance(he.field)

    mrow = ttk.Frame(panel); mrow.pack(anchor=tk.W, fill=tk.X, pady=(6, 0))
    ttk.Button(mrow, text="Measure", command=_begin_measure).pack(side=tk.LEFT)
    acc_btn = ttk.Button(mrow, text="Accept", state="disabled")
    acc_btn.pack(side=tk.LEFT, padx=4)
    ttk.Button(mrow, text="Type value…", command=_type_value).pack(
        side=tk.LEFT, padx=(4, 0))

    # Length-closure line (booster: stages + fairing should tile the overall
    # length).  Warn-only by design — a mismatch is information (wrong claimed
    # total, mis-clicked invisible joint, real gap), never auto-corrected.
    _closure_applies = bool(im.closure_segments(prompts))
    closure_var = tk.StringVar(value="")
    closure_lbl = ttk.Label(panel, textvariable=closure_var, foreground="#888",
                            wraplength=250, justify=tk.LEFT)
    if _closure_applies:
        closure_lbl.pack(anchor=tk.W, pady=(6, 0))

    # Cross-view consistency line (multi-view only): the length measured in
    # BOTH views audits the two independent scale anchors — disagreement
    # means one anchor is wrong, and the span would inherit the error.
    viewcheck_var = tk.StringVar(value="")
    viewcheck_lbl = ttk.Label(panel, textvariable=viewcheck_var,
                              foreground="#888", wraplength=250,
                              justify=tk.LEFT)
    if multiview:
        viewcheck_lbl.pack(anchor=tk.W, pady=(6, 0))

    def _refresh_viewcheck():
        if not multiview:
            return
        vc = im.view_consistency(state["accepted"])
        if vc is None:
            viewcheck_var.set("view cross-check: measure the length in BOTH "
                              "views to audit the two scales")
            viewcheck_lbl.config(foreground="#888")
        else:
            viewcheck_var.set(im.view_consistency_note(vc))
            viewcheck_lbl.config(
                foreground="#b00" if abs(vc["rel"]) > 0.02 else "#2a7")

    # Angle cross-checks (warn-only, like the closure): every measurable
    # angle has an independent twin — derived-from-lengths identities and the
    # two-flank symmetry test (a failed symmetry impeaches the whole image:
    # tilt/perspective, the R9 screening upgraded to a measurement).
    _has_angles = any(p.get("angle") for p in prompts)
    anglecheck_var = tk.StringVar(value="")
    anglecheck_lbl = ttk.Label(panel, textvariable=anglecheck_var,
                               foreground="#888", wraplength=250,
                               justify=tk.LEFT)
    if _has_angles:
        anglecheck_lbl.pack(anchor=tk.W, pady=(6, 0))

    def _refresh_angle_checks():
        if not _has_angles:
            return
        acc = state["accepted"]
        notes = []
        if "_wing_sweep_var" in acc and "_wing_root_var" in acc \
                and "_wing_span_var" in acc:
            d = im.sweep_from_planform(acc["_wing_root_var"],
                                       acc["_wing_span_var"])
            notes.append(im.angle_check_note(acc["_wing_sweep_var"], d,
                                             "wing sweep (delta TE assumed)"))
        if "fin_sweep" in acc and "fin_root" in acc and "fin_span" in acc:
            d = im.sweep_from_planform(acc["fin_root"], acc["fin_span"],
                                       acc.get("fin_tip", 0.0))
            notes.append(im.angle_check_note(acc["fin_sweep"], d, "fin sweep"))
        up, lo = acc.get(im.FLANK_UPPER_FIELD), acc.get(im.FLANK_LOWER_FIELD)
        if up is not None and lo is not None:
            notes.append(im.symmetry_note(up, lo))
            if "_dia_var" in acc and "_len_var" in acc:
                d = im.cone_half_angle_from_lengths(acc["_dia_var"],
                                                    acc["_len_var"])
                notes.append(im.angle_check_note(0.5 * (up + lo), d,
                                                 "mean flank vs ⌀/L"))
        bad = any("DISAGREES" in n or "ASYMMETRIC" in n for n in notes if n)
        anglecheck_var.set("\n".join(n for n in notes if n)
                           or "angle checks: measure an angle (and its "
                           "lengths) to cross-check")
        anglecheck_lbl.config(foreground=("#b00" if bad
                                          else "#2a7" if notes else "#888"))
    _refresh_angle_checks()

    def _refresh_closure():
        _refresh_viewcheck()
        _refresh_angle_checks()      # a late-accepted length completes a twin
        if not _closure_applies:
            return
        total = (state["accepted"].get(im.OVERALL_LEN_CHECK_FIELD)
                 or views.get("side", {}).get("anchor_total"))
        c = im.length_closure(state["accepted"], prompts, total)
        if c is None:
            # No idle nag — the prompt sequence already asks for the overall
            # length; the line lights up once there is a result to show.
            closure_var.set("")
        else:
            closure_var.set(im.closure_note(c))
            closure_lbl.config(
                foreground=("#888" if not c["complete"]
                            else "#b00" if abs(c["rel"]) > 0.02 else "#2a7"))
    state["_refresh_closure"] = _refresh_closure
    _refresh_closure()

    ttk.Separator(dlg, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=8)
    af = ttk.Frame(dlg, padding=(8, 6)); af.pack(fill=tk.X)

    def _do_apply():
        # The stamp's primary scale is the side view's (or the first set one).
        order = ["side"] + [v for v in views_needed if v != "side"]
        sc = next((views[vn]["scale"] for vn in order
                   if vn in views and views[vn]["scale"] is not None), None)
        apply_fn(state["accepted"], state["measurements"], sc)
        dlg.destroy()

    def _apply():
        """R8 delta preview: before ANYTHING is written, show field →
        current editor value → proposed → Δ%, biggest deltas first.  A large
        delta is a FINDING (a hand-entered number contradicted by the image,
        or vice versa) — surfaced, never silently resolved."""
        acc = state["accepted"]
        if not acc or current_fn is None:
            _do_apply()
            return
        rows = im.apply_deltas(acc, current_fn(list(acc)))
        if not rows:
            _do_apply()
            return
        units = {p["field"]: ("°" if p.get("angle") else "m") for p in prompts}
        kinds = {m.field: ("entered" if getattr(m, "hand_entered", False)
                           else "measured") for m in state["measurements"]}
        pv = tk.Toplevel(dlg)
        pv.title("Apply preview — old vs new")
        pv.grab_set()
        frm = ttk.Frame(pv, padding=12); frm.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frm, text="Nothing is written until you confirm.  Large "
                  "deltas are findings — a stored number contradicted by the "
                  "image (or a mis-click), not noise to smooth over.",
                  wraplength=460, foreground="#555",
                  justify=tk.LEFT).grid(row=0, column=0, columnspan=5,
                                        sticky=tk.W, pady=(0, 8))
        for col, head in enumerate(("field", "current", "proposed", "Δ", "")):
            ttk.Label(frm, text=head, font=("TkDefaultFont", 9, "bold")).grid(
                row=1, column=col, sticky=tk.W, padx=(0, 12))
        for i, r in enumerate(rows, start=2):
            u = units.get(r["field"], "m")
            big = r["rel"] is not None and abs(r["rel"]) > im.DELTA_WARN_REL
            fg = "#b00" if big else "#333"
            old_txt = "—" if r["old"] is None else f"{r['old']:.4g} {u}"
            rel_txt = ("new" if r["rel"] is None else f"{r['rel']:+.1%}")
            for col, txt in enumerate(
                    (r["field"], old_txt, f"{r['new']:.4g} {u}", rel_txt,
                     kinds.get(r["field"], ""))):
                ttk.Label(frm, text=txt, foreground=fg).grid(
                    row=i, column=col, sticky=tk.W, padx=(0, 12), pady=1)
        skipped = len(acc) - len(rows)
        if skipped:
            ttk.Label(frm, text=f"({skipped} cross-check value"
                      f"{'s' if skipped != 1 else ''} not written — "
                      "audit only)", foreground="#888").grid(
                row=len(rows) + 2, column=0, columnspan=5, sticky=tk.W,
                pady=(6, 0))
        bf = ttk.Frame(pv, padding=(12, 8)); bf.pack(fill=tk.X)

        def _confirm():
            pv.destroy()
            _do_apply()
        ttk.Button(bf, text=f"Write {len(rows)} field"
                   f"{'s' if len(rows) != 1 else ''}",
                   command=_confirm).pack(side=tk.LEFT)
        ttk.Button(bf, text="Back", command=pv.destroy).pack(
            side=tk.LEFT, padx=6)

    # No anchor-free banner (cut 2026-08-18): each accepted value carries its
    # own anchor-free / scale-anchor provenance flag, so the static note was
    # boilerplate here.
    ttk.Button(af, text="Apply to editor", command=_apply).pack(side=tk.LEFT)
    ttk.Button(af, text="Cancel", command=dlg.destroy).pack(side=tk.LEFT, padx=6)

    # OS drag-and-drop — OPPORTUNISTIC: enabled only when the optional
    # tkinterdnd2 package is installed (it bundles the tkdnd Tk extension);
    # silently absent otherwise.  Load/Paste are the guaranteed paths, so
    # this is a convenience, not a dependency.
    def _on_drop(ev):
        paths = dlg.tk.splitlist(ev.data)
        if paths:
            _load_path(paths[0])
    dnd_on = False
    try:
        from tkinterdnd2 import TkinterDnD, DND_FILES
        TkinterDnD._require(dlg)
        canvas.drop_target_register(DND_FILES)
        canvas.dnd_bind("<<Drop>>", _on_drop)
        dnd_on = True
    except Exception:
        pass
    status.set("Load an image to begin"
               + (", drop one on the canvas," if dnd_on else "")
               + " or paste one (⌘V / Ctrl-V).")

    # The checklist is available from the start: Type value… works with no
    # image and no scale (only Measure needs those), so a known dimension can
    # be entered before — or instead of — any clicking.
    _refresh_prompts()

    # Test hooks — the loaders live in a closure; expose the seams the GUI
    # tests exercise (paste path, image-resets-scale, drop availability,
    # hand entry, views, zoom, measure gating).
    dlg._im_state = state
    dlg._im_views = views
    dlg._im_paste = _paste
    dlg._im_load_path = _load_path
    dlg._im_set_image = _set_image
    dlg._im_dnd = dnd_on
    dlg._im_type_value = _type_value
    dlg._im_select_view = _select_view
    dlg._im_begin_measure = _begin_measure
    dlg._im_zoom_at = _zoom_at
    dlg._im_fit = _fit
    dlg._im_wheel_steps = _wheel_steps
    dlg._im_wheel_zoom = _wheel_zoom
    dlg._im_wheel_pan = _wheel_pan
    dlg._im_key_zoom = _key_zoom
    dlg._im_key_pan = _key_pan
    dlg._im_pan_by = _pan_by
    dlg._im_route_wheel = _route_wheel
    dlg._im_canvas = canvas
    dlg._im_prompt_var = prompt_var
    dlg._im_diag = diag_canvas
    dlg._im_on_prompt = _on_prompt
    return dlg


class ROEditorDialog(tk.Toplevel):
    """Modal dialog for creating or editing an ROParams object.

    Usage::
        dlg = ROEditorDialog(parent, ro=existing_ro, mass_kg=ro_mass)
        parent.wait_window(dlg)
        if dlg.result is not None:
            self._ro = dlg.result
    """

    _GUIDANCE_LABELS = {
        "ballistic":               "Ballistic (drag · gravity · rotation)",
        "equilibrium_glide":       "Equilibrium glide (Tracy)",
        "equilibrium_glide_acton": "Non-oscillatory glide (Acton)",
        "skip_glide":              "Phugoid / skip-glide",
        "damped_glide":            "Damped phugoid glide",
        "dynamic_equilibrium_glide": "Dynamic equilibrium glide",
        # skip_to_equilibrium retired -> aliased to damped_glide on load.
        "skip_to_equilibrium":     "Damped phugoid glide",
    }

    # ROParams.body_form display labels (the label carries the ⌀ convention).
    _BODY_FORM_LABELS = {
        "axisymmetric": "body of revolution",
        "wedge":        "flattened wedge (⌀ = base depth)",
        "half_cone":    "half-cone lifting body",
    }

    def __init__(self, parent, ro=None, mass_kg=500.0, plan_sep=None,
                 booster=None):
        super().__init__(parent)
        # plan_sep: the ACTIVE reentry plan's separation mode ('body' |
        # 'separating_ro').  Separation is a plan choice, not an object property,
        # so the editor shows the plan's mode (and thus the body-only front-end
        # fields) even for a brand-new object — otherwise a non-separating plan
        # would open the editor in separating mode and hide the nose fields.
        self._plan_sep_override = plan_sep
        # booster: the current sidebar booster (raw).  For a non-separating body
        # the L/D and β are DERIVED from the whole airframe (stage + fins + nose),
        # which the object alone does not carry — so the editor composes its live
        # values onto this booster to preview them inline next to the fields.
        self._booster = booster
        self.title("Edit Reentry Object" if ro is not None else "New Reentry Object")
        self.resizable(False, True)
        self.grab_set()
        self._result = None
        # Pull-up g-limit and re-entry βₛ moved to the Reentry Plan editor
        # (they are "how it's flown", not hardware); preserved here from the
        # object being edited so a save round-trips them unchanged.
        self._orig_ro = ro

        # Two-column layout: identity + geometry down the left; maneuvering,
        # TPS and provenance in labeled groups on the right.  Same fields,
        # same variable names — only the arrangement differs, so everything
        # stays visible at once without the single-column scroll.
        cols = ttk.Frame(self, padding=(12, 8, 12, 0))
        cols.pack(fill=tk.BOTH, expand=True)
        left = ttk.Frame(cols)
        left.grid(row=0, column=0, sticky="nw")
        right = ttk.Frame(cols)
        right.grid(row=0, column=1, sticky="new", padx=(18, 0))
        cols.columnconfigure(1, weight=1)

        frm = ttk.Frame(left)
        frm.pack(fill=tk.X)
        # Geometry group — the shape and every dimensional field, plus the
        # image tool that populates them.
        geo = ttk.LabelFrame(left, text="Geometry", padding=(8, 4, 8, 6))
        geo.pack(fill=tk.X, pady=(8, 0))

        def _lbl(row, text, parent=frm):
            ttk.Label(parent, text=text).grid(
                row=row, column=0, sticky=tk.W, padx=(0, 8), pady=3)

        def _entry(row, var, width=14, parent=frm):
            e = ttk.Entry(parent, textvariable=var, width=width)
            e.grid(row=row, column=1, sticky=tk.W, pady=3)
            return e

        # Separation is a reentry-PLAN decision (how the object is flown, not
        # what it is): it is set with the Separation control in the sidebar's
        # Reentry Object section and persisted with the plan.  Shown read-only
        # here so the greyed-out inherited fields below are explicable.
        _lbl(0, "Separation:")
        # The active plan's separation wins when supplied (it is the run
        # authority); else fall back to the object's stored mode / the default.
        self._plan_sep = (self._plan_sep_override
                          or (getattr(ro, 'separation_mode', 'separating_ro')
                              if ro else 'separating_ro'))
        _sep_txt = ("Non-separating — body reenters with the final stage"
                    if self._plan_sep == 'body'
                    else "Separates at burnout")
        ttk.Label(frm, text=f"{_sep_txt}   (set in the sidebar — reentry plan)",
                  foreground="#888888").grid(row=0, column=1, sticky=tk.W, pady=3)

        # Name
        _lbl(1, "Name:")
        self._name_var = tk.StringVar(value=ro.name if ro else "")
        self._name_entry = _entry(1, self._name_var)

        # Mass.  Inherited (read-only) from the booster's last stage in body
        # mode — the airframe IS that stage — so the label says so; a separating
        # RV owns its mass.  (Separation is fixed for the life of this dialog.)
        # For a body the front end ALSO owns an ADDED payload (warhead / bus)
        # that rides on top of the inherited airframe burnout mass — entered
        # inline here, with a live "total reentry mass" truth line below.
        _from_booster = " (from booster)" if self._plan_sep == 'body' else ""
        _lbl(2, "Mass (kg):" + _from_booster)
        _mass_row = ttk.Frame(frm)
        _mass_row.grid(row=2, column=1, sticky=tk.W, pady=3)
        self._mass_var = tk.StringVar(
            value=f"{ro.mass_kg:.0f}" if ro else f"{mass_kg:.0f}")
        self._mass_entry = ttk.Entry(_mass_row, textvariable=self._mass_var,
                                     width=10)
        self._mass_entry.pack(side=tk.LEFT)
        # Payload (body only) — packed/hidden in _update_separation_state.
        self._payload_lbl = ttk.Label(_mass_row, text="   + payload:")
        self._payload_var = tk.StringVar(
            value=f"{ro.payload_kg:.0f}" if (ro and ro.payload_kg > 0) else "0")
        self._payload_entry = ttk.Entry(_mass_row, textvariable=self._payload_var,
                                        width=8)
        self._payload_unit = ttk.Label(_mass_row, text="kg")
        # Live total-reentry-mass truth line (airframe burnout + payload),
        # packed inline after the payload so the sum the run flies is visible
        # where it is entered.
        self._payload_total_lbl = ttk.Label(_mass_row, text="",
                                            foreground="#2a7")
        self._payload_var.trace_add(
            "write", lambda *_a: self._refresh_payload_total())

        # β with Estimate button.  For a non-separating body β is emergent from
        # the whole airframe (derived β(Mach) at run time), so it defaults to the
        # 0 = derive sentinel rather than a typed placeholder; a separating RV
        # keeps its designed-value default.  (FRONT_END_DESIGN.md Part II §10.)
        _lbl(3, "β (kg/m²):")
        self._beta_var = tk.StringVar(
            value=(f"{ro.beta_kg_m2:.0f}" if (ro and ro.beta_kg_m2 > 0)
                   else ("0" if self._plan_sep == 'body' else "10000")))
        _beta_row = ttk.Frame(frm)
        _beta_row.grid(row=3, column=1, sticky=tk.W, pady=3)
        self._beta_entry = ttk.Entry(_beta_row, textvariable=self._beta_var, width=10)
        self._beta_entry.pack(side=tk.LEFT)
        ttk.Label(_beta_row, text=" kg/m²").pack(side=tk.LEFT)
        ttk.Button(_beta_row, text="Estimate…",
                   command=self._calc_beta).pack(side=tk.LEFT, padx=(6, 0))
        # Inline derived-β preview for a non-separating body (managed in
        # _update_separation_state).  Mirrors the L/D field: with β left at the
        # sentinel 0 the label shows the geometry-derived β the run will use —
        # nose-first β(Mach) from the airframe Cd0, or the spinning-cylinder β
        # if the trim gate declares the body unstable (tumbling).  A trace keeps
        # it live as the user edits.  (FRONT_END_DESIGN.md Part II §10.)
        self._beta_derive_lbl = ttk.Label(
            _beta_row, text="  0 = derive (nose-first, or tumbling if unstable)",
            foreground="#2a7", wraplength=360, justify=tk.LEFT)
        self._beta_var.trace_add(
            "write", lambda *_a: self._refresh_beta_preview())

        # Shape — a SINGLE selector merging the nose profile (axisymmetric)
        # with the lifting-body forms.  From the user's view "shape" and "body
        # form" are one question; they only differ in physics.  Picking an
        # axisymmetric profile implies a body of revolution; the two lifting-
        # body entries set body_form and make the nose profile moot (a wedge is
        # flat, a half-cone is conical → 'cone' kept internally).  The data
        # model still stores shape and body_form separately; only the editor
        # merges them.  _shape_opts is the ONE ordered map both the display and
        # _build_ro read, so label ↔ (shape, body_form) never drifts.
        _lbl(0, "Shape:", parent=geo)
        self._shape_opts = self._shape_form_options()   # (label, shape, form)
        _cur_form = str(getattr(ro, 'body_form', '') or 'axisymmetric') if ro \
            else 'axisymmetric'
        _cur_shape = (ro.shape if ro else 'cone') or 'cone'
        self._shape_var = tk.StringVar(
            value=self._shape_label_for(
                _cur_shape, _cur_form,
                bool(getattr(ro, 'biconic', False)) if ro else False))
        self._shape_combo = ttk.Combobox(
            geo, textvariable=self._shape_var,
            values=[o[0] for o in self._shape_opts], state="readonly", width=30)
        self._shape_combo.grid(row=0, column=1, sticky=tk.W, pady=3)
        self._shape_combo.bind("<<ComboboxSelected>>",
                               lambda _e: self._update_body_form_state())
        # Populate geometry by clicking dimensions off an image (Phase A).
        ttk.Button(geo, text="Measure from image…",
                   command=self._measure_from_image).grid(
            row=1, column=1, sticky=tk.W, pady=(0, 3))

        # Diameter + length — inherited (read-only) from the booster's last
        # stage in body mode, labeled "from booster" to match.
        _lbl(2, "Diameter (m):" + _from_booster, parent=geo)
        self._dia_var = tk.StringVar(
            value=f"{ro.diameter_m:.2f}" if ro else "0.5")
        self._dia_entry = _entry(2, self._dia_var, width=10, parent=geo)

        _lbl(3, "Length (m):" + _from_booster, parent=geo)
        self._len_var = tk.StringVar(
            value=f"{ro.length_m:.2f}" if ro else "2.0")
        self._len_entry = _entry(3, self._len_var, width=10, parent=geo)

        # Nose-tip radius — drives Sutton-Graves stagnation heating (∝ 1/√RN).
        # Shown as the EFFECTIVE radius: an explicit value, else the
        # shape/diameter screening default (so the field is never a misleading
        # 0.000 for an auto RV).
        _lbl(4, "Nose radius (m):", parent=geo)
        self._nose_var = tk.StringVar(
            value=f"{ro.effective_nose_radius_m():.3f}" if ro else "0.050")
        self._nose_entry = _entry(4, self._nose_var, width=10, parent=geo)

        # Body nose length — the forward taper of a NON-SEPARATING body, carved
        # SUBTRACTIVELY from the airframe (whose length = the last stage's at run
        # time; see FRONT_END_DESIGN.md).  Active only in body mode, where the
        # RV's own length above is inherited from the stage; a separating RV uses
        # that length and ignores this.  0 = the schematic draws a flagged
        # shape-appropriate default.
        _lbl(5, "Body nose length (m):", parent=geo)
        self._body_nose_var = tk.StringVar(
            value=f"{getattr(ro, 'body_nose_length_m', 0.0):.2f}" if ro else "0")
        self._body_nose_entry = _entry(5, self._body_nose_var, width=10, parent=geo)

        # Biconic (two-cone) body — fore cone + aft frustum.  Only length and
        # break diameter are entered; the half-angles derive from these against
        # the base diameter / length (feeds the two-cone β estimator).  The
        # SELECTION lives in the merged Shape dropdown (its own entry — the
        # standalone checkbox failed discoverability twice); this var is the
        # backing state, driven by _update_body_form_state.
        self._biconic_var = tk.BooleanVar(
            value=bool(getattr(ro, 'biconic', False)) if ro else False)
        _lbl(6, "Fore-cone length (m):", parent=geo)
        self._fore_len_var = tk.StringVar(
            value=f"{getattr(ro, 'fore_length_m', 0.0):.2f}" if ro else "0")
        self._fore_len_entry = _entry(6, self._fore_len_var, width=10, parent=geo)
        _lbl(7, "Break diameter (m):", parent=geo)
        self._break_dia_var = tk.StringVar(
            value=f"{getattr(ro, 'break_diameter_m', 0.0):.2f}" if ro else "0")
        self._break_dia_entry = _entry(7, self._break_dia_var, width=10, parent=geo)

        # Planform span — WEDGE only (body geometry: the wedge's body IS its
        # lifting surface; tip-to-tip base width).  Distinct from the wing
        # planform below, which is a wing ON a body.  A half-cone's span is its
        # diameter; a body of revolution has none — so the field greys out for
        # every form but the wedge.
        _lbl(8, "Planform span (m):", parent=geo)
        self._body_span_var = tk.StringVar(
            value=f"{getattr(ro, 'body_span_m', 0.0):.2f}" if ro else "0")
        self._body_span_entry = _entry(8, self._body_span_var, width=10, parent=geo)

        # Wing planform + derived S/AR — GEOMETRY (hardware on the body), even
        # though it is STORED only with the maneuvering model (wings exist in
        # Thrusty to anchor the glide polar).  The rows grey out until
        # Maneuvering is ticked — the same declared-topology pattern as the
        # biconic fields above greying until their checkbox.
        def _we(row, label, default, unit=""):
            ttk.Label(geo, text=label).grid(
                row=row, column=0, sticky=tk.W, padx=(0, 8), pady=2)
            var = tk.StringVar(value=default)
            inner = ttk.Frame(geo)
            inner.grid(row=row, column=1, sticky=tk.W, pady=2)
            ent = ttk.Entry(inner, textvariable=var, width=10)
            ent.pack(side=tk.LEFT)
            if unit:
                ttk.Label(inner, text=f" {unit}").pack(side=tk.LEFT)
            return var, ent

        _wa = f"{ro.wing_area_m2:g}" if (ro and ro.wing_area_m2 > 0) else "0"
        _war = f"{ro.wing_aspect_ratio:g}" if (ro and ro.wing_aspect_ratio > 0) else "0"
        _wrc = f"{ro.wing_root_chord_m:g}" if (ro and getattr(ro, 'wing_root_chord_m', 0) > 0) else "0"
        _wss = f"{ro.wing_span_exposed_m:g}" if (ro and getattr(ro, 'wing_span_exposed_m', 0) > 0) else "0"
        _wsw = f"{ro.wing_sweep_deg:g}" if (ro and getattr(ro, 'wing_sweep_deg', 0) > 0) else "0"
        self._wing_root_var, self._wing_root_ent = _we(9, "Wing root chord:", _wrc, unit="m")
        self._wing_span_var, self._wing_span_ent = _we(10, "  exposed span:", _wss, unit="m")
        self._wing_sweep_var, self._wing_sweep_ent = _we(11, "  LE sweep:", _wsw, unit="°")
        self._wing_area_var, self._wing_area_ent = _we(12, "  wing area:", _wa, unit="m²")
        self._wing_ar_var, self._wing_ar_ent = _we(13, "  aspect ratio:", _war)
        # Panel thickness + count — GEOMETRY only (feed the 3-D Blender
        # export; the polar uses area/AR/sweep, not these).  A C-HGB carries
        # 4 flaps; a delta glider 2.
        _wt = (f"{ro.wing_thickness_m:g}"
               if (ro and getattr(ro, 'wing_thickness_m', 0) > 0) else "0")
        self._wing_thick_var, self._wing_thick_ent = _we(
            14, "  panel thickness:", _wt, unit="m")
        _nw = f"{getattr(ro, 'n_wings', 4)}" if ro else "4"
        self._wing_n_var, self._wing_n_ent = _we(
            15, "  panel count:", _nw)

        # Reentry CG (body only): where the CG sits decides nose-first trim vs
        # tumbling (hence the whole glide).  The auto-estimate treats the empty
        # airframe as a uniform tube (CG at the centroid); a real warhead-forward
        # missile sits ahead of that.  0 = auto.  Shown/gated for a body in
        # _update_separation_state.  (FRONT_END_DESIGN.md Part II.)
        self._reentry_cg_lbl = ttk.Label(geo, text="Reentry CG (m):")
        self._reentry_cg_lbl.grid(row=16, column=0, sticky=tk.W,
                                  padx=(0, 8), pady=2)
        self._reentry_cg_var = tk.StringVar(
            value=f"{getattr(ro, 'reentry_cg_m', 0.0):.2f}" if ro else "0")
        self._reentry_cg_entry = _entry(16, self._reentry_cg_var, width=10,
                                        parent=geo)
        self._reentry_cg_hint = ttk.Label(
            geo, text="0 = auto (uniform-airframe centroid; set forward for a "
            "warhead-forward body)", foreground="#2a7", wraplength=320,
            justify=tk.LEFT)
        self._reentry_cg_hint.grid(row=17, column=0, columnspan=3, sticky=tk.W)
        # All wing entries, for gating: enabled iff Maneuvering is on AND
        # the form can carry a wing (the WEDGE disables them — its body is the
        # lifting surface; a wing entered here would double-count through the
        # polar's e_pull while invisible to the user).
        self._wing_entries = (self._wing_area_ent, self._wing_ar_ent,
                              self._wing_root_ent, self._wing_span_ent,
                              self._wing_sweep_ent, self._wing_thick_ent,
                              self._wing_n_ent)
        # Live "derived" indicator + traces: when a planform is present, S/AR
        # are recomputed and locked; otherwise they are editable.
        self._wing_derived_lbl = ttk.Label(
            geo, text="", foreground="#2a7", justify=tk.LEFT, wraplength=300)
        self._wing_derived_lbl.grid(row=16, column=0, columnspan=2,
                                    sticky=tk.W, pady=(0, 0))
        for _v in (self._wing_root_var, self._wing_span_var,
                   self._wing_sweep_var, self._dia_var):
            _v.trace_add("write", lambda *_a: self._sync_wing_derived())

        # Biconic is a body-of-revolution concept; a lifting-body Shape unticks
        # and disables it.  (Body form now lives in the merged Shape selector
        # above — no separate dropdown.)
        self._update_body_form_state()
        self._update_biconic_state()

        # Sync the read-only state of mass/diameter/length to separation mode
        self._update_separation_state()

        # ── Maneuvering (glider / HGV) — HOW it flies, not what it is.  The
        # wing PLANFORM lives in Geometry (it is hardware); this group keeps
        # only the capability flag and L/D.  The checkbox IS the group label.
        self._glider_var = tk.BooleanVar(value=ro.glider_enabled if ro else False)
        _glider_chk = ttk.Checkbutton(right, text="Maneuvering (glider / HGV)",
                                      variable=self._glider_var,
                                      command=self._update_glider_state)
        _man_box = ttk.LabelFrame(right, labelwidget=_glider_chk)
        _man_box.pack(fill=tk.X)
        self._glider_frm = ttk.Frame(_man_box, padding=(8, 2, 8, 6))
        self._glider_frm.pack(fill=tk.X)
        self._glider_frm.columnconfigure(1, weight=1)

        ttk.Label(self._glider_frm, text="Lift/drag (L/D):").grid(
            row=0, column=0, sticky=tk.W, padx=(0, 8), pady=2)
        # Non-separating body: L/D is emergent from the airframe (derived at run
        # time via the trim gate), so default to 0 = derive; a separating RV
        # keeps its designed-value default.
        _LD = (f"{ro.glider_LD:.2f}" if (ro and ro.glider_LD > 0)
               else ("0" if self._plan_sep == 'body' else "2.5"))
        self._LD_var = tk.StringVar(value=_LD)
        self._LD_entry = ttk.Entry(self._glider_frm, textvariable=self._LD_var,
                                   width=10)
        self._LD_entry.grid(row=0, column=1, sticky=tk.W, pady=2)
        # Body-only: an inline "Estimate" button and a live derived-value label
        # next to the field (managed in _update_separation_state).  L/D for a
        # non-separating body is derived from the WHOLE airframe (stage + fins +
        # nose), so the estimate composes the live object onto the current
        # booster; a 0 in the field then reads as "0 → derives ~2.57".
        self._LD_est_btn = ttk.Button(self._glider_frm, text="Estimate",
                                      width=9, command=self._show_ld_estimate)
        self._LD_derive_lbl = ttk.Label(
            self._glider_frm, text="", foreground="#2a7",
            wraplength=340, justify=tk.LEFT)
        self._LD_var.trace_add("write", lambda *_a: self._refresh_ld_preview())
        # Estimator trim row (Phase 3): α* feeds the windward-α consistency
        # guard, C_L0 the offset polar.  Set only by "Use β and L/D" (lifting
        # forms); shown read-only so the stored state is never invisible.
        self._trim_alpha_val = float(getattr(ro, 'trim_alpha_deg', 0.0) or 0.0) \
            if ro else 0.0
        self._trim_cl0_val = float(getattr(ro, 'trim_CL0', 0.0) or 0.0) \
            if ro else 0.0
        self._trim_lbl = ttk.Label(self._glider_frm, text="",
                                   foreground="#2a7", justify=tk.LEFT)
        self._trim_lbl.grid(row=1, column=0, columnspan=2, sticky=tk.W)
        self._sync_trim_label()
        ttk.Label(self._glider_frm,
                  text="Wings (in Geometry) anchor the drag polar: enter the "
                       "planform and S / AR derive; without one, type S / AR "
                       "directly (flagged tab).  Pull-up g-limit and re-entry "
                       "βₛ are in the Reentry Plan editor.",
                  foreground="#888888", justify=tk.LEFT, wraplength=340).grid(
                      row=2, column=0, columnspan=2, sticky=tk.W, pady=(2, 0))

        self._sync_wing_derived()
        self._update_glider_state()
        self._update_body_form_state()   # apply wedge gating to the wing rows
        # Re-run separation gating now that the glider widgets exist, so the
        # body-only inline L/D Estimate button + derived-value label appear and
        # the preview populates on open.
        self._update_separation_state()

        # ── Thermal protection (TPS) materials — per location (§10) ──────
        _tps_box = ttk.LabelFrame(right, text="Thermal protection (TPS) materials",
                                  padding=(8, 2, 8, 6))
        _tps_box.pack(fill=tk.X, pady=(8, 0))
        tps_frm = ttk.Frame(_tps_box)
        tps_frm.pack(fill=tk.X)
        tps_frm.columnconfigure(1, weight=1)

        # Choice list shared by both locations: "(none)" + full grouped catalog
        # + "Custom…".  _mat_map maps the display string back to the material key.
        self._mat_choices, self._mat_map = self._material_choices()

        def _mat_row(row, label, cur_key):
            ttk.Label(tps_frm, text=label).grid(
                row=row, column=0, sticky=tk.W, padx=(0, 8), pady=2)
            var = tk.StringVar(value=self._display_for_key(cur_key))
            cb = ttk.Combobox(tps_frm, textvariable=var, values=self._mat_choices,
                              state="readonly", width=34)
            cb.grid(row=row, column=1, sticky=tk.W, pady=2)
            return var, cb

        _nk = (ro.nose_tps_material or ro.tps_material) if ro else ""
        _bk = (ro.body_tps_material or ro.tps_material) if ro else ""
        self._nose_mat_var, self._nose_mat_cb = _mat_row(0, "Nose / leading edge:", _nk)
        self._nose_cust_frm, self._nose_cust = self._build_custom_fields(
            tps_frm, 1, ro.nose_tps_custom if ro else None)
        self._body_mat_var, self._body_mat_cb = _mat_row(2, "Body / acreage:", _bk)
        self._body_cust_frm, self._body_cust = self._build_custom_fields(
            tps_frm, 3, ro.body_tps_custom if ro else None)

        ttk.Label(tps_frm, text="Body layer thickness:").grid(
            row=4, column=0, sticky=tk.W, padx=(0, 8), pady=2)
        _bt = f"{ro.body_tps_thickness_m:.4f}" if (ro and ro.body_tps_thickness_m > 0) else "0"
        self._body_thick_var = tk.StringVar(value=_bt)
        _bt_in = ttk.Frame(tps_frm); _bt_in.grid(row=4, column=1, sticky=tk.W, pady=2)
        ttk.Entry(_bt_in, textvariable=self._body_thick_var, width=10).pack(side=tk.LEFT)
        ttk.Label(_bt_in, text=" m  (0 = auto)").pack(side=tk.LEFT)

        ttk.Label(tps_frm, text="Emissivity:").grid(
            row=5, column=0, sticky=tk.W, padx=(0, 8), pady=2)
        self._emiss_var = tk.StringVar(value=f"{ro.emissivity:.2f}" if ro else "0.85")
        _em_in = ttk.Frame(tps_frm); _em_in.grid(row=5, column=1, sticky=tk.W, pady=2)
        ttk.Entry(_em_in, textvariable=self._emiss_var, width=10).pack(side=tk.LEFT)
        ttk.Label(_em_in, text="  (0.85 typical; range 0.75–0.90)").pack(side=tk.LEFT)

        self._nose_mat_cb.bind("<<ComboboxSelected>>",
                               lambda _e: self._update_custom_state("nose"))
        self._body_mat_cb.bind("<<ComboboxSelected>>",
                               lambda _e: self._update_custom_state("body"))
        self._update_custom_state("nose")
        self._update_custom_state("body")

        # ── Provenance — where these numbers came from / how firm they are ──
        _prov_box = ttk.LabelFrame(right, text="Provenance", padding=(8, 2, 8, 6))
        _prov_box.pack(fill=tk.X, pady=(8, 0))
        prov_frm = ttk.Frame(_prov_box)
        prov_frm.pack(fill=tk.X)
        prov_frm.columnconfigure(1, weight=1)
        ttk.Label(prov_frm, text="Source:").grid(row=0, column=0, sticky=tk.W, padx=(0, 8), pady=2)
        self._source_var = tk.StringVar(value=(ro.source if ro else ""))
        ttk.Entry(prov_frm, textvariable=self._source_var, width=42).grid(
            row=0, column=1, sticky=tk.EW, pady=2)
        ttk.Label(prov_frm, text="Notes:").grid(row=1, column=0, sticky=tk.NW, padx=(0, 8), pady=2)
        self._notes_text = tk.Text(prov_frm, width=42, height=4, wrap=tk.WORD)
        self._notes_text.grid(row=1, column=1, sticky=tk.EW, pady=2)
        if ro and ro.notes:
            self._notes_text.insert("1.0", ro.notes)

        # OK / Save to Library / Cancel
        ttk.Separator(self, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=12, pady=8)
        btn_frm = ttk.Frame(self, padding=(12, 0, 12, 12))
        btn_frm.pack(fill=tk.X)
        ttk.Button(btn_frm, text="OK", command=self._ok).pack(side=tk.LEFT)
        ttk.Button(btn_frm, text="Save to Library…",
                   command=self._save_to_library).pack(side=tk.LEFT, padx=6)
        ttk.Button(btn_frm, text="Cancel", command=self.destroy).pack(
            side=tk.LEFT, padx=6)

    # ------------------------------------------------------------------
    def _calc_beta(self):
        """Open the β estimator sub-dialog.  Body-form aware: a lifting body
        (wedge / half-cone) gets the α-sweep estimator (β AND L/D, trim-
        consistent); a body of revolution gets the zero-AoA cone/biconic
        build-up."""
        if self._body_form_key() in ("wedge", "half_cone"):
            return self._calc_beta_lifting(self._body_form_key())
        import math
        dlg = tk.Toplevel(self)
        dlg.title("Estimate Object β")
        dlg.resizable(False, False)
        dlg.grab_set()

        # Pre-fill from current diameter / length
        try:
            _d = float(self._dia_var.get())
            _l = float(self._len_var.get())
            if _d > 0 and _l > 0:
                _theta0   = f"{math.degrees(math.atan(1.0 / (2.0 * _l / _d))):.1f}"
                _dia_dflt = self._dia_var.get()
            else:
                raise ValueError
        except Exception:
            _theta0 = "10.0"; _dia_dflt = "0"

        frm = ttk.Frame(dlg, padding=12)
        frm.pack(fill=tk.X)
        frm.columnconfigure(1, weight=1)

        def _lbl(row, text):
            ttk.Label(frm, text=text).grid(
                row=row, column=0, sticky=tk.W, padx=(0, 8), pady=3)

        _lbl(0, "Object mass (kg):")
        # For a non-separating body the reentering mass is the inherited airframe
        # burnout PLUS the added payload (§A2); β = m/(Cd·A) must use that total,
        # not the airframe alone.  A separating RV owns its whole mass in _mass_var.
        _seed_mass = self._mass_var.get()
        if self._plan_sep == 'body':
            try:
                _total = (float(self._mass_var.get() or 0.0)
                          + float(self._payload_var.get() or 0.0))
                _seed_mass = f"{_total:.0f}"
            except (ValueError, tk.TclError):
                pass
        mass_var = tk.StringVar(value=_seed_mass)
        ttk.Entry(frm, textvariable=mass_var, width=10).grid(row=0, column=1, sticky=tk.W)

        _lbl(1, "Reentry object base diameter (m):")
        dia_var = tk.StringVar(value=_dia_dflt)
        ttk.Entry(frm, textvariable=dia_var, width=10).grid(row=1, column=1, sticky=tk.W)

        _lbl(2, "Fore-cone half-angle (°):")
        theta_var = tk.StringVar(value=_theta0)
        ttk.Entry(frm, textvariable=theta_var, width=10).grid(row=2, column=1, sticky=tk.W)

        _lbl(3, "Nose radius / base radius:")
        try:
            _d_eps = float(self._dia_var.get())
            _rn_eps = float(self._nose_var.get())
            _eps_dflt = (f"{2.0 * _rn_eps / _d_eps:.3f}"
                         if _d_eps > 0 and _rn_eps >= 0 else "0.0")
        except (ValueError, AttributeError):
            _eps_dflt = "0.0"
        eps_var  = tk.StringVar(value=_eps_dflt)
        eps_row  = ttk.Frame(frm)
        eps_row.grid(row=3, column=1, sticky=tk.W)
        ttk.Entry(eps_row, textvariable=eps_var, width=10).pack(side=tk.LEFT)
        ttk.Label(eps_row, text="  (0 = sharp tip,  1 = hemisphere)",
                  foreground="gray50").pack(side=tk.LEFT)

        _lbl(4, "Eval Mach:")
        mach_var = tk.StringVar(value="10")
        ttk.Entry(frm, textvariable=mach_var, width=10).grid(
            row=4, column=1, sticky=tk.W)

        # Wing area — pre-filled from the object editor's wing field.  A winged
        # vehicle is draggier than the bare cone, so its estimated β is lower;
        # this is the advisory half of the wing decoupling (Level 2).
        _lbl(5, "Wing area (m²):")
        # Prefer the DERIVED area when a planform is present (matches the polar),
        # falling back to whatever is in the wing-area field otherwise.
        try:
            _S_wa, _AR_wa, _src_wa = self._current_wing_geometry()
            _wa_dflt = f"{_S_wa:.4g}" if (_src_wa and _S_wa > 0) \
                else self._wing_area_var.get()
        except AttributeError:
            _wa_dflt = "0"
        wing_var = tk.StringVar(value=_wa_dflt or "0")
        _wf = ttk.Frame(frm); _wf.grid(row=5, column=1, sticky=tk.W)
        ttk.Entry(_wf, textvariable=wing_var, width=10).pack(side=tk.LEFT)
        ttk.Label(_wf, text="  (0 = bare cone)",
                  foreground="gray50").pack(side=tk.LEFT)

        # ── Biconic (two-cone) option ─────────────────────────────────────
        # The half-angle above is the FORE cone; ticking biconic adds the aft
        # frustum (aft half-angle + break-diameter ratio).  Pre-filled from the
        # object's stored biconic geometry when it has one.
        _th2_dflt, _br_dflt = "6.0", "0.7"
        _bic0 = bool(getattr(self, '_biconic_var', None) and self._biconic_var.get())
        if _bic0:
            try:
                _r = _biconic_angles(float(self._dia_var.get()),
                                     float(self._len_var.get()),
                                     float(self._fore_len_var.get()),
                                     float(self._break_dia_var.get()),
                                     float(self._nose_var.get()))
                if _r:
                    theta_var.set(f"{_r[0]:.1f}")
                    _th2_dflt, _br_dflt = f"{_r[1]:.1f}", f"{_r[2]:.3f}"
            except (ValueError, AttributeError):
                pass
        _lbl(6, "Biconic (two-cone):")
        bicon_var = tk.BooleanVar(value=_bic0)
        _bcf = ttk.Frame(frm); _bcf.grid(row=6, column=1, sticky=tk.W)
        ttk.Checkbutton(_bcf, variable=bicon_var,
                        text="add aft frustum").pack(side=tk.LEFT)
        _lbl(7, "Aft half-angle (°):")
        th2_var = tk.StringVar(value=_th2_dflt)
        th2_entry = ttk.Entry(frm, textvariable=th2_var, width=10)
        th2_entry.grid(row=7, column=1, sticky=tk.W)
        _lbl(8, "Break ⌀ / base ⌀:")
        br_var = tk.StringVar(value=_br_dflt)
        br_entry = ttk.Entry(frm, textvariable=br_var, width=10)
        br_entry.grid(row=8, column=1, sticky=tk.W)

        def _toggle_bicon(*_):
            st = "normal" if bicon_var.get() else "disabled"
            th2_entry.config(state=st); br_entry.config(state=st)
        bicon_var.trace_add("write", _toggle_bicon)
        _toggle_bicon()

        ttk.Separator(dlg, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=12)

        res = ttk.Frame(dlg, padding=(12, 8))
        res.pack(fill=tk.X)
        res.columnconfigure(1, weight=1)

        def _res_row(row, label):
            ttk.Label(res, text=label).grid(
                row=row, column=0, sticky=tk.W, padx=(0, 8), pady=2)
            lbl = ttk.Label(res, text="—", foreground="gray40")
            lbl.grid(row=row, column=1, sticky=tk.W)
            return lbl

        cdw_lbl  = _res_row(0, "Cd pressure (Newtonian):")
        cdf_lbl  = _res_row(1, f"Cd friction (Cf {_CONE_CF:g} · S_wet/A):")
        cdb_lbl  = _res_row(2, "Cd base (2/γM²):")
        cdwing_lbl = _res_row(3, "Cd wing (Cf · 2·S_w/A):")
        cd_lbl   = _res_row(4, "Cd total:")
        area_lbl = _res_row(5, "Reference area (m²):")
        beta_lbl = ttk.Label(res, text="—", font=("", 11, "bold"), foreground="navy")
        ttk.Label(res, text="β = m / (Cd · A):").grid(
            row=6, column=0, sticky=tk.W, padx=(0, 8), pady=2)
        beta_lbl.grid(row=6, column=1, sticky=tk.W)
        ttk.Label(res,
                  text="Newtonian pressure 2·sin²θ + ε²·cos⁴θ (Wells &\n"
                       "Armstrong TR R-127) + turbulent skin friction +\n"
                       "hypersonic base drag + wing friction.  Cf is a screening\n"
                       "constant (0.0008–0.0015 band); wing wave drag is not carried.",
                  foreground="gray50", justify=tk.LEFT).grid(
            row=7, column=0, columnspan=2, sticky=tk.W, pady=(4, 0))

        _result = [None]

        def _compute(*_):
            try:
                m = float(mass_var.get()); d = float(dia_var.get())
                th = float(theta_var.get()); ep = float(eps_var.get())
                mk = float(mach_var.get())
                wa = max(0.0, float(wing_var.get() or 0.0))
                if d <= 0 or m <= 0 or th <= 0 or mk <= 0:
                    raise ValueError
            except ValueError:
                for _l in (cdw_lbl, cdf_lbl, cdb_lbl, cdwing_lbl, cd_lbl, area_lbl):
                    _l.config(text="—")
                beta_lbl.config(text="invalid input"); _result[0] = None
                return
            area = math.pi * (d / 2.0) ** 2
            _war = (wa / area if area > 0 else 0.0)
            if bicon_var.get():
                try:
                    th2 = float(th2_var.get()); br = float(br_var.get())
                    if not (0.0 < br < 1.0) or th2 <= 0:
                        raise ValueError
                except ValueError:
                    for _l in (cdw_lbl, cdf_lbl, cdb_lbl, cdwing_lbl, cd_lbl, area_lbl):
                        _l.config(text="—")
                    beta_lbl.config(text="invalid biconic input"); _result[0] = None
                    return
                c = _cd_biconic_hypersonic(th, th2, br, ep, mach=mk,
                                           wing_area_ratio=_war)
            else:
                c = _cd_cone_hypersonic(th, ep, mach=mk, wing_area_ratio=_war)
            beta = m / (c['total'] * area) if c['total'] > 0 else float('inf')
            cdw_lbl.config(text=f"{c['pressure']:.4f}")
            cdf_lbl.config(text=f"{c['friction']:.4f}")
            cdb_lbl.config(text=f"{c['base']:.4f}")
            cdwing_lbl.config(text=f"{c.get('wing', 0.0):.4f}")
            cd_lbl.config(text=f"{c['total']:.4f}")
            area_lbl.config(text=f"{area:.4f} m²")
            beta_lbl.config(text=f"{beta:,.0f} kg/m²")
            _result[0] = beta

        for _v in (mass_var, dia_var, theta_var, eps_var, mach_var,
                   bicon_var, th2_var, br_var, wing_var):
            _v.trace_add("write", _compute)
        _compute()

        ttk.Separator(dlg, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=12)
        bf = ttk.Frame(dlg, padding=(12, 8))
        bf.pack(fill=tk.X)

        def _use():
            if _result[0] is not None and _result[0] != float('inf'):
                self._beta_var.set(f"{_result[0]:.0f}")
                # Also stamp the absolute nose-tip radius (eps × base radius)
                # back into the editor so heating uses the same geometry.
                try:
                    _dia = float(dia_var.get())
                    _ep  = float(eps_var.get())
                    if _dia > 0 and _ep >= 0:
                        self._nose_var.set(f"{_ep * _dia / 2.0:.3f}")
                except (ValueError, AttributeError):
                    pass
            dlg.destroy()

        ttk.Button(bf, text="Use this value", command=_use).pack(side=tk.LEFT)
        ttk.Button(bf, text="Cancel", command=dlg.destroy).pack(side=tk.LEFT, padx=6)

    # ------------------------------------------------------------------
    def _calc_beta_lifting(self, form):
        """α-sweep β + L/D estimator for the lifting-body forms (Phase 2c).

        Presentation over booster_models.lifting_body_sweep(): modified-
        Newtonian pressure + Eckert reference-temperature friction + base
        drag, swept over α, returning ONE consistent trim row — β(α=0) is
        what the drag polar wants, (L/D)max is quoted WITH the α* that
        produces it (the Tracy & Wright attitude/L-D inconsistency, per
        Candler & Leyva, cannot be re-created from this dialog).  The
        conditions line states M, Re, boundary-layer state, base-drag
        convention, A_ref, and K — an L/D without its conditions is not a
        number (Fetterman).
        """
        import math
        dlg = tk.Toplevel(self)
        dlg.title("Estimate β and L/D — " + ("flattened wedge" if form == "wedge"
                                             else "half-cone lifting body"))
        dlg.resizable(False, False)
        dlg.grab_set()

        frm = ttk.Frame(dlg, padding=12)
        frm.pack(fill=tk.X)
        frm.columnconfigure(1, weight=1)

        def _lbl(row, text):
            ttk.Label(frm, text=text).grid(
                row=row, column=0, sticky=tk.W, padx=(0, 8), pady=3)

        def _entry_row(row, label, default, hint=""):
            _lbl(row, label)
            var = tk.StringVar(value=default)
            inner = ttk.Frame(frm); inner.grid(row=row, column=1, sticky=tk.W)
            ttk.Entry(inner, textvariable=var, width=10).pack(side=tk.LEFT)
            if hint:
                ttk.Label(inner, text=f"  {hint}",
                          foreground="gray50").pack(side=tk.LEFT)
            return var

        # Body: seed with the total reentry mass (airframe burnout + payload),
        # not the airframe alone — β = m/(Cd·A) uses the mass that reenters.
        _seed_mass = self._mass_var.get()
        if self._plan_sep == 'body':
            try:
                _seed_mass = f"{float(self._mass_var.get() or 0.0) + float(self._payload_var.get() or 0.0):.0f}"
            except (ValueError, tk.TclError):
                pass
        mass_var = _entry_row(0, "Object mass (kg):", _seed_mass)
        len_var = _entry_row(1, "Length (m):", self._len_var.get() or "0")
        if form == "wedge":
            # ⌀ field = base depth for the wedge (editor labels carry this).
            depth_var = _entry_row(2, "Base depth (m):", self._dia_var.get() or "0")
            try:                       # pre-fill from the editor's span field
                _span0 = f"{max(0.0, float(self._body_span_var.get() or 0.0)):g}"
            except (ValueError, AttributeError):
                _span0 = "0"
            span_var = _entry_row(3, "Planform span (m):", _span0,
                                  "(REQUIRED — measurable off an image)")
            # Phase 2b: swept-cylinder leading edge (AEDC §2.1.3).  For a
            # wedge the "nose" IS the leading edge, so the RO nose-radius
            # field pre-fills it; 0 = sharp (the documented upper bound).
            try:
                _rle0 = f"{max(0.0, float(self._nose_var.get() or 0.0)):g}"
            except (ValueError, AttributeError):
                _rle0 = "0"
            rle_var = _entry_row(4, "Leading-edge radius (m):", _rle0,
                                 "(0 = sharp; gives the sweep its cost)")
            geo_vars = (depth_var, span_var, rle_var)
        else:
            dia_var = _entry_row(2, "Base diameter (m):", self._dia_var.get() or "0")
            theta_lbl = ttk.Label(frm, text="—", foreground="gray40")
            _lbl(3, "Half-cone angle θ (derived):")
            theta_lbl.grid(row=3, column=1, sticky=tk.W)
            # Phase 2b: wing-body composite — the editor's wing planform (the
            # Fetterman half-cone + delta-wing configuration) joins the
            # sweep as a coplanar lower surface.  Derived, shown, not typed.
            wing_lbl = ttk.Label(frm, text="—", foreground="gray40")
            _lbl(4, "Wing planform (from editor):")
            wing_lbl.grid(row=4, column=1, sticky=tk.W)
            geo_vars = (dia_var,)
        mach_var = _entry_row(5, "Eval Mach:", "10")
        re_var = _entry_row(6, "Reynolds number (length):", "1e7",
                            "(0 = inviscid ceiling — anchors only)")
        turb_var = tk.BooleanVar(value=True)
        _lbl(7, "Boundary layer:")
        ttk.Checkbutton(frm, variable=turb_var,
                        text="turbulent (untick = laminar)").grid(
            row=7, column=1, sticky=tk.W)
        tw_var = _entry_row(8, "Wall / freestream temp ratio:", "1.0",
                            "(Eckert T*; cold wall = 1)")
        base_var = tk.BooleanVar(value=True)
        _lbl(9, "Base drag 2/(γM²):")
        ttk.Checkbutton(frm, variable=base_var,
                        text="include (untick to match base-corrected data)"
                        ).grid(row=9, column=1, sticky=tk.W)

        ttk.Separator(dlg, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=12)
        res = ttk.Frame(dlg, padding=(12, 8))
        res.pack(fill=tk.X)
        res.columnconfigure(1, weight=1)

        def _res_row(row, label):
            ttk.Label(res, text=label).grid(
                row=row, column=0, sticky=tk.W, padx=(0, 8), pady=2)
            lbl = ttk.Label(res, text="—", foreground="gray40")
            lbl.grid(row=row, column=1, sticky=tk.W)
            return lbl

        ld_lbl = ttk.Label(res, text="—", font=("", 11, "bold"), foreground="navy")
        ttk.Label(res, text="(L/D)max at α*:").grid(
            row=0, column=0, sticky=tk.W, padx=(0, 8), pady=2)
        ld_lbl.grid(row=0, column=1, sticky=tk.W)
        beta_lbl = ttk.Label(res, text="—", font=("", 11, "bold"), foreground="navy")
        ttk.Label(res, text="β (zero lift) = m/(C_D0·A_ref):").grid(
            row=1, column=0, sticky=tk.W, padx=(0, 8), pady=2)
        beta_lbl.grid(row=1, column=1, sticky=tk.W)
        btrim_lbl = _res_row(2, "β at trim α*:")
        cl0_lbl = _res_row(3, "Camber offset C_L0 (at min drag):")
        cond_lbl = ttk.Label(res, text="—", foreground="gray50", justify=tk.LEFT)
        cond_lbl.grid(row=4, column=0, columnspan=2, sticky=tk.W, pady=(4, 0))
        warn_lbl = ttk.Label(res, text="", foreground="#a33", justify=tk.LEFT)
        warn_lbl.grid(row=5, column=0, columnspan=2, sticky=tk.W)

        _result = [None]           # (beta_zero_lift, LD_max)

        def _clear(msg):
            for _l in (btrim_lbl, cl0_lbl):
                _l.config(text="—")
            ld_lbl.config(text="—"); beta_lbl.config(text=msg)
            cond_lbl.config(text="—"); warn_lbl.config(text="")
            _result[0] = None

        def _compute(*_):
            try:
                m = float(mass_var.get()); L = float(len_var.get())
                mk = float(mach_var.get()); re = float(re_var.get() or 0.0)
                tw = float(tw_var.get())
                if m <= 0 or L <= 0 or mk <= 0:
                    raise ValueError
                if form == "wedge":
                    t = float(depth_var.get()); b = float(span_var.get())
                    if t <= 0 or b <= 0:
                        return _clear("enter base depth and span")
                    r_le = max(0.0, float(rle_var.get() or 0.0))
                    kw = dict(length_m=L, depth_m=t, span_m=b, r_le_m=r_le)
                    a_ref = None                       # planform (derived)
                else:
                    d = float(dia_var.get())
                    if d <= 0:
                        return _clear("enter base diameter")
                    th = math.degrees(math.atan2(d / 2.0, L))
                    theta_lbl.config(text=f"{th:.2f}°  (atan(r_b/ℓ))")
                    # Wing composite: exposed panels (c_r + c_t)·s_e from the
                    # editor's planform, only when Maneuvering declares wings.
                    s_exp = 0.0
                    try:
                        if bool(self._glider_var.get()):
                            c_r = float(self._wing_root_var.get() or 0.0)
                            s_e = float(self._wing_span_var.get() or 0.0)
                            swp = float(self._wing_sweep_var.get() or 0.0)
                            if c_r > 0.0 and s_e > 0.0:
                                c_t = max(0.0, c_r - s_e * math.tan(
                                    math.radians(swp)))
                                s_exp = (c_r + c_t) * s_e
                    except (ValueError, AttributeError, tk.TclError):
                        s_exp = 0.0
                    wing_lbl.config(
                        text=(f"S_exp = {s_exp:.3g} m² (composite, Fetterman "
                              "config)" if s_exp > 0.0
                              else "none (body alone)"))
                    kw = dict(theta_deg=th, wing_exposed_m2=s_exp)
                    a_ref = math.pi * (d / 2.0) ** 2   # full-circle base area
            except (ValueError, AttributeError):
                return _clear("invalid input")
            r = _lifting_body_sweep(form, mach=mk,
                                    reynolds_length=(re if re > 0 else None),
                                    wall_temp_ratio=tw,
                                    turbulent=bool(turb_var.get()),
                                    base_drag=bool(base_var.get()),
                                    mass_kg=m, a_ref_m2=a_ref, **kw)
            tr, cond = r['trim'], r['conditions']
            ld_lbl.config(text=f"{tr['LD_max']:.2f}   at α* = "
                               f"{tr['alpha_star_deg']:.1f}°")
            beta_lbl.config(text=f"{tr['beta_zero_lift']:,.0f} kg/m²")
            btrim_lbl.config(text=f"{tr['beta_trim']:,.0f} kg/m²")
            cl0_lbl.config(text=f"{tr['C_L0']:+.4f}")
            bl = "inviscid (Cf=0)" if cond['inviscid'] else \
                ("turbulent" if cond['turbulent'] else "laminar")
            extra = ""
            if cond.get('wing_ratio'):
                extra += f" · wing S/A_b {cond['wing_ratio']:.2f}"
            if cond.get('r_le_m'):
                extra += f" · LE r {cond['r_le_m']:g} m (swept-cyl)"
            sharp_note = ("blunt-LE swept-cylinder term included"
                          if cond.get('r_le_m') else
                          "sharp-body Newtonian: an UPPER BOUND on L/D\n"
                          "(real blunt/viscous vehicles run ~30–40% lower)")
            cond_lbl.config(text=(
                f"conditions: M {cond['mach']:g} · Re "
                f"{cond['reynolds_length'] or 0:.3g} · {bl} · Cf "
                f"{cond['cf']:.5f} · base drag "
                f"{'on' if cond['base_drag'] else 'off'} ·\n"
                f"A_ref {cond['a_ref_m2']:.4g} m² ({cond['a_ref_kind']}) · "
                f"K {cond['K']:g}{extra}  — {sharp_note}"))
            warns = []
            if cond['inviscid']:
                warns.append("Cf = 0: inviscid ceiling — for anchor "
                             "comparisons only, not a vehicle estimate.")
            if tr['LD_max'] > 6.0:
                warns.append("(L/D)max exceeds the viscous-optimized-waverider "
                             "band (≈6–7): treat as non-physical.")
            warn_lbl.config(text="\n".join(warns))
            _result[0] = (tr['beta_zero_lift'], tr['LD_max'],
                          tr['alpha_star_deg'], tr['C_L0'])

        for _v in (mass_var, len_var, mach_var, re_var, tw_var, turb_var,
                   base_var) + geo_vars:
            _v.trace_add("write", _compute)
        _compute()

        ttk.Separator(dlg, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=12)
        bf = ttk.Frame(dlg, padding=(12, 8))
        bf.pack(fill=tk.X)

        def _use():
            if _result[0] is not None:
                beta0, ldmax, a_star, cl0 = _result[0]
                self._beta_var.set(f"{beta0:.0f}")
                # (L/D)max pre-fills the glider L/D — the capability figure the
                # airframe supports, quoted from the SAME α* as the β above.
                try:
                    self._LD_var.set(f"{ldmax:.2f}")
                except (AttributeError, tk.TclError):
                    pass
                # Phase 3: the trim row persists on the RO from the SAME sweep
                # — α* feeds the windward-α consistency guard, C_L0 the offset
                # polar (sweep-native coefficients; _build_ro stores them).
                self._trim_alpha_val = float(a_star)
                self._trim_cl0_val = float(cl0)
                self._sync_trim_label()
                # The span the estimate used persists on the wedge (body_span_m)
                # — the same number the β/L-D pair was computed from.
                if form == "wedge":
                    try:
                        self._body_span_var.set(f"{float(span_var.get()):g}")
                    except (ValueError, AttributeError, tk.TclError):
                        pass
            dlg.destroy()

        ttk.Button(bf, text="Use β and L/D", command=_use).pack(side=tk.LEFT)
        ttk.Button(bf, text="Cancel", command=dlg.destroy).pack(side=tk.LEFT, padx=6)

    # ------------------------------------------------------------------
    # image_measure field → editor StringVar map (one place, so the dialog and
    # the tests agree on what a measured field writes to).  Wing S/AR are NOT
    # here: they derive from the planform vars via _sync_wing_derived.
    # UNITS CONTRACT: the tool writes METRES; every var listed must be a
    # metre-labelled field backed by a *_m model attribute (audited
    # 2026-08-01).  Degrees fields (LE sweep, half-angles) stay out — a
    # two-point distance tool cannot measure an angle anyway.
    _IMG_FIELD_VARS = ("_len_var", "_dia_var", "_nose_var", "_body_span_var",
                       "_fore_len_var", "_break_dia_var",
                       "_wing_root_var", "_wing_span_var")
    # Angle targets, in DEGREES — written by the 3-click angle prompts, which
    # never pass through the metre CONVENTIONS.  The flank check fields are
    # deliberately absent (check-only, never stored).
    _IMG_ANGLE_FIELD_VARS = ("_wing_sweep_var",)

    def _img_field_var(self, field):
        """The editor StringVar an image-tool field writes to, or None for a
        field this editor cannot store (check-only cross-checks).  Shared by
        apply and the R8 delta preview."""
        if field in self._IMG_FIELD_VARS + self._IMG_ANGLE_FIELD_VARS:
            return getattr(self, field, None)
        return None

    def _current_image_values(self, fields):
        """{field: current editor value} for the R8 delta preview — only the
        fields this editor can actually write; None = blank/unparseable."""
        out = {}
        for f in fields:
            var = self._img_field_var(f)
            if var is None:
                continue
            try:
                out[f] = float(var.get())
            except (ValueError, tk.TclError):
                out[f] = None
        return out

    def _apply_image_measurements(self, accepted, measurements, scale):
        """Write accepted image measurements into this editor's fields and
        append the provenance stamp to notes (the image tool's ONLY durable
        outputs — decision 1).  Extracted so it is testable without the canvas.
        `accepted` is {field_var_name: value_m}; refused measurements never
        reach here."""
        import datetime
        import image_measure as im
        for field in self._IMG_FIELD_VARS + self._IMG_ANGLE_FIELD_VARS:
            if field in accepted:
                var = getattr(self, field, None)
                if var is not None:
                    var.set(f"{float(accepted[field]):.4g}")
        # The nose-tip click sees the sphere-cone TANGENCY circle, whose
        # width is 2·R_N·cos(θ), not the full sphere diameter — so on a CONE
        # the half-width convention under-reads R_N by cos(θ).  Corrected
        # here with θ derived from the just-written length/diameter (biconic:
        # the FORE cone carries the tip, so its length/break-⌀ set θ).
        # Curved profiles (ogive/Haack/…) keep the plain hemisphere
        # convention — their blend slope is not the L-⌀ cone angle — and a
        # blunt cylinder is the θ=0 identity case.
        if "_nose_var" in accepted and self._shape_key() == "cone":
            try:
                if self._biconic_var.get():
                    th = im.cone_half_angle_from_lengths(
                        float(self._break_dia_var.get() or 0.0),
                        float(self._fore_len_var.get() or 0.0))
                else:
                    th = im.cone_half_angle_from_lengths(
                        float(self._dia_var.get() or 0.0),
                        float(self._len_var.get() or 0.0))
                r_n = im.nose_radius_from_tip_width(
                    float(accepted["_nose_var"]), th)
                self._nose_var.set(f"{r_n:.4g}")
            except (ValueError, tk.TclError):
                pass
        # Wing sweep is DERIVED from the planform, not measured as an angle:
        # tan Λ = (root − tip)/span (a length ratio — anchor-free, no
        # complement).  Fires whenever any planform length was measured; a
        # pointed delta (no tip measured) uses tip = 0.  Straight-TE
        # assumption (the same wing_geometry uses).
        if any(f in accepted for f in ("_wing_root_var", "_wing_span_var",
                                       "_wing_tip_derive")):
            try:
                root = float(self._wing_root_var.get() or 0.0)
                span = float(self._wing_span_var.get() or 0.0)
                tip = float(accepted.get("_wing_tip_derive", 0.0) or 0.0)
                sw = im.sweep_from_planform(root, span, tip)
                if sw is not None:
                    self._wing_sweep_var.set(f"{sw:.4g}")
            except (ValueError, tk.TclError):
                pass
        # Measured wing geometry enables the Maneuvering section (wings are
        # stored with the maneuvering model) — the same measured-it-so-show-it
        # rule as the booster's fairing/fin sections.  The derived S/AR fire
        # from the freshly written planform via _update_wing_state.
        if any(f in accepted for f in ("_wing_root_var", "_wing_span_var",
                                       "_wing_tip_derive", "_wing_sweep_var")) \
                and hasattr(self, "_glider_var") \
                and not self._glider_var.get():
            self._glider_var.set(True)
            self._update_glider_state()
        if measurements and scale is not None:
            stamp = im.provenance_stamp(measurements, scale,
                                        datetime.date.today().isoformat())
            try:
                cur = self._notes_text.get("1.0", "end-1c")
                self._notes_text.insert(
                    "end", ("\n" if cur.strip() else "") + stamp)
            except (AttributeError, tk.TclError):
                pass

    def _measure_from_image(self):
        """Open the shared image-measure dialog for the reentry object: prompts
        come from the declared body_form, and Apply writes fields + stamps
        notes via _apply_image_measurements.  The editor's LOADED dimensions
        ride along in ctx so the diagrams draw this vehicle's proportions
        (a new/empty object falls back to the representative art)."""
        import image_measure as im
        form = self._body_form_key()

        def _fv(name):
            try:
                return float(getattr(self, name).get())
            except (ValueError, tk.TclError, AttributeError):
                return 0.0
        dims = dict(length_m=_fv("_len_var"), diameter_m=_fv("_dia_var"),
                    nose_radius_m=_fv("_nose_var"),
                    shape=self._shape_key(),
                    biconic=bool(self._biconic_var.get()),
                    fore_length_m=_fv("_fore_len_var"),
                    break_diameter_m=_fv("_break_dia_var"),
                    wing_root_m=_fv("_wing_root_var"),
                    wing_span_m=_fv("_wing_span_var"),
                    wing_sweep_deg=_fv("_wing_sweep_var"))
        _open_image_measure_dialog(
            self, "Measure from image — reentry object",
            im.ro_prompts(form, biconic=bool(self._biconic_var.get()))
            + im.ro_angle_prompts(form),
            self._apply_image_measurements,
            current_fn=self._current_image_values,
            ctx=dict(nose_shape=self._shape_key(),
                     biconic=bool(self._biconic_var.get()),
                     dims=dims))

    # ------------------------------------------------------------------
    def _current_wing_geometry(self):
        """(S, AR, source) for the wing as the editor fields currently stand.

        Builds a throwaway shim from the live entry values and defers to
        booster_models.wing_geometry() — the ONE place that decides whether a
        planform is present and what reference S / AR it implies — so the
        editor read-out, the schematic, and the drag polar never diverge.
        """
        def _f(var):
            try:
                return max(0.0, float(var.get() or 0.0))
            except (ValueError, AttributeError):
                return 0.0
        shim = types.SimpleNamespace(
            wing_root_chord_m=_f(self._wing_root_var),
            wing_span_exposed_m=_f(self._wing_span_var),
            wing_sweep_deg=_f(self._wing_sweep_var),
            wing_area_m2=_f(self._wing_area_var),
            wing_aspect_ratio=_f(self._wing_ar_var),
            diameter_m=_f(self._dia_var),
        )
        return _wing_geometry(shim)

    def _sync_wing_derived(self):
        """When a planform is present, DERIVE S / AR from it and lock the two
        entries read-only; otherwise leave them editable (direct-entry
        fallback).  Keeps a green indicator line in step.  No-op for the wedge
        (wing rows are disabled — _update_body_form_state owns their state)."""
        if getattr(self, '_shape_opts', None) and self._body_form_key() == "wedge":
            return
        if getattr(self, '_glider_var', None) is not None \
                and not self._glider_var.get():
            return                       # rows disabled — _update_wing_state owns them
        try:
            c_r = float(self._wing_root_var.get() or 0.0)
            s_e = float(self._wing_span_var.get() or 0.0)
        except (ValueError, AttributeError):
            c_r = s_e = 0.0
        has_planform = c_r > 0.0 and s_e > 0.0
        if has_planform:
            S, AR, _src = self._current_wing_geometry()
            # Populate then lock, so the stored ROParams carries the derived
            # numbers even though the user can't hand-edit them.
            self._wing_area_ent.configure(state="normal")
            self._wing_ar_ent.configure(state="normal")
            self._wing_area_var.set(f"{S:.4g}")
            self._wing_ar_var.set(f"{AR:.3g}")
            self._wing_area_ent.configure(state="readonly")
            self._wing_ar_ent.configure(state="readonly")
            self._wing_derived_lbl.configure(
                text=f"↳ derived from planform: S = {S:.3g} m² · "
                     f"AR = {AR:.2g}  (read-only)")
        else:
            self._wing_area_ent.configure(state="normal")
            self._wing_ar_ent.configure(state="normal")
            self._wing_derived_lbl.configure(text="")

    # ------------------------------------------------------------------
    def _build_ro(self):
        """Validate fields and build an ROParams.  Returns None on input error
        (after showing a messagebox).
        """
        try:
            name    = self._name_var.get().strip() or "(unnamed)"
            mass_kg = float(self._mass_var.get())
            beta    = float(self._beta_var.get())
            shape   = self._shape_key()          # merged Shape selector
            dia     = float(self._dia_var.get())
            length  = float(self._len_var.get())
            nose_rn = float(self._nose_var.get())
        except ValueError:
            messagebox.showerror(
                "Invalid input",
                "Mass, β, diameter, length, and nose radius must be numbers.",
                parent=self)
            return None

        body_form = self._body_form_key()
        # The editor disables the tick for lifting bodies; the guard makes the
        # rule hold even if state drifted (biconic = body of revolution only).
        biconic = bool(self._biconic_var.get()) and body_form == "axisymmetric"
        fore_len_m = break_dia_m = 0.0
        if biconic:
            try:
                fore_len_m = max(0.0, float(self._fore_len_var.get()))
                break_dia_m = max(0.0, float(self._break_dia_var.get()))
            except ValueError:
                messagebox.showerror(
                    "Invalid input",
                    "Biconic fore-cone length and break diameter must be numbers.",
                    parent=self)
                return None
            if not (0.0 < fore_len_m < length and 0.0 < break_dia_m < dia):
                messagebox.showerror(
                    "Invalid biconic",
                    "Fore-cone length must be less than total length, and break "
                    "diameter less than the base diameter.",
                    parent=self)
                return None

        # Wedge-only planform span (body geometry; disabled for other forms).
        body_span = 0.0
        if body_form == "wedge":
            try:
                body_span = max(0.0, float(self._body_span_var.get() or 0.0))
            except ValueError:
                messagebox.showerror(
                    "Invalid input", "Planform span must be a number.",
                    parent=self)
                return None

        try:
            body_nose = max(0.0, float(self._body_nose_var.get() or 0.0))
        except ValueError:
            messagebox.showerror(
                "Invalid input", "Body nose length must be a number.",
                parent=self)
            return None

        try:
            reentry_cg = max(0.0, float(self._reentry_cg_var.get() or 0.0))
        except ValueError:
            messagebox.showerror(
                "Invalid input", "Reentry CG must be a number.", parent=self)
            return None

        try:
            payload = max(0.0, float(self._payload_var.get() or 0.0))
        except ValueError:
            messagebox.showerror(
                "Invalid input", "Payload mass must be a number.", parent=self)
            return None

        glider_on = bool(self._glider_var.get())
        wing_area = wing_ar = 0.0
        wing_root = wing_span = wing_sweep = 0.0
        LD = 0.0
        if glider_on:
            try:
                LD = float(self._LD_var.get())
            except ValueError:
                messagebox.showerror(
                    "Invalid input", "L/D must be a number.", parent=self)
                return None
        # Wing fields are stored only when VISIBLE (not for the wedge, whose
        # rows are disabled): hidden-but-active wing physics through the
        # polar's e_pull would be dishonest.  What you see is what's stored.
        wing_thick = 0.0
        wing_n = 4
        if glider_on and body_form != "wedge":
            try:
                wing_area = max(0.0, float(self._wing_area_var.get() or 0.0))
                wing_ar   = max(0.0, float(self._wing_ar_var.get() or 0.0))
                wing_root  = max(0.0, float(self._wing_root_var.get() or 0.0))
                wing_span  = max(0.0, float(self._wing_span_var.get() or 0.0))
                wing_sweep = max(0.0, float(self._wing_sweep_var.get() or 0.0))
                wing_thick = max(0.0, float(self._wing_thick_var.get() or 0.0))
                wing_n = max(2, int(float(self._wing_n_var.get() or 4)))
            except ValueError:
                messagebox.showerror(
                    "Invalid input", "Wing area / aspect ratio / planform "
                    "fields must be numbers.",
                    parent=self)
                return None
            # Whenever a planform is present, S / AR are DERIVED from it — never
            # trust a hand-typed area that might have drifted (the editor locks
            # those fields, but a JSON loaded straight in could still carry a
            # stale pair).  wing_geometry() is the single source of truth.
            if wing_root > 0.0 and wing_span > 0.0:
                wing_area, wing_ar, _ = self._current_wing_geometry()

        # --- TPS materials (§10): resolve each dropdown to a catalog key, or a
        # 'custom_*' sentinel with the bespoke properties stored on the RV. ---
        _nose_disp = self._nose_mat_var.get()
        _body_disp = self._body_mat_var.get()
        nose_custom = body_custom = None
        if _nose_disp == self._MAT_CUSTOM_LABEL:
            nose_custom = self._custom_dict_from(self._nose_cust)
            nose_key = heating.CUSTOM_NOSE_KEY if nose_custom else ""
        else:
            nose_key = self._mat_map.get(_nose_disp, "")
        if _body_disp == self._MAT_CUSTOM_LABEL:
            body_custom = self._custom_dict_from(self._body_cust)
            body_key = heating.CUSTOM_BODY_KEY if body_custom else ""
        else:
            body_key = self._mat_map.get(_body_disp, "")
        try:
            emiss = float(self._emiss_var.get())
            body_thick = float(self._body_thick_var.get())
        except ValueError:
            messagebox.showerror(
                "Invalid input",
                "Emissivity and body layer thickness must be numbers.",
                parent=self)
            return None

        ro_new = ROParams(
            name=name, mass_kg=mass_kg, beta_kg_m2=beta,
            shape=shape, diameter_m=dia, length_m=length,
            nose_radius_m=nose_rn,
            biconic=biconic, fore_length_m=fore_len_m,
            break_diameter_m=break_dia_m,
            body_form=body_form, body_span_m=body_span,
            body_nose_length_m=body_nose, reentry_cg_m=reentry_cg,
            payload_kg=payload,
            glider_enabled=glider_on,
            glider_LD=LD,
            wing_area_m2=wing_area,
            wing_aspect_ratio=wing_ar,
            wing_root_chord_m=wing_root,
            wing_span_exposed_m=wing_span,
            wing_sweep_deg=wing_sweep,
            wing_thickness_m=wing_thick,   # geometry only (3-D export)
            n_wings=wing_n,                # geometry only (3-D export)
            # Estimator trim row (Phase 3): kept only for the lifting forms
            # its sweep produced it for — a stale offset carried onto a body
            # of revolution would silently skew the polar.
            trim_alpha_deg=(self._trim_alpha_val
                            if body_form in ("wedge", "half_cone") else 0.0),
            trim_CL0=(self._trim_cl0_val
                      if body_form in ("wedge", "half_cone") else 0.0),
            emissivity=emiss,
            nose_tps_material=nose_key,
            body_tps_material=body_key,
            body_tps_thickness_m=body_thick,
            nose_tps_custom=nose_custom,
            body_tps_custom=body_custom,
            source=self._source_var.get().strip(),
            notes=self._notes_text.get("1.0", "end-1c").strip(),
        )
        # This dialog edits HARDWARE.  Carry every reentry-plan field through
        # from the object being edited, wholesale — otherwise saving a hardware
        # tweak would silently reset the stored plan (glide law, dive, banks,
        # ζ, separation, attitude, …) to dataclass defaults, because
        # _save_ro_to_library re-extracts the plan from what we return here.
        # glider_enabled is the one plan key the dialog owns (its checkbox).
        if self._orig_ro is not None:
            import dataclasses as _dc
            _carry = {k: (list(_v) if isinstance(_v := getattr(self._orig_ro, k),
                                                 list) else _v)
                      for k in mm._REENTRY_PLAN_KEYS if k != 'glider_enabled'}
            ro_new = _dc.replace(ro_new, **_carry)
        return ro_new

    def _update_glider_state(self):
        # Grey-out, not hide: the L/D row (and the wing rows in Geometry)
        # disable until Maneuvering is ticked — the same declared-topology
        # pattern as the biconic fields.  Hiding the frame would collapse the
        # LabelFrame to nothing, checkbox and all.
        self._LD_entry.config(
            state="normal" if self._glider_var.get() else "disabled")
        self._update_wing_state()

    def _update_wing_state(self):
        """The wing rows live in the GEOMETRY group (a wing is hardware) but
        are declared topology of the maneuvering model: enabled iff
        Maneuvering is on AND the form can carry a wing ON the body (not the
        wedge, whose body IS the lifting surface — hidden-but-active wing
        physics through the polar's e_pull would be dishonest)."""
        if not hasattr(self, '_wing_entries') or not hasattr(self, '_glider_var'):
            return                       # __init__ mid-build; re-run at the end
        on = (bool(self._glider_var.get())
              and self._body_form_key() != "wedge")
        if on:
            for e in self._wing_entries:
                e.config(state="normal")
            self._sync_wing_derived()   # restore derived-S/AR readonly
        else:
            for e in self._wing_entries:
                e.config(state="disabled")
            self._wing_derived_lbl.configure(text="")

    def _update_biconic_state(self):
        st = "normal" if self._biconic_var.get() else "disabled"
        self._fore_len_entry.config(state=st)
        self._break_dia_entry.config(state=st)

    def _sync_trim_label(self):
        """Read-only display of the stored estimator trim row (Phase 3)."""
        if not hasattr(self, '_trim_lbl'):
            return
        if self._trim_alpha_val or self._trim_cl0_val:
            self._trim_lbl.configure(
                text=f"↳ estimator trim: α* = {self._trim_alpha_val:.1f}° · "
                     f"C_L0 = {self._trim_cl0_val:+.4f}  (offset polar + "
                     "windward α)")
        else:
            self._trim_lbl.configure(text="")

    # Biconic lives IN the Shape selector: to the user a two-cone IS a shape
    # (the standalone checkbox failed discoverability twice), even though the
    # model stores it as a flag on a body of revolution.
    _BICONIC_LABEL = "Biconic — two-cone (fore + aft)"

    def _shape_form_options(self):
        """Ordered (display_label, shape_key, body_form_key) for the merged
        Shape selector: the axisymmetric nose profiles (each implying a body
        of revolution) with the biconic entry right after the plain cone,
        then the two lifting-body forms (nose profile moot → 'cone').  The
        single source both the dropdown and _build_ro read."""
        opts = [(v, k, 'axisymmetric') for k, v in NOSE_SHAPE_LABELS.items()]
        i = next((j for j, o in enumerate(opts) if o[1] == 'cone'), 0) + 1
        opts.insert(i, (self._BICONIC_LABEL, 'cone', 'axisymmetric'))
        opts.append((self._BODY_FORM_LABELS['wedge'], 'cone', 'wedge'))
        opts.append((self._BODY_FORM_LABELS['half_cone'], 'cone', 'half_cone'))
        return opts

    def _shape_label_for(self, shape_key, body_form_key, biconic=False):
        """The merged-selector label for a stored (shape, body_form, biconic):
        a lifting body_form wins (nose profile is moot for it); then the
        biconic flag (a two-cone reads as a shape); otherwise the nose
        profile.  So an ogive-nosed round body shows 'Tangent Ogive', a wedge
        shows 'Flattened wedge', and a biconic shows 'Biconic' regardless of
        the stored nose shape (re-saving normalizes it to 'cone', as the
        lifting forms already do)."""
        if body_form_key in ('wedge', 'half_cone'):
            return self._BODY_FORM_LABELS[body_form_key]
        if biconic:
            return self._BICONIC_LABEL
        return NOSE_SHAPE_LABELS.get(shape_key or 'cone', NOSE_SHAPE_LABELS['cone'])

    def _shape_key(self):
        """The ROParams.shape (nose-profile) key for the current selection."""
        disp = self._shape_var.get()
        return next((s for lbl, s, _f in self._shape_opts if lbl == disp), 'cone')

    def _body_form_key(self):
        """The ROParams.body_form key for the current Shape selection."""
        disp = self._shape_var.get()
        return next((f for lbl, _s, f in self._shape_opts if lbl == disp),
                    'axisymmetric')

    def _update_body_form_state(self):
        """Per-form field gating.  Biconic is its own Shape-dropdown entry
        (a body-of-revolution concept, exclusive with the lifting forms by
        construction); the selection drives the flag.  The planform-span
        field is WEDGE-only (body geometry); the wing rows are disabled FOR the wedge
        (its body is the lifting surface — hidden-but-active wing physics
        would be dishonest).  A half-cone keeps its wing rows: half-cone +
        delta wing is a real configuration (Fetterman TN D-2942)."""
        form = self._body_form_key()
        # Biconic is a Shape-dropdown entry (body-of-revolution only, so a
        # lifting selection is exclusive with it by construction): the
        # selection drives the flag and the fore/break field gating.
        self._biconic_var.set(form == "axisymmetric"
                              and self._shape_var.get() == self._BICONIC_LABEL)
        self._update_biconic_state()
        if hasattr(self, '_body_span_entry'):
            self._body_span_entry.config(
                state="normal" if form == "wedge" else "disabled")
        self._update_wing_state()

    # ---- TPS material dropdown helpers (§10 materials dropdown) --------
    _MAT_NONE_LABEL   = "(none — numbers only)"
    _MAT_CUSTOM_LABEL = "Custom…"
    _GROUP_TITLES = {"metal": "Metal / heat-sink", "hot_structure": "Hot structure",
                     "insulative": "Insulative tile", "ablative": "Ablator"}

    def _material_choices(self):
        """Return (display_list, {display: material_key}) for the grouped catalog,
        bracketed by "(none)" and "Custom…"."""
        by_group = heating.materials_by_group()
        displays = [self._MAT_NONE_LABEL]
        mapping = {self._MAT_NONE_LABEL: ""}
        for grp in heating.TPS_MATERIAL_GROUPS:
            title = self._GROUP_TITLES.get(grp, grp)
            for key, label in by_group.get(grp, []):
                disp = f"{title} · {label}"
                displays.append(disp)
                mapping[disp] = key
        displays.append(self._MAT_CUSTOM_LABEL)
        return displays, mapping

    def _display_for_key(self, key):
        """Reverse-map a stored material key to its dropdown display string."""
        if key in (heating.CUSTOM_NOSE_KEY, heating.CUSTOM_BODY_KEY):
            return self._MAT_CUSTOM_LABEL
        for disp, k in self._mat_map.items():
            if k == key and k != "":
                return disp
        return self._MAT_NONE_LABEL

    def _build_custom_fields(self, parent, row, cust):
        """Build the hidden per-location 'Custom…' property sub-frame; return
        (frame, vars_dict).  `cust` prefills from an existing custom dict."""
        cust = cust or {}
        frm = ttk.Frame(parent, padding=(16, 2, 0, 2))
        frm.grid(row=row, column=0, columnspan=2, sticky=tk.W)
        v = {}
        v['name'] = tk.StringVar(value=str(cust.get('label', '') or ''))
        v['ablator'] = tk.BooleanVar(value=bool(cust.get('is_ablator', False)))
        v['limit'] = tk.StringVar(
            value=str(cust.get('continuous_K') or cust.get('peak_K') or ''))
        v['density'] = tk.StringVar(value=str(cust.get('density_kg_m3', '') or ''))
        v['heff'] = tk.StringVar(value=str(cust.get('H_eff_MJ_kg', '') or ''))
        r = 0
        ttk.Label(frm, text="Name:").grid(row=r, column=0, sticky=tk.W, padx=(0, 6))
        ttk.Entry(frm, textvariable=v['name'], width=22).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Checkbutton(frm, text="Ablator (recedes)",
                        variable=v['ablator']).grid(row=r, column=1, sticky=tk.W, pady=1); r += 1
        ttk.Label(frm, text="Temp. limit (K):").grid(row=r, column=0, sticky=tk.W, padx=(0, 6))
        ttk.Entry(frm, textvariable=v['limit'], width=10).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Label(frm, text="Density (kg/m³):").grid(row=r, column=0, sticky=tk.W, padx=(0, 6))
        ttk.Entry(frm, textvariable=v['density'], width=10).grid(row=r, column=1, sticky=tk.W); r += 1
        ttk.Label(frm, text="Heat of ablation (MJ/kg):").grid(row=r, column=0, sticky=tk.W, padx=(0, 6))
        ttk.Entry(frm, textvariable=v['heff'], width=10).grid(row=r, column=1, sticky=tk.W); r += 1
        frm.grid_remove()
        return frm, v

    def _update_custom_state(self, loc):
        """Show the Custom… sub-frame only when that location's dropdown is Custom…."""
        var = self._nose_mat_var if loc == "nose" else self._body_mat_var
        frm = self._nose_cust_frm if loc == "nose" else self._body_cust_frm
        if var.get() == self._MAT_CUSTOM_LABEL:
            frm.grid()
        else:
            frm.grid_remove()

    def _custom_dict_from(self, v):
        """Build a custom-material dict from the sub-frame vars, or None if blank."""
        def _f(s):
            try:
                return float(s)
            except (TypeError, ValueError):
                return None
        limit = _f(v['limit'].get())
        if limit is None and not v['name'].get().strip():
            return None
        return {
            'label': v['name'].get().strip() or 'Custom material',
            'is_ablator': bool(v['ablator'].get()),
            'continuous_K': limit,
            'peak_K': limit,
            'density_kg_m3': _f(v['density'].get()),
            'H_eff_MJ_kg': _f(v['heff'].get()),
        }

    def _update_separation_state(self):
        """When the active reentry plan says 'body' (no separation), the airframe
        IS the last stage, so mass/diameter/length are inherited from the
        booster's Stage panel (read-only, "from booster") and the aero (β, L/D)
        is derived from the whole stack (0 = derive).  A separating RV owns all
        of these as designed inputs.  FRONT_END_DESIGN.md Part II §8–§10."""
        is_body = (self._plan_sep == 'body')
        state = 'disabled' if is_body else 'normal'
        for w in (getattr(self, '_mass_entry', None),
                  getattr(self, '_dia_entry', None),
                  getattr(self, '_len_entry', None)):
            if w is not None:
                w.configure(state=state)
        # Payload (body only): the added warhead/bus mass the front end owns,
        # on top of the inherited airframe burnout mass.  Shown just for a body,
        # where the mass field is the read-only airframe; a separating RV owns
        # its whole mass in the mass field, so the payload widgets stay hidden.
        for w in (getattr(self, '_payload_lbl', None),
                  getattr(self, '_payload_entry', None),
                  getattr(self, '_payload_unit', None),
                  getattr(self, '_payload_total_lbl', None)):
            if w is not None:
                w.pack(side=tk.LEFT) if is_body else w.pack_forget()
        if is_body:
            self._refresh_payload_total()
        # Body nose length is the inverse: meaningful ONLY for a body (it carves
        # the airframe's nose); a separating RV uses its own length instead.
        _bn = getattr(self, '_body_nose_entry', None)
        if _bn is not None:
            _bn.configure(state='normal' if is_body else 'disabled')
        # Reentry CG: body-only (it sets the reentry static margin).  Hide the
        # whole row for a separating RV, whose stability is its own designed
        # property rather than the airframe's.
        _cg_lbl = getattr(self, '_reentry_cg_lbl', None)
        _cg_ent = getattr(self, '_reentry_cg_entry', None)
        _cg_hint = getattr(self, '_reentry_cg_hint', None)
        if _cg_lbl is not None and _cg_ent is not None:
            if is_body:
                _cg_lbl.grid(row=16, column=0, sticky=tk.W, padx=(0, 8), pady=2)
                _cg_ent.grid()
                if _cg_hint is not None:
                    _cg_hint.grid(row=17, column=0, columnspan=3, sticky=tk.W)
            else:
                _cg_lbl.grid_remove(); _cg_ent.grid_remove()
                if _cg_hint is not None:
                    _cg_hint.grid_remove()
        # "0 = derive" hints / inline previews beside β and L/D — shown only for
        # a body, where the sentinel 0 triggers the geometry derivation; hidden
        # for a separating RV whose β/L-D are designed inputs.
        _bl = getattr(self, '_beta_derive_lbl', None)      # packed in _beta_row
        if _bl is not None:
            _bl.pack(side=tk.LEFT) if is_body else _bl.pack_forget()
        if is_body:
            self._refresh_beta_preview()
        # L/D: an Estimate button (row 0, col 2) + a live derived-value label
        # (row 1, spanning) — the inline "0 → derives ~2.57" preview.
        _lb = getattr(self, '_LD_est_btn', None)
        _ll = getattr(self, '_LD_derive_lbl', None)
        if _lb is not None:
            if is_body:
                _lb.grid(row=0, column=2, sticky=tk.W, padx=(8, 0))
            else:
                _lb.grid_remove()
        if _ll is not None:
            if is_body:
                _ll.grid(row=1, column=0, columnspan=3, sticky=tk.W)
            else:
                _ll.grid_remove()
        if is_body:
            self._refresh_ld_preview()

    def _preview_ro(self):
        """A lightweight ROParams from the LIVE editor fields (no validation
        dialogs) for the inline body-mode L/D preview.  Only the fields the trim
        gate reads matter (shape, nose taper, reentry CG); mass/diameter are
        inherited from the booster stage at run time."""
        def _f(var, d=0.0):
            try:
                return float(var.get() or 0.0)
            except (ValueError, tk.TclError):
                return d
        def _fv(name):
            var = getattr(self, name, None)
            return _f(var) if var is not None else 0.0
        # Mirror _build_ro's shape resolution: the merged dropdown stores a
        # biconic as shape='cone' + the biconic flag (axisymmetric only), and
        # the wing planform rides along so the trim gate judges the object as
        # built, not a stripped stand-in.
        _bic = (bool(getattr(self, '_biconic_var', None)
                     and self._biconic_var.get())
                and self._body_form_key() == "axisymmetric")
        return mm.ROParams(
            name="preview", mass_kg=max(_f(self._mass_var, 1.0), 1.0),
            beta_kg_m2=0.0, shape=self._shape_key(),
            diameter_m=_f(self._dia_var), length_m=_f(self._len_var),
            nose_radius_m=_f(self._nose_var, 0.0),
            biconic=_bic,
            fore_length_m=_fv('_fore_len_var'),
            break_diameter_m=_fv('_break_dia_var'),
            wing_area_m2=_fv('_wing_area_var'),
            wing_aspect_ratio=_fv('_wing_ar_var'),
            wing_sweep_deg=_fv('_wing_sweep_var'),
            wing_root_chord_m=_fv('_wing_root_var'),
            wing_span_exposed_m=_fv('_wing_span_var'),
            body_nose_length_m=_fv('_body_nose_var'),
            reentry_cg_m=_fv('_reentry_cg_var'),
            payload_kg=_fv('_payload_var'),
            separation_mode='body', glider_enabled=True, glider_LD=0.0)

    def _refresh_ld_preview(self):
        """Update the inline L/D preview next to the field (body mode): the
        geometry-derived, trim-gated L/D the run will use, composed from the live
        object onto the current booster."""
        lbl = getattr(self, '_LD_derive_lbl', None)
        if lbl is None or self._plan_sep != 'body':
            return
        try:
            _ld = float(self._LD_var.get() or 0.0)
        except (ValueError, tk.TclError):
            _ld = 0.0
        if _ld > 0.0:
            lbl.config(text="typed value — clear to 0 to derive from geometry")
            return
        if self._booster is None:
            lbl.config(text="0 = derive from geometry (nose + body + fins)")
            return
        try:
            import glider_ld as _gld
            import trim_gate as _tg
            ro = self._preview_ro()
            p = mm.compose_loadout(self._booster, ro, 1)
            p.ro = ro
            _cg = float(getattr(ro, 'reentry_cg_m', 0.0) or 0.0)
            g = _tg.trim_gate(p, mach=_gld.GLIDE_MACH_REF,
                              x_cg_m=(_cg if _cg > 0.0 else None))
            if g.get('error'):
                lbl.config(text="0 = derive — set body geometry / fins")
                return
            _sm = g['static_margin_cal']
            if g['LD_achievable'] <= 0.0:
                # Unstable → the run tumbles, but still report the aero
                # (L/D)_max the shape could reach if trimmed — the estimate
                # must always return a value, plus the lever that unlocks it.
                _lm = float(g.get('LD_max', 0.0) or 0.0)
                _tail = (f" — aero max ~{_lm:.2f} if trimmed; "
                         f"set Reentry CG forward" if _lm > 0.0 else "")
                lbl.config(text=f"0 → tumbles — unstable "
                                f"(SM {_sm:+.1f} cal){_tail}")
            else:
                lbl.config(text=f"0 → derives ~{g['LD_achievable']:.2f} "
                                f"(SM {_sm:+.1f} cal, M{_gld.GLIDE_MACH_REF:.0f})")
        except Exception:
            lbl.config(text="0 = derive from geometry")

    def _refresh_beta_preview(self):
        """Update the inline β preview next to the field (body mode): the
        geometry-derived ballistic coefficient the run will use, composed from
        the live object onto the current booster.  This mirrors the L/D preview
        and the runtime derivation (trajectory.py §β-derive): a body flagged
        unstable by the trim gate tumbles, so the spinning-cylinder β wins over
        the nose-first β(Mach) table."""
        lbl = getattr(self, '_beta_derive_lbl', None)
        if lbl is None or self._plan_sep != 'body':
            return
        try:
            _b = float(self._beta_var.get() or 0.0)
        except (ValueError, tk.TclError):
            _b = 0.0
        if _b > 0.0:
            lbl.config(text="  typed value — clear to 0 to derive from geometry")
            return
        if self._booster is None:
            lbl.config(text="  0 = derive β(Mach) from the airframe")
            return
        try:
            import glider_ld as _gld
            import trim_gate as _tg
            import dataclasses as _dc
            import math as _math
            ro = self._preview_ro()
            p = mm.compose_loadout(self._booster, ro, 1)
            p.ro = ro
            _cg = float(getattr(ro, 'reentry_cg_m', 0.0) or 0.0)
            g = _tg.trim_gate(p, mach=_gld.GLIDE_MACH_REF,
                              x_cg_m=(_cg if _cg > 0.0 else None))
            # Unstable → the run forces tumbling; the spinning-cylinder β (from
            # effective_ro's tumbling branch) is what actually flies, not the
            # nose-first table.  Report that so the preview stays honest.
            if (not g.get('error')) and g.get('LD_achievable', 0.0) <= 0.0:
                p_t = mm.compose_loadout(self._booster, ro, 1)
                p_t.ro = _dc.replace(ro, reentry_attitude='tumbling')
                _et = mm.effective_ro(p_t)
                _bt = float(getattr(_et, 'beta_kg_m2', 0.0) or 0.0)
                _sm = g['static_margin_cal']
                if _bt > 0.0:
                    lbl.config(text=f"  0 → tumbles: β ~{_bt:,.0f} kg/m² "
                                    f"(unstable, SM {_sm:+.1f} cal)")
                else:
                    lbl.config(text=f"  0 → tumbles (unstable, SM {_sm:+.1f} cal)")
                return
            # Nose-first: β(Mach) = m / (Cd0_body(M)·A_base), sampled at the
            # reference Mach with its M2–12 spread (matches trajectory.py).
            _last = _gld._last_stage(p)
            _mass = float(mm.effective_ro(p).mass_kg)
            _A = _math.pi * (float(_last.diameter_m) / 2.0) ** 2
            if _A <= 0.0 or _mass <= 0.0:
                lbl.config(text="  0 = derive — set body geometry")
                return

            def _bfn(M):
                _cd = float(_gld.body_cd0(p, M))
                return _mass / (_cd * _A) if _cd > 0.0 else 0.0
            b_ref = _bfn(_gld.GLIDE_MACH_REF)
            b_lo, b_hi = _bfn(2.0), _bfn(12.0)
            if b_ref <= 0.0:
                lbl.config(text="  0 = derive from geometry")
                return
            lbl.config(text=f"  0 → derives ~{b_ref:,.0f} kg/m²  "
                            f"({min(b_lo, b_hi):,.0f}–{max(b_lo, b_hi):,.0f} "
                            f"over M2–12)")
        except Exception:
            lbl.config(text="  0 = derive β(Mach) from geometry")

    def _refresh_payload_total(self):
        """Live total-reentry-mass line for a body: airframe burnout (the
        inherited, read-only mass field) + the added payload = what the run
        actually reenters with."""
        lbl = getattr(self, '_payload_total_lbl', None)
        if lbl is None or self._plan_sep != 'body':
            return
        def _f(var):
            try:
                return max(0.0, float(var.get() or 0.0))
            except (ValueError, tk.TclError):
                return 0.0
        airframe = _f(self._mass_var)
        payload = _f(self._payload_var)
        if payload > 0.0:
            lbl.config(text=f"  = {airframe + payload:,.0f} kg reentry "
                            f"({airframe:,.0f} airframe + {payload:,.0f})")
        else:
            lbl.config(text=f"  = {airframe:,.0f} kg reentry (airframe only)")

    def _show_ld_estimate(self):
        """The L/D Estimate button (body mode): the full whole-airframe
        build-up in a dialog — Cd0, C_Nα, (L/D)_max, best-glide α — plus the
        trim-gate verdict.  Unlike the passive inline label this ALWAYS
        returns a value (the aero maximum), even when a typed L/D sits in the
        field or the gate says the body tumbles; the verdict line then says
        why the run won't reach it.  (The former sidebar "Estimate body
        L/D…" dialog, moved next to the field it informs.)"""
        self._refresh_ld_preview()          # keep the inline label in step
        if self._booster is None:
            messagebox.showinfo(
                "L/D Estimate",
                "No booster in context — the body L/D is derived from the "
                "whole airframe (stage + nose + fins).", parent=self)
            return
        try:
            import glider_ld as _gld
            import trim_gate as _tg
            ro = self._preview_ro()
            p = mm.compose_loadout(self._booster, ro, 1)
            p.ro = ro
            r = _gld.whole_booster_LD(p, mach=_gld.GLIDE_MACH_REF)
        except Exception as exc:
            messagebox.showerror("L/D Estimate",
                                 f"Estimate failed:\n{exc}", parent=self)
            return
        if r.get("error"):
            messagebox.showinfo("L/D Estimate", r["error"], parent=self)
            return
        fins = (f" + fins {r['c_na_fin']:.2f}" if r["fin_planform_m2"] > 0
                else " (body only)")
        # Trim-gate verdict: whether the run can actually FLY that L/D.
        _verdict = ""
        try:
            _cg = float(getattr(ro, 'reentry_cg_m', 0.0) or 0.0)
            g = _tg.trim_gate(p, mach=_gld.GLIDE_MACH_REF,
                              x_cg_m=(_cg if _cg > 0.0 else None))
            if not g.get('error'):
                _verdict = (f"\nTrim gate (SM {g['static_margin_cal']:+.1f} "
                            f"cal): {g['verdict']}")
        except Exception:
            pass
        messagebox.showinfo(
            "Whole-body L/D estimate (Jorgensen + Allen-Perkins + N-K-P)",
            f"Mach {_gld.GLIDE_MACH_REF:.0f}   |   body Cd₀ = {r['cd0']:.3f}\n"
            f"C_Nα (potential) = {r['c_na_pot']:.2f} /rad  "
            f"(body {r['c_na_body']:.2f}{fins})\n"
            f"wing-body factor (1+r/s)² = {r['k_sum']:.2f}\n"
            f"\n"
            f"Max L/D ≈ {r['ld_max']:.2f}   at α ≈ {r['alpha_deg']:.0f}°\n"
            f"{_verdict}\n"
            f"\n"
            f"Leave the L/D field at 0 to fly the derived, trim-gated value.",
            parent=self)

    def _ok(self):
        ro = self._build_ro()
        if ro is None:
            return
        self._result = ro
        self.destroy()

    def _save_to_library(self):
        """Validate + write to a .ro.json file in the library, then close."""
        ro = self._build_ro()
        if ro is None:
            return
        from tkinter import filedialog
        _safe_name = "".join(c if c.isalnum() or c in "-_" else "_"
                             for c in ro.name).strip("_") or "RO"
        path = filedialog.asksaveasfilename(
            parent=self,
            title="Save Reentry Object to Library",
            initialdir=str(_ensure_dir(_RO_LIBRARY_PATH)),
            initialfile=f"{_safe_name}.ro.json",
            defaultextension=".json",
            filetypes=[("reentry-object files (*.ro.json)", "*.json"), ("All files", "*.*")])
        if not path:
            return
        try:
            # Hardware-only object file; the reentry plan is written by name to
            # the reentry-plan library, where _load_ro_library merges it back on.
            Path(path).write_text(
                json.dumps(ro_to_dict(ro, include_reentry_plan=False), indent=2))
            save_reentry_plan(ro.name, extract_reentry_plan(ro),
                              _REENTRY_PLAN_LIBRARY_PATH)
        except Exception as exc:
            messagebox.showerror("Save Reentry Object",
                                 f"Could not write reentry-object file:\n{exc}",
                                 parent=self)
            return
        self._result = ro
        self.destroy()

    @property
    def result(self):
        return self._result


# ---------------------------------------------------------------------------
# Reentry-plan editor dialog — the "how it's flown" half of a reentry object,
# the down-leg analogue of FlightPlanDialog.  It edits the reentry-plan FILE
# fields that are NOT quick picks on the strip: the commanded L/D (clamped to
# the airframe's L/D capability — fly it worse, never better), the pull-up
# g-limit and re-entry βₛ, the control-flap deflection, the reentry attitude,
# the ζ damping/tracking knob (with its estimator), the bank schedule, the
# dive-at-target trigger, and provenance.  The sidebar strip keeps the quick
# run-to-run picks: glide law, separation, terminal-dive altitude, aero model
# and skip count.
# ---------------------------------------------------------------------------

class ReentryPlanDialog(tk.Toplevel):
    """Edit the reentry plan for one reentry object; returns a partial plan dict
    (only the fields this dialog owns) to merge over the active plan."""

    def __init__(self, parent, title, plan, ld_capability):
        super().__init__(parent)
        self._app = parent          # for the ζ estimator and target picker
        self._result = None
        self._cap = float(ld_capability or 0.0)
        # The plan's integration family (its identity) decides which tuning
        # fields exist: ζ / banks / dive-at-target are NUMERICAL capabilities;
        # β_S (Acton Phase 3) is ANALYTIC.  See REENTRY_FAMILY_DESIGN.md.
        self._family = mm.glide_family(plan.get('glider_guidance'))
        self.title(f"Reentry Plan — {title}")
        self.transient(parent)
        self.resizable(False, False)
        self.grab_set()

        frm = ttk.Frame(self, padding=12)
        frm.pack(fill=tk.BOTH, expand=True)
        frm.columnconfigure(1, weight=1)
        r = 0

        ttk.Label(frm, text=f"Integration family: "
                            f"{_FAMILY_LABELS[self._family]}  "
                            f"(fixed for this plan)",
                  foreground="#888888").grid(
            row=r, column=0, columnspan=2, sticky=tk.W, pady=(0, 6)); r += 1

        def _f(key, default):
            v = plan.get(key)
            return default if v is None else v

        # Commanded L/D — clamped to the airframe capability on save.
        ttk.Label(frm, text="Commanded L/D:").grid(row=r, column=0, sticky=tk.W, pady=3)
        _cmd = _f('commanded_LD', self._cap)
        self._cmd_ld_var = tk.StringVar(value=f"{float(_cmd):g}")
        _cf = ttk.Frame(frm); _cf.grid(row=r, column=1, sticky=tk.W, pady=3)
        ttk.Entry(_cf, textvariable=self._cmd_ld_var, width=10).pack(side=tk.LEFT)
        ttk.Label(_cf, text=f"  (≤ {self._cap:g} airframe max — fly it worse, not better)",
                  foreground="#888888").pack(side=tk.LEFT); r += 1

        ttk.Label(frm, text="Pull-up g-limit:").grid(row=r, column=0, sticky=tk.W, pady=3)
        self._pullup_var = tk.StringVar(value=f"{float(_f('glider_pullup_g_max', 10.0)):g}")
        _pf = ttk.Frame(frm); _pf.grid(row=r, column=1, sticky=tk.W, pady=3)
        ttk.Entry(_pf, textvariable=self._pullup_var, width=10).pack(side=tk.LEFT)
        ttk.Label(_pf, text="  g", foreground="#888888").pack(side=tk.LEFT); r += 1

        # β_S is the Acton Phase-3 direct-reentry ballistic coefficient — an
        # ANALYTIC-family field; the numerical EOM never reads it.
        self._beta_s_var = tk.StringVar(value=f"{float(_f('glider_beta_entry_kg_m2', 0.0)):g}")
        if self._family == 'analytic':
            ttk.Label(frm, text="Re-entry βₛ:").grid(row=r, column=0, sticky=tk.W, pady=3)
            _bf = ttk.Frame(frm); _bf.grid(row=r, column=1, sticky=tk.W, pady=3)
            ttk.Entry(_bf, textvariable=self._beta_s_var, width=10).pack(side=tk.LEFT)
            ttk.Label(_bf, text="  kg/m²  (Acton Phase 3; 0 = Tracy)",
                      foreground="#888888").pack(side=tk.LEFT); r += 1

        ttk.Label(frm, text="Flap deflection:").grid(row=r, column=0, sticky=tk.W, pady=3)
        self._flap_var = tk.StringVar(value=f"{float(_f('glider_flap_deflection_deg', 0.0)):g}")
        _ff = ttk.Frame(frm); _ff.grid(row=r, column=1, sticky=tk.W, pady=3)
        ttk.Entry(_ff, textvariable=self._flap_var, width=10).pack(side=tk.LEFT)
        ttk.Label(_ff, text="  °  (0 = 12° default)",
                  foreground="#888888").pack(side=tk.LEFT); r += 1

        # Reentry attitude — trimmed (controlled) vs tumbling (uncontrolled).
        # Tumbling zeroes lift and derives β from geometry as a tumbling
        # cylinder (two-orientation Hoerner form); trim keeps the aeroshell β.
        # For a non-separating body the run-time static-margin gate flags
        # SM ≤ 0 and forces tumbling regardless of this setting.
        ttk.Label(frm, text="Reentry attitude:").grid(row=r, column=0, sticky=tk.W, pady=3)
        self._ATT_LABELS = {'trim': "Trimmed (controlled)",
                            'tumbling': "Tumbling (uncontrolled)"}
        _att = str(_f('reentry_attitude', 'trim'))
        self._att_var = tk.StringVar(
            value=self._ATT_LABELS.get(_att, self._ATT_LABELS['trim']))
        _af = ttk.Frame(frm); _af.grid(row=r, column=1, sticky=tk.W, pady=3)
        ttk.Combobox(_af, textvariable=self._att_var,
                     values=list(self._ATT_LABELS.values()),
                     state="readonly", width=22).pack(side=tk.LEFT)
        ttk.Label(_af, text="  (tumbling: L/D = 0, β derived from geometry;"
                            "\n   unstable bodies tumble regardless — SM gate)",
                  foreground="#888888", justify=tk.LEFT).pack(side=tk.LEFT); r += 1

        # ζ, bank schedule and dive-at-target are NUMERICAL-family capabilities
        # (the closed-form analytic laws cannot bank, steer to a target, or
        # damp a phugoid — there isn't one).  Vars always exist so _save can
        # read them; the widgets are built only for a numerical plan.
        _guid = str(plan.get('glider_guidance', '') or '')
        # Terminal-dive altitude and aero model are set-once tuning that live
        # in this editor; ζ (the per-run knob) is on the sidebar strip and is
        # NOT edited here — _save omits it so the merge preserves the strip's
        # value.  Vars init from the plan.
        self._dive_alt_var = tk.StringVar(
            value=f"{float(_f('glider_terminal_alt_km', 0.0)):g}")
        self._pullup_alt_var = tk.StringVar(
            value=f"{float(_f('glider_pullup_start_alt_km', 0.0)):g}")
        self._aero_var = tk.StringVar(
            value=("Drag polar (realistic)"
                   if str(_f('glider_aero_model', 'polar')) == 'polar'
                   else "Fixed L/D (idealized)"))
        _sched = list(_f('glider_bank_schedule', []) or [])
        self._bank_vars = [{'start': tk.StringVar(), 'end': tk.StringVar(),
                            'bank': tk.StringVar()} for _ in range(3)]
        for _i, _bv in enumerate(self._bank_vars):
            if _i < len(_sched):
                try:
                    _s, _e, _b = _sched[_i]
                    _bv['start'].set(f"{float(_s):g}"); _bv['end'].set(f"{float(_e):g}")
                    _bv['bank'].set(f"{float(_b):g}")
                except (ValueError, TypeError):
                    pass
        _dt_on = float(_f('glider_dive_target_radius_km', 0.0) or 0.0) > 0.0
        self._dt_on_var = tk.BooleanVar(value=_dt_on)
        self._dt_lat_var = tk.StringVar(value=f"{float(_f('glider_dive_target_lat_deg', 0.0)):g}")
        self._dt_lon_var = tk.StringVar(value=f"{float(_f('glider_dive_target_lon_deg', 0.0)):g}")
        self._dt_rad_var = tk.StringVar(
            value=f"{float(_f('glider_dive_target_radius_km', 20.0)) or 20.0:g}")

        if self._family == 'numerical':
            # Commanded pull-up — fall at zero lift to this altitude, then
            # pull at full authority (≤ g-cap) and hand off to the glide law.
            ttk.Label(frm, text="Pull-up start:").grid(
                row=r, column=0, sticky=tk.W, pady=3)
            _pf = ttk.Frame(frm); _pf.grid(row=r, column=1, sticky=tk.W, pady=3)
            ttk.Entry(_pf, textvariable=self._pullup_alt_var, width=8).pack(side=tk.LEFT)
            ttk.Label(_pf, text="  km  (0 = capture by glide law)",
                      foreground="#888888").pack(side=tk.LEFT); r += 1

            # Terminal dive — glide until this altitude, then pitch into a
            # steep terminal dive (0 = glide all the way to impact).  A
            # numerical-EOM capability; the closed-form laws fly their arc.
            ttk.Label(frm, text="Terminal dive below:").grid(
                row=r, column=0, sticky=tk.W, pady=3)
            _tf = ttk.Frame(frm); _tf.grid(row=r, column=1, sticky=tk.W, pady=3)
            ttk.Entry(_tf, textvariable=self._dive_alt_var, width=8).pack(side=tk.LEFT)
            ttk.Label(_tf, text="  km  (0 = glide to impact)",
                      foreground="#888888").pack(side=tk.LEFT); r += 1

            # Aero model — realistic drag polar vs an idealized fixed L/D.
            ttk.Label(frm, text="Aero model:").grid(
                row=r, column=0, sticky=tk.W, pady=3)
            _ef = ttk.Frame(frm); _ef.grid(row=r, column=1, sticky=tk.W, pady=3)
            ttk.Combobox(_ef, textvariable=self._aero_var,
                         values=["Drag polar (realistic)",
                                 "Fixed L/D (idealized)"],
                         state="readonly", width=22).pack(side=tk.LEFT); r += 1

            # Bank schedule — up to three (start s, end s, bank °) segments.
            ttk.Label(frm, text="Bank schedule:").grid(row=r, column=0, sticky=tk.NW, pady=3)
            _bkf = ttk.Frame(frm); _bkf.grid(row=r, column=1, sticky=tk.W, pady=3)
            for _mc, _hdr in enumerate(["start s", "end s", "bank °"], start=1):
                ttk.Label(_bkf, text=_hdr, foreground="#888888").grid(
                    row=0, column=_mc, padx=3, pady=(0, 1))
            for _i, _bv in enumerate(self._bank_vars):
                ttk.Label(_bkf, text=f"#{_i+1}").grid(row=_i+1, column=0, sticky=tk.W, padx=(0, 4))
                for _mc, _k in enumerate(['start', 'end', 'bank'], start=1):
                    ttk.Entry(_bkf, textvariable=_bv[_k], width=7).grid(
                        row=_i+1, column=_mc, padx=3, pady=1)
            r += 1

            # Dive-at-target — steer to a lat/lon then dive (radius = trigger).
            ttk.Checkbutton(frm, text="Dive at target (lat/lon)",
                            variable=self._dt_on_var).grid(
                row=r, column=0, sticky=tk.W, pady=(6, 1)); r += 1
            _dtf = ttk.Frame(frm); _dtf.grid(row=r, column=0, columnspan=2,
                                             sticky=tk.W, padx=(16, 0), pady=(0, 3))
            ttk.Label(_dtf, text="Lat:").pack(side=tk.LEFT)
            ttk.Entry(_dtf, textvariable=self._dt_lat_var, width=8).pack(side=tk.LEFT, padx=2)
            ttk.Label(_dtf, text="°  Lon:").pack(side=tk.LEFT, padx=(4, 0))
            ttk.Entry(_dtf, textvariable=self._dt_lon_var, width=8).pack(side=tk.LEFT, padx=2)
            ttk.Label(_dtf, text="°  Radius:").pack(side=tk.LEFT, padx=(4, 0))
            ttk.Entry(_dtf, textvariable=self._dt_rad_var, width=6).pack(side=tk.LEFT, padx=2)
            ttk.Label(_dtf, text="km").pack(side=tk.LEFT)
            ttk.Button(_dtf, text="Find…", width=7,
                       command=lambda: self._app._pick_location(
                           self._dt_lat_var, self._dt_lon_var)).pack(side=tk.LEFT, padx=(6, 0))
            r += 1
        else:
            ttk.Label(frm, text="(Closed-form analytic: constant L/D, no banking,\n"
                                "no dive-at-target, no phugoid ζ — the pull-up arc\n"
                                "and glide are the Tracy/Acton formulas.)",
                      foreground="#888888", justify=tk.LEFT).grid(
                row=r, column=0, columnspan=2, sticky=tk.W, pady=(6, 1)); r += 1

        ttk.Label(frm, text="(Glide law, separation and the ζ damping/tracking\n"
                            "knob are on the sidebar strip.)",
                  foreground="#888888", justify=tk.LEFT).grid(
            row=r, column=0, columnspan=2, sticky=tk.W, pady=(6, 4)); r += 1

        ttk.Label(frm, text="Source:").grid(row=r, column=0, sticky=tk.W, pady=3)
        self._source_var = tk.StringVar(value=str(plan.get('source', '') or ''))
        ttk.Entry(frm, textvariable=self._source_var).grid(
            row=r, column=1, sticky=tk.EW, pady=3); r += 1
        ttk.Label(frm, text="Notes:").grid(row=r, column=0, sticky=tk.NW, pady=3)
        self._notes_text = tk.Text(frm, height=3, width=40, wrap=tk.WORD)
        self._notes_text.grid(row=r, column=1, sticky=tk.EW, pady=3)
        self._notes_text.insert("1.0", str(plan.get('notes', '') or '')); r += 1

        bf = ttk.Frame(frm)
        bf.grid(row=r, column=0, columnspan=2, sticky=tk.E, pady=(10, 0))
        ttk.Button(bf, text="Cancel", command=self.destroy).pack(side=tk.RIGHT, padx=(4, 0))
        ttk.Button(bf, text="Save", command=self._save).pack(side=tk.RIGHT)

    def _save(self):
        def _num(sv, default):
            try:
                return float(sv.get().strip())
            except (ValueError, AttributeError):
                return default
        cmd = _num(self._cmd_ld_var, self._cap)
        if self._cap > 0:
            cmd = min(cmd, self._cap)          # clamp: fly it worse, never better
        # Bank schedule: keep only fully-filled rows.
        _banks = []
        for _bv in self._bank_vars:
            _s, _e, _b = (_bv['start'].get().strip(), _bv['end'].get().strip(),
                          _bv['bank'].get().strip())
            if _s and _e and _b:
                try:
                    _banks.append([float(_s), float(_e), float(_b)])
                except ValueError:
                    pass
        # Dive-at-target: radius 0 disables it.
        _dt_rad = _num(self._dt_rad_var, 0.0) if self._dt_on_var.get() else 0.0
        self._result = {
            'commanded_LD':             cmd,
            'glider_pullup_g_max':      _num(self._pullup_var, 10.0),
            'glider_beta_entry_kg_m2':  _num(self._beta_s_var, 0.0),
            'glider_flap_deflection_deg': _num(self._flap_var, 0.0),
            'reentry_attitude': ('tumbling'
                                 if self._att_var.get()
                                 == self._ATT_LABELS['tumbling']
                                 else 'trim'),
            # ζ is owned by the sidebar strip — deliberately omitted so the
            # merge over the (strip-snapshotted) plan preserves it.
            'glider_terminal_alt_km':   max(0.0, _num(self._dive_alt_var, 0.0)),
            'glider_pullup_start_alt_km': max(0.0, _num(self._pullup_alt_var, 0.0)),
            'glider_aero_model': ('polar'
                                  if 'polar' in self._aero_var.get().lower()
                                  else 'constant_LD'),
            'glider_bank_schedule':     _banks,
            'glider_dive_target_lat_deg':   _num(self._dt_lat_var, 0.0),
            'glider_dive_target_lon_deg':   _num(self._dt_lon_var, 0.0),
            'glider_dive_target_radius_km': _dt_rad,
            'source': self._source_var.get().strip(),
            'notes':  self._notes_text.get("1.0", "end-1c").strip(),
        }
        self.destroy()

    @property
    def result(self):
        return self._result


# ---------------------------------------------------------------------------
# Flight-plan editor dialog — the "how it's flown" half of a booster, given the
# same New/Edit/Delete dialog treatment as boosters and reentry objects.  It
# edits the flight-plan FILE fields: ascent mode, launch elevation, the
# per-stage pitch schedule (turn start/stop, burnout angle, coast, end-burn),
# and the subsystem-deployment events (fairing / booster drop / grid-fin),
# each shown only when the booster actually carries that hardware, plus
# provenance.  Yaw/dogleg stays on the sidebar for now (a separate global
# representation) and is reconciled when the profile store is unified.
# ---------------------------------------------------------------------------

class FlightPlanDialog(tk.Toplevel):
    """Edit the flight plan for one booster; returns an updated plan dict."""

    def __init__(self, parent, booster_name, plan, booster):
        super().__init__(parent)
        self._result = None
        self._booster = booster
        self._name = booster_name
        self.title(f"Flight Plan — {booster_name}")
        self.transient(parent)
        self.resizable(False, False)

        # Stage chain (root first) — drives the per-stage rows and which events
        # apply.
        self._stages = []
        _n = booster
        while _n is not None:
            self._stages.append(_n)
            _n = getattr(_n, 'stage2', None)

        frm = ttk.Frame(self, padding=12)
        frm.pack(fill=tk.BOTH, expand=True)
        r = 0

        # ── Mode ─────────────────────────────────────────────────────────
        # The guidance LAW is the plan's identity, fixed when the plan was
        # created (New Flight Plan) — it cannot be changed here.  For a
        # pitch-program plan the combobox still toggles Simple/Advanced (same
        # law, different parameterisation); for gravity-turn / orbital plans it
        # is a fixed, disabled display.
        ttk.Label(frm, text="Mode:").grid(row=r, column=0, sticky=tk.W, pady=3)
        self._mode_var = tk.StringVar()
        _g = str(plan.get('guidance', 'pitch_program'))
        # advanced pitch = any per-stage angle override present
        _adv = any(s.get('stage_burnout_angle_deg') is not None
                   for s in plan.get('stages', []))
        self._mode_var.set(
            "Advanced pitch profile" if (_g == 'pitch_program' and _adv) else
            "Simple pitch profile"   if _g == 'pitch_program' else
            "Gravity turn"           if _g == 'true_gravity_turn' else
            "Orbital insertion"      if _g == 'orbital_insertion' else
            "Simple pitch profile")
        _law_choices = (["Simple pitch profile", "Advanced pitch profile"]
                        if _g not in ("true_gravity_turn", "orbital_insertion")
                        else [self._mode_var.get()])
        self._mode_cb = ttk.Combobox(
            frm, textvariable=self._mode_var, width=22,
            state=("readonly" if len(_law_choices) > 1 else "disabled"),
            values=_law_choices)
        self._mode_cb.grid(row=r, column=1, sticky=tk.W, pady=3)
        self._mode_cb.bind("<<ComboboxSelected>>", self._on_mode_changed)
        ttk.Label(frm, text="(law fixed when the plan was created)",
                  foreground="#888888").grid(row=r, column=2, sticky=tk.W,
                                             padx=(6, 0), pady=3)
        r += 1

        # ── Launch elevation (always shown) ──────────────────────────────
        ttk.Label(frm, text="Launch elev. (°):").grid(row=r, column=0, sticky=tk.W, pady=3)
        self._launch_el_var = tk.StringVar(value=f"{float(plan.get('launch_elevation_deg', 90.0)):g}")
        ttk.Entry(frm, textvariable=self._launch_el_var, width=10).grid(
            row=r, column=1, sticky=tk.W, pady=3); r += 1

        # ── Burnout / boost angle (mode-dependent) + Wheelon estimator ───
        self._burnout_lbl = ttk.Label(frm, text="Burnout angle (°):")
        self._burnout_lbl.grid(row=r, column=0, sticky=tk.W, pady=3)
        self._burnout_frame = ttk.Frame(frm)
        self._burnout_frame.grid(row=r, column=1, sticky=tk.W, pady=3)
        self._burnout_var = tk.StringVar(value=f"{float(plan.get('burnout_angle_deg', 45.0)):g}")
        ttk.Entry(self._burnout_frame, textvariable=self._burnout_var, width=10).pack(side=tk.LEFT)
        self._wheelon_btn = ttk.Button(self._burnout_frame, text="Estimate ε*",
                                       command=self._estimate_wheelon)
        self._wheelon_btn.pack(side=tk.LEFT, padx=6); r += 1

        # ── Global turn window (used by simple / gravity / orbital) ──────
        self._turn_start_lbl = ttk.Label(frm, text="Turn start (s):")
        self._turn_start_lbl.grid(row=r, column=0, sticky=tk.W, pady=3)
        self._turn_start_var = tk.StringVar(
            value=("" if plan.get('gt_turn_start_s') is None
                   else f"{float(plan.get('gt_turn_start_s')):g}"))
        self._turn_start_ent = ttk.Entry(frm, textvariable=self._turn_start_var, width=10)
        self._turn_start_ent.grid(row=r, column=1, sticky=tk.W, pady=3); r += 1
        self._turn_stop_lbl = ttk.Label(frm, text="Turn stop (s):")
        self._turn_stop_lbl.grid(row=r, column=0, sticky=tk.W, pady=3)
        self._turn_stop_var = tk.StringVar(
            value=("" if plan.get('gt_turn_stop_s') is None
                   else f"{float(plan.get('gt_turn_stop_s')):g}"))
        self._turn_stop_ent = ttk.Entry(frm, textvariable=self._turn_stop_var, width=10)
        self._turn_stop_ent.grid(row=r, column=1, sticky=tk.W, pady=3); r += 1

        self._mode_note = ttk.Label(frm, text="", foreground="#555555")
        self._mode_note.grid(row=r, column=0, columnspan=2, sticky=tk.W, padx=(0, 4)); r += 1

        # ── Per-stage pitch table (mode-dependent) ───────────────────────
        self._pstage_lbl = ttk.Label(frm, text="Per-stage pitch program",
                                     font=("TkDefaultFont", 10, "bold"))
        self._pstage_lbl.grid(row=r, column=0, columnspan=2, sticky=tk.W, pady=(10, 2)); r += 1
        tbl = ttk.Frame(frm)
        self._pstage_tbl = tbl
        tbl.grid(row=r, column=0, columnspan=2, sticky=tk.W); r += 1
        self._angle_hdr = None
        for c, h in enumerate(["Stage", "Turn start", "Turn stop", "Angle",
                               "Coast", "End-burn"]):
            _hl = ttk.Label(tbl, text=h, foreground="#555555")
            _hl.grid(row=0, column=c, padx=3, pady=(2, 1))
            if h == "Angle":
                self._angle_hdr = _hl
        self._stage_rows = []
        for i, st in enumerate(plan.get('stages', [])):
            ttk.Label(tbl, text=f"{i + 1}").grid(row=i + 1, column=0, padx=3)
            row = {}
            for c, key in enumerate(['stage_turn_start_s', 'stage_turn_stop_s',
                                     'stage_burnout_angle_deg', 'coast_time_s',
                                     'stage_cutoff_s'], start=1):
                v = st.get(key)
                sv = tk.StringVar(value="" if v is None else f"{float(v):g}")
                ttk.Entry(tbl, textvariable=sv, width=8).grid(
                    row=i + 1, column=c, padx=3, pady=1)
                row[key] = sv
            self._stage_rows.append(row)
        self._on_mode_changed()   # apply initial mode visibility

        # ── Events (only for hardware the booster actually carries) ──────
        self._ev_vars = {}
        _ev = ttk.LabelFrame(frm, text="Deployment events", padding=6)
        _has_ev = False
        er = 0
        if getattr(booster, 'shroud_mass_kg', 0.0) > 0:
            ttk.Label(_ev, text="Fairing jettison (km):").grid(row=er, column=0, sticky=tk.W, pady=2)
            _sj = float(plan.get('shroud_jettison_alt_km', 0.0) or 0.0)
            self._ev_vars['shroud'] = tk.StringVar(value=f"{_sj:g}" if _sj > 0 else "")
            ttk.Entry(_ev, textvariable=self._ev_vars['shroud'], width=10).grid(row=er, column=1, sticky=tk.W)
            ttk.Label(_ev, text="blank = heating", foreground="#555555").grid(row=er, column=2, sticky=tk.W, padx=4)
            er += 1; _has_ev = True
        if getattr(booster, 'n_boosters', 0) > 0:
            ttk.Label(_ev, text="Booster drop (s):").grid(row=er, column=0, sticky=tk.W, pady=2)
            _bj = float(plan.get('booster_jettison_s', 0.0) or 0.0)
            self._ev_vars['booster'] = tk.StringVar(value=f"{_bj:g}" if _bj > 0 else "")
            ttk.Entry(_ev, textvariable=self._ev_vars['booster'], width=10).grid(row=er, column=1, sticky=tk.W)
            ttk.Label(_ev, text="blank = at burnout", foreground="#555555").grid(row=er, column=2, sticky=tk.W, padx=4)
            er += 1; _has_ev = True
        self._ev_gridfin = {}
        for i, s in enumerate(self._stages):
            if getattr(s, 'has_grid_fins', False) and getattr(s, 'n_grid_fins', 0) > 0:
                ttk.Label(_ev, text=f"Grid-fin deploy, stage {i + 1} (t:n,…):").grid(
                    row=er, column=0, sticky=tk.W, pady=2)
                sched = []
                if i < len(plan.get('stages', [])):
                    sched = plan['stages'][i].get('grid_fin_deploy_schedule') or []
                sv = tk.StringVar(value=_format_deploy_schedule(sched))
                ttk.Entry(_ev, textvariable=sv, width=18).grid(row=er, column=1, columnspan=2, sticky=tk.W)
                self._ev_gridfin[i] = sv
                er += 1; _has_ev = True
        if _has_ev:
            _ev.grid(row=r, column=0, columnspan=2, sticky=tk.EW, pady=(10, 2)); r += 1

        # ── Yaw / dogleg program (up to three azimuth maneuvers) ─────────
        _yf = ttk.LabelFrame(frm, text="Yaw / dogleg program", padding=6)
        _yf.grid(row=r, column=0, columnspan=2, sticky=tk.EW, pady=(10, 2)); r += 1
        for _c, _h in enumerate(["", "Start (s)", "Stop (s)", "Final az (°)"]):
            ttk.Label(_yf, text=_h, foreground="#555555").grid(row=0, column=_c, padx=3, pady=(0, 1))
        _yaw = list(plan.get('yaw_maneuvers') or [])
        # Surface any orphaned per-stage yaw (stage_yaw_*) into the visible grid
        # so a legacy baked dogleg is editable here too; _save clears the
        # per-stage fields, leaving the global grid the sole yaw authority.
        _ign_t = 0.0
        for _si, _sd in enumerate(plan.get('stages', [])):
            _fa = _sd.get('stage_yaw_final_az_deg')
            if _fa is not None:
                _ys = _sd.get('stage_yaw_start_s')
                _ye = _sd.get('stage_yaw_stop_s')
                _yaw.append([_ys if _ys is not None else _ign_t,
                             _ye if _ye is not None else _ign_t, _fa])
            if _si < len(self._stages):
                _st = self._stages[_si]
                _ign_t += (getattr(_st, 'burn_time_s', 0.0)
                           + getattr(_st, 'coast_time_s', 0.0))
        self._yaw_rows = []
        for _i in range(3):
            man = _yaw[_i] if _i < len(_yaw) else [None, None, None]
            ttk.Label(_yf, text=f"#{_i + 1}").grid(row=_i + 1, column=0, padx=3)
            trip = {}
            for _c, _k in enumerate(('start', 'stop', 'final_az'), start=1):
                v = man[_c - 1] if _c - 1 < len(man) else None
                sv = tk.StringVar(value="" if v in (None, "") else f"{float(v):g}")
                ttk.Entry(_yf, textvariable=sv, width=8).grid(row=_i + 1, column=_c, padx=3, pady=1)
                trip[_k] = sv
            self._yaw_rows.append(trip)

        # Angle-of-attack envelope + induced-drag toggle (NASA SP-8099), sitting
        # with the dogleg they govern.  The α limit clamps the commanded thrust
        # to a cone about the velocity while dynamic pressure is significant, so
        # a sharp dogleg slews over the time it physically needs instead of
        # being flown instantly at α≈90°; blank = no limit (the maneuver is
        # flown as commanded and only flagged when it exceeds the envelope).
        # Default 10° (SP-8099 mid-envelope) when the plan has never set the
        # key; a stored None (user cleared it) stays blank.
        if 'alpha_limit_deg' in plan:
            _alv = plan.get('alpha_limit_deg')
            _alv_str = "" if _alv in (None, "") else f"{float(_alv):g}"
        else:
            _alv_str = "10"
        self._alpha_limit_var = tk.StringVar(value=_alv_str)
        ttk.Label(_yf, text="α limit (°):").grid(
            row=4, column=0, sticky=tk.E, padx=3, pady=(8, 1))
        ttk.Entry(_yf, textvariable=self._alpha_limit_var, width=8).grid(
            row=4, column=1, padx=3, pady=(8, 1))
        ttk.Label(_yf, text="at max-q (SP-8099 5–10); blank = warn only",
                  foreground="#555555").grid(
            row=4, column=2, columnspan=2, sticky=tk.W, padx=3, pady=(8, 1))
        self._alpha_induced_var = tk.BooleanVar(
            value=bool(plan.get('alpha_induced_drag', False)))
        ttk.Checkbutton(
            _yf, text="α induced drag (bleeds energy, reshapes q)",
            variable=self._alpha_induced_var).grid(
            row=5, column=0, columnspan=4, sticky=tk.W, padx=3, pady=(1, 2))

        # ── Provenance ───────────────────────────────────────────────────
        ttk.Label(frm, text="Source:").grid(row=r, column=0, sticky=tk.W, pady=(10, 2))
        self._source_var = tk.StringVar(value=str(plan.get('source', '')))
        ttk.Entry(frm, textvariable=self._source_var, width=44).grid(
            row=r, column=1, sticky=tk.EW, pady=(10, 2)); r += 1
        ttk.Label(frm, text="Notes:").grid(row=r, column=0, sticky=tk.NW, pady=2)
        self._notes_text = tk.Text(frm, width=44, height=3, wrap=tk.WORD)
        self._notes_text.grid(row=r, column=1, sticky=tk.EW, pady=2)
        if plan.get('notes'):
            self._notes_text.insert("1.0", str(plan['notes']))
        r += 1

        bf = ttk.Frame(frm)
        bf.grid(row=r, column=0, columnspan=2, sticky=tk.E, pady=(12, 0))
        ttk.Button(bf, text="Cancel", command=self.destroy).pack(side=tk.RIGHT, padx=4)
        ttk.Button(bf, text="Save", command=self._save).pack(side=tk.RIGHT)

        self._base_plan = plan
        self.grab_set()

    @staticmethod
    def _f(sv):
        s = sv.get().strip()
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None

    def _on_mode_changed(self, *_):
        """Show only the controls the selected ascent mode actually uses."""
        m = self._mode_var.get()
        show_angle   = m in ("Simple pitch profile", "Orbital insertion")
        show_wheelon = m == "Simple pitch profile"
        show_table   = m in ("Advanced pitch profile", "Gravity turn")
        self._burnout_lbl.config(
            text="Boost angle (°):" if m == "Orbital insertion" else "Burnout angle (°):")
        if show_angle:
            self._burnout_lbl.grid(); self._burnout_frame.grid()
        else:
            self._burnout_lbl.grid_remove(); self._burnout_frame.grid_remove()
        if show_wheelon:
            self._wheelon_btn.pack(side=tk.LEFT, padx=6)
        else:
            self._wheelon_btn.pack_forget()
        if show_table:
            self._pstage_lbl.grid(); self._pstage_tbl.grid()
        else:
            self._pstage_lbl.grid_remove(); self._pstage_tbl.grid_remove()
        # Global turn window applies to every mode except advanced pitch (which
        # sets turn start/stop per stage in the table above).
        _show_turn = not show_table
        for _w in (self._turn_start_lbl, self._turn_start_ent,
                   self._turn_stop_lbl, self._turn_stop_ent):
            (_w.grid() if _show_turn else _w.grid_remove())
        # In gravity turn the per-stage "Angle" column is the eta kick angle.
        if self._angle_hdr is not None:
            self._angle_hdr.config(text="η (°)" if m == "Gravity turn" else "Angle")
        self._mode_note.config(text={
            "Simple pitch profile":
                "Linear pitch to the burnout angle; Estimate ε* fills the Wheelon optimum.",
            "Advanced pitch profile":
                "Per-stage pitch schedule below.",
            "Gravity turn":
                "Thrust follows velocity from the launch elevation; per-stage η is an optional kick.",
            "Orbital insertion":
                "Use Plan Orbit on the main panel to solve the boost angle for a target orbit.",
        }.get(m, ""))

    def _estimate_wheelon(self):
        """Fill the burnout-angle field with the Wheelon-optimal estimate."""
        try:
            self._burnout_var.set(f"{wheelon_burnout_angle(self._booster):.1f}")
        except Exception as exc:
            messagebox.showerror("Estimate ε*", str(exc), parent=self)

    def _save(self):
        plan = dict(self._base_plan)   # keep unknown keys (loft rate, yaw, …)
        _mode = self._mode_var.get()
        plan['guidance'] = ("true_gravity_turn" if _mode == "Gravity turn" else
                            "orbital_insertion" if _mode == "Orbital insertion" else
                            "pitch_program")
        le = self._f(self._launch_el_var)
        if le is not None:
            plan['launch_elevation_deg'] = le
        bo = self._f(self._burnout_var)
        if bo is not None:
            plan['burnout_angle_deg'] = bo
        # Global turn window (GUI keys the run path reads).
        plan['gt_turn_start_s'] = self._f(self._turn_start_var)
        plan['gt_turn_stop_s']  = self._f(self._turn_stop_var)

        stages = [dict(s) for s in plan.get('stages', [])]
        while len(stages) < len(self._stage_rows):
            stages.append({})
        _advanced = (_mode == "Advanced pitch profile")
        for i, row in enumerate(self._stage_rows):
            for key, sv in row.items():
                val = self._f(sv)
                # In non-advanced modes, a blank per-stage angle clears the
                # override so the global burnout angle governs.
                if key == 'stage_burnout_angle_deg' and not _advanced and val is None:
                    stages[i][key] = None
                elif val is not None:
                    stages[i][key] = val
                elif key in ('coast_time_s',):
                    stages[i][key] = 0.0
        # events
        if 'shroud' in self._ev_vars:
            v = self._f(self._ev_vars['shroud'])
            plan['shroud_jettison_alt_km'] = v if v is not None else 0.0
        if 'booster' in self._ev_vars:
            v = self._f(self._ev_vars['booster'])
            plan['booster_jettison_s'] = v if v is not None else 0.0
        for i, sv in self._ev_gridfin.items():
            try:
                stages[i]['grid_fin_deploy_schedule'] = _parse_deploy_schedule(sv.get())
            except ValueError:
                messagebox.showerror("Flight Plan",
                                     f"Grid-fin schedule for stage {i + 1} must be 't:n, t:n'.",
                                     parent=self)
                return
        plan['stages'] = stages
        # Yaw / dogleg program -> yaw_maneuvers list (GUI key); a maneuver
        # counts only when it has a final azimuth.
        yaw = []
        for trip in self._yaw_rows:
            fa = self._f(trip['final_az'])
            if fa is not None:
                yaw.append([self._f(trip['start']), self._f(trip['stop']), fa])
        plan['yaw_maneuvers'] = yaw
        plan['adv_yaw_on'] = bool(yaw)
        # α envelope + induced-drag toggle travel with the dogleg.
        plan['alpha_limit_deg'] = self._f(self._alpha_limit_var)
        plan['alpha_induced_drag'] = bool(self._alpha_induced_var.get())
        # Yaw is owned by the global grid above; clear any per-stage stage_yaw_*
        # (surfaced into the grid on load) so it can't override it on reload.
        for st in stages:
            st['stage_yaw_start_s'] = None
            st['stage_yaw_stop_s'] = None
            st['stage_yaw_final_az_deg'] = None
        plan['source'] = self._source_var.get().strip()
        plan['notes'] = self._notes_text.get("1.0", "end").strip()
        self._result = plan
        self.destroy()

    @property
    def result(self):
        return self._result


# ---------------------------------------------------------------------------
# Range-ring dialog
# ---------------------------------------------------------------------------

class RangeRingDialog(tk.Toplevel):
    """Compute and export a maximum-range ring for the current booster.

    Sweeps 72 azimuths (every 5°) using maximize_range(), collects the
    impact point for each direction, then renders the closed polygon on a
    Cartopy map using the shared projection picker.
    """

    _N_AZ = 72   # number of azimuths → 5° spacing

    def __init__(self, app):
        super().__init__(app)
        self._app   = app
        self._ring  = None   # list of (lon, lat) impact points once computed
        self._stop  = threading.Event()

        self.title("Range Ring")
        self.resizable(False, False)
        self.grab_set()

        frm = ttk.Frame(self, padding=12)
        frm.pack(fill=tk.BOTH)
        frm.columnconfigure(1, weight=1)

        # Booster label (informational)
        ttk.Label(frm, text="Booster:").grid(
            row=0, column=0, sticky=tk.W, padx=(0, 8), pady=3)
        self._booster_lbl = ttk.Label(frm, text=app._booster_var.get(),
                                      foreground="navy")
        self._booster_lbl.grid(row=0, column=1, sticky=tk.W)

        # Launch lat
        ttk.Label(frm, text="Launch lat (°N):").grid(
            row=1, column=0, sticky=tk.W, padx=(0, 8), pady=3)
        self._lat_var = tk.StringVar(value=app._launch_lat.get())
        ttk.Entry(frm, textvariable=self._lat_var, width=12).grid(
            row=1, column=1, sticky=tk.W)

        # Launch lon
        ttk.Label(frm, text="Launch lon (°E):").grid(
            row=2, column=0, sticky=tk.W, padx=(0, 8), pady=3)
        self._lon_var = tk.StringVar(value=app._launch_lon.get())
        ttk.Entry(frm, textvariable=self._lon_var, width=12).grid(
            row=2, column=1, sticky=tk.W)

        ttk.Separator(self, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=12)

        # Progress bar + status label
        prog_frm = ttk.Frame(self, padding=(12, 6))
        prog_frm.pack(fill=tk.X)
        self._prog_var = tk.StringVar(value="Press Compute to start.")
        ttk.Label(prog_frm, textvariable=self._prog_var).pack(anchor=tk.W)
        self._pbar = ttk.Progressbar(prog_frm, maximum=self._N_AZ,
                                     mode="determinate")
        self._pbar.pack(fill=tk.X, pady=(4, 0))

        ttk.Separator(self, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=12)

        # Buttons
        btn_frm = ttk.Frame(self, padding=(12, 8))
        btn_frm.pack(fill=tk.X)
        self._compute_btn = ttk.Button(btn_frm, text="Compute Ring",
                                       command=self._compute)
        self._compute_btn.pack(side=tk.LEFT)
        self._cancel_btn = ttk.Button(btn_frm, text="Stop",
                                      command=self._cancel,
                                      state=tk.DISABLED)
        self._cancel_btn.pack(side=tk.LEFT, padx=6)
        self._export_btn = ttk.Button(btn_frm, text="Export Map…",
                                      command=self._export,
                                      state=tk.DISABLED)
        self._export_btn.pack(side=tk.LEFT, padx=(24, 0))
        ttk.Button(btn_frm, text="Close",
                   command=self.destroy).pack(side=tk.RIGHT)

        app._center_dialog(self)

    # ------------------------------------------------------------------
    def _compute(self):
        try:
            lat = float(self._lat_var.get())
            lon = float(self._lon_var.get())
        except ValueError:
            messagebox.showerror("Input error",
                                 "Enter valid launch lat/lon.", parent=self)
            return

        try:
            (booster, guidance, _la, _lo, _az, cutoff, la,
             gt_start_s, gt_stop_s, _orb,
             _yaw_maneuvers, launch_elevation_deg) = self._app._get_inputs()
        except Exception as e:
            messagebox.showerror("Input error", str(e), parent=self)
            return

        self._ring = None
        self._stop.clear()
        self._pbar["value"] = 0
        self._prog_var.set(f"Computing 0 / {self._N_AZ}…")
        self._compute_btn.config(state=tk.DISABLED)
        self._cancel_btn.config(state=tk.NORMAL)
        self._export_btn.config(state=tk.DISABLED)

        threading.Thread(
            target=self._worker,
            args=(booster, guidance, lat, lon, la,
                  gt_start_s, gt_stop_s, launch_elevation_deg),
            daemon=True,
        ).start()

    def _cancel(self):
        self._stop.set()

    def _worker(self, booster, guidance, lat, lon, la,
                gt_start_s, gt_stop_s, launch_elevation_deg):
        azimuths = np.linspace(0.0, 360.0, self._N_AZ, endpoint=False)
        points   = []   # (az, impact_lon, impact_lat)

        for i, az in enumerate(azimuths):
            if self._stop.is_set():
                self.after(0, self._on_cancelled)
                return
            try:
                result = maximize_range(
                    booster, lat, lon, az,
                    guidance=guidance,
                    burnout_angle_deg=la,
                    gt_turn_start_s=gt_start_s,
                    gt_turn_stop_s=gt_stop_s,
                )
                ms_list = result.get('milestones', [])
                impact  = next(
                    (m for m in ms_list
                     if 'impact' in m.get('event', '').lower()
                     and not m.get('is_debris', False)),
                    None)
                if impact:
                    t_arr  = np.asarray(result['t'])
                    la_arr = np.asarray(result['lat'])
                    lo_arr = np.asarray(result['lon'])
                    imp_lat = float(np.interp(impact['t_s'], t_arr, la_arr))
                    imp_lon = float(np.interp(impact['t_s'], t_arr, lo_arr))
                    points.append((az, imp_lon, imp_lat))
            except Exception:
                pass   # skip failed azimuths silently

            self.after(0, self._on_progress, i + 1, len(points))

        self.after(0, self._on_done, points, lat, lon)

    def _on_progress(self, done, n_ok):
        if not self.winfo_exists():
            return
        self._pbar["value"] = done
        self._prog_var.set(
            f"Computing {done} / {self._N_AZ}… ({n_ok} points OK)")

    def _on_cancelled(self):
        if not self.winfo_exists():
            return
        self._prog_var.set("Cancelled.")
        self._compute_btn.config(state=tk.NORMAL)
        self._cancel_btn.config(state=tk.DISABLED)

    def _on_done(self, points, launch_lat, launch_lon):
        if not self.winfo_exists():
            return
        self._cancel_btn.config(state=tk.DISABLED)
        self._compute_btn.config(state=tk.NORMAL)
        if len(points) < 3:
            self._prog_var.set(
                f"Too few valid azimuths ({len(points)}). Check booster params.")
            return
        avg_range = float(np.mean([
            np.sqrt((p[2] - launch_lat)**2 + (p[1] - launch_lon)**2)
            for p in points]))
        self._ring         = points
        self._launch_lat   = launch_lat
        self._launch_lon   = launch_lon
        self._prog_var.set(
            f"Done — {len(points)} / {self._N_AZ} azimuths succeeded.")
        self._export_btn.config(state=tk.NORMAL)

    # ------------------------------------------------------------------
    def _export(self):
        if not self._ring:
            return
        try:
            import cartopy.crs as ccrs
            import cartopy.feature as cfeature
            import matplotlib.patheffects as pe
            from matplotlib.backends.backend_agg import FigureCanvasAgg
        except ImportError as _e:
            messagebox.showerror("Missing package",
                                 f"Cartopy not installed.\n{_e}", parent=self)
            return

        # Reuse the full export-options dialog (projection + map extent).
        mid_lon = float(np.mean([p[1] for p in self._ring]))
        mid_lat = float(np.mean([p[2] for p in self._ring]))
        proj, extent_spec = self._app._pick_cartopy_export_options(mid_lon, mid_lat)
        if proj is None:
            return

        import datetime as _dt
        from tkinter.filedialog import asksaveasfilename
        ts      = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        booster = _safe_name(self._app._booster_var.get())
        path = asksaveasfilename(
            defaultextension=".png",
            initialdir=str(_ensure_dir(_DIR_MAPS)),
            initialfile=f"{ts}_{booster}_rangering.cartopy.png",
            filetypes=[("PNG image",    "*.png"), ("PDF document", "*.pdf"),
                       ("SVG image",    "*.svg"), ("All files",    "*.*")],
            title="Save range-ring map",
            parent=self,
        )
        if not path:
            return

        geo    = ccrs.Geodetic()
        fig    = Figure(figsize=(10, 8), dpi=300)
        canvas = FigureCanvasAgg(fig)
        ax     = fig.add_subplot(1, 1, 1, projection=proj)

        # ── Map extent ────────────────────────────────────────────────
        if extent_spec is None:
            ax.set_global()
        elif extent_spec[0] == 'auto':
            pad_frac = extent_spec[1] / 100.0
            ring_lats = [p[2] for p in self._ring] + [self._launch_lat]
            ring_lons = [p[1] for p in self._ring] + [self._launch_lon]
            lat_span = max(float(max(ring_lats) - min(ring_lats)), 2.0)
            lon_span = max(float(max(ring_lons) - min(ring_lons)), 2.0)
            ax.set_extent([
                max(-180.0, min(ring_lons) - lon_span * pad_frac),
                min(+180.0, max(ring_lons) + lon_span * pad_frac),
                max( -90.0, min(ring_lats) - lat_span * pad_frac),
                min( +90.0, max(ring_lats) + lat_span * pad_frac),
            ], crs=ccrs.PlateCarree())
        else:
            ax.set_extent(list(extent_spec), crs=ccrs.PlateCarree())

        ax.add_feature(cfeature.OCEAN,     facecolor="#d6e8f5", zorder=0)
        ax.add_feature(cfeature.LAND,      facecolor="#e8e4d8", zorder=1)
        ax.add_feature(cfeature.COASTLINE, linewidth=0.5, edgecolor="#555555",
                       zorder=2)
        ax.add_feature(cfeature.BORDERS,   linewidth=0.3, edgecolor="#888888",
                       linestyle=":", zorder=2)
        ax.add_feature(cfeature.LAKES,     facecolor="#d6e8f5", linewidth=0.3,
                       edgecolor="#555555", zorder=2)
        ax.gridlines(color="white", linewidth=0.4, linestyle="--", alpha=0.6,
                     zorder=3)

        # Ring polygon — close it by repeating the first point.
        ring_lons = [p[1] for p in self._ring] + [self._ring[0][1]]
        ring_lats = [p[2] for p in self._ring] + [self._ring[0][2]]
        ax.fill(ring_lons, ring_lats,
                color="crimson", alpha=0.12, transform=geo, zorder=4)
        ax.plot(ring_lons, ring_lats,
                color="crimson", linewidth=1.6,
                path_effects=[pe.withStroke(linewidth=3, foreground="white")],
                transform=geo, zorder=5)

        # Launch point
        ax.plot(self._launch_lon, self._launch_lat,
                marker="^", markersize=8, color="black",
                markeredgecolor="white", markeredgewidth=1.0,
                transform=geo, zorder=6)

        # Estimate average max range for the title.
        ranges_km = []
        for _, imp_lon, imp_lat in self._ring:
            try:
                km = range_between(
                    np.radians(self._launch_lat), np.radians(self._launch_lon),
                    np.radians(imp_lat), np.radians(imp_lon)) / 1000.0
                ranges_km.append(km)
            except Exception:
                pass
        rng_str = (f"~{np.mean(ranges_km):.0f} km max range"
                   if ranges_km else "max range")
        booster_name = self._app._booster_var.get()
        ax.set_title(f"{booster_name}  ·  {rng_str}", fontsize=11, pad=8)

        fig.tight_layout()
        canvas.print_figure(path, bbox_inches="tight")
        self._app._status_var.set(f"Range ring saved: {path}")
        _open_file(path)


# ---------------------------------------------------------------------------
# Parametric sweep / sensitivity-analysis dialog
# ---------------------------------------------------------------------------

class ParametricSweepDialog(tk.Toplevel):
    """Non-modal dialog for 1-D parametric trajectory sweep.

    Reproduces the analyses Forden performs in all three worked examples:
      • Table 2  — Range vs azimuth (vary azimuth, fixed burnout / cutoff)
      • Figure 7 — Range vs burnout angle (vary burnout_angle_deg)
      • Ad hoc   — Range vs cutoff time (vary engine cutoff)

    The user picks which parameter to vary plus a start/stop/step range;
    the remaining parameters are taken from the main window at the moment
    "Run Sweep" is clicked.  Results appear incrementally in a live plot
    and a scrollable table.  An "Overplot trajectories" option shows all
    altitude-vs-range profiles on one axes (≤ 20 curves).
    """

    _PARAM_INFO = {
        "Azimuth":     dict(key="azimuth",    lo=0.0,  hi=360.0, step=5.0,  unit="°"),
        "Burnout Angle": dict(key="burnout_angle", lo=10.0, hi=80.0, step=5.0, unit="°"),
        "Cutoff Time": dict(key="cutoff",     lo=None, hi=None,  step=5.0,  unit="s"),
        "Turn Stop":   dict(key="turn_stop",  lo=None, hi=None,  step=10.0, unit="s"),
    }

    def __init__(self, parent_app):
        super().__init__(parent_app)
        self.title("Parametric Sweep / Sensitivity Analysis")
        self.geometry("820x680")
        self.resizable(True, True)
        self._app        = parent_app
        self._stop_evt   = threading.Event()
        self._results    = []          # list of (param_val, range_km, apogee_km)
        self._traj_store = []          # list of (param_val, result_dict), for overplot
        self._build()

    # ------------------------------------------------------------------
    def _build(self):
        pad = dict(padx=8, pady=4)

        # ── Sweep configuration ────────────────────────────────────────
        cf = ttk.LabelFrame(self, text="Sweep Configuration")
        cf.pack(fill=tk.X, **pad)

        row0 = ttk.Frame(cf)
        row0.pack(fill=tk.X, padx=6, pady=(4, 2))

        ttk.Label(row0, text="Vary:").pack(side=tk.LEFT)
        self._param_var = tk.StringVar(value="Azimuth")
        pcb = ttk.Combobox(row0, textvariable=self._param_var,
                           values=list(self._PARAM_INFO.keys()),
                           state="readonly", width=14)
        pcb.pack(side=tk.LEFT, padx=(4, 12))
        pcb.bind("<<ComboboxSelected>>", self._on_param_changed)

        self._lo_var   = tk.StringVar(value="0.0")
        self._hi_var   = tk.StringVar(value="360.0")
        self._step_var = tk.StringVar(value="5.0")
        for lbl, var in [("From:", self._lo_var), ("To:", self._hi_var),
                          ("Step:", self._step_var)]:
            ttk.Label(row0, text=lbl).pack(side=tk.LEFT, padx=(4, 1))
            ttk.Entry(row0, textvariable=var, width=7).pack(side=tk.LEFT)

        opts = ttk.Frame(cf)
        opts.pack(fill=tk.X, padx=6, pady=(2, 6))
        ttk.Label(opts, text="Show:").pack(side=tk.LEFT)
        self._show_range  = tk.BooleanVar(value=True)
        self._show_apogee = tk.BooleanVar(value=True)
        # Heating axes (SURVIVABILITY_REPORT_DESIGN.md §7): peak stagnation
        # flux q̇ and integrated load Q per sweep point — checking either
        # switches the plot to the heating view, where the loft/depress trade
        # is the crossing of the two curves (flux rises with burnout angle
        # while load falls).
        self._show_qpeak  = tk.BooleanVar(value=False)
        self._show_load   = tk.BooleanVar(value=False)
        self._overplot    = tk.BooleanVar(value=False)
        ttk.Checkbutton(opts, text="Range",  variable=self._show_range ).pack(side=tk.LEFT, padx=4)
        ttk.Checkbutton(opts, text="Apogee", variable=self._show_apogee).pack(side=tk.LEFT, padx=4)
        ttk.Checkbutton(opts, text="Peak flux q̇", variable=self._show_qpeak).pack(side=tk.LEFT, padx=4)
        ttk.Checkbutton(opts, text="Heat load Q", variable=self._show_load).pack(side=tk.LEFT, padx=4)
        ttk.Separator(opts, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=8, pady=2)
        ttk.Checkbutton(opts, text="Overplot trajectory profiles (≤ 20 pts)",
                        variable=self._overplot).pack(side=tk.LEFT, padx=4)

        # ── Buttons + progress ─────────────────────────────────────────
        bf = ttk.Frame(self)
        bf.pack(fill=tk.X, padx=8, pady=(0, 4))
        self._run_btn = ttk.Button(bf, text="▶  Run Sweep", command=self._run)
        self._run_btn.pack(side=tk.LEFT, padx=(0, 4))
        self._cancel_btn = ttk.Button(bf, text="■  Cancel",
                                      command=self._cancel, state=tk.DISABLED)
        self._cancel_btn.pack(side=tk.LEFT, padx=4)
        ttk.Button(bf, text="Export CSV…", command=self._export).pack(side=tk.LEFT, padx=4)
        self._prog_lbl = tk.StringVar(value="")
        ttk.Label(bf, textvariable=self._prog_lbl).pack(side=tk.LEFT, padx=8)
        self._progressbar = ttk.Progressbar(bf, mode="determinate", length=180)
        self._progressbar.pack(side=tk.RIGHT, padx=(4, 0))

        # ── Embedded matplotlib figure ─────────────────────────────────
        self._fig     = Figure(figsize=(8, 3.2), dpi=96)
        self._canvas  = FigureCanvasTkAgg(self._fig, master=self)
        self._canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=8)
        NavigationToolbar2Tk(self._canvas, self).update()
        self._init_plot()

        # ── Results table ──────────────────────────────────────────────
        tf = ttk.LabelFrame(self, text="Results Table")
        tf.pack(fill=tk.X, padx=8, pady=(4, 8))
        self._tree = ttk.Treeview(tf,
                                  columns=("param", "range", "apogee",
                                           "qpeak", "load"),
                                  show="headings", height=6)
        self._tree.heading("param",  text="Parameter")
        self._tree.heading("range",  text="Range (km)")
        self._tree.heading("apogee", text="Apogee (km)")
        self._tree.heading("qpeak",  text="q̇ peak (MW/m²)")
        self._tree.heading("load",   text="Q load (MJ/m²)")
        for col in ("param", "range", "apogee", "qpeak", "load"):
            self._tree.column(col, width=105, anchor=tk.CENTER)
        vsb = ttk.Scrollbar(tf, orient=tk.VERTICAL, command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self._tree.pack(fill=tk.X, padx=4, pady=4)

    # ------------------------------------------------------------------
    def _init_plot(self):
        self._fig.clf()
        ax = self._fig.add_subplot(111)
        ax.set_title("Run a sweep to see results", fontsize=9)
        ax.grid(True, alpha=0.35)
        self._fig.tight_layout(pad=2.5)
        self._canvas.draw()

    # ------------------------------------------------------------------
    def _on_param_changed(self, _event=None):
        info = self._PARAM_INFO[self._param_var.get()]
        lo   = info["lo"]
        hi   = info["hi"]
        if lo is None or hi is None:
            # Cutoff: derive from selected booster
            try:
                booster, *_ = self._app._get_inputs()
                lo, hi = 5.0, float(int(total_burn_time(booster)))
            except Exception:
                lo, hi = 5.0, 100.0
        self._lo_var  .set(str(lo))
        self._hi_var  .set(str(hi))
        self._step_var.set(str(info["step"]))
        # Update table column header
        self._tree.heading("param", text=f"{self._param_var.get()} ({info['unit']})")

    # ------------------------------------------------------------------
    def _make_points(self):
        lo   = float(self._lo_var.get())
        hi   = float(self._hi_var.get())
        step = float(self._step_var.get())
        if step <= 0:
            raise ValueError("Step must be > 0.")
        n = max(2, int(round((hi - lo) / step)) + 1)
        return np.linspace(lo, hi, n)

    # ------------------------------------------------------------------
    def _run(self):
        self._stop_evt.clear()
        try:
            (booster, guidance, lat, lon, az, cutoff, la,
             gt_start_s, gt_stop_s, _orb,
             _yaw_maneuvers, launch_elevation_deg) = self._app._get_inputs()
        except Exception as e:
            messagebox.showerror("Input error", str(e), parent=self)
            return
        try:
            points = self._make_points()
        except Exception as e:
            messagebox.showerror("Sweep range error", str(e), parent=self)
            return

        if cutoff is None:
            cutoff = total_burn_time(booster)

        overplot = self._overplot.get()
        if overplot and len(points) > 20:
            messagebox.showwarning(
                "Too many points for overplot",
                f"Overplot is limited to 20 trajectory profiles.\n"
                f"Your sweep has {len(points)} points — overplot will be skipped.\n"
                "Increase the step size or disable 'Overplot trajectory profiles'.",
                parent=self)
            overplot = False

        param_key = self._PARAM_INFO[self._param_var.get()]["key"]
        self._results    = []
        self._traj_store = []
        self._tree.delete(*self._tree.get_children())
        self._progressbar["maximum"] = len(points)
        self._progressbar["value"]   = 0
        self._prog_lbl.set(f"0 / {len(points)}")
        self._run_btn   .config(state=tk.DISABLED)
        self._cancel_btn.config(state=tk.NORMAL)
        self._init_plot()

        threading.Thread(
            target=self._sweep_worker,
            args=(booster, guidance, lat, lon, az, la, cutoff,
                  param_key, points, overplot, gt_start_s, gt_stop_s,
                  launch_elevation_deg),
            daemon=True,
        ).start()

    # ------------------------------------------------------------------
    def _sweep_worker(self, booster, guidance, lat, lon, az, la, cutoff,
                      param_key, points, store_trajs, gt_start_s=5.0, gt_stop_s=None,
                      launch_elevation_deg=90.0):
        for i, val in enumerate(points):
            if self._stop_evt.is_set():
                break
            run_az      = val if param_key == "azimuth"    else az
            run_la      = val if param_key == "burnout_angle" else la
            run_cut     = val if param_key == "cutoff"     else cutoff
            run_gt_stop = val if param_key == "turn_stop"  else gt_stop_s
            try:
                r = integrate_trajectory(
                    booster, lat, lon, run_az,
                    guidance=guidance,
                    burnout_angle_deg=run_la,
                    cutoff_time_s=run_cut,
                    gt_turn_start_s=gt_start_s,
                    gt_turn_stop_s=run_gt_stop,
                    launch_elevation_deg=launch_elevation_deg,
                )
                # Heating axes come free: integrate_trajectory computes the
                # survivability FOM per run (result['heating_fom']).
                _fom = r.get("heating_fom") or {}
                row  = (val, r["range_km"] if r["range_km"] is not None else float("nan"),
                        r["apogee_km"],
                        float(_fom.get("q_peak_MW_m2") or float("nan")),
                        float(_fom.get("integrated_load_MJ_m2") or float("nan")))
                traj = (val, r) if store_trajs else None
            except Exception:
                row  = (val, float("nan"), float("nan"),
                        float("nan"), float("nan"))
                traj = None
            self.after(0, self._add_point, row, traj, i + 1, len(points))
        self.after(0, self._sweep_done)

    # ------------------------------------------------------------------
    def _add_point(self, row, traj, done, total):
        self._results.append(row)
        if traj is not None:
            self._traj_store.append(traj)
        val, rng, apo, qpk, qld = row
        self._tree.insert("", tk.END, values=(
            f"{val:.2f}",
            f"{rng:.1f}"  if np.isfinite(rng) else "—",
            f"{apo:.1f}"  if np.isfinite(apo) else "—",
            f"{qpk:.1f}"  if np.isfinite(qpk) else "—",
            f"{qld:.0f}"  if np.isfinite(qld) else "—",
        ))
        self._tree.yview_moveto(1.0)
        self._progressbar["value"] = done
        self._prog_lbl.set(f"{done} / {total}")
        self._redraw()

    # ------------------------------------------------------------------
    def _sweep_done(self):
        self._run_btn   .config(state=tk.NORMAL)
        self._cancel_btn.config(state=tk.DISABLED)
        n = len(self._results)
        cancelled = self._stop_evt.is_set()
        self._prog_lbl.set(
            f"{'Cancelled after' if cancelled else 'Done —'} {n} point{'s' if n != 1 else ''}.")
        # Stash the sweep for the Reentry Survivability report's Form A
        # loft/depress context line (design decision 2: sweep-fed only).
        if n >= 2:
            try:
                self._app._last_heating_sweep = dict(
                    param=self._param_var.get(),
                    unit=self._PARAM_INFO[self._param_var.get()]["unit"],
                    booster=self._app._booster_var.get(),
                    rows=list(self._results))
            except Exception:
                pass
        self._redraw()

    # ------------------------------------------------------------------
    def _cancel(self):
        self._stop_evt.set()

    # ------------------------------------------------------------------
    def _redraw(self):
        if not self._results:
            return

        info   = self._PARAM_INFO[self._param_var.get()]
        xlabel = f"{self._param_var.get()} ({info['unit']})"

        xs          = [r[0] for r in self._results]
        ys_range    = [r[1] for r in self._results]
        ys_apogee   = [r[2] for r in self._results]
        ys_qpeak    = [r[3] for r in self._results]
        ys_load     = [r[4] for r in self._results]

        self._fig.clf()

        # ── Heating view (flux/load vs parameter) ─────────────────────
        # Takes precedence over range/apogee when either heating box is
        # checked (and overplot is off): the loft/depress trade IS the
        # crossing of the two curves.
        _sq, _sQ = self._show_qpeak.get(), self._show_load.get()
        if (_sq or _sQ) and not self._traj_store:
            ax = self._fig.add_subplot(111)
            ax.set_xlabel(xlabel, fontsize=8)
            ax.grid(True, alpha=0.35)
            ax.tick_params(labelsize=7)
            if _sq and _sQ:
                ax2 = ax.twinx()
                ax .plot(xs, ys_qpeak, "-o", color="#aa2222", markersize=3,
                         linewidth=1.5, label="Peak flux q̇")
                ax2.plot(xs, ys_load,  "--s", color="#2255aa", markersize=3,
                         linewidth=1.5, label="Heat load Q")
                ax .set_ylabel("Peak flux:  q̇  MW/m²", fontsize=8,
                               color="#aa2222")
                ax2.set_ylabel("Integrated load:  Q  MJ/m²", fontsize=8,
                               color="#2255aa")
                ax2.tick_params(labelsize=7)
                ax.set_title("Heating vs trajectory shaping — "
                             "flux rises where load falls (the loft/depress "
                             "trade)", fontsize=9)
            elif _sq:
                ax.plot(xs, ys_qpeak, "-o", color="#aa2222", markersize=3,
                        linewidth=1.5)
                ax.set_ylabel("Peak flux:  q̇  MW/m²", fontsize=8,
                              color="#aa2222")
                ax.set_title("Peak stagnation flux vs parameter", fontsize=9)
            else:
                ax.plot(xs, ys_load, "--s", color="#2255aa", markersize=3,
                        linewidth=1.5)
                ax.set_ylabel("Integrated load:  Q  MJ/m²", fontsize=8,
                              color="#2255aa")
                ax.set_title("Integrated heat load vs parameter",
                             fontsize=9)
            self._fig.tight_layout(pad=2.5)
            self._canvas.draw()
            return

        if self._traj_store:
            # ── Overplot trajectory profiles ──────────────────────────
            ax   = self._fig.add_subplot(111)
            cmap = matplotlib.cm.viridis
            vals = [t[0] for t in self._traj_store]
            vmin, vmax = min(vals), max(vals)
            span = max(vmax - vmin, 1e-9)
            for pval, r in self._traj_store:
                color = cmap((pval - vmin) / span)
                ax.plot(r["range"] / 1000.0, r["alt"] / 1000.0,
                        color=color, linewidth=1.0, alpha=0.85,
                        label=f"{pval:.1f}")
            ax.set_xlabel("Downrange (km)", fontsize=8)
            ax.set_ylabel("Altitude (km)",  fontsize=8)
            ax.set_title(f"Trajectory Profiles  —  {self._param_var.get()} sweep",
                         fontsize=9)
            ax.grid(True, alpha=0.35)
            ax.tick_params(labelsize=7)
            if len(self._traj_store) <= 12:
                ax.legend(title=info["unit"], fontsize=6, title_fontsize=6,
                          loc="upper right")
            sm = matplotlib.cm.ScalarMappable(
                cmap=cmap,
                norm=matplotlib.colors.Normalize(vmin=vmin, vmax=vmax))
            sm.set_array([])
            cb = self._fig.colorbar(sm, ax=ax, pad=0.02)
            cb.set_label(f"{self._param_var.get()} ({info['unit']})", fontsize=7)
            cb.ax.tick_params(labelsize=6)
        else:
            # ── Range / apogee vs parameter ───────────────────────────
            sr = self._show_range.get()
            sa = self._show_apogee.get()

            if sr and sa:
                ax  = self._fig.add_subplot(111)
                ax2 = ax.twinx()
                ax .plot(xs, ys_range,  "b-o",  markersize=3, linewidth=1.5, label="Range")
                ax2.plot(xs, ys_apogee, "r--s", markersize=3, linewidth=1.2, label="Apogee")
                ax .set_ylabel("Range (km)",  color="royalblue",  fontsize=8)
                ax2.set_ylabel("Apogee (km)", color="firebrick", fontsize=8)
                ax .tick_params(axis="y", labelcolor="royalblue",  labelsize=7)
                ax2.tick_params(axis="y", labelcolor="firebrick", labelsize=7)
                lines  = ax.get_lines() + ax2.get_lines()
                ax.legend(lines, [l.get_label() for l in lines], fontsize=7)
            elif sr:
                ax = self._fig.add_subplot(111)
                ax.plot(xs, ys_range, "b-o", markersize=3, linewidth=1.5)
                ax.set_ylabel("Range (km)", fontsize=8)
            else:
                ax = self._fig.add_subplot(111)
                ax.plot(xs, ys_apogee, "r-s", markersize=3, linewidth=1.5)
                ax.set_ylabel("Apogee (km)", fontsize=8)

            ax.set_xlabel(xlabel, fontsize=8)
            ax.set_title(f"{self._param_var.get()} Sweep", fontsize=9)
            ax.grid(True, alpha=0.35)
            ax.tick_params(labelsize=7)

        self._fig.tight_layout(pad=2.5)
        self._canvas.draw_idle()

    # ------------------------------------------------------------------
    def _export(self):
        if not self._results:
            messagebox.showinfo("No data", "Run a sweep first.", parent=self)
            return
        from tkinter.filedialog import asksaveasfilename
        path = asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            title="Export sweep results",
            parent=self,
        )
        if not path:
            return
        info   = self._PARAM_INFO[self._param_var.get()]
        header = (f"{self._param_var.get()}_{info['unit']},Range_km,Apogee_km,"
                  f"qpeak_MW_m2,Qload_MJ_m2")
        np.savetxt(path, np.array(self._results),
                   delimiter=",", header=header, comments="")
        self._app._status_var.set(f"Sweep exported: {path}")


# ---------------------------------------------------------------------------
# Helper: labelled decimal-degree row in a grid parent
# ---------------------------------------------------------------------------
def _dd_row(parent, label, row, default="0.0"):
    """Pack a label + decimal-degree Entry into a grid row; return the StringVar."""
    ttk.Label(parent, text=label).grid(row=row, column=0,
                                       sticky=tk.W, padx=(8, 2), pady=2)
    var = tk.StringVar(value=default)
    inner = ttk.Frame(parent)
    inner.grid(row=row, column=1, sticky=tk.W, padx=(0, 8), pady=2)
    ttk.Entry(inner, textvariable=var, width=10).pack(side=tk.LEFT)
    ttk.Label(inner, text="°").pack(side=tk.LEFT, padx=(2, 0))
    return var


class FootprintDialog(tk.Toplevel):
    """Generate an HGV maneuver footprint by sweeping glider bank angle.

    Runs N integrate_trajectory calls with the current booster/launch params
    but varying the glider bank angle (held for the full flight).  Negative
    banks turn left, positive banks turn right; zero produces the wings-
    level reference trajectory.  Results are written to a Folium HTML map
    and opened in the default browser.
    """

    def __init__(self, parent_app):
        super().__init__(parent_app)
        self.title("HGV Footprint Sweep")
        self.resizable(False, False)
        self._app      = parent_app
        self._stop_evt = threading.Event()
        self._results  = []   # list of (bank_deg, result_dict | None)
        self._map_path = None

        pad = dict(padx=8, pady=4)

        cf = ttk.LabelFrame(self, text="Sweep Configuration")
        cf.pack(fill=tk.X, **pad)

        r0 = ttk.Frame(cf)
        r0.pack(fill=tk.X, padx=6, pady=(4, 2))
        ttk.Label(r0, text="Bank angle — From:").pack(side=tk.LEFT)
        self._lo_var   = tk.StringVar(value="-45")
        self._hi_var   = tk.StringVar(value="45")
        self._step_var = tk.StringVar(value="10")
        for lbl, var, w in [("", self._lo_var, 6),
                             ("To:", self._hi_var, 6),
                             ("Step:", self._step_var, 5)]:
            if lbl:
                ttk.Label(r0, text=lbl).pack(side=tk.LEFT, padx=(6, 1))
            ttk.Entry(r0, textvariable=var, width=w).pack(side=tk.LEFT)
        ttk.Label(r0, text="°").pack(side=tk.LEFT, padx=(2, 0))

        bf = ttk.Frame(self)
        bf.pack(fill=tk.X, padx=8, pady=(0, 6))
        self._run_btn = ttk.Button(bf, text="▶  Run Footprint", command=self._run)
        self._run_btn.pack(side=tk.LEFT, padx=(0, 4))
        self._cancel_btn = ttk.Button(bf, text="■  Cancel",
                                      command=self._cancel, state=tk.DISABLED)
        self._cancel_btn.pack(side=tk.LEFT, padx=4)
        self._map_btn = ttk.Button(bf, text="Open Map…",
                                   command=self._open_map, state=tk.DISABLED)
        self._map_btn.pack(side=tk.LEFT, padx=4)
        self._prog_lbl = tk.StringVar(value="")
        ttk.Label(bf, textvariable=self._prog_lbl).pack(side=tk.LEFT, padx=8)
        self._progressbar = ttk.Progressbar(bf, mode="determinate", length=160)
        self._progressbar.pack(side=tk.RIGHT, padx=(4, 0))

    # ------------------------------------------------------------------
    def _run(self):
        try:
            lo   = float(self._lo_var.get())
            hi   = float(self._hi_var.get())
            step = float(self._step_var.get())
            if step <= 0 or lo > hi:
                raise ValueError
        except ValueError:
            messagebox.showerror("Input error",
                                 "Check sweep range (lo ≤ hi, step > 0).", parent=self)
            return

        try:
            (booster, guidance, lat, lon, az, cutoff, la,
             gt_start, gt_stop, orb, yaw, el) = self._app._get_inputs()
        except Exception as exc:
            messagebox.showerror("Input error", str(exc), parent=self)
            return

        bank_angles = []
        v = lo
        while v <= hi + 1e-9:
            bank_angles.append(round(v, 6))
            v += step

        self._results = []
        self._map_path = None
        self._stop_evt.clear()
        self._run_btn.config(state=tk.DISABLED)
        self._cancel_btn.config(state=tk.NORMAL)
        self._map_btn.config(state=tk.DISABLED)
        self._progressbar["maximum"] = max(len(bank_angles), 1)
        self._progressbar["value"]   = 0
        self._prog_lbl.set(f"0 / {len(bank_angles)}")

        threading.Thread(
            target=self._worker,
            args=(booster, guidance, lat, lon, az, cutoff, la,
                  gt_start, gt_stop, orb, yaw, el, bank_angles),
            daemon=True,
        ).start()

    def _cancel(self):
        self._stop_evt.set()

    def _worker(self, booster, guidance, lat, lon, az, cutoff, la,
                gt_start, gt_stop, orb, yaw, el, bank_angles):
        import copy, dataclasses
        from trajectory import integrate_trajectory
        from booster_models import effective_ro

        _max_t = 3600.0
        results = []

        for i, bk in enumerate(bank_angles):
            if self._stop_evt.is_set():
                break

            m = copy.deepcopy(booster)
            _ero = effective_ro(m)
            if _ero is not None:
                # Hold the swept bank angle for the entire flight.  The bank
                # is only applied while the glider is active (post-pierce
                # lift block in _eom), so the [0, _max_t] window safely
                # covers the whole glide phase regardless of when it starts.
                new_ro = dataclasses.replace(
                    _ero,
                    glider_enabled=True,
                    glider_bank_schedule=[(0.0, _max_t, float(bk))],
                )
                node = m
                while node is not None:
                    if node.ro is not None:
                        node.ro = new_ro
                        break
                    node = node.stage2

            try:
                r = integrate_trajectory(
                    m, lat, lon, az,
                    guidance=guidance,
                    burnout_angle_deg=la,
                    cutoff_time_s=cutoff,
                    gt_turn_start_s=gt_start,
                    gt_turn_stop_s=gt_stop,
                    yaw_maneuvers=yaw,
                    launch_elevation_deg=el,
                    alpha_limit_deg=self._app._alpha_limit_value(),
                    alpha_induced_drag=self._app._alpha_induced_value(),
                    terrain_dem=self._app._terrain_dem_value(),
                    launch_elev_m=self._app._launch_elev_value(),
                    max_time_s=_max_t,
                )
            except Exception:
                r = None

            results.append((bk, r))
            done = i + 1
            self.after(0, lambda d=done, tot=len(bank_angles): (
                self._prog_lbl.set(f"{d} / {tot}"),
                self._progressbar.__setitem__("value", d),
            ))

        self._results = results
        self.after(0, self._on_done)

    def _on_done(self):
        self._run_btn.config(state=tk.NORMAL)
        self._cancel_btn.config(state=tk.DISABLED)
        if self._results:
            self._build_and_open_map()

    def _build_and_open_map(self):
        import folium, os, tempfile, webbrowser

        try:
            (_, _, lat_deg, lon_deg, *_) = self._app._get_inputs()
        except Exception:
            lat_deg, lon_deg = 0.0, 0.0

        launch_lat = float(lat_deg)
        launch_lon = float(lon_deg)

        # Collect valid results and compute map centre
        valid = [(bk, r) for bk, r in self._results
                 if r is not None and r.get('lat') is not None
                 and len(r.get('lat', [])) > 0]
        if not valid:
            messagebox.showinfo("No results", "All trajectories failed.", parent=self)
            return

        all_lats = [launch_lat] + [r['lat'][-1] for _, r in valid]
        all_lons = [launch_lon] + [r['lon'][-1] for _, r in valid]
        centre = [sum(all_lats) / len(all_lats), sum(all_lons) / len(all_lons)]

        # Light basemap, matching the single-trajectory map (_export_folium).
        m = folium.Map(location=centre, zoom_start=4,
                       tiles='CartoDB positron')

        # Diverging palette keyed to the BANK ANGLE, not the index: bank is a
        # diverging quantity (left turn ↔ straight ↔ right turn), so the
        # restrained, meaningful encoding is blue↔gray↔red — cool for left
        # (negative) banks, brick-red (the app accent) for right (positive),
        # a mid-gray for near-straight.  Replaces the old rainbow HSL sweep,
        # which coded index order rather than turn direction.
        _COOL = (0x2a, 0x6f, 0xb0)     # steel blue  — max left bank
        _MID  = (0x7a, 0x7a, 0x76)     # mid gray    — straight
        _WARM = (0xc0, 0x39, 0x2b)     # brick red   — max right bank
        _b_max = max((abs(float(bk)) for bk, _ in valid), default=1.0) or 1.0

        def _bank_color(bk):
            t = max(-1.0, min(1.0, float(bk) / _b_max))    # −1 … +1
            a, b = (_MID, _WARM) if t >= 0 else (_MID, _COOL)
            f = abs(t)
            rgb = tuple(round(a[k] + (b[k] - a[k]) * f) for k in range(3))
            return "#%02x%02x%02x" % rgb

        # Trajectories
        for i, (bk, r) in enumerate(valid):
            lats = list(r['lat'])
            lons = list(r['lon'])
            coords = list(zip(lats, lons))
            col = _bank_color(bk)
            folium.PolyLine(
                coords, color=col, weight=2, opacity=0.85,
                tooltip=f"Bank {bk:+.0f}°"
            ).add_to(m)
            # Impact marker
            folium.CircleMarker(
                [lats[-1], lons[-1]],
                radius=4, color=col, fill=True, fill_opacity=1.0,
                tooltip=f"Impact (bank {bk:+.0f}°)",
            ).add_to(m)

        # Footprint envelope (convex hull of impact points).
        # Drawing the envelope as the convex hull instead of sweep-order
        # is robust against bank angles outside ±90°: those produce
        # inverted-lift dive trajectories whose impacts crash short, so
        # they fall *inside* the hull and don't distort the boundary.
        if len(valid) >= 3:
            try:
                from scipy.spatial import ConvexHull
                _pts = np.array([[r['lat'][-1], r['lon'][-1]] for _, r in valid])
                _hull = ConvexHull(_pts)
                env = [_pts[i].tolist() for i in _hull.vertices]
                env.append(env[0])
            except Exception:
                # Fallback: sweep-order polyline (degenerate hull, e.g. all
                # impacts collinear).
                env = [[r['lat'][-1], r['lon'][-1]] for _, r in valid]
                env.append(env[0])
            # The envelope is the headline of this map (the reachable
            # footprint), so it reads as an AREA, not a faint outline: a very
            # light neutral fill inside a dashed boundary.  The 10 % fill is a
            # wash the weight-2 trajectories (drawn earlier) still read
            # through, so the left/right spread stays legible over the area.
            folium.Polygon(env, color="#555555", weight=1.2,
                           dash_array="4 4", opacity=0.7,
                           fill=True, fill_color="#8a8a86",
                           fill_opacity=0.10).add_to(m)

        # Launch marker
        folium.Marker(
            [launch_lat, launch_lon],
            icon=folium.Icon(color="blue", icon="map-marker"),
            tooltip="Launch",
        ).add_to(m)

        # Save and open
        fd, path = tempfile.mkstemp(suffix="_footprint.html")
        os.close(fd)
        m.save(path)
        self._map_path = path
        self._map_btn.config(state=tk.NORMAL)
        webbrowser.open(f"file://{path}")
        self._prog_lbl.set(f"Done — {len(valid)} trajectories")

    def _open_map(self):
        import webbrowser
        if self._map_path:
            webbrowser.open(f"file://{self._map_path}")


def _glide_state_from_result(res):
    """Mid-glide (V_kms, alt_km) from a trajectory result dict, or None.

    Takes the median speed/altitude over the post-apogee in-atmosphere glide
    (25-55 km) so the damping estimate can anchor on a flown state."""
    try:
        import numpy as _np
        alt = _np.asarray(res['alt']).ravel() / 1000.0
        spd = _np.asarray(res['speed']).ravel() / 1000.0
        if alt.size < 5:
            return None
        iap = int(_np.argmax(alt))
        a, v = alt[iap:], spd[iap:]
        m = (a >= 25.0) & (a <= 55.0)
        if not m.any():
            return None
        return float(_np.median(v[m])), float(_np.median(a[m]))
    except Exception:
        return None


class DampingEstimatorDialog(tk.Toplevel):
    """Suggest a phugoid damping ratio ζ for the active glide RV.

    Wraps damping_estimate.estimate_damping (docs/damping_estimate_spec.md):
    the achievable ζ ceiling (lift-authority limited) with a modeling-
    uncertainty band, from β, L/D and a control-surface descriptor.  Pre-filled
    from the active terminal vehicle; the user may add control-surface area /
    deflection and a glide state (auto-filled if a trajectory has been flown,
    with a Restore button).  "Apply" writes the central ζ to the knob and stores
    the inputs on the RV.  The band is a capability boundary: ζ below it is a
    free design choice, inside it the authority limit, above it unphysical.
    """

    def __init__(self, app, zeta_var=None):
        super().__init__(app)
        self._app = app
        # Where to write the applied ζ: a caller-supplied var (the Reentry Plan
        # dialog's field) or, by default, the sidebar strip's var.
        self._zeta_target = zeta_var if zeta_var is not None else getattr(
            app, '_main_zeta_var', None)
        self.title("Estimate damping ratio ζ")
        self.resizable(False, False)
        self._result = None
        ro = getattr(app, "_ro", None)
        self._flown = None
        store = getattr(app, "_traj_store", None)
        if store:
            try:
                self._flown = _glide_state_from_result(store[-1][1])
            except Exception:
                self._flown = None
        self._build(ro)
        self._compute()
        try:
            app._center_dialog(self)
        except Exception:
            pass

    def _build(self, ro):
        pad = dict(padx=6, pady=3)
        frm = ttk.Frame(self)
        frm.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        self._r = 0

        def _row(label, default, unit=""):
            ttk.Label(frm, text=label).grid(row=self._r, column=0, sticky=tk.W, **pad)
            var = tk.StringVar(value=default)
            ttk.Entry(frm, textvariable=var, width=10).grid(
                row=self._r, column=1, sticky=tk.W, **pad)
            if unit:
                ttk.Label(frm, text=unit).grid(row=self._r, column=2, sticky=tk.W)
            self._r += 1
            return var

        def _sep():
            ttk.Separator(frm, orient=tk.HORIZONTAL).grid(
                row=self._r, column=0, columnspan=3, sticky='ew', pady=4)
            self._r += 1

        ttk.Label(frm, text="Reentry object (from reentry-object editor; editable)",
                  font=("TkDefaultFont", 9, "bold")).grid(
            row=self._r, column=0, columnspan=3, sticky=tk.W, **pad)
        self._r += 1
        beta = getattr(ro, "beta_kg_m2", 0.0) if ro else 0.0
        ld = getattr(ro, "glider_LD", 0.0) if ro else 0.0
        nmax = getattr(ro, "glider_pullup_g_max", 10.0) if ro else 10.0
        self._beta = _row("Ballistic coeff β", f"{beta:g}", "kg/m²")
        self._ld = _row("Lift-to-drag L/D", f"{ld:g}")
        self._nmax = _row("Max pull-up", f"{nmax:g}", "g")

        _sep()
        ttk.Label(frm, text="Control surfaces").grid(
            row=self._r, column=0, sticky=tk.W, **pad)
        self._ctrl = tk.StringVar(
            value=(getattr(ro, "glider_control_surfaces", "unknown") if ro else "unknown"))
        ttk.Combobox(frm, textvariable=self._ctrl, state="readonly", width=12,
                     values=["unknown", "none", "small", "substantial"]).grid(
            row=self._r, column=1, columnspan=2, sticky=tk.W, **pad)
        self._r += 1
        far = getattr(ro, "glider_flap_area_ratio", 0.0) if ro else 0.0
        dfl = getattr(ro, "glider_flap_deflection_deg", 0.0) if ro else 0.0
        self._far = _row("  Flap area S_flap/S_ref (opt.)", f"{far:g}" if far > 0 else "")
        self._dfl = _row("  Flap deflection (opt.)", f"{dfl:g}" if dfl > 0 else "", "deg")

        _sep()
        gl = "from flown trajectory" if self._flown else \
            "no trajectory flown — blank ⇒ swept 3–5 km/s"
        ttk.Label(frm, text=f"Glide state ({gl})").grid(
            row=self._r, column=0, columnspan=3, sticky=tk.W, **pad)
        self._r += 1
        self._v = _row("  Glide speed (opt.)",
                       (f"{self._flown[0]:.2f}" if self._flown else ""), "km/s")
        self._h = _row("  Glide altitude (opt.)",
                       (f"{self._flown[1]:.0f}" if self._flown else ""), "km")
        ttk.Button(frm, text="Restore", command=self._restore).grid(
            row=self._r, column=1, sticky=tk.W, **pad)
        self._r += 1

        _sep()
        ttk.Button(frm, text="Compute", command=self._compute).grid(
            row=self._r, column=0, sticky=tk.W, **pad)
        ttk.Button(frm, text="Apply ζ to knob", command=self._apply).grid(
            row=self._r, column=1, columnspan=2, sticky=tk.W, **pad)
        self._r += 1
        self._out = ttk.Label(frm, text="", font=("TkDefaultFont", 11, "bold"))
        self._out.grid(row=self._r, column=0, columnspan=3, sticky=tk.W, **pad)
        self._r += 1
        self._note = ttk.Label(frm, text="", foreground="#555555",
                               wraplength=380, justify=tk.LEFT)
        self._note.grid(row=self._r, column=0, columnspan=3, sticky=tk.W, **pad)
        self._r += 1

    def _restore(self):
        if self._flown:
            self._v.set(f"{self._flown[0]:.2f}")
            self._h.set(f"{self._flown[1]:.0f}")
        else:
            self._v.set("")
            self._h.set("")

    @staticmethod
    def _fval(var, default=None):
        s = var.get().strip()
        if not s:
            return default
        try:
            return float(s)
        except ValueError:
            return default

    def _compute(self):
        from damping_estimate import estimate_damping
        v = self._fval(self._v, None)
        h = self._fval(self._h, None)
        r = estimate_damping(
            self._fval(self._beta, 0.0) or 0.0,
            self._fval(self._ld, 0.0) or 0.0,
            nmax=self._fval(self._nmax, 10.0) or 10.0,
            control=self._ctrl.get(),
            flap_area_ratio=self._fval(self._far, 0.0) or 0.0,
            flap_deflection_deg=self._fval(self._dfl, 0.0) or 0.0,
            v_glide=(v * 1000.0 if v else None),
            h_glide=(h * 1000.0 if h else None))
        self._result = r
        self._out.config(text=r.text())
        self._note.config(text=r.notes)

    def _apply(self):
        if self._result is None:
            self._compute()
        r = self._result
        try:
            if self._zeta_target is not None:
                self._zeta_target.set(f"{r.zeta:.2f}")
        except Exception:
            pass
        ro = getattr(self._app, "_ro", None)
        if ro is not None:
            try:
                ro.glider_control_surfaces = self._ctrl.get()
                ro.glider_flap_area_ratio = self._fval(self._far, 0.0) or 0.0
                ro.glider_flap_deflection_deg = self._fval(self._dfl, 0.0) or 0.0
            except Exception:
                pass
        self.destroy()


class MassEstimatorDialog(tk.Toplevel):
    """Estimate a stage's dry (inert) mass from parameters and report how far
    the booster's stated burnout mass diverges from known mass relationships.

    Wraps mass_estimator.py: component-level Wilhite-school MERs (Akin/UMD)
    plus aggregate relations (Pietrobon hydrolox, structural coefficient,
    Zandbergen solid-stage regressions).  Stage parameters are prefilled from
    the currently selected booster; the user picks the propellant combination
    (liquid) or casing material (solid) and any unknowns, then recomputes.
    """

    def __init__(self, app):
        super().__init__(app)
        self._app = app
        self.title("Dry Mass Estimator")
        self.resizable(True, True)

        name = app._booster_var.get()
        try:
            self._params = get_booster(name)
        except Exception as exc:                       # pragma: no cover
            messagebox.showerror("Dry Mass Estimator",
                                 f"Could not load booster {name!r}:\n{exc}",
                                 parent=app)
            self.destroy()
            return
        self._stages = self._decompose(self._params)

        self._build(name)
        self._load_stage(0)
        self._compute()
        app._center_dialog(self)

    # ------------------------------------------------------------------
    @staticmethod
    def _decompose(p):
        """Per-stage (fueled, dry, prop, thrust, geometry) for a booster.

        Mirrors BoosterDialog._prefill: mass_initial is cumulative, so per-stage
        fueled mass is recovered by differencing adjacent stages and stripping
        payload / shroud; the stage's own dry (structural) mass is
        fueled − propellant, independent of the possibly-cumulative mass_final.
        """
        payload     = p.payload_kg
        shroud_mass = p.shroud_mass_kg
        out, node, idx = [], p, 0
        while node is not None:
            nxt      = node.stage2
            is_first = (idx == 0)
            is_last  = (nxt is None)
            if is_last and is_first:
                fueled = node.mass_initial - payload - shroud_mass
            elif is_last:
                fueled = node.mass_initial - payload
            elif is_first:
                fueled = node.mass_initial - shroud_mass - nxt.mass_initial
            else:
                fueled = node.mass_initial - nxt.mass_initial
            out.append({
                "prop":   node.mass_propellant,
                "dry":    fueled - node.mass_propellant,
                "fueled": fueled,
                "thrust": node.thrust_N,
                "dia":    node.diameter_m,
                "length": node.length_m,
                "solid":  bool(getattr(node, "solid_motor", False)
                               or getattr(node, "grain_type", "")),
            })
            node, idx = nxt, idx + 1
        return out

    # ------------------------------------------------------------------
    def _build(self, name):
        pad = dict(padx=6, pady=3)
        top = ttk.Frame(self, padding=10)
        top.pack(fill=tk.X)
        top.columnconfigure(1, weight=1)
        top.columnconfigure(3, weight=1)

        ttk.Label(top, text="Booster:").grid(row=0, column=0, sticky=tk.W, **pad)
        ttk.Label(top, text=name, foreground="navy").grid(
            row=0, column=1, sticky=tk.W, **pad)

        ttk.Label(top, text="Stage:").grid(row=0, column=2, sticky=tk.E, **pad)
        self._stage_var = tk.StringVar()
        stages = [f"Stage {i+1}" for i in range(len(self._stages))]
        self._stage_cb = ttk.Combobox(top, textvariable=self._stage_var,
                                      values=stages, state="readonly", width=10)
        self._stage_cb.grid(row=0, column=3, sticky=tk.W, **pad)
        self._stage_cb.current(0)
        self._stage_cb.bind("<<ComboboxSelected>>",
                            lambda _e: (self._load_stage(self._stage_cb.current()),
                                        self._compute()))

        # Propulsion type
        ttk.Label(top, text="Propulsion:").grid(row=1, column=0, sticky=tk.W, **pad)
        self._type_var = tk.StringVar(value="Liquid")
        tfrm = ttk.Frame(top)
        tfrm.grid(row=1, column=1, columnspan=3, sticky=tk.W, **pad)
        for t in ("Liquid", "Solid"):
            ttk.Radiobutton(tfrm, text=t, value=t, variable=self._type_var,
                            command=self._on_type).pack(side=tk.LEFT, padx=(0, 10))

        # Common parameter grid
        grid = ttk.LabelFrame(self, text="Stage parameters", padding=8)
        grid.pack(fill=tk.X, padx=10, pady=(0, 6))
        for c in (1, 3):
            grid.columnconfigure(c, weight=1)

        def _entry(parent, label, r, c, default="", width=12, unit=""):
            ttk.Label(parent, text=label).grid(row=r, column=c, sticky=tk.W, **pad)
            var = tk.StringVar(value=default)
            ent = ttk.Entry(parent, textvariable=var, width=width)
            ent.grid(row=r, column=c + 1, sticky=tk.W, **pad)
            if unit:
                ttk.Label(parent, text=unit).grid(row=r, column=c + 2,
                                                  sticky=tk.W)
            return var

        self._prop_mass_var = _entry(grid, "Propellant mass:", 0, 0, unit="kg")
        self._thrust_var    = _entry(grid, "Total thrust:",    0, 3, unit="kN")
        self._dia_var       = _entry(grid, "Diameter:",        1, 0, unit="m")
        self._len_var       = _entry(grid, "Length:",          1, 3, unit="m")
        self._gross_var     = _entry(grid, "Gross (wet) mass:",2, 0, unit="kg")
        self._pc_var        = _entry(grid, "Chamber pressure:",2, 3,
                                     default="6.9", unit="MPa")
        self._stated_var    = _entry(grid, "Stated dry mass:", 3, 0, unit="kg")

        # Guidance avionics: one package per vehicle, on the upper stage only
        # (never the bus).  Default on for the last stage, off for boosters.
        self._avionics_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(grid, text="Carries guidance avionics (upper stage)",
                        variable=self._avionics_var,
                        command=self._compute).grid(
            row=3, column=3, columnspan=3, sticky=tk.W, **pad)

        # Liquid-specific controls
        self._liq_frm = ttk.Frame(grid)
        self._liq_frm.grid(row=4, column=0, columnspan=6, sticky=tk.W + tk.E,
                           pady=(4, 0))
        ttk.Label(self._liq_frm, text="Propellant:").pack(side=tk.LEFT, padx=(0, 4))
        self._combo_var = tk.StringVar(value="LOX/RP1")
        ttk.Combobox(self._liq_frm, textvariable=self._combo_var,
                     values=mest.available_propellants(), state="readonly",
                     width=12).pack(side=tk.LEFT, padx=(0, 12))
        ttk.Label(self._liq_frm, text="Tank material:").pack(side=tk.LEFT,
                                                             padx=(0, 4))
        self._tankmat_var = tk.StringVar(value="aluminium")
        ttk.Combobox(self._liq_frm, textvariable=self._tankmat_var,
                     values=["aluminium", "al-li", "composite", "steel"],
                     state="readonly", width=10).pack(side=tk.LEFT, padx=(0, 12))
        ttk.Label(self._liq_frm, text="# engines:").pack(side=tk.LEFT, padx=(0, 4))
        self._neng_var = tk.StringVar(value="1")
        ttk.Entry(self._liq_frm, textvariable=self._neng_var, width=5).pack(
            side=tk.LEFT, padx=(0, 12))
        ttk.Label(self._liq_frm, text="Expansion ε:").pack(side=tk.LEFT, padx=(0, 4))
        self._exp_var = tk.StringVar(value="30")
        ttk.Entry(self._liq_frm, textvariable=self._exp_var, width=6).pack(
            side=tk.LEFT, padx=(0, 12))
        ttk.Label(self._liq_frm, text="Fairing area:").pack(side=tk.LEFT, padx=(0, 4))
        self._fair_var = tk.StringVar(value="0")
        ttk.Entry(self._liq_frm, textvariable=self._fair_var, width=7).pack(
            side=tk.LEFT)
        ttk.Label(self._liq_frm, text="m²").pack(side=tk.LEFT, padx=(2, 0))

        # Second liquid row: physics-based (GT-STRESS) tank sizing.
        self._liq_frm2 = ttk.Frame(grid)
        self._liq_frm2.grid(row=5, column=0, columnspan=6, sticky=tk.W,
                            pady=(2, 0))
        ttk.Label(self._liq_frm2, text="Tank model:").pack(side=tk.LEFT,
                                                           padx=(0, 4))
        self._tankmodel_var = tk.StringVar(value="akin_volume")
        ttk.Combobox(self._liq_frm2, textvariable=self._tankmodel_var,
                     values=["akin_volume", "akin_offset", "physics",
                             "averaged"], state="readonly", width=12).pack(
            side=tk.LEFT, padx=(0, 12))
        ttk.Label(self._liq_frm2, text="Lateral g:").pack(side=tk.LEFT, padx=(0, 4))
        self._latg_var = tk.StringVar(value="0.5")
        ttk.Entry(self._liq_frm2, textvariable=self._latg_var, width=5).pack(
            side=tk.LEFT, padx=(0, 8))
        ttk.Label(self._liq_frm2, text="Ullage MPa:").pack(side=tk.LEFT, padx=(0, 4))
        self._ullage_var = tk.StringVar(value="0.25")
        ttk.Entry(self._liq_frm2, textvariable=self._ullage_var, width=5).pack(
            side=tk.LEFT, padx=(0, 8))
        ttk.Label(self._liq_frm2, text="κ_E:").pack(side=tk.LEFT, padx=(0, 4))
        self._kappae_var = tk.StringVar(value="0")
        ttk.Entry(self._liq_frm2, textvariable=self._kappae_var, width=5).pack(
            side=tk.LEFT)

        # Solid-specific controls
        self._sol_frm = ttk.Frame(grid)
        self._sol_frm.grid(row=4, column=0, columnspan=6, sticky=tk.W,
                           pady=(4, 0))
        ttk.Label(self._sol_frm, text="Casing:").pack(side=tk.LEFT, padx=(0, 4))
        self._casing_var = tk.StringVar(value="steel")
        ttk.Combobox(self._sol_frm, textvariable=self._casing_var,
                     values=["steel", "composite"], state="readonly",
                     width=10).pack(side=tk.LEFT, padx=(0, 12))

        # Buttons
        btns = ttk.Frame(self, padding=(10, 0))
        btns.pack(fill=tk.X)
        ttk.Button(btns, text="Compute", command=self._compute).pack(side=tk.LEFT)
        ttk.Button(btns, text="Close", command=self.destroy).pack(
            side=tk.RIGHT)

        # Output
        outfrm = ttk.Frame(self, padding=(10, 6))
        outfrm.pack(fill=tk.BOTH, expand=True)
        self._out = tk.Text(outfrm, width=86, height=26, wrap="none",
                            font=("Courier", 9))
        vsb = ttk.Scrollbar(outfrm, orient=tk.VERTICAL, command=self._out.yview)
        self._out.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self._out.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    # ------------------------------------------------------------------
    def _on_type(self):
        solid = (self._type_var.get() == "Solid")
        # Show only the relevant sub-frame (both share the same grid cell).
        self._sol_frm.grid() if solid else self._sol_frm.grid_remove()
        for f in (self._liq_frm, self._liq_frm2):
            f.grid_remove() if solid else f.grid()

    def _load_stage(self, idx):
        sd = self._stages[idx]
        self._type_var.set("Solid" if sd["solid"] else "Liquid")
        self._prop_mass_var.set(f"{sd['prop']:.0f}")
        self._thrust_var.set(f"{sd['thrust']/1e3:.1f}")    # kN
        self._dia_var.set(f"{sd['dia']:.3f}")
        self._len_var.set(f"{sd['length']:.2f}")
        self._gross_var.set(f"{sd['fueled']:.0f}")
        self._stated_var.set(f"{sd['dry']:.0f}")
        # Avionics defaults on for the upper (last) stage only.
        self._avionics_var.set(idx == len(self._stages) - 1)
        self._on_type()

    # ------------------------------------------------------------------
    @staticmethod
    def _f(var, default=0.0):
        try:
            return float(var.get())
        except (ValueError, tk.TclError):
            return default

    def _compute(self):
        try:
            prop_mass = self._f(self._prop_mass_var)
            thrust_n  = self._f(self._thrust_var) * 1e3      # kN → N
            dia       = self._f(self._dia_var)
            length    = self._f(self._len_var)
            gross     = self._f(self._gross_var)
            pc_pa     = self._f(self._pc_var, 6.9) * 1e6     # MPa → Pa
            stated    = self._f(self._stated_var)

            avionics = bool(self._avionics_var.get())
            glow     = self._params.mass_initial   # vehicle GLOW for avionics
            if self._type_var.get() == "Solid":
                inp = mest.SolidStageInputs(
                    prop_mass_kg=prop_mass, thrust_n=thrust_n,
                    chamber_pressure_pa=pc_pa, casing=self._casing_var.get(),
                    diameter_m=dia, length_m=length, gross_mass_kg=gross,
                    include_avionics=avionics, vehicle_gross_kg=glow)
                estimates, report = mest.analyse_solid(inp, stated)
            else:
                inp = mest.LiquidStageInputs(
                    propellant=self._combo_var.get(), prop_mass_kg=prop_mass,
                    thrust_n=thrust_n, n_engines=int(self._f(self._neng_var, 1)),
                    expansion_ratio=self._f(self._exp_var, 30.0),
                    chamber_pressure_pa=pc_pa, diameter_m=dia, length_m=length,
                    gross_mass_kg=gross,
                    fairing_area_m2=self._f(self._fair_var),
                    tank_material=self._tankmat_var.get(),
                    include_avionics=avionics, vehicle_gross_kg=glow,
                    tank_model=self._tankmodel_var.get(),
                    lateral_g=self._f(self._latg_var, 0.5),
                    ullage_pa=self._f(self._ullage_var, 0.25) * 1e6,
                    kappa_e=self._f(self._kappae_var, 0.0),
                    stage_role=("lower" if self._stage_cb.current() == 0
                                else "upper"))
                estimates, report = mest.analyse_liquid(inp, stated)

            lines = []
            for e in estimates:
                lines.append(e.table())
                lines.append("")
            if report:
                lines.append("Divergence of stated dry mass vs. estimates:")
                lines.append(mest.format_divergence(report))
            text = "\n".join(lines)
        except Exception as exc:                            # pragma: no cover
            text = f"Could not compute estimate:\n{exc}"

        self._out.configure(state=tk.NORMAL)
        self._out.delete("1.0", tk.END)
        self._out.insert("1.0", text)
        self._out.configure(state=tk.DISABLED)


class ScreeningEnvelopeDialog(tk.Toplevel):
    """View and adjust the screening-envelope benchmark numbers (thresholds.py).

    The shipped defaults are frozen: an edit lives only in the overlay file,
    and "Restore All Defaults" discards it.  Each row shows the current value,
    its units, the greyed shipped default, and the default's citation of
    record, plus a "modified" flag when the two differ.  A changed number
    self-discloses in the survivability report, so the citation on the row is
    the *default's* provenance, not a warranty for the edited value.

    Design decision (README "Adjustable screening thresholds"): a policy
    modeler is likelier to adjust the ENVELOPE (glide endurance, maneuver
    ceiling) in light of a new flight than to re-coupon a material — so this
    curated set of ~9 envelope numbers is the editable surface; the material
    catalog and anchor datasets are deferred to a future spreadsheet project.
    """

    def __init__(self, app):
        super().__init__(app)
        self._app = app
        self.title("Screening Envelope — benchmark thresholds")
        self.resizable(False, False)
        self._vars = {}      # key -> StringVar (entry text)
        self._badges = {}    # key -> Label (the "modified" flag)
        self._build()
        try:
            app._center_dialog(self)
        except Exception:
            pass

    def _build(self):
        pad = dict(padx=6, pady=2)
        outer = ttk.Frame(self)
        outer.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        ttk.Label(
            outer, wraplength=560, justify=tk.LEFT,
            text=("Screening-envelope benchmarks. Shipped defaults are frozen; "
                  "an edit lives only in the overlay file and can always be "
                  "restored. A changed number self-discloses in the "
                  "survivability report — the citation shown is the default's "
                  "provenance, not a warranty for an edited value.")
        ).pack(anchor=tk.W, pady=(0, 8))

        grid = ttk.Frame(outer)
        grid.pack(fill=tk.BOTH, expand=True)
        cur = thresholds.current()
        r = 0
        last_group = None
        _grey = "#777777"
        for e in thresholds.REGISTRY:
            if e["group"] != last_group:
                if last_group is not None:
                    ttk.Separator(grid, orient=tk.HORIZONTAL).grid(
                        row=r, column=0, columnspan=4, sticky="ew", pady=(6, 2))
                    r += 1
                ttk.Label(grid, text=e["group"],
                          font=("TkDefaultFont", 9, "bold")).grid(
                    row=r, column=0, columnspan=4, sticky=tk.W, **pad)
                r += 1
                last_group = e["group"]

            key = e["key"]
            ttk.Label(grid, text=e["label"]).grid(
                row=r, column=0, sticky=tk.W, **pad)
            var = tk.StringVar(value=f"{cur[key]:g}")
            self._vars[key] = var
            ent = ttk.Entry(grid, textvariable=var, width=11)
            ent.grid(row=r, column=1, sticky=tk.W, **pad)
            ttk.Label(grid, text=e["units"], foreground=_grey).grid(
                row=r, column=2, sticky=tk.W, **pad)
            badge = ttk.Label(grid, text="", foreground="#2f6fb0",
                              font=("TkDefaultFont", 8, "bold"))
            badge.grid(row=r, column=3, sticky=tk.W, **pad)
            self._badges[key] = badge
            var.trace_add("write", lambda *_a, k=key: self._refresh_badge(k))
            r += 1

            ttk.Label(grid, text=f"default {e['default']:g} {e['units']}".strip(),
                      foreground=_grey, font=("TkDefaultFont", 8)).grid(
                row=r, column=1, columnspan=3, sticky=tk.W, padx=6)
            r += 1
            ttk.Label(grid, text=e["source"], foreground=_grey, wraplength=520,
                      justify=tk.LEFT, font=("TkDefaultFont", 8, "italic")).grid(
                row=r, column=1, columnspan=3, sticky=tk.W, padx=6, pady=(0, 2))
            r += 1
            self._refresh_badge(key)

        btns = ttk.Frame(outer)
        btns.pack(fill=tk.X, pady=(10, 0))
        ttk.Button(btns, text="Restore All Defaults",
                   command=self._restore_defaults).pack(side=tk.LEFT)
        ttk.Button(btns, text="Cancel", command=self.destroy).pack(side=tk.RIGHT)
        ttk.Button(btns, text="Save", command=self._save).pack(
            side=tk.RIGHT, padx=(0, 6))

    def _refresh_badge(self, key):
        """Show a 'modified' flag when the entry differs from the default."""
        e = thresholds._BY_KEY[key]
        badge = self._badges.get(key)
        if badge is None:
            return
        try:
            v = float(self._vars[key].get())
            changed = (v != float(e["default"]))
        except ValueError:
            changed = True   # unparseable text is not the default
        badge.configure(text="● modified" if changed else "")

    def _restore_defaults(self):
        thresholds.reset()
        for key, var in self._vars.items():
            var.set(f"{thresholds._BY_KEY[key]['default']:g}")
            self._refresh_badge(key)

    def _save(self):
        # Parse every field; clamp/store via set_override (value==default clears).
        bad = []
        for e in thresholds.REGISTRY:
            txt = self._vars[e["key"]].get().strip()
            try:
                thresholds.set_override(e["key"], float(txt))
            except ValueError:
                bad.append(e["label"])
        if bad:
            messagebox.showerror(
                "Invalid value",
                "These fields are not numbers:\n  • " + "\n  • ".join(bad),
                parent=self)
            return
        thresholds.save()      # writes the overlay (or deletes it at defaults)
        thresholds.apply()     # push into the live models
        # Re-render the survivability report so any change (and the modified-
        # benchmark disclosure) shows immediately.
        if getattr(self._app, "_result", None) is not None:
            try:
                self._app._populate_survivability(self._app._result)
            except Exception:
                pass
        self.destroy()


# ---------------------------------------------------------------------------
# Main application window
# ---------------------------------------------------------------------------
class BoosterFlyoutApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title("Thrusty — A Booster Performance Calculator")
        # Disable macOS window-tabbing so the system doesn't inject
        # Hide/Show Tab Bar into the View menu (Thrusty is single-window).
        try:
            self.tk.call('::tk::unsupported::MacWindowStyle',
                         'tabbingMode', self._w, 'disallowed')
        except tk.TclError:
            pass   # non-macOS platforms ignore this silently
        # Bespoke app icon (assets/thrusty.png, transparent PNG).  On macOS
        # aqua Tk this sets the running app's DOCK icon (replacing the stock
        # Python rocket); on Windows/Linux it sets the window/taskbar icon.
        # Silently keeps the stock icon when the asset is absent.
        try:
            _icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                      "assets", "thrusty.png")
            if os.path.exists(_icon_path):
                self._app_icon = tk.PhotoImage(file=_icon_path)  # keep a ref
                self.iconphoto(True, self._app_icon)
        except tk.TclError:
            pass
        self.minsize(900, 700)
        # Size to 92 % of the available screen, capped at 1600 × 1050.
        self.update_idletasks()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        w = min(1600, int(sw * 0.92))
        h = min(1050, int(sh * 0.92))
        x = (sw - w) // 2
        y = max(0, (sh - h) // 2 - 24)   # shift up slightly for macOS menu bar
        self.geometry(f"{w}x{h}+{x}+{y}")

        self._result         = None
        self._running        = False
        self._cancel_event   = threading.Event()
        # Max Range state: writes its optimum to the reserved "max-range" plan
        # variant instead of mutating the active plan (see _write_max_range_variant).
        self._max_range_pending   = False
        self._max_range_base_plan = None
        self._max_range_base_law  = "pitch_program"
        self._max_range_on_reserved = False
        self._max_range_context   = ""
        self._plan_orbit_base_plan = None
        self._plan_orbit_context   = ""
        self._notam_overlay  = None   # list of GeoJSON-style polygon rings, or None
        self._units_var      = tk.StringVar(value="km")  # plot display units

        _migrate_terminal_dive_default()  # migrate: dive default 30 km -> 0 (glide to impact)
        _migrate_analytic_family()   # migrate: banked/targeted analytic plans -> numerical family
        _load_ro_library()           # populate RO_DB from ro_library/*.ro.json
        _load_custom_boosters()      # restore any user-saved boosters
        thresholds.load()            # restore any saved screening-threshold overrides
        thresholds.apply()           # push them (or the shipped defaults) into the live models
        # Restore per-booster named flight-plan selections.
        try:
            if _ACTIVE_PLANS_PATH.exists():
                mm.ACTIVE_FLIGHT_PLANS.update(
                    json.loads(_ACTIVE_PLANS_PATH.read_text()))
        except Exception as exc:
            print(f"Warning: could not load active flight plans: {exc}")
        # Restore per-object named reentry-plan selections.
        try:
            if _ACTIVE_REENTRY_PLANS_PATH.exists():
                mm.ACTIVE_REENTRY_PLANS.update(
                    json.loads(_ACTIVE_REENTRY_PLANS_PATH.read_text()))
        except Exception as exc:
            print(f"Warning: could not load active reentry plans: {exc}")
        _extract_ros_from_boosters() # migrate: pull embedded RVs into the library

        self._build_menu()
        self._build_ui()
        self._on_booster_changed()   # populate params tab with default booster

    # ------------------------------------------------------------------
    # Utility
    # ------------------------------------------------------------------
    def _center_dialog(self, dlg):
        """Centre a Toplevel dialog over the main window."""
        dlg.update_idletasks()
        px = self.winfo_rootx() + (self.winfo_width()  - dlg.winfo_reqwidth())  // 2
        py = self.winfo_rooty() + (self.winfo_height() - dlg.winfo_reqheight()) // 2
        dlg.geometry(f"+{max(0, px)}+{max(0, py)}")

    # ------------------------------------------------------------------
    # Menu
    # ------------------------------------------------------------------
    def _build_menu(self):
        menubar = tk.Menu(self)

        file_menu = tk.Menu(menubar, tearoff=0)
        # ── Modeling inputs (load/save) ───────────────────────────────
        file_menu.add_command(label="Load Booster…",            command=self._load_booster)
        file_menu.add_command(label="Save Booster…",            command=self._export_booster)
        file_menu.add_command(label="Load Booster from XLSX…",  command=self._import_booster_xlsx)
        file_menu.add_command(label="Save Booster to XLSX…",    command=self._export_booster_xlsx)
        file_menu.add_command(label="New Booster XLSX Template…", command=self._new_booster_template)
        file_menu.add_command(label="Export 3-D Model (Blender/OBJ)…",
                              command=self._export_blender)
        file_menu.add_separator()
        file_menu.add_command(label="Load Reentry Object…",                 command=self._load_ro)
        file_menu.add_command(label="Save Reentry Object…",                 command=self._export_ro)
        file_menu.add_command(label="Load Reentry Object from XLSX…",       command=self._import_ro_xlsx)
        file_menu.add_command(label="Save Reentry Object to XLSX…",         command=self._export_ro_xlsx)
        file_menu.add_command(label="New Reentry Object XLSX Template…",    command=self._new_ro_template)
        file_menu.add_separator()
        file_menu.add_command(label="Load Flight Plan…",        command=self._import_flight_plan_file)
        file_menu.add_command(label="Save Flight Plan…",        command=self._export_flight_plan_file)
        file_menu.add_separator()
        file_menu.add_command(label="Load Launch Site…",        command=self._load_site)
        file_menu.add_command(label="Save Launch Site…",        command=self._export_site)
        file_menu.add_separator()
        file_menu.add_command(label="Load Scenario…",           command=self._load_scenario)
        file_menu.add_command(label="Save Scenario…",           command=self._save_scenario)
        file_menu.add_separator()
        # ── Trajectory outcomes (export only) ─────────────────────────
        file_menu.add_command(label="Export Trajectory CSV…",   command=self._save_trajectory)
        file_menu.add_command(label="Export Trajectory XLSX…",  command=self._export_trajectory_xlsx)
        file_menu.add_command(label="Export Trajectory KML…",   command=self._export_kml)
        file_menu.add_separator()
        # ── Flight events (export only) ───────────────────────────────
        file_menu.add_command(label="Export Flight Events CSV…",  command=self._export_timeline)
        file_menu.add_command(label="Export Flight Events XLSX…", command=self._export_timeline_xlsx)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.quit)
        menubar.add_cascade(label="File", menu=file_menu)

        analysis_menu = tk.Menu(menubar, tearoff=0)
        analysis_menu.add_command(label="Parametric Sweep…",        command=self._open_sweep)
        analysis_menu.add_command(label="Aim at Target (liquid)…",  command=self._aim_at_target)
        analysis_menu.add_command(label="Engine Cutoff (liquid)…",  command=self._set_engine_cutoff)
        analysis_menu.add_command(label="Re-entry Query…",          command=self._set_reentry_query)
        analysis_menu.add_command(label="Dry Mass Estimator…",      command=self._open_mass_estimator)
        analysis_menu.add_command(label="Screening Envelope…",      command=self._open_screening_envelope)

        # Reference Data — swap the empirical source behind a model term.
        # Built automatically from booster_models.MODEL_OPTIONS so new toggles
        # (atmosphere, etc.) appear here without touching this code.
        ref_menu = tk.Menu(analysis_menu, tearoff=0)
        self._model_option_vars = {}
        for _key, _spec in mm.MODEL_OPTIONS.items():
            _sub = tk.Menu(ref_menu, tearoff=0)
            _var = tk.StringVar(value=mm.get_model_option(_key))
            self._model_option_vars[_key] = _var
            for _choice in _spec["choices"]:
                _sub.add_radiobutton(
                    label=_spec["labels"][_choice], value=_choice, variable=_var,
                    command=lambda k=_key, v=_var: self._set_model_option(k, v.get()))
            ref_menu.add_cascade(label=_spec["label"], menu=_sub)
        # Offline gazetteer: an OPT-IN build for serious users (10.4 M
        # named places, worldwide, official BGN lineage — a multi-minute
        # / ~2.5 GB one-time index).  Kept out of the Find Location
        # dialog so a casual user is never ambushed by the wait.
        ref_menu.add_separator()
        ref_menu.add_command(label="Offline Gazetteer…",
                             command=self._manage_gazetteer)
        analysis_menu.add_separator()
        analysis_menu.add_cascade(label="Reference Data", menu=ref_menu)
        menubar.add_cascade(label="Analysis", menu=analysis_menu)

        # Plots menu mirrors the matplotlib navigation toolbar so the icon
        # strip doesn't have to live at the bottom of the plot panel.
        plots_menu = tk.Menu(menubar, tearoff=0)
        plots_menu.add_command(label="Home (reset view)",
                               command=lambda: self._plot_toolbar.home())
        plots_menu.add_command(label="Back",
                               command=lambda: self._plot_toolbar.back())
        plots_menu.add_command(label="Forward",
                               command=lambda: self._plot_toolbar.forward())
        plots_menu.add_separator()
        plots_menu.add_command(label="Pan",
                               command=lambda: self._plot_toolbar.pan())
        plots_menu.add_command(label="Zoom to rectangle",
                               command=lambda: self._plot_toolbar.zoom())
        plots_menu.add_separator()
        plots_menu.add_command(label="Configure subplots…",
                               command=lambda: self._plot_toolbar.configure_subplots())
        plots_menu.add_command(label="Save figure…",
                               command=lambda: self._plot_toolbar.save_figure())
        plots_menu.add_separator()
        plots_menu.add_command(label="Export Figures…", command=self._export_figures)
        menubar.add_cascade(label="Plots", menu=plots_menu)

        # Cartography — anything that produces or overlays a map: map
        # exports, NOTAM overlays, and the map-based analysis tools.
        carto_menu = tk.Menu(menubar, tearoff=0)
        carto_menu.add_command(label="Open Folium Map…",         command=self._export_folium)
        carto_menu.add_command(label="Export Cartopy Map…",      command=self._export_cartopy)
        carto_menu.add_command(label="Nearby Places…",           command=self._nearby_places)
        carto_menu.add_command(label="Gazetteer Explorer…",      command=self._gazetteer_explorer)
        carto_menu.add_separator()
        carto_menu.add_command(label="HGV Footprint…",           command=self._open_footprint)
        carto_menu.add_command(label="Range Ring (Cartopy)…",    command=self._open_range_ring)
        carto_menu.add_separator()
        carto_menu.add_command(label="Load NOTAM overlay…",      command=self._load_notam_overlay)
        carto_menu.add_command(label="Clear NOTAM overlay",      command=self._clear_notam_overlay)
        menubar.add_cascade(label="Cartography", menu=carto_menu)

        view_menu = tk.Menu(menubar, tearoff=0)
        view_menu.add_radiobutton(label="km  (metric)",       variable=self._units_var, value="km")
        view_menu.add_radiobutton(label="nmi  (nautical)",    variable=self._units_var, value="nm")
        view_menu.add_radiobutton(label="miles  (statute)",   variable=self._units_var, value="mi")
        menubar.add_cascade(label="View", menu=view_menu)

        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="About…", command=self._show_about)
        menubar.add_cascade(label="Help", menu=help_menu)

        self.config(menu=menubar)

    # ------------------------------------------------------------------
    # Top-level layout
    # ------------------------------------------------------------------
    def _build_ui(self):
        top = ttk.Frame(self)
        top.pack(fill=tk.BOTH, expand=True, padx=6, pady=4)

        # Resizable split — drag the sash to widen the control panel.
        paned = ttk.PanedWindow(top, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True)

        # Left control panel — vertically scrollable; width is user-adjustable.
        LEFT_W = 500
        self._left_hints = []                 # hint labels with dynamic wraplength
        self._left_wrap = LEFT_W - 28
        left_outer = ttk.Frame(paned, width=LEFT_W)
        left_outer.pack_propagate(False)

        left_canvas = tk.Canvas(left_outer, highlightthickness=0)
        left_vsb = ttk.Scrollbar(left_outer, orient=tk.VERTICAL,
                                  command=left_canvas.yview)
        left_canvas.configure(yscrollcommand=left_vsb.set)
        left_vsb.pack(side=tk.RIGHT, fill=tk.Y)
        left_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        left = ttk.Frame(left_canvas)
        _left_win = left_canvas.create_window((0, 0), window=left, anchor="nw")

        def _left_on_frame(event):
            left_canvas.configure(scrollregion=left_canvas.bbox("all"))
        def _left_on_canvas(event):
            left_canvas.itemconfig(_left_win, width=event.width)
            # keep every registered hint label wrapped to the current width so
            # all text stays visible at any pane size (A + C).
            self._left_wrap = max(160, event.width - 28)
            for _lbl in self._left_hints:
                try:
                    _lbl.configure(wraplength=self._left_wrap)
                except tk.TclError:
                    pass
        left.bind("<Configure>", _left_on_frame)
        left_canvas.bind("<Configure>", _left_on_canvas)

        for seq in ("<MouseWheel>", "<Button-4>", "<Button-5>"):
            left_canvas.bind(seq,
                lambda e, c=left_canvas: c.yview_scroll(
                    -1 if e.num == 4 else (1 if e.num == 5
                    else -1 * (e.delta // 120)), "units"))

        self._build_control_panel(left)

        # Right panel — tabbed notebook (Plots | Flight Timeline)
        right = ttk.Frame(paned)
        paned.add(left_outer, weight=0)
        paned.add(right, weight=1)

        # Set the initial sash position only once the paned window actually has
        # a width.  On macOS (and whenever the toplevel isn't mapped yet) the
        # first after_idle fires while winfo_width() is still ~1 px, and Tk
        # *clamps* sashpos(0, LEFT_W) down to that width — collapsing the left
        # control panel to zero.  Retry until the pane is realized, then set it.
        def _init_sash(attempt=0):
            if not paned.winfo_exists():
                return
            w = paned.winfo_width()
            if w <= 1 and attempt < 50:          # not realized yet — wait
                self.after(30, lambda: _init_sash(attempt + 1))
                return
            # On a normal window w >> LEFT_W so the sash sits at LEFT_W; on a
            # very narrow window degrade gracefully (leave ~80 px for the right
            # pane) instead of letting Tk clamp the left pane to zero.
            target = min(LEFT_W, w - 80) if w > 1 else LEFT_W
            paned.sashpos(0, max(120, target))
        self.after_idle(_init_sash)

        # Pinned results strip — always visible above the notebook tabs
        self._results_strip_var = tk.StringVar(value="")
        results_strip = ttk.Frame(right, relief=tk.GROOVE, borderwidth=1)
        results_strip.pack(fill=tk.X, padx=2, pady=(0, 3))
        ttk.Label(results_strip, textvariable=self._results_strip_var,
                  anchor=tk.W).pack(
            fill=tk.X, padx=8, pady=3)

        self._right_nb = ttk.Notebook(right)
        self._right_nb.pack(fill=tk.BOTH, expand=True)

        plots_tab    = ttk.Frame(self._right_nb)
        timeline_tab = ttk.Frame(self._right_nb)
        params_tab   = ttk.Frame(self._right_nb)
        schem_tab    = ttk.Frame(self._right_nb)
        slv_tab      = ttk.Frame(self._right_nb)
        surv_tab     = ttk.Frame(self._right_nb)
        self._right_nb.add(plots_tab,    text="  Plots  ")
        self._right_nb.add(timeline_tab, text="  Flight Timeline  ")
        self._right_nb.add(params_tab,   text="  Booster Parameters  ")
        self._right_nb.add(schem_tab,    text="  Schematic  ")
        self._right_nb.add(slv_tab,      text="  SLV Performance  ")
        self._right_nb.add(surv_tab,     text="  Reentry Survivability  ")

        self._build_plot_panel(plots_tab)
        self._build_timeline_panel(timeline_tab)
        self._build_params_tab(params_tab)
        self._build_schematic_tab(schem_tab)
        self._build_slv_tab(slv_tab)
        self._build_surv_tab(surv_tab)

        # Status bar
        self._status_var = tk.StringVar(value="Ready.")
        ttk.Label(self, textvariable=self._status_var,
                  relief=tk.SUNKEN, anchor=tk.W).pack(
            side=tk.BOTTOM, fill=tk.X, padx=4, pady=2)

    # ------------------------------------------------------------------
    # Control panel  (mirrors Forden's left-side panel)
    # ------------------------------------------------------------------
    def _build_control_panel(self, parent):
        # ── Booster type ───────────────────────────────────────────────
        mf = ttk.LabelFrame(parent, text="Booster Type")
        mf.pack(fill=tk.X, padx=6, pady=3)
        _cb_values   = list(BOOSTER_DB.keys())
        _first_valid = _cb_values[0] if _cb_values else ""
        self._last_valid_booster = _first_valid
        self._booster_var = tk.StringVar(value=_first_valid)
        self._booster_cb = ttk.Combobox(mf, textvariable=self._booster_var,
                                        values=_cb_values,
                                        state="readonly", width=24)
        self._booster_cb.pack(padx=6, pady=(4, 2))
        self._booster_cb.bind("<<ComboboxSelected>>", self._on_booster_changed)
        _bind_typeahead(self._booster_cb)

        mb = ttk.Frame(mf)
        mb.pack(padx=6, pady=(0, 4))
        ttk.Button(mb, text="New",    width=7,
                   command=self._new_booster).pack(side=tk.LEFT, padx=2)
        ttk.Button(mb, text="Edit…",  width=7,
                   command=self._edit_booster).pack(side=tk.LEFT, padx=2)
        self._del_btn = ttk.Button(mb, text="Delete", width=7,
                                   command=self._delete_booster,
                                   state=tk.DISABLED)
        self._del_btn.pack(side=tk.LEFT, padx=2)

        # ── Flight plan — a booster can fly many named plans ───────────
        # "(default)" is the booster-named plan file (undeletable); variants
        # are <booster>__<plan>.flightplan.json in the user library.  The
        # active selection feeds get_booster via ACTIVE_FLIGHT_PLANS, so the
        # whole panel (and every run) follows the chosen plan.
        # The Flight Plan section is created here but packed later, after Launch
        # Site, so the sidebar runs in flight order: Booster, Reentry Object,
        # Launch Site, Flight Plan, Reentry Plan.  It also hosts the ascent
        # strip (built below), consolidating what were two separate panels.
        fpf = ttk.LabelFrame(parent, text="Flight Plan")
        self._flight_plan_section = fpf
        self._fp_var = tk.StringVar(value=mm.DEFAULT_PLAN_LABEL)
        self._fp_cb = ttk.Combobox(fpf, textvariable=self._fp_var,
                                   values=[mm.DEFAULT_PLAN_LABEL],
                                   state="readonly", width=24)
        self._fp_cb.pack(padx=6, pady=(4, 2))
        self._fp_cb.bind("<<ComboboxSelected>>", self._on_flight_plan_selected)
        _bind_typeahead(self._fp_cb)
        fpb = ttk.Frame(fpf)
        fpb.pack(padx=6, pady=(0, 2))
        ttk.Button(fpb, text="New",   width=7,
                   command=self._new_flight_plan).pack(side=tk.LEFT, padx=2)
        ttk.Button(fpb, text="Edit…", width=7,
                   command=self._edit_flight_plan_main).pack(side=tk.LEFT, padx=2)
        self._fp_del_btn = ttk.Button(fpb, text="Delete", width=7,
                                      command=self._delete_flight_plan,
                                      state=tk.DISABLED)
        self._fp_del_btn.pack(side=tk.LEFT, padx=2)
        self._fp_summary_var = tk.StringVar(value="")
        ttk.Label(fpf, textvariable=self._fp_summary_var,
                  foreground="#555555").pack(padx=8, pady=(0, 4), anchor=tk.W)

        # ── Reentry vehicle (payload) ─────────────────────────────────
        # The RV library is independent of any booster; the run-time
        # selection here overrides whatever RV the booster was saved with.
        rf = ttk.LabelFrame(parent, text="Reentry Object")
        rf.pack(fill=tk.X, padx=6, pady=3)
        self._ro_main_var = tk.StringVar(value="(booster default)")
        self._ro_main_cb = ttk.Combobox(rf, textvariable=self._ro_main_var,
                                        values=self._ro_combo_values(),
                                        state="readonly", width=24)
        self._ro_main_cb.pack(padx=6, pady=(4, 2))
        self._ro_main_cb.bind("<<ComboboxSelected>>", self._on_ro_selected_main)
        _bind_typeahead(self._ro_main_cb)

        rb = ttk.Frame(rf)
        rb.pack(padx=6, pady=(0, 4))
        ttk.Button(rb, text="New",   width=7,
                   command=self._new_ro_main).pack(side=tk.LEFT, padx=2)
        ttk.Button(rb, text="Edit…", width=7,
                   command=self._edit_ro_main).pack(side=tk.LEFT, padx=2)
        self._ro_del_btn = ttk.Button(rb, text="Delete", width=7,
                                      command=self._delete_ro_main,
                                      state=tk.DISABLED)
        self._ro_del_btn.pack(side=tk.LEFT, padx=2)

        # Loadout: how many of the selected object the stack carries through
        # boost (throw weight = bus + N × object mass, composed onto the
        # stage chain at run time).  One object is modeled on the way back —
        # its arc represents the pattern.  Non-separating (body) runs force
        # N = 1: a multi-object integrated warhead is meaningless.
        _lo = ttk.Frame(rf)
        _lo.pack(padx=6, pady=(0, 4))
        ttk.Label(_lo, text="Loadout:").pack(side=tk.LEFT)
        self._loadout_n_var = tk.StringVar(value="1")
        self._loadout_spin = ttk.Spinbox(
            _lo, textvariable=self._loadout_n_var, from_=1, to=24, width=4,
            command=lambda: self._update_params_display())
        self._loadout_spin.pack(side=tk.LEFT, padx=(4, 0))
        ttk.Label(_lo, text="× object carried through boost").pack(
            side=tk.LEFT, padx=(4, 0))
        # (The whole-body L/D estimate lives inside the Reentry Object editor,
        # inline next to the L/D field — not here; the sidebar space is scarce.)

        # ── Launch site ────────────────────────────────────────────────
        lf = ttk.LabelFrame(parent, text="Launch Site")
        lf.pack(fill=tk.X, padx=6, pady=3)

        _site_values, self._site_map = _load_launch_sites()
        self._site_var = tk.StringVar(value="")
        self._site_cb = ttk.Combobox(lf, textvariable=self._site_var,
                                     values=_site_values, state="readonly", width=26)
        self._site_cb.pack(padx=6, pady=(4, 2))
        self._site_cb.bind("<<ComboboxSelected>>", self._on_site_selected)
        _bind_typeahead(self._site_cb)

        sb = ttk.Frame(lf)
        sb.pack(padx=6, pady=(0, 4))
        ttk.Button(sb, text="New",    width=7,
                   command=self._new_site).pack(side=tk.LEFT, padx=2)
        ttk.Button(sb, text="Edit…",  width=7,
                   command=self._edit_site).pack(side=tk.LEFT, padx=2)
        self._site_del_btn = ttk.Button(sb, text="Delete", width=7,
                                        command=self._delete_site,
                                        state=tk.DISABLED)
        self._site_del_btn.pack(side=tk.LEFT, padx=2)

        lf_grid = ttk.Frame(lf)
        lf_grid.pack(fill=tk.X)
        self._launch_lat = _dd_row(lf_grid, "Latitude:",  row=0, default="0.0")
        self._launch_lon = _dd_row(lf_grid, "Longitude:", row=1, default="0.0")
        # Buttons live in a shared column (2) so Find… and Estimate… align.
        ttk.Button(lf_grid, text="Find…", width=10,
                   command=lambda: self._pick_location(
                       self._launch_lat, self._launch_lon)
                   ).grid(row=1, column=2, sticky=tk.W, padx=4, pady=2)

        ttk.Label(lf_grid, text="Azimuth:").grid(row=2, column=0,
                                                  sticky=tk.W, padx=(8, 2), pady=2)
        az_frame = ttk.Frame(lf_grid)
        az_frame.grid(row=2, column=1, sticky=tk.W, padx=(0, 8), pady=2)
        self._azimuth_var = tk.StringVar(value="0.0")
        # Entry width matches the lat/lon rows (_dd_row uses 10).
        ttk.Entry(az_frame, textvariable=self._azimuth_var, width=10).pack(side=tk.LEFT)
        ttk.Label(az_frame, text="°  (from N)").pack(side=tk.LEFT, padx=(2, 0))
        ttk.Button(lf_grid, text="Estimate…", width=10,
                   command=self._estimate_azimuth
                   ).grid(row=2, column=2, sticky=tk.W, padx=4, pady=2)

        # Terrain (DEM): launch from real ground elevation and terminate on real
        # ground height (terrain.py — bundled coarse grid, offline).  Off by
        # default = flat sea-level Earth (the Forden benchmark condition).
        self._terrain_dem_var = tk.BooleanVar(value=False)
        _terr_row = ttk.Frame(lf_grid)
        _terr_row.grid(row=3, column=0, columnspan=3, sticky=tk.W,
                       padx=(8, 2), pady=(2, 4))
        ttk.Checkbutton(_terr_row, text="Use terrain (DEM)",
                        variable=self._terrain_dem_var,
                        command=self._on_terrain_toggled).pack(side=tk.LEFT)
        self._terrain_elev_lbl = ttk.Label(_terr_row, text="",
                                           foreground="#555555")
        self._terrain_elev_lbl.pack(side=tk.LEFT, padx=(6, 0))

        # Flight Plan section takes its place now (after Launch Site), then the
        # ascent strip is built inside it as a plain frame — one consolidated
        # panel instead of a separate "Ascent Mode" box.
        fpf.pack(fill=tk.X, padx=6, pady=3)

        # ── Ascent strip — mode + hot-loop pitch controls ────────────
        gf = ttk.Frame(fpf)
        gf.pack(fill=tk.X)
        self._guidance_frame = gf          # saved for dynamic grid management
        gf.columnconfigure(1, weight=1)    # column 1 fills available width

        self._guidance_var = tk.StringVar(value="pitch_program")
        gmode_frame = ttk.Frame(gf)
        gmode_frame.grid(row=0, column=0, columnspan=2, sticky=tk.EW,
                         padx=6, pady=(4, 2))
        ttk.Label(gmode_frame, text="Mode:").pack(side=tk.LEFT, padx=(0, 4))
        # Values are restricted to the active plan's guidance law by
        # _sync_ascent_mode_display — the law is plan identity and cannot be
        # switched here; only Simple/Advanced (same law) toggles in place.
        self._guidance_cb = ttk.Combobox(
            gmode_frame,
            values=["Simple pitch profile",
                    "Advanced pitch profile"],
            state="readonly",
            width=22,
        )
        self._guidance_cb.set("Simple pitch profile")
        self._guidance_cb.pack(side=tk.LEFT)
        self._guidance_cb.bind("<<ComboboxSelected>>",
                               lambda _e: self._on_guidance_changed())

        self._launch_el_lbl = ttk.Label(gf, text="Launch elev.:")
        self._launch_el_lbl.grid(row=1, column=0, sticky=tk.W, padx=(8, 2), pady=2)
        le_frame = ttk.Frame(gf)
        le_frame.grid(row=1, column=1, sticky=tk.W, padx=(0, 8), pady=2)
        self._launch_el_frame = le_frame
        self._launch_el_var = tk.StringVar(value="90.0")
        ttk.Entry(le_frame, textvariable=self._launch_el_var, width=8).pack(side=tk.LEFT)
        ttk.Label(le_frame, text="°  (90 = vertical)").pack(side=tk.LEFT, padx=2)

        self._loft_angle_lbl = ttk.Label(gf, text="Burnout Angle:")
        self._loft_angle_lbl.grid(row=3, column=0, sticky=tk.W, padx=(8, 2), pady=2)
        la_frame = ttk.Frame(gf)
        la_frame.grid(row=3, column=1, sticky=tk.W, padx=(0, 8), pady=2)
        self._loft_angle_frame = la_frame
        self._loft_angle_var = tk.StringVar(value="45.0")
        ttk.Entry(la_frame, textvariable=self._loft_angle_var, width=8).pack(side=tk.LEFT)
        self._loft_angle_unit_lbl = ttk.Label(la_frame, text="°  (final elev.)")
        self._loft_angle_unit_lbl.pack(side=tk.LEFT, padx=2)
        # Wheelon optimal-angle estimator — fills the field without a full
        # sweep.  Pitch-program law only; hidden otherwise (_update_guidance_labels).
        self._wheelon_main_btn = ttk.Button(
            la_frame, text="ε*", width=3, command=self._estimate_wheelon_main)
        self._wheelon_main_btn.pack(side=tk.LEFT, padx=(4, 0))

        self._gt_turn_start_lbl = ttk.Label(gf, text="Turn Start:")
        self._gt_turn_start_lbl.grid(row=5, column=0, sticky=tk.W, padx=(8, 2), pady=2)
        gt_ts_frame = ttk.Frame(gf)
        gt_ts_frame.grid(row=5, column=1, sticky=tk.W, padx=(0, 8), pady=2)
        self._gt_turn_start_frame = gt_ts_frame
        self._gt_turn_start_var = tk.StringVar(value="5.0")
        ttk.Entry(gt_ts_frame, textvariable=self._gt_turn_start_var, width=8).pack(side=tk.LEFT)
        ttk.Label(gt_ts_frame, text="s").pack(side=tk.LEFT, padx=2)

        self._gt_turn_stop_lbl = ttk.Label(gf, text="Turn Stop:")
        self._gt_turn_stop_lbl.grid(row=6, column=0, sticky=tk.W, padx=(8, 2), pady=2)
        gt_te_frame = ttk.Frame(gf)
        gt_te_frame.grid(row=6, column=1, sticky=tk.W, padx=(0, 8), pady=2)
        self._gt_turn_stop_frame = gt_te_frame
        self._gt_turn_stop_var = tk.StringVar(value="")
        # Autopopulated with the booster burn time; while untouched it still
        # reads as "optimize me" for Max Range (see _maximize_range).
        self._gt_turn_stop_auto = None
        ttk.Entry(gt_te_frame, textvariable=self._gt_turn_stop_var, width=8).pack(side=tk.LEFT)
        ttk.Label(gt_te_frame, text="s  (default = burn time)").pack(side=tk.LEFT, padx=2)

        self._orbit_alt_lbl = ttk.Label(gf, text="Target orbit alt:")
        self._orbit_alt_lbl.grid(row=7, column=0, sticky=tk.W, padx=(8, 2), pady=2)
        orb_frame = ttk.Frame(gf)
        orb_frame.grid(row=7, column=1, sticky=tk.W, padx=(0, 8), pady=2)
        self._orbit_alt_frame = orb_frame
        self._orbit_alt_var = tk.StringVar(value="400")
        ttk.Entry(orb_frame, textvariable=self._orbit_alt_var, width=8).pack(side=tk.LEFT)
        ttk.Label(orb_frame, text="km").pack(side=tk.LEFT, padx=2)

        # Row 8: Plan Orbit button — placed directly in gf (no wrapper Frame)
        # so grid_forget/grid reliably shows and hides it.
        self._plan_orbit_btn = ttk.Button(gf, text="Plan Orbit",
                                          command=self._plan_orbit)
        self._plan_orbit_btn.grid(row=8, column=0, columnspan=2,
                                  sticky=tk.EW, padx=8, pady=(4, 6), ipadx=2, ipady=4)

        # Row 9: Advanced pitch program toggle (gravity_turn / orbital_insertion only)
        self._adv_pitch_var = tk.BooleanVar(value=False)
        self._adv_pitch_chk = ttk.Checkbutton(
            gf, text="Advanced pitch program (per-stage)",
            variable=self._adv_pitch_var,
            command=self._on_adv_pitch_toggled)
        self._adv_pitch_chk.grid(row=9, column=0, columnspan=2,
                                  sticky=tk.W, padx=8, pady=(0, 2))

        # Row 10: Per-stage inline rows — rebuilt whenever booster changes
        self._adv_pitch_frame = ttk.Frame(gf)
        self._stage_rows = []   # list of dicts with StringVars per stage

        # Row 11: Yaw / dogleg program toggle (gravity_turn / orbital_insertion only)
        self._adv_yaw_var = tk.BooleanVar(value=False)
        self._adv_yaw_chk = ttk.Checkbutton(
            gf, text="Yaw / dogleg program",
            variable=self._adv_yaw_var,
            command=self._on_adv_yaw_toggled)

        # Row 10: Global yaw fields — shown when checkbox enabled
        # Three maneuvers laid out as a grid: rows=field, cols=maneuver
        yf = ttk.Frame(gf)
        self._adv_yaw_frame = yf
        self._yaw_vars = [
            {'start': tk.StringVar(value=""),
             'stop':  tk.StringVar(value=""),
             'final_az': tk.StringVar(value="")}
            for _ in range(3)
        ]
        for _mc, _hdr in enumerate(["#1", "#2", "#3"], start=1):
            ttk.Label(yf, text=_hdr, foreground="#555555").grid(
                row=0, column=_mc, padx=4, pady=(4, 1))
        for _yr, _lbl, _key, _unit in [
                (1, "Yaw start:", "start",    "s"),
                (2, "Yaw end:",   "stop",     "s"),
                (3, "Final az:",  "final_az", "°")]:
            ttk.Label(yf, text=_lbl).grid(
                row=_yr, column=0, sticky=tk.W, padx=(8, 2), pady=1)
            for _mc, _yvars in enumerate(self._yaw_vars, start=1):
                ttk.Entry(yf, textvariable=_yvars[_key], width=6).grid(
                    row=_yr, column=_mc, padx=3, pady=1)
            ttk.Label(yf, text=_unit).grid(
                row=_yr, column=4, sticky=tk.W, padx=(2, 8), pady=1)
        # Sustained angle-of-attack envelope enforced on the commanded attitude
        # while dynamic pressure is significant (NASA SP-8099 §2.1.2.2:
        # 5–10° at max q is the preliminary-design condition).  Blank = no
        # limit: the maneuver is flown as commanded and only flagged in the
        # Flight Timeline when it exceeds the envelope.
        self._alpha_limit_var = tk.StringVar(value="")
        ttk.Label(yf, text="α limit:").grid(
            row=4, column=0, sticky=tk.W, padx=(8, 2), pady=(4, 1))
        ttk.Entry(yf, textvariable=self._alpha_limit_var, width=6).grid(
            row=4, column=1, padx=3, pady=(4, 1))
        ttk.Label(yf, text="° (SP-8099: 5–10)", foreground="#555555").grid(
            row=4, column=2, columnspan=3, sticky=tk.W, padx=(2, 8),
            pady=(4, 1))
        # α-induced drag: the angle of attack develops an aerodynamic normal
        # force whose induced-drag component bleeds energy and reshapes q
        # (Jorgensen cross-flow; Fresconi 2017 / Kim 2013).  Off = the α load
        # is reported but does not feed back on the trajectory.
        self._alpha_induced_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            yf, text="α induced drag (bleeds energy, reshapes q)",
            variable=self._alpha_induced_var).grid(
            row=5, column=0, columnspan=5, sticky=tk.W, padx=(8, 8),
            pady=(1, 2))

        # Row 11: Reset trajectory button — always visible in Ascent Mode
        self._reset_traj_btn = ttk.Button(
            gf, text="Reset trajectory to defaults",
            command=self._reset_traj_profile)
        self._reset_traj_btn.grid(row=11, column=0, columnspan=2,
                                  sticky=tk.EW, padx=8, pady=(4, 2))

        # Initialise ascent-mode row visibility for the default mode.
        self._orbit_alt_lbl.grid_forget()
        self._orbit_alt_frame.grid_forget()
        self._plan_orbit_btn.grid_forget()
        self._adv_pitch_chk.grid_forget()
        self._adv_yaw_chk.grid_forget()
        self._update_guidance_labels("pitch_program")

        # Fairing jettison altitude is a flight-plan field edited in the Flight
        # Plan dialog (shown there only when the booster carries a shroud); it no
        # longer has its own sidebar box.  Keep the StringVar so _get_inputs and
        # the profile snapshot still read the value the plan supplies.
        self._shroud_jett_var = tk.StringVar(value="")

        # ── Reentry Mode ──────────────────────────────────────────────
        rf = ttk.LabelFrame(parent, text="Reentry Plan")
        rf.pack(fill=tk.X, padx=6, pady=3)
        self._reentry_frame = rf
        rf.columnconfigure(1, weight=1)

        # Row 0/1: reentry-plan variant selector, then its New/Edit…/Delete on
        # the row below — the down-leg analogue of the Flight Plan dropdown, laid
        # out like every other library section (combo on its own line, buttons
        # under it).  A reentry object can carry many named reentry plans
        # (default + variants); this switches between them, and the glider
        # controls below are the live editor (write-through on every run).
        self._rp_var = tk.StringVar(value=mm.DEFAULT_PLAN_LABEL)
        self._rp_cb = ttk.Combobox(rf, textvariable=self._rp_var,
                                   values=[mm.DEFAULT_PLAN_LABEL],
                                   state="readonly", width=24)
        self._rp_cb.grid(row=0, column=0, columnspan=2, padx=6, pady=(4, 2))
        self._rp_cb.bind("<<ComboboxSelected>>", self._on_reentry_plan_selected)
        _bind_typeahead(self._rp_cb)
        _rpbar = ttk.Frame(rf)
        _rpbar.grid(row=1, column=0, columnspan=2, pady=(0, 2))
        ttk.Button(_rpbar, text="New", width=7,
                   command=self._new_reentry_plan).pack(side=tk.LEFT, padx=2)
        ttk.Button(_rpbar, text="Edit…", width=7,
                   command=self._edit_reentry_plan_main).pack(side=tk.LEFT, padx=2)
        self._rp_del_btn = ttk.Button(_rpbar, text="Delete", width=7,
                                      command=self._delete_reentry_plan,
                                      state=tk.DISABLED)
        self._rp_del_btn.pack(side=tk.LEFT, padx=2)

        # Row 2: Separation — the run-level choice of whether the reentry
        # object separates at burnout or the last stage reenters whole
        # (Hwasong-11 / MaRV class).  A reentry-PLAN field like the glide law:
        # live here, written through to the active plan on every run, so the
        # same aeroshell can be A/B'd separating vs. integrated in two clicks.
        # Separation and Mode rows share fixed label width + combobox width so
        # the two dropdowns align as one column with left-justified labels.
        _sepbar = ttk.Frame(rf)
        _sepbar.grid(row=2, column=0, columnspan=2, sticky=tk.W, padx=6, pady=(3, 0))
        ttk.Label(_sepbar, text="Separation:", width=11,
                  anchor=tk.W).pack(side=tk.LEFT)
        self._SEP_LABELS = {
            'separating_ro': "Separates at burnout",
            'body':          "Non-separating (body reenters)",
        }
        self._main_sep_var = tk.StringVar(
            value=self._SEP_LABELS['separating_ro'])
        self._main_sep_cb = ttk.Combobox(
            _sepbar, textvariable=self._main_sep_var,
            values=list(self._SEP_LABELS.values()),
            state="readonly", width=32)
        self._main_sep_cb.pack(side=tk.LEFT, padx=(4, 0))
        self._main_sep_cb.bind(
            "<<ComboboxSelected>>",
            lambda _e: (self._refresh_glider_status_line(),
                        self._update_loadout_state(),
                        self._update_params_display()))

        # No status line: the strip explains only non-obvious inputs (house
        # rule) — vehicle properties (L/D, g-limit) live in Edit Reentry
        # Object….  The StringVar survives because several refresh paths
        # still write it (harmlessly, no widget displays it).
        self._glider_status_var = tk.StringVar(value="")

        # Row 1: reentry-mode detail frame.  Always visible; combobox at the
        # top selects the mode, rest of rows show/hide per selection.
        _gmf = ttk.Frame(rf)
        self._glider_main_frame = _gmf
        _gmf.columnconfigure(1, weight=1)

        # Reentry mode row — labeled like the Separation control above it.
        # Primary modes first; the analytic equilibrium-glide laws are kept below
        # a (non-selectable) separator as legacy/comparison modes.
        # FAMILY-SCOPED: the values list holds only the active plan's family
        # (numerical EOM or closed-form analytic), set by _scope_mode_choices —
        # the family is the plan's identity, chosen in New Reentry Plan and not
        # crossable from the strip.  Initial values = numerical (the default
        # family); populate re-scopes from the plan's mode.  (The former
        # "family: … — fixed for this plan" caption is gone: the family is
        # visible in New Reentry Plan where it is chosen, and the scoped
        # dropdown already enforces it.)
        _modebar = ttk.Frame(_gmf)
        _modebar.grid(row=0, column=0, columnspan=2,
                      sticky=tk.W, padx=6, pady=(2, 0))
        ttk.Label(_modebar, text="Mode:", width=11,
                  anchor=tk.W).pack(side=tk.LEFT)
        self._main_guidance_var = tk.StringVar(value="Ballistic (drag · gravity · rotation)")
        self._main_guidance_cb = ttk.Combobox(
            _modebar, textvariable=self._main_guidance_var,
            values=[lbl for _k, lbl in self._REENTRY_MODE_NUMERICAL],
            state="readonly", width=32)
        self._main_guidance_cb.pack(side=tk.LEFT, padx=(4, 0))
        self._main_guidance_cb.bind("<<ComboboxSelected>>",
                                     lambda _e: self._on_glider_guidance_changed())

        # Skip count belonged only to the retired skip_to_equilibrium mode; the
        # control is gone.  The var stays as an inert conduit (scenario
        # save/restore and the plan schema still carry glider_skip_count,
        # defaulting to 1) so nothing downstream needs to change.
        self._main_skip_count_var = tk.StringVar(value="1")

        # ζ is the one glide knob you iterate per-run for a damped or
        # dynamic-equilibrium plan (damping ratio / tracking gain), so it lives
        # on the strip with its estimator.  Terminal-dive altitude and the aero
        # model are set-once tuning and live in the Reentry Plan editor (Edit…);
        # their vars stay here as the in-memory conduit (populate ← plan, run →
        # reads them, dialog edits them through the plan file).
        self._main_zeta_var = tk.StringVar(value="0.7")
        self._main_dive_alt_var = tk.StringVar(value="0")
        # Commanded pull-up start altitude — set-once tuning in the plan editor;
        # this var is the in-memory conduit (populate ← plan, run → reads it,
        # dialog edits it via the plan file), exactly like _main_dive_alt_var.
        self._main_pullup_alt_var = tk.StringVar(value="0")
        self._main_aero_var = tk.StringVar(value="Drag polar (realistic)")

        # ζ row — shown only for the ζ-using glide laws (damped / dynamic).
        _r2 = ttk.Frame(_gmf)
        _r2.grid(row=2, column=0, columnspan=2, sticky=tk.W, padx=(8, 0), pady=1)
        self._main_glide_detail_frm = _r2
        self._main_zeta_lbl = ttk.Label(_r2, text="Damping ratio ζ:")
        self._main_zeta_lbl.pack(side=tk.LEFT)
        ttk.Entry(_r2, textvariable=self._main_zeta_var, width=6).pack(
            side=tk.LEFT, padx=2)
        self._main_zeta_est_btn = ttk.Button(
            _r2, text="Estimate…", width=10, command=self._estimate_main_zeta)
        self._main_zeta_est_btn.pack(side=tk.LEFT, padx=(4, 0))

        # Bank schedule and dive-at-target are detail (a 3×3 grid and a
        # lat/lon/radius block) that clutter the strip and are rarely re-tuned,
        # so they move into the Reentry Plan editor (Edit…).  Their vars stay
        # here as the in-memory conduit (populate ← plan, run → reads them,
        # dialog edits them via the plan file).  A one-line pointer replaces the
        # widgets, shown for any glide mode.
        self._main_bank_sched_var = tk.BooleanVar(value=False)
        self._main_bank_vars = [{'start': tk.StringVar(value=""),
                                  'end':   tk.StringVar(value=""),
                                  'bank':  tk.StringVar(value="")}
                                 for _ in range(3)]
        self._main_dive_target_var = tk.BooleanVar(value=False)
        self._main_dt_lat_var = tk.StringVar(value="0.0")
        self._main_dt_lon_var = tk.StringVar(value="0.0")
        self._main_dt_radius_var = tk.StringVar(value="20")

        # (No editor-pointer hint: house rule — only explain inputs that
        # might not be obvious.  Edit… is the affordance.)

        # Grid _glider_main_frame in its parent (rf); set mode-driven visibility.
        self._on_glider_guidance_changed()
        self._glider_main_frame.grid(row=4, column=0, columnspan=2,
                                      sticky=tk.EW, padx=0, pady=(0, 4))

        # ── Engine cutoff ─────────────────────────────────────────────
        # Liquid-engines-only and rarely used, so it lives under
        # Analysis ▸ Engine Cutoff… instead of taking left-panel space.
        # The variable stays: Aim-at-Target writes its computed cutoff here,
        # and scenarios/profiles save and restore it.
        self._cutoff_var = tk.StringVar(value="")

        # ── Re-entry query altitude ────────────────────────────────────
        # A per-run diagnostic ("report the state vector at N km on descent"),
        # not a plan field — it lives under Analysis ▸ Re-entry Query… (like
        # Engine Cutoff) instead of taking sidebar space.  The variables stay:
        # the run path reads them, and scenarios save/restore them.
        self._query_alt_enable = tk.BooleanVar(value=False)
        self._query_alt_km_var = tk.StringVar(value="50")

        # ── Run buttons ───────────────────────────────────────────────
        btn_frame = ttk.Frame(parent)
        btn_frame.pack(fill=tk.X, padx=6, pady=6)
        ttk.Button(btn_frame, text="Run Flyout",
                   command=self._run_flyout).pack(
            side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 2), ipady=4)
        ttk.Button(btn_frame, text="Max Range",
                   command=self._maximize_range).pack(
            side=tk.LEFT, expand=True, fill=tk.X, padx=2, ipady=4)
        self._cancel_max_btn = ttk.Button(btn_frame, text="■ Stop",
                                          command=self._cancel_max_range,
                                          state=tk.DISABLED)
        self._cancel_max_btn.pack(side=tk.LEFT, expand=False, padx=(0, 2), ipady=4)
        ttk.Button(btn_frame, text="Sweep…",
                   command=self._open_sweep).pack(
            side=tk.LEFT, expand=True, fill=tk.X, padx=(2, 0), ipady=4)

        # (Booster Parameters moved to right-panel notebook tab)

    # ------------------------------------------------------------------
    # SLV Performance tab  (Schilling / Townsend algebraic analysis)
    # ------------------------------------------------------------------
    def _build_slv_tab(self, parent):
        # ── Target orbit input ────────────────────────────────────────
        of = ttk.LabelFrame(parent, text="Target Orbit")
        of.pack(fill=tk.X, padx=8, pady=(8, 4))

        of_grid = ttk.Frame(of)
        of_grid.pack(padx=8, pady=6, anchor=tk.W)

        ttk.Label(of_grid, text="Perigee:").grid(
            row=0, column=0, sticky=tk.E, padx=(0, 4))
        self._slv_alt_var = tk.StringVar(value="500")
        ttk.Entry(of_grid, textvariable=self._slv_alt_var,
                  width=8).grid(row=0, column=1)
        ttk.Label(of_grid, text="km").grid(
            row=0, column=2, sticky=tk.W, padx=(4, 0))

        ttk.Label(of_grid, text="Apogee:").grid(
            row=1, column=0, sticky=tk.E, padx=(0, 4), pady=(4, 0))
        self._slv_apo_var = tk.StringVar(value="")
        ttk.Entry(of_grid, textvariable=self._slv_apo_var,
                  width=8).grid(row=1, column=1, pady=(4, 0))
        ttk.Label(of_grid, text="km  (blank = circular)").grid(
            row=1, column=2, sticky=tk.W, padx=(4, 0), pady=(4, 0))

        ttk.Button(of, text="Analyze SLV Performance",
                   command=self._run_slv_analysis).pack(pady=(0, 6))

        # ── Results ───────────────────────────────────────────────────
        rf = ttk.LabelFrame(parent, text="Results  (Schilling / Townsend method)")
        rf.pack(fill=tk.BOTH, expand=True, padx=8, pady=(4, 8))

        self._slv_text = tk.Text(
            rf, state=tk.DISABLED, font=("TkFixedFont", 9),
            wrap=tk.NONE, relief=tk.FLAT, background="#f8f8f8",
            foreground="#222222", selectbackground="#c0d8f0")
        _fam = tkfont.nametofont("TkFixedFont").actual()["family"]
        _tab = tkfont.Font(family=_fam, size=9).measure("x" * 32)
        self._slv_text.configure(tabs=(_tab,))
        vsb = ttk.Scrollbar(rf, orient=tk.VERTICAL,
                            command=self._slv_text.yview)
        self._slv_text.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self._slv_text.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        # Logo — overlaid on the results panel, bottom-left
        import os as _os
        _logo_path = _os.path.join(_os.path.dirname(__file__), "data", "Thrusty3.png")
        try:
            from PIL import Image, ImageTk as _ITk
            _img = Image.open(_logo_path)
            _h = 195
            _w = int(_img.width * _h / _img.height)
            _img = _img.resize((_w, _h), Image.LANCZOS)
            self._slv_logo_photo = _ITk.PhotoImage(_img)
            tk.Label(self._slv_text, image=self._slv_logo_photo,
                     borderwidth=0, highlightthickness=0, bg="#f8f8f8"
                     ).place(relx=1.0, rely=1.0, anchor="se", x=-6, y=-4)
        except Exception:
            pass

        # Tag for the headline verdict line
        self._slv_text.tag_configure("yes", foreground="#006600",
                                     font=("TkFixedFont", 9, "bold"))
        self._slv_text.tag_configure("no",  foreground="#aa0000",
                                     font=("TkFixedFont", 9, "bold"))
        self._slv_text.tag_configure("hdr", font=("TkFixedFont", 9, "bold"))

        self._slv_set_text(
            "Select a booster, set the launch site and azimuth in the left\n"
            "panel, enter a target orbit above, then click Analyze.\n\n"
            "For a circular orbit enter only the perigee altitude (apogee\n"
            "blank or equal to perigee).  For a Hohmann transfer or GTO set\n"
            "apogee higher than perigee; the booster burns to the perigee\n"
            "injection speed and coasts to apogee.\n\n"
            "The launch azimuth determines the orbital inclination and the\n"
            "Earth-rotation benefit (maximum for a due-east launch).\n\n"
            "Accuracy: ~260 m/s RMS in total mission ΔV; typically < 10 %\n"
            "error in payload capacity  (Schilling 2009).",
            verdict=None)

    # ------------------------------------------------------------------
    # Reentry Survivability tab — the mode-keyed report + flux/load plot
    # (SURVIVABILITY_REPORT_DESIGN.md; the down-leg Schilling panel).
    # Assembled by survivability_report.build_report over the numbers the
    # run already computed (result['heating_fom'] + ['heating_arc']).
    # ------------------------------------------------------------------
    def _build_surv_tab(self, parent):
        rf = ttk.LabelFrame(
            parent,
            text="Reentry Survivability  (mode-keyed screening report — "
                 "not a TPS design verdict)")
        rf.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        # Toolbar: one-click loft/depress heating sweep (ballistic RVs) — opens
        # the Parametric Sweep preset to Burnout Angle with the flux/load axes
        # checked, so the shaping trade is reachable from where the report is
        # read.  Its result feeds the Form A "Sweep context" block.
        _tb = ttk.Frame(rf)
        _tb.pack(fill=tk.X, padx=4, pady=(4, 0))
        ttk.Button(_tb, text="Loft / Depress heating sweep…",
                   command=lambda: self._open_sweep(param="Burnout Angle",
                                                    heating=True)
                   ).pack(side=tk.LEFT)
        ttk.Label(_tb, text="  sweeps burnout angle → peak flux vs integrated "
                            "load (the loft/depress trade)",
                  foreground="#888888").pack(side=tk.LEFT)

        # Top: flux(t) / load(t) plot — the mode's signature pulse shape.
        self._surv_fig = Figure(figsize=(6.0, 2.4), dpi=96)
        self._surv_canvas = FigureCanvasTkAgg(self._surv_fig, master=rf)
        self._surv_canvas.get_tk_widget().pack(fill=tk.X, padx=4, pady=(4, 2))

        # Bottom: the report text.
        self._surv_text = tk.Text(
            rf, state=tk.DISABLED, font=("TkFixedFont", 9),
            wrap=tk.WORD, relief=tk.FLAT, background="#f8f8f8",
            foreground="#222222", selectbackground="#c0d8f0")
        vsb = ttk.Scrollbar(rf, orient=tk.VERTICAL,
                            command=self._surv_text.yview)
        self._surv_text.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self._surv_text.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        import survivability_report as _sr
        _tier_tags = [(k, v[1]) for k, v in _sr.SURVIVAL_TIERS.items()]
        for tag, colour in (_tier_tags + [("none", "#555555")]):
            self._surv_text.tag_configure(
                tag, foreground=colour, font=("TkFixedFont", 11, "bold"))
        # Survival-map cell tags: tier colours at body size, colorized in
        # place.  Columns are TAB-separated and aligned with pixel tab stops
        # measured from the BOLD cell font (bold fixed fonts don't share
        # regular metrics on every platform — a character grid drifts).
        for tag, colour in _tier_tags:
            self._surv_text.tag_configure(
                "map_" + tag, foreground=colour,
                font=("TkFixedFont", 9, "bold"))
        import tkinter.font as _tkfont
        _cellf = _tkfont.Font(font=("TkFixedFont", 9, "bold"))
        _t0 = _cellf.measure("  Windward flank  ")     # widest row label
        _tw = _cellf.measure("X" * 22)                 # widest cell + gap
        self._surv_text.tag_configure(
            "mapgrid", tabs=(_t0, _t0 + _tw, _t0 + 2 * _tw))

        self._surv_set_text(
            None, "",
            "Fly a trajectory (Launch / Max Range).  This panel then reports\n"
            "the reentry object's heating survivability, keyed to its reentry\n"
            "mode: a ballistic RV is judged on the nose-recession accuracy\n"
            "ladder, a glider on survival-time vs glide-time and the NRC-2008\n"
            "TPS duration ladder.  Banking, a terminal dive and a commanded\n"
            "lift cap each add their own block when the plan carries them.\n"
            "The plot shows the flux pulse q̇(t) and the running load Q(t) —\n"
            "the pulse shape is the mode's signature.")

    def _surv_set_text(self, tier, headline, body, map_spans=None):
        self._surv_text.configure(state=tk.NORMAL)
        self._surv_text.delete("1.0", tk.END)
        if tier is not None:
            mark = {"experience": "✓  ", "design": "◆  ", "beyond": "△  ",
                    "fail": "✗  ", "none": "•  "}.get(tier, "")
            self._surv_text.insert(tk.END, mark + headline + "\n\n", tier)
        self._surv_text.insert(tk.END, body)
        # Colorize the survival-map cells: spans are (line-within-block, c0,
        # c1, tier) relative to the "─── Survival map" header line, so the
        # block is self-locating however the body around it is composed.
        if map_spans:
            idx = self._surv_text.search("─── Survival map", "1.0")
            if idx:
                base = int(idx.split(".")[0])
                # Tab-stop alignment on every tabbed line of the block.
                ln = base + 1
                while "\t" in self._surv_text.get(f"{ln}.0", f"{ln}.end"):
                    self._surv_text.tag_add("mapgrid", f"{ln}.0", f"{ln}.end")
                    ln += 1
                for ln_, c0, c1, tk_key in map_spans:
                    self._surv_text.tag_add(
                        "map_" + tk_key, f"{base + ln_}.{c0}",
                        f"{base + ln_}.{c1}")
        self._surv_text.configure(state=tk.DISABLED)

    def _populate_survivability(self, r):
        """Fill the Reentry Survivability tab from the trajectory result."""
        import survivability_report as _sr
        rep = _sr.build_report(r or {})
        body = rep['body']
        # Ballistic-RV loft/depress context — fed ONLY from the most recent
        # sweep (design decision 2), when it swept this booster with heating data.
        _sw = getattr(self, '_last_heating_sweep', None)
        if (rep.get('form') == 'ballistic' and _sw
                and _sw.get('booster') == self._booster_var.get()
                and len(_sw.get('rows') or []) >= 2):
            rows = [row for row in _sw['rows']
                    if np.isfinite(row[3]) and np.isfinite(row[4])]
            if len(rows) >= 2:
                q_lo = min(rows, key=lambda w: w[3]); q_hi = max(rows, key=lambda w: w[3])
                l_lo = min(rows, key=lambda w: w[4]); l_hi = max(rows, key=lambda w: w[4])
                u = _sw.get('unit', '')
                body += (
                    f"\n─── Sweep context ({_sw['param']}, "
                    f"{rows[0][0]:g}–{rows[-1][0]:g} {u}) ─────────────\n"
                    f"  Peak flux spans {q_lo[3]:.1f} → {q_hi[3]:.1f} MW/m² "
                    f"(min at {q_lo[0]:g}{u}, max at {q_hi[0]:g}{u})\n"
                    f"  Heat load spans {l_lo[4]:.0f} → {l_hi[4]:.0f} MJ/m² "
                    f"(min at {l_lo[0]:g}{u}, max at {l_hi[0]:g}{u})\n"
                    f"  Trajectory shaping trades flux against load "
                    f"(lofted = flux-stressed, depressed = load-stressed).\n")
        self._surv_set_text(
            rep.get('tier') if rep.get('form') else 'none',
            rep['headline'], body, rep.get('map_spans'))

        self._surv_fig.clf()
        pl = rep.get('plot')
        if pl is not None and len(pl['t']) > 1:
            ax = self._surv_fig.add_subplot(111)
            ax2 = ax.twinx()
            # UHTC envelope-coverage shading (SRD §11.3): green = protected
            # (below the 1650 °C glass ceiling), amber = inside the
            # demonstrated envelope consuming recession margin, red =
            # beyond the demonstrated record (too hot / too long — see the
            # report text; NOT a failure prediction by itself).  The bands
            # classify TIME segments by nose surface temperature — they
            # correspond to neither y-axis; the caption says so.
            if pl.get('bands'):
                import survivability_report as _sr
                # bands now carry survival-tier keys; fall back to legacy keys.
                _band_col = {k: v[1] for k, v in _sr.SURVIVAL_TIERS.items()}
                _band_col.update({'green': "#2e8b57", 'amber': "#2e8b57",
                                  'red': "#2f6fb0"})
                for _b0, _b1, _c in pl['bands']:
                    if _b1 > _b0:
                        ax.axvspan(_b0, _b1, color=_band_col.get(_c, "#888888"),
                                   alpha=0.14, linewidth=0)
                ax.set_title(
                    "shading = nose temperature vs the survival ladder:  "
                    "green within experience · blue within design · yellow "
                    "beyond design",
                    fontsize=6.5, color="#555555", loc="left")
            ax.plot(pl['t'], pl['q_MW'], color="#aa2222", linewidth=1.4,
                    label="q̇ (MW/m²)")
            ax2.plot(pl['t'], pl['Q_MJ'], color="#2255aa", linewidth=1.4,
                     linestyle="--", label="Q (MJ/m²)")
            ax.set_xlabel("time from reentry-arc start (s)", fontsize=8)
            ax.set_ylabel("Flux:  q̇  MW/m²", fontsize=8, color="#aa2222")
            ax2.set_ylabel("Integrated load:  Q  MJ/m²", fontsize=8,
                           color="#2255aa")
            ax.tick_params(labelsize=7)
            ax2.tick_params(labelsize=7)
            if pl.get('t_fail') is not None:
                ax.axvline(pl['t_fail'], color="#aa0000", linewidth=1.0,
                           linestyle=":", alpha=0.9)
                ax.text(pl['t_fail'], ax.get_ylim()[1] * 0.95, " t_fail",
                        fontsize=7, color="#aa0000", va="top")
            # NRC duration ticks for gliders (300 / 800 / 3000 / 3600 s),
            # only those inside the plotted span.
            if pl.get('glide_s') and pl.get('tiers'):
                _tmax = float(pl['t'][-1])
                for _lbl, _sec, _mat, _tier in pl['tiers']:
                    if _sec and _sec <= _tmax:
                        ax.axvline(_sec, color="#888888", linewidth=0.7,
                                   linestyle=":", alpha=0.6)
                        ax.text(_sec, ax.get_ylim()[1] * 0.02, f" {_sec:g}s",
                                fontsize=6, color="#888888", va="bottom")
            self._surv_fig.tight_layout()
        self._surv_canvas.draw_idle()

    def _slv_set_text(self, body: str, verdict=None):
        """Replace the SLV results text widget contents."""
        self._slv_text.configure(state=tk.NORMAL)
        self._slv_text.delete("1.0", tk.END)
        if verdict is not None:
            tag = "yes" if verdict else "no"
            mark = "✓  CAN reach orbit" if verdict else "✗  CANNOT reach orbit"
            self._slv_text.insert(tk.END, mark + "\n", tag)
            self._slv_text.insert(tk.END, "\n")
        self._slv_text.insert(tk.END, body)
        self._slv_text.configure(state=tk.DISABLED)

    def _run_slv_analysis(self):
        # ── Parse inputs ──────────────────────────────────────────────
        try:
            perigee_km = float(self._slv_alt_var.get())
        except ValueError:
            messagebox.showerror("Input error",
                                 "Perigee altitude must be a number (km).",
                                 parent=self)
            return
        if perigee_km <= 0:
            messagebox.showerror("Input error",
                                 "Perigee altitude must be positive.", parent=self)
            return

        apo_str = self._slv_apo_var.get().strip()
        if apo_str == "" or apo_str == str(perigee_km):
            apogee_km = None           # circular
        else:
            try:
                apogee_km = float(apo_str)
            except ValueError:
                messagebox.showerror("Input error",
                                     "Apogee altitude must be a number or blank.",
                                     parent=self)
                return
            if apogee_km < perigee_km:
                messagebox.showerror("Input error",
                                     "Apogee must be ≥ perigee.", parent=self)
                return

        try:
            lat = float(self._launch_lat.get())
            az  = float(self._azimuth_var.get())
        except ValueError:
            messagebox.showerror("Input error",
                                 "Check launch latitude and azimuth.", parent=self)
            return

        booster = get_booster(self._booster_var.get())

        try:
            r = schilling_performance(booster, perigee_km, lat, az,
                                      target_apogee_km=apogee_km)
        except Exception as exc:
            messagebox.showerror("Analysis error", str(exc), parent=self)
            return

        # ── Format results ────────────────────────────────────────────
        from slv_performance import stage_delta_v as _sdv

        n_stages = 0
        s = booster
        while s:
            n_stages += 1
            s = s.stage2

        stage_lines = []
        s, i = booster, 1
        while s:
            stage_lines.append(
                f"    Stage {i} ({s.isp_s:.0f} s Isp):\t{_sdv(s):8.0f} m/s")
            s = s.stage2
            i += 1

        is_circular = (apogee_km is None or
                       apogee_km == perigee_km)
        if is_circular:
            orbit_desc = f"{perigee_km:.0f} km circular"
            inj_label  = "Circular orbit speed:"
        else:
            orbit_desc = (f"{perigee_km:.0f} × {r['orbit_apogee_km']:.0f} km  "
                          f"(e = {r['orbit_eccentricity']:.4f})")
            inj_label  = "Injection speed (perigee):"

        margin_sign = "+" if r['dv_margin_ms'] >= 0 else ""

        payload_line = ""
        if booster.payload_kg > 0:
            pm   = r['payload_margin_kg'] or 0.0
            sign = "+" if pm >= 0 else ""
            payload_line = (
                f"  Claimed payload:\t{booster.payload_kg:8.0f} kg\n"
                f"  Payload margin:\t{sign}{pm:7.0f} kg\n"
            )

        body = (
            f"Booster:  {booster.name}  ({n_stages}-stage)\n"
            f"Target:   {orbit_desc}\n"
            f"Launch:   {lat:.2f}° lat,  azimuth {az:.1f}°\n"
            "\n"
            "─── Delta-V Budget ─────────────────────────────────────────\n"
            "  Available ΔV (rocket eq.):\n"
            + "\n".join(stage_lines) + "\n"
            f"  Total available:\t{r['dv_available_ms']:8.0f} m/s\n"
            "\n"
            "  Required to reach orbit:\n"
            f"    {inj_label}\t{r['v_injection_ms']:8.0f} m/s\n"
            f"    Loss penalty (eq. 5):\t{r['dv_penalty_ms']:8.0f} m/s\n"
            f"    Earth rotation benefit:\t{-r['v_rotation_ms']:8.0f} m/s\n"
            f"  Total required:\t{r['dv_required_ms']:8.0f} m/s\n"
            "\n"
            f"  Margin:\t{margin_sign}{r['dv_margin_ms']:7.0f} m/s\n"
            "\n"
            "─── Payload Capacity ───────────────────────────────────────\n"
            f"  Maximum payload:\t{r['max_payload_kg']:8.0f} kg\n"
            + payload_line +
            "\n"
            "─── Schilling Timing Parameters ────────────────────────────\n"
            f"  Actual burn time  (Tₐ):\t{r['t_actual_s']:8.1f} s\n"
            f"  3-stage equiv.    (T₃ₛ):\t{r['t_3stage_s']:8.1f} s\n"
            f"  Blended time    (T_mix):\t{r['t_mix_s']:8.1f} s\n"
            f"  Initial accel.    (A₀):\t{r['a0_ms2']:8.2f} m/s²"
            f"  ({r['a0_ms2'] / 9.80665:.2f} g)\n"
            "\n"
            "Method accuracy: ±260 m/s RMS in mission ΔV; < 10 % in payload.\n"
            "Ref: Schilling (2009), Townsend / Martin-Marietta (1962)."
        )

        self._slv_set_text(body, verdict=r['can_reach_orbit'])
        self._right_nb.select(3)

    # ------------------------------------------------------------------
    # Plot panel  (6-subplot grid; slot [2,1] reserved for future use)
    # ------------------------------------------------------------------
    def _build_plot_panel(self, parent):
        self._fig = Figure(figsize=(8, 8.5), dpi=110)
        gs = self._fig.add_gridspec(3, 2, hspace=0.52, wspace=0.38,
                                    left=0.10, right=0.95,
                                    top=0.95, bottom=0.06)
        self._ax_alt  = self._fig.add_subplot(gs[0, 0])  # alt vs time
        self._ax_spd       = self._fig.add_subplot(gs[0, 1])  # speed vs time
        self._ax_spd_twin  = self._ax_spd.twinx()             # Mach axis
        self._ax_traj = self._fig.add_subplot(gs[1, 0])  # alt vs range
        self._ax_trk  = self._fig.add_subplot(gs[1, 1])  # ground track
        self._ax_guid      = self._fig.add_subplot(gs[2, 0])  # pitch / azimuth
        self._ax_guid_twin = self._ax_guid.twinx()            # azimuth axis (created once)
        self._ax_qmach      = self._fig.add_subplot(gs[2, 1]) # q + Mach (burn period)
        self._ax_qmach_twin = self._ax_qmach.twinx()          # Mach axis

        self._canvas = FigureCanvasTkAgg(self._fig, master=parent)
        self._canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        # The standard matplotlib toolbar is kept around for its underlying
        # actions (home / pan / zoom / save) but hidden — the Plots menu
        # drives those same actions, which keeps the chrome out of the way.
        # pack_toolbar=False (mpl ≥ 3.3) builds a headless toolbar with
        # clean event wiring; pack_forget() after construction is fragile.
        try:
            self._plot_toolbar = NavigationToolbar2Tk(
                self._canvas, parent, pack_toolbar=False)
        except TypeError:
            self._plot_toolbar = NavigationToolbar2Tk(self._canvas, parent)
            self._plot_toolbar.pack_forget()
        self._plot_toolbar.update()

        # Initialise axes with placeholder labels
        self._init_axes()
        self._canvas.draw()

    def _init_axes(self):
        for ax, title, xl, yl in [
            (self._ax_alt,   "Altitude vs Time",       "Time (s)",       "Altitude (km)"),
            (self._ax_spd,   "Speed vs Time",           "Time (s)",       "Speed (km/s)"),
            (self._ax_traj,  "Altitude vs Range",       "Downrange (km)", "Altitude (km)"),
            (self._ax_trk,   "Ground Track",            "Longitude (°E)", "Latitude (°N)"),
            (self._ax_guid,  "Pitch, Azimuth vs. Time",          "Time (s)", "Elevation (°)"),
            (self._ax_qmach, "Dyn. Pressure, Mach vs. Time",     "Time (s)", "q  (kPa)"),
        ]:
            ax.set_title(title, fontsize=9)
            ax.set_xlabel(xl, fontsize=8)
            ax.set_ylabel(yl, fontsize=8)
            ax.grid(True, alpha=0.35)
            ax.tick_params(labelsize=7)
        self._ax_spd_twin.set_ylabel('Mach', fontsize=8, color='steelblue')
        self._ax_spd_twin.tick_params(labelsize=7, colors='steelblue')
        self._ax_guid_twin.set_ylabel('Azimuth (°)', fontsize=7, color='darkorange')
        self._ax_guid_twin.tick_params(labelsize=7, colors='darkorange')
        self._ax_qmach_twin.set_ylabel('Mach', fontsize=7, color='darkorange')
        self._ax_qmach_twin.tick_params(labelsize=7, colors='darkorange')

    # ------------------------------------------------------------------
    # Flight Timeline panel
    # ------------------------------------------------------------------
    _TL_COLS = [
        ("event",       "Event",             180, tk.W),
        ("t_s",         "Time (s)",           72, tk.E),
        ("alt_km",      "Alt (km)",            72, tk.E),
        ("range_km",    "Range (km)",          80, tk.E),
        ("gnd_speed",   "Gnd Spd (km/s)",      90, tk.E),
        ("inrtl_speed", "Inrtl Spd (km/s)",    95, tk.E),
        ("accel",       "Accel (m/s²)",         80, tk.E),
        ("mass",        "Mass (t)",             72, tk.E),
    ]

    def _build_timeline_panel(self, parent):
        # Summary block (mirrors left-panel results, visible without switching back)
        sf = ttk.LabelFrame(parent, text="Summary")
        sf.pack(fill=tk.X, padx=6, pady=(6, 2))
        self._tl_summary_var = tk.StringVar(
            value="Run a simulation to populate the flight timeline.")
        ttk.Label(sf, textvariable=self._tl_summary_var,
                  justify=tk.LEFT, anchor=tk.W).pack(
            fill=tk.X, padx=8, pady=4)

        # Timeline table
        tf = ttk.LabelFrame(parent, text="Flight Event Timeline")
        tf.pack(fill=tk.BOTH, expand=True, padx=6, pady=(2, 6))

        col_ids = [c[0] for c in self._TL_COLS]
        self._tl_tree = ttk.Treeview(tf, columns=col_ids, show="headings",
                                     height=14)
        for col_id, heading, width, anchor in self._TL_COLS:
            self._tl_tree.heading(col_id, text=heading)
            self._tl_tree.column(col_id, width=width, anchor=anchor,
                                 stretch=(col_id == "event"))

        vsb = ttk.Scrollbar(tf, orient=tk.VERTICAL,
                            command=self._tl_tree.yview)
        self._tl_tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self._tl_tree.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        # Alternating row colours; no explicit font — inherits system default
        self._tl_tree.tag_configure("odd",    background="#f5f5f5")
        self._tl_tree.tag_configure("even",   background="#ffffff")
        # NOTE: no per-tag font here — font="bold" is parsed by Tk as a font
        # NAMED "bold" (a fallback face at a different size), which rendered
        # the key rows smaller than the rest.  The tint alone marks them.
        self._tl_tree.tag_configure("key",    background="#ddeeff")
        self._tl_tree.tag_configure("debris", background="#fff3cd")

        # Logo — overlaid on the treeview using place() so it sits inside
        # the white area, bottom-left, regardless of platform theme.
        import os as _os
        _logo_path = _os.path.join(_os.path.dirname(__file__), "data", "Thrusty3.png")
        try:
            from PIL import Image, ImageTk as _ITk
            _img = Image.open(_logo_path)
            _h = 195
            _w = int(_img.width * _h / _img.height)
            _img = _img.resize((_w, _h), Image.LANCZOS)
            self._tl_logo_photo = _ITk.PhotoImage(_img)
            tk.Label(self._tl_tree, image=self._tl_logo_photo,
                     borderwidth=0, highlightthickness=0, bg="white"
                     ).place(relx=0.0, rely=1.0, anchor="sw", x=6, y=-4)
        except Exception:
            pass   # logo absent or Pillow unavailable — silent skip

    # ------------------------------------------------------------------
    # Booster Parameters tab
    # ------------------------------------------------------------------
    def _build_params_tab(self, parent):
        """Scrollable structured display — rebuilt on each booster change."""
        self._params_canvas = tk.Canvas(
            parent, borderwidth=0, highlightthickness=0)
        vsb = ttk.Scrollbar(parent, orient="vertical",
                            command=self._params_canvas.yview)
        self._params_canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self._params_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self._params_inner = ttk.Frame(self._params_canvas)
        self._params_win_id = self._params_canvas.create_window(
            (0, 0), window=self._params_inner, anchor="nw")

        self._params_inner.bind(
            "<Configure>",
            lambda _e: self._params_canvas.configure(
                scrollregion=self._params_canvas.bbox("all")))
        self._params_canvas.bind(
            "<Configure>",
            lambda e: self._params_canvas.itemconfig(
                self._params_win_id, width=e.width))

        def _mw(event):
            self._params_canvas.yview_scroll(
                int(-1 * (event.delta / 120)), "units")

        self._params_canvas.bind("<Enter>",
            lambda _e: self._params_canvas.bind_all("<MouseWheel>", _mw))
        self._params_canvas.bind("<Leave>",
            lambda _e: self._params_canvas.unbind_all("<MouseWheel>"))

    def _build_schematic_tab(self, parent):
        """To-scale side elevation of the stack as it will fly (see
        booster_schematic.py).  Redrawn by _update_params_display, so it
        tracks every booster change, edit, and loadout choice live."""
        self._schem_fig = Figure(figsize=(4.5, 8.5), dpi=110)
        self._schem_ax = self._schem_fig.add_subplot(111)
        self._schem_fig.subplots_adjust(left=0.02, right=0.98,
                                        top=0.94, bottom=0.03)
        self._schem_canvas = FigureCanvasTkAgg(self._schem_fig, master=parent)
        self._schem_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def _update_schematic(self, p, name):
        # Cache the composed as-it-will-fly stack for the Blender export —
        # the export ships exactly what the Schematic tab shows.
        self._schem_params, self._schem_name = p, name
        if not hasattr(self, '_schem_ax'):
            return          # sidebar init may fire before the tab exists
        import booster_schematic as bsch
        info = bsch.draw_booster(self._schem_ax, p, title=name)
        self._schem_ax.set_title(
            f"{name}   —   {info['total_height_m']:.1f} m",
            fontsize=11, weight="bold")
        self._schem_canvas.draw_idle()

    def _export_blender(self):
        """File → Export Schematic to Blender…: write the composed
        as-it-will-fly stack (exactly what the Schematic tab shows) as
        discrete, named 3-D objects — true cylinders/frustums for stages
        and interstages, real revolved nose and fairing profiles, plates
        for fins, the RO beside the stack.  Default is a Wavefront .obj
        that Blender IMPORTS directly (File → Import → Wavefront); a .py
        Blender script is offered as the alternative extension (run it in
        the Scripting tab), for anyone who wants editable primitives."""
        from tkinter import filedialog
        p = getattr(self, "_schem_params", None)
        name = getattr(self, "_schem_name", "") or "vehicle"
        if p is None:
            messagebox.showwarning(
                "Export Schematic to Blender",
                "No vehicle composed yet — select a booster first.",
                parent=self)
            return
        path = filedialog.asksaveasfilename(
            title="Export 3-D model for Blender",
            defaultextension=".obj",
            initialdir=str(_ensure_dir(_DIR_BLENDER)),
            initialfile=f"{name.replace(' ', '_')}.obj",
            filetypes=[("Wavefront OBJ (File → Import in Blender)", "*.obj"),
                       ("Blender Python script (run in Scripting tab)",
                        "*.py"),
                       ("All files", "*.*")], parent=self)
        if not path:
            return
        import blender_export as bx
        import os as _os
        if path.lower().endswith(".py"):
            text, info = bx.bpy_script(p, title=name)
            with open(path, "w") as f:
                f.write(text)
            tex = bx.copy_figure_texture(_os.path.dirname(path))
            sidecars = [tex] if tex else []
            how = "In Blender: Scripting tab → Open → Run Script."
        else:
            info = bx.write_obj_bundle(path, p, title=name)
            sidecars = info["files"][1:]
            how = ("In Blender: File → Import → Wavefront (.obj) — the "
                   "DEFAULT import settings work;\nthe scene arrives as "
                   "composed (missile lying, Thrusty standing behind, "
                   "RO lying beyond the nose).\n"
                   "Import at scale 1.0; the model is already in metres.")
        msg = (f"Wrote {info['n_objects']} objects "
               f"({info['total_height_m']:.1f} m stack) to:\n{path}\n\n"
               f"{how}\nUnits are metres, +Z up; every element is a "
               "separate named object.")
        if sidecars:
            msg += ("\n\nSidecar written beside it (keep them together — "
                    "it colours the 3-D Thrusty figure):\n• "
                    + "\n• ".join(_os.path.basename(s) for s in sidecars))
        if info["flags"]:
            msg += ("\n\nFallbacks for unset data (also in the file "
                    "header):\n• " + "\n• ".join(info["flags"]))
        messagebox.showinfo("Export 3-D model for Blender", msg, parent=self)

    # ------------------------------------------------------------------
    # Booster selection
    # ------------------------------------------------------------------
    def _on_booster_changed(self, _event=None):
        name = self._booster_var.get()
        if name not in BOOSTER_DB:
            return
        self._last_valid_booster = name
        # Refresh the Flight Plan combobox for this booster (the plan set and
        # active selection are per booster) before reading the applied booster.
        if hasattr(self, '_fp_cb'):
            self._refresh_flight_plan_list()
        p = get_booster(name)

        # The panel populates from the ACTIVE flight plan (single store).
        # Scalars come from the applied booster `p`; GUI-only run-args (global
        # turn window, cutoff, advanced-pitch toggle, yaw program) come from the
        # raw plan file, with the legacy trajectory_profiles store as a one-time
        # migration fallback.
        raw = self._raw_active_plan(name)
        legacy = _load_traj_profiles().get(name, {})

        def _gk(key, default=None):
            if key in raw:
                return raw[key]
            if key in legacy:
                return legacy[key]
            return default

        self._guidance_var.set(p.guidance)
        self._loft_angle_var.set(f"{p.burnout_angle_deg:.4g}")   # 38.1, not 38.1000
        self._launch_el_var.set(f"{p.launch_elevation_deg:.1f}")
        # Turn start defaults to 0 s (from liftoff); a saved plan value wins.
        gt_start = _gk('gt_turn_start_s', 0.0)
        self._gt_turn_start_var.set(str(gt_start) if gt_start is not None else "0")
        # Turn stop autopopulates with the booster burn time so the field is
        # never blank, but until the user changes or saves it we keep it flagged
        # as an auto-default so Max Range still treats it as "optimize me".
        # Orbital insertion is exempt: there a blank turn-stop makes the boost
        # pitch end just before final-stage ignition (not at full burn), so
        # autopopulating burn time would alter the pitch program -- leave blank.
        gt_stop = _gk('gt_turn_stop_s')
        if gt_stop is not None:
            self._gt_turn_stop_var.set(str(gt_stop))
            self._gt_turn_stop_auto = None
        elif p.guidance == "orbital_insertion":
            self._gt_turn_stop_var.set("")
            self._gt_turn_stop_auto = None
        else:
            _auto = f"{total_burn_time(p):g}"
            self._gt_turn_stop_var.set(_auto)
            self._gt_turn_stop_auto = _auto
        cutoff = _gk('cutoff_time_s')
        self._cutoff_var.set(str(int(cutoff)) if cutoff is not None
                             else str(int(total_burn_time(p))))
        _orb_alt = _gk('target_orbit_km')
        if hasattr(self, '_orbit_alt_var'):
            self._orbit_alt_var.set(f"{float(_orb_alt):g}" if _orb_alt else "400")

        # Fairing jettison altitude (from the applied plan); <=0 = heating
        # default.  Kept in a hidden StringVar so _get_inputs applies it; it is
        # edited in the Flight Plan dialog, not the sidebar.
        try:
            _sj = float(p.shroud_jettison_alt_km)
        except (TypeError, ValueError):
            _sj = 0.0
        self._shroud_jett_var.set(f"{_sj:g}" if _sj > 0 else "")

        self._update_guidance_labels(self._guidance_var.get())
        self._update_params_display(p)
        self._del_btn.config(state=tk.NORMAL)

        # Advanced-pitch toggle + per-stage rows.  The toggle is a GUI key;
        # default it to "on" when the plan carries any per-stage angle override.
        _pstages = []
        _n = p
        while _n is not None:
            _pstages.append(_n)
            _n = getattr(_n, 'stage2', None)
        _adv_default = any(getattr(s, 'stage_burnout_angle_deg', None) is not None
                           for s in _pstages)
        self._adv_pitch_var.set(bool(_gk('adv_pitch_on', _adv_default)))
        self._on_adv_pitch_toggled()
        if self._adv_pitch_var.get() and getattr(self, '_stage_rows', None):
            # Fill the per-stage rows from the applied booster's stage chain.
            for row, s in zip(self._stage_rows, _pstages):
                def _s(v):
                    return "" if v is None else f"{float(v):g}"
                row['start'].set(_s(getattr(s, 'stage_turn_start_s', None)))
                row['stop'].set(_s(getattr(s, 'stage_turn_stop_s', None)))
                row['angle'].set(_s(getattr(s, 'stage_burnout_angle_deg', None)))
                row['coast'].set(_s(getattr(s, 'coast_time_s', None)))
                if not row.get('solid') and 'cutoff' in row:
                    row['cutoff'].set(_s(getattr(s, 'stage_cutoff_s', None)))

        # Yaw / dogleg program.  Prefer the plan's yaw_maneuvers list; fall back
        # to the legacy yaw_rows (strings) for migration.
        _yaw = _gk('yaw_maneuvers')
        if _yaw is None and 'yaw_rows' in legacy:
            _yaw = [[m.get('start'), m.get('stop'), m.get('final_az')]
                    for m in (legacy.get('yaw_rows') or [])]
        _yaw = list(_yaw) if _yaw else []
        # Surface any orphaned PER-STAGE yaw (stage_yaw_*) into the global grid.
        # Older builds baked a per-stage dogleg onto the booster; _yaw_program
        # applies it even with the grid empty, so it flew invisibly.  Absorb it
        # here (start/stop default to the stage's ignition time) so the user can
        # see and edit it; _get_inputs / snapshot then clear the per-stage
        # fields so the visible grid is the sole yaw authority.
        _ign_t = 0.0
        for _st in _pstages:
            _fa = getattr(_st, 'stage_yaw_final_az_deg', None)
            if _fa is not None:
                _ys = getattr(_st, 'stage_yaw_start_s', None)
                _ye = getattr(_st, 'stage_yaw_stop_s', None)
                _yaw.append([_ys if _ys is not None else _ign_t,
                             _ye if _ye is not None else _ign_t, _fa])
            _ign_t += (getattr(_st, 'burn_time_s', 0.0)
                       + getattr(_st, 'coast_time_s', 0.0))
        self._adv_yaw_var.set(bool(_gk('adv_yaw_on', bool(_yaw))) or bool(_yaw))
        self._on_adv_yaw_toggled()
        if self._adv_yaw_var.get() and _yaw and getattr(self, '_yaw_vars', None):
            for yv, man in zip(self._yaw_vars, _yaw):
                def _m(v):
                    return "" if v in (None, "") else str(v)
                yv['start'].set(_m(man[0] if len(man) > 0 else ""))
                yv['stop'].set(_m(man[1] if len(man) > 1 else ""))
                yv['final_az'].set(_m(man[2] if len(man) > 2 else ""))
        # Sustained-α envelope (SP-8099) stored with the yaw program.  Default
        # 10° when the plan has never set it (key absent → _gk returns the 10.0
        # default); a plan that stored None (user cleared it) keeps blank.
        if getattr(self, '_alpha_limit_var', None):
            _al = _gk('alpha_limit_deg', 10.0)
            self._alpha_limit_var.set("" if _al in (None, "") else f"{float(_al):g}")
        # α-induced boost drag toggle stored with the yaw program.
        if getattr(self, '_alpha_induced_var', None):
            self._alpha_induced_var.set(bool(_gk('alpha_induced_drag', False)))

        # Seed glider mission-control fields from the RV if it has glider
        # enabled.  Vehicle properties (L/D, g-limit, βₛ, separation_mode)
        # belong to the RV editor — they're not displayed here; the status
        # line summarises them.  When the booster loads with an RV that has
        # glider_enabled, populate self._ro so the status line and the
        # mission-control frame appear automatically.
        _p_ero = effective_ro(p)
        # If the booster carries a named RV that we have in the library, point
        # the main-panel combobox at it.  Otherwise leave the user's current
        # selection (or the sentinel) alone — they may want the same RV across
        # different boosters.
        if (hasattr(self, '_ro_main_cb')
                and bool(getattr(p, 'body_reenters', False))
                and not self._ro_is_body_compatible(self._ro_main_var.get())):
            # Non-separating vehicle carrying an incompatible (separating)
            # object: fall back to the booster default BEFORE the active object
            # is read.  A separating HGV on a body drew a body AND a corner RV
            # at once — a category error.  Body-mode library objects (the
            # booster-default front end, saved) are compatible and kept.
            self._ro_main_var.set(self._RO_DEFAULT_SENTINEL)
            self._ro_del_btn.config(state=tk.DISABLED)
        elif (hasattr(self, '_ro_main_cb') and _p_ero is not None
                and _p_ero.name and _p_ero.name in RO_DB
                and self._ro_main_var.get() == self._RO_DEFAULT_SENTINEL):
            self._ro_main_var.set(_p_ero.name)
            self._ro_del_btn.config(state=tk.NORMAL)
        # Refresh the reentry-plan dropdown for the now-current object, then take
        # the variant-applied object as the single source of truth for the
        # glider panel, the run, and the write-through (mirror of get_booster
        # applying the active flight-plan variant).
        self._refresh_reentry_plan_list()
        _rp_name, _rp_ero = self._active_reentry_object()
        if _rp_ero is not None:
            _p_ero = _rp_ero
        if _p_ero is not None:
            self._populate_glider_panel(_p_ero)
        if hasattr(self, '_glider_status_var'):
            self._refresh_glider_status_line()
            self._on_main_bank_toggled()
            self._on_main_dive_target_toggled()
            self._on_glider_guidance_changed()

        self._sync_ascent_mode_display(
            self._guidance_var.get(), self._adv_pitch_var.get())

    # ------------------------------------------------------------------
    # Advanced per-stage pitch program
    # ------------------------------------------------------------------
    def _on_adv_pitch_toggled(self):
        """Show or hide the per-stage pitch rows; hide basic rows when active."""
        _basic = [
            self._launch_el_lbl,    self._launch_el_frame,
            self._loft_angle_lbl,   self._loft_angle_frame,
            self._gt_turn_start_lbl, self._gt_turn_start_frame,
            self._gt_turn_stop_lbl,  self._gt_turn_stop_frame,
        ]
        if self._adv_pitch_var.get():
            for w in _basic:
                w.grid_remove()
            self._rebuild_stage_rows()
            self._adv_pitch_frame.grid(row=8, column=0, columnspan=2,
                                        sticky=tk.EW, padx=0, pady=(0, 4))
        else:
            self._adv_pitch_frame.grid_forget()
            for w in _basic:
                w.grid()
            # Re-apply the quick-strip layout: the blanket restore above
            # resurrects Turn Start at its build row (below Turn Stop, which
            # the quick strip moved up) — the source of the inverted
            # Stop-above-Start ordering.  Let the mode logic re-hide/place.
            self._update_guidance_labels(self._guidance_var.get())

    def _on_adv_yaw_toggled(self):
        """Yaw / dogleg now lives in the Flight Plan dialog; keep its sidebar
        frame hidden regardless of the (hidden) checkbox state.  The yaw
        StringVars remain live so _get_inputs still applies the program."""
        self._adv_yaw_frame.grid_remove()

    def _estimate_main_zeta(self):
        """Open the damping estimator, writing its result into the strip's ζ
        field (its default target is self._main_zeta_var)."""
        DampingEstimatorDialog(self)

    def _update_loadout_state(self):
        """Loadout N is a separating-run concept: a non-separating (body)
        vehicle IS its single warhead, so body mode pins N = 1 and greys
        the spinbox.  compose_loadout enforces the same rule model-side.

        The reentry-object selector is scoped the same way: for a body run the
        dropdown offers only body-compatible entries (the booster default plus
        any body-mode library objects), and a separating object left selected
        from a previous booster falls back to the booster default.  Otherwise
        the schematic draws a body AND a corner RV, and the run flies a front
        end the airframe does not have."""
        if not hasattr(self, '_loadout_spin'):
            return
        body = (getattr(self, '_main_sep_var', None) is not None
                and self._main_sep_var.get() == self._SEP_LABELS['body'])
        if body:
            self._loadout_n_var.set("1")
            self._loadout_spin.config(state="disabled")
        else:
            self._loadout_spin.config(state="normal")
        # Re-scope the reentry-object dropdown to the current separation.
        if hasattr(self, '_ro_main_cb'):
            _sel = self._ro_main_var.get()
            self._ro_main_cb.configure(values=self._ro_combo_values())
            if (body and not self._ro_is_body_compatible(_sel)
                    and not getattr(self, '_ro_rescoping', False)):
                # Drop the incompatible object and re-sync the panel from the
                # booster default.  The guard stops the re-sync from recursing
                # back into here (_on_ro_selected_main -> _populate_glider_panel
                # -> _update_loadout_state); one pass is enough because the
                # sentinel it lands on is always compatible.
                self._ro_rescoping = True
                try:
                    self._ro_main_var.set(self._RO_DEFAULT_SENTINEL)
                    self._ro_del_btn.config(state=tk.DISABLED)
                    self._on_ro_selected_main()
                finally:
                    self._ro_rescoping = False

    def _refresh_glider_status_line(self):
        """Update the Glider/HGV status label.  The reentry-mode combobox is
        always shown; this line summarises the terminal vehicle's properties."""
        ro = getattr(self, '_ro', None)
        if ro and ro.glider_enabled and ro.glider_LD > 0:
            # Just the glider numbers: the object's NAME is the Reentry Object
            # dropdown above and separation is the Separation control — this
            # line only carries what is shown nowhere else in the strip.
            self._glider_status_var.set(
                f"L/D {ro.glider_LD:.2f} · g-limit "
                f"{ro.glider_pullup_g_max:.0f}   (Edit Reentry Object…)")
        else:
            self._glider_status_var.set(
                "No L/D set — Edit Reentry Object…")
        # Reentry mode combobox is always visible regardless of glider config.
        self._glider_main_frame.grid(row=4, column=0, columnspan=2,
                                      sticky=tk.EW, padx=0, pady=(0, 4))

    def _is_glider_active(self) -> bool:
        """True iff the active terminal vehicle is a maneuvering glider."""
        ro = getattr(self, '_ro', None)
        return bool(ro and ro.glider_enabled and ro.glider_LD > 0)

    def _on_main_bank_toggled(self):
        # Bank schedule lives in the Reentry Plan editor now; its sidebar frame
        # is gone.  Retained (with a guard) because populate/guidance-changed
        # still call it — a harmless no-op when the widget is absent.
        _frm = getattr(self, '_main_bank_frm', None)
        if _frm is None:
            return
        if self._is_glider_active() and self._main_bank_sched_var.get():
            _frm.grid()
        else:
            _frm.grid_remove()

    def _on_main_dive_target_toggled(self):
        # Dive-at-target lives in the Reentry Plan editor now; guarded no-op if
        # the sidebar frame is absent (same as _on_main_bank_toggled).
        _frm = getattr(self, '_main_dive_target_frm', None)
        if _frm is None:
            return
        if self._is_glider_active() and self._main_dive_target_var.get():
            _frm.grid()
        else:
            _frm.grid_remove()

    def _scope_mode_choices(self):
        """Scope the strip's mode dropdown to the CURRENT mode's integration
        family and refresh the family caption.  This single rule is the Level-2
        identity mechanism (REENTRY_FAMILY_DESIGN.md): the dropdown never
        offers a cross-family law, so the family is fixed for the plan's life
        (New Reentry Plan chooses it); programmatic mode changes (scenario
        restore) legitimately re-scope."""
        _fam = mm.glide_family(self._current_reentry_mode_key())
        _fam_modes = (self._REENTRY_MODE_ANALYTIC if _fam == 'analytic'
                      else self._REENTRY_MODE_NUMERICAL)
        _vals = [lbl for _k, lbl in _fam_modes]
        try:
            if list(self._main_guidance_cb['values']) != _vals:
                self._main_guidance_cb.configure(values=_vals)
        except tk.TclError:
            pass

    def _on_glider_guidance_changed(self):
        raw = self._main_guidance_var.get()
        self._prev_guidance = raw
        self._scope_mode_choices()
        label = raw.lower()
        is_ballistic  = "ballistic"    in label
        key = self._current_reentry_mode_key()
        _uses_zeta = key in ("damped_glide", "dynamic_equilibrium_glide")
        _is_dyn    = key == "dynamic_equilibrium_glide"

        # The ζ row is the only per-mode strip knob: shown for the ζ-using glide
        # laws (damped ratio / dynamic tracking gain), hidden otherwise.  The
        # phugoid-damping estimator applies only to the damped case.
        _zr = getattr(self, '_main_glide_detail_frm', None)
        if _zr is not None:
            if _uses_zeta:
                self._main_zeta_lbl.config(
                    text="Tracking gain ζ:" if _is_dyn else "Damping ratio ζ:")
                # Estimator applies only to phugoid damping (damped case).
                self._main_zeta_est_btn.pack_forget()
                if not _is_dyn:
                    self._main_zeta_est_btn.pack(side=tk.LEFT, padx=(4, 0))
                _zr.grid()
            else:
                _zr.grid_remove()

    def _estimate_body_LD(self):
        """Derive the terminal/glide vehicle's max L/D from geometry, via the
        rigorous whole-booster build-up (glider_ld.py): Jorgensen TR R-474 body
        normal force + Allen-Perkins NACA 1048 viscous crossflow + N-K-P NACA
        1307 wing-body interference, with Jorgensen's sin(2a)/(2a) high-AoA
        correction, maximised over angle of attack.

        For a NO-SEPARATION body the fins are part of the lifting vehicle and are
        included.  This is the value to use as a no-sep glider's L/D; for a
        SEPARATING REENTRY OBJECT, set the reentry object's own designed L/D instead.  Hypersonic regime
        (not Barrowman fin theory, which is for booster static margin)."""
        import glider_ld
        try:
            p = get_booster(self._booster_var.get())
        except Exception:
            return
        # Compose the current reentry object + loadout so the build-up sees the
        # real front-end nose (the RO's, for a body) and the composed mass — the
        # same stack the run derives its L/D from.
        try:
            _name, _uro = self._active_reentry_object()
            if _uro is not None:
                _n = (max(1, int(self._loadout_n_var.get()))
                      if hasattr(self, '_loadout_n_var') else 1)
                p = mm.compose_loadout(p, _uro, _n)
                p.ro = _uro
        except Exception:
            pass
        mach_ref = glider_ld.GLIDE_MACH_REF
        r = glider_ld.whole_booster_LD(p, mach=mach_ref)
        if r.get("error"):
            messagebox.showinfo("L/D Estimate", r["error"])
            return
        fins = (f" + fins {r['c_na_fin']:.2f}" if r["fin_planform_m2"] > 0
                else " (body only)")
        messagebox.showinfo(
            "Whole-body L/D estimate (Jorgensen + Allen-Perkins + N-K-P)",
            f"Mach {mach_ref:.0f}   |   body Cd₀ = {r['cd0']:.3f}\n"
            f"C_Nα (potential) = {r['c_na_pot']:.2f} /rad  "
            f"(body {r['c_na_body']:.2f}{fins})\n"
            f"wing-body factor (1+r/s)² = {r['k_sum']:.2f}\n"
            f"\n"
            f"Max L/D ≈ {r['ld_max']:.2f}   at α ≈ {r['alpha_deg']:.0f}°\n"
            f"\n"
            f"Derived from geometry — the no-sep body L/D.  For a SEPARATING REENTRY OBJECT, "
            f"use the reentry object's own designed L/D instead.",
        )

    def _derived_beta_preview(self, p, ero):
        """Booster-Parameters preview for a non-separating body whose β is left
        at 0 = derive: the geometry-derived β at the reference Mach, its M2–12
        range, and a screening caveat.  None-safe (returns a plain fallback
        string on any failure)."""
        try:
            import glider_ld as _gld
            import math as _math
            d = float(getattr(ero, 'diameter_m', 0.0) or 0.0)
            m = float(getattr(ero, 'mass_kg', 0.0) or 0.0)
            A = _math.pi * (d / 2.0) ** 2
            if A <= 0.0 or m <= 0.0:
                return "derive from geometry"

            def _b(M):
                cd = float(_gld.body_cd0(p, M))
                return m / (cd * A) if cd > 0.0 else 0.0
            b_ref = _b(_gld.GLIDE_MACH_REF)
            b_lo, b_hi = _b(2.0), _b(12.0)
            if b_ref <= 0.0:
                return "derive from geometry"
            return (f"~{b_ref:,.0f} kg/m² (derived, screening; "
                    f"{min(b_lo, b_hi):,.0f}–{max(b_lo, b_hi):,.0f} over M2–12)")
        except Exception:
            return "derive from geometry"

    def _derived_ld_preview(self, p, ero):
        """Booster-Parameters preview for a non-separating body whose L/D is left
        at 0 = derive: the geometry-derived, trim-gated L/D with its static
        margin (or the tumbling verdict).  None-safe."""
        try:
            import glider_ld as _gld
            import trim_gate as _tg
            _cg = float(getattr(ero, 'reentry_cg_m', 0.0) or 0.0)
            g = _tg.trim_gate(p, mach=_gld.GLIDE_MACH_REF,
                              x_cg_m=(_cg if _cg > 0.0 else None))
            if g.get('error'):
                return "derive from geometry"
            _ach = g['LD_achievable']
            _sm = g['static_margin_cal']
            if _ach <= 0.0:
                return f"tumbles — unstable (SM {_sm:+.1f} cal)  (derived)"
            return f"~{_ach:.2f}  (derived; SM {_sm:+.1f} cal)"
        except Exception:
            return "derive from geometry"

    def _rebuild_stage_rows(self):
        """Rebuild inline per-stage pitch rows from the current booster."""
        for w in self._adv_pitch_frame.winfo_children():
            w.destroy()
        self._stage_rows = []

        p = get_booster(self._booster_var.get())
        if p is None:
            return

        # Walk stage chain, record absolute ignition / burnout times
        stages, node, t_ign = [], p, 0.0
        while node is not None:
            t_burn = t_ign + node.burn_time_s
            stages.append({'node': node, 't_ign': t_ign, 't_burn': t_burn})
            t_ign = t_burn + node.coast_time_s
            node = node.stage2

        # Defaults from simple-mode fields
        try:
            g_angle = float(self._loft_angle_var.get())
        except ValueError:
            g_angle = 45.0
        try:
            g_start = float(self._gt_turn_start_var.get())
        except ValueError:
            g_start = 5.0
        try:
            g_stop = float(self._gt_turn_stop_var.get())
        except (ValueError, AttributeError):
            g_stop = None

        _gui_guidance = self._guidance_var.get()
        n_stg = len(stages)

        # Column header
        af = self._adv_pitch_frame
        af.columnconfigure(0, minsize=55)
        # Quick strip: only the two hot-loop columns are shown per stage
        # (Angle, Turn stop).  Turn start / coast / end-burn are set-once fields
        # edited in the Flight Plan dialog; their StringVars still exist here so
        # _get_inputs keeps working, but the entries are not shown.
        _angle_hdr = "η (°)" if _gui_guidance == "true_gravity_turn" else "Angle (°)"
        _headers = [(0, "Stage"), (1, _angle_hdr), (2, "Turn stop (s)"),
                    (3, "Burn window")]
        for col, hdr in _headers:
            ttk.Label(af, text=hdr, foreground="#555555").grid(
                row=0, column=col, padx=(8 if col == 0 else 4, 4),
                pady=(4, 1), sticky=tk.W)

        for i, s in enumerate(stages):
            node = s['node']
            t_i, t_b = s['t_ign'], s['t_burn']
            is_last = (i == n_stg - 1)

            # Seed per-stage values: use stored overrides if present,
            # otherwise derive sensible defaults from simple-mode params.
            if node.stage_turn_start_s is not None:
                def_start = node.stage_turn_start_s
            elif _gui_guidance == "orbital_insertion":
                # All stages share a single global turn-start so the
                # continuous two-phase pitch is reproduced exactly.
                def_start = g_start
            else:
                def_start = g_start if i == 0 else t_i

            if node.stage_turn_stop_s is not None:
                def_stop = node.stage_turn_stop_s
            elif is_last and _gui_guidance == "orbital_insertion":
                # Final stage burns horizontally — stop pitch just before ignition.
                def_stop = max(0.0, t_i - 1.0)
            elif _gui_guidance == "orbital_insertion" and g_stop is not None:
                # Pre-final orbital stages use Plan Orbit's global turn_stop.
                def_stop = g_stop
            else:
                def_stop = max(t_i, t_b - 5.0)

            if node.stage_burnout_angle_deg is not None:
                def_angle = node.stage_burnout_angle_deg
            elif is_last and _gui_guidance == "orbital_insertion":
                def_angle = 0.0   # final stage burns horizontally
            else:
                def_angle = g_angle

            sv_start = tk.StringVar(value=f"{def_start:.1f}")
            sv_stop  = tk.StringVar(value=f"{def_stop:.1f}")
            sv_angle = tk.StringVar(value=f"{def_angle:.1f}")

            # Coast — pre-populate from booster definition; blank for last stage
            sv_coast = tk.StringVar(
                value=f"{node.coast_time_s:.1f}" if not is_last else "")

            # Engine cutoff — commanded burn duration for this stage (blank =
            # full burn).  Liquid engines only; solid motors burn to completion.
            sv_cutoff = tk.StringVar(
                value=f"{node.stage_cutoff_s:.1f}"
                      if node.stage_cutoff_s is not None else "")

            # Solid motors cannot be shut down — no per-stage cutoff.
            if node.solid_motor:
                sv_cutoff.set("")

            row = i + 1
            ttk.Label(af, text=f"Stage {i+1}:").grid(
                row=row, column=0, sticky=tk.W, padx=(8, 4), pady=1)
            # Only Angle and Turn stop are shown (the hot loop); sv_start /
            # sv_coast / sv_cutoff remain live for _get_inputs but are edited in
            # the Flight Plan dialog.
            ttk.Entry(af, textvariable=sv_angle, width=5).grid(
                row=row, column=1, padx=3, pady=1)
            ttk.Entry(af, textvariable=sv_stop,  width=5).grid(
                row=row, column=2, padx=3, pady=1)
            ttk.Label(af, text=f"({t_i:.0f}–{t_b:.0f} s)",
                      foreground="#888888").grid(
                row=row, column=3, sticky=tk.W, padx=(4, 8), pady=1)

            self._stage_rows.append(
                {'start': sv_start, 'stop': sv_stop, 'angle': sv_angle,
                 'coast': sv_coast, 'cutoff': sv_cutoff,
                 'solid': node.solid_motor, 'node': node})

    # ------------------------------------------------------------------
    def _on_site_selected(self, _event=None):
        name = self._site_var.get()
        site = self._site_map.get(name)
        if site is None:          # country-header row clicked — revert
            self._site_var.set("")
            return
        self._launch_lat.set(f"{site['lat']:.4f}")
        self._launch_lon.set(f"{site['lon']:.4f}")
        is_user = name not in _BUNDLED_SITE_NAMES
        self._site_del_btn.config(state=tk.NORMAL if is_user else tk.DISABLED)
        self._on_terrain_toggled()

    def _new_site(self):
        """Clear the site selector and lat/lon fields for fresh entry."""
        self._site_var.set("")
        self._launch_lat.set("")
        self._launch_lon.set("")
        self._site_del_btn.config(state=tk.DISABLED)

    def _edit_site(self):
        """Save current lat/lon as a named user site."""
        lat_str = self._launch_lat.get().strip()
        lon_str = self._launch_lon.get().strip()
        try:
            lat = float(lat_str)
            lon = float(lon_str)
        except ValueError:
            messagebox.showerror("Invalid coordinates",
                                 "Enter valid lat/lon before saving.", parent=self)
            return

        dlg = tk.Toplevel(self)
        dlg.title("Save Site")
        dlg.resizable(False, False)
        dlg.grab_set()
        ttk.Label(dlg, text="Name:").grid(   row=0, column=0, sticky=tk.W, padx=(10,4), pady=(10,2))
        ttk.Label(dlg, text="Country:").grid(row=1, column=0, sticky=tk.W, padx=(10,4), pady=2)
        name_var    = tk.StringVar(value=self._site_var.get()
                                   if self._site_var.get() in self._site_map else "")
        country_var = tk.StringVar(value=self._site_map.get(
                                   self._site_var.get(), {}).get("country", ""))
        ttk.Entry(dlg, textvariable=name_var,    width=28).grid(row=0, column=1, padx=(0,10), pady=(10,2))
        ttk.Entry(dlg, textvariable=country_var, width=28).grid(row=1, column=1, padx=(0,10), pady=2)

        def _do_save():
            name    = name_var.get().strip()
            country = country_var.get().strip()
            if not name or not country:
                messagebox.showerror("Missing fields",
                                     "Name and country are required.", parent=dlg)
                return
            user_sites = _load_user_sites()
            # Update in place if name already exists in user list
            for s in user_sites:
                if s["name"] == name:
                    s.update({"country": country, "lat": lat, "lon": lon})
                    break
            else:
                user_sites.append({"name": name, "country": country,
                                   "lat": lat, "lon": lon})
            _save_user_sites(user_sites)
            self._site_map, cb_values = {}, []
            new_values, new_map = _load_launch_sites()
            self._site_map = new_map
            self._site_cb.config(values=new_values)
            self._site_var.set(name)
            is_user = name not in _BUNDLED_SITE_NAMES
            self._site_del_btn.config(state=tk.NORMAL if is_user else tk.DISABLED)
            self._status_var.set(f"Site '{name}' saved.")
            dlg.destroy()

        bf = ttk.Frame(dlg)
        bf.grid(row=2, column=0, columnspan=2, pady=(6, 10))
        ttk.Button(bf, text="Save",   command=_do_save).pack(side=tk.LEFT, padx=6)
        ttk.Button(bf, text="Cancel", command=dlg.destroy).pack(side=tk.LEFT, padx=6)
        dlg.bind("<Return>", lambda _e: _do_save())

    def _delete_site(self):
        name = self._site_var.get()
        if name in _BUNDLED_SITE_NAMES:
            return
        if not messagebox.askyesno("Delete site",
                                   f"Permanently delete '{name}'?", parent=self):
            return
        user_sites = [s for s in _load_user_sites() if s["name"] != name]
        _save_user_sites(user_sites)
        new_values, new_map = _load_launch_sites()
        self._site_map = new_map
        self._site_cb.config(values=new_values)
        self._site_var.set("")
        self._site_del_btn.config(state=tk.DISABLED)
        self._status_var.set(f"Site '{name}' deleted.")

    def _toggle_query_alt(self):
        # The sidebar Re-entry Query panel moved to Analysis ▸ Re-entry Query…;
        # kept as a guarded no-op for any stale caller.
        _e = getattr(self, '_query_alt_entry', None)
        if _e is not None:
            _e.config(state="normal" if self._query_alt_enable.get()
                      else "disabled")

    def _on_guidance_changed(self):
        """Called when the user selects an ascent mode from the dropdown."""
        display = self._guidance_cb.get()
        if display == "Simple pitch profile":
            # Dropping to Simple discards the per-stage pitch table on the next
            # run/save (per-stage angles fly whenever present, so they must be
            # cleared for the global profile to govern).  Confirm when the
            # table actually holds values.
            if (self._adv_pitch_var.get()
                    and any(r['angle'].get().strip()
                            for r in getattr(self, '_stage_rows', []))
                    and not messagebox.askyesno(
                        "Simple pitch profile",
                        "Switching to Simple clears the per-stage pitch table "
                        "for this plan on the next run or save.\n\nContinue?",
                        parent=self)):
                self._sync_ascent_mode_display(self._guidance_var.get(),
                                               self._adv_pitch_var.get())
                return
            self._guidance_var.set("pitch_program")
            self._adv_pitch_var.set(False)
            self._on_adv_pitch_toggled()
        elif display == "Advanced pitch profile":
            self._guidance_var.set("pitch_program")
            self._adv_pitch_var.set(True)
            self._on_adv_pitch_toggled()
        elif display == "Gravity turn":
            self._guidance_var.set("true_gravity_turn")
        elif display == "Orbital insertion":
            self._guidance_var.set("orbital_insertion")
        self._update_guidance_labels(self._guidance_var.get())

    def _sync_ascent_mode_display(self, guidance: str, adv_pitch: bool):
        """Keep the ascent-mode combobox in sync with backend state.

        The guidance LAW is the flight plan's identity, chosen when the plan is
        created and never changed in place — so the combobox only offers the
        entries valid for the active plan's law: Simple/Advanced for a
        pitch-program plan (same law, different parameterisation), and a single
        fixed entry for gravity-turn / orbital plans.  Switching laws means
        switching (or creating) a flight plan.
        """
        if guidance == "pitch_program":
            display = "Advanced pitch profile" if adv_pitch else "Simple pitch profile"
            choices = ["Simple pitch profile", "Advanced pitch profile"]
        elif guidance == "true_gravity_turn":
            display = "Gravity turn"
            choices = ["Gravity turn"]
        elif guidance == "orbital_insertion":
            display = "Orbital insertion"
            choices = ["Orbital insertion"]
        else:
            display = "Simple pitch profile"
            choices = ["Simple pitch profile", "Advanced pitch profile"]
        if hasattr(self, '_guidance_cb'):
            self._guidance_cb.config(values=choices)
            self._guidance_cb.set(display)

    # ------------------------------------------------------------------
    def _update_guidance_labels(self, guidance: str):
        """Relabel the main-panel guidance fields to match the active mode."""
        if guidance in ("pitch_program", "true_gravity_turn", "orbital_insertion"):
            if guidance == "true_gravity_turn":
                self._loft_angle_lbl.config(text="η kick angle:")
                self._loft_angle_unit_lbl.config(text="°  (+ pitches down)")
            else:
                self._loft_angle_lbl.config(text="Burnout Angle:")
                self._loft_angle_unit_lbl.config(text="°  (Wheelon ε*)")
            # Wheelon ε* estimates the optimal pitch-program burnout angle; in
            # gravity turn the field is the η kick and in orbital the boost
            # angle comes from Plan Orbit — the estimator is meaningless there.
            if hasattr(self, '_wheelon_main_btn'):
                if guidance == "pitch_program":
                    self._wheelon_main_btn.pack(side=tk.LEFT, padx=(4, 0))
                else:
                    self._wheelon_main_btn.pack_forget()
            # Quick strip: burnout angle + turn stop only.  Launch elevation,
            # turn start, coast, and cutoff are set-once fields edited in the
            # Flight Plan dialog; their widgets stay hidden here (grid_remove so
            # their StringVars keep feeding _get_inputs) rather than shown.
            self._gt_turn_start_lbl.grid_remove()
            self._gt_turn_start_frame.grid_remove()
            if not self._adv_pitch_var.get():
                self._loft_angle_lbl.grid(
                    row=2, column=0, sticky=tk.W, padx=(8, 2), pady=2)
                self._loft_angle_frame.grid(
                    row=2, column=1, sticky=tk.W, padx=(0, 8), pady=2)
                self._gt_turn_stop_lbl.grid(
                    row=3, column=0, sticky=tk.W, padx=(8, 2), pady=2)
                self._gt_turn_stop_frame.grid(
                    row=3, column=1, sticky=tk.W, padx=(0, 8), pady=2)
        if guidance == "orbital_insertion":
            self._orbit_alt_lbl.grid(
                row=5, column=0, sticky=tk.W, padx=(8, 2), pady=2)
            self._orbit_alt_frame.grid(
                row=5, column=1, sticky=tk.W, padx=(0, 8), pady=2)
            self._plan_orbit_btn.grid(
                row=6, column=0, columnspan=2,
                sticky=tk.EW, padx=8, pady=(4, 6), ipadx=2, ipady=4)
        else:
            self._orbit_alt_lbl.grid_forget()
            self._orbit_alt_frame.grid_forget()
            self._plan_orbit_btn.grid_forget()

        # Advanced pitch checkbox — shown for gravity_turn/orbital_insertion;
        # for pitch_program the dropdown ("Simple"/"Advanced") handles it.
        if guidance in ("pitch_program", "true_gravity_turn", "orbital_insertion"):
            if guidance == "pitch_program":
                self._adv_pitch_chk.grid_forget()
            else:
                self._adv_pitch_chk.grid(row=7, column=0, columnspan=2,
                                          sticky=tk.W, padx=8, pady=(0, 2))
            if self._adv_pitch_var.get():
                self._adv_pitch_frame.grid(row=8, column=0, columnspan=2,
                                            sticky=tk.EW, padx=0, pady=(0, 4))
            # Yaw / dogleg is a set-once program edited in the Flight Plan
            # dialog; its sidebar checkbox/frame stay hidden.
            self._adv_yaw_chk.grid_remove()
            self._adv_yaw_frame.grid_remove()
        else:
            self._adv_pitch_chk.grid_forget()
            self._adv_pitch_frame.grid_forget()
            self._adv_yaw_chk.grid_remove()
            self._adv_yaw_frame.grid_remove()

        # Reentry Mode frame: orbital insertion has no descent phase.
        if hasattr(self, '_main_guidance_cb'):
            if guidance == "orbital_insertion":
                self._main_guidance_cb.config(state='disabled')
                self._glider_status_var.set(
                    "Not applicable — orbital trajectories have no reentry phase")
            else:
                self._main_guidance_cb.config(state='readonly')
                # Status line is updated by _refresh_glider_status_line when
                # the active terminal vehicle changes; just ensure it's readable.

    # ------------------------------------------------------------------
    # Custom booster management
    # ------------------------------------------------------------------
    def _refresh_booster_list(self, select_name=None):
        """Rebuild the combobox values from the current BOOSTER_DB."""
        names = list(BOOSTER_DB.keys())
        self._booster_cb.configure(values=names)
        target = select_name or self._booster_var.get()
        if target not in BOOSTER_DB:
            target = names[0] if names else ""
        self._booster_var.set(target)
        self._del_btn.config(state=tk.NORMAL if target else tk.DISABLED)
        if target:
            self._on_booster_changed()

    def _on_booster_saved(self, p):
        """Callback invoked by BoosterDialog when the user clicks Save."""
        name = p.name
        BOOSTER_DB[name] = lambda _p=p: _p
        _save_custom_boosters()
        # Deployment timing (fairing / booster-drop / grid-fin) is flight-plan
        # data owned by the Flight Plan editor, so a booster save must NOT write
        # it -- the plan file is its sole source of truth.
        # Snapshot trajectory panel so saving the booster doesn't reset it.
        self._snapshot_traj_profile(name)
        self._refresh_booster_list(select_name=name)
        self._status_var.set(f"Booster '{name}' saved.")

    def _new_booster(self):
        BoosterDialog(self, on_save=self._on_booster_saved)

    def _edit_booster(self):
        name = self._booster_var.get()
        BoosterDialog(self, on_save=self._on_booster_saved, existing_name=name)

    def _delete_booster(self):
        name = self._booster_var.get()
        if not name or name not in BOOSTER_DB:
            return
        if not messagebox.askyesno("Delete booster",
                                   f"Permanently delete '{name}'?",
                                   parent=self):
            return
        del BOOSTER_DB[name]
        _save_custom_boosters()
        profiles = _load_traj_profiles()
        if name in profiles:
            del profiles[name]
            _save_traj_profiles(profiles)
        self._refresh_booster_list()
        self._status_var.set(f"Booster '{name}' deleted.")

    # ------------------------------------------------------------------
    # Reentry vehicle (payload) selection
    # ------------------------------------------------------------------
    _RO_DEFAULT_SENTINEL = "(booster default)"

    def _ro_is_body_compatible(self, name):
        """Can this reentry-object name be the front end of a NON-SEPARATING
        vehicle?  The booster default always can (it IS the body).  A library
        object can only if it is itself body-mode: a separating RV/HGV grafted
        onto a body is a category error — the schematic then draws a body AND a
        corner object, and the run flies a front end the airframe doesn't have.
        """
        if name == self._RO_DEFAULT_SENTINEL or name not in RO_DB:
            return True
        try:
            return (getattr(RO_DB[name](), 'separation_mode',
                            'separating_ro') == 'body')
        except Exception:
            return True          # unreadable entry: don't hide it

    def _run_is_non_separating(self):
        """True when the current run reenters as a body — either the booster is
        marked non-separating (the vehicle-level master) or the reentry plan's
        Separation control says so."""
        try:
            if bool(getattr(get_booster(self._booster_var.get()),
                            'body_reenters', False)):
                return True
        except Exception:
            pass
        return (getattr(self, '_main_sep_var', None) is not None
                and self._main_sep_var.get() == self._SEP_LABELS['body'])

    def _ro_combo_values(self):
        """Combobox values: the sentinel plus every name in RO_DB — filtered to
        the body-compatible entries when the run is non-separating, so a
        separating object cannot be selected onto a body in the first place."""
        names = sorted(RO_DB.keys())
        try:
            if self._run_is_non_separating():
                names = [n for n in names if self._ro_is_body_compatible(n)]
        except Exception:
            pass
        return [self._RO_DEFAULT_SENTINEL] + names

    def _refresh_ro_list(self, select_name=None):
        """Rebuild the RV combobox after a library change."""
        _load_ro_library()
        self._ro_main_cb.configure(values=self._ro_combo_values())
        target = select_name or self._ro_main_var.get()
        if target not in RO_DB and target != self._RO_DEFAULT_SENTINEL:
            target = self._RO_DEFAULT_SENTINEL
        self._ro_main_var.set(target)
        self._ro_del_btn.config(
            state=tk.NORMAL if target in RO_DB else tk.DISABLED)
        self._on_ro_selected_main()

    def _on_ro_selected_main(self, _event=None):
        """Sync self._ro to the selected library entry and refresh the
        glider mission-control panel."""
        sel = self._ro_main_var.get()
        if sel in RO_DB:
            self._ro = RO_DB[sel]()
            self._ro_del_btn.config(state=tk.NORMAL)
            # Any object flies on any booster now: separation is a run-level
            # plan choice (the sidebar Separation control), so the old
            # "separating object needs a separating booster" refusal is gone.
        else:
            self._ro_del_btn.config(state=tk.DISABLED)
            # Fall back to whatever the active booster carries.
            try:
                p = get_booster(self._booster_var.get())
                self._ro = effective_ro(p)
            except Exception:
                self._ro = None
        # The reentry object changed, so its reentry-plan variants and the
        # active selection change too — repopulate the dropdown and the glider
        # controls from the variant-applied object.
        if hasattr(self, '_rp_cb'):
            self._refresh_reentry_plan_list()
            _n, _ero = self._active_reentry_object()
            if _ero is not None:
                self._ro = _ero
                self._populate_glider_panel(_ero)
        if hasattr(self, '_glider_status_var'):
            self._refresh_glider_status_line()
            if hasattr(self, '_main_bank_sched_var'):
                self._on_main_bank_toggled()
            if hasattr(self, '_main_dive_target_var'):
                self._on_main_dive_target_toggled()
            self._on_glider_guidance_changed()
        # Loadout tally + composed launch mass follow the selected object.
        self._update_params_display()

    def _current_booster(self):
        """The current sidebar booster (raw, un-composed), or None — passed to
        the RO editor so its inline body-mode L/D preview can compose the live
        object onto the whole airframe."""
        try:
            return get_booster(self._booster_var.get())
        except Exception:
            return None

    def _current_plan_sep(self):
        """The active reentry plan's separation mode ('body' | 'separating_ro'),
        read back from the sidebar Separation control."""
        lbl = self._main_sep_var.get() if hasattr(self, '_main_sep_var') else ""
        for _k, _v in getattr(self, '_SEP_LABELS', {}).items():
            if _v == lbl:
                return _k
        return 'separating_ro'

    def _seed_booster_default_ro(self):
        """A booster-default reentry object: the booster's front end made into an
        editable object so a non-separating plan has somewhere to enter the nose
        shape / taper / β / L-D.  Geometry is inherited (read-only) from the last
        stage; β and L/D default to 0 = derive."""
        _sep = self._current_plan_sep()
        try:
            bp = get_booster(self._booster_var.get())
        except Exception:
            bp = None
        if _sep == 'body' and bp is not None:
            _last = bp
            while getattr(_last, 'stage2', None) is not None:
                _last = _last.stage2
            return mm.ROParams(
                name=f"{self._booster_var.get()} front end",
                mass_kg=max(float(getattr(_last, 'mass_final', 0.0) or 0.0), 1.0),
                beta_kg_m2=0.0, shape="cone",
                diameter_m=float(getattr(_last, 'diameter_m', 0.0) or 0.5),
                length_m=float(getattr(_last, 'length_m', 0.0) or 2.0),
                separation_mode='body', glider_enabled=True, glider_LD=0.0)
        return mm.ROParams(name="New RV", mass_kg=500.0, beta_kg_m2=10000.0,
                           shape="cone", diameter_m=0.5, length_m=2.0,
                           separation_mode=_sep)

    def _new_ro_main(self):
        """Create an RV in the editor; on Save, write it to the library.  Seeds
        from the booster's front end so a non-separating plan opens straight
        into the body's nose fields."""
        dlg = ROEditorDialog(self, ro=self._seed_booster_default_ro(),
                             mass_kg=500.0, plan_sep=self._current_plan_sep(),
                             booster=self._current_booster())
        self.wait_window(dlg)
        if dlg.result is None:
            return
        try:
            _save_ro_to_library(dlg.result)
        except Exception as exc:
            messagebox.showerror("Save Reentry Object",
                                 f"Could not write reentry-object file:\n{exc}", parent=self)
            return
        self._refresh_ro_list(select_name=dlg.result.name)
        self._status_var.set(f"Reentry object '{dlg.result.name}' saved to library.")

    def _edit_ro_main(self):
        """Edit the currently selected RV in place; rewrite the library file."""
        sel = self._ro_main_var.get()
        if sel not in RO_DB:
            if sel == self._RO_DEFAULT_SENTINEL:
                # No library object selected — seed one from the booster's front
                # end so the user can enter the nose data (the schematic's
                # subtractive nose, β/L-D, TPS) instead of hitting a dead end.
                base = self._seed_booster_default_ro()
            else:
                messagebox.showinfo(
                    "Edit Reentry Object",
                    "Select a reentry object from the library to edit, "
                    "or use 'New' to create one.", parent=self)
                return
        else:
            base = RO_DB[sel]()
        dlg = ROEditorDialog(self, ro=base, mass_kg=base.mass_kg,
                             plan_sep=self._current_plan_sep(),
                             booster=self._current_booster())
        self.wait_window(dlg)
        if dlg.result is None:
            return
        # If the user renamed it, delete the old file so we don't leak orphans.
        if dlg.result.name != base.name:
            _stem = _safe_name(base.name)
            for _ext in (".ro.json", ".ro.json"):   # sweep legacy form too
                try:
                    (_RO_LIBRARY_PATH / f"{_stem}{_ext}").unlink(missing_ok=True)
                    (_LEGACY_RO_LIBRARY_PATH / f"{_stem}{_ext}").unlink(missing_ok=True)
                except Exception:
                    pass
        try:
            _save_ro_to_library(dlg.result)
        except Exception as exc:
            messagebox.showerror("Save Reentry Object",
                                 f"Could not write reentry-object file:\n{exc}", parent=self)
            return
        self._refresh_ro_list(select_name=dlg.result.name)
        self._status_var.set(f"Reentry object '{dlg.result.name}' updated.")

    def _delete_ro_main(self):
        """Remove the selected RV from RO_DB and from ro_library/."""
        sel = self._ro_main_var.get()
        if sel not in RO_DB:
            return
        if not messagebox.askyesno("Delete Reentry Object",
                                   f"Permanently delete '{sel}' from the library?",
                                   parent=self):
            return
        _stem = _safe_name(sel)
        try:
            for _ext in (".ro.json", ".ro.json"):   # sweep legacy form too
                (_RO_LIBRARY_PATH / f"{_stem}{_ext}").unlink(missing_ok=True)
                (_LEGACY_RO_LIBRARY_PATH / f"{_stem}{_ext}").unlink(missing_ok=True)
        except Exception as exc:
            messagebox.showerror("Delete Reentry Object",
                                 f"Could not delete reentry-object file:\n{exc}", parent=self)
            return
        RO_DB.pop(sel, None)
        self._refresh_ro_list(select_name=self._RO_DEFAULT_SENTINEL)
        self._status_var.set(f"Reentry object '{sel}' deleted.")

    def _estimate_wheelon_main(self):
        """Fill the main-panel burnout-angle field with the Wheelon estimate."""
        name = self._booster_var.get()
        if name not in BOOSTER_DB:
            return
        try:
            self._loft_angle_var.set(f"{wheelon_burnout_angle(get_booster(name)):.1f}")
            self._status_var.set("Burnout angle set to the Wheelon ε* estimate.")
        except Exception as exc:
            messagebox.showerror("Estimate ε*", str(exc), parent=self)

    def _active_plan_name(self):
        """Name of the active flight-plan variant, or None for the default."""
        sel = self._fp_var.get() if hasattr(self, '_fp_var') else mm.DEFAULT_PLAN_LABEL
        return None if sel == mm.DEFAULT_PLAN_LABEL else sel

    def _resolve_generator_base(self, active):
        """The plan a generator (Max Range / Plan Orbit) should build on.

        Normally the active plan itself; but when the active plan IS a
        generated variant, rebase onto the plan it was generated from (its
        stored 'base_plan'), so re-running the generator doesn't quietly
        forget the curated base's events/yaw and rebuild from (default).
        """
        if active in (mm.MAX_RANGE_PLAN_LABEL, mm.ORBITAL_PLAN_LABEL):
            return self._raw_active_plan(
                self._booster_var.get(), active).get('base_plan')
        return active

    def _raw_active_plan(self, name, plan_name=None, use_active=True):
        """Raw merged plan-file content for booster `name`: the default plan
        merged with the named (or active) variant.

        This is the ONLY faithful source of the whole plan: GUI-only keys
        (yaw_maneuvers, turn window, cutoff, adv toggles, target orbit) and
        provenance (source/notes) exist solely in the file — extract_flight_plan
        cannot know about them.  Every flow that rebuilds or seeds a plan must
        start from this, not from extract_flight_plan alone, or those keys are
        silently destroyed on the next save.
        """
        raw = mm.load_flight_plan(name, extra_dirs=mm.USER_FLIGHT_PLAN_DIRS) or {}
        pn = plan_name if plan_name is not None else (
            self._active_plan_name() if use_active else None)
        if pn:
            raw = mm._merge_flight_plans(
                raw, mm.load_flight_plan(name, extra_dirs=mm.USER_FLIGHT_PLAN_DIRS,
                                         plan=pn) or {})
        return raw

    @staticmethod
    def _fnum(sv):
        s = sv.get().strip()
        try:
            return float(s) if s else None
        except ValueError:
            return None

    def _snapshot_traj_profile(self, booster_name: str) -> None:
        """Persist the trajectory-panel state into the active flight plan.

        The panel and the flight-plan dialog are two views of ONE store: this
        merges the panel's fields over the current plan (preserving the
        dialog-owned events grid-fin / strap-on) and writes the active plan
        file.  GUI-only run-args with no BoosterParams home — global turn
        window, engine cutoff, the advanced-pitch toggle, and the yaw program —
        ride along as extra keys in the plan file, read back on load.
        """
        # Start from the raw plan file merged over the extraction, so keys the
        # panel doesn't own (source/notes, future additions) survive the
        # rebuild; the panel's fields are then written over the top.
        base = mm._merge_flight_plans(
            extract_flight_plan(get_booster(booster_name)),
            self._raw_active_plan(booster_name))
        base['guidance'] = self._guidance_var.get()
        for key, var in (('launch_elevation_deg', self._launch_el_var),
                         ('burnout_angle_deg', self._loft_angle_var)):
            v = self._fnum(var)
            if v is not None:
                base[key] = v
        _sj = self._shroud_jett_var.get().strip()
        try:
            base['shroud_jettison_alt_km'] = float(_sj) if _sj else 0.0
        except ValueError:
            pass
        # GUI-only run-args (extra keys; apply_flight_plan ignores them).
        base['gt_turn_start_s'] = self._fnum(self._gt_turn_start_var) or 0.0
        base['target_orbit_km'] = self._fnum(getattr(self, '_orbit_alt_var',
                                                     tk.StringVar()))
        # Turn stop autopopulates with the booster burn time; while it is still
        # the untouched auto-default, persist None so it keeps reading as
        # "optimize me" for Max Range rather than locking in the burn time.
        _auto = getattr(self, '_gt_turn_stop_auto', None)
        if _auto is not None and self._gt_turn_stop_var.get().strip() == _auto:
            base['gt_turn_stop_s'] = None
        else:
            base['gt_turn_stop_s'] = self._fnum(self._gt_turn_stop_var)
        base['cutoff_time_s']   = self._fnum(self._cutoff_var)
        adv_pitch = bool(getattr(self, '_adv_pitch_var', tk.BooleanVar()).get())
        base['adv_pitch_on'] = adv_pitch

        stages = base.get('stages', [])
        if adv_pitch and getattr(self, '_stage_rows', None):
            for i, row in enumerate(self._stage_rows):
                if i >= len(stages):
                    break
                stages[i]['stage_turn_start_s']     = self._fnum(row['start'])
                stages[i]['stage_turn_stop_s']      = self._fnum(row['stop'])
                stages[i]['stage_burnout_angle_deg'] = self._fnum(row['angle'])
                _c = self._fnum(row.get('coast', tk.StringVar()))
                if _c is not None:
                    stages[i]['coast_time_s'] = _c
                stages[i]['stage_cutoff_s'] = (None if row.get('solid')
                                               else self._fnum(row.get('cutoff', tk.StringVar())))
        else:
            # Simple mode: clear per-stage angle overrides so the global
            # burnout angle governs the whole ascent.
            for st in stages:
                st['stage_burnout_angle_deg'] = None
        base['stages'] = stages

        adv_yaw = bool(getattr(self, '_adv_yaw_var', tk.BooleanVar()).get())
        base['adv_yaw_on'] = adv_yaw
        yaw = []
        if adv_yaw and getattr(self, '_yaw_vars', None):
            for yv in self._yaw_vars:
                yaw.append([self._fnum(yv['start']), self._fnum(yv['stop']),
                            self._fnum(yv['final_az'])])
        base['yaw_maneuvers'] = yaw
        # α envelope + induced-drag toggle are surfaced from the plan into these
        # vars on load and edited in the Flight Plan dialog; snapshot them back
        # faithfully (independent of the removed sidebar yaw checkbox, and of
        # whether a dogleg is present — an α limit or induced drag can apply to
        # a straight ascent too).  None = no limit.
        base['alpha_limit_deg'] = (self._fnum(self._alpha_limit_var)
                                   if getattr(self, '_alpha_limit_var', None)
                                   else None)
        base['alpha_induced_drag'] = bool(
            getattr(self, '_alpha_induced_var', None)
            and self._alpha_induced_var.get())
        # Yaw is owned by the global program (yaw_maneuvers) above.  Clear any
        # per-stage stage_yaw_* so a legacy baked dogleg can't override the
        # visible grid on the next load/run (it was surfaced into the grid in
        # _on_booster_changed).
        for st in stages:
            st['stage_yaw_start_s'] = None
            st['stage_yaw_stop_s'] = None
            st['stage_yaw_final_az_deg'] = None

        try:
            save_flight_plan(booster_name, base, _FLIGHT_PLAN_LIBRARY_PATH,
                             plan=self._active_plan_name())
        except Exception as exc:
            print(f"Warning: could not save flight plan for '{booster_name}': {exc}")

    # ------------------------------------------------------------------
    # Reentry plan — the down-leg analogue of the flight plan.  The sidebar
    # glider controls are a live view of the active reentry object's
    # .reentryplan.json; every run writes them through, exactly like the
    # flight plan (single-store model).
    # ------------------------------------------------------------------
    def _reentry_plan_kwargs(self) -> dict:
        """Read the sidebar glider controls into the reentry-plan mission-time
        fields.  The SINGLE source of what the run flies AND what is persisted,
        so the reentry object on disk can never disagree with the flight."""
        label = self._main_guidance_var.get().lower()
        key = ("ballistic"                 if "ballistic"    in label else
               "damped_glide"              if "damped"       in label else
               "dynamic_equilibrium_glide" if "dynamic"      in label else
               "skip_glide"                if "skip"         in label else
               "equilibrium_glide_acton"   if "acton"        in label else
               "equilibrium_glide")
        # Separation is a plan field on EVERY path — a ballistic reentry
        # object still either separates at burnout or reenters as the body.
        _sep = ('body'
                if getattr(self, '_main_sep_var', None) is not None
                and self._main_sep_var.get() == self._SEP_LABELS['body']
                else 'separating_ro')
        if key == "ballistic":
            # no lift, regardless of RV config
            return {'glider_enabled': False, 'separation_mode': _sep}
        try:
            dalt = float(self._main_dive_alt_var.get())
        except (ValueError, AttributeError):
            dalt = 0.0                          # blank = glide to impact
        skip = 1                                # (retired skip_to_equilibrium N)
        zeta = 0.7
        if key in ("damped_glide", "dynamic_equilibrium_glide"):
            try:    zeta = max(0.0, float(self._main_zeta_var.get()))
            except (ValueError, AttributeError): zeta = 0.7
        bank = []
        if self._main_bank_sched_var.get():
            for bv in self._main_bank_vars:
                try:
                    bank.append((float(bv['start'].get()), float(bv['end'].get()),
                                 float(bv['bank'].get())))
                except (ValueError, AttributeError):
                    pass
        aero = ("polar" if "polar" in self._main_aero_var.get().lower()
                else "constant_LD")
        dt_lat = dt_lon = dt_rad = 0.0
        if (getattr(self, '_main_dive_target_var', None)
                and self._main_dive_target_var.get()):
            try:    dt_lat = float(self._main_dt_lat_var.get())
            except (ValueError, AttributeError): pass
            try:    dt_lon = float(self._main_dt_lon_var.get())
            except (ValueError, AttributeError): pass
            try:    dt_rad = float(self._main_dt_radius_var.get())
            except (ValueError, AttributeError): pass
        try:
            pull_alt = max(0.0, float(self._main_pullup_alt_var.get()))
        except (ValueError, AttributeError):
            pull_alt = 0.0
        return dict(
            glider_enabled=True, glider_guidance=key, glider_skip_count=skip,
            glider_damping_zeta=zeta, glider_terminal_dive=True,
            glider_terminal_alt_km=dalt, glider_bank_schedule=bank,
            glider_pullup_start_alt_km=pull_alt,
            glider_aero_model=aero, glider_dive_target_lat_deg=dt_lat,
            glider_dive_target_lon_deg=dt_lon, glider_dive_target_radius_km=dt_rad,
            separation_mode=_sep)

    def _populate_glider_panel(self, ro):
        """Fill the sidebar glider controls from reentry object ``ro`` (with its
        reentry plan already applied).  Shared by booster-change and
        reentry-object/plan-change population so all paths agree."""
        self._ro = ro           # _refresh_glider_status_line picks this up
        if hasattr(self, '_main_sep_var'):
            # A booster marked non-separating (body_reenters) is the master: it
            # locks the reentry-plan separation to "body" and greys the combo,
            # so the separation choice lives with the missile.  Otherwise the
            # plan/RO owns it as before.
            _locked = False
            try:
                _b = self._current_booster()
                _locked = bool(getattr(_b, 'body_reenters', False)) if _b else False
            except Exception:
                _locked = False
            if _locked:
                self._main_sep_var.set(self._SEP_LABELS['body'])
            else:
                self._main_sep_var.set(self._SEP_LABELS[
                    'body' if getattr(ro, 'separation_mode',
                                      'separating_ro') == 'body'
                    else 'separating_ro'])
            if hasattr(self, '_main_sep_cb'):
                self._main_sep_cb.configure(
                    state='disabled' if _locked else 'readonly')
            self._update_loadout_state()
        _guid = ro.glider_guidance if ro.glider_enabled else "ballistic"
        # skip_to_equilibrium is retired (aliased to damped_glide on load), so
        # it never reaches here; azimuth_command still maps to skip-glide.
        self._main_guidance_var.set(
            "Phugoid / skip-glide"
            if _guid in ("skip_glide", "azimuth_command")
            else "Damped phugoid glide"
            if _guid in ("damped_glide", "skip_to_equilibrium")
            else "Dynamic equilibrium glide"
            if _guid == "dynamic_equilibrium_glide"
            else "Non-oscillatory glide (Acton)"
            if _guid == "equilibrium_glide_acton"
            else "Equilibrium glide (Tracy)"
            if _guid == "equilibrium_glide"
            else "Ballistic (drag · gravity · rotation)")
        self._main_dive_alt_var.set(f"{ro.glider_terminal_alt_km:.0f}")
        self._main_pullup_alt_var.set(
            f"{getattr(ro, 'glider_pullup_start_alt_km', 0.0):g}")
        _sched = ro.glider_bank_schedule or []
        self._main_bank_sched_var.set(bool(_sched))
        for _i, _bvars in enumerate(self._main_bank_vars):
            if _i < len(_sched):
                _bs, _be, _bk = _sched[_i]
                _bvars['start'].set(f"{_bs:.0f}")
                _bvars['end'].set(f"{_be:.0f}")
                _bvars['bank'].set(f"{_bk:.0f}")
            else:
                _bvars['start'].set('')
                _bvars['end'].set('')
                _bvars['bank'].set('')
        if hasattr(self, '_main_aero_var'):
            self._main_aero_var.set(
                "Drag polar (realistic)"
                if getattr(ro, 'glider_aero_model', 'polar') == 'polar'
                else "Fixed L/D (idealized)")
        if hasattr(self, '_main_dive_target_var'):
            _dt_r = float(getattr(ro, 'glider_dive_target_radius_km', 0.0) or 0.0)
            self._main_dive_target_var.set(_dt_r > 0.0)
            self._main_dt_lat_var.set(
                f"{getattr(ro, 'glider_dive_target_lat_deg', 0.0):.4f}")
            self._main_dt_lon_var.set(
                f"{getattr(ro, 'glider_dive_target_lon_deg', 0.0):.4f}")
            self._main_dt_radius_var.set(
                f"{_dt_r:.0f}" if _dt_r > 0.0 else "20")
        if hasattr(self, '_main_skip_count_var'):
            self._main_skip_count_var.set(str(getattr(ro, 'glider_skip_count', 1)))
        if hasattr(self, '_main_zeta_var'):
            self._main_zeta_var.set(f"{getattr(ro, 'glider_damping_zeta', 0.7):g}")

    def _active_reentry_plan_name(self):
        """Name of the active reentry-plan variant, or None for the default."""
        sel = self._rp_var.get() if hasattr(self, '_rp_var') else mm.DEFAULT_PLAN_LABEL
        return None if sel == mm.DEFAULT_PLAN_LABEL else sel

    def _active_reentry_object(self):
        """(name, plan-applied base ROParams) for the effective reentry object
        given the current sidebar selection, or (None, None) if there is none.

        A library object selected in the sidebar wins; otherwise the booster's
        own reentry object.  The active reentry-plan VARIANT (default + the
        selected variant's diffs) is applied on top, so the base already
        reflects the store — the down-leg analogue of get_booster."""
        sel = self._ro_main_var.get() if hasattr(self, '_ro_main_var') else ""
        if sel in RO_DB:
            ro = RO_DB[sel]()
            name = ro.name or sel
        else:
            try:
                booster = get_booster(self._booster_var.get())
            except Exception:
                return None, None
            ro = effective_ro(booster)
            if ro is None or not ro.name:
                return None, None
            name = ro.name
        rp = load_reentry_plan(name, extra_dirs=mm.USER_REENTRY_PLAN_DIRS,
                               plan=self._active_reentry_plan_name())
        if rp is not None:
            ro = apply_reentry_plan(ro, rp)
        return name, ro

    def _snapshot_reentry_plan(self) -> None:
        """Persist the sidebar glider controls into the active reentry object's
        reentry plan (write-through, mirror of _snapshot_traj_profile)."""
        name, base_ro = self._active_reentry_object()
        if base_ro is None:
            return
        plan = extract_reentry_plan(base_ro)          # full keyset + commanded_LD
        plan.update(self._reentry_plan_kwargs())      # panel mission-time fields
        _pv = self._active_reentry_plan_name()
        try:
            save_reentry_plan(name, plan, _REENTRY_PLAN_LIBRARY_PATH, plan=_pv)
        except Exception as exc:
            print(f"Warning: could not save reentry plan for '{name}': {exc}")
            return
        # Refresh the library entry so re-selecting the object reflects the save
        # (only the DEFAULT plan is baked into RO_DB; a variant is applied on
        # top at read time by _active_reentry_object, so don't bake it here).
        if _pv is None and name in RO_DB:
            RO_DB[name] = lambda _r=apply_reentry_plan(base_ro, plan): _r

    # ── Reentry-plan variant management (mirror of the flight-plan dropdown) ──
    def _save_active_reentry_plans(self):
        try:
            _ACTIVE_REENTRY_PLANS_PATH.parent.mkdir(parents=True, exist_ok=True)
            _ACTIVE_REENTRY_PLANS_PATH.write_text(
                json.dumps(mm.ACTIVE_REENTRY_PLANS, indent=2))
        except Exception as exc:
            print(f"Warning: could not save active reentry plans: {exc}")

    def _refresh_reentry_plan_list(self, select=None):
        """Repopulate the Reentry Plan combobox for the active reentry object."""
        if not hasattr(self, '_rp_cb'):
            return
        name, _ = self._active_reentry_object()
        if name is None:
            self._rp_cb.config(values=[mm.DEFAULT_PLAN_LABEL])
            self._rp_var.set(mm.DEFAULT_PLAN_LABEL)
            self._rp_del_btn.config(state=tk.DISABLED)
            return
        plans = mm.list_reentry_plans(name, extra_dirs=mm.USER_REENTRY_PLAN_DIRS)
        self._rp_cb.config(values=plans)
        active = select or mm.ACTIVE_REENTRY_PLANS.get(name, mm.DEFAULT_PLAN_LABEL)
        if active not in plans:
            active = mm.DEFAULT_PLAN_LABEL
        self._rp_var.set(active)
        self._rp_del_btn.config(
            state=tk.NORMAL if active != mm.DEFAULT_PLAN_LABEL else tk.DISABLED)

    def _on_reentry_plan_selected(self, _event=None):
        name, _ = self._active_reentry_object()
        sel = self._rp_var.get()
        if name is not None:
            if sel == mm.DEFAULT_PLAN_LABEL:
                mm.ACTIVE_REENTRY_PLANS.pop(name, None)
            else:
                mm.ACTIVE_REENTRY_PLANS[name] = sel
            self._save_active_reentry_plans()
        self._rp_del_btn.config(
            state=tk.NORMAL if sel != mm.DEFAULT_PLAN_LABEL else tk.DISABLED)
        self._on_booster_changed()   # repopulate glider controls from the variant

    # Reentry-mode picklist, split exactly like the sidebar strip by HOW the
    # trajectory is integrated: NUMERICAL (EOM, step-by-step lift/drag) vs.
    # CLOSED-FORM ANALYTIC (Tracy/Acton pull-up arc + range formula).  Shared by
    # the New Reentry Plan dialog.  Unlike a flight-plan law (fixed for the
    # plan's life), a reentry mode is the plan's STARTING law and stays
    # switchable on the strip afterward (the hybrid).
    _REENTRY_MODE_NUMERICAL = (
        ("ballistic",                  "Ballistic (drag · gravity · rotation)"),
        ("skip_glide",                 "Phugoid / skip-glide"),
        ("damped_glide",               "Damped phugoid glide"),
        ("dynamic_equilibrium_glide",  "Dynamic equilibrium glide"),
    )
    _REENTRY_MODE_ANALYTIC = (
        ("equilibrium_glide_acton",    "Non-oscillatory glide (Acton)"),
        ("equilibrium_glide",          "Equilibrium glide (Tracy)"),
    )
    _REENTRY_MODE_CHOICES = _REENTRY_MODE_NUMERICAL + _REENTRY_MODE_ANALYTIC

    def _current_reentry_mode_key(self) -> str:
        """The guidance key the strip is currently showing (its default seed)."""
        label = self._main_guidance_var.get().lower()
        return ("ballistic"                 if "ballistic"    in label else
                "damped_glide"              if "damped"       in label else
                "dynamic_equilibrium_glide" if "dynamic"      in label else
                "skip_glide"                if "skip"         in label else
                "equilibrium_glide_acton"   if "acton"        in label else
                "equilibrium_glide")

    def _ask_new_reentry_plan_name_and_mode(self, object_name):
        """Modal prompt for a new reentry plan's name, its INTEGRATION FAMILY
        (the plan's identity — numerical EOM vs closed-form analytic, fixed for
        the plan's life), and its starting law within that family (switchable
        later on the strip, within the family).  Seeded from the object's
        current mode.  Returns (name, mode_label) or (None, None) on cancel."""
        dlg = tk.Toplevel(self)
        dlg.title("New Reentry Plan")
        dlg.resizable(False, False)
        dlg.grab_set()
        frm = ttk.Frame(dlg, padding=12)
        frm.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frm, text=f"Name for the new reentry plan for '{object_name}':"
                  ).grid(row=0, column=0, sticky=tk.W, pady=(0, 4))
        name_var = tk.StringVar()
        ent = ttk.Entry(frm, textvariable=name_var, width=36)
        ent.grid(row=1, column=0, sticky=tk.EW, pady=(0, 10))

        _cur = self._current_reentry_mode_key()
        _cur_fam = mm.glide_family(_cur)
        ttk.Label(frm, text="Integration family (fixed for the life of the plan):"
                  ).grid(row=2, column=0, sticky=tk.W, pady=(0, 4))
        fam_var = tk.StringVar(value=_cur_fam)
        mode_var = tk.StringVar()

        def _fam_modes():
            return (self._REENTRY_MODE_ANALYTIC if fam_var.get() == 'analytic'
                    else self._REENTRY_MODE_NUMERICAL)

        def _on_family(*_):
            _modes = _fam_modes()
            _labels = [lbl for _k, lbl in _modes]
            mode_cb.configure(values=_labels)
            # keep the current mode if it belongs to the family; else default
            # to the family's first (numerical: Ballistic; analytic: Acton).
            _keep = next((lbl for k, lbl in _modes if k == _cur), None)
            mode_var.set(_keep if _keep is not None else _labels[0])

        for _i, (_key, _lbl, _hint) in enumerate((
                ('numerical', "Numerical (EOM)",
                 "step-by-step integration; banking, dive-at-target, "
                 "Mach-varying L/D; honest capture (lofted entries plunge)"),
                ('analytic', "Closed-form analytic",
                 "Tracy/Acton pull-up arc + range formula; constant L/D, "
                 "always captures; fast comparison law"))):
            ttk.Radiobutton(frm, text=_lbl, variable=fam_var, value=_key,
                            command=_on_family).grid(
                row=3 + 2 * _i, column=0, sticky=tk.W, padx=(4, 0))
            ttk.Label(frm, text=_hint, foreground="#888888",
                      wraplength=360, justify=tk.LEFT).grid(
                row=4 + 2 * _i, column=0, sticky=tk.W, padx=(24, 0))

        ttk.Label(frm, text="Starting law (switchable within the family):"
                  ).grid(row=7, column=0, sticky=tk.W, pady=(8, 4))
        mode_cb = ttk.Combobox(frm, textvariable=mode_var,
                               state="readonly", width=34)
        mode_cb.grid(row=8, column=0, sticky=tk.EW, pady=(0, 4))
        _on_family()

        out = {}
        def _ok(*_):
            _m = mode_var.get()
            if not _m:
                return
            out['name'] = name_var.get().strip()
            out['mode'] = _m
            dlg.destroy()
        bf = ttk.Frame(frm)
        bf.grid(row=9, column=0, sticky=tk.E, pady=(12, 0))
        ttk.Button(bf, text="Cancel", command=dlg.destroy).pack(side=tk.RIGHT, padx=(4, 0))
        ttk.Button(bf, text="Create", command=_ok).pack(side=tk.RIGHT)
        ent.bind("<Return>", _ok)
        ent.focus_set()
        self.wait_window(dlg)
        return (out.get('name') or None, out.get('mode'))

    def _new_reentry_plan(self):
        """Create a reentry-plan variant: name + starting mode, seeded from the
        active plan's other fields (write-through)."""
        name, _ = self._active_reentry_object()
        if name is None:
            messagebox.showinfo("Reentry Plan",
                                "Select a maneuvering reentry object first.",
                                parent=self)
            return
        new_name, mode_lbl = self._ask_new_reentry_plan_name_and_mode(name)
        if not new_name:
            return
        if new_name == mm.DEFAULT_PLAN_LABEL:
            messagebox.showerror("Reentry Plan",
                                 f"'{new_name}' is reserved.", parent=self)
            return
        # Stamp the chosen starting mode onto the strip, then write the variant
        # through from the current panel state.  The strip dropdown still
        # switches it afterward — the mode is a coherent seed, not a lock.
        if mode_lbl:
            self._main_guidance_var.set(mode_lbl)
            self._on_glider_guidance_changed()
        mm.ACTIVE_REENTRY_PLANS[name] = new_name
        self._rp_var.set(new_name)
        self._snapshot_reentry_plan()
        self._save_active_reentry_plans()
        self._refresh_reentry_plan_list(select=new_name)
        self._on_booster_changed()
        self._status_var.set(f"Reentry plan '{new_name}' created for '{name}'.")

    def _ro_hardware(self, name):
        """The hardware ROParams for ``name`` (its .ro.json, no reentry plan
        applied) — the source of the TRUE L/D capability, which a plan's
        commanded_LD clamp must never erode.  Falls back to the booster's own
        object (get_booster applies the flight plan, not the reentry plan, so
        its glider_LD is still the capability)."""
        safe = _safe_name(name)
        for d in (_RO_LIBRARY_PATH, _BUNDLED_RO_LIBRARY_PATH, _LEGACY_RO_LIBRARY_PATH):
            fp = d / f"{safe}.ro.json"
            if fp.exists():
                try:
                    return ro_from_dict(json.loads(fp.read_text()))
                except Exception:
                    pass
        try:
            return effective_ro(get_booster(self._booster_var.get()))
        except Exception:
            return None

    def _edit_reentry_plan_main(self):
        """Open the Reentry Plan editor for the active object + variant, then
        write the edited fields through to that plan and repopulate."""
        name, applied = self._active_reentry_object()
        if name is None:
            messagebox.showinfo("Reentry Plan",
                                "Select a maneuvering reentry object first.",
                                parent=self)
            return
        pv = self._active_reentry_plan_name()
        # Snapshot the sidebar strip first so the dialog edits the same plan the
        # panel shows (strip + dialog are two views of one plan file).
        self._snapshot_reentry_plan()
        plan = (mm.load_reentry_plan(name, extra_dirs=mm.USER_REENTRY_PLAN_DIRS, plan=pv)
                or extract_reentry_plan(applied))
        # Capability is the HARDWARE L/D, not the (possibly clamped) applied one.
        _hw = self._ro_hardware(name)
        cap = getattr(_hw, 'glider_LD', getattr(applied, 'glider_LD', 0.0))
        title = name if pv is None else f"{name} — {pv}"
        dlg = ReentryPlanDialog(self, title, plan, cap)
        self.wait_window(dlg)
        if dlg.result is None:
            return
        merged = {**plan, **dlg.result}
        try:
            save_reentry_plan(name, merged, _REENTRY_PLAN_LIBRARY_PATH, plan=pv)
        except Exception as exc:
            messagebox.showerror("Reentry Plan",
                                 f"Could not save reentry plan:\n{exc}", parent=self)
            return
        # Refresh the library baseline when editing the default plan — apply to
        # the HARDWARE object so the stored capability is never eroded.
        if pv is None and name in RO_DB and _hw is not None:
            RO_DB[name] = lambda _r=apply_reentry_plan(_hw, merged): _r
        self._on_booster_changed()
        self._status_var.set(f"Reentry plan for '{title}' saved.")

    def _delete_reentry_plan(self):
        name, _ = self._active_reentry_object()
        sel = self._rp_var.get()
        if name is None or sel == mm.DEFAULT_PLAN_LABEL:
            return
        if not messagebox.askyesno(
                "Delete Reentry Plan",
                f"Delete reentry plan '{sel}' for '{name}'?", parent=self):
            return
        fp = Path(_REENTRY_PLAN_LIBRARY_PATH) / mm.reentry_plan_filename(name, sel)
        try:
            fp.unlink(missing_ok=True)
        except Exception as exc:
            messagebox.showerror("Reentry Plan",
                                 f"Could not delete reentry plan:\n{exc}", parent=self)
            return
        mm.ACTIVE_REENTRY_PLANS.pop(name, None)
        self._save_active_reentry_plans()
        self._refresh_reentry_plan_list(select=mm.DEFAULT_PLAN_LABEL)
        self._on_booster_changed()
        self._status_var.set(f"Reentry plan '{sel}' deleted.")

    def _reset_traj_profile(self) -> None:
        """Revert the active flight plan to the shipped/bundled default by
        removing the user override, then repopulate the panel."""
        name = self._booster_var.get()
        if not name:
            return
        # Drop the user override for the active plan and the legacy profile.
        try:
            (Path(_FLIGHT_PLAN_LIBRARY_PATH)
             / mm.flight_plan_filename(name, self._active_plan_name())).unlink(missing_ok=True)
        except Exception as exc:
            print(f"Warning: could not reset flight plan for '{name}': {exc}")
        profiles = _load_traj_profiles()
        if name in profiles:
            del profiles[name]
            _save_traj_profiles(profiles)
        self._on_booster_changed()
        self._status_var.set(f"Trajectory reset to '{name}' defaults.")

    # Legacy pre-flight-plan guidance snapshot keys (read on import only).
    _GUIDANCE_KEYS = frozenset({
        'guidance', 'burnout_angle_deg', 'launch_elevation_deg',
        'gt_turn_start_s', 'gt_turn_stop_s', 'cutoff_s',
        'adv_pitch', 'stage_overrides', 'adv_yaw', 'yaw_maneuvers',
        'azimuth_deg', 'launch_lat', 'launch_lon',
    })

    def _export_flight_plan_file(self):
        """Save the active flight plan (as shown in the panel) to a JSON file.

        The panel is snapshotted to the flight-plan library first, so the
        exported file is exactly the plan the next Run would fly -- the same
        .flightplan.json format the library uses, portable between machines.
        """
        from tkinter import filedialog
        name = self._booster_var.get()
        if not name:
            messagebox.showinfo("Flight Plan", "Select a booster first.", parent=self)
            return
        self._snapshot_traj_profile(name)
        active = self._active_plan_name()
        path = filedialog.asksaveasfilename(
            title="Save flight plan",
            defaultextension=".json",
            initialdir=str(_ensure_dir(_THRUSTY_ROOT)),
            initialfile=mm.flight_plan_filename(name, active),
            filetypes=[("Flight plan JSON", "*.flightplan.json"),
                       ("JSON files", "*.json"), ("All files", "*.*")],
            parent=self,
        )
        if not path:
            return
        plan = dict(self._raw_active_plan(name, active))
        plan['_type']   = 'flight_plan'
        plan['booster'] = name
        if active:
            plan['name'] = active
        try:
            with open(path, 'w') as fh:
                json.dump(plan, fh, indent=2)
            self._status_var.set(
                f"Flight plan saved: {os.path.basename(path)}")
        except Exception as exc:
            messagebox.showerror("Export error", str(exc), parent=self)

    def _import_flight_plan_file(self):
        """Load a flight plan from a JSON file into the active plan slot.

        Reads both the current .flightplan.json format and the legacy
        .guidance.json snapshots (pre-flight-plan 'Save Guidance…' exports).
        Either way the result is written to the flight-plan library, so the
        import persists like any other plan edit.
        """
        from tkinter import filedialog
        name = self._booster_var.get()
        if not name:
            messagebox.showinfo("Flight Plan", "Select a booster first.", parent=self)
            return
        _legacy_dir = _DIR_GUIDANCE if _DIR_GUIDANCE.is_dir() else _THRUSTY_ROOT
        path = filedialog.askopenfilename(
            title="Load flight plan",
            initialdir=str(_ensure_dir(_legacy_dir)),
            filetypes=[("Flight plan / guidance JSON", "*.json"),
                       ("All files", "*.*")],
            parent=self,
        )
        if not path:
            return
        try:
            with open(path) as fh:
                data = json.load(fh)
            # Import lands as a NEW named variant, never overwriting the active
            # plan — an imported plan carries its own guidance law, and the law
            # is plan identity, so it must arrive as its own artifact.
            _stem = os.path.basename(path).split('.')[0]
            variant = self._unique_plan_name(
                name, str(data.get('name') or _stem or 'imported'))
            if data.get('_type') == 'guidance_program':
                # Legacy snapshot: apply to the panel, switch to the new
                # variant, then persist the panel through write-through.
                self._apply_trajectory_metadata(data)
                mm.ACTIVE_FLIGHT_PLANS[name] = variant
                self._fp_var.set(variant)
                self._snapshot_traj_profile(name)
            else:
                for k in ('_type', 'booster', 'name', 'base_plan'):
                    data.pop(k, None)
                save_flight_plan(name, data, _FLIGHT_PLAN_LIBRARY_PATH, plan=variant)
                mm.ACTIVE_FLIGHT_PLANS[name] = variant
            self._save_active_plans()
            self._refresh_flight_plan_list(select=variant)
            self._on_booster_changed()
            self._status_var.set(
                f"Flight plan imported as '{variant}': {os.path.basename(path)}")
        except Exception as exc:
            messagebox.showerror("Import error", str(exc), parent=self)

    def _unique_plan_name(self, booster_name, desired):
        """A flight-plan variant name derived from `desired` that is neither
        reserved nor already in use (appends ' (2)', ' (3)', … as needed)."""
        desired = (desired or 'imported').strip() or 'imported'
        reserved = set(mm.RESERVED_PLAN_NAMES)
        existing = set(mm.list_flight_plans(
            booster_name, extra_dirs=mm.USER_FLIGHT_PLAN_DIRS))
        if desired not in reserved and desired not in existing:
            return desired
        i = 2
        while f"{desired} ({i})" in existing or f"{desired} ({i})" in reserved:
            i += 1
        return f"{desired} ({i})"

    # ------------------------------------------------------------------
    # Scenario save / load (bundle: booster + RV + site + guidance)
    # ------------------------------------------------------------------
    _CSV_SCENARIO_PREFIX = "# scenario: "
    _XLSX_SCENARIO_SHEET = "Scenario"

    def _scenario_dict(self) -> dict:
        """Return _trajectory_metadata() stamped as a scenario record."""
        data = dict(self._trajectory_metadata())
        data['_type']    = 'scenario'
        data['_version'] = 1
        return data

    @staticmethod
    def _read_scenario_from_csv(path: Path) -> dict:
        """Parse the leading '# scenario: <json>' line from a trajectory CSV."""
        with open(path, encoding="utf-8-sig") as fh:
            first = fh.readline().strip()
        prefix = BoosterFlyoutApp._CSV_SCENARIO_PREFIX.strip()
        if not first.startswith(prefix):
            raise ValueError("CSV does not start with a scenario header line.")
        return json.loads(first[len(prefix):].strip())

    @staticmethod
    def _read_scenario_from_xlsx(path: Path) -> dict:
        """Read the 'Scenario' sheet of a trajectory XLSX into a dict."""
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        sheet = BoosterFlyoutApp._XLSX_SCENARIO_SHEET
        if sheet not in wb.sheetnames:
            raise ValueError(f"XLSX has no '{sheet}' sheet.")
        ws = wb[sheet]
        data: dict = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            k, v = row[0], (row[1] if len(row) > 1 else None)
            if isinstance(v, str) and v[:1] in ("[", "{"):
                try:
                    v = json.loads(v)
                except Exception:
                    pass
            data[k] = v
        return data

    def _save_scenario(self):
        """Save the full input state — booster name, RV name, launch site,
        azimuth, and every guidance / glider field — as a .scenario.json."""
        from tkinter import filedialog
        import datetime as _dt
        ts      = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        booster = _safe_name(self._booster_var.get())
        path = filedialog.asksaveasfilename(
            parent=self,
            title="Save scenario",
            defaultextension=".json",
            initialdir=str(_ensure_dir(_DIR_SCENARIOS)),
            initialfile=f"{ts}_{booster}.scenario.json",
            filetypes=[("Scenario files (*.scenario.json)", "*.json"),
                       ("All files", "*.*")],
        )
        if not path:
            return
        try:
            Path(path).write_text(json.dumps(self._scenario_dict(), indent=2))
        except Exception as exc:
            messagebox.showerror("Save scenario",
                                 f"Could not write file:\n{exc}", parent=self)
            return
        self._status_var.set(f"Scenario saved: {os.path.basename(path)}")

    def _load_scenario(self):
        """Load a scenario from a .scenario.json, a trajectory CSV with an
        embedded scenario header, or a trajectory XLSX with a Scenario sheet,
        and apply it to the GUI."""
        try:
            from tkinter import filedialog
            path = filedialog.askopenfilename(
                parent=self,
                title="Load scenario",
                initialdir=str(_ensure_dir(_DIR_SCENARIOS)),
                filetypes=[
                    ("Scenario / trajectory",
                        "*.json *.csv *.xlsx"),
                    ("Scenario JSON",   "*.json"),
                    ("Trajectory CSV",  "*.csv"),
                    ("Trajectory XLSX", "*.xlsx"),
                    ("All files",       "*.*"),
                ],
            )
        except Exception as exc:
            messagebox.showerror("Load scenario",
                                 f"Could not open file dialog:\n{exc}", parent=self)
            return
        if not path:
            return
        p   = Path(path)
        ext = p.suffix.lower()
        try:
            if ext == '.csv':
                data = self._read_scenario_from_csv(p)
            elif ext == '.xlsx':
                data = self._read_scenario_from_xlsx(p)
            else:
                data = json.loads(p.read_text())
        except Exception as exc:
            messagebox.showerror("Load scenario",
                                 f"Could not read file:\n{exc}", parent=self)
            return
        if not isinstance(data, dict):
            messagebox.showerror("Load scenario",
                                 "File does not contain a scenario object.",
                                 parent=self)
            return
        try:
            self._apply_trajectory_metadata(data)
        except Exception as exc:
            messagebox.showerror("Load scenario",
                                 f"Could not apply scenario:\n{exc}", parent=self)
            return
        _isolated = self._isolate_scenario_law_if_needed()
        _msg = f"Scenario loaded: {os.path.basename(path)}"
        if _isolated:
            _msg += " (guidance isolated into the 'scenario' flight plan)"
        self._status_var.set(_msg)

    def _isolate_scenario_law_if_needed(self):
        """If a just-loaded scenario's guidance law differs from the active
        plan's, retarget to the reserved 'scenario' variant so the subsequent
        write-through can't rewrite the curated plan's law (which is its
        identity).  Returns True if isolation happened."""
        name = self._booster_var.get()
        if name not in BOOSTER_DB:
            return False
        panel_law = self._guidance_var.get()
        active_law = get_booster(name).guidance
        if panel_law == active_law:
            return False
        mm.ACTIVE_FLIGHT_PLANS[name] = mm.SCENARIO_PLAN_LABEL
        self._fp_var.set(mm.SCENARIO_PLAN_LABEL)
        self._snapshot_traj_profile(name)   # persist panel (scenario law) here
        self._save_active_plans()
        self._refresh_flight_plan_list(select=mm.SCENARIO_PLAN_LABEL)
        return True

    def _update_params_display(self, p=None):
        """Rebuild the Booster Parameters tab with structured label rows.

        The tab shows the stack AS IT WILL FLY: the sidebar's reentry object
        and Loadout N are composed onto a display copy (raw library boosters
        carry no object and — for stack-only builds — no payload at all), so
        launch mass, T/W and the throw-weight tally track the run setup."""
        if not hasattr(self, '_params_inner'):
            return          # sidebar init may fire before the tab exists
        if p is None:
            try:
                p = get_booster(self._booster_var.get())
            except (KeyError, ValueError):
                return
        _ro_name = self._ro_main_var.get() if hasattr(self, '_ro_main_var') else ""
        if _ro_name in RO_DB:
            _uro = RO_DB[_ro_name]()
            try:
                _n = max(1, int(self._loadout_n_var.get())) \
                    if hasattr(self, '_loadout_n_var') else 1
            except (ValueError, tk.TclError):
                _n = 1
            p = mm.compose_loadout(p, _uro, _n)
            p.ro = _uro

        # The Schematic tab draws the same composed as-it-will-fly stack.
        self._update_schematic(p, self._booster_var.get())

        _G0 = 9.80665
        pad = dict(padx=8, pady=4)

        # Clear previous content
        for w in self._params_inner.winfo_children():
            w.destroy()

        def _row(frame, row, label, value):
            ttk.Label(frame, text=label).grid(
                row=row, column=0, sticky=tk.W, padx=(6, 2), pady=2)
            ttk.Label(frame, text=value).grid(
                row=row, column=1, sticky=tk.W, padx=(0, 6), pady=2)

        def _row2(frame, row, lab1, val1, lab2='', val2=''):
            """Two label:value pairs on a single row (4-column layout)."""
            ttk.Label(frame, text=lab1).grid(
                row=row, column=0, sticky=tk.W, padx=(6, 2), pady=2)
            ttk.Label(frame, text=val1).grid(
                row=row, column=1, sticky=tk.W, padx=(0, 10), pady=2)
            if lab2:
                ttk.Label(frame, text=lab2).grid(
                    row=row, column=2, sticky=tk.W, padx=(8, 2), pady=2)
                ttk.Label(frame, text=val2).grid(
                    row=row, column=3, sticky=tk.W, padx=(0, 6), pady=2)

        # ── Compute totals used in summary ────────────────────────────
        total_prop = p.mass_propellant
        node = p.stage2
        while node is not None:
            total_prop += node.mass_propellant
            node = node.stage2

        _node_l, _sn_l, _stage_lengths = p, 1, []
        while _node_l is not None:
            if _node_l.length_m > 0:
                _stage_lengths.append((_sn_l, _node_l.length_m))
            _node_l = _node_l.stage2
            _sn_l += 1
        _total_len = sum(l for _, l in _stage_lengths)
        if p.shroud_length_m > 0:
            _total_len += p.shroud_length_m

        liftoff_tw = (p.thrust_N / (p.mass_initial * _G0)
                      if p.mass_initial > 0 else 0.0)

        # ── Summary (4-col, 2 pairs per row) ──────────────────────────
        sf = ttk.LabelFrame(self._params_inner, text="Summary")
        sf.pack(fill=tk.X, **pad)
        sf.columnconfigure(1, weight=1)
        sf.columnconfigure(3, weight=1)

        r = 0
        ttk.Label(sf, text="Name:").grid(
            row=r, column=0, sticky=tk.W, padx=(6, 2), pady=2)
        ttk.Label(sf, text=p.name).grid(
            row=r, column=1, columnspan=3, sticky=tk.W, padx=(0, 6), pady=2)
        r += 1
        _len_str = f"{_total_len:.1f} m" if _total_len > 0 else "—"
        _row2(sf, r, "Launch mass:", f"{p.mass_initial:,.0f} kg",
              "Total length:", _len_str); r += 1
        _row2(sf, r, "Total propellant:", f"{total_prop:,.0f} kg",
              "Liftoff T/W:", f"{liftoff_tw:.2f}"); r += 1
        if p.payload_kg > 0:
            # Throw weight is a computed tally of the composed loadout:
            # bus + N × object mass, carried through boost.
            if p.ro_mass_kg > 0:
                _lo = f"{p.num_ros} × {p.ro_mass_kg:,.0f} kg"
                if p.bus_mass_kg > 0:
                    _lo += f" + PBV {p.bus_mass_kg:,.0f} kg"
                _row2(sf, r, "Throw weight:", f"{p.payload_kg:,.0f} kg",
                      "Loadout:", _lo); r += 1
            else:
                _row2(sf, r, "Throw weight:", f"{p.payload_kg:,.0f} kg"); r += 1

        # ── Per-stage blocks ──────────────────────────────────────────
        sn = 1
        node = p
        while node is not None:
            is_last = node.stage2 is None
            lf = ttk.LabelFrame(self._params_inner,
                                text=f"Stage {sn}" if sn > 1 else "Stage 1")
            lf.pack(fill=tk.X, **pad)

            prop = node.mass_propellant
            tw   = (node.thrust_N / (node.mass_initial * _G0)
                    if node.mass_initial > 0 else 0.0)

            # Recover this stage's own fueled mass (payload excluded).
            # For non-last stages mass_final is the jettisoned dry mass only,
            # so stage_fueled = mass_final + propellant — no stack arithmetic needed.
            # For the last stage (and single-stage) we strip payload (and shroud
            # if single-stage) from mass_initial, which requires payload_kg to be
            # set correctly on the top-level node.
            is_first = (sn == 1)
            if is_last and is_first:
                # Single-stage: mass_initial = fueled + payload + shroud
                stage_fueled = node.mass_initial - p.payload_kg - p.shroud_mass_kg
            elif is_last:
                # Last of multi: mass_initial = fueled + payload
                stage_fueled = node.mass_initial - p.payload_kg
            else:
                # Non-last: mass_final = jettisoned dry mass only
                stage_fueled = node.mass_final + prop
            stage_dry = stage_fueled - prop
            dry_pct   = stage_dry / stage_fueled * 100 if stage_fueled > 0 else 0.0

            # Two-column layout inside the stage LabelFrame.
            # Left: Dimensions & Masses   Right: Engine Performance
            lf.columnconfigure(0, weight=1)
            lf.columnconfigure(2, weight=1)
            left  = ttk.Frame(lf)
            left.grid( row=0, column=0, sticky="nsew", padx=(4, 0), pady=4)
            ttk.Separator(lf, orient="vertical").grid(
                row=0, column=1, sticky="ns", padx=4)
            right = ttk.Frame(lf)
            right.grid(row=0, column=2, sticky="nsew", padx=(0, 4), pady=4)

            # ── Left: Dimensions & Masses ─────────────────────────────
            r = 0
            _row(left, r, "Diameter (m):",      f"{node.diameter_m:.2f}");       r += 1
            _row(left, r, "Length (m):",         f"{node.length_m:.2f}");         r += 1
            _row(left, r, "Fueled mass (kg):",   f"{stage_fueled:,.0f}");         r += 1
            _row(left, r, "Propellant (kg):",    f"{prop:,.0f}  (computed)");     r += 1
            _row(left, r, "Dry mass (kg):",      f"{stage_dry:,.0f}");            r += 1
            _row(left, r, "Dry mass %:",         f"{dry_pct:.1f}%");              r += 1
            if not is_last:
                _row(left, r, "Coast (s):",      f"{node.coast_time_s:.0f}");     r += 1
            # Debris β for jettisoned stage bodies.  Whether the LAST stage
            # body becomes debris is the run-level separation choice (the
            # sidebar Separation control), not a booster property.
            _sep_body = (getattr(self, '_main_sep_var', None) is not None
                         and self._main_sep_var.get()
                         == self._SEP_LABELS['body'])
            _body_jettisoned = (not is_last) or not _sep_body
            if _body_jettisoned:
                if is_last:
                    # Casing = burnout mass minus the departing loadout
                    # (mirrors the debris-arc arithmetic in trajectory.py).
                    _m_bo = (node.mass_initial - node.mass_propellant
                             if node.mass_propellant > 0 else node.mass_final)
                    _cas = (_m_bo - p.payload_kg
                            if _m_bo > p.payload_kg > 0 else node.mass_final)
                else:
                    _cas = node.mass_final
                beta = tumbling_cylinder_beta(_cas,
                                              node.diameter_m, node.length_m)
                if beta > 0:
                    _row(left, r, "Empty β (kg/m²):", f"{beta:,.0f}");            r += 1

            # ── Right: Engine Performance ─────────────────────────────
            mdot = (node.thrust_N / (node.isp_s * _G0)
                    if node.isp_s > 0 else 0.0)
            r = 0
            _row(right, r, "Thrust (kN):",       f"{node.thrust_N/1000:,.0f}");  r += 1
            _row(right, r, "ISP (s):",            f"{node.isp_s:.0f}");           r += 1
            _row(right, r, "Nozzle area (m²):",  f"{node.nozzle_exit_area_m2:.4f}"); r += 1
            _row(right, r, "Burntime (s):",       f"{node.burn_time_s:.1f}  (computed)"); r += 1
            _row(right, r, "Mass flow (kg/s):",   f"{mdot:.1f}  (computed)");     r += 1
            _row(right, r, "T/W ratio:",          f"{tw:.2f}  (computed)");       r += 1

            sn  += 1
            node = node.stage2

        # ── Front End ─────────────────────────────────────────────────
        af = ttk.LabelFrame(self._params_inner, text="Front End")
        af.pack(fill=tk.X, **pad)
        af.columnconfigure(1, weight=1)
        af.columnconfigure(3, weight=1)

        r = 0
        # Legacy hand-entered payload shape (older saved boosters only; new
        # builds derive ascent shape from the fairing or the run's object).
        _pd_m     = getattr(p, 'payload_diameter_m', 0.0)
        _pl_m     = p.nose_length_m
        _fe_shape = NOSE_SHAPE_LABELS.get(p.nose_shape, p.nose_shape)
        if _fe_shape or _pd_m > 0:
            _row2(af, r, "Payload shape:", _fe_shape if _fe_shape else "—",
                  "Payload diameter:", f"{_pd_m:.2f} m" if _pd_m > 0 else "—"); r += 1
            if _pl_m > 0:
                _ref_d = _pd_m if _pd_m > 0 else p.diameter_m
                _ld_str = f"  (L/D {_pl_m / _ref_d:.2f})" if _ref_d > 0 else ""
                _row2(af, r, "Payload length:", f"{_pl_m:.2f} m{_ld_str}"); r += 1

        _aero_LD = float(getattr(p, 'aerospike_LD', 0.0) or 0.0)
        _aero_dD = float(getattr(p, 'aerospike_dD', 0.0) or 0.0)
        if _aero_LD > 0:
            _row2(af, r, "Aerospike L/D:", f"{_aero_LD:.2f}",
                  "Aerodisk d/D:",
                  f"{_aero_dD:.2f}" if _aero_dD > 0 else "— (pointed)"); r += 1

        _ero = effective_ro(p)
        if _ero is not None:
            _row2(af, r, "Loadout:", f"{p.num_ros} × {_ero.name}",
                  "Per-object mass:", f"{_ero.mass_kg:,.0f} kg"); r += 1
            _ro_beta = _ero.beta_kg_m2
            _pbv_m   = p.bus_mass_kg
            # Non-separating body with β left at 0 = derive: preview the
            # geometry-derived β(Mach) — the value the run will use — at the
            # reference Mach, with its M2–12 range and a screening caveat.
            _is_body_ro = (getattr(_ero, 'separation_mode', 'separating_ro')
                           == 'body')
            _beta_txt = (f"{_ro_beta:,.0f} kg/m²" if _ro_beta > 0
                         else (self._derived_beta_preview(p, _ero)
                               if _is_body_ro else None))
            if _pbv_m > 0:
                _row2(af, r, "PBV mass:", f"{_pbv_m:,.0f} kg",
                      "Object β:", _beta_txt or "—"); r += 1
            elif _beta_txt:
                _row2(af, r, "Object β:", _beta_txt); r += 1
            _ro_shape_s = NOSE_SHAPE_LABELS.get(_ero.shape, NOSE_SHAPE_LABELS['cone'])
            _row2(af, r, "Reentry object shape:", _ro_shape_s,
                  "Object diameter:",
                  f"{_ero.diameter_m:.2f} m" if _ero.diameter_m > 0 else "—"); r += 1
            if _ero.length_m > 0:
                _row2(af, r, "Object length:", f"{_ero.length_m:.2f} m"); r += 1
            # Guidance (the glide law) and whether the object glides at all are
            # REENTRY-PLAN choices, not object hardware — the run flies the
            # sidebar strip (_reentry_plan_kwargs), so read the same live store
            # here rather than the object's baked-in default.  The sidebar
            # guidance variable already holds the exact display label; fall back
            # to the object only when there is no sidebar (headless render).
            _GUID_LABELS = {
                "equilibrium_glide":          "Equilibrium glide (Tracy)",
                "equilibrium_glide_acton":    "Non-oscillatory glide (Acton)",
                "damped_glide":               "Damped phugoid glide",
                "skip_to_equilibrium":        "Damped phugoid glide",
                "dynamic_equilibrium_glide":  "Dynamic equilibrium glide",
                "skip_glide":                 "Phugoid / skip-glide",
                "azimuth_command":            "Phugoid / skip-glide",
            }
            if hasattr(self, '_main_guidance_var'):
                _guid_lbl = self._main_guidance_var.get()
                _glide_on = "ballistic" not in _guid_lbl.lower()
            else:
                _guid_lbl = _GUID_LABELS.get(_ero.glider_guidance,
                                             _ero.glider_guidance)
                _glide_on = _ero.glider_enabled
            if _glide_on:
                # glider_LD is the airframe CAPABILITY; a plan's commanded L/D
                # may fly it lower (never higher), so label it as the ceiling.
                # For a non-separating body with L/D=0 = derive, preview the
                # geometry-derived (trim-gated) value the run will use.
                _ld_txt = (f"{_ero.glider_LD:.2f}  (capability)"
                           if _ero.glider_LD > 0
                           else (self._derived_ld_preview(p, _ero)
                                 if _is_body_ro else "0.00  (capability)"))
                _row2(af, r, "Glider L/D:", _ld_txt,
                      "Guidance:", _guid_lbl); r += 1

        # ── Shroud ────────────────────────────────────────────────────
        if p.shroud_mass_kg > 0:
            ff = ttk.LabelFrame(self._params_inner, text="Fairing")
            ff.pack(fill=tk.X, **pad)
            r = 0
            _row(ff, r, "Mass (kg):",          f"{p.shroud_mass_kg:,.0f}"); r += 1
            _row(ff, r, "Jettison alt (km):",  f"{p.shroud_jettison_alt_km:.0f}"); r += 1
            if p.shroud_diameter_m > 0:
                _sd = p.shroud_diameter_m
                _row(ff, r, "Diameter (m):",   f"{_sd:.2f}"); r += 1
                _area_ratio = (_sd / p.diameter_m) ** 2
                _row(ff, r, "Area vs body:",   f"{_area_ratio:.2f}×  (drag pre-jettison)"); r += 1
            else:
                _sd = p.diameter_m
            if p.shroud_nose_shape not in ('', 'forden'):
                _row(ff, r, "Nose shape:",
                     NOSE_SHAPE_LABELS.get(p.shroud_nose_shape, p.shroud_nose_shape)); r += 1
                if p.shroud_nose_length_m > 0 and _sd > 0:
                    _sld = p.shroud_nose_length_m / _sd
                    _row(ff, r, "Nose length (m):",
                         f"{p.shroud_nose_length_m:.2f}  (L/D = {_sld:.2f})"); r += 1
            if p.shroud_length_m > 0:
                _row(ff, r, "Length (m):",     f"{p.shroud_length_m:.2f}"); r += 1
                beta = tumbling_cylinder_beta(p.shroud_mass_kg, _sd, p.shroud_length_m)
                if beta > 0:
                    _row(ff, r, "Fairing β (kg/m²):", f"{beta:,.0f}"); r += 1


    # ------------------------------------------------------------------
    # Shared location picker — used by launch site, aim-at-target, and
    # estimate-azimuth.  Writes to lat_var / lon_var on Apply.
    # ------------------------------------------------------------------
    def _pick_location(self, lat_var, lon_var, parent=None):
        """
        Open a small modal search dialog.  The user types a place name;
        results come from geonamescache (offline, instant) or Nominatim
        (online, threaded).  Selecting a row and clicking Apply writes
        decimal-degree strings into lat_var and lon_var.

        Direct lat/lon entry in the caller is unaffected — this is purely
        an optional convenience.
        """
        if parent is None:
            parent = self

        dlg = tk.Toplevel(parent)
        dlg.title("Find Location")
        dlg.resizable(True, False)
        dlg.grab_set()
        dlg.minsize(460, 300)

        frm = ttk.Frame(dlg, padding=10)
        frm.pack(fill=tk.BOTH, expand=True)
        frm.columnconfigure(1, weight=1)
        frm.rowconfigure(1, weight=1)

        # Search row — entry + column of action buttons
        ttk.Label(frm, text="Search:").grid(
            row=0, column=0, sticky=tk.W, padx=(0, 4), pady=(0, 4))
        search_var = tk.StringVar()
        search_entry = ttk.Entry(frm, textvariable=search_var, width=28)
        search_entry.grid(row=0, column=1, sticky=tk.EW, pady=(0, 4))
        # Both action buttons share a frame in column 2 so they stack neatly.
        _btn_col = ttk.Frame(frm)
        _btn_col.grid(row=0, column=2, sticky=tk.NW, padx=(6, 0), pady=(0, 4))
        online_btn = ttk.Button(_btn_col, text="Search online…", width=15)
        online_btn.pack(fill=tk.X, pady=(0, 2))
        # Map-picker button added after _pick_on_map is defined below.

        # Results listbox
        lb_frm = ttk.Frame(frm)
        lb_frm.grid(row=1, column=0, columnspan=3, sticky=tk.NSEW, pady=(0, 4))
        lb_frm.columnconfigure(0, weight=1)
        lb_frm.rowconfigure(0, weight=1)
        sb = ttk.Scrollbar(lb_frm, orient=tk.VERTICAL)
        lb = tk.Listbox(lb_frm, height=10, yscrollcommand=sb.set,
                        selectmode=tk.SINGLE, font=("TkFixedFont", 9),
                        activestyle="dotbox")
        sb.config(command=lb.yview)
        lb.grid(row=0, column=0, sticky=tk.NSEW)
        sb.grid(row=0, column=1, sticky=tk.NS)

        # Status / selection preview line
        status_var = tk.StringVar(value="")
        ttk.Label(frm, textvariable=status_var, font=("TkDefaultFont", 9),
                  justify=tk.LEFT).grid(
            row=2, column=0, columnspan=3, sticky=tk.W, pady=2)

        # ── Data ──────────────────────────────────────────────────────
        _rows = []   # [(display_name, lat_dd, lon_dd), …]

        # Offline search order: the BUNDLED gazetteer (data/gazetteer/
        # packs — official BGN-lineage sources, variants searchable) →
        # geonamescache if installed (legacy) → online Nominatim.
        # The gazetteer's SQLite index is a multi-minute / ~2.5 GB build
        # for the worldwide packs, so this dialog NEVER triggers it: it
        # uses the index only if a serious user already built it via
        # Analysis ▸ Reference Data ▸ Offline Gazetteer.  A casual user
        # who never builds it just gets online search, no surprise wait.
        _gaz_db = [None]
        _gc_cities = None
        import gazetteer as _gz
        if _gz.available() and _gz.index_ready():
            try:
                _gaz_db[0] = _gz.ensure_index()      # instant: already built
            except Exception:
                _gaz_db[0] = None
        elif _gz.available():
            status_var.set("Offline gazetteer not built — Analysis ▸ "
                           "Reference Data ▸ Offline Gazetteer (or Search "
                           "online…)")
        else:
            try:
                import geonamescache as _gnc
                _gc_cities = list(_gnc.GeonamesCache().get_cities().values())
            except ImportError:
                status_var.set("No offline gazetteer — use Search online…")

        def _do_offline(query):
            q = query.strip()
            if not q:
                return []
            if _gaz_db[0] is not None:
                hits = []
                for r in _gz.search(q, limit=50, db=_gaz_db[0]):
                    name = r["primary"]
                    if r["matched"] != r["primary"]:
                        name += f"  (= {r['matched']})"
                    hits.append((name, r["lat"], r["lon"],
                                 (r["admin"] or "")[:14], 0))
                return hits
            if not _gc_cities:
                return []
            ql = q.lower()
            hits = []
            for c in _gc_cities:
                if ql in c['name'].lower():
                    hits.append((
                        c['name'],
                        float(c['latitude']),
                        float(c['longitude']),
                        c.get('countrycode', ''),
                        int(c.get('population') or 0),
                    ))
            hits.sort(key=lambda x: -x[4])
            return hits[:50]

        def _fmt_row(name, lat, lon, cc):
            ns = 'N' if lat >= 0 else 'S'
            ew = 'E' if lon >= 0 else 'W'
            return f"{name}, {cc:<3}  {abs(lat):>7.2f}°{ns}  {abs(lon):>8.2f}°{ew}"

        def _populate(results):
            _rows.clear()
            lb.delete(0, tk.END)
            for name, lat, lon, cc, *_ in results:
                lb.insert(tk.END, _fmt_row(name, lat, lon, cc))
                _rows.append((name, lat, lon))

        # ── Search scheduling ──────────────────────────────────────────
        _after_id = [None]

        def _schedule(*_):
            if _after_id[0]:
                dlg.after_cancel(_after_id[0])
            _after_id[0] = dlg.after(200, _run_offline)

        def _run_offline():
            q = search_var.get()
            results = _do_offline(q)
            _populate(results)
            if results:
                status_var.set(
                    f"{len(results)} result(s) — select one and click Apply")
            elif q.strip() and (_gaz_db[0] is not None or _gc_cities):
                status_var.set("No offline match — try Search online…")
            # Online stays available whenever there is a query: the
            # bundled gazetteer nearly always finds SOMETHING offline,
            # and the old only-when-empty gating would leave the button
            # permanently grey (observed 2026-08-17).
            online_btn.config(state=tk.NORMAL if q.strip() else tk.DISABLED)

        def _run_online():
            q = search_var.get().strip()
            if not q:
                return
            online_btn.config(state=tk.DISABLED, text="Searching…")
            status_var.set("Querying Nominatim (OpenStreetMap)…")

            def _thread():
                try:
                    from geopy.geocoders import Nominatim
                    geo  = Nominatim(user_agent="thrusty-location-picker")
                    locs = geo.geocode(q, exactly_one=False, limit=15) or []
                    rows = [(loc.address[:52], loc.latitude,
                             loc.longitude, '', 0) for loc in locs]
                    msg  = (f"{len(rows)} result(s) — select one and click Apply"
                            if rows else "No results from Nominatim")
                except ImportError:
                    rows, msg = [], "geopy not installed  (pip install geopy)"
                except Exception as ex:
                    rows, msg = [], f"Online search failed: {ex}"
                dlg.after(0, lambda: (
                    _populate(rows),
                    status_var.set(msg),
                    online_btn.config(state=tk.NORMAL, text="Search online…"),
                ))

            threading.Thread(target=_thread, daemon=True).start()

        online_btn.config(command=_run_online, state=tk.DISABLED)

        # ── Map picker ─────────────────────────────────────────────────
        def _pick_on_map():
            try:
                import cartopy.crs      as ccrs
                import cartopy.feature  as cfeature
            except ImportError:
                messagebox.showinfo(
                    "Missing package",
                    "cartopy is not installed.\n\nRun:  pip install cartopy",
                    parent=dlg)
                return

            map_win = tk.Toplevel(dlg)
            map_win.title("Click to select location — close to cancel")
            map_win.grab_set()

            fig = Figure(figsize=(9, 4.5), tight_layout=True)
            ax  = fig.add_subplot(1, 1, 1, projection=ccrs.PlateCarree())
            ax.set_global()
            ax.add_feature(cfeature.LAND,   facecolor='#f0ede8', zorder=0)
            ax.add_feature(cfeature.OCEAN,  facecolor='#d0e8f0', zorder=0)
            ax.coastlines(resolution='110m', linewidth=0.5, color='#444')
            ax.add_feature(cfeature.BORDERS, linewidth=0.3,
                           edgecolor='#888', zorder=1)
            ax.gridlines(draw_labels=True, linewidth=0.3, alpha=0.5,
                         xlocs=range(-180, 181, 30),
                         ylocs=range(-90,   91, 30))
            ax.set_title("Click anywhere to select — close window to cancel",
                         fontsize=9)

            canvas = FigureCanvasTkAgg(fig, master=map_win)
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            canvas.draw()
            ttk.Label(map_win,
                      text="Click the map to select a point.  "
                           "Close this window to cancel.",
                      font=("TkDefaultFont", 9)).pack(pady=4)

            _clicked = [False]

            def _on_map_click(event):
                if event.inaxes is not ax or event.xdata is None:
                    return
                if _clicked[0]:
                    return          # guard against duplicate events
                _clicked[0] = True
                lon_c, lat_c = event.xdata, event.ydata
                ax.plot(lon_c, lat_c, 'r+', markersize=14,
                        markeredgewidth=2,
                        transform=ccrs.PlateCarree(), zorder=10)
                canvas.draw()
                map_win.after(350, lambda: _confirm_map(lat_c, lon_c))

            def _confirm_map(lat_c, lon_c):
                lat_var.set(f"{lat_c:.4f}")
                lon_var.set(f"{lon_c:.4f}")
                ns = 'N' if lat_c >= 0 else 'S'
                ew = 'E' if lon_c >= 0 else 'W'
                status_var.set(
                    f"Map pick  →  {abs(lat_c):.4f}°{ns},  "
                    f"{abs(lon_c):.4f}°{ew}")
                map_win.destroy()

            canvas.mpl_connect('button_press_event', _on_map_click)

        ttk.Button(_btn_col, text="Pick on map…", width=15,
                   command=_pick_on_map).pack(fill=tk.X)

        search_var.trace_add("write", _schedule)

        # ── Selection & apply ──────────────────────────────────────────
        def _on_select(_evt=None):
            sel = lb.curselection()
            if not sel:
                return
            name, lat, lon = _rows[sel[0]]
            ns = 'N' if lat >= 0 else 'S'
            ew = 'E' if lon >= 0 else 'W'
            status_var.set(
                f"{name}  →  {abs(lat):.4f}°{ns},  {abs(lon):.4f}°{ew}")

        def _apply():
            sel = lb.curselection()
            if not sel:
                messagebox.showinfo("No selection",
                                    "Select a location from the list.",
                                    parent=dlg)
                return
            _, lat, lon = _rows[sel[0]]
            lat_var.set(f"{lat:.4f}")
            lon_var.set(f"{lon:.4f}")
            dlg.destroy()

        lb.bind("<<ListboxSelect>>", _on_select)
        lb.bind("<Double-1>",        lambda _e: _apply())

        btn_frm = ttk.Frame(frm)
        btn_frm.grid(row=3, column=0, columnspan=3, pady=(6, 0))
        ttk.Button(btn_frm, text="Apply",  width=8,
                   command=_apply        ).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frm, text="Cancel", width=8,
                   command=dlg.destroy   ).pack(side=tk.LEFT, padx=4)

        dlg.bind("<Return>", lambda _e: _apply())
        dlg.bind("<Escape>", lambda _e: dlg.destroy())
        search_entry.focus_set()

    # ------------------------------------------------------------------
    # Estimate azimuth from target coordinates (with Earth-rotation correction)
    # ------------------------------------------------------------------
    def _estimate_azimuth(self):
        """
        Open a dialog to compute the launch azimuth needed to hit a target
        at a given lat/lon, optionally corrected for Earth's rotation during
        flight.  Sets the azimuth entry on Apply.
        """
        dlg = tk.Toplevel(self)
        dlg.title("Estimate Launch Azimuth")
        dlg.resizable(False, False)
        dlg.grab_set()

        frm = ttk.Frame(dlg, padding=10)
        frm.pack(fill=tk.BOTH, expand=True)

        # Target coordinates
        ttk.Label(frm, text="Target latitude (°):").grid(
            row=0, column=0, sticky=tk.W, padx=(0, 6), pady=4)
        lat_var = tk.StringVar(value="0.0")
        ttk.Entry(frm, textvariable=lat_var, width=12).grid(
            row=0, column=1, sticky=tk.W, pady=4)

        ttk.Label(frm, text="Target longitude (°):").grid(
            row=1, column=0, sticky=tk.W, padx=(0, 6), pady=4)
        lon_var = tk.StringVar(value="0.0")
        ttk.Entry(frm, textvariable=lon_var, width=12).grid(
            row=1, column=1, sticky=tk.W, pady=4)
        ttk.Button(frm, text="Find…", width=7,
                   command=lambda: self._pick_location(lat_var, lon_var, dlg)
                   ).grid(row=1, column=2, sticky=tk.W, padx=4, pady=4)

        ttk.Separator(frm, orient=tk.HORIZONTAL).grid(
            row=2, column=0, columnspan=3, sticky=tk.EW, pady=6)
        ttk.Label(frm, text="Earth-rotation correction:").grid(
            row=3, column=0, columnspan=3, sticky=tk.W)

        # Correction method
        method_var = tk.StringVar(value="ballistic")
        user_t_var = tk.StringVar(value="10.0")
        preview_var = tk.StringVar(value="")

        last_t = (self._result.get('time_of_flight_s')
                  if self._result is not None else None)

        OMEGA = 7.2921e-5   # Earth rotation rate (rad/s, sidereal)
        G_MS2 = 9.81

        def _compute():
            try:
                lat1 = np.radians(float(self._launch_lat.get()))
                lon1 = np.radians(float(self._launch_lon.get()))
                lat2 = np.radians(float(lat_var.get()))
                lon2 = np.radians(float(lon_var.get()))
            except (ValueError, AttributeError):
                return None

            rng_m  = float(range_between(lat1, lon1, lat2, lon2))
            rng_km = rng_m / 1000.0

            method = method_var.get()
            if method == "none":
                T = 0.0
            elif method == "ballistic":
                # Minimum-energy ballistic flight time: T = sqrt(2 R / g)
                T = float(np.sqrt(2.0 * rng_m / G_MS2))
            elif method == "last_sim":
                T = float(last_t) if last_t else 0.0
            elif method == "user_t":
                try:
                    T = float(user_t_var.get()) * 60.0
                except ValueError:
                    T = 0.0
            else:
                T = 0.0

            # Target drifts east by Ω·T during flight; aim at where it
            # WILL BE, i.e. shift the aim longitude east by the same amount.
            # (Equivalently: the booster's inertial trajectory must end
            # at the target's future ECI position.)
            dlon_corr = OMEGA * T

            def _bearing(la1, lo1, la2, lo2):
                dl = lo2 - lo1
                x  = np.sin(dl) * np.cos(la2)
                y  = (np.cos(la1) * np.sin(la2)
                      - np.sin(la1) * np.cos(la2) * np.cos(dl))
                return float(np.degrees(np.arctan2(x, y)) % 360.0)

            az_u = _bearing(lat1, lon1, lat2, lon2)
            az   = _bearing(lat1, lon1, lat2, lon2 + dlon_corr)
            return rng_km, T, float(np.degrees(dlon_corr)), az, az_u

        def _update(*_):
            r = _compute()
            if r is None:
                preview_var.set("(enter valid target coordinates)")
                return
            rng_km, T, dlon_deg, az, az_u = r
            preview_var.set(
                f"Range:            {rng_km:>8.1f} km\n"
                f"Flight time T:    {T:>8.0f} s  ({T/60:.1f} min)\n"
                f"Earth rotation:   {dlon_deg:>8.2f}°  east\n"
                f"Bearing (no corr.):{az_u:>7.2f}°\n"
                f"Bearing (corr.):  {az:>8.2f}°"
            )

        ttk.Radiobutton(
            frm, text="Ballistic minimum-energy estimate  (T = √(2R/g))",
            variable=method_var, value="ballistic",
            command=_update
        ).grid(row=4, column=0, columnspan=3, sticky=tk.W, padx=(20, 0))

        last_label = (f"Use last simulation  (T = {last_t:.0f} s)"
                      if last_t else "Use last simulation  (no result yet)")
        ttk.Radiobutton(
            frm, text=last_label,
            variable=method_var, value="last_sim",
            state=(tk.NORMAL if last_t else tk.DISABLED),
            command=_update
        ).grid(row=5, column=0, columnspan=3, sticky=tk.W, padx=(20, 0))

        user_t_row = ttk.Frame(frm)
        user_t_row.grid(row=6, column=0, columnspan=3, sticky=tk.W, padx=(20, 0))
        ttk.Radiobutton(
            user_t_row, text="User flight time (min):",
            variable=method_var, value="user_t",
            command=_update
        ).pack(side=tk.LEFT)
        ttk.Entry(user_t_row, textvariable=user_t_var, width=8
                  ).pack(side=tk.LEFT, padx=4)

        ttk.Radiobutton(
            frm, text="No correction (instantaneous bearing)",
            variable=method_var, value="none",
            command=_update
        ).grid(row=7, column=0, columnspan=3, sticky=tk.W, padx=(20, 0))

        ttk.Separator(frm, orient=tk.HORIZONTAL).grid(
            row=8, column=0, columnspan=3, sticky=tk.EW, pady=6)
        ttk.Label(frm, textvariable=preview_var,
                  font=("TkFixedFont", 9), justify=tk.LEFT
                  ).grid(row=9, column=0, columnspan=3, sticky=tk.W, pady=2)

        def _apply():
            r = _compute()
            if r is None:
                messagebox.showerror(
                    "Input error",
                    "Target latitude and longitude must be numbers.",
                    parent=dlg)
                return
            _, _, _, az, _ = r
            self._azimuth_var.set(f"{az:.2f}")
            dlg.destroy()

        btn_frm = ttk.Frame(frm)
        btn_frm.grid(row=10, column=0, columnspan=3, pady=(8, 0))
        ttk.Button(btn_frm, text="Apply",  width=8, command=_apply
                   ).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frm, text="Cancel", width=8, command=dlg.destroy
                   ).pack(side=tk.LEFT, padx=4)

        # Live update as user edits coordinates or user-flight-time field
        lat_var.trace_add("write", _update)
        lon_var.trace_add("write", _update)
        user_t_var.trace_add("write", _update)

        _update()
        dlg.bind("<Return>", lambda e: _apply())
        dlg.bind("<Escape>", lambda e: dlg.destroy())

    # ------------------------------------------------------------------
    # Aim at target
    # ------------------------------------------------------------------
    def _aim_at_target(self):
        """
        Prompt for target lat/lon, then compute great-circle azimuth and
        bisect cutoff time to hit the target range.
        """
        dlg = tk.Toplevel(self)
        dlg.title("Aim at Target (liquid)")
        dlg.resizable(False, False)
        dlg.grab_set()

        frm = ttk.Frame(dlg, padding=10)
        frm.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frm, text="Target Latitude (°):").grid(
            row=0, column=0, sticky=tk.W, padx=(0, 6), pady=4)
        lat_var = tk.StringVar(value="0.0")
        ttk.Entry(frm, textvariable=lat_var, width=12).grid(
            row=0, column=1, sticky=tk.W, pady=4)

        ttk.Label(frm, text="Target Longitude (°):").grid(
            row=1, column=0, sticky=tk.W, padx=(0, 6), pady=4)
        lon_var = tk.StringVar(value="0.0")
        ttk.Entry(frm, textvariable=lon_var, width=12).grid(
            row=1, column=1, sticky=tk.W, pady=4)
        ttk.Button(frm, text="Find…", width=7,
                   command=lambda: self._pick_location(lat_var, lon_var, dlg)
                   ).grid(row=1, column=2, sticky=tk.W, padx=4, pady=4)

        result = {}

        def _ok():
            try:
                result['lat'] = float(lat_var.get())
                result['lon'] = float(lon_var.get())
            except ValueError:
                messagebox.showerror("Input error",
                                     "Latitude and longitude must be numbers.",
                                     parent=dlg)
                return
            dlg.destroy()

        def _cancel():
            dlg.destroy()

        btn_frm = ttk.Frame(frm)
        btn_frm.grid(row=2, column=0, columnspan=3, pady=(8, 0))
        ttk.Button(btn_frm, text="OK",     width=8, command=_ok    ).pack(side=tk.LEFT, padx=4)
        ttk.Button(btn_frm, text="Cancel", width=8, command=_cancel).pack(side=tk.LEFT, padx=4)

        dlg.bind("<Return>", lambda e: _ok())
        dlg.bind("<Escape>", lambda e: _cancel())
        self.wait_window(dlg)

        if 'lat' not in result:
            return

        try:
            lat1_dd = float(self._launch_lat.get())
            lon1_dd = float(self._launch_lon.get())
            lat2_dd = result['lat']
            lon2_dd = result['lon']

            lat1 = np.radians(lat1_dd)
            lon1 = np.radians(lon1_dd)
            lat2 = np.radians(lat2_dd)
            lon2 = np.radians(lon2_dd)

            dlon = lon2 - lon1
            x = np.sin(dlon) * np.cos(lat2)
            y = np.cos(lat1)*np.sin(lat2) - np.sin(lat1)*np.cos(lat2)*np.cos(dlon)
            az = np.degrees(np.arctan2(x, y)) % 360
            self._azimuth_var.set(f"{az:.2f}")

            rng_km = range_between(lat1, lon1, lat2, lon2) / 1000.0
            self._status_var.set(
                f"Target: {rng_km:.1f} km  |  Azimuth: {az:.1f}°  —  "
                "Computing cutoff time…")

            booster  = get_booster(self._booster_var.get())
            guidance = self._guidance_var.get()
            la           = self._field_float("Loft / burnout angle (°)",
                                             self._loft_angle_var.get(), required=True)
            gt_start_s   = self._field_float("Turn start (s)",
                                             self._gt_turn_start_var.get(), default=5.0)
            gt_stop_s    = self._field_float("Turn stop (s)",
                                             self._gt_turn_stop_var.get())
            try:
                booster.launch_elevation_deg = float(self._launch_el_var.get())
            except (ValueError, AttributeError):
                pass
            threading.Thread(
                target=self._aim_thread,
                args=(booster, guidance, lat1_dd, lon1_dd, az, rng_km, la,
                      gt_start_s, gt_stop_s),
                daemon=True,
            ).start()

        except Exception as e:
            messagebox.showerror("Aim error", str(e))

    def _aim_thread(self, booster, guidance, lat, lon, az, rng_km, la,
                    gt_start_s=5.0, gt_stop_s=None):
        try:
            cutoff = aim_booster(booster, lat, lon, az, rng_km,
                                 guidance=guidance,
                                 burnout_angle_deg=la,
                                 gt_turn_start_s=gt_start_s,
                                 gt_turn_stop_s=gt_stop_s)
            self.after(0, lambda: self._cutoff_var.set(f"{cutoff:.1f}"))
            self.after(0, lambda: self._status_var.set(
                f"Target: {rng_km:.1f} km  |  Azimuth: {az:.1f}°  |  "
                f"Cutoff: {cutoff:.1f} s"))
        except Exception as e:
            self.after(0, lambda: self._status_var.set(f"Aim failed: {e}"))

    # ------------------------------------------------------------------
    # Run buttons
    # ------------------------------------------------------------------
    def _save_active_plans(self):
        try:
            _ACTIVE_PLANS_PATH.parent.mkdir(parents=True, exist_ok=True)
            _ACTIVE_PLANS_PATH.write_text(json.dumps(mm.ACTIVE_FLIGHT_PLANS, indent=2))
        except Exception as exc:
            print(f"Warning: could not save active flight plans: {exc}")

    def _refresh_flight_plan_list(self, select=None):
        """Repopulate the Flight Plan combobox for the current booster."""
        name = self._booster_var.get()
        plans = (mm.list_flight_plans(name, extra_dirs=mm.USER_FLIGHT_PLAN_DIRS)
                 if name in BOOSTER_DB else [mm.DEFAULT_PLAN_LABEL])
        self._fp_cb.config(values=plans)
        active = select or mm.ACTIVE_FLIGHT_PLANS.get(name, mm.DEFAULT_PLAN_LABEL)
        if active not in plans:
            active = mm.DEFAULT_PLAN_LABEL
        self._fp_var.set(active)
        self._fp_del_btn.config(
            state=tk.NORMAL if active != mm.DEFAULT_PLAN_LABEL else tk.DISABLED)
        self._update_flight_plan_summary()

    def _on_flight_plan_selected(self, _event=None):
        name = self._booster_var.get()
        sel = self._fp_var.get()
        if sel == mm.DEFAULT_PLAN_LABEL:
            mm.ACTIVE_FLIGHT_PLANS.pop(name, None)
        else:
            mm.ACTIVE_FLIGHT_PLANS[name] = sel
        self._save_active_plans()
        self._fp_del_btn.config(
            state=tk.NORMAL if sel != mm.DEFAULT_PLAN_LABEL else tk.DISABLED)
        self._on_booster_changed()

    def _update_flight_plan_summary(self):
        """Caption under the plan combobox.  House rule: only say what is
        NOT evident in the fields below — mode and burnout angle are the
        Mode/Burnout widgets right there, so the caption carries only the
        fairing jettison (shown nowhere else in the strip), or nothing."""
        try:
            p = get_booster(self._booster_var.get())
        except Exception:
            self._fp_summary_var.set("")
            return
        if getattr(p, 'shroud_mass_kg', 0.0) > 0:
            self._fp_summary_var.set(
                "fairing jettison @ "
                + (f"{p.shroud_jettison_alt_km:g} km"
                   if p.shroud_jettison_alt_km > 0 else "heating-based"))
        else:
            self._fp_summary_var.set("")

    def _ask_new_plan_name_and_law(self, booster_name):
        """Modal prompt for a new flight plan's name AND its guidance law.

        The law is the plan's identity — chosen here, immutable afterwards
        (the editor and sidebar only toggle Simple/Advanced within a pitch
        plan).  Returns (name, law_key) or (None, None) on cancel.
        """
        dlg = tk.Toplevel(self)
        dlg.title("New Flight Plan")
        dlg.resizable(False, False)
        dlg.grab_set()
        frm = ttk.Frame(dlg, padding=12)
        frm.pack(fill=tk.BOTH, expand=True)
        ttk.Label(frm, text=f"Name for the new flight plan for '{booster_name}':").grid(
            row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 4))
        name_var = tk.StringVar()
        ent = ttk.Entry(frm, textvariable=name_var, width=32)
        ent.grid(row=1, column=0, columnspan=2, sticky=tk.EW, pady=(0, 10))
        ttk.Label(frm, text="Guidance law (fixed for the life of the plan):").grid(
            row=2, column=0, columnspan=2, sticky=tk.W, pady=(0, 4))
        law_var = tk.StringVar(value=get_booster(booster_name).guidance)
        for i, (key, lbl, hint) in enumerate((
                ("pitch_program", "Pitch program",
                 "commanded pitch to a burnout angle; Simple or Advanced per-stage"),
                ("true_gravity_turn", "Gravity turn",
                 "thrust along velocity from the launch elevation; η kick"),
                ("orbital_insertion", "Orbital insertion",
                 "two-phase boost to a target orbit; solve with Plan Orbit"))):
            ttk.Radiobutton(frm, text=lbl, variable=law_var, value=key).grid(
                row=3 + 2 * i, column=0, columnspan=2, sticky=tk.W, padx=(4, 0))
            ttk.Label(frm, text=hint, foreground="#888888").grid(
                row=4 + 2 * i, column=0, columnspan=2, sticky=tk.W, padx=(24, 0))
        out = {}
        def _ok(*_):
            out['name'] = name_var.get().strip()
            out['law'] = law_var.get()
            dlg.destroy()
        bf = ttk.Frame(frm)
        bf.grid(row=9, column=0, columnspan=2, sticky=tk.E, pady=(12, 0))
        ttk.Button(bf, text="Cancel", command=dlg.destroy).pack(side=tk.RIGHT, padx=(4, 0))
        ttk.Button(bf, text="Create", command=_ok).pack(side=tk.RIGHT)
        ent.bind("<Return>", _ok)
        ent.focus_set()
        self.wait_window(dlg)
        return (out.get('name') or None, out.get('law'))

    def _new_flight_plan(self):
        """Create a plan variant: name + guidance law, seeded from the active
        plan's non-law fields (events, yaw, launch elevation carry over)."""
        name = self._booster_var.get()
        if name not in BOOSTER_DB:
            messagebox.showinfo("Flight Plan", "Select a booster first.", parent=self)
            return
        new_name, law = self._ask_new_plan_name_and_law(name)
        if not new_name:
            return
        if new_name in mm.RESERVED_PLAN_NAMES:
            messagebox.showerror("Flight Plan",
                                 f"'{new_name}' is reserved.", parent=self)
            return
        booster = get_booster(name)          # active plan as the starting point
        plan = mm._merge_flight_plans(extract_flight_plan(booster),
                                      self._raw_active_plan(name))
        if law != plan.get('guidance'):
            # Crossing laws: per-stage pitch angles from the old law would be
            # wrong-valued (pitch angle vs η) or masked — start those clean.
            plan['guidance'] = law
            for st in plan.get('stages', []):
                st['stage_burnout_angle_deg'] = None
        dlg = FlightPlanDialog(self, f"{name} — {new_name}", plan, booster)
        self.wait_window(dlg)
        if dlg.result is None:
            return
        try:
            save_flight_plan(name, dlg.result, _FLIGHT_PLAN_LIBRARY_PATH, plan=new_name)
        except Exception as exc:
            messagebox.showerror("Flight Plan",
                                 f"Could not save flight plan:\n{exc}", parent=self)
            return
        mm.ACTIVE_FLIGHT_PLANS[name] = new_name
        self._save_active_plans()
        self._refresh_flight_plan_list(select=new_name)
        self._on_booster_changed()
        self._status_var.set(f"Flight plan '{new_name}' created for '{name}'.")

    def _delete_flight_plan(self):
        name = self._booster_var.get()
        sel = self._fp_var.get()
        if sel == mm.DEFAULT_PLAN_LABEL:
            return
        if not messagebox.askyesno(
                "Delete Flight Plan",
                f"Delete flight plan '{sel}' for '{name}'?", parent=self):
            return
        fp = Path(_FLIGHT_PLAN_LIBRARY_PATH) / mm.flight_plan_filename(name, sel)
        try:
            fp.unlink(missing_ok=True)
        except Exception as exc:
            messagebox.showerror("Flight Plan",
                                 f"Could not delete flight plan:\n{exc}", parent=self)
            return
        mm.ACTIVE_FLIGHT_PLANS.pop(name, None)
        self._save_active_plans()
        self._refresh_flight_plan_list(select=mm.DEFAULT_PLAN_LABEL)
        self._on_booster_changed()
        self._status_var.set(f"Flight plan '{sel}' deleted.")

    def _edit_flight_plan_main(self):
        """Open the flight-plan editor for the ACTIVE plan of the selected
        booster and, on save, write it back (default file or the named variant)."""
        name = self._booster_var.get()
        if name not in BOOSTER_DB:
            messagebox.showinfo("Flight Plan", "Select a booster first.", parent=self)
            return
        sel = self._fp_var.get()
        booster = get_booster(name)
        # Seed from the raw plan file merged over the extraction — the file is
        # the only source of the GUI keys (yaw, turn window) and provenance;
        # seeding from extract alone wiped them on the next Save.
        plan = mm._merge_flight_plans(
            extract_flight_plan(booster),
            self._raw_active_plan(name,
                                  None if sel == mm.DEFAULT_PLAN_LABEL else sel))
        title_name = name if sel == mm.DEFAULT_PLAN_LABEL else f"{name} — {sel}"
        dlg = FlightPlanDialog(self, title_name, plan, booster)
        self.wait_window(dlg)
        if dlg.result is None:
            return
        try:
            save_flight_plan(name, dlg.result, _FLIGHT_PLAN_LIBRARY_PATH,
                             plan=None if sel == mm.DEFAULT_PLAN_LABEL else sel)
        except Exception as exc:
            messagebox.showerror("Flight Plan",
                                 f"Could not save flight plan:\n{exc}", parent=self)
            return
        self._on_booster_changed()
        self._status_var.set(f"Flight plan for '{title_name}' saved.")

    def _set_engine_cutoff(self):
        """Analysis ▸ Engine Cutoff… — view/set the early-cutoff time.

        Liquid engines only: solid motors burn to completion regardless
        (the integrator enforces this).  Blank = full burn.  Aim-at-Target
        writes its computed cutoff into the same setting."""
        from tkinter import simpledialog
        cur = self._cutoff_var.get().strip()
        s = simpledialog.askstring(
            "Engine Cutoff (liquid)",
            "Command engine cutoff at time t after launch (s).\n"
            "Liquid engines only — solid motors burn to completion.\n"
            "Leave blank for full burn.",
            initialvalue=cur, parent=self)
        if s is None:
            return                              # cancelled
        s = s.strip()
        if s:
            try:
                self._field_float("Engine cutoff (s)", s, required=True)
            except ValueError as e:
                messagebox.showerror("Engine Cutoff", str(e), parent=self)
                return
        self._cutoff_var.set(s)
        self._status_var.set(f"Engine cutoff: {s} s" if s
                             else "Engine cutoff cleared — full burn.")

    def _set_reentry_query(self):
        """Analysis ▸ Re-entry Query… — report the state vector at an altitude
        on descent.  A per-run diagnostic (formerly a sidebar panel); blank
        disables the query."""
        from tkinter import simpledialog
        cur = (self._query_alt_km_var.get().strip()
               if self._query_alt_enable.get() else "")
        s = simpledialog.askstring(
            "Re-entry Query",
            "Report the trajectory state at this altitude on descent (km).\n"
            "Adds a milestone row with speed, angle and time at that height.\n"
            "Leave blank to disable.",
            initialvalue=cur, parent=self)
        if s is None:
            return                              # cancelled
        s = s.strip()
        if s:
            try:
                self._field_float("Re-entry query altitude (km)", s,
                                  required=True)
            except ValueError as e:
                messagebox.showerror("Re-entry Query", str(e), parent=self)
                return
            self._query_alt_km_var.set(s)
            self._query_alt_enable.set(True)
            self._status_var.set(f"Re-entry query: report state at {s} km "
                                 f"(descent).")
        else:
            self._query_alt_enable.set(False)
            self._status_var.set("Re-entry query disabled.")

    def _field_float(self, label, s, default=None, required=False):
        """Parse one sidebar numeric field, naming the field in any error.

        Blank fields return `default` (or raise if required).  A bad value
        produces "Turn start (s): '1-3' is not a number — enter a single
        number, not a range" instead of the bare float() message, so the
        user can tell WHICH field to fix."""
        s = (s or "").strip()
        if not s:
            if required:
                raise ValueError(f"{label} is required.")
            return default
        try:
            return float(s)
        except ValueError:
            import re as _re
            hint = (" — enter a single number, not a range"
                    if _re.search(r'\d\s*[-–—:]\s*\d', s) else "")
            raise ValueError(f"{label}: '{s}' is not a number{hint}") from None

    def _alpha_limit_value(self):
        """Sustained-α envelope (°) for the run, or None (no limit).

        The dogleg lives in the Flight Plan dialog (the sidebar yaw controls
        were removed); the α limit and induced-drag toggle now travel with it
        in the plan JSON.  On plan load they are surfaced into these vars
        (reset from the plan every time, so no stale value leaks), which are
        the transfer bus the run path reads — so this is NOT gated on the
        former sidebar checkbox.  SP-8099 §2.1.2.2 gives 5–10° at max q.
        """
        try:
            v = float(self._alpha_limit_var.get().strip())
        except (ValueError, AttributeError):
            return None
        return v if v > 0.0 else None

    def _alpha_induced_value(self):
        """Whether α-induced boost drag is enabled for the run.

        Surfaced from the plan into this var on load (see _alpha_limit_value);
        read directly, not gated on the removed sidebar yaw checkbox.
        """
        _v = getattr(self, '_alpha_induced_var', None)
        return bool(_v and _v.get())

    def _terrain_dem_value(self):
        """Whether the run uses real terrain (DEM) for launch and impact."""
        _v = getattr(self, '_terrain_dem_var', None)
        return bool(_v and _v.get())

    def _launch_elev_value(self):
        """Baked launch-pad elevation (m) for the current lat/lon, or None.

        launch_sites.json carries a per-site `elev_m` sampled once at high
        resolution (see its `elev_source`), which is finer than the bundled
        coarse grid the integrator falls back to.  Use it only while the
        lat/lon fields still sit on that site — once the user moves the pad,
        the baked number belongs to somewhere else and terrain.py resolves
        the elevation from the grid instead.
        """
        site = self._site_map.get(self._site_var.get())
        if site is None or site.get('elev_m') is None:
            return None
        try:
            lat = float(self._launch_lat.get())
            lon = float(self._launch_lon.get())
        except (ValueError, AttributeError, tk.TclError):
            return None
        if abs(lat - float(site['lat'])) > 1e-4 or abs(lon - float(site['lon'])) > 1e-4:
            return None
        return float(site['elev_m'])

    def _on_terrain_toggled(self, _event=None):
        """Refresh the elevation readout beside the "Use terrain (DEM)" box."""
        lbl = getattr(self, '_terrain_elev_lbl', None)
        if lbl is None:
            return
        if not self._terrain_dem_value():
            lbl.config(text="")
            return
        elev = self._launch_elev_value()
        if elev is not None:
            lbl.config(text=f"pad {elev:,.0f} m (site)")
            return
        # No baked site value — sample the bundled coarse grid for the readout,
        # which is exactly what the integrator will use for this pad.
        try:
            import terrain as _terrain
            lat = float(self._launch_lat.get())
            lon = float(self._launch_lon.get())
            lbl.config(
                text=f"pad {_terrain.ground_elevation(lat, lon):,.0f} m (grid)")
        except Exception:
            lbl.config(text="pad elevation unavailable")

    def _get_inputs(self):
        booster  = get_booster(self._booster_var.get())

        # Yaw is driven solely by the global dogleg grid (yaw_maneuvers below).
        # Strip any legacy per-stage stage_yaw_* off the run booster so a baked
        # dogleg can't override what the Flight Plan dialog shows.  (It is
        # surfaced into the grid on load; see _on_booster_changed.)
        def _stage_chain(_b):
            _n = _b
            while _n is not None:
                yield _n
                _n = getattr(_n, 'stage2', None)
        if any(getattr(_n, 'stage_yaw_final_az_deg', None) is not None
               for _n in _stage_chain(booster)):
            booster = copy.deepcopy(booster)
            for _n in _stage_chain(booster):
                _n.stage_yaw_start_s = None
                _n.stage_yaw_stop_s = None
                _n.stage_yaw_final_az_deg = None

        # Supply the reentry object from the sidebar library and COMPOSE the
        # run-level loadout onto the stage chain.  Any library object flies
        # on any booster — separation is a plan choice, not a compatibility
        # constraint.  The stack carries the whole front end through boost
        # (bus + N × object mass; compose_loadout adjusts every stage's
        # launch mass by the delta against whatever payload the chain was
        # built with), while one object is modeled on the way back.  Casing
        # debris strips the full loadout so nothing is counted twice.
        _ro_sel = getattr(self, '_ro_main_var', None)
        _ro_name = _ro_sel.get() if _ro_sel is not None else ""
        if _ro_name in RO_DB:
            _user_ro = RO_DB[_ro_name]()
            try:
                _n_load = max(1, int(self._loadout_n_var.get())) \
                    if hasattr(self, '_loadout_n_var') else 1
            except (ValueError, tk.TclError):
                _n_load = 1
            booster = mm.compose_loadout(booster, _user_ro, _n_load)
            _node, _placed = booster, False
            while _node is not None:
                if _node.ro is not None:
                    _node.ro = _user_ro
                    _placed = True
                    break
                _node = getattr(_node, 'stage2', None)
            if not _placed:
                booster.ro = _user_ro

        guidance = self._guidance_var.get()
        lat      = self._field_float("Launch latitude (°)",  self._launch_lat.get(),  required=True)
        lon      = self._field_float("Launch longitude (°)", self._launch_lon.get(),  required=True)
        az       = self._field_float("Azimuth (°)",          self._azimuth_var.get(), required=True)
        cutoff   = self._field_float("Engine cutoff (s)",    self._cutoff_var.get())
        la       = self._field_float("Loft / burnout angle (°)",
                                     self._loft_angle_var.get(), required=True)
        gt_start_s = self._field_float("Turn start (s)", self._gt_turn_start_var.get(),
                                       default=0.0)
        gt_stop_s  = self._field_float("Turn stop (s)",  self._gt_turn_stop_var.get())
        target_orbit_km = (self._field_float("Target orbit altitude (km)",
                                             self._orbit_alt_var.get())
                           if guidance == "orbital_insertion" else None)

        # Advanced per-stage pitch: deep-copy the booster and stamp each
        # stage object with the values from the inline rows.
        if (self._adv_pitch_var.get()
                and guidance in ("pitch_program", "orbital_insertion")
                and self._stage_rows):
            booster = copy.deepcopy(booster)
            node = booster
            for row in self._stage_rows:
                if node is None:
                    break
                try:
                    node.stage_turn_start_s      = float(row['start'].get())
                    node.stage_turn_stop_s        = float(row['stop'].get())
                    node.stage_burnout_angle_deg  = float(row['angle'].get())
                except ValueError:
                    pass  # leave existing/None values if field is blank
                coast_s = row.get('coast', tk.StringVar()).get().strip()
                if coast_s:
                    try:
                        node.coast_time_s = float(coast_s)
                    except ValueError:
                        pass
                # Per-stage engine cutoff — liquid only; blank clears it.
                cut_s = row.get('cutoff', tk.StringVar()).get().strip()
                if cut_s and not node.solid_motor:
                    try:
                        node.stage_cutoff_s = self._field_float(
                            f"Stage cutoff (s)", cut_s)
                    except ValueError:
                        node.stage_cutoff_s = None
                else:
                    node.stage_cutoff_s = None
                node = node.stage2

        # Global yaw program (checkbox + three fields)
        def _fon(sv):
            try:
                s = sv.get().strip()
                return float(s) if s else None
            except Exception:
                return None
        _yaw_chk = getattr(self, '_adv_yaw_var', None)
        yaw_enabled = (bool(_yaw_chk and _yaw_chk.get())
                       and guidance in ("pitch_program", "orbital_insertion"))
        yaw_maneuvers = []
        if yaw_enabled:
            for _yvars in self._yaw_vars:
                fa = _fon(_yvars['final_az'])
                if fa is not None:
                    yaw_maneuvers.append((_fon(_yvars['start']),
                                          _fon(_yvars['stop']),
                                          fa))

        try:
            launch_elevation_deg = float(self._launch_el_var.get())
        except (ValueError, AttributeError):
            launch_elevation_deg = 90.0
        booster.launch_elevation_deg = launch_elevation_deg

        # Fairing jettison is a flight-plan choice: stamp the panel value onto
        # a shrouded booster (blank = 0 = heating-flux default).  Left untouched
        # for boosters with no shroud, where the value is inert anyway.
        if getattr(booster, 'shroud_mass_kg', 0.0) > 0:
            _sj = self._shroud_jett_var.get().strip()
            try:
                booster.shroud_jettison_alt_km = float(_sj) if _sj else 0.0
            except ValueError:
                pass

        # Glider / HGV mission control — write the sidebar strip's plan fields
        # (glide law, dive, banks, ζ, dive-target, SEPARATION) onto the RAW
        # reentry object in the stack.  Hardware (glider_LD capability, β,
        # TPS, mass) is owned by the object editor and untouched.  Stamp the
        # raw ro, NOT effective_ro's fusion: separation is flippable from the
        # strip now, and stamping the fused body (stage mass/geometry already
        # inherited) would let a body→separating flip fly the whole stage's
        # burnout mass as if it were the separated object.  integrate_trajectory
        # calls effective_ro itself, after this plan is applied.
        _g_node = booster
        _g_raw = None
        while _g_node is not None:
            if _g_node.ro is not None:
                _g_raw = _g_node.ro
                break
            _g_node = getattr(_g_node, 'stage2', None)
        if _g_raw is not None:
            import dataclasses as _dc
            booster = copy.deepcopy(booster)
            _g_new_ro = _dc.replace(_g_raw, **self._reentry_plan_kwargs())
            # Write back into wherever the ro lives in the (copied) stack
            _g_node = booster
            _g_saved = False
            while _g_node is not None:
                if _g_node.ro is not None:
                    _g_node.ro = _g_new_ro
                    _g_saved = True
                    break
                _g_node = _g_node.stage2
            if not _g_saved:
                booster.ro = _g_new_ro

        return (booster, guidance, lat, lon, az, cutoff, la,
                gt_start_s, gt_stop_s, target_orbit_km,
                yaw_maneuvers, launch_elevation_deg)

    def _open_sweep(self, param=None, heating=False):
        """Open the Parametric Sweep.  Optional param preselects the swept
        variable (e.g. 'Burnout Angle') and heating=True checks the flux/load
        Show boxes — used by the Reentry Survivability tab's loft/depress
        shortcut so the heating sweep is one click from where it's read."""
        dlg = ParametricSweepDialog(self)
        if param and param in dlg._PARAM_INFO:
            dlg._param_var.set(param)
            dlg._on_param_changed()
        if heating:
            dlg._show_qpeak.set(True)
            dlg._show_load.set(True)
        return dlg

    def _open_footprint(self):
        FootprintDialog(self)

    def _open_mass_estimator(self):
        MassEstimatorDialog(self)

    def _open_screening_envelope(self):
        ScreeningEnvelopeDialog(self)

    def _set_model_option(self, key, value):
        """Switch the empirical source behind a model term (Analysis ▸ Reference
        Data) and prompt a re-run, since results depend on it."""
        mm.set_model_option(key, value)
        label = mm.MODEL_OPTIONS[key]["labels"].get(value, value)
        self._status_var.set(
            f"{mm.MODEL_OPTIONS[key]['label']} source → {label}. Re-run to apply.")

    def _manage_gazetteer(self):
        """Analysis ▸ Reference Data ▸ Offline Gazetteer… — build (or
        rebuild) the bundled worldwide place index on demand.  The build
        is O(minutes) and writes a ~2.5 GB cache, so it is an explicit,
        opt-in action here rather than a surprise inside Find Location."""
        import gazetteer as _gz
        if not _gz.available():
            messagebox.showinfo(
                "Offline Gazetteer",
                "No gazetteer packs are bundled in this build "
                "(data/gazetteer/).  Find Location will use online "
                "search (Nominatim) instead.", parent=self)
            return
        ready = _gz.index_ready()
        try:
            import gazetteer_build as _gb   # noqa: F401  (presence check)
        except Exception:
            pass
        n = _gz.feature_count()
        if ready:
            if not messagebox.askyesno(
                    "Offline Gazetteer",
                    f"The offline index is already built and current "
                    f"({n:,} named places, worldwide).\n\n"
                    "Rebuild it from the bundled packs?  This takes a few "
                    "minutes and rewrites a ~2.5 GB cache.", parent=self):
                return
            try:
                _gz._DB_PATH.unlink()
            except OSError:
                pass
        else:
            if not messagebox.askyesno(
                    "Offline Gazetteer",
                    f"Build the offline place index now?\n\n"
                    f"{n:,} named places worldwide (official BGN sources: "
                    "NGA GNS + USGS GNIS + BGN Antarctic), searchable by "
                    "every name variant with no network.\n\n"
                    "One-time build: a few minutes, ~2.5 GB cache in "
                    "~/.gui_missile_flyout.  Until it is built, Find "
                    "Location uses online search.", parent=self):
                return
        prog = tk.Toplevel(self)
        prog.title("Building offline gazetteer")
        prog.resizable(False, False)
        # No grab_set(): a modal grab here once contradicted the label's
        # "keep working" promise (user report 2026-08-18).  The window
        # stays open, the app stays interactive — though the build is
        # CPU-heavy, so say so instead of overpromising.
        ttk.Label(prog, padding=16, justify=tk.LEFT,
                  text="Building the worldwide place index…\n"
                  "This takes several minutes, and the app may be "
                  "sluggish meanwhile.\n"
                  "The window closes itself when done.").pack()
        pb = ttk.Progressbar(prog, mode="indeterminate", length=320)
        pb.pack(padx=16, pady=(0, 16)); pb.start(12)
        done = [None]

        def _bg():
            try:
                db = _gz.ensure_index()
                db.close()
                done[0] = True
            except Exception as ex:       # noqa: BLE001
                done[0] = ex

        def _poll():
            if done[0] is None:
                prog.after(300, _poll); return
            pb.stop(); prog.grab_release(); prog.destroy()
            if done[0] is True:
                self._status_var.set(
                    f"Offline gazetteer ready — {n:,} places searchable in "
                    "Find Location.")
            else:
                messagebox.showerror("Offline Gazetteer",
                                     f"Build failed:\n{done[0]}", parent=self)
        threading.Thread(target=_bg, daemon=True).start()
        prog.after(300, _poll)

    def _open_damping_estimator(self):
        DampingEstimatorDialog(self)

    def _mk_hint(self, parent, text):
        """A greyed hint label that wraps to the control-panel width, so its
        text is always fully visible (registered for dynamic re-wrapping)."""
        lbl = ttk.Label(parent, text=text, foreground="#555555", justify=tk.LEFT)
        try:
            lbl.configure(wraplength=getattr(self, "_left_wrap", 460))
        except tk.TclError:
            pass
        if hasattr(self, "_left_hints"):
            self._left_hints.append(lbl)
        return lbl

    def _open_range_ring(self):
        RangeRingDialog(self)

    def _run_flyout(self):
        if self._running:
            return
        # Write-through FIRST: the plan you fly is the plan on disk.  The
        # sidebar is a live view of the active flight plan, so every run
        # persists it, and _get_inputs below builds the booster from the
        # just-persisted file — file and flight can never disagree (e.g.
        # Simple mode clears per-stage angles before they could fly).
        self._snapshot_traj_profile(self._booster_var.get())
        self._snapshot_reentry_plan()
        try:
            (booster, guidance, lat, lon, az, cutoff, la,
             gt_start_s, gt_stop_s, target_orbit_km,
             yaw_maneuvers, launch_elevation_deg) = self._get_inputs()
        except ValueError as e:
            messagebox.showerror("Input error", str(e))
            return
        self._running = True
        self._status_var.set("Running simulation…")
        threading.Thread(
            target=self._run_thread,
            args=(booster, guidance, lat, lon, az, cutoff, la,
                  gt_start_s, gt_stop_s, target_orbit_km,
                  yaw_maneuvers, launch_elevation_deg, False),
            daemon=True,
        ).start()

    def _cancel_max_range(self):
        self._cancel_event.set()
        self._cancel_max_btn.config(state=tk.DISABLED)
        self._status_var.set("Cancelling…")

    def _maximize_range(self):
        if self._running:
            return
        # Max Range optimises the SIMPLE (global) pitch profile — burnout angle
        # and turn-stop.  Per-stage overrides mask those globals in the guidance
        # law, so on an advanced-pitch plan the sweep would optimise knobs that
        # don't fly.  Refuse rather than report a meaningless optimum.
        if self._adv_pitch_var.get():
            messagebox.showinfo(
                "Max Range",
                "Max Range optimises the simple pitch profile (burnout angle "
                "and turn-stop).\n\nThe active flight plan uses Advanced pitch, "
                "whose per-stage angles override those globals, so the sweep "
                "cannot improve it.\n\nSwitch Mode to “Simple pitch profile” to "
                "run Max Range.", parent=self)
            return
        # Orbital-insertion plans have a goal (a target orbit), not a range to
        # maximise — that's Plan Orbit's job.
        if self._guidance_var.get() == "orbital_insertion":
            messagebox.showinfo(
                "Max Range",
                "The active flight plan is an orbital-insertion plan; use "
                "Plan Orbit to solve it for a target orbit.\n\nTo find the "
                "booster's maximum range, switch to (or create) a pitch-program "
                "or gravity-turn flight plan.", parent=self)
            return
        # Write-through before optimising: panel tweaks are persisted to the
        # active plan (same as Run / Plan Orbit) so nothing is lost when the
        # dropdown switches to the max-range variant afterwards.
        self._snapshot_traj_profile(self._booster_var.get())
        self._snapshot_reentry_plan()
        try:
            (booster, guidance, lat, lon, az, cutoff, la,
             gt_start_s, gt_stop_s, target_orbit_km,
             yaw_maneuvers, launch_elevation_deg) = self._get_inputs()
        except ValueError as e:
            messagebox.showerror("Input error", str(e))
            return
        # Max Range does NOT mutate the active plan.  It optimises, then writes
        # the result to the reserved "max-range" variant and switches to it, so
        # the loaded plan is preserved one click away.  Record the base plan and
        # the context the optimum is valid for (stamped into the variant notes).
        _active = self._active_plan_name()
        self._max_range_base_plan = self._resolve_generator_base(_active)
        self._max_range_on_reserved = _active == mm.MAX_RANGE_PLAN_LABEL
        self._max_range_base_law = self._guidance_var.get()   # law carries to the variant
        self._max_range_context = (
            f"site={self._site_var.get() or f'{lat:.3f},{lon:.3f}'}, "
            f"az={az:.1f}°, reentry object={self._ro_main_var.get() or '(booster default)'}")
        # Turn stop is autopopulated with the booster burn time, but until the
        # user changes or saves it we treat it as unset and let the optimiser
        # pick the turn-stop.  A user-changed/saved value is honoured verbatim —
        # except when re-running on the max-range variant itself, where we
        # always re-optimise both knobs (the variant carries the last optimum).
        if self._max_range_on_reserved:
            opt_stop = None
        else:
            _auto = getattr(self, '_gt_turn_stop_auto', None)
            _untouched = _auto is not None and self._gt_turn_stop_var.get().strip() == _auto
            opt_stop = None if _untouched else gt_stop_s
        self._max_range_pending = True
        self._cancel_event.clear()
        self._running = True
        self._cancel_max_btn.config(state=tk.NORMAL)
        self._status_var.set("Optimising for maximum range…")
        threading.Thread(
            target=self._run_thread,
            args=(booster, guidance, lat, lon, az, cutoff, la,
                  gt_start_s, opt_stop, target_orbit_km,
                  yaw_maneuvers, launch_elevation_deg, True),
            daemon=True,
        ).start()

    def _write_max_range_variant(self, r):
        """Save a Max Range result to the reserved 'max-range' plan variant and
        switch to it, leaving the base plan the user loaded untouched.

        The variant is a simple-profile plan (the optimiser's knobs are the
        global burnout angle and turn-stop); it starts from the base plan so
        non-guidance fields (fairing, events, etc.) carry over, then stamps the
        optimum and the launch context the optimum is valid for.
        """
        name = self._booster_var.get()
        base = getattr(self, '_max_range_base_plan', None)
        if base in (mm.MAX_RANGE_PLAN_LABEL, mm.ORBITAL_PLAN_LABEL):
            base = None   # never rebase a generated variant onto itself
        # Base plan fully merged (bundled + user default, then the base variant).
        plan = dict(self._raw_active_plan(name, plan_name=base, use_active=False))
        plan['stages'] = [dict(s) for s in plan.get('stages', [])]
        # Remember the plan this optimum was generated FROM, so a re-run while
        # the variant is active rebases onto it instead of onto (default).
        if base:
            plan['base_plan'] = base
        else:
            plan.pop('base_plan', None)
        # Simple profile under the base plan's own law (pitch program or
        # gravity turn — the optimiser swept that law's global knobs).
        plan['guidance'] = getattr(self, '_max_range_base_law', 'pitch_program')
        plan['adv_pitch_on'] = False
        if r.get('optimal_burnout_angle_deg') is not None:
            plan['burnout_angle_deg'] = float(r['optimal_burnout_angle_deg'])
        if r.get('optimal_gt_turn_stop_s') is not None:
            plan['gt_turn_stop_s'] = float(r['optimal_gt_turn_stop_s'])
        # Clear per-stage angle overrides so the global profile governs.
        for st in plan['stages']:
            st['stage_burnout_angle_deg'] = None
        plan['source'] = 'Auto-generated by Max Range'
        plan['notes'] = (
            f"Optimum for {getattr(self, '_max_range_context', 'the last run')}. "
            f"Regenerated on every Max Range run; the optimum shifts with launch "
            f"site, azimuth, and reentry object.")
        try:
            save_flight_plan(name, plan, _FLIGHT_PLAN_LIBRARY_PATH,
                             plan=mm.MAX_RANGE_PLAN_LABEL)
        except Exception as exc:
            self._status_var.set(f"Max Range: could not save variant: {exc}")
            return
        mm.ACTIVE_FLIGHT_PLANS[name] = mm.MAX_RANGE_PLAN_LABEL
        self._save_active_plans()
        self._refresh_flight_plan_list(select=mm.MAX_RANGE_PLAN_LABEL)
        self._on_booster_changed()

    def _plan_orbit(self):
        """Handler for the Plan Orbit button.

        Same generator-not-editor contract as Max Range: the solved two-phase
        boost program is written to the reserved 'orbital' plan variant and the
        dropdown switches to it — the plan the user loaded is never mutated.
        """
        if self._running:
            return
        self._snapshot_traj_profile(self._booster_var.get())   # write-through first
        self._snapshot_reentry_plan()
        try:
            (booster, guidance, lat, lon, az, cutoff, la,
             gt_start_s, gt_stop_s, target_orbit_km,
             _yaw_maneuvers, _launch_el) = self._get_inputs()
        except ValueError as e:
            messagebox.showerror("Input error", str(e))
            return
        if target_orbit_km is None:
            messagebox.showerror("Input error",
                                 "Enter a target orbit altitude (km) first.")
            return
        self._plan_orbit_base_plan = self._resolve_generator_base(
            self._active_plan_name())
        self._plan_orbit_context = (
            f"target {target_orbit_km:g} km orbit, "
            f"site={self._site_var.get() or f'{lat:.3f},{lon:.3f}'}, az={az:.1f}°")
        self._running = True
        self._status_var.set(
            f"Planning orbital trajectory to {target_orbit_km:.0f} km…")
        threading.Thread(
            target=self._plan_orbit_thread,
            args=(booster, lat, lon, az, target_orbit_km, gt_start_s),
            daemon=True,
        ).start()

    def _write_orbital_variant(self, boost_angle, turn_stop, gt_start_s,
                               target_orbit_km, perigee_km, apogee_km):
        """Save a Plan Orbit solution to the reserved 'orbital' plan variant
        and switch to it (mirror of _write_max_range_variant)."""
        name = self._booster_var.get()
        base = getattr(self, '_plan_orbit_base_plan', None)
        if base in (mm.MAX_RANGE_PLAN_LABEL, mm.ORBITAL_PLAN_LABEL):
            base = None   # never rebase a generated variant onto itself
        plan = dict(self._raw_active_plan(name, plan_name=base, use_active=False))
        plan['stages'] = [dict(s) for s in plan.get('stages', [])]
        if base:
            plan['base_plan'] = base
        else:
            plan.pop('base_plan', None)
        plan['guidance'] = 'orbital_insertion'
        plan['adv_pitch_on'] = False
        plan['burnout_angle_deg'] = float(boost_angle)
        plan['gt_turn_start_s'] = float(gt_start_s)
        plan['gt_turn_stop_s'] = float(turn_stop)
        plan['target_orbit_km'] = float(target_orbit_km)
        # The two-phase program is fully described by the globals above; clear
        # per-stage pitch overrides carried over from the base plan.
        for st in plan['stages']:
            st['stage_burnout_angle_deg'] = None
        plan['source'] = 'Auto-generated by Plan Orbit'
        plan['notes'] = (
            f"Solution for {getattr(self, '_plan_orbit_context', 'the last run')} "
            f"→ {perigee_km:.0f}×{apogee_km:.0f} km. Regenerated on every "
            f"Plan Orbit run; the solution shifts with launch site, azimuth, "
            f"and target altitude.")
        try:
            save_flight_plan(name, plan, _FLIGHT_PLAN_LIBRARY_PATH,
                             plan=mm.ORBITAL_PLAN_LABEL)
        except Exception as exc:
            self._status_var.set(f"Plan Orbit: could not save variant: {exc}")
            return
        mm.ACTIVE_FLIGHT_PLANS[name] = mm.ORBITAL_PLAN_LABEL
        self._save_active_plans()
        self._refresh_flight_plan_list(select=mm.ORBITAL_PLAN_LABEL)
        self._on_booster_changed()

    def _plan_orbit_thread(self, booster, lat, lon, az,
                           target_orbit_km, gt_start_s):
        """Worker: runs plan_orbital_insertion then fires the full simulation."""
        try:
            plan = plan_orbital_insertion(
                booster, lat, lon, az, target_orbit_km,
                gt_turn_start_s=gt_start_s)
        except Exception as e:
            self._running = False
            self.after(0, lambda: messagebox.showerror(
                "Planner error", str(e)))
            return

        if not plan['success']:
            self._running = False
            self.after(0, lambda: messagebox.showerror(
                "No solution", plan['message']))
            return

        boost_angle = plan['boost_angle_deg']
        turn_stop   = plan['turn_stop_s']

        # On the main thread: persist the solution to the reserved 'orbital'
        # variant (switching the dropdown to it — the loaded plan is untouched),
        # then fly it.
        def _apply_and_run():
            self._write_orbital_variant(boost_angle, turn_stop, gt_start_s,
                                        target_orbit_km,
                                        plan['perigee_km'], plan['apogee_km'])
            self._status_var.set(
                f"Plan found: boost {boost_angle:.0f}°  →  "
                f"{plan['perigee_km']:.0f}×{plan['apogee_km']:.0f} km  "
                f"— running simulation…")

            # The panel now shows the orbital variant; fly exactly that.
            try:
                (m_run, guidance_run, lat_run, lon_run, az_run,
                 cutoff_run, la_run,
                 gts_run, gtstp_run, orb_run,
                 yaw_maneuvers_run, el_run) = self._get_inputs()
            except ValueError:
                # Fallback: use the solved parameters directly.
                m_run, guidance_run = booster, "orbital_insertion"
                lat_run, lon_run, az_run = lat, lon, az
                cutoff_run, la_run = None, boost_angle
                gts_run, gtstp_run, orb_run = gt_start_s, turn_stop, target_orbit_km
                yaw_maneuvers_run = []
                el_run = 90.0

            # _run_thread checks self._running; it's still True from _plan_orbit
            threading.Thread(
                target=self._run_thread,
                args=(m_run, guidance_run, lat_run, lon_run, az_run,
                      cutoff_run, la_run,
                      gts_run, gtstp_run, orb_run,
                      yaw_maneuvers_run, el_run, False),
                daemon=True,
            ).start()

        self.after(0, _apply_and_run)

    def _run_thread(self, booster, guidance, lat, lon, az, cutoff, la,
                    gt_start_s, gt_stop_s, target_orbit_km,
                    yaw_maneuvers, launch_elevation_deg, maximise):
        q_str = self._query_alt_km_var.get().strip()
        q_alt = float(q_str) if (self._query_alt_enable.get() and q_str) else None
        try:
            if maximise:
                result = maximize_range(booster, lat, lon, az, guidance=guidance,
                                        cutoff_time_s=cutoff,
                                        gt_turn_start_s=gt_start_s,
                                        gt_turn_stop_s=gt_stop_s,
                                        reentry_query_alt_km=q_alt,
                                        cancel_event=self._cancel_event,
                                        terrain_dem=self._terrain_dem_value(),
                                        launch_elev_m=self._launch_elev_value())
            else:
                # Orbital insertion trajectories can have very long flight
                # times: a highly elliptical transfer orbit peaks at thousands
                # of km and takes 90+ minutes to come back down.  Use 3 hours
                # so the integrator always reaches the ground.
                _max_t = 10800.0 if guidance == "orbital_insertion" else 3600.0
                result = integrate_trajectory(
                    booster, lat, lon, az,
                    guidance=guidance,
                    burnout_angle_deg=la,
                    cutoff_time_s=cutoff,
                    gt_turn_start_s=gt_start_s,
                    gt_turn_stop_s=gt_stop_s,
                    reentry_query_alt_km=q_alt,
                    target_orbit_alt_km=target_orbit_km,
                    yaw_maneuvers=yaw_maneuvers,
                    launch_elevation_deg=launch_elevation_deg,
                    alpha_limit_deg=self._alpha_limit_value(),
                    alpha_induced_drag=self._alpha_induced_value(),
                    terrain_dem=self._terrain_dem_value(),
                    launch_elev_m=self._launch_elev_value(),
                    max_time_s=_max_t)
            self._result = result
            self.after(0, self._on_result_ready)
        except MaxRangeCancelled:
            self.after(0, lambda: self._status_var.set("Max Range cancelled."))
        except Exception as e:
            _err_msg = str(e)
            import traceback as _tb
            _tb.print_exc()
            self.after(0, lambda m=_err_msg: messagebox.showerror("Simulation error", m))
        finally:
            self._running = False
            self.after(0, lambda: self._cancel_max_btn.config(state=tk.DISABLED))

    # ------------------------------------------------------------------
    # Display results
    # ------------------------------------------------------------------
    def _on_result_ready(self):
        r = self._result

        # If this was a Max Range run, persist the optimum to the reserved
        # "max-range" flight-plan variant and switch to it — the loaded plan is
        # left untouched, one dropdown click away.  _write_max_range_variant
        # repopulates the panel from the saved variant, so the fields update
        # too.  (Non-Max-Range runs never set the flag.)
        if getattr(self, '_max_range_pending', False):
            self._max_range_pending = False
            self._write_max_range_variant(r)

        orbital   = r.get('orbital', False)
        rng_km    = r['range_km']
        rng_nm    = rng_km / 1.852   if rng_km    is not None else None
        rng_mi    = rng_km / 1.60934 if rng_km    is not None else None
        apogee_km = r['apogee_km']

        tof_s       = r['time_of_flight_s']
        imp_spd_kms = r['impact_speed_ms'] / 1000.0 if r['impact_speed_ms'] is not None else None
        apo_lat     = r.get('apogee_lat_deg')
        apo_lon     = r.get('apogee_lon_deg')

        units = self._units_var.get()
        scale_map = {"km": (1.0, "km"), "nm": (1/1.852, "nmi"), "mi": (1/1.60934, "mi")}
        scale, ulbl = scale_map[units]

        oe = r.get('orbital_elements')
        _oe_str = ""
        if oe:
            _oe_str = (f"  |  {oe['perigee_km']:.0f}×{oe['apogee_km']:.0f} km"
                       f"  i={oe['inclination_deg']:.1f}°"
                       f"  e={oe['eccentricity']:.4f}"
                       f"  T={oe['period_min']:.1f} min")

        if orbital and r.get('max_range_km') is None:
            _strip = (f"No sub-orbital solution — exceeds orbital velocity.  "
                      f"Apogee: {apogee_km*scale:.1f} {ulbl}")
            self._status_var.set("Max Range: " + _strip)
        elif orbital:
            _strip = (f"In orbit.  Apogee: {apogee_km*scale:.1f} {ulbl}" + _oe_str)
            self._status_var.set(_strip)
        else:
            _spd_str = f"{imp_spd_kms:.2f} km/s" if imp_spd_kms is not None else "—"
            _strip = (f"Range: {rng_km*scale:.1f} {ulbl}  |  "
                      f"Apogee: {apogee_km*scale:.1f} {ulbl}  |  "
                      f"ToF: {tof_s:.0f} s  |  "
                      f"Impact: {r['impact_lat']:.2f}°N, {r['impact_lon']:.2f}°E  |  "
                      f"Impact spd: {_spd_str}")
            self._status_var.set("Done.  " + _strip)
        # Surface the static-margin / trim-gate verdict when it changed the
        # reentry: an unstable body was flipped to a tumbling (ballistic)
        # descent, or a stable body was L/D-limited by its control authority.
        _tg = r.get('reentry_trim')
        if _tg and ('UNSTABLE' in _tg['verdict'] or 'CONTROL-LIMITED' in _tg['verdict']):
            _sm = _tg['static_margin_cal']
            self._status_var.set(
                self._status_var.get()
                + f"   ⚠ reentry: SM {_sm:+.1f} cal — {_tg['verdict']}")
        self._results_strip_var.set(_strip)
        # Rendering runs here (scheduled via after(), OUTSIDE the flyout's
        # try/except).  Guard it: an unhandled exception here fires after
        # _plot_results has already cla()'d every axis but before its final
        # canvas.draw(), so it would silently blank every plot and leave the
        # previous run's image on screen — which reads as "all plots broken /
        # every booster shows the same hardcoded plot".  On failure, force a
        # draw so partial plots show, and surface the traceback instead of
        # hiding it in the terminal.
        try:
            self._plot_results(r, scale, ulbl)
            self._populate_timeline(r)
            self._populate_survivability(r)
        except Exception as exc:
            import traceback as _tb
            _tb_str = _tb.format_exc()
            _tb.print_exc()
            try:
                self._canvas.draw()
            except Exception:
                pass
            self._status_var.set(f"Plot error: {exc}")
            messagebox.showerror(
                "Plot error",
                f"The trajectory ran, but rendering the plots failed:\n\n{exc}"
                f"\n\n{_tb_str}",
                parent=self)

    # ------------------------------------------------------------------
    def _populate_timeline(self, r):
        """Fill the Flight Timeline tab from the milestones list."""
        # Clear existing rows
        self._tl_tree.delete(*self._tl_tree.get_children())

        rng_km    = r['range_km']
        apogee_km = r['apogee_km']
        tof_s     = r['time_of_flight_s']
        _orbital  = r.get('orbital', False)
        if rng_km is not None:
            rng_nm = rng_km / 1.852
            rng_mi = rng_km / 1.60934
        imp_spd   = r['impact_speed_ms'] / 1000.0 if r['impact_speed_ms'] is not None else None

        if _orbital:
            oe = r.get('orbital_elements')
            _oe_line = ""
            if oe:
                _oe_line = (f"\nOrbit: {oe['perigee_km']:.0f}×{oe['apogee_km']:.0f} km"
                            f"   i={oe['inclination_deg']:.1f}°"
                            f"   e={oe['eccentricity']:.4f}"
                            f"   T={oe['period_min']:.1f} min")
            _apo_loc = (f"{r['apogee_lat_deg']:.2f}°N  {r['apogee_lon_deg']:.2f}°E"
                        if r.get('apogee_lat_deg') is not None else "—")
            self._tl_summary_var.set(
                f"In orbit — no ground impact within integration window\n"
                f"Apogee: {apogee_km:.1f} km   "
                f"Apogee loc: {_apo_loc}"
                + _oe_line
            )
        else:
            _apo_loc = (f"{r['apogee_lat_deg']:.2f}°N  {r['apogee_lon_deg']:.2f}°E"
                        if r.get('apogee_lat_deg') is not None else "—")
            self._tl_summary_var.set(
                f"Range: {rng_km:.1f} km  /  {rng_nm:.1f} nmi  /  {rng_mi:.1f} mi\n"
                f"Apogee: {apogee_km:.1f} km   "
                f"Apogee loc: {_apo_loc}\n"
                f"Impact: {r['impact_lat']:.2f}°N  {r['impact_lon']:.2f}°E   "
                f"Flight time: {tof_s:.0f} s   "
                f"Impact speed: {f'{imp_spd:.2f} km/s' if imp_spd is not None else '—'}"
            )

        # Key events highlighted differently; debris impact rows get their own tag
        _key_prefixes = ("Ignition", "Apogee", "Perigee", "Impact", "Orbital insertion")

        for idx, m in enumerate(r.get('milestones', [])):
            if m.get('is_debris'):
                tag = "debris"
            elif m['event'].startswith(_key_prefixes):
                tag = "key"
            else:
                tag = "odd" if idx % 2 else "even"
            # Acceleration at Impact is dominated by drag spike — show as blank
            accel_str = (f"{m['accel_ms2']:+.1f}"
                         if not m['event'].startswith("Impact") else "—")
            self._tl_tree.insert("", tk.END, tags=(tag,), values=(
                m['event'],
                f"{m['t_s']:.1f}",
                f"{m['alt_km']:.1f}",
                f"{m['range_km']:.1f}",
                f"{m['speed_kms']:.3f}",
                f"{m['inertial_speed_kms']:.3f}",
                accel_str,
                f"{m['mass_t']:.3f}",
            ))

    def _plot_results(self, r, scale, ulbl):
        t   = np.asarray(r['t'])
        alt = np.asarray(r['alt']) / 1000.0 * scale
        spd = np.asarray(r['speed']) / 1000.0   # always km/s
        rng = np.asarray(r['range']) / 1000.0 * scale
        lat_arr = np.asarray(r['lat'])
        lon_arr = np.asarray(r['lon'])
        orbital = r.get('orbital', False)

        for ax in (self._ax_alt, self._ax_spd, self._ax_spd_twin,
                   self._ax_traj, self._ax_trk,
                   self._ax_guid, self._ax_guid_twin,
                   self._ax_qmach, self._ax_qmach_twin):
            ax.cla()
            ax.grid(True, alpha=0.35)
            ax.tick_params(labelsize=7)

        # cla() resets twinx positioning — restore right-side y-axis on all twin axes.
        # Also suppress their redundant gridlines (primary axes already provide them).
        for _twin in (self._ax_spd_twin, self._ax_guid_twin, self._ax_qmach_twin):
            _twin.yaxis.tick_right()
            _twin.yaxis.set_label_position('right')
            _twin.grid(False)

        # ── Find key event times for orbital trajectories ────────────
        _ins_t = _apo_t = _peri_t = None
        if orbital:
            for ms in r.get('milestones', []):
                ev = ms.get('event', '').lower()
                if 'orbital insertion' in ev and _ins_t is None:
                    _ins_t = ms['t_s']
                elif ev.startswith('apogee') and _apo_t is None:
                    _apo_t = ms['t_s']
                elif ev.startswith('perigee') and _peri_t is None:
                    _peri_t = ms['t_s']
            # Circular-orbit fallback: no perigee found → show one full period
            if _peri_t is None and _ins_t is not None:
                oe = r.get('orbital_elements')
                if oe:
                    _peri_t = _ins_t + oe['period_min'] * 60.0

        # Array indices for truncation
        _ins_idx  = (int(np.searchsorted(t, _ins_t))
                     if _ins_t is not None else len(t) - 1)
        _peri_idx = (int(np.searchsorted(t, _peri_t))
                     if _peri_t is not None else len(t) - 1)
        # Clamp to valid range
        _ins_idx  = min(_ins_idx,  len(t) - 1)
        _peri_idx = min(_peri_idx, len(t) - 1)

        # ── Altitude vs Time (truncate at insertion for orbital) ──────
        _sl = slice(0, _ins_idx + 1) if orbital else slice(None)
        self._ax_alt.plot(t[_sl], alt[_sl], color='royalblue', linewidth=1.5)
        self._ax_alt.set_xlabel("Time (s)", fontsize=8)
        self._ax_alt.set_ylabel(f"Altitude ({ulbl})", fontsize=8)
        self._ax_alt.set_title("Altitude vs Time", fontsize=9)
        self._ax_alt.fill_between(t[_sl], 0, alt[_sl],
                                  alpha=0.12, color='royalblue')

        # ── Speed vs Time ─────────────────────────────────────────────
        from atmosphere import atmosphere as _atm_fn
        self._ax_spd.plot(t, spd, color='firebrick', linewidth=1.5, label='Speed')
        self._ax_spd.set_xlabel("Time (s)", fontsize=8)
        self._ax_spd.set_ylabel("Speed (km/s)", fontsize=8)
        self._ax_spd.set_title("Speed vs Time", fontsize=9)
        # Altitude-corrected Mach on twin axis
        _alt_m_s = np.asarray(r.get('alt', []))
        _spd_ms  = np.asarray(r['speed'])
        _mach_s  = np.full(len(_alt_m_s), np.nan)
        for _i, _h in enumerate(_alt_m_s):
            _, _, _, _snd = _atm_fn(float(_h))
            if _snd > 10.0:          # NaN above ~86 km where atmosphere model → 0
                _mach_s[_i] = _spd_ms[_i] / _snd
        _ax_m = self._ax_spd_twin
        _ax_m.plot(t, _mach_s, color='steelblue', linewidth=1.2, ls='--', label='Mach')
        _ax_m.set_ylabel("Mach", fontsize=8, color='steelblue')
        _ax_m.tick_params(labelsize=7, colors='steelblue')
        _ax_m.yaxis.set_major_locator(matplotlib.ticker.MaxNLocator(integer=True))

        # ── Altitude vs Range (truncate at insertion for orbital) ─────
        self._ax_traj.plot(rng[_sl], alt[_sl], color='seagreen', linewidth=1.5)
        self._ax_traj.set_xlabel(f"Downrange ({ulbl})", fontsize=8)
        self._ax_traj.set_ylabel(f"Altitude ({ulbl})", fontsize=8)
        self._ax_traj.set_title("Altitude vs Range", fontsize=9)
        self._ax_traj.fill_between(rng[_sl], 0, alt[_sl],
                                   alpha=0.12, color='seagreen')

        # ── Ground Track (truncate at perigee / one orbit for orbital) ─
        center_lon = float(lon_arr[0])          # launch meridian as origin

        # Truncate ground-track arrays
        _trk_sl = slice(0, _peri_idx + 1) if orbital else slice(None)
        lon_trk = lon_arr[_trk_sl]
        lat_trk = lat_arr[_trk_sl]
        lon_c   = ((lon_trk - center_lon + 180.0) % 360.0) - 180.0

        # NaN-break any residual jumps > 180° (multi-hemisphere trajectories)
        lon_c = list(lon_c)
        lat_c = list(lat_trk)
        i = 1
        while i < len(lon_c):
            if abs(lon_c[i] - lon_c[i - 1]) > 180:
                lon_c.insert(i, np.nan)
                lat_c.insert(i, np.nan)
                i += 2
            else:
                i += 1

        self._ax_trk.plot(lon_c, lat_c, color='black', linewidth=1.2, zorder=2)
        self._ax_trk.plot(0.0, lat_arr[0], 'go', markersize=7,
                          label="Launch", zorder=5)

        if not orbital:
            impact_lon_c = ((lon_arr[-1] - center_lon + 180.0) % 360.0) - 180.0
            self._ax_trk.plot(impact_lon_c, lat_arr[-1], 'r*', markersize=9,
                              label="Impact", zorder=5)

        # Orbital event markers (insertion ◆, apogee ▲, perigee ▼)
        if orbital:
            for _t_ev, mkr, col, lbl in [
                (_ins_t,  'D', '#003580', 'Insertion'),
                (_apo_t,  '^', '#6600bb', 'Apogee'),
                (_peri_t, 'v', '#006655', 'Perigee'),
            ]:
                if _t_ev is None or _t_ev > t[-1]:
                    continue
                _ev_lat = float(np.interp(_t_ev, t, lat_arr))
                _ev_lon = float(np.interp(_t_ev, t, lon_arr))
                _ev_lon_c = ((_ev_lon - center_lon + 180.0) % 360.0) - 180.0
                self._ax_trk.plot(_ev_lon_c, _ev_lat, mkr, color=col,
                                  markersize=7, label=lbl, zorder=6)

        # Debris impact locations — red crosses, one per shed stage / fairing.
        _debris_plotted = False
        for m in r.get('milestones', []):
            if not m.get('is_debris'):
                continue
            d_lat = m.get('impact_lat')
            d_lon = m.get('impact_lon')
            if d_lat is None or d_lon is None:
                continue
            d_lon_c = ((d_lon - center_lon + 180.0) % 360.0) - 180.0
            self._ax_trk.plot(d_lon_c, d_lat, 'rx', markersize=8,
                              markeredgewidth=1.8,
                              label="Debris" if not _debris_plotted else "_nolegend_",
                              zorder=5)
            _debris_plotted = True

        # Capture the trajectory-fitted limits, draw borders, then restore so
        # the world-spanning border lines cannot expand the view.
        # Add 20% padding on each side beyond matplotlib's default 5% margin.
        self._ax_trk.autoscale()
        xlo, xhi = self._ax_trk.get_xlim()
        ylo, yhi = self._ax_trk.get_ylim()
        xpad = (xhi - xlo) * 0.20
        ypad = (yhi - ylo) * 0.20
        _draw_borders(self._ax_trk, center_lon)
        self._ax_trk.set_xlim(xlo - xpad, xhi + xpad)
        self._ax_trk.set_ylim(ylo - ypad, yhi + ypad)

        # Tick labels show absolute longitudes (convert back from centred frame)
        self._ax_trk.xaxis.set_major_formatter(matplotlib.ticker.FuncFormatter(
            lambda v, _: f"{((v + center_lon + 180) % 360) - 180:.0f}°"))
        self._ax_trk.set_xlabel("Longitude (°E)", fontsize=8)
        self._ax_trk.set_ylabel("Latitude (°N)", fontsize=8)
        self._ax_trk.set_title("Ground Track", fontsize=9)
        self._ax_trk.legend(fontsize=7)

        # ── Pitch, Azimuth vs. Time ───────────────────────────────────
        ax_g  = self._ax_guid
        ax_g2 = self._ax_guid_twin   # reuse pre-created twin (avoids stacking)
        t_plot = np.asarray(r.get('t', []))
        pc     = np.asarray(r.get('pitch_cmd_deg', []))
        ac     = np.asarray(r.get('az_cmd_deg', []))

        if len(t_plot) > 0 and len(pc) == len(t_plot):
            ax_g.plot(t_plot, pc, color='royalblue', lw=1.4, label='Pitch (°)')
            # Boost angle of attack — thrust axis vs velocity (SP-8099).  Same
            # degree axis as pitch; NaN outside powered flight so it gaps.
            # Solid = FLOWN α (after the q·α clamp); dashed = COMMANDED α (the
            # raw maneuver demand).  They coincide unless the limit engaged, in
            # which case the gap between them is exactly what the clamp held
            # back — "demanded X°, flew Y°".
            _al_g  = np.asarray(r.get('alpha_deg', []))
            _alc_g = np.asarray(r.get('alpha_cmd_deg', []))
            if len(_al_g) == len(t_plot) and np.any(np.isfinite(_al_g)):
                _al_eng = r.get('alpha_limit_engaged')
                # Draw the commanded (demand) trace only when it diverges from
                # the flown one — i.e. the clamp actually bit — so an
                # unclamped run keeps a single clean α line.
                if (_al_eng and len(_alc_g) == len(t_plot)
                        and np.any(np.isfinite(_alc_g)
                                   & (_alc_g > _al_g + 0.5))):
                    ax_g.plot(t_plot, _alc_g, color='forestgreen', lw=1.0,
                              ls='--', alpha=0.55, label='α cmd (°)')
                # α is green — distinct from pitch (blue) and azimuth (orange),
                # and clearly on the left (Elevation / α) axis.
                ax_g.plot(t_plot, _al_g, color='forestgreen', lw=1.1,
                          label='α flown (°)' if _al_eng else 'α (°)')
            if len(ac) == len(t_plot):
                ax_g2.plot(t_plot, ac, color='darkorange', lw=1.4,
                           ls='--', label='Azimuth (°)')
                # Adaptive ticks.  A fixed 5° step (MultipleLocator(5)) crams 30+
                # overlapping labels onto the axis whenever the heading sweeps a
                # wide range — e.g. long-range / near-polar flights where the
                # great-circle azimuth changes by >100°.  MaxNLocator keeps ~6
                # readable ticks at any range.  Limits are set from the finite
                # samples so the NaN gaps (coast/glide, where no azimuth is
                # commanded) don't break autoscaling.
                _ac_fin = ac[np.isfinite(ac)]
                if _ac_fin.size:
                    _lo, _hi = float(_ac_fin.min()), float(_ac_fin.max())
                    if _hi - _lo < 1.0:                 # near-constant heading
                        _lo, _hi = _lo - 2.0, _hi + 2.0
                    _pad = 0.05 * (_hi - _lo)
                    ax_g2.set_ylim(_lo - _pad, _hi + _pad)
                ax_g2.yaxis.set_major_locator(
                    matplotlib.ticker.MaxNLocator(nbins=6,
                                                  steps=[1, 2, 2.5, 5, 10]))
                ax_g2.yaxis.set_label_position('right')
                ax_g2.yaxis.set_ticks_position('right')
                ax_g2.set_ylabel('Azimuth (°)', fontsize=7, color='darkorange')
                ax_g2.tick_params(labelsize=7, colors='darkorange')
                # Combined legend on the primary axis
                _l1, _lb1 = ax_g.get_legend_handles_labels()
                _l2, _lb2 = ax_g2.get_legend_handles_labels()
                ax_g.legend(_l1 + _l2, _lb1 + _lb2,
                             fontsize=7, loc='upper right')
            # Stage separation and yaw event lines
            for ms in r.get('milestones', []):
                _ev = ms.get('event', '').lower()
                _t  = ms.get('t_s', None)
                if _t is None:
                    continue
                if 'burnout' in _ev or 'ignition' in _ev:
                    ax_g.axvline(_t, color='#aaaaaa', lw=0.8, ls=':')
        ax_g.set_xlabel('Time (s)', fontsize=7)
        # The left axis carries BOTH the pitch (elevation) and α traces — they
        # are angles in the same degrees — so label it for both and keep the
        # ticks a neutral colour rather than pitch-blue.
        ax_g.set_ylabel('Elevation / α (°)', fontsize=7, color='#333333')
        ax_g.tick_params(labelsize=7, colors='#333333')
        ax_g.set_title('Pitch, α, Azimuth vs. Time', fontsize=8)
        ax_g.grid(True, alpha=0.35)

        # ── Dyn. Pressure & Mach (burn period only) ──────────────────
        from atmosphere import atmosphere as _atm
        _alt_m  = np.asarray(r.get('alt', []))
        _vel_ec = np.asarray(r.get('vel_ecef', []))
        _t_aero = np.asarray(r['t'])

        if len(_alt_m) > 1 and _vel_ec.ndim == 2 and len(_vel_ec) == len(_alt_m):
            _spd_ms = np.asarray(r['speed'])
            _rho    = np.empty(len(_alt_m))
            _sound  = np.empty(len(_alt_m))
            for _i, _h in enumerate(_alt_m):
                _, _, _rho[_i], _sound[_i] = _atm(float(_h))
            _q_kpa = 0.5 * _rho * _spd_ms**2 / 1e3
            _mach  = _spd_ms / np.where(_sound > 0, _sound, 1.0)

            # Restrict to burn period (t ≤ last burnout milestone)
            _ms = r.get('milestones', [])
            _bo_times = [float(m['t_s']) for m in _ms
                         if any(k in m.get('event', '').lower()
                                for k in ('burnout', 'cutoff', 'burn out'))]
            _t_cutoff = max(_bo_times) if _bo_times else float(_t_aero[-1])
            _mask = _t_aero <= _t_cutoff
            _tb   = _t_aero[_mask]
            _qb   = _q_kpa[_mask]
            _mb   = _mach[_mask]

            ax_qm  = self._ax_qmach
            ax_mch = self._ax_qmach_twin
            ax_qm.fill_between(_tb, _qb, alpha=0.18, color='steelblue')
            ax_qm.plot(_tb, _qb, color='steelblue', lw=1.3, label='q (kPa)')
            ax_mch.plot(_tb, _mb, color='darkorange', lw=1.2, ls='--', label='Mach')
            # Annotate max-q only.  The q·α peak (the SP-8099 combined-load
            # metric) is reported in the Flight Timeline as a "Max q·α"
            # milestone rather than on this plot — it is proportional to the
            # lateral-g readout below, so annotating both here was redundant.
            _qmax_i = int(np.argmax(_qb))
            ax_qm.axvline(_tb[_qmax_i], color='steelblue', lw=0.8, ls=':', alpha=0.7)
            ax_qm.annotate(
                f"max-q\n{_qb[_qmax_i]:.1f} kPa\nM {_mb[_qmax_i]:.1f}",
                xy=(_tb[_qmax_i], _qb[_qmax_i]),
                xytext=(6, -4), textcoords='offset points',
                fontsize=6, color='steelblue', va='top')
            # Peak lateral load factor (applied aero side load) — the metric
            # payload user's guides tabulate; a small corner readout.
            _lg = np.asarray(r.get('lateral_g', []))
            if len(_lg) == len(_t_aero):
                _lg_b = _lg[_mask]
                _lg_f = _lg_b[np.isfinite(_lg_b)]
                if _lg_f.size and float(_lg_f.max()) > 0.0:
                    ax_qm.text(
                        0.02, 0.97, f"peak lateral {_lg_f.max():.2f} g",
                        transform=ax_qm.transAxes, fontsize=6, color='crimson',
                        va='top', ha='left')
            _l1, _lb1 = ax_qm.get_legend_handles_labels()
            _l2, _lb2 = ax_mch.get_legend_handles_labels()
            ax_qm.legend(_l1 + _l2, _lb1 + _lb2, fontsize=6, loc='upper right')
            ax_qm.set_xlabel('Time (s)', fontsize=7)
            ax_qm.set_ylabel('q  (kPa)', fontsize=7, color='steelblue')
            ax_qm.tick_params(labelsize=7, colors='steelblue')
            ax_qm.set_title('Dyn. Pressure, Mach vs. Time', fontsize=8)
            ax_qm.grid(True, alpha=0.35)
            ax_mch.set_ylabel('Mach', fontsize=7, color='darkorange')
            ax_mch.tick_params(labelsize=7, colors='darkorange')
            ax_mch.yaxis.set_label_position('right')
            ax_mch.yaxis.set_ticks_position('right')

        self._canvas.draw()

    # ------------------------------------------------------------------
    # File / Help actions
    # ------------------------------------------------------------------
    def _export_figures(self):
        """Save the trajectory plots to PNG, PDF, or SVG."""
        import datetime as _dt
        from tkinter.filedialog import asksaveasfilename
        ts      = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        booster = _safe_name(self._booster_var.get())
        rng_km  = self._result.get('range_km') if self._result else None
        rng_sfx = f"_{rng_km:.0f}km" if rng_km is not None else ""
        path = asksaveasfilename(
            defaultextension=".png",
            initialdir=str(_ensure_dir(_DIR_PLOTS)),
            initialfile=f"{ts}_{booster}{rng_sfx}.figures.png",
            filetypes=[
                ("PNG image",    "*.png"),
                ("PDF document", "*.pdf"),
                ("SVG vector",   "*.svg"),
                ("All files",    "*.*"),
            ],
            title="Export figures",
        )
        if not path:
            return
        self._fig.savefig(path, dpi=150, bbox_inches="tight")
        self._status_var.set(f"Figures exported: {path}")

    # ------------------------------------------------------------------
    def _trajectory_metadata(self):
        """Return a dict of all guidance/launch settings for CSV header embedding."""
        _ro_name = (self._ro_main_var.get()
                    if hasattr(self, '_ro_main_var') else '')
        _site_name = (self._site_var.get()
                      if (hasattr(self, '_site_var')
                          and self._site_var.get() in getattr(self, '_site_map', {}))
                      else '')
        meta = {
            'booster':              self._booster_var.get(),
            'ro':                   _ro_name,
            'site_name':            _site_name,
            'launch_lat':           self._launch_lat.get(),
            'launch_lon':           self._launch_lon.get(),
            'azimuth_deg':          self._azimuth_var.get(),
            'guidance':        self._guidance_var.get(),
            'burnout_angle_deg': self._loft_angle_var.get(),
            'gt_turn_start_s': self._gt_turn_start_var.get(),
            'gt_turn_stop_s':       self._gt_turn_stop_var.get(),
            'cutoff_s':             self._cutoff_var.get(),
            'launch_elevation_deg': getattr(self, '_launch_el_var',
                                            tk.StringVar(value='90')).get(),
            'adv_pitch':            self._adv_pitch_var.get(),
            'adv_yaw':              self._adv_yaw_var.get(),
            'yaw_maneuvers': [
                {'start':    v['start'].get(),
                 'stop':     v['stop'].get(),
                 'final_az': v['final_az'].get()}
                for v in self._yaw_vars
            ],
            'glider_guid': getattr(self, '_main_guidance_var', tk.StringVar(value='')).get(),
            'glider_skip_count': getattr(self, '_main_skip_count_var', tk.StringVar(value='1')).get(),
            'glider_dive_alt': getattr(self, '_main_dive_alt_var', tk.StringVar(value='0')).get(),
            'glider_pullup_alt': getattr(self, '_main_pullup_alt_var', tk.StringVar(value='0')).get(),
            'glider_bank_on':  getattr(self, '_main_bank_sched_var', tk.BooleanVar()).get(),
            'glider_banks': [
                {'start': v['start'].get(), 'end': v['end'].get(), 'bank': v['bank'].get()}
                for v in getattr(self, '_main_bank_vars', [])
            ],
            'glider_aero': getattr(self, '_main_aero_var',
                                    tk.StringVar(value='Drag polar (realistic)')).get(),
            'glider_dive_target_on': getattr(self, '_main_dive_target_var',
                                             tk.BooleanVar()).get(),
            'glider_dt_lat':    getattr(self, '_main_dt_lat_var',    tk.StringVar(value='0.0')).get(),
            'glider_dt_lon':    getattr(self, '_main_dt_lon_var',    tk.StringVar(value='0.0')).get(),
            'glider_dt_radius': getattr(self, '_main_dt_radius_var', tk.StringVar(value='20')).get(),
        }
        # Per-stage pitch / yaw overrides
        if self._adv_pitch_var.get() and self._stage_rows:
            meta['stage_overrides'] = [
                {
                    'start': row['start'].get(),
                    'stop':  row['stop'].get(),
                    'angle': row['angle'].get(),
                    'coast': row.get('coast', tk.StringVar()).get(),
                }
                for row in self._stage_rows
            ]
        return meta

    def _apply_trajectory_metadata(self, meta):
        """Restore GUI fields from a metadata dict loaded from a CSV header."""
        name = meta.get('booster', meta.get('missile', ''))
        if name and name in BOOSTER_DB:
            self._booster_var.set(name)
            self._on_booster_changed()
        elif name:
            # Don't silently leave the previously-selected booster in place —
            # the rest of the scenario would then be applied to the wrong
            # booster and "produce no (or wrong) results".  Say so plainly.
            messagebox.showwarning(
                "Load scenario",
                f"Booster '{name}' is not available, so the scenario was applied "
                f"to whatever booster is currently selected.\n\n"
                f"It is not a shipped booster and is not in your saved boosters "
                f"(~/.gui_missile_flyout/custom_boosters.json).  Load or recreate "
                f"'{name}', select it, then reload the scenario.",
                parent=self)
        # RV selection (added with the scenario schema; absent in older files)
        if hasattr(self, '_ro_main_var'):
            _ro_name = meta.get('ro', '')
            if _ro_name in RO_DB:
                self._ro_main_var.set(_ro_name)
            else:
                self._ro_main_var.set(self._RO_DEFAULT_SENTINEL)
            self._on_ro_selected_main()
        # Launch coordinates are authoritative; site_name is for display only.
        self._launch_lat.set(meta.get('launch_lat', ''))
        self._launch_lon.set(meta.get('launch_lon', ''))
        if hasattr(self, '_site_var'):
            _site_name = meta.get('site_name', '')
            if _site_name and _site_name in getattr(self, '_site_map', {}):
                self._site_var.set(_site_name)
            else:
                self._site_var.set('')
        self._azimuth_var.set(meta.get('azimuth_deg', '0.0'))
        guidance = meta.get('guidance', 'pitch_program')
        if guidance == 'gravity_turn':       # legacy key from pre-rename saves
            guidance = 'pitch_program'
        self._guidance_var.set(guidance)
        self._update_guidance_labels(guidance)
        self._loft_angle_var.set(meta.get('burnout_angle_deg', '45.0'))
        self._gt_turn_start_var.set(meta.get('gt_turn_start_s', '5.0'))
        self._gt_turn_stop_var.set(meta.get('gt_turn_stop_s', ''))
        self._cutoff_var.set(meta.get('cutoff_s', ''))
        if hasattr(self, '_launch_el_var'):
            self._launch_el_var.set(meta.get('launch_elevation_deg', '90.0'))
        self._adv_yaw_var.set(bool(meta.get('adv_yaw', False)))
        saved_yaw = meta.get('yaw_maneuvers', [])
        # Back-compat: old single-maneuver keys
        if not saved_yaw and meta.get('yaw_final_az_deg', ''):
            saved_yaw = [{'start': meta.get('yaw_start_s', ''),
                          'stop':  meta.get('yaw_stop_s', ''),
                          'final_az': meta.get('yaw_final_az_deg', '')}]
        for _i, _yvars in enumerate(self._yaw_vars):
            _d = saved_yaw[_i] if _i < len(saved_yaw) else {}
            _yvars['start'].set(_d.get('start', ''))
            _yvars['stop'].set(_d.get('stop', ''))
            _yvars['final_az'].set(_d.get('final_az', ''))
        # Per-stage overrides — expand the panel then fill row by row
        adv = bool(meta.get('adv_pitch', False))
        self._adv_pitch_var.set(adv)
        self._on_adv_pitch_toggled()
        self._sync_ascent_mode_display(guidance, adv)
        overrides = meta.get('stage_overrides', [])
        if adv and overrides and self._stage_rows:
            for row, ov in zip(self._stage_rows, overrides):
                row['start'].set(ov.get('start', ''))
                row['stop'].set(ov.get('stop', ''))
                row['angle'].set(ov.get('angle', ''))
                if 'coast' in row:
                    row['coast'].set(ov.get('coast', ''))
        if hasattr(self, '_main_guidance_var'):
            self._main_guidance_var.set(meta.get('glider_guid', 'Equilibrium glide (Tracy)'))
            if hasattr(self, '_main_skip_count_var'):
                self._main_skip_count_var.set(str(meta.get('glider_skip_count', '1')))
            self._main_dive_alt_var.set(meta.get('glider_dive_alt', '0'))
            if hasattr(self, '_main_pullup_alt_var'):
                self._main_pullup_alt_var.set(meta.get('glider_pullup_alt', '0'))
            self._main_bank_sched_var.set(bool(meta.get('glider_bank_on', False)))
            saved_banks = meta.get('glider_banks', [])
            for _i, _bvars in enumerate(self._main_bank_vars):
                _bd = saved_banks[_i] if _i < len(saved_banks) else {}
                _bvars['start'].set(_bd.get('start', ''))
                _bvars['end'].set(_bd.get('end', ''))
                _bvars['bank'].set(_bd.get('bank', ''))
            if hasattr(self, '_main_aero_var'):
                # Normalize legacy labels ("Constant L/D (trim)" /
                # "Slender-body polar") onto the current combobox values.
                _aero_meta = str(meta.get('glider_aero', 'Drag polar (realistic)'))
                self._main_aero_var.set(
                    "Drag polar (realistic)" if "polar" in _aero_meta.lower()
                    else "Fixed L/D (idealized)")
            if hasattr(self, '_main_dive_target_var'):
                self._main_dive_target_var.set(
                    bool(meta.get('glider_dive_target_on', False)))
                self._main_dt_lat_var.set(meta.get('glider_dt_lat', '0.0'))
                self._main_dt_lon_var.set(meta.get('glider_dt_lon', '0.0'))
                self._main_dt_radius_var.set(meta.get('glider_dt_radius', '20'))
            if hasattr(self, '_glider_status_var'):
                self._refresh_glider_status_line()
            self._on_main_bank_toggled()
            self._on_main_dive_target_toggled()
            self._on_glider_guidance_changed()

    def _save_trajectory(self):
        if self._result is None:
            messagebox.showinfo("No data", "Run a simulation first.")
            return
        import datetime as _dt
        from tkinter.filedialog import asksaveasfilename
        ts      = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        booster = _safe_name(self._booster_var.get())
        rng_km  = self._result.get('range_km')
        rng_sfx = f"_{rng_km:.0f}km" if rng_km is not None else ""
        path = asksaveasfilename(
            parent=self,
            defaultextension=".csv",
            initialdir=str(_ensure_dir(_DIR_TRAJECTORIES)),
            initialfile=f"{ts}_{booster}{rng_sfx}.traj.csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            title="Export Trajectory",
        )
        if not path:
            return
        try:
            r = self._result
            rows = [
                self._CSV_SCENARIO_PREFIX + json.dumps(self._scenario_dict()),
                "piece,time_s,lat_deg,lon_deg,alt_m,speed_ms,range_km",
            ]
            for i, ti in enumerate(r['t']):
                rows.append(f"primary,{ti:.3f},{r['lat'][i]:.6f},{r['lon'][i]:.6f},"
                            f"{r['alt'][i]:.1f},{r['speed'][i]:.2f},{r['range'][i]/1000.0:.3f}")
            for d in r.get('debris_trajectories', []):
                label = str(d.get('label', 'debris')).replace(',', ' ')
                for i, ti in enumerate(d['t']):
                    rows.append(f"{label},{ti:.3f},{d['lat'][i]:.6f},{d['lon'][i]:.6f},"
                                f"{d['alt'][i]:.1f},,")
            with open(path, "w", encoding="utf-8") as fh:
                fh.write("\n".join(rows) + "\n")
        except Exception as exc:
            import traceback
            messagebox.showerror(
                "Export Trajectory",
                f"Could not export trajectory:\n{exc}\n\n{traceback.format_exc()}",
                parent=self)
            return
        self._status_var.set(f"Trajectory CSV exported: {path}")

    def _export_trajectory_xlsx(self):
        """Export the trajectory time-series to an XLSX workbook."""
        if self._result is None:
            messagebox.showinfo("No data", "Run a simulation first.")
            return
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            messagebox.showerror("Missing dependency",
                                 f"openpyxl is required:\n{exc}")
            return
        import datetime as _dt
        from tkinter.filedialog import asksaveasfilename
        ts      = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        booster = _safe_name(self._booster_var.get())
        rng_km  = self._result.get('range_km')
        rng_sfx = f"_{rng_km:.0f}km" if rng_km is not None else ""
        path = asksaveasfilename(
            defaultextension=".xlsx",
            initialdir=str(_ensure_dir(_DIR_TRAJECTORIES)),
            initialfile=f"{ts}_{booster}{rng_sfx}.traj.xlsx",
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
            title="Export Trajectory XLSX",
        )
        if not path:
            return
        r = self._result
        wb = Workbook()
        ws = wb.active
        ws.title = "Trajectory"
        ws.append(["piece", "time_s", "lat_deg", "lon_deg",
                   "alt_m", "speed_ms", "range_km"])
        for i, ti in enumerate(r['t']):
            ws.append(["primary", float(ti),
                       float(r['lat'][i]), float(r['lon'][i]),
                       float(r['alt'][i]), float(r['speed'][i]),
                       float(r['range'][i]) / 1000.0])
        for d in r.get('debris_trajectories', []):
            label = d['label']
            for i, ti in enumerate(d['t']):
                ws.append([label, float(ti),
                           float(d['lat'][i]), float(d['lon'][i]),
                           float(d['alt'][i]), None, None])
        scn = wb.create_sheet(self._XLSX_SCENARIO_SHEET)
        scn.append(["key", "value"])
        for k, v in self._scenario_dict().items():
            scn.append([k, json.dumps(v) if isinstance(v, (list, dict)) else v])
        wb.save(path)
        self._status_var.set(f"Trajectory XLSX exported: {path}")

    def _export_kml(self):
        """Export the ground track and 3-D trajectory path as a KML file."""
        if self._result is None:
            messagebox.showinfo("No data", "Run a simulation first.")
            return
        import datetime as _dt
        from tkinter.filedialog import asksaveasfilename
        ts      = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        booster = _safe_name(self._booster_var.get())
        rng_km  = self._result.get('range_km')
        rng_sfx = f"_{rng_km:.0f}km" if rng_km is not None else ""
        path = asksaveasfilename(
            defaultextension=".kml",
            initialdir=str(_ensure_dir(_DIR_TRAJECTORIES)),
            initialfile=f"{ts}_{booster}{rng_sfx}.traj.kml",
            filetypes=[("KML files", "*.kml"), ("All files", "*.*")],
            title="Export trajectory KML",
        )
        if not path:
            return

        r   = self._result
        lat = np.asarray(r['lat'])
        lon = np.asarray(r['lon'])
        alt = np.asarray(r['alt'])

        # 3-D trajectory (absolute altitude)
        coords_3d = " ".join(
            f"{lo:.6f},{la:.6f},{a:.1f}"
            for lo, la, a in zip(lon, lat, alt)
        )
        # Ground track (clamped to ground)
        coords_gnd = " ".join(
            f"{lo:.6f},{la:.6f},0"
            for lo, la in zip(lon, lat)
        )

        booster_name = self._booster_var.get()

        # Build debris Placemarks
        debris_placemarks = []
        for d in r.get('debris_trajectories', []):
            d_lat = np.asarray(d['lat'])
            d_lon = np.asarray(d['lon'])
            d_alt = np.asarray(d['alt'])
            label = d['label']
            c3d = " ".join(f"{lo:.6f},{la:.6f},{a:.1f}"
                           for lo, la, a in zip(d_lon, d_lat, d_alt))
            cgnd = " ".join(f"{lo:.6f},{la:.6f},0"
                            for lo, la in zip(d_lon, d_lat))
            debris_placemarks.append(f"""
    <Placemark>
      <name>{label} (3-D)</name>
      <styleUrl>#debrisTraj</styleUrl>
      <LineString>
        <altitudeMode>absolute</altitudeMode>
        <tessellate>0</tessellate>
        <coordinates>{c3d}</coordinates>
      </LineString>
    </Placemark>

    <Placemark>
      <name>{label} ground track</name>
      <styleUrl>#debrisGnd</styleUrl>
      <LineString>
        <altitudeMode>clampToGround</altitudeMode>
        <tessellate>1</tessellate>
        <coordinates>{cgnd}</coordinates>
      </LineString>
    </Placemark>

    <Placemark>
      <name>{label} impact</name>
      <Point>
        <altitudeMode>clampToGround</altitudeMode>
        <coordinates>{d_lon[-1]:.6f},{d_lat[-1]:.6f},0</coordinates>
      </Point>
    </Placemark>""")

        debris_xml = "".join(debris_placemarks)

        kml = f"""<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2">
  <Document>
    <name>{booster_name} Trajectory</name>

    <Style id="traj3d">
      <LineStyle><color>ffff0000</color><width>2</width></LineStyle>
    </Style>
    <Style id="trajGnd">
      <LineStyle><color>880000ff</color><width>1</width></LineStyle>
    </Style>
    <Style id="debrisTraj">
      <LineStyle><color>ff00aaff</color><width>1</width></LineStyle>
    </Style>
    <Style id="debrisGnd">
      <LineStyle><color>8800aaff</color><width>1</width></LineStyle>
    </Style>

    <Placemark>
      <name>Launch</name>
      <Point>
        <altitudeMode>clampToGround</altitudeMode>
        <coordinates>{lon[0]:.6f},{lat[0]:.6f},0</coordinates>
      </Point>
    </Placemark>

    <Placemark>
      <name>Impact</name>
      <Point>
        <altitudeMode>clampToGround</altitudeMode>
        <coordinates>{lon[-1]:.6f},{lat[-1]:.6f},0</coordinates>
      </Point>
    </Placemark>

    <Placemark>
      <name>Trajectory (3-D)</name>
      <styleUrl>#traj3d</styleUrl>
      <LineString>
        <altitudeMode>absolute</altitudeMode>
        <tessellate>0</tessellate>
        <coordinates>{coords_3d}</coordinates>
      </LineString>
    </Placemark>

    <Placemark>
      <name>Ground Track</name>
      <styleUrl>#trajGnd</styleUrl>
      <LineString>
        <altitudeMode>clampToGround</altitudeMode>
        <tessellate>1</tessellate>
        <coordinates>{coords_gnd}</coordinates>
      </LineString>
    </Placemark>
{debris_xml}
  </Document>
</kml>"""

        with open(path, "w", encoding="utf-8") as fh:
            fh.write(kml)
        self._status_var.set(f"KML exported: {path}")

    # ------------------------------------------------------------------
    # NOTAM overlay load / clear
    # ------------------------------------------------------------------

    def _load_notam_overlay(self):
        """Parse a KML or KMZ file and store polygon rings for Folium rendering."""
        from tkinter.filedialog import askopenfilename
        import xml.etree.ElementTree as ET
        import zipfile, io

        path = askopenfilename(
            title="Load NOTAM overlay",
            filetypes=[("KML / KMZ files", "*.kml *.kmz"), ("All files", "*.*")],
        )
        if not path:
            return

        # KMZ is a ZIP containing a .kml file.
        if path.lower().endswith(".kmz"):
            with zipfile.ZipFile(path) as zf:
                kml_names = [n for n in zf.namelist() if n.lower().endswith(".kml")]
                if not kml_names:
                    messagebox.showerror("NOTAM overlay",
                                         "No .kml file found inside the .kmz archive.")
                    return
                kml_text = zf.read(kml_names[0])
        else:
            with open(path, "rb") as fh:
                kml_text = fh.read()

        # KML uses a namespace; strip it so tag names are plain.
        kml_text = kml_text.replace(b'xmlns="http://www.opengis.net/kml/2.2"', b"")
        kml_text = kml_text.replace(b'xmlns="http://earth.google.com/kml/2.1"', b"")
        kml_text = kml_text.replace(b'xmlns="http://earth.google.com/kml/2.0"', b"")

        try:
            root = ET.fromstring(kml_text)
        except ET.ParseError as exc:
            messagebox.showerror("NOTAM overlay", f"KML parse error:\n{exc}")
            return

        polygons = []
        for poly_el in root.iter("Polygon"):
            outer = poly_el.find(".//outerBoundaryIs/LinearRing/coordinates")
            if outer is None or not outer.text:
                continue
            coords = []
            for token in outer.text.split():
                parts = token.split(",")
                if len(parts) >= 2:
                    try:
                        lon, lat = float(parts[0]), float(parts[1])
                        coords.append([lon, lat])
                    except ValueError:
                        pass
            if len(coords) >= 3:
                # GeoJSON polygon rings must be closed
                if coords[0] != coords[-1]:
                    coords.append(coords[0])
                polygons.append(coords)

        if not polygons:
            messagebox.showwarning("NOTAM overlay",
                                    "No polygon features found in the file.")
            return

        self._notam_overlay = polygons
        n = len(polygons)
        self._status_var.set(
            f"NOTAM overlay loaded: {n} polygon{'s' if n != 1 else ''} "
            f"from {Path(path).name}"
        )

    def _clear_notam_overlay(self):
        self._notam_overlay = None
        self._status_var.set("NOTAM overlay cleared.")

    # Projection catalogue used by the Cartopy export dialog.
    # Each entry: (display label, factory callable(mid_lon, mid_lat) → CRS)
    _CARTOPY_PROJECTIONS = [
        ("Orthographic (globe)",
         lambda lo, la: __import__('cartopy.crs', fromlist=['Orthographic'])
                        .Orthographic(central_longitude=lo, central_latitude=la)),
        ("Azimuthal Equidistant (true distances from centre)",
         lambda lo, la: __import__('cartopy.crs', fromlist=['AzimuthalEquidistant'])
                        .AzimuthalEquidistant(central_longitude=lo, central_latitude=la)),
        ("Lambert Conformal Conic (mid-latitude)",
         lambda lo, la: __import__('cartopy.crs', fromlist=['LambertConformal'])
                        .LambertConformal(central_longitude=lo, central_latitude=la)),
        ("Plate Carrée (equirectangular)",
         lambda lo, la: __import__('cartopy.crs', fromlist=['PlateCarree'])
                        .PlateCarree()),
        ("Mercator",
         lambda lo, la: __import__('cartopy.crs', fromlist=['Mercator'])
                        .Mercator()),
        ("Robinson (global overview)",
         lambda lo, la: __import__('cartopy.crs', fromlist=['Robinson'])
                        .Robinson(central_longitude=lo)),
        ("Equal Earth",
         lambda lo, la: __import__('cartopy.crs', fromlist=['EqualEarth'])
                        .EqualEarth(central_longitude=lo)),
        ("North Polar Stereographic",
         lambda lo, la: __import__('cartopy.crs', fromlist=['NorthPolarStereo'])
                        .NorthPolarStereo(central_longitude=lo)),
        ("South Polar Stereographic",
         lambda lo, la: __import__('cartopy.crs', fromlist=['SouthPolarStereo'])
                        .SouthPolarStereo(central_longitude=lo)),
    ]

    def _pick_cartopy_projection(self, mid_lon, mid_lat):
        """Modal dialog to choose a Cartopy projection. Returns a CRS or None."""
        dlg = tk.Toplevel(self)
        dlg.title("Choose Projection")
        dlg.resizable(False, False)
        dlg.grab_set()

        ttk.Label(dlg, text="Projection:", padding=(12, 10, 12, 4)).pack(anchor=tk.W)

        lb_frame = ttk.Frame(dlg)
        lb_frame.pack(fill=tk.BOTH, padx=12)
        vsb = ttk.Scrollbar(lb_frame, orient=tk.VERTICAL)
        lb  = tk.Listbox(lb_frame, yscrollcommand=vsb.set, activestyle="dotbox",
                         width=52, height=len(self._CARTOPY_PROJECTIONS),
                         selectmode=tk.SINGLE)
        vsb.config(command=lb.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        lb.pack(side=tk.LEFT, fill=tk.BOTH)

        for label, _ in self._CARTOPY_PROJECTIONS:
            lb.insert(tk.END, label)
        lb.selection_set(0)   # default: Orthographic

        result = [None]

        def _ok(*_):
            sel = lb.curselection()
            if sel:
                _, factory = self._CARTOPY_PROJECTIONS[sel[0]]
                result[0] = factory(mid_lon, mid_lat)
            dlg.destroy()

        lb.bind("<Double-Button-1>", _ok)

        btn_frm = ttk.Frame(dlg, padding=(12, 8))
        btn_frm.pack(fill=tk.X)
        ttk.Button(btn_frm, text="OK",     command=_ok).pack(side=tk.LEFT)
        ttk.Button(btn_frm, text="Cancel", command=dlg.destroy).pack(
            side=tk.LEFT, padx=6)

        self._center_dialog(dlg)
        self.wait_window(dlg)
        return result[0]

    def _pick_cartopy_export_options(self, mid_lon, mid_lat):
        """Combined projection + map-extent dialog.

        Returns (proj, extent_spec) on OK, or (None, None) on cancel.
        extent_spec is one of:
          None                       → global (ax.set_global)
          ('auto', pad_pct)          → auto-fit with % padding
          (lon_min, lon_max, lat_min, lat_max) → explicit bounds
        """
        dlg = tk.Toplevel(self)
        dlg.title("Cartopy Export Options")
        dlg.resizable(False, False)
        dlg.grab_set()

        # ── Projection list ───────────────────────────────────────────
        ttk.Label(dlg, text="Projection:", padding=(12, 10, 12, 4)).pack(anchor=tk.W)
        lb_frame = ttk.Frame(dlg)
        lb_frame.pack(fill=tk.BOTH, padx=12)
        vsb = ttk.Scrollbar(lb_frame, orient=tk.VERTICAL)
        lb  = tk.Listbox(lb_frame, yscrollcommand=vsb.set, activestyle="dotbox",
                         width=52, height=len(self._CARTOPY_PROJECTIONS),
                         selectmode=tk.SINGLE)
        vsb.config(command=lb.yview)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        lb.pack(side=tk.LEFT, fill=tk.BOTH)
        for _lbl, _ in self._CARTOPY_PROJECTIONS:
            lb.insert(tk.END, _lbl)
        lb.selection_set(0)

        ttk.Separator(dlg, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=12, pady=(8, 4))

        # ── Map extent ────────────────────────────────────────────────
        ttk.Label(dlg, text="Map extent:", padding=(12, 0, 12, 4)).pack(anchor=tk.W)
        extent_var = tk.StringVar(value="auto")
        ef = ttk.Frame(dlg, padding=(12, 0, 12, 4))
        ef.pack(fill=tk.X)

        ttk.Radiobutton(ef, text="Global (full world)", variable=extent_var,
                        value="global").grid(row=0, column=0, columnspan=6,
                                            sticky=tk.W, pady=2)

        # Auto-fit row
        af = ttk.Frame(ef)
        af.grid(row=1, column=0, columnspan=6, sticky=tk.W, pady=2)
        ttk.Radiobutton(af, text="Auto-fit to trajectory  —  padding:",
                        variable=extent_var, value="auto").pack(side=tk.LEFT)
        pad_var = tk.StringVar(value="25")
        ttk.Entry(af, textvariable=pad_var, width=4).pack(side=tk.LEFT, padx=2)
        ttk.Label(af, text="%").pack(side=tk.LEFT)

        ttk.Radiobutton(ef, text="Custom bounds:", variable=extent_var,
                        value="custom").grid(row=2, column=0, columnspan=6,
                                            sticky=tk.W, pady=(6, 2))

        # Custom bounds sub-grid
        cf = ttk.Frame(ef)
        cf.grid(row=3, column=0, columnspan=6, sticky=tk.W, padx=20, pady=(0, 4))
        ttk.Label(cf, text="N:").grid(row=0, column=0, padx=(0, 2))
        n_var = tk.StringVar(value="")
        ttk.Entry(cf, textvariable=n_var, width=7).grid(row=0, column=1, padx=2)
        ttk.Label(cf, text="°").grid(row=0, column=2)
        ttk.Label(cf, text="S:").grid(row=0, column=3, padx=(8, 2))
        s_var = tk.StringVar(value="")
        ttk.Entry(cf, textvariable=s_var, width=7).grid(row=0, column=4, padx=2)
        ttk.Label(cf, text="°").grid(row=0, column=5)
        ttk.Label(cf, text="W:").grid(row=1, column=0, padx=(0, 2), pady=2)
        w_var = tk.StringVar(value="")
        ttk.Entry(cf, textvariable=w_var, width=7).grid(row=1, column=1, padx=2)
        ttk.Label(cf, text="°").grid(row=1, column=2)
        ttk.Label(cf, text="E:").grid(row=1, column=3, padx=(8, 2))
        e_var = tk.StringVar(value="")
        ttk.Entry(cf, textvariable=e_var, width=7).grid(row=1, column=4, padx=2)
        ttk.Label(cf, text="°").grid(row=1, column=5)

        result = [None, None]

        def _ok(*_):
            sel = lb.curselection()
            if not sel:
                dlg.destroy()
                return
            _, factory = self._CARTOPY_PROJECTIONS[sel[0]]
            result[0] = factory(mid_lon, mid_lat)
            mode = extent_var.get()
            if mode == "global":
                result[1] = None
            elif mode == "auto":
                try:
                    pad = max(0.0, float(pad_var.get()))
                except ValueError:
                    pad = 25.0
                result[1] = ('auto', pad)
            else:
                try:
                    n = float(n_var.get())
                    s = float(s_var.get())
                    w = float(w_var.get())
                    e = float(e_var.get())
                    if s >= n or w >= e:
                        raise ValueError("degenerate bounds")
                    result[1] = (w, e, s, n)
                except ValueError:
                    messagebox.showerror(
                        "Invalid bounds",
                        "Enter numeric values where N > S and E > W.",
                        parent=dlg,
                    )
                    return
            dlg.destroy()

        lb.bind("<Double-Button-1>", _ok)
        btn_frm = ttk.Frame(dlg, padding=(12, 8))
        btn_frm.pack(fill=tk.X)
        ttk.Button(btn_frm, text="OK",     command=_ok).pack(side=tk.LEFT)
        ttk.Button(btn_frm, text="Cancel", command=dlg.destroy).pack(
            side=tk.LEFT, padx=6)

        self._center_dialog(dlg)
        self.wait_window(dlg)
        return result[0], result[1]

    def _export_cartopy(self):
        """Export a static Cartopy map of the current trajectory."""
        try:
            import cartopy.crs as ccrs
            import cartopy.feature as cfeature
            import matplotlib.patheffects as pe
            from matplotlib.backends.backend_agg import FigureCanvasAgg
        except ImportError as _e:
            messagebox.showerror(
                "Missing package",
                f"Cartopy is not installed.\n\n{_e}\n\nRun:  pip install cartopy",
            )
            return

        if self._result is None:
            messagebox.showinfo("No data", "Run a simulation first.")
            return

        r   = self._result
        lat = np.asarray(r['lat'], dtype=float)
        lon = np.asarray(r['lon'], dtype=float)
        t   = np.asarray(r['t'],   dtype=float)

        mid_lat = float(np.mean(lat))
        mid_lon = float(np.mean(lon))

        proj, extent_spec = self._pick_cartopy_export_options(mid_lon, mid_lat)
        if proj is None:
            return   # user cancelled

        import datetime as _dt
        from tkinter.filedialog import asksaveasfilename
        ts      = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        booster = _safe_name(self._booster_var.get())
        rng_km  = r.get('range_km')
        rng_sfx = f"_{rng_km:.0f}km" if rng_km is not None else ""
        path = asksaveasfilename(
            defaultextension=".png",
            initialdir=str(_ensure_dir(_DIR_MAPS)),
            initialfile=f"{ts}_{booster}{rng_sfx}.cartopy.png",
            filetypes=[("PNG image", "*.png"), ("PDF document", "*.pdf"),
                       ("SVG image", "*.svg"), ("All files", "*.*")],
            title="Save Cartopy map",
        )
        if not path:
            return

        geo = ccrs.Geodetic()

        fig    = Figure(figsize=(10, 8), dpi=300)
        canvas = FigureCanvasAgg(fig)
        ax     = fig.add_subplot(1, 1, 1, projection=proj)

        # ── Map extent ────────────────────────────────────────────────
        if extent_spec is None:
            ax.set_global()
        elif extent_spec[0] == 'auto':
            pad_frac = extent_spec[1] / 100.0
            # Include debris track points in bounding box
            _all_lat = [lat]
            _all_lon = [lon]
            for _d in r.get('debris_trajectories', []):
                _all_lat.append(np.asarray(_d['lat'], dtype=float))
                _all_lon.append(np.asarray(_d['lon'], dtype=float))
            _flat = np.concatenate(_all_lat)
            _flon = np.concatenate(_all_lon)
            lat_span = max(float(np.max(_flat) - np.min(_flat)), 2.0)
            lon_span = max(float(np.max(_flon) - np.min(_flon)), 2.0)
            ax.set_extent([
                max(-180.0, float(np.min(_flon)) - lon_span * pad_frac),
                min(+180.0, float(np.max(_flon)) + lon_span * pad_frac),
                max( -90.0, float(np.min(_flat)) - lat_span * pad_frac),
                min( +90.0, float(np.max(_flat)) + lat_span * pad_frac),
            ], crs=ccrs.PlateCarree())
        else:
            ax.set_extent(list(extent_spec), crs=ccrs.PlateCarree())

        # ── Background features ───────────────────────────────────────
        ax.add_feature(cfeature.OCEAN,     facecolor="#d6e8f5", zorder=0)
        ax.add_feature(cfeature.LAND,      facecolor="#e8e4d8", zorder=1)
        ax.add_feature(cfeature.COASTLINE, linewidth=0.5, edgecolor="#555555",
                       zorder=2)
        ax.add_feature(cfeature.BORDERS,   linewidth=0.3, edgecolor="#888888",
                       linestyle=":", zorder=2)
        ax.add_feature(cfeature.LAKES,     facecolor="#d6e8f5", linewidth=0.3,
                       edgecolor="#555555", zorder=2)
        ax.gridlines(color="white", linewidth=0.4, linestyle="--", alpha=0.6,
                     zorder=3)

        # ── NOTAM overlay ─────────────────────────────────────────────
        if self._notam_overlay:
            try:
                from shapely.geometry import Polygon as _ShapelyPoly
                from cartopy.feature import ShapelyFeature as _ShapelyFeat
                _polys = [_ShapelyPoly(ring) for ring in self._notam_overlay]
                ax.add_feature(_ShapelyFeat(
                    _polys, ccrs.PlateCarree(),
                    facecolor="#f5f5f5", edgecolor="#c0392b",
                    linewidth=1.5, alpha=0.7, zorder=3,
                ))
            except Exception:
                pass

        _OUTLINE = [pe.withStroke(linewidth=2.5, foreground="white")]

        # ── Main ground track ─────────────────────────────────────────
        _ins_t = next(
            (ms['t_s'] for ms in r.get('milestones', [])
             if 'orbital insertion' in ms.get('event', '').lower()),
            None)
        if _ins_t is not None:
            _sp = int(np.searchsorted(t, _ins_t))
            ax.plot(lon[:_sp + 1], lat[:_sp + 1], color="black",
                    linewidth=1.8, transform=geo, zorder=4,
                    path_effects=_OUTLINE)
            ax.plot(lon[_sp:], lat[_sp:], color="#555555",
                    linewidth=1.2, linestyle="--", transform=geo, zorder=4)
        else:
            ax.plot(lon, lat, color="black", linewidth=1.8,
                    transform=geo, zorder=4, path_effects=_OUTLINE)

        # ── Debris arcs ───────────────────────────────────────────────
        for d in r.get('debris_trajectories', []):
            ax.plot(np.asarray(d['lon'], dtype=float),
                    np.asarray(d['lat'], dtype=float),
                    color="black", linewidth=1.0, alpha=0.5,
                    transform=geo, zorder=4)

        # ── Milestone markers and tick marks ──────────────────────────

        def _show_labeled(e, is_debris, ms):
            if is_debris:
                return (('empty impact' in e or 'shroud impact' in e
                         or 'fairing impact' in e)   # old + new labels
                        and 'impact_lat' in ms)
            # Dots for ground events only: launch ignition and impacts.
            # "Ignition" (no boosters, t=0) and "Launch" (strap-on boosters,
            # t=0) are on the pad.  "Core ignition" (t=delay, airborne) and
            # "Stage N ignition" (in flight) get tick marks instead.
            _launch_ignition = (e == 'launch' or
                                ('ignition' in e and
                                 'stage' not in e and
                                 'core'  not in e))
            return _launch_ignition or 'impact' in e

        def _show_tick(e, is_debris):
            if is_debris:
                return False
            return not (e == 'launch' or
                        'impact' in e or
                        ('ignition' in e and
                         'stage' not in e and
                         'core'  not in e))

        def _mk_pos(ms):
            if ms.get('is_debris') and 'impact_lat' in ms:
                return ms['impact_lat'], ms['impact_lon']
            return (float(np.interp(ms['t_s'], t, lat)),
                    float(np.interp(ms['t_s'], t, lon)))

        # Tick half-length: ~0.8 % of the map's latitude span so ticks scale
        # with the map extent (same visual weight at all ranges) while staying
        # small relative to the launch/impact circles.
        try:
            _ext = ax.get_extent(crs=ccrs.PlateCarree())
            _lat_span = max(1.0, _ext[3] - _ext[2])
        except Exception:
            _lat_span = max(1.0, float(np.ptp(lat)))
        _tick_half = max(0.12, _lat_span * 0.008)

        for ms in r.get('milestones', []):
            is_debris = ms.get('is_debris', False)
            e         = ms['event'].lower()
            if _show_labeled(e, is_debris, ms):
                mk_lat, mk_lon = _mk_pos(ms)
                is_impact = 'impact' in e and not is_debris
                ax.plot(mk_lon, mk_lat, marker="o",
                        markersize=7 if is_impact else 5,
                        color="crimson" if is_impact else "white",
                        markeredgecolor="black", markeredgewidth=0.8,
                        transform=geo, zorder=6)
            elif _show_tick(e, is_debris):
                mk_lat, mk_lon = _mk_pos(ms)
                # Find nearest trajectory index to get the local tangent.
                _i = int(np.argmin(np.abs(t - ms['t_s'])))
                _i = int(np.clip(_i, 1, len(t) - 2))
                _dlat = float(lat[_i + 1] - lat[_i - 1])
                _dlon = float(lon[_i + 1] - lon[_i - 1])
                _cos  = np.cos(np.radians(mk_lat)) or 1e-9
                # Magnitude in (north, east) space
                _mag  = np.hypot(_dlat, _dlon * _cos) or 1e-9
                # Perpendicular unit vector (CW rotation): (−E, N) / mag
                _pn = -_dlon * _cos / _mag   # northward component
                _pe =  _dlat       / _mag    # eastward component
                # Offset in geographic degrees
                _dlat_t = _tick_half * _pn
                _dlon_t = _tick_half * _pe / _cos
                ax.plot([mk_lon - _dlon_t, mk_lon + _dlon_t],
                        [mk_lat - _dlat_t, mk_lat + _dlat_t],
                        color="#333333", linewidth=1.5,
                        transform=geo, zorder=6)

        # ── Title ─────────────────────────────────────────────────────
        parts = [self._booster_var.get()]
        rng = r.get('range_km')
        apo = r.get('apogee_km')
        if rng is not None: parts.append(f"Range {rng:.0f} km")
        if apo is not None: parts.append(f"Apogee {apo:.0f} km")
        ax.set_title("  ·  ".join(parts), fontsize=11, pad=8)

        fig.tight_layout()
        canvas.print_figure(path, bbox_inches="tight")
        self._status_var.set(f"Cartopy map saved: {path}")
        _open_file(path)

    def _export_folium(self):
        """Generate an interactive Folium HTML map and open it in the browser."""
        try:
            import folium
        except ImportError:
            messagebox.showerror(
                "Missing package",
                "folium is not installed.\n\nRun:  pip install folium",
            )
            return

        if self._result is None:
            messagebox.showinfo("No data", "Run a simulation first.")
            return

        import datetime as _dt
        from tkinter.filedialog import asksaveasfilename
        ts      = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        booster = _safe_name(self._booster_var.get())
        rng_km  = self._result.get('range_km')
        rng_sfx = f"_{rng_km:.0f}km" if rng_km is not None else ""
        path = asksaveasfilename(
            defaultextension=".html",
            initialdir=str(_ensure_dir(_DIR_MAPS)),
            initialfile=f"{ts}_{booster}{rng_sfx}.folium.html",
            filetypes=[("HTML files", "*.html"), ("All files", "*.*")],
            title="Save Folium map",
        )
        if not path:
            return

        r   = self._result
        lat = np.asarray(r['lat'])
        lon = np.asarray(r['lon'])
        t   = np.asarray(r['t'])
        alt = np.asarray(r.get('alt', []))   # metres; used for prefer_above

        # Unwrap longitude so the polyline never jumps across the antimeridian.
        # Values may exceed ±180°; Leaflet renders them on the correct world copy.
        _diffs = np.diff(lon)
        _diffs = (_diffs + 180.0) % 360.0 - 180.0
        lon_uw = np.empty_like(lon, dtype=float)
        lon_uw[0] = lon[0]
        lon_uw[1:] = lon[0] + np.cumsum(_diffs)

        mid_lat  = float(np.mean(lat))
        mid_lon  = float(np.mean(lon_uw))
        lon_uw_min = float(lon_uw.min())
        lon_uw_max = float(lon_uw.max())

        fmap = folium.Map(location=[mid_lat, mid_lon], zoom_start=4,
                          tiles="CartoDB positron")

        # ── NOTAM overlay (loaded via File → Load NOTAM overlay…) ─────
        if self._notam_overlay:
            _notam_geojson = {
                "type": "FeatureCollection",
                "features": [
                    {
                        "type": "Feature",
                        "geometry": {"type": "Polygon", "coordinates": [ring]},
                        "properties": {},
                    }
                    for ring in self._notam_overlay
                ],
            }
            folium.GeoJson(
                _notam_geojson,
                style_function=lambda _: {
                    "color":       "#c0392b",
                    "weight":      1.5,
                    "opacity":     0.8,
                    "fillColor":   "#f5f5f5",
                    "fillOpacity": 0.30,
                },
            ).add_to(fmap)

        # ── Ground track ──────────────────────────────────────────────
        # For orbital insertions split the track: boost phase in black,
        # orbital phase in dark grey, sharing one point at the junction.
        _ins_t = next(
            (ms['t_s'] for ms in r.get('milestones', [])
             if 'orbital insertion' in ms.get('event', '').lower()),
            None)
        if _ins_t is not None:
            _sp = int(np.searchsorted(t, _ins_t))
            folium.PolyLine(
                list(zip(lat[:_sp + 1].tolist(), lon_uw[:_sp + 1].tolist())),
                color="black", weight=2.0, opacity=0.8,
                tooltip="Boost phase",
            ).add_to(fmap)
            folium.PolyLine(
                list(zip(lat[_sp:].tolist(), lon_uw[_sp:].tolist())),
                color="#555555", weight=1.5, opacity=0.7,
                tooltip="Orbital phase",
            ).add_to(fmap)
        else:
            folium.PolyLine(
                list(zip(lat.tolist(), lon_uw.tolist())),
                color="black", weight=2.0, opacity=0.8,
                tooltip="Ground track",
            ).add_to(fmap)

        # ── Debris ground tracks ──────────────────────────────────────
        for d in r.get('debris_trajectories', []):
            _d_lon = np.asarray(d['lon'], dtype=float)
            _d_diffs = np.diff(_d_lon)
            _d_diffs = (_d_diffs + 180.0) % 360.0 - 180.0
            _d_lon_uw = np.empty_like(_d_lon)
            _d_lon_uw[0] = _d_lon[0]
            if len(_d_diffs):
                _d_lon_uw[1:] = _d_lon[0] + np.cumsum(_d_diffs)
            folium.PolyLine(
                list(zip(d['lat'].tolist(), _d_lon_uw.tolist())),
                color="black", weight=1.5, opacity=0.5,
                tooltip=d['label'],
            ).add_to(fmap)

        # ── Merge simultaneous milestones (coast_time_s == 0) ────────
        raw_milestones = r.get('milestones', [])
        merged = []
        i = 0
        while i < len(raw_milestones):
            ms = raw_milestones[i]
            group = [ms]
            if not ms.get('is_debris', False):
                j = i + 1
                while j < len(raw_milestones):
                    nxt = raw_milestones[j]
                    if (not nxt.get('is_debris', False) and
                            abs(nxt['t_s'] - ms['t_s']) < 0.1):
                        group.append(nxt)
                        j += 1
                    else:
                        break
            merged.append(group)
            i += len(group)

        def _is_ro_impact(group):
            return any('impact' in g['event'].lower() and
                       not g.get('is_debris', False) for g in group)

        merged.sort(key=lambda g: (1 if _is_ro_impact(g) else 0, g[0]['t_s']))

        # ── Circle markers + label data collection ────────────────────
        # Labeled events (filled circle + label + popup):
        #   Launch, stage empty impacts, fairing impact, warhead impact.
        # Tick-mark events (SVG perpendicular line, no circle, no label):
        #   All other non-debris flight events (apogee, re-entry, burnouts…).
        def _show_labeled(e, is_debris, ms):
            if is_debris:
                return (('empty impact' in e or 'shroud impact' in e
                         or 'fairing impact' in e)   # old + new labels
                        and 'impact_lat' in ms)
            return (('ignition' in e and 'stage' not in e) or
                    ('impact'   in e and 'empty' not in e
                                     and 'shroud' not in e
                                     and 'fairing' not in e))

        def _show_tick(e, is_debris):
            if is_debris:
                return False
            return ('apogee'       in e or
                    're-entry'     in e or
                    'burnout'      in e or
                    ('ignition' in e and 'stage' in e) or
                    'jettison'     in e or
                    'bank'         in e or
                    'pull-up start' in e or
                    'glide start'  in e or
                    'peak heating' in e or
                    'max-g'        in e or
                    'terminal dive' in e)

        def _is_major(e, is_debris):
            return not is_debris

        import re as _re_ev, json as _json

        def _name_only(raw):
            """Strip time/data parentheticals; rename Ignition → Launch."""
            name = _re_ev.sub(r'\s*\(\d[^)]*\)\s*$', '', raw).strip()
            if name.lower() == 'ignition':
                return 'Launch'
            return name

        _label_data = []   # [{lat, lon, text, t_s, prefer_above}] for JS labels
        _tick_data  = []   # [{lat, lon}] for JS tick marks

        def _prefer_above(ms, e, is_debris, mk_lat, mk_lon):
            """
            Hint for initial vertical label placement.
            Debris impacts: place the label on the SAME SIDE of the main
            trajectory as the dot, so no trajectory sits between dot and label.
            Other impact events: always above (dot at ground, label toward arc).
            Other events: above while ascending, below while descending.
            """
            if is_debris and 'impact' in e and 'impact_lat' in ms:
                # Find nearest main-trajectory point and compare latitudes.
                dlat2 = (lat - mk_lat) ** 2
                dlon2 = (lon_uw - mk_lon) ** 2
                ni    = int(np.argmin(dlat2 + dlon2))
                diff  = mk_lat - float(lat[ni])
                if abs(diff) < 1e-5:
                    return True   # essentially on the main track → default above
                return bool(diff > 0)   # north of main track → above in screen
            if 'impact' in e:
                return True
            if len(alt) < 2:
                return True
            ti  = float(ms['t_s'])
            ic  = int(np.searchsorted(t, ti))
            i0  = max(0, ic - 5)
            i1  = min(len(t) - 1, ic + 5)
            return bool(alt[i1] >= alt[i0])

        for group in merged:
            ms        = group[0]
            is_debris = ms.get('is_debris', False)
            label     = " / ".join(g['event'] for g in group)
            e         = label.lower()

            if is_debris and 'impact_lat' in ms:
                mk_lat = ms['impact_lat']
                mk_lon = ms['impact_lon']
            else:
                mk_lat = float(np.interp(ms['t_s'], t, lat))
                mk_lon = float(np.interp(ms['t_s'], t, lon_uw))

            if _show_labeled(e, is_debris, ms):
                display_name = _name_only(label)
                popup_html = (
                    f"<b>{display_name}</b><br>"
                    f"t = {ms['t_s']:.1f} s<br>"
                    f"Alt: {ms['alt_km']:.1f} km<br>"
                    f"Range: {ms['range_km']:.1f} km<br>"
                    f"Speed: {ms['speed_kms']:.2f} km/s"
                )
                popup = folium.Popup(popup_html, max_width=220)
                folium.CircleMarker(
                    [mk_lat, mk_lon], radius=5,
                    color="black", weight=1,
                    fill=True, fill_color="black", fill_opacity=1.0,
                    popup=popup, tooltip=display_name,
                ).add_to(fmap)
                _label_data.append({'lat':   mk_lat, 'lon':  mk_lon,
                                    'text':  display_name, 't_s': ms['t_s'],
                                    'major': _is_major(e, is_debris)})
            elif _show_tick(e, is_debris):
                display_name = _name_only(label)
                popup_html = (
                    f"<b>{display_name}</b><br>"
                    f"t = {ms['t_s']:.1f} s<br>"
                    f"Alt: {ms['alt_km']:.1f} km<br>"
                    f"Range: {ms['range_km']:.1f} km<br>"
                    f"Speed: {ms['speed_kms']:.2f} km/s"
                )
                folium.CircleMarker(
                    [mk_lat, mk_lon], radius=8,
                    color="black", weight=0,
                    fill=True, fill_color="black", fill_opacity=0.0,
                    opacity=0.0,
                    popup=folium.Popup(popup_html, max_width=220),
                    tooltip=display_name,
                ).add_to(fmap)
                _tick_data.append({'lat': mk_lat, 'lon': mk_lon})

        # ── Trajectory skeleton for tick-mark perpendicular computation ──
        _n_traj   = min(200, len(lat))
        _idx_traj = np.round(np.linspace(0, len(lat) - 1, _n_traj)).astype(int)
        _traj_pts = [{'lat': float(lat[i]), 'lon': float(lon_uw[i])}
                     for i in _idx_traj]
        traj_json = _json.dumps(_traj_pts)
        tick_json = _json.dumps(_tick_data)

        # All trajectory polylines (main + debris arcs) for collision detection.
        # Passed to JS so that labels are not separated from their dots by ANY arc.
        _all_traj_polys = [_traj_pts]
        for _d in r.get('debris_trajectories', []):
            _dl  = np.asarray(_d['lat'])
            _dlo = np.asarray(_d['lon'])
            _dd  = np.diff(_dlo)
            _dd  = (_dd + 180.0) % 360.0 - 180.0
            _dlu = np.empty_like(_dlo)
            _dlu[0] = _dlo[0]
            if len(_dd):
                _dlu[1:] = _dlo[0] + np.cumsum(_dd)
            _nd  = min(100, len(_dl))
            _ixd = np.round(np.linspace(0, len(_dl) - 1, _nd)).astype(int)
            _all_traj_polys.append([{'lat': float(_dl[i]), 'lon': float(_dlu[i])}
                                    for i in _ixd])
        all_traj_json = _json.dumps(_all_traj_polys)

        # ── Leader-line labels + tick marks (pure JS, update on zoom+pan) ──
        # Labels are name-only; full detail is in the click popup.
        # Tick marks are drawn perpendicular to the trajectory skeleton.
        map_var    = fmap.get_name()
        _apo_ms    = next((m for m in r.get('milestones', [])
                           if 'apogee' in m.get('event', '').lower()), None)
        apogee_t_s = float(_apo_ms['t_s']) if _apo_ms else 1e9
        label_json = _json.dumps(_label_data)
        leader_js  = f"""
        <script>
        (function() {{
            var LABELS    = {label_json};
            var APOGEE_T  = {apogee_t_s};
            var TICKS     = {tick_json};
            var TRAJ      = {traj_json};
            var ALL_TRAJ  = {all_traj_json};
            var H_GAP     = 10;   // px between dot centre and label edge
            var V_ABOVE   = 4;    // px between dot and nearest edge of label
            var STACK_GAP = 3;    // px between stacked labels
            var PAD       = 2;    // extra padding around each label box
            var TICK_HALF = 5;    // px: half-length of tick mark (≈ circle radius)
            var CLUSTER_R = 60;   // px: dot-distance threshold for new stack group

            var _svg = null, _con = null, _divs = [], _labelsLayer = null;

            function _init(map) {{
                var mc = map.getContainer();
                _svg = document.createElementNS('http://www.w3.org/2000/svg', 'svg');
                _svg.style.cssText = 'position:absolute;top:0;left:0;' +
                    'width:100%;height:100%;pointer-events:none;z-index:450;' +
                    'overflow:visible;';
                mc.appendChild(_svg);
                _con = document.createElement('div');
                _con.style.cssText = 'position:absolute;top:0;left:0;' +
                    'width:0;height:0;pointer-events:none;z-index:500;';
                mc.appendChild(_con);
                LABELS.forEach(function(lb) {{
                    var d = document.createElement('div');
                    d.style.cssText = 'position:absolute;font-size:11px;' +
                        'font-family:sans-serif;font-weight:bold;' +
                        'white-space:nowrap;padding:1px 4px;display:none;';
                    d.textContent = lb.text;
                    _con.appendChild(d);
                    _divs.push(d);
                }});

                // Dummy LayerGroup used purely as a toggle handle for the
                // Labels overlay (tick marks + leader-line labels).
                // Added to the map so it starts checked in the layer control.
                _labelsLayer = L.layerGroup().addTo(map);
                L.control.layers(
                    {{}},
                    {{'Labels': _labelsLayer}},
                    {{collapsed: false}}
                ).addTo(map);
            }}

            function _update(map) {{
                if (!_con) return;
                _svg.innerHTML = '';
                // If the Labels overlay is unchecked, hide all label divs
                // (tick SVG was already cleared above) and stop.
                if (_labelsLayer && !map.hasLayer(_labelsLayer)) {{
                    _divs.forEach(function(d) {{ d.style.display = 'none'; }});
                    return;
                }}

                var mc   = map.getContainer();
                var mapW = mc.offsetWidth  || 800;
                var mapH = mc.offsetHeight || 600;
                var EDGE = 40;

                // ── Convert trajectory skeleton to container points ───────
                var tPts = TRAJ.map(function(tp) {{
                    return map.latLngToContainerPoint([tp.lat, tp.lon]);
                }});
                // All polylines (main + debris arcs) projected for collision checks.
                var allPts = ALL_TRAJ.map(function(poly) {{
                    return poly.map(function(tp) {{
                        return map.latLngToContainerPoint([tp.lat, tp.lon]);
                    }});
                }});

                // ── Draw tick marks ───────────────────────────────────────
                TICKS.forEach(function(tk) {{
                    var tp = map.latLngToContainerPoint([tk.lat, tk.lon]);
                    // Find nearest TRAJ point by geographic distance so the
                    // result is zoom-independent (screen distances compress at
                    // low zoom, causing the wrong segment to be selected).
                    var bestD2 = Infinity, bestI = 0;
                    for (var i = 0; i < TRAJ.length; i++) {{
                        var dlat = tk.lat - TRAJ[i].lat;
                        var dlon = tk.lon - TRAJ[i].lon;
                        var d2   = dlat*dlat + dlon*dlon;
                        if (d2 < bestD2) {{ bestD2 = d2; bestI = i; }}
                    }}
                    // Project the two neighbouring geographic points to screen
                    // to get the tangent in screen space.
                    var i0  = Math.min(bestI, TRAJ.length - 2);
                    var p0  = map.latLngToContainerPoint([TRAJ[i0].lat,   TRAJ[i0].lon]);
                    var p1  = map.latLngToContainerPoint([TRAJ[i0+1].lat, TRAJ[i0+1].lon]);
                    var tdx = p1.x - p0.x;
                    var tdy = p1.y - p0.y;
                    var tlen = Math.sqrt(tdx*tdx + tdy*tdy) || 1;
                    // Perpendicular unit vector (rotated 90°).
                    var px = -tdy / tlen, py = tdx / tlen;
                    var tl = document.createElementNS(
                        'http://www.w3.org/2000/svg', 'line');
                    tl.setAttribute('x1', tp.x - px * TICK_HALF);
                    tl.setAttribute('y1', tp.y - py * TICK_HALF);
                    tl.setAttribute('x2', tp.x + px * TICK_HALF);
                    tl.setAttribute('y2', tp.y + py * TICK_HALF);
                    tl.setAttribute('stroke', '#333');
                    tl.setAttribute('stroke-width', '1.5');
                    tl.setAttribute('opacity', '0.75');
                    _svg.appendChild(tl);
                }});

                // ── Label markers ─────────────────────────────────────────
                var pts = LABELS.map(function(lb) {{
                    return map.latLngToContainerPoint([lb.lat, lb.lon]);
                }});

                // ── Collision-detection helpers ───────────────────────────
                // Liang-Barsky segment / axis-aligned rect intersection.
                function segHitsRect(x1,y1,x2,y2,rx,ry,rw,rh) {{
                    var dx=x2-x1,dy=y2-y1;
                    var p=[-dx,dx,-dy,dy],q=[x1-rx,rx+rw-x1,y1-ry,ry+rh-y1];
                    var u0=0,u1=1;
                    for(var k=0;k<4;k++){{
                        if(p[k]===0){{if(q[k]<0)return false;}}
                        else{{var u=q[k]/p[k];if(p[k]<0)u0=Math.max(u0,u);else u1=Math.min(u1,u);}}
                        if(u0>u1)return false;
                    }}
                    return true;
                }}
                // Check rect against ALL trajectory polylines (main + debris arcs).
                function labelHitsTraj(lx,ly,lw,lh) {{
                    for(var p=0;p<allPts.length;p++){{
                        var ap=allPts[p];
                        for(var i=0;i<ap.length-1;i++){{
                            if(segHitsRect(ap[i].x,ap[i].y,ap[i+1].x,ap[i+1].y,
                                           lx,ly,lw,lh)) return true;
                        }}
                    }}
                    return false;
                }}
                // Check the GAP corridor between the dot and the label rectangle.
                // A trajectory can pass through the gap without entering the label
                // rect; this catches that case.  The immediate dot vicinity (DOT_R)
                // is excluded so the trajectory passing through the dot itself does
                // not generate a false positive.
                var DOT_R=8;
                function corridorHitsTraj(pt,lx,ly,lw,lh,lRight,above){{
                    var cx,cy,cw,ch;
                    if(above){{cy=ly;ch=pt.y-DOT_R-ly;}}
                    else{{cy=pt.y+DOT_R;ch=ly+lh-(pt.y+DOT_R);}}
                    if(lRight){{cx=pt.x+DOT_R;cw=lx+lw-(pt.x+DOT_R);}}
                    else{{cx=lx;cw=pt.x-DOT_R-lx;}}
                    if(cw<=0||ch<=0)return false;
                    return labelHitsTraj(cx,cy,cw,ch);
                }}
                function rectsOverlap(ax,ay,aw,ah,bx,by,bw,bh){{
                    return ax<bx+bw&&ax+aw>bx&&ay<by+bh&&ay+ah>by;
                }}

                // Vertical side: label above dot when more trajectory points
                // are below it (more traj below → label above), and vice-versa.
                function goAbovePt(pt){{
                    var bC=0,aC=0;
                    for(var i=0;i<tPts.length;i++){{
                        var dy=tPts[i].y-pt.y;
                        if(dy>8)bC++;else if(dy<-8)aC++;
                    }}
                    return bC>=aC;
                }}

                // Local trajectory tangent at the nearest skeleton point.
                function trajTan(pt){{
                    var best=Infinity,bi=0;
                    for(var i=0;i<tPts.length;i++){{
                        var dx=tPts[i].x-pt.x,dy=tPts[i].y-pt.y;
                        var d=dx*dx+dy*dy;
                        if(d<best){{best=d;bi=i;}}
                    }}
                    var i0=Math.max(0,bi-1),i1=Math.min(tPts.length-1,bi+1);
                    return {{dx:tPts[i1].x-tPts[i0].x,dy:tPts[i1].y-tPts[i0].y}};
                }}

                // ── Label placement ───────────────────────────────────────
                // Pre-apogee labels go LEFT (stack downward from dot).
                // Post-apogee labels go RIGHT (stack upward from dot).
                // Within each side: major events nearest the dot, minor further.
                _divs.forEach(function(d){{d.style.display='none';}});

                var order=pts.map(function(_,i){{return i;}}).filter(function(i){{
                    var p=pts[i];
                    return p.x>=-EDGE&&p.x<=mapW+EDGE&&p.y>=-EDGE&&p.y<=mapH+EDGE;
                }});

                var topY={{}}, left={{}}, lw_cache={{}};

                // Split into left-side (pre-apogee) and right-side (post-apogee).
                // Within each side: major events first (nearest dot), minor second.
                var leftIdx=[], rightIdx=[];
                order.forEach(function(i){{
                    (LABELS[i].t_s < APOGEE_T ? leftIdx : rightIdx).push(i);
                }});
                function byMajorFirst(a,b){{
                    if(LABELS[b].major !== LABELS[a].major)
                        return LABELS[b].major ? 1 : -1;
                    return LABELS[a].t_s - LABELS[b].t_s;
                }}
                leftIdx.sort(byMajorFirst);
                rightIdx.sort(byMajorFirst);

                // Left-side: stack downward from dot.
                var prevBottomLeft=null, prevPtLeft=null;
                leftIdx.forEach(function(idx){{
                    var pt=pts[idx];
                    _divs[idx].style.display='block';
                    var lw=(_divs[idx].offsetWidth ||80)+PAD*2;
                    var lh=(_divs[idx].offsetHeight||14)+PAD*2;
                    lw_cache[idx]=lw;

                    if(prevPtLeft){{
                        var dx=pt.x-prevPtLeft.x, dy=pt.y-prevPtLeft.y;
                        if(Math.sqrt(dx*dx+dy*dy)>CLUSTER_R) prevBottomLeft=null;
                    }}

                    var idealTop=pt.y-lh/2;
                    var candidate=(prevBottomLeft===null)
                        ? idealTop
                        : prevBottomLeft+STACK_GAP;
                    if(candidate+lh>mapH-EDGE) candidate=mapH-EDGE-lh;

                    topY[idx]=candidate;
                    left[idx]=true;
                    prevBottomLeft=candidate+lh;
                    prevPtLeft=pt;
                }});

                // Right-side: stack upward from dot.
                var prevTopRight=null, prevPtRight=null;
                rightIdx.forEach(function(idx){{
                    var pt=pts[idx];
                    _divs[idx].style.display='block';
                    var lh=(_divs[idx].offsetHeight||14)+PAD*2;
                    var lw=(_divs[idx].offsetWidth ||80)+PAD*2;
                    lw_cache[idx]=lw;

                    if(prevPtRight){{
                        var dx=pt.x-prevPtRight.x, dy=pt.y-prevPtRight.y;
                        if(Math.sqrt(dx*dx+dy*dy)>CLUSTER_R) prevTopRight=null;
                    }}

                    var idealTop=pt.y-lh-V_ABOVE;
                    var candidate=(prevTopRight===null)
                        ? idealTop
                        : Math.min(idealTop, prevTopRight-lh-STACK_GAP);
                    if(candidate<EDGE){{ candidate=pt.y+V_ABOVE; prevTopRight=null; }}

                    topY[idx]=candidate;
                    left[idx]=false;
                    prevTopRight=candidate;
                    prevPtRight=pt;
                }});

                // Render: position each label div and draw a leader line.
                order.forEach(function(idx){{
                    var pt =pts[idx];
                    var lh =(_divs[idx].offsetHeight||14)+PAD*2;
                    var lw =lw_cache[idx]||(_divs[idx].offsetWidth||80)+PAD*2;
                    var ly =topY[idx];
                    var lx =left[idx]
                        ? pt.x-H_GAP-lw
                        : pt.x+H_GAP;
                    _divs[idx].style.left=lx+'px';
                    _divs[idx].style.top =ly+'px';

                    var line=document.createElementNS(
                        'http://www.w3.org/2000/svg','line');
                    line.setAttribute('x1',pt.x);
                    line.setAttribute('y1',pt.y);
                    line.setAttribute('x2',left[idx]?lx+lw:lx);
                    line.setAttribute('y2',ly+lh/2);
                    line.setAttribute('stroke','black');
                    line.setAttribute('stroke-width','0.7');
                    line.setAttribute('opacity','0.5');
                    _svg.appendChild(line);
                }});
            }}

            var _poll = setInterval(function() {{
                var map = window["{map_var}"];
                if (map && map.getZoom) {{
                    clearInterval(_poll);
                    _init(map);
                    // Snap back only when the centre drifts well outside the
                    // trajectory extent (±90° margin).  This lets the user pan
                    // freely across the full track without a sudden jump.
                    var SNAP_MIN = {lon_uw_min:.3f} - 90;
                    var SNAP_MAX = {lon_uw_max:.3f} + 90;
                    map.on('moveend', function() {{
                        var c = this.getCenter(), lng = c.lng;
                        if (lng < SNAP_MIN || lng > SNAP_MAX) {{
                            this.setView(
                                [c.lat, ((lng % 360) + 540) % 360 - 180],
                                this.getZoom(), {{animate: false}});
                        }}
                    }}, map);
                    map.on('moveend zoomend overlayadd overlayremove',
                           function() {{ _update(map); }});
                    _update(map);
                }}
            }}, 50);
        }})();
        </script>"""
        fmap.get_root().html.add_child(folium.Element(leader_js))
        # (Gazetteer dot/label overlays were tried here and WITHDRAWN —
        # user decision 2026-08-18: the basemap tiles already label
        # places, and the field test cluttered.  The gazetteer serves
        # the Nearby Places lookup instead.)

        # ── Logo overlay (lower-left, ~1/8–1/7 of viewport height) ──────
        import base64 as _b64, os as _os
        _logo_path = _os.path.join(_os.path.dirname(__file__), "data", "Thrusty.png")
        if _os.path.exists(_logo_path):
            with open(_logo_path, "rb") as _lf:
                _logo_b64 = _b64.b64encode(_lf.read()).decode()
            _logo_html = (
                '<img src="data:image/png;base64,' + _logo_b64 + '" '
                'style="position:fixed;bottom:12px;left:12px;'
                'height:35vh;width:auto;z-index:1000;pointer-events:none;" />'
            )
            fmap.get_root().html.add_child(folium.Element(_logo_html))

        fmap.save(path)
        import webbrowser
        webbrowser.open(f"file://{path}")
        self._status_var.set(f"Folium map saved and opened: {path}")

    def _nearby_places(self):
        """Cartography → Nearby Places…: the gazetteer as a LOOKUP, not
        a map layer (user decision 2026-08-18, after the overlay field
        test: basemap tiles already label places; what the 10.4 M-name
        index is uniquely good for is turning coordinates into words).
        For every key trajectory event — launch, stage/fairing/debris
        impacts, apogee ground point, reentry, the impact — the nearest
        POPULATED place, with distance and compass direction, as a
        table plus copyable sentences."""
        if self._result is None:
            messagebox.showinfo("No data", "Run a simulation first.")
            return
        import gazetteer as _gz
        if not (_gz.available() and _gz.index_ready()):
            messagebox.showinfo(
                "Nearby Places",
                "The offline gazetteer index is not built yet.\n\n"
                "Analysis ▸ Reference Data ▸ Offline Gazetteer… builds "
                "it once (several minutes); after that this lookup is "
                "instant.", parent=self)
            return
        r = self._result
        t = np.asarray(r['t'])
        la = np.asarray(r['lat'])
        lo = np.asarray(r['lon'])
        # unwrap for interpolation, wrap the sample back to [-180, 180]
        _d = np.diff(lo)
        _d = (_d + 180.0) % 360.0 - 180.0
        lo_uw = np.empty_like(lo, dtype=float)
        lo_uw[0] = lo[0]
        if len(_d):
            lo_uw[1:] = lo[0] + np.cumsum(_d)

        import re as _re

        def _at(ts):
            lon = float(np.interp(ts, t, lo_uw))
            return (float(np.interp(ts, t, la)),
                    ((lon + 180.0) % 360.0) - 180.0)

        def _short(e):
            return _re.sub(r'\s*\(\d[^)]*\)\s*$', '', e).strip()

        events = [("Launch", float(la[0]),
                   ((float(lo[0]) + 180.0) % 360.0) - 180.0, 0.0)]
        for ms in r.get('milestones', []):
            e, ts = ms.get('event', ''), float(ms.get('t_s', 0.0))
            el = e.lower()
            if ms.get('is_debris', False):
                if 'impact' in el and 'impact_lat' in ms:
                    events.append((_short(e), float(ms['impact_lat']),
                                   float(ms['impact_lon']), ts))
                continue
            if any(k in el for k in ('impact', 'apogee', 're-entry',
                                     'insertion')):
                ela, elo = _at(ts)
                events.append((_short(e), ela, elo, ts))

        db = _gz.ensure_index()
        dlg = tk.Toplevel(self)
        dlg.title("Nearby Places — nearest populated place and nearest "
                  "named feature per event")
        cols = ("event", "t", "coords", "place", "feat")
        tree = ttk.Treeview(dlg, columns=cols, show="headings",
                            height=min(14, max(4, len(events))))
        for c, (h, w) in zip(cols, (("Event", 165), ("t (s)", 60),
                                    ("Lat / Lon", 165),
                                    ("Nearest populated place", 290),
                                    ("Nearest named feature (any)",
                                     320))):
            tree.heading(c, text=h)
            tree.column(c, width=w, anchor=tk.W)

        def _fmt(p):
            km = (f"{p['km']:.1f}" if p['km'] < 100 else f"{p['km']:.0f}")
            return f"{p['primary']} — {km} km {p['dir']}"

        lines = []
        for name, elat, elon, ts in events:
            ns = 'N' if elat >= 0 else 'S'
            ew = 'E' if elon >= 0 else 'W'
            coord = f"{abs(elat):.4f}°{ns}  {abs(elon):.4f}°{ew}"
            pop = _gz.nearest_populated(elat, elon, n=1, db=db)
            any_ = _gz.nearest(elat, elon, n=1, db=db)
            cell_p = cell_f = "(no index hit)"
            line = f"{name} (t={ts:.0f} s, {coord}):"
            if pop:
                p = pop[0]
                cell_p = f"{_fmt(p)}  ({p['admin']})"
                line += (f" ~{p['km']:.0f} km {p['dir']} of "
                         f"{p['primary']} ({p['admin']}) [{p['ext_id']}]")
            if any_:
                f = any_[0]
                f["dir"] = _gz.compass_8(f["lat"], f["lon"], elat, elon)
                word = _gz.class_word(f["fclass"], f["source"])
                cell_f = f"{_fmt(f)}  ({word})"
                line += (f"; nearest feature ~{f['km']:.0f} km "
                         f"{f['dir']} of {f['primary']} ({word}) "
                         f"[{f['ext_id']}]")
            tree.insert("", tk.END, values=(name, f"{ts:.0f}", coord,
                                            cell_p, cell_f))
            lines.append(line)
        tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        ttk.Label(dlg, foreground="#888", justify=tk.LEFT, padding=(8, 0),
                  text="Distances are great-circle to the feature's "
                  "gazetteer coordinate; direction is the event as seen "
                  "FROM the feature.  'Any' spans every class — over "
                  "water that is the nearest seamount/trough/island; "
                  "over land often a stream or hill.  Sources: NGA GNS "
                  "/ USGS GNIS (ids in the copied text).").pack(
                      anchor=tk.W)
        bf = ttk.Frame(dlg)
        bf.pack(pady=(4, 8))

        def _copy():
            dlg.clipboard_clear()
            dlg.clipboard_append("\n".join(lines))
            self._status_var.set("Nearby-places report copied to "
                                 "clipboard")
        ttk.Button(bf, text="Copy report", command=_copy).pack(
            side=tk.LEFT, padx=4)
        ttk.Button(bf, text="Close", command=dlg.destroy).pack(
            side=tk.LEFT, padx=4)

    _GZX_COLORS = {"populated": "#4d4d4d", "facilities": "#b2182b",
                   "water": "#2166ac", "terrain": "#8c510a",
                   "undersea": "#35978f", "other": "#bbbbbb"}

    def _gazetteer_explorer(self):
        """Cartography → Gazetteer Explorer…: a zoomable working map of
        ALL 10.4 M gazetteer features (user request 2026-08-18) — a
        lookup tool, not a presentation product.  matplotlib + the
        bundled NE coastlines (no cartopy, no network).  Every zoom/pan
        re-queries the SQLite index for the viewport; above the point
        budget it draws an unbiased 1-in-k id sample and SAYS SO in the
        corner; zoomed in far enough you see literally everything, with
        names once fewer than ~150 features are in view.  Click a dot
        to identify it (class in words, source id, nearest populated
        place).  Optional ground-track overlay from the last run."""
        import gazetteer as _gz
        if not (_gz.available() and _gz.index_ready()):
            messagebox.showinfo(
                "Gazetteer Explorer",
                "The offline gazetteer index is not built yet.\n\n"
                "Analysis ▸ Reference Data ▸ Offline Gazetteer… builds "
                "it once (several minutes).", parent=self)
            return
        db = _gz.ensure_index()
        win = tk.Toplevel(self)
        win.title("Gazetteer Explorer — every named place, zoom to see "
                  "all of them")
        fig = Figure(figsize=(11, 5.6), tight_layout=True)
        ax = fig.add_subplot(111)
        ax.set_xlim(-180, 180)
        ax.set_ylim(-90, 90)
        ax.set_xlabel("Longitude (°E)", fontsize=10)
        ax.set_ylabel("Latitude (°N)", fontsize=10)
        ax.tick_params(labelsize=9)
        _draw_borders(ax, 0.0)
        canvas = FigureCanvasTkAgg(fig, master=win)
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        # Keep our own toolbar handle: a manually embedded toolbar is
        # NOT attached to canvas.toolbar, so checking that attribute in
        # the click handler raised on every click (mac field report
        # 2026-08-18 — "click doesn't work").
        toolbar = NavigationToolbar2Tk(canvas, win)
        toolbar.update()

        bar = ttk.Frame(win)
        bar.pack(fill=tk.X, padx=6)
        fam_vars = {}
        for fam in self._GZX_COLORS:
            v = tk.BooleanVar(value=True)
            fam_vars[fam] = v
            ttk.Checkbutton(bar, text=fam, variable=v,
                            command=lambda: _refresh(force=True)).pack(
                side=tk.LEFT, padx=2)
        trk_var = tk.BooleanVar(value=self._result is not None)
        if self._result is not None:
            ttk.Checkbutton(bar, text="ground track", variable=trk_var,
                            command=lambda: _refresh(force=True)).pack(
                side=tk.LEFT, padx=8)
        status = tk.StringVar(value="loading…")
        ttk.Label(win, textvariable=status, foreground="#666",
                  justify=tk.LEFT, padding=(8, 2)).pack(anchor=tk.W)

        state = {"lims": None, "busy": False, "k": 64,
                 "artists": [], "rows": []}

        def _apply(rows, k, lims):
            for a in state["artists"]:
                try:
                    a.remove()
                except Exception:
                    pass
            state["artists"], state["rows"] = [], rows
            shown = 0
            for fam in self._GZX_COLORS:
                if not fam_vars[fam].get():
                    continue
                pts = [(r["lon"], r["lat"]) for r in rows
                       if _gz.family(r["fclass"], r["source"]) == fam]
                if not pts:
                    continue
                shown += len(pts)
                state["artists"].append(ax.scatter(
                    [p[0] for p in pts], [p[1] for p in pts], s=7,
                    c=self._GZX_COLORS[fam], zorder=4, linewidths=0))
            if k == 1 and shown <= 150:
                for r in rows:
                    fam = _gz.family(r["fclass"], r["source"])
                    if fam_vars[fam].get():
                        state["artists"].append(ax.annotate(
                            r["primary"], (r["lon"], r["lat"]),
                            xytext=(4, 2), textcoords="offset points",
                            fontsize=9, zorder=5,
                            color=self._GZX_COLORS[fam]))
            if trk_var.get() and self._result is not None:
                rr = self._result
                tlo = np.asarray(rr['lon'], dtype=float)
                tla = np.asarray(rr['lat'], dtype=float)
                brk = np.where(np.abs(np.diff(tlo)) > 180.0)[0]
                tlo, tla = tlo.copy(), tla.copy()
                tlo[brk] = np.nan
                state["artists"].append(ax.plot(
                    tlo, tla, color="black", lw=1.2, zorder=6)[0])
            state["k"] = k
            note = (f"{shown:,} features in view — complete" if k == 1
                    else f"showing a 1-in-{k} sample ({shown:,} dots) — "
                    "zoom in for all")
            status.set(note + "   ·   click a dot to identify")
            canvas.draw_idle()

        def _refresh(force=False):
            lims = (ax.get_xlim(), ax.get_ylim())
            if state["busy"] or (not force and lims == state["lims"]):
                return
            state["lims"] = lims
            state["busy"] = True
            (x0, x1), (y0, y1) = lims
            status.set("querying index…")

            def _work():
                try:
                    rows, k = _gz.viewport_sample(
                        max(-90, y0), min(90, y1), max(-180, x0),
                        min(180, x1), budget=20000, db=None,
                        k_hint=state["k"])
                except Exception as ex:
                    rows, k = [], 1
                    win.after(0, lambda: status.set(f"query failed: {ex}"))
                def _done():
                    state["busy"] = False
                    _apply(rows, k, lims)
                    if (ax.get_xlim(), ax.get_ylim()) != lims:
                        _refresh()          # user moved on meanwhile
                win.after(0, _done)
            threading.Thread(target=_work, daemon=True).start()

        # viewport_sample opens its own connection in the worker thread
        # (sqlite connections are not shared across threads); the main
        # thread keeps `db` only for the click-identify lookups.
        def _on_draw(_ev):
            win.after(60, _refresh)
        canvas.mpl_connect("draw_event", _on_draw)

        def _on_click(ev):
            if (ev.inaxes is not ax or ev.xdata is None
                    or getattr(toolbar, "mode", "")):
                return
            try:
                best, bkm = None, float("inf")
                import math as _m
                for r in state["rows"]:
                    d = _m.hypot((r["lon"] - ev.xdata)
                                 * _m.cos(_m.radians(ev.ydata)),
                                 r["lat"] - ev.ydata) * 111.32
                    if d < bkm:
                        best, bkm = r, d
                if best is None:
                    return
                word = _gz.class_word(best["fclass"], best["source"])
                txt = (f"{best['primary']} — {word} ({best['admin']}) "
                       f"[{best['ext_id']}]")
                pop = _gz.nearest_populated(best["lat"], best["lon"],
                                            n=1, db=db)
                if pop and pop[0]["ext_id"] != best["ext_id"]:
                    p = pop[0]
                    txt += (f"   ·   nearest populated: {p['primary']} "
                            f"({p['admin']}), {p['km']:.0f} km "
                            f"{p['dir']}")
                status.set(txt)
            except Exception as ex:
                status.set(f"identify failed: {ex}")
        canvas.mpl_connect("button_press_event", _on_click)
        _refresh(force=True)

    def _export_timeline(self):
        """Export the flight event timeline to CSV."""
        if self._result is None:
            messagebox.showinfo("No data", "Run a simulation first.")
            return
        milestones = self._result.get("milestones", [])
        if not milestones:
            messagebox.showinfo("No data", "No timeline events in last result.")
            return
        import datetime as _dt
        from tkinter.filedialog import asksaveasfilename
        ts      = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        booster = _safe_name(self._booster_var.get())
        rng_km  = self._result.get('range_km')
        rng_sfx = f"_{rng_km:.0f}km" if rng_km is not None else ""
        path = asksaveasfilename(
            defaultextension=".csv",
            initialdir=str(_ensure_dir(_DIR_EVENTS)),
            initialfile=f"{ts}_{booster}{rng_sfx}.events.csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            title="Export flight events",
        )
        if not path:
            return
        header = "event,time_s,alt_km,range_km,gnd_speed_kms,inrtl_speed_kms,accel_ms2,mass_t"
        rows = []
        for m in milestones:
            rows.append(",".join([
                f'"{m.get("event","")}"',
                f'{m.get("t_s", ""):g}',
                f'{m.get("alt_km", ""):g}',
                f'{m.get("range_km", ""):g}',
                f'{m.get("speed_kms", ""):g}',
                f'{m.get("inertial_speed_kms", ""):g}',
                f'{m.get("accel_ms2", ""):g}',
                f'{m.get("mass_t", ""):g}',
            ]))
        Path(path).write_text(header + "\n" + "\n".join(rows))
        self._status_var.set(f"Timeline exported: {path}")

    def _export_timeline_xlsx(self):
        """Export the flight event timeline to an XLSX workbook."""
        if self._result is None:
            messagebox.showinfo("No data", "Run a simulation first.")
            return
        milestones = self._result.get("milestones", [])
        if not milestones:
            messagebox.showinfo("No data", "No timeline events in last result.")
            return
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            messagebox.showerror("Missing dependency",
                                 f"openpyxl is required:\n{exc}")
            return
        import datetime as _dt
        from tkinter.filedialog import asksaveasfilename
        ts      = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        booster = _safe_name(self._booster_var.get())
        rng_km  = self._result.get('range_km')
        rng_sfx = f"_{rng_km:.0f}km" if rng_km is not None else ""
        path = asksaveasfilename(
            defaultextension=".xlsx",
            initialdir=str(_ensure_dir(_DIR_EVENTS)),
            initialfile=f"{ts}_{booster}{rng_sfx}.events.xlsx",
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
            title="Export flight events XLSX",
        )
        if not path:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Timeline"
        ws.append(["event", "time_s", "alt_km", "range_km",
                   "gnd_speed_kms", "inrtl_speed_kms",
                   "accel_ms2", "mass_t"])
        def _num(v):
            return float(v) if isinstance(v, (int, float)) else None
        for m in milestones:
            ws.append([
                m.get("event", ""),
                _num(m.get("t_s")),
                _num(m.get("alt_km")),
                _num(m.get("range_km")),
                _num(m.get("speed_kms")),
                _num(m.get("inertial_speed_kms")),
                _num(m.get("accel_ms2")),
                _num(m.get("mass_t")),
            ])
        wb.save(path)
        self._status_var.set(f"Timeline XLSX exported: {path}")

    def _export_booster(self):
        """Export the current booster definition to a .booster.json file."""
        name = self._booster_var.get()
        if not name or name not in BOOSTER_DB:
            messagebox.showinfo("No booster", "Select a booster first.")
            return
        from tkinter.filedialog import asksaveasfilename
        safe = _safe_name(name)
        path = asksaveasfilename(
            defaultextension=".json",
            initialdir=str(_boosters_dir()),
            initialfile=f"{safe}.booster.json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Export Booster",
        )
        if not path:
            return
        # get_booster applies the booster's flight plan, so the extracted plan
        # reflects the shipped guidance rather than hardware defaults.
        p = get_booster(name)
        # Booster files are hardware-only; the flight plan travels beside them
        # as a companion .flightplan.json so guidance never lives on hardware.
        data = booster_to_dict(p, include_flight_plan=False)
        Path(path).write_text(json.dumps(data, indent=2))
        fp_path = save_flight_plan(name, extract_flight_plan(p), Path(path).parent)
        self._status_var.set(
            f"Booster exported: {path}  (+ flight plan {Path(fp_path).name})")

    def _load_ro(self):
        """Import a .ro.json file into the RV library (parallel to Load Booster)."""
        from tkinter.filedialog import askopenfilename
        path = askopenfilename(
            initialdir=str(_ensure_dir(_RO_LIBRARY_PATH)),
            filetypes=[("reentry-object files", "*.ro.json"), ("JSON files", "*.json"),
                       ("All files", "*.*")],
            title="Load Reentry Object",
        )
        if not path:
            return
        try:
            ro = ro_from_dict(json.loads(Path(path).read_text()))
        except Exception as e:
            messagebox.showerror("Load error", f"Could not parse reentry-object file:\n{e}")
            return
        # Object files are hardware-only; apply a companion reentry plan if the
        # export dropped one beside it, so import is the inverse of export.
        _sib = _ro_plan_sibling(path)
        if _sib.exists():
            try:
                ro = apply_reentry_plan(ro, json.loads(_sib.read_text()))
            except Exception:
                pass
        name = ro.name or Path(path).stem.replace('.ro', '')
        if not name:
            messagebox.showerror("Load error", "reentry-object file has no name field.")
            return
        if name in RO_DB and not messagebox.askyesno(
                "Overwrite?", f"'{name}' already exists. Overwrite?"):
            return
        try:
            _save_ro_to_library(ro)        # copy into the writable user library
        except Exception as exc:
            messagebox.showerror("Load Reentry Object", f"Could not write reentry-object file:\n{exc}")
            return
        self._refresh_ro_list(select_name=name)
        self._status_var.set(f"Reentry object '{name}' loaded from {Path(path).name}")

    def _export_ro(self):
        """Export the selected RV (or the booster's RV) to a .ro.json file."""
        sel = self._ro_main_var.get()
        ro = RO_DB[sel]() if sel in RO_DB else getattr(self, '_ro', None)
        if ro is None or not getattr(ro, 'name', ''):
            messagebox.showinfo("No Reentry Object", "Select a reentry object first.", parent=self)
            return
        from tkinter.filedialog import asksaveasfilename
        path = asksaveasfilename(
            parent=self,
            defaultextension=".json",
            initialdir=str(_ensure_dir(_RO_LIBRARY_PATH)),
            initialfile=f"{_safe_name(ro.name)}.ro.json",
            filetypes=[("reentry-object files", "*.ro.json"), ("JSON files", "*.json"),
                       ("All files", "*.*")],
            title="Export Reentry Object",
        )
        if not path:
            return
        try:
            # Hardware-only object file plus a companion reentry plan beside it,
            # so the exported pair round-trips without embedding the plan.
            Path(path).write_text(
                json.dumps(ro_to_dict(ro, include_reentry_plan=False), indent=2))
            _ro_plan_sibling(path).write_text(
                json.dumps(extract_reentry_plan(ro), indent=2) + "\n")
        except Exception as exc:
            messagebox.showerror("Save Reentry Object",
                                 f"Could not write reentry-object file:\n{exc}", parent=self)
            return
        self._status_var.set(f"Reentry object exported: {path}")

    def _export_ro_xlsx(self):
        """Export the selected RV to a fillable XLSX spreadsheet."""
        sel = self._ro_main_var.get()
        ro = RO_DB[sel]() if sel in RO_DB else getattr(self, '_ro', None)
        if ro is None or not getattr(ro, 'name', ''):
            messagebox.showinfo("No Reentry Object", "Select a reentry object first.", parent=self)
            return
        try:
            from ro_xlsx import export_ro_xlsx
        except ImportError as exc:
            messagebox.showerror("Missing dependency", str(exc), parent=self)
            return
        from tkinter.filedialog import asksaveasfilename
        path = asksaveasfilename(
            parent=self, title="Save Reentry Object to XLSX",
            defaultextension=".xlsx",
            initialdir=str(_ensure_dir(_RO_LIBRARY_PATH)),
            initialfile=f"{_safe_name(ro.name)}.ro.xlsx",
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")])
        if not path:
            return
        try:
            export_ro_xlsx(path, ro)
            self._status_var.set(f"Reentry object exported: {os.path.basename(path)}")
        except Exception as exc:
            messagebox.showerror("Export error", str(exc), parent=self)

    def _import_ro_xlsx(self):
        """Import an RV from a filled XLSX spreadsheet into the library."""
        try:
            from ro_xlsx import import_ro_xlsx
        except ImportError as exc:
            messagebox.showerror("Missing dependency", str(exc), parent=self)
            return
        from tkinter.filedialog import askopenfilename
        path = askopenfilename(
            parent=self, title="Load Reentry Object from XLSX",
            initialdir=str(_ensure_dir(_RO_LIBRARY_PATH)),
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")])
        if not path:
            return
        try:
            ro = import_ro_xlsx(path)
        except Exception as exc:
            messagebox.showerror("Import error", str(exc), parent=self)
            return
        if not ro.name:
            messagebox.showwarning(
                "Import warning",
                "Reentry object name is blank — fill in the Name field in the XLSX and "
                "re-import.", parent=self)
            return
        try:
            _save_ro_to_library(ro)
        except Exception as exc:
            messagebox.showerror("Save Reentry Object",
                                 f"Could not write reentry-object file:\n{exc}", parent=self)
            return
        self._refresh_ro_list(select_name=ro.name)
        self._status_var.set(f"Reentry object imported: {ro.name}")

    def _new_ro_template(self):
        """Save a blank RV XLSX template the user fills in from scratch."""
        try:
            from ro_xlsx import make_blank_ro_template
        except ImportError as exc:
            messagebox.showerror("Missing dependency", str(exc), parent=self)
            return
        from tkinter.filedialog import asksaveasfilename
        path = asksaveasfilename(
            parent=self, title="Save Blank Reentry Object Template",
            defaultextension=".xlsx",
            initialdir=str(_ensure_dir(_RO_LIBRARY_PATH)),
            initialfile="new_ro.ro.xlsx",
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")])
        if not path:
            return
        try:
            make_blank_ro_template(path)
            self._status_var.set(f"Template saved: {os.path.basename(path)}")
        except Exception as exc:
            messagebox.showerror("Template error", str(exc), parent=self)

    def _export_booster_xlsx(self):
        """Export current booster to a filled-in XLSX template."""
        name = self._booster_var.get()
        if not name or name not in BOOSTER_DB:
            messagebox.showinfo("No booster", "Select a booster first.", parent=self)
            return
        try:
            from booster_xlsx import export_booster_xlsx
        except ImportError as exc:
            messagebox.showerror("Missing dependency", str(exc), parent=self)
            return
        from tkinter.filedialog import asksaveasfilename
        safe = _safe_name(name)
        path = asksaveasfilename(
            title="Export Booster to XLSX",
            defaultextension=".xlsx",
            initialdir=str(_boosters_dir()),
            initialfile=f"{safe}.booster.xlsx",
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
            parent=self,
        )
        if not path:
            return
        try:
            export_booster_xlsx(path, BOOSTER_DB[name]())
            self._status_var.set(f"Booster exported: {os.path.basename(path)}")
        except Exception as exc:
            messagebox.showerror("Export error", str(exc), parent=self)

    def _import_booster_xlsx(self):
        """Import a booster from a filled XLSX template."""
        try:
            from booster_xlsx import import_booster_xlsx
        except ImportError as exc:
            messagebox.showerror("Missing dependency", str(exc), parent=self)
            return
        from tkinter.filedialog import askopenfilename
        path = askopenfilename(
            title="Import Booster from XLSX",
            initialdir=str(_boosters_dir()),
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
            parent=self,
        )
        if not path:
            return
        try:
            params = import_booster_xlsx(path)
        except Exception as exc:
            messagebox.showerror("Import error", str(exc), parent=self)
            return
        if not params.name:
            messagebox.showwarning("Import warning",
                                   "Booster name is blank — please fill in "
                                   "the Name field in the XLSX and re-import.",
                                   parent=self)
            return
        if params.name in BOOSTER_DB and not messagebox.askyesno(
                "Overwrite?", f"'{params.name}' already exists. Overwrite?",
                parent=self):
            return
        BOOSTER_DB[params.name] = lambda p=params: p
        _save_custom_boosters()
        self._refresh_booster_list(select_name=params.name)
        self._status_var.set(f"Booster imported: {params.name}")

    def _new_booster_template(self):
        """Save a blank XLSX template the user fills in from scratch."""
        try:
            from booster_xlsx import make_blank_template
        except ImportError as exc:
            messagebox.showerror("Missing dependency", str(exc), parent=self)
            return
        from tkinter.filedialog import asksaveasfilename
        path = asksaveasfilename(
            title="Save Blank Booster Template",
            defaultextension=".xlsx",
            initialdir=str(_boosters_dir()),
            initialfile="booster_template.xlsx",
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
            parent=self,
        )
        if not path:
            return
        try:
            make_blank_template(path)
            self._status_var.set(f"Template saved: {os.path.basename(path)}")
        except Exception as exc:
            messagebox.showerror("Template error", str(exc), parent=self)

    def _load_booster(self):
        """Import a .booster.json file into the custom booster library."""
        from tkinter.filedialog import askopenfilename
        path = askopenfilename(
            initialdir=str(_boosters_dir()),
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Load Booster",
        )
        if not path:
            return
        try:
            data = json.loads(Path(path).read_text())
            p    = booster_from_dict(data)
        except Exception as e:
            messagebox.showerror("Load error", f"Could not parse booster file:\n{e}")
            return
        # Booster files are hardware-only; if the export dropped a companion
        # flight plan beside it, apply it so the imported booster keeps its
        # guidance (making import the exact inverse of export).
        _fp_sibling = Path(path).with_name(
            Path(path).name.replace('.booster.json', '').replace('.json', '')
            + '.flightplan.json')
        if _fp_sibling.exists():
            try:
                p = apply_flight_plan(p, json.loads(_fp_sibling.read_text()))
            except Exception:
                pass
        name = data.get('name') or Path(path).stem.replace('.booster', '').replace('.booster', '')
        if not name:
            messagebox.showerror("Load error", "Booster file has no name field.")
            return
        if name in BOOSTER_DB and not messagebox.askyesno(
                "Overwrite?", f"'{name}' already exists. Overwrite?"):
            return
        BOOSTER_DB[name] = lambda p=p: p
        _save_custom_boosters()
        self._refresh_booster_list(select_name=name)
        self._status_var.set(f"Booster '{name}' loaded from {Path(path).name}")

    def _export_site(self):
        """Export the current launch site to a .site.json file."""
        name = self._site_var.get()
        lat_s = self._launch_lat.get().strip()
        lon_s = self._launch_lon.get().strip()
        if not lat_s or not lon_s:
            messagebox.showinfo("No site", "Enter a launch site location first.")
            return
        from tkinter.filedialog import asksaveasfilename
        safe = _safe_name(name or "site")
        path = asksaveasfilename(
            defaultextension=".json",
            initialdir=str(_ensure_dir(_DIR_SITES)),
            initialfile=f"{safe}.site.json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Export Launch Site",
        )
        if not path:
            return
        data = {"name": name, "lat": float(lat_s), "lon": float(lon_s)}
        Path(path).write_text(json.dumps(data, indent=2))
        self._status_var.set(f"Site exported: {path}")

    def _load_site(self):
        """Import a .site.json file into the custom launch-site library."""
        from tkinter.filedialog import askopenfilename
        path = askopenfilename(
            initialdir=str(_ensure_dir(_DIR_SITES)),
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Load Launch Site",
        )
        if not path:
            return
        try:
            data = json.loads(Path(path).read_text())
            name = data['name']
            lat  = float(data['lat'])
            lon  = float(data['lon'])
        except Exception as e:
            messagebox.showerror("Load error", f"Could not parse site file:\n{e}")
            return
        user_sites = _load_user_sites()
        if any(s['name'] == name for s in user_sites):
            if not messagebox.askyesno("Overwrite?",
                                       f"'{name}' already exists. Overwrite?"):
                return
            user_sites = [s for s in user_sites if s['name'] != name]
        user_sites.append({"name": name, "lat": lat, "lon": lon})
        _save_user_sites(user_sites)
        new_values, new_map = _load_launch_sites()
        self._site_map = new_map
        self._site_cb.config(values=new_values)
        self._site_var.set(name)
        self._launch_lat.set(f"{lat:.4f}")
        self._launch_lon.set(f"{lon:.4f}")
        self._status_var.set(f"Site '{name}' loaded from {Path(path).name}")

    def _show_about(self):
        messagebox.showinfo(
            "About Thrusty",
            "Thrusty\n\n"
            "Based on the MATLAB application by Geoffrey Forden\n"
            "G. Forden, Science & Global Security 15 (2007)\n\n"
            "3-DOF trajectory integration:\n"
            "  • COESA 1976 standard atmosphere\n"
            "  • WGS-84 J2 gravity (ECEF)\n"
            "  • Coriolis & centrifugal corrections\n"
            "  • Gravity-turn guidance with per-stage pitch profiles\n"
            "  • Up to 4 stages with inter-stage coast\n\n"
            "Packaged boosters (Forden Table 1 + extension):\n"
            "  Scud-B, Al Hussein, No-dong,\n"
            "  Taepodong-I, Taepodong-II (3-stage),\n"
            "  Shahab-3, Generic ICBM\n"
        )


def _bg_spatial_upgrade():
    """One-time background upgrade of an ALREADY-BUILT gazetteer index:
    add the R-Tree and populated-places helper tables that make the
    Nearby Places / Explorer nearest-neighbour queries millisecond-fast
    (field report 2026-08-18: 'the nearest place search is slow').
    Runs silently off the UI thread at startup; fresh index builds get
    the helpers directly; until this finishes, queries use the slower
    correct fallback.  Never triggers the multi-minute index build
    itself — users who haven't opted into the gazetteer pay nothing."""
    try:
        import gazetteer as _gz
        if not (_gz.available() and _gz.index_ready()):
            return
        db = _gz.ensure_index()          # thread-local connection
        _gz.ensure_rtree(db)
        _gz.ensure_pplaces(db)
        db.close()
    except Exception:
        pass                             # cosmetic speed-up only


# ---------------------------------------------------------------------------
def main():
    app = BoosterFlyoutApp()
    threading.Thread(target=_bg_spatial_upgrade, daemon=True).start()
    app.mainloop()


if __name__ == "__main__":
    main()

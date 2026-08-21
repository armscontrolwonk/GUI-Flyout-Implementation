"""
ro_xlsx.py — XLSX import/export for Thrusty reentry objects (ROParams).

The reentry-object counterpart to booster_xlsx.py: edit a reentry object in a
familiar spreadsheet grid instead of hand-editing ro.json.  Reuses that
module's low-level cell writers/readers so the two stay visually consistent.

Scope note: this is a self-contained full-object exchange format — one sheet
carries both the hardware (mass, β, TPS, L/D capability) AND the reentry-plan
fields (glide law, commanded L/D, pull-up g, βₛ, dives, separation), even
though the live app now stores those in separate .ro.json / .reentryplan.json
files.  Keeping them together here is deliberate: a spreadsheet is a portable
snapshot of a whole vehicle, not the live store.  An imported object therefore
arrives with its plan fields as defaults on the object, which the reentry-plan
library then owns; re-export round-trips them.

Sheet layout
------------
  Sheet 1 "RO"        — every ROParams field, fields-as-rows, values in col D
  Sheet 2 "Reference" — read-only TPS material catalog + emissivity guidance

Public API
----------
  export_ro_xlsx(path, ro)      -> None
  import_ro_xlsx(path)          -> ROParams
  make_blank_ro_template(path)  -> None

Google Sheets: no dedicated integration is needed — in Google Sheets use
File > Download > Microsoft Excel (.xlsx), then Load Reentry Object from XLSX.  A native
Sheets read/write is a possible future enhancement.
"""

from __future__ import annotations

from booster_xlsx import (
    _xl, _section, _label, _inputs, _dropdown, _yn,
    _rnum, _rint, _rstr, _rbool,
    _NOSE_OPTS, _NOSE_LABEL, _NOSE_KEY,
)

# ---------------------------------------------------------------------------
# Row registry (1-based).  Writer and reader share it so the mapping never
# drifts.  All values live in column D (4); labels/units/notes in B/C/I.
# ---------------------------------------------------------------------------
_R: dict[str, int] = {
    'name':        4,
    # geometry & mass
    'mass':        7,
    'beta':        8,
    'shape':       9,
    'diam':       10,
    'length':     11,
    'nose_rn':    12,
    'sep':        13,
    # maneuvering / glider
    'g_on':       16,
    'g_ld':       17,
    'g_gmax':     18,
    'g_betaS':    19,
    'g_guid':     20,
    'g_zeta':     21,
    'g_skip':     22,
    'g_aero':     23,
    'g_tdive':    24,
    'g_talt':     25,
    # TPS
    'emiss':      28,
    'nose_mat':   29,
    'body_mat':   30,
    'body_thk':   31,
    'struct_mat': 32,
    'struct_lim': 33,
    # custom nose material
    'cn_label':   36,
    'cn_abl':     37,
    'cn_lim':     38,
    'cn_dens':    39,
    'cn_heff':    40,
    # custom body material
    'cb_label':   43,
    'cb_abl':     44,
    'cb_lim':     45,
    'cb_dens':    46,
    'cb_heff':    47,
    # provenance
    'source':     50,
    'notes':      51,
    # ── rows below were APPENDED after the original layout: old workbooks
    # simply have empty cells here, and the reader defaults them (axisymmetric,
    # no biconic, no wings) — a pre-upgrade file imports unchanged. ──
    # body form & biconic hardware
    'body_form':  54,
    'biconic':    55,
    'fore_len':   56,
    'break_dia':  57,
    'body_span':  58,
    # wings (drag-polar anchor + planform; planform is the primary data —
    # wing_geometry() derives S/AR from it at every consumer)
    'wing_area':  60,
    'wing_ar':    61,
    'wing_root':  62,
    'wing_span':  63,
    'wing_sweep': 64,
    # trim row from the lifting-body estimator (sweep-native coefficients)
    'trim_alpha': 65,
    'trim_cl0':   66,
    # wing geometry for the 3-D Blender export (not the polar)
    'wing_thick': 67,
    'n_wings':    68,
    # non-separating body: forward-taper (nose) length carved from the body
    'body_nose':  70,
}

_VAL_COL = 4   # column D

_SEP_OPTS   = ['separating_ro', 'body']
_YESNO_OPTS = ['YES', 'NO']
_CUSTOM     = 'custom'          # material-cell sentinel for a bespoke material


def _material_opts():
    """Dropdown options for a material cell: blank, every catalog key, custom."""
    import heating
    return [''] + list(heating.TPS_MATERIALS.keys()) + [_CUSTOM]


# ---------------------------------------------------------------------------
# Writer
# ---------------------------------------------------------------------------
def _build_ro_sheet(ws, ro) -> None:
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 26
    ws.column_dimensions['I'].width = 40

    ws.merge_cells('B1:D1')
    t = ws.cell(row=1, column=2, value='Thrusty — Reentry Object')
    from openpyxl.styles import Font
    t.font = Font(bold=True, size=13)

    mat_opts = _material_opts()

    def put(rk, value, fmt='General'):
        _inputs(ws, _R[rk], [_VAL_COL], [value], fmt=fmt)

    # Identity
    _section(ws, 3, 'Identity')
    _label(ws, _R['name'], 'Name')
    put('name', ro.name)

    # Geometry & mass
    _section(ws, 6, 'Geometry & mass')
    _label(ws, _R['mass'],   'Mass', 'kg')
    _label(ws, _R['beta'],   'Ballistic coeff β', 'kg/m²',
           'm/(Cd·A); scales with mass at fixed Cd·A')
    _label(ws, _R['shape'],  'Nose shape')
    _label(ws, _R['diam'],   'Base diameter', 'm')
    _label(ws, _R['length'], 'Length', 'm')
    _label(ws, _R['nose_rn'],'Nose-tip radius', 'm')
    _label(ws, _R['sep'],    'Separation mode', '', 'separating_ro or body')
    put('mass', ro.mass_kg); put('beta', ro.beta_kg_m2)
    put('diam', ro.diameter_m); put('length', ro.length_m)
    put('nose_rn', ro.nose_radius_m)
    _dropdown(ws, _R['shape'], _VAL_COL, _NOSE_OPTS)
    ws.cell(row=_R['shape'], column=_VAL_COL, value=_NOSE_LABEL.get(ro.shape, 'Cone'))
    _dropdown(ws, _R['sep'], _VAL_COL, _SEP_OPTS)
    ws.cell(row=_R['sep'], column=_VAL_COL, value=(ro.separation_mode or 'separating_ro'))

    # Maneuvering / glider
    _section(ws, 15, 'Maneuvering (glider / HGV)')
    _label(ws, _R['g_on'],    'Glider enabled')
    _label(ws, _R['g_ld'],    'Lift/drag L/D')
    _label(ws, _R['g_gmax'],  'Pull-up g-limit', 'g')
    _label(ws, _R['g_betaS'], 'Re-entry βₛ', 'kg/m²', '0 = Tracy')
    _label(ws, _R['g_guid'],  'Guidance mode', '', 'e.g. damped_glide, skip_glide')
    _label(ws, _R['g_zeta'],  'Damping ratio ζ', '', 'damped_glide only')
    _label(ws, _R['g_skip'],  'Skip count')
    _label(ws, _R['g_aero'],  'Aero model', '', 'e.g. polar')
    _label(ws, _R['g_tdive'], 'Terminal dive')
    _label(ws, _R['g_talt'],  'Terminal alt', 'km')
    _dropdown(ws, _R['g_on'], _VAL_COL, _YESNO_OPTS)
    ws.cell(row=_R['g_on'], column=_VAL_COL, value=_yn(ro.glider_enabled))
    put('g_ld', ro.glider_LD); put('g_gmax', ro.glider_pullup_g_max)
    put('g_betaS', ro.glider_beta_entry_kg_m2)
    put('g_guid', ro.glider_guidance); put('g_zeta', ro.glider_damping_zeta)
    put('g_skip', ro.glider_skip_count)
    put('g_aero', getattr(ro, 'glider_aero_model', 'polar'))
    _dropdown(ws, _R['g_tdive'], _VAL_COL, _YESNO_OPTS)
    ws.cell(row=_R['g_tdive'], column=_VAL_COL, value=_yn(ro.glider_terminal_dive))
    put('g_talt', ro.glider_terminal_alt_km)

    # TPS materials
    _section(ws, 27, 'Thermal protection (TPS)')
    _label(ws, _R['emiss'],     'Emissivity', '', '0.85 typical (0.75–0.90)')
    _label(ws, _R['nose_mat'],  'Nose material')
    _label(ws, _R['body_mat'],  'Body material')
    _label(ws, _R['body_thk'],  'Body layer thickness', 'm', '0 = auto')
    _label(ws, _R['struct_mat'],'Structure material')
    _label(ws, _R['struct_lim'],'Structure limit', 'K')
    put('emiss', ro.emissivity); put('body_thk', ro.body_tps_thickness_m)
    put('struct_lim', ro.structure_limit_K)
    import heating
    for rk, key in (('nose_mat', ro.nose_tps_material),
                    ('body_mat', ro.body_tps_material),
                    ('struct_mat', ro.structure_material)):
        _dropdown(ws, _R[rk], _VAL_COL, mat_opts)
        cell_val = key
        if key == heating.CUSTOM_NOSE_KEY or key == heating.CUSTOM_BODY_KEY:
            cell_val = _CUSTOM
        ws.cell(row=_R[rk], column=_VAL_COL, value=cell_val)

    # Custom material sub-blocks
    def _put_custom(prefix, cust):
        cust = cust or {}
        _inputs(ws, _R[prefix + '_label'], [_VAL_COL], [cust.get('label', '')])
        _dropdown(ws, _R[prefix + '_abl'], _VAL_COL, _YESNO_OPTS)
        ws.cell(row=_R[prefix + '_abl'], column=_VAL_COL,
                value=_yn(bool(cust.get('is_ablator', False))))
        _inputs(ws, _R[prefix + '_lim'],  [_VAL_COL],
                [cust.get('continuous_K') or cust.get('peak_K') or ''])
        _inputs(ws, _R[prefix + '_dens'], [_VAL_COL], [cust.get('density_kg_m3', '')])
        _inputs(ws, _R[prefix + '_heff'], [_VAL_COL], [cust.get('H_eff_MJ_kg', '')])

    _section(ws, 35, 'Custom nose material  (only if Nose material = custom)')
    for rk, lb, un in (('cn_label', 'Name', ''), ('cn_abl', 'Ablator', ''),
                       ('cn_lim', 'Temp. limit', 'K'), ('cn_dens', 'Density', 'kg/m³'),
                       ('cn_heff', 'Heat of ablation', 'MJ/kg')):
        _label(ws, _R[rk], lb, un)
    _put_custom('cn', ro.nose_tps_custom)

    _section(ws, 42, 'Custom body material  (only if Body material = custom)')
    for rk, lb, un in (('cb_label', 'Name', ''), ('cb_abl', 'Ablator', ''),
                       ('cb_lim', 'Temp. limit', 'K'), ('cb_dens', 'Density', 'kg/m³'),
                       ('cb_heff', 'Heat of ablation', 'MJ/kg')):
        _label(ws, _R[rk], lb, un)
    _put_custom('cb', ro.body_tps_custom)

    # Provenance
    _section(ws, 49, 'Provenance')
    _label(ws, _R['source'], 'Source', '', 'short citation')
    _label(ws, _R['notes'],  'Notes', '', 'free-form; confidence, assumptions')
    _inputs(ws, _R['source'], [_VAL_COL], [ro.source])
    _inputs(ws, _R['notes'],  [_VAL_COL], [ro.notes])

    # Body form & biconic (appended rows — see registry note)
    from booster_models import BODY_FORMS
    _section(ws, 53, 'Body form & biconic')
    _label(ws, _R['body_form'], 'Body form', '',
           'axisymmetric / wedge (⌀ = base depth) / half_cone')
    _label(ws, _R['biconic'],   'Biconic (two-cone)', '',
           'axisymmetric only; ignored for lifting forms')
    _label(ws, _R['fore_len'],  'Fore-cone length', 'm')
    _label(ws, _R['break_dia'], 'Break diameter', 'm')
    _label(ws, _R['body_span'], 'Planform span', 'm',
           'wedge only: tip-to-tip base width (body IS the wing)')
    _dropdown(ws, _R['body_form'], _VAL_COL, list(BODY_FORMS))
    ws.cell(row=_R['body_form'], column=_VAL_COL,
            value=(getattr(ro, 'body_form', '') or 'axisymmetric'))
    _dropdown(ws, _R['biconic'], _VAL_COL, _YESNO_OPTS)
    ws.cell(row=_R['biconic'], column=_VAL_COL,
            value=_yn(bool(getattr(ro, 'biconic', False))))
    put('fore_len', getattr(ro, 'fore_length_m', 0.0))
    put('break_dia', getattr(ro, 'break_diameter_m', 0.0))
    put('body_span', getattr(ro, 'body_span_m', 0.0))
    _label(ws, _R['body_nose'], 'Body nose length', 'm',
           'non-separating body only: forward taper carved from the body '
           '(0 = auto); separating RVs use their own length instead')
    put('body_nose', getattr(ro, 'body_nose_length_m', 0.0))

    # Wings (appended rows).  The planform (root chord + exposed span + sweep)
    # is the PRIMARY data: when present, S and AR are DERIVED from it by
    # wing_geometry() at every consumer — the S/AR cells here are the direct-
    # entry fallback and are overridden by a planform, exactly as in the app.
    _section(ws, 59, 'Wings  (planform is primary; S/AR direct-entry fallback)')
    _label(ws, _R['wing_area'],  'Wing area S', 'm²', 'fallback if no planform')
    _label(ws, _R['wing_ar'],    'Aspect ratio AR', '', 'fallback if no planform')
    _label(ws, _R['wing_root'],  'Root chord', 'm', 'planform')
    _label(ws, _R['wing_span'],  'Exposed span', 'm', 'planform')
    _label(ws, _R['wing_sweep'], 'LE sweep', '°', 'planform')
    put('wing_area',  getattr(ro, 'wing_area_m2', 0.0))
    put('wing_ar',    getattr(ro, 'wing_aspect_ratio', 0.0))
    put('wing_root',  getattr(ro, 'wing_root_chord_m', 0.0))
    put('wing_span',  getattr(ro, 'wing_span_exposed_m', 0.0))
    put('wing_sweep', getattr(ro, 'wing_sweep_deg', 0.0))
    # Estimator trim row (Phase 3: offset polar + windward-α guard).
    _label(ws, _R['trim_alpha'], 'Trim α* (estimator)', '°',
           'from the α-sweep; 0 = absent')
    _label(ws, _R['trim_cl0'], 'Camber offset C_L0', '',
           'sweep-native coefficients; 0 = symmetric polar')
    put('trim_alpha', getattr(ro, 'trim_alpha_deg', 0.0))
    put('trim_cl0',   getattr(ro, 'trim_CL0', 0.0))
    # Wing geometry for the 3-D Blender export (panel thickness + count);
    # geometry only — the polar never reads these.
    _label(ws, _R['wing_thick'], 'Wing panel thickness', 'm', '3-D export')
    _label(ws, _R['n_wings'],    'Wing panel count', '', '3-D export (e.g. 4)')
    put('wing_thick', getattr(ro, 'wing_thickness_m', 0.0))
    put('n_wings',    getattr(ro, 'n_wings', 4))


def _build_ro_reference_sheet(ws) -> None:
    import heating
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    hdr = ['key', 'label', 'group', 'limit K', 'ablator?']
    for j, h in enumerate(hdr, start=1):
        from openpyxl.styles import Font
        c = ws.cell(row=1, column=j, value=h); c.font = Font(bold=True)
    r = 2
    for key, m in heating.TPS_MATERIALS.items():
        ws.cell(row=r, column=1, value=key)
        ws.cell(row=r, column=2, value=m.get('label', ''))
        ws.cell(row=r, column=3, value=m.get('group', ''))
        ws.cell(row=r, column=4, value=m.get('continuous_K') or m.get('peak_K'))
        ws.cell(row=r, column=5, value='yes' if m.get('is_ablator') else 'no')
        r += 1
    ws.cell(row=r + 1, column=1, value='Emissivity typical: 0.85  (range 0.75–0.90)')


# ---------------------------------------------------------------------------
# Reader
# ---------------------------------------------------------------------------
def _read_material(ws, rk, prefix, sentinel):
    """Return (material_key, custom_dict_or_None) for a material cell."""
    val = _rstr(ws, _R[rk], _VAL_COL)
    if val.lower() == _CUSTOM:
        lim = _rnum(ws, _R[prefix + '_lim'], _VAL_COL, 0.0)
        label = _rstr(ws, _R[prefix + '_label'], _VAL_COL, 'Custom material')
        if lim <= 0 and not label:
            return '', None
        cust = {
            'label': label or 'Custom material',
            'is_ablator': _rbool(ws, _R[prefix + '_abl'], _VAL_COL),
            'continuous_K': lim, 'peak_K': lim,
            'density_kg_m3': _rnum(ws, _R[prefix + '_dens'], _VAL_COL, 0.0) or None,
            'H_eff_MJ_kg': _rnum(ws, _R[prefix + '_heff'], _VAL_COL, 0.0) or None,
        }
        return sentinel, cust
    return val, None


def import_ro_xlsx(path: str):
    """Read a reentry-object XLSX and return an ROParams.

    Backward compatible: workbooks written before the appended 'Body form &
    biconic' / 'Wings' sections have empty cells at those rows, which read as
    defaults (axisymmetric, no biconic, no wings) — a pre-upgrade file
    imports exactly as it always did.
    """
    import heating
    from booster_models import ROParams, _norm_sep_mode, BODY_FORMS
    xl = _xl()
    wb = xl.load_workbook(path, data_only=True)
    # New workbooks use sheet "RO"; accept the legacy "RV" name too.
    ws = wb['RO'] if 'RO' in wb.sheetnames else wb['RV']

    nose_key, nose_cust = _read_material(ws, 'nose_mat', 'cn', heating.CUSTOM_NOSE_KEY)
    body_key, body_cust = _read_material(ws, 'body_mat', 'cb', heating.CUSTOM_BODY_KEY)
    struct_key = _rstr(ws, _R['struct_mat'], _VAL_COL)
    if struct_key.lower() == _CUSTOM:
        struct_key = ''      # custom not supported for the structure slot

    # Unknown/legacy body-form strings normalise to the default, as in
    # ro_from_dict — never crash, never propagate an invented form.
    _bf = _rstr(ws, _R['body_form'], _VAL_COL, 'axisymmetric')
    body_form = _bf if _bf in BODY_FORMS else 'axisymmetric'

    return ROParams(
        name=_rstr(ws, _R['name'], _VAL_COL, 'Unnamed'),
        mass_kg=_rnum(ws, _R['mass'], _VAL_COL),
        beta_kg_m2=_rnum(ws, _R['beta'], _VAL_COL),
        shape=_NOSE_KEY.get(_rstr(ws, _R['shape'], _VAL_COL), 'cone') or 'cone',
        diameter_m=_rnum(ws, _R['diam'], _VAL_COL),
        length_m=_rnum(ws, _R['length'], _VAL_COL),
        nose_radius_m=_rnum(ws, _R['nose_rn'], _VAL_COL),
        body_form=body_form,
        biconic=_rbool(ws, _R['biconic'], _VAL_COL),
        fore_length_m=_rnum(ws, _R['fore_len'], _VAL_COL),
        break_diameter_m=_rnum(ws, _R['break_dia'], _VAL_COL),
        body_span_m=_rnum(ws, _R['body_span'], _VAL_COL),
        body_nose_length_m=_rnum(ws, _R['body_nose'], _VAL_COL),
        wing_area_m2=_rnum(ws, _R['wing_area'], _VAL_COL),
        wing_aspect_ratio=_rnum(ws, _R['wing_ar'], _VAL_COL),
        wing_root_chord_m=_rnum(ws, _R['wing_root'], _VAL_COL),
        wing_span_exposed_m=_rnum(ws, _R['wing_span'], _VAL_COL),
        wing_sweep_deg=_rnum(ws, _R['wing_sweep'], _VAL_COL),
        wing_thickness_m=_rnum(ws, _R['wing_thick'], _VAL_COL),
        n_wings=int(_rnum(ws, _R['n_wings'], _VAL_COL) or 4),
        trim_alpha_deg=_rnum(ws, _R['trim_alpha'], _VAL_COL),
        trim_CL0=_rnum(ws, _R['trim_cl0'], _VAL_COL),
        separation_mode=_norm_sep_mode(_rstr(ws, _R['sep'], _VAL_COL, 'separating_ro')),
        glider_enabled=_rbool(ws, _R['g_on'], _VAL_COL),
        glider_LD=_rnum(ws, _R['g_ld'], _VAL_COL),
        glider_pullup_g_max=_rnum(ws, _R['g_gmax'], _VAL_COL, 10.0),
        glider_beta_entry_kg_m2=_rnum(ws, _R['g_betaS'], _VAL_COL),
        glider_guidance=_rstr(ws, _R['g_guid'], _VAL_COL, 'equilibrium_glide') or 'equilibrium_glide',
        glider_damping_zeta=_rnum(ws, _R['g_zeta'], _VAL_COL, 0.7),
        glider_skip_count=_rint(ws, _R['g_skip'], _VAL_COL, 1),
        glider_aero_model=_rstr(ws, _R['g_aero'], _VAL_COL, 'polar') or 'polar',
        glider_terminal_dive=_rbool(ws, _R['g_tdive'], _VAL_COL),
        glider_terminal_alt_km=_rnum(ws, _R['g_talt'], _VAL_COL, 30.0),
        emissivity=_rnum(ws, _R['emiss'], _VAL_COL, 0.85),
        nose_tps_material=nose_key,
        body_tps_material=body_key,
        body_tps_thickness_m=_rnum(ws, _R['body_thk'], _VAL_COL),
        structure_material=struct_key,
        structure_limit_K=_rnum(ws, _R['struct_lim'], _VAL_COL),
        nose_tps_custom=nose_cust,
        body_tps_custom=body_cust,
        source=_rstr(ws, _R['source'], _VAL_COL),
        notes=_rstr(ws, _R['notes'], _VAL_COL),
    )


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
def export_ro_xlsx(path: str, ro) -> None:
    """Write an ROParams to a fully filled-in XLSX at path."""
    xl = _xl()
    wb = xl.Workbook()
    ws = wb.active
    ws.title = 'RO'
    _build_ro_sheet(ws, ro)
    _build_ro_reference_sheet(wb.create_sheet('Reference'))
    wb.save(path)


def make_blank_ro_template(path: str) -> None:
    """Write a blank RV template for the user to fill in from scratch."""
    from booster_models import ROParams
    export_ro_xlsx(path, ROParams(name='', mass_kg=0, beta_kg_m2=0))

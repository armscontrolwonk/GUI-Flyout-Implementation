"""XLSX round-trip for reentry objects (ro_xlsx).

The gap this closes (TODO item): the spreadsheet exchange format silently
dropped the biconic trio, the wing fields, and body_form — a wedge exported
to XLSX came back a plain cone.  The new rows are APPENDED after the original
layout, so a pre-upgrade workbook (empty cells there) must import exactly as
it always did: axisymmetric, no biconic, no wings.
"""

import dataclasses

import pytest

pytest.importorskip("openpyxl", reason="openpyxl not installed")

from booster_models import ROParams, wing_geometry
from ro_xlsx import export_ro_xlsx, import_ro_xlsx, make_blank_ro_template


def _ro(**kw):
    base = dict(name="RT", mass_kg=450.0, beta_kg_m2=15000.0, shape="cone",
                diameter_m=0.58, length_m=1.5, nose_radius_m=0.02,
                glider_enabled=True, glider_LD=2.0, source="test", notes="n")
    base.update(kw)
    return ROParams(**base)


_NEW_FIELDS = ("body_form", "biconic", "fore_length_m", "break_diameter_m",
               "body_span_m",
               "wing_area_m2", "wing_aspect_ratio", "wing_root_chord_m",
               "wing_span_exposed_m", "wing_sweep_deg",
               "wing_thickness_m", "n_wings")


def test_biconic_with_planform_wings_round_trips(tmp_path):
    """The exact loss case: a biconic with planform wings must come back
    field-for-field, and wing_geometry() must derive the same S/AR from the
    re-imported planform (single source of truth survives the spreadsheet)."""
    p = tmp_path / "ro.xlsx"
    ro = _ro(biconic=True, fore_length_m=0.5, break_diameter_m=0.42,
             wing_area_m2=0.2, wing_aspect_ratio=0.0,
             wing_root_chord_m=0.6, wing_span_exposed_m=0.15,
             wing_sweep_deg=65.0, wing_thickness_m=0.04, n_wings=4)
    export_ro_xlsx(str(p), ro)
    back = import_ro_xlsx(str(p))
    for f in _NEW_FIELDS:
        assert getattr(back, f) == getattr(ro, f), f
    assert wing_geometry(back) == wing_geometry(ro)


@pytest.mark.parametrize("form", ["axisymmetric", "wedge", "half_cone"])
def test_body_form_round_trips(tmp_path, form):
    p = tmp_path / "ro.xlsx"
    export_ro_xlsx(str(p), _ro(body_form=form))
    assert import_ro_xlsx(str(p)).body_form == form


def test_wedge_body_span_round_trips(tmp_path):
    p = tmp_path / "ro.xlsx"
    export_ro_xlsx(str(p), _ro(body_form="wedge", body_span_m=0.9))
    back = import_ro_xlsx(str(p))
    assert back.body_form == "wedge" and back.body_span_m == 0.9


def test_unknown_body_form_normalises_not_crashes(tmp_path):
    """A hand-edited workbook with a junk body-form string imports as the
    default, mirroring ro_from_dict."""
    import openpyxl
    from ro_xlsx import _R, _VAL_COL
    p = tmp_path / "ro.xlsx"
    export_ro_xlsx(str(p), _ro())
    wb = openpyxl.load_workbook(str(p))
    wb["RO"].cell(row=_R["body_form"], column=_VAL_COL, value="waverider")
    wb.save(str(p))
    assert import_ro_xlsx(str(p)).body_form == "axisymmetric"


def test_pre_upgrade_workbook_imports_with_defaults(tmp_path):
    """Emulate a workbook written BEFORE the appended sections by deleting
    every appended row: the import must fall back to the old behavior —
    axisymmetric, no biconic, no wings — not crash or invent values."""
    import openpyxl
    p = tmp_path / "ro.xlsx"
    export_ro_xlsx(str(p), _ro(biconic=True, fore_length_m=0.5,
                               break_diameter_m=0.42, body_form="wedge",
                               wing_root_chord_m=0.6, wing_span_exposed_m=0.15))
    wb = openpyxl.load_workbook(str(p))
    wb["RO"].delete_rows(52, 20)               # drop all appended rows
    wb.save(str(p))
    back = import_ro_xlsx(str(p))
    assert back.body_form == "axisymmetric"
    assert back.biconic is False and back.fore_length_m == 0.0
    assert wing_geometry(back) == (0.0, 0.0, None)
    # the ORIGINAL fields still read correctly
    assert back.mass_kg == 450.0 and back.diameter_m == 0.58


def test_blank_template_still_builds_and_imports(tmp_path):
    p = tmp_path / "blank.xlsx"
    make_blank_ro_template(str(p))
    back = import_ro_xlsx(str(p))
    assert back.body_form == "axisymmetric" and not back.biconic

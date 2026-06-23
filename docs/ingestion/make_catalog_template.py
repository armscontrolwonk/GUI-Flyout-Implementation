"""Generate the draft Thrusty catalog workbook (rockets + payloads).

Two-sheet normalized format:
  - `rockets`  : one row per STAGE, grouped by model_id (sorted by stage_no).
  - `payloads` : one row per reentry vehicle / glide body, referenced by
                 rockets via payload_id (reusable, mirrors rv_library).
  - `enums`    : allowed values (documentation) — also wired as dropdowns.

Blank numeric cells are intentional: they are exactly what the ingest resolver
estimates and flags (estimate-with-VERIFY). Column choices mirror
missile.schema.json / rv.schema.json.

Run:  python docs/ingestion/make_catalog_template.py
Out:  docs/ingestion/catalog_template_draft.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "docs/ingestion/catalog_template_draft.xlsx"

# ---- enum value lists (mirror the JSON schemas) ----------------------------
ENUMS = {
    "missile_class":   ["srbm", "mrbm", "irbm", "icbm", "slv", "sounding", "other"],
    "propellant_type": ["solid_composite", "solid_double_base", "liquid_storable",
                        "liquid_lox_rp1", "liquid_lox_lh2", "liquid_lox_ch4", "unknown"],
    "guidance":        ["pitch_program", "true_gravity_turn", "orbital_insertion"],
    "grain_type":      ["tubular", "rod_tube", "double_anchor", "star",
                        "multi_fin", "dual_composition"],
    "terminal_mode":   ["reentry", "orbital", "suborbital_payload"],
    "yes_no":          ["YES", "NO"],
    "rv_kind":         ["ballistic", "marv_body", "glider", "decoy"],
    "glider_guidance": ["skip_glide", "damped_glide", "equilibrium_glide_acton",
                        "equilibrium_glide", "skip_to_equilibrium"],
    "glider_aero_model": ["polar", "constant_LD"],
    "shape":           ["cone", "tangent_ogive", "von_karman", "lv_haack",
                        "parabola", "blunt_cylinder"],
}

# ---- sheet column definitions ----------------------------------------------
# (header, dropdown_enum_key_or_None)
ROCKETS_COLS = [
    ("model_id", None),            # groups a model's stages
    ("name", None),
    ("missile_class", "missile_class"),     # model-level: fill on stage_no=1
    ("country", None),                       # model-level
    ("stage_no", None),
    ("propellant_type", "propellant_type"),
    ("mass_initial_kg", None),
    ("mass_propellant_kg", None),
    ("mass_final_kg", None),
    ("diameter_m", None),
    ("length_m", None),
    ("burn_time_s", None),
    ("isp_s", None),
    ("thrust_N", None),
    ("solid_motor", "yes_no"),
    ("grain_type", "grain_type"),
    ("guidance", "guidance"),
    ("burnout_angle_deg", None),
    ("loft_angle_rate_deg_s", None),
    ("coast_time_s", None),
    ("payload_id", None),          # model-level: link into payloads (blank = orbital/none)
    ("terminal_mode", "terminal_mode"),      # model-level
    ("payload_kg", None),          # model-level: orbital payload mass (no RV)
    ("source_citation", None),     # model-level
    ("source_url", None),
    ("license", None),
]

PAYLOADS_COLS = [
    ("payload_id", None),
    ("name", None),
    ("rv_kind", "rv_kind"),
    ("mass_kg", None),             # blank for marv_body (inherited from burnout)
    ("beta_kg_m2", None),
    ("shape", "shape"),
    ("diameter_m", None),
    ("length_m", None),
    ("nose_radius_m", None),
    ("glider_LD", None),           # glider only
    ("glider_guidance", "glider_guidance"),
    ("glider_damping_zeta", None), # damped_glide only (default 0.7)
    ("glider_pullup_g_max", None),
    ("glider_terminal_dive", "yes_no"),
    ("glider_terminal_alt_km", None),
    ("glider_aero_model", "glider_aero_model"),
    ("glider_beta_entry_kg_m2", None),  # equilibrium_glide_acton only
    ("glider_skip_count", None),        # skip_to_equilibrium only
    ("source_citation", None),
    ("source_url", None),
    ("license", None),
]

# ---- worked examples -------------------------------------------------------
# Rockets: one row per stage. Model-level fields filled on the stage_no=1 row.
ROCKETS_ROWS = [
    # Scud-B — single-stage SRBM, ballistic RV (numbers blank where the
    # resolver would estimate are deliberately left in here since they are known)
    ["scud_b", "Scud-B (R-17)", "srbm", "USSR", 1, "liquid_storable",
     5897, 3699, 2198, 0.84, 11.25, 75, 230, 111200, "NO", "", "pitch_program",
     47.66, 1.348, 0, "scud_rv", "reentry", "", "Lewis 2024", "", "CC-BY-4.0"],

    # DF-21 — two-stage solid MRBM, MaRV. Isp/thrust LEFT BLANK on purpose:
    # the resolver fills them from propellant_type + class and flags VERIFY.
    ["df21", "DF-21", "mrbm", "China", 1, "solid_composite",
     14700, 11000, "", 1.4, 10.7, 60, "", "", "YES", "tubular", "pitch_program",
     "", "", 0, "df21_marv", "reentry", "", "example", "", "CC-BY-4.0"],
    ["df21", "DF-21", "", "", 2, "solid_composite",
     2500, 1800, "", 1.4, 2.6, 55, "", "", "YES", "star", "pitch_program",
     "", "", 0, "", "", "", "", "", ""],

    # AUR + HGB — two-stage booster carrying a hypersonic glide body
    ["aur_hgb", "AUR + HGB", "irbm", "example", 1, "solid_composite",
     "", "", "", 1.5, "", "", "", "", "YES", "", "pitch_program",
     "", "", 0, "hgb1", "reentry", "", "example", "", "CC-BY-4.0"],
    ["aur_hgb", "AUR + HGB", "", "", 2, "solid_composite",
     "", "", "", 1.2, "", "", "", "", "YES", "", "pitch_program",
     "", "", 0, "", "", "", "", "", ""],

    # Generic SLV — three solid stages to orbit. terminal_mode=orbital, NO
    # payload_id (a satellite is not an RV); orbital payload mass in payload_kg.
    ["slv_demo", "Generic SLV", "slv", "example", 1, "solid_composite",
     "", "", "", 1.6, "", "", "", "", "YES", "", "pitch_program",
     "", "", 5, "", "orbital", 350, "example", "", "CC-BY-4.0"],
    ["slv_demo", "Generic SLV", "", "", 2, "solid_composite",
     "", "", "", 1.2, "", "", "", "", "YES", "", "pitch_program",
     "", "", 5, "", "", "", "", "", ""],
    ["slv_demo", "Generic SLV", "", "", 3, "solid_composite",
     "", "", "", 0.9, "", "", "", "", "YES", "", "orbital_insertion",
     "", "", 0, "", "", "", "", "", ""],
]

# Payloads: one row per RV / glide body.
PAYLOADS_ROWS = [
    # ballistic RV
    ["scud_rv", "Scud warhead", "ballistic", 989, "", "cone", 0.84, "", 0.05,
     "", "", "", "", "NO", "", "constant_LD", "", "", "Lewis 2024", "", "CC-BY-4.0"],
    # MaRV — mass/diameter INHERITED from the missile burnout body (left blank)
    ["df21_marv", "DF-21 MaRV", "marv_body", "", 8000, "cone", "", "", 0.05,
     "", "", "", "", "NO", "", "constant_LD", "", "", "example", "", "CC-BY-4.0"],
    # HGB glide body — damped_glide
    ["hgb1", "HGB", "glider", 450, 15000, "cone", 0.5, 2.0, 0.05,
     2.6, "damped_glide", 0.7, 10, "NO", 30, "constant_LD", "", "", "example", "", "CC-BY-4.0"],
]

# ---- styling helpers -------------------------------------------------------
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF")
MODEL_FILL = PatternFill("solid", fgColor="FCE4D6")  # tint model-level cols
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MODEL_LEVEL_ROCKET = {"missile_class", "country", "payload_id", "terminal_mode",
                      "payload_kg", "source_citation", "source_url", "license"}


def write_sheet(ws, cols, rows, note, model_level=None):
    model_level = model_level or set()
    ws.append([note])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    ws.cell(1, 1).font = Font(italic=True, color="808080")
    ws.cell(1, 1).alignment = Alignment(wrap_text=False)

    headers = [c[0] for c in cols]
    ws.append(headers)
    for j, h in enumerate(headers, 1):
        c = ws.cell(2, j)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    for r in rows:
        ws.append(list(r))

    # tint model-level columns
    for j, h in enumerate(headers, 1):
        if h in model_level:
            for i in range(3, 3 + len(rows)):
                ws.cell(i, j).fill = MODEL_FILL

    # borders + widths
    for j, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(j)].width = max(11, len(h) + 2)
        for i in range(3, 3 + len(rows)):
            ws.cell(i, j).border = BORDER

    ws.freeze_panes = "A3"

    # dropdowns (rows 3..400)
    for j, (h, ekey) in enumerate(cols, 1):
        if not ekey:
            continue
        vals = ",".join(ENUMS[ekey])
        dv = DataValidation(type="list", formula1=f'"{vals}"', allow_blank=True)
        ws.add_data_validation(dv)
        col = get_column_letter(j)
        dv.add(f"{col}3:{col}400")


def write_enums(ws):
    ws.append(["enums — allowed values (mirrors the JSON schemas)"])
    ws.cell(1, 1).font = Font(bold=True)
    keys = list(ENUMS.keys())
    ws.append(keys)
    for j, k in enumerate(keys, 1):
        c = ws.cell(2, j)
        c.fill = HDR_FILL
        c.font = HDR_FONT
    maxlen = max(len(v) for v in ENUMS.values())
    for i in range(maxlen):
        ws.append([ENUMS[k][i] if i < len(ENUMS[k]) else "" for k in keys])
    for j, k in enumerate(keys, 1):
        ws.column_dimensions[get_column_letter(j)].width = max(13, len(k) + 2)


def write_readme(ws):
    lines = [
        ("Thrusty catalog — DRAFT format", True),
        ("", False),
        ("Two record types live in two tabs:", True),
        ("  • rockets  — one row per STAGE; rows sharing model_id form the stage chain (sorted by stage_no).", False),
        ("  • payloads — one row per reentry vehicle / glide body; reused across rockets via payload_id.", False),
        ("", False),
        ("Conventions", True),
        ("  • BLANK numeric cells are intentional — the ingest resolver estimates them and flags VERIFY.", False),
        ("  • Peach-tinted columns in `rockets` are MODEL-level: fill them once on the stage_no=1 row.", False),
        ("  • payload_id links a rocket to a payloads row. Leave blank for orbital launches (a", False),
        ("    satellite is not an RV) and set terminal_mode=orbital + payload_kg instead.", False),
        ("  • marv_body payloads inherit mass/diameter/length from the missile burnout — leave blank.", False),
        ("  • solid_motor=YES enables grain_type; guidance=orbital_insertion belongs on the final SLV stage.", False),
        ("", False),
        ("Running in Thrusty", True),
        ("  • Live:    File → Share (anyone-with-link viewer), then give Thrusty the sheet URL;", False),
        ("             it fetches  …/export?format=xlsx  (both tabs, no auth).", False),
        ("  • Upload:  File → Download → .xlsx  and open the file in Thrusty (same importer).", False),
    ]
    for text, bold in lines:
        ws.append([text])
        ws.cell(ws.max_row, 1).font = Font(bold=bold)
    ws.column_dimensions["A"].width = 100


def main():
    wb = Workbook()
    ws_r = wb.active
    ws_r.title = "rockets"
    write_sheet(ws_r, ROCKETS_COLS, ROCKETS_ROWS,
                "DRAFT — one row per stage. Blank numbers = resolver estimates & flags. "
                "Peach columns are model-level (fill on the stage_no=1 row).",
                model_level=MODEL_LEVEL_ROCKET)

    ws_p = wb.create_sheet("payloads")
    write_sheet(ws_p, PAYLOADS_COLS, PAYLOADS_ROWS,
                "DRAFT — one row per reentry vehicle / glide body. Referenced by rockets via payload_id.")

    write_enums(wb.create_sheet("enums"))
    write_readme(wb.create_sheet("README"))

    wb.save(OUT)
    print(f"wrote {OUT}: tabs = {wb.sheetnames}")
    print(f"  rockets: {len(ROCKETS_ROWS)} stage-rows; payloads: {len(PAYLOADS_ROWS)} rows")


if __name__ == "__main__":
    main()

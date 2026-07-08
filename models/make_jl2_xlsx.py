"""Build the JL-2 SLBM + RV workbooks from parade-photo photogrammetry.

Produces two files in models/:
  JL-2.xlsx          — detailed single-missile workbook (missile_xlsx format;
                       RV parameters live in its PAYLOAD & REENTRY VEHICLE block)
  JL-2_catalog.xlsx  — catalog long-format (docs/ingestion format: one `rockets`
                       row per stage + a `payloads` row for the RV)

Measurement chain (Jeffrey Lewis, WebPlotDigitizer on 2019-parade JL-3
canister photo — proxy geometry for the JL-2 family):
  Dist3 = 154.6746 px vertical = canister radius 1.1 m  → scale 7.112 mm/px
  Dist1 = 620.8316 px = stage-1 raceway (motor cylinder) = 4.415 m
  Dist0 = 324.1810 px = stage-2 raceway                  = 2.306 m
  Dist2 =  44.4954 px = motor dome height                = 0.316 m
Each motor = cylinder + two half-ellipsoid domes of height Dist2.

Assumption set (P35-anchored; Wei & Bai, "Design and Development of P35
Solid Rocket Motor for Long March 11 Launch Vehicle", GLEX 2017, #36288 —
same steel-case HTPB technology family):
  stage 1 & 2 diam    2.0 m   (canister OD 2.2 m discounted to the missile;
                               P35 is likewise 2.0 m)
  stage 3 diameter    1.8 m   (modest 0.9x taper; necked-down upper stage)
  grain density       1800 kg/m³ (HTPB/AP composite; mass_estimator.py APCP)
  volumetric loading  0.85
  mass ratio          0.88    (prop/wet; P35 flight value → dry/wet = 0.12)
  Isp (vacuum)        272 s   (P35 ground Isp 248 s + P_a·A_e/(mdot·g0))
  burn times          CALCULATED — P35 fixes the family mass flow (35 t /
                      71 s = 493 kg/s at 2.0 m); throat ∝ diameter², so
                      mdot(d) = 493·(d/2)² and burn = propellant/mdot.
                      → S1 47 s, S2 27 s, S3 (1.8 m) 18 s.
  stage-1 nozzle A_e  sized so sea-level Isp comes out at exactly 248 s
Stage 3 propellant is sized relative to stages 1 & 2 (geometric mass taper
P2/P1 ≈ 0.56); its diameter is the one free choice (1.8 m).  The warhead is
out of scope — RV mass/beta are left blank and the RV mass is NOT carried
in the stack.

CAVEAT: per-stage `mass_initial` in the detailed workbook is the motor
stack only (S1+S2+S3 wet) — it excludes the warhead, so add the RV mass
when it is known.

Run from the repo root:  python models/make_jl2_xlsx.py
"""
import math
import os
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)
sys.path.insert(0, os.path.join(ROOT, 'docs', 'ingestion'))

G0 = 9.80665
PA_SL = 101_325.0            # Pa, sea level

# ── WebPlotDigitizer measurements (px) ──────────────────────────────────────
PX_RADIUS  = 154.6745886654478   # Dist3: canister radius
PX_STAGE1  = 620.8316474783319   # Dist1: stage-1 raceway
PX_STAGE2  = 324.18098720292505  # Dist0: stage-2 raceway
PX_DOME    = 44.49542961608768   # Dist2: dome height
PX_SHROUD  = 608.0               # forward shroud (contains stage 3 + warhead)
CANISTER_RADIUS_M = 1.1

# ── Assumptions (see module docstring for provenance) ───────────────────────
DIAMETER_M     = 2.0             # stages 1 & 2 (missile OD inside 2.2 m canister)
DIAMETER3_M    = 1.8            # stage 3 — modest 0.9x taper (necked-down upper stage)
RHO_GRAIN      = 1800.0          # kg/m³
VOL_LOADING    = 0.85
MASS_RATIO     = 0.88            # propellant / wet  (P35)
ISP_VAC_S      = 272.0
ISP_GROUND_S   = 248.0           # P35 ground Isp — used only to size A_e

# Burn times are CALCULATED, not assumed.  P35 fixes the family mass-flow
# rate (35 t / 71 s = 493 kg/s at 2.0 m).  For geometrically similar
# motors the throat area — and thus mass flow — scales with diameter², so
#   mdot(d) = MDOT_REF · (d / 2.0)²   and   burn = propellant / mdot(d).
# This reproduces P35's quoted ~1176 kN thrust and puts stage 1 at ~47 s.
P35_PROP_KG    = 35_000.0
P35_BURN_S     = 71.0
P35_DIAM_M     = 2.0
MDOT_REF       = P35_PROP_KG / P35_BURN_S        # 493 kg/s at 2.0 m

SCALE = CANISTER_RADIUS_M / PX_RADIUS          # m per px
DOME_H = PX_DOME * SCALE

def stage_masses(px_raceway: float, dia: float = DIAMETER_M):
    """(length_m, prop_kg, wet_kg, dry_kg) for one motor from its raceway px."""
    r = dia / 2.0
    l_cyl = px_raceway * SCALE
    vol = math.pi * r * r * l_cyl + 2.0 * (2.0 / 3.0) * math.pi * r * r * DOME_H
    prop = vol * RHO_GRAIN * VOL_LOADING
    wet = prop / MASS_RATIO
    return l_cyl + 2.0 * DOME_H, prop, wet, wet - prop

def length_for_prop(prop: float, dia: float):
    """Total motor length (cylinder + 2 domes) holding `prop` at diameter `dia`."""
    r = dia / 2.0
    vol = prop / (RHO_GRAIN * VOL_LOADING)
    dome2 = 2.0 * (2.0 / 3.0) * math.pi * r * r * DOME_H
    return (vol - dome2) / (math.pi * r * r) + 2.0 * DOME_H

def mdot_for(dia: float) -> float:
    """Mass flow (kg/s): P35 anchor scaled by throat area ∝ diameter²."""
    return MDOT_REF * (dia / P35_DIAM_M) ** 2

def burn_thrust(prop: float, dia: float):
    """(burn_s, vacuum_thrust_N) for a motor of `prop` kg at diameter `dia`."""
    mdot = mdot_for(dia)
    return prop / mdot, ISP_VAC_S * G0 * mdot

L1, PROP1, WET1, DRY1 = stage_masses(PX_STAGE1)
L2, PROP2, WET2, DRY2 = stage_masses(PX_STAGE2)

# Stage 3: sized relative to stages 1 & 2, not measured directly. The
# propellant follows the geometric mass taper P2/P1 (~0.56); at 1.8 m
# that fixes the motor length, and the remaining shroud is the warhead
# bay (ignored per current scope). Diameter is the only free choice.
PROP3 = PROP2 * (PROP2 / PROP1)
WET3  = PROP3 / MASS_RATIO
DRY3  = WET3 - PROP3
L3    = length_for_prop(PROP3, DIAMETER3_M)
SHROUD_M   = PX_SHROUD * SCALE
WARHEAD_BAY_M = SHROUD_M - L3       # length left in shroud for RV + bus (out of scope)

CITATION = ('Photogrammetry of 2019-parade JL-3 canister (proxy geometry); '
            'motor anchors from P35 (Wei & Bai, GLEX 2017 #36288): '
            'steel case HTPB, mass ratio 0.88, ground Isp 248 s')


# Burn times and vacuum thrusts, all from the P35 mass-flow anchor.
BURN1_S, THRUST1_N = burn_thrust(PROP1, DIAMETER_M)
BURN2_S, THRUST2_N = burn_thrust(PROP2, DIAMETER_M)
BURN3_S, THRUST3_N = burn_thrust(PROP3, DIAMETER3_M)

# Stage 1 alone flies through sea level, so its nozzle A_e is sized to make
# the vacuum Isp collapse to the P35 ground Isp (248 s) at lift-off.
AE1_M2 = (ISP_VAC_S - ISP_GROUND_S) * mdot_for(DIAMETER_M) * G0 / PA_SL


def build_detailed(path: str) -> None:
    from missile_models import MissileParams
    from missile_xlsx import export_missile_xlsx, _R

    def mk(name, mass_init, prop, wet_dry, length, thrust, burn, ae, dia):
        p = MissileParams(
            name=name, mass_initial=round(mass_init), mass_propellant=round(prop),
            mass_final=round(wet_dry), diameter_m=dia, length_m=round(length, 2),
            thrust_N=round(thrust), burn_time_s=burn, isp_s=ISP_VAC_S,
        )
        p.solid_motor = True
        p.nozzle_exit_area_m2 = ae
        return p

    # mass_initial = wet mass of this stage + everything above it. Payload
    # (warhead) is out of scope, so the RV mass is NOT included in the stack.
    s1 = mk('JL-2', WET1 + WET2 + WET3, PROP1, DRY1, L1, THRUST1_N, BURN1_S, round(AE1_M2, 2), DIAMETER_M)
    s2 = mk('JL-2 S2', WET2 + WET3, PROP2, DRY2, L2, THRUST2_N, BURN2_S, 0.0, DIAMETER_M)
    s3 = mk('JL-2 S3', WET3, PROP3, DRY3, L3, THRUST3_N, BURN3_S, 0.0, DIAMETER3_M)
    s1.stage2 = s2
    s2.stage2 = s3
    # RV block — mass/beta/geometry to be filled from a future measurement
    s1.num_rvs = 1
    s1.rv_separates = True
    export_missile_xlsx(path, s1)

    # rv_shape only serialises alongside a positive beta, which we don't
    # have yet — write the shape cell directly so the intent is recorded.
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb['Missile']
    ws.cell(row=_R['rv_shape'], column=4, value='Cone')
    # Body nose shape: blunted SLBM ogive, approximated as tangent ogive.
    # With a shape set, the Chin (1961) component drag model governs.
    ws.cell(row=_R['nose_shape'], column=4, value='Tangent Ogive')
    # The exporter still pre-fills the legacy Forden Cd table, which the
    # importer would treat as an explicit override — clear it so the
    # nose-shape (Chin) model is what the workbook actually encodes.
    wc = wb['Cd Table']
    for row in wc.iter_rows(min_row=4, max_col=2):
        for cell in row:
            cell.value = None
    wb.save(path)


def build_catalog(path: str) -> None:
    import make_catalog_template as cat

    cat.ROCKETS_ROWS = [
        ['jl2', 'JL-2 (SLBM)', 'icbm', 'China', 1, 'solid_composite',
         '', round(PROP1), '', DIAMETER_M, round(L1, 2), round(BURN1_S, 1), ISP_VAC_S,
         round(THRUST1_N), 'YES', '', 'true_gravity_turn',
         '', '', 0, 'jl2_rv', 'reentry', '', CITATION, '', ''],
        ['jl2', 'JL-2 (SLBM)', '', '', 2, 'solid_composite',
         '', round(PROP2), '', DIAMETER_M, round(L2, 2), round(BURN2_S, 1), ISP_VAC_S,
         round(THRUST2_N), 'YES', '', 'true_gravity_turn', '', '', 0, '', '', '', '', '', ''],
        # Stage 3 sized relative to stages 1 & 2 (geometric mass taper); the
        # 1.8 m taper and P35 mass-flow anchor fix its length and burn time.
        ['jl2', 'JL-2 (SLBM)', '', '', 3, 'solid_composite',
         '', round(PROP3), '', DIAMETER3_M, round(L3, 2), round(BURN3_S, 1), ISP_VAC_S,
         round(THRUST3_N), 'YES', '', 'true_gravity_turn',
         '', '', 0, '', '', '', '', '', ''],
    ]
    cat.PAYLOADS_ROWS = [
        ['jl2_rv', 'JL-2 RV', 'ballistic', '', '', 'cone', '', '', '',
         '', '', '', '', 'NO', '', '', '', '', CITATION, '', ''],
    ]
    cat.OUT = path
    cat.main()


if __name__ == '__main__':
    print(f'scale {SCALE*1000:.3f} mm/px, dome h {DOME_H:.3f} m')
    rows = (('S1', L1, PROP1, WET1, DRY1, BURN1_S, THRUST1_N, DIAMETER_M),
            ('S2', L2, PROP2, WET2, DRY2, BURN2_S, THRUST2_N, DIAMETER_M),
            ('S3', L3, PROP3, WET3, DRY3, BURN3_S, THRUST3_N, DIAMETER3_M))
    for tag, L, p, w, d, tb, th, dia in rows:
        print(f'  {tag}: dia {dia} m, length {L:.2f} m, prop {p/1000:.2f} t, '
              f'wet {w/1000:.2f} t, dry {d/1000:.2f} t | burn {tb:.1f} s, thrust(vac) {th/1000:.0f} kN')
    print(f'  S1 A_e {AE1_M2:.2f} m² '
          f'(ground Isp check: {ISP_VAC_S - PA_SL*AE1_M2/(mdot_for(DIAMETER_M)*G0):.1f} s)')
    print(f'  shroud {SHROUD_M:.2f} m -> stage 3 {L3:.2f} m + warhead bay {WARHEAD_BAY_M:.2f} m (out of scope)')
    print(f'  motor stack wet: {(WET1+WET2+WET3)/1000:.1f} t')
    out = os.path.dirname(os.path.abspath(__file__))
    build_detailed(os.path.join(out, 'JL-2.xlsx'))
    build_catalog(os.path.join(out, 'JL-2_catalog.xlsx'))
    print('wrote JL-2.xlsx and JL-2_catalog.xlsx')

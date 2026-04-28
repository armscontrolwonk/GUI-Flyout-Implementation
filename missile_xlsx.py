"""
missile_xlsx.py — XLSX import/export for Thrusty missile parameters.

Requires openpyxl (pip install openpyxl).  Import is lazy so the rest of
the app starts without it; the user sees a friendly error only when they
attempt an XLSX operation.

Sheet layout
------------
  Sheet 1 "Missile"  — all parameters, fields-as-rows, stages-as-columns
  Sheet 2 "Cd Table" — Mach/Cd lookup table (Forden defaults pre-filled)
  Sheet 3 "Reference"— read-only reference values (Isp, mass fractions, T/W)

Public API
----------
  export_missile_xlsx(path, params)   -> None
  import_missile_xlsx(path)           -> MissileParams
  make_blank_template(path)           -> None
"""

from __future__ import annotations
from typing import Optional

# ---------------------------------------------------------------------------
# Row-number registry (1-based).  Both writer and reader use these constants
# so the mapping never gets out of sync.
# ---------------------------------------------------------------------------
_R: dict[str, int] = {
    # IDENTITY
    'name':        3,
    # STAGES (per-stage input rows; columns D/E/F/G = stages 1-4)
    'mass_init':   7,
    'mass_prop':   8,
    'mass_final':  9,
    'diam':        10,
    'length':      11,
    'thrust':      12,
    'burn':        13,
    'isp':         14,
    'nozzle':      15,
    'solid':       16,
    'grain':       17,
    'peak_thr':    18,
    # BOOSTERS (single-value; column D only)
    'b_n':         29,
    'b_thr':       30,
    'b_burn':      31,
    'b_inert':     32,
    'b_prop':      33,
    'b_isp':       34,
    'b_nozzle':    35,
    'b_diam':      36,
    'b_len':       37,
    'b_cd':        38,
    'b_delay':     39,
    # PAYLOAD & RV
    'payload':     47,
    'bus_mass':    48,
    'n_rvs':       49,
    'rv_mass':     50,
    'rv_sep':      51,
    'rv_beta':     52,
    'rv_shape':    53,
    'rv_diam':     54,
    'rv_len':      55,
    'pbv_diam':    56,
    'pbv_len':     57,
    # SHROUD / FAIRING
    'shr_mass':    60,
    'shr_alt':     61,
    'shr_len':     62,
    'shr_diam':    63,
    'shr_shape':   64,
    'shr_noselen': 65,
    # AERODYNAMICS
    'nose_shape':  68,
    'nose_len':    69,
    'pay_diam':    70,
}

# Computed-section anchor rows (writer only)
_RC_STAGES  = 20   # section header; data rows 21-25 (prop chk, frac, dV, T/W, derived thrust)
_RC_BOOSTERS= 41   # section header; data rows 42-44

# Stage column indices (openpyxl 1-based) and letters
_SCOLS  = [4, 5, 6, 7]          # D E F G
_SLETRS = ['D', 'E', 'F', 'G']

# ---------------------------------------------------------------------------
# Dropdown option lists (human-readable strings)
# ---------------------------------------------------------------------------
_NOSE_OPTS = [
    '', 'Cone', 'Tangent Ogive', 'Von Karman (LD-Haack)',
    'LV-Haack (Sears-Haack)', 'Parabola', 'Blunt Cylinder',
]
_GRAIN_OPTS = [
    '', 'Tubular (progressive)', 'Rod and tube (neutral)',
    'Double anchor (regressive)', 'Star (neutral)',
    'Multi-fin (two-phase)', 'Dual composition (two-phase)',
]
_GUID_OPTS  = ['Gravity Turn', 'Loft']
_YESNO_OPTS = ['YES', 'NO']

# Internal-key → human label
_NOSE_LABEL = {
    '': '', 'cone': 'Cone', 'tangent_ogive': 'Tangent Ogive',
    'von_karman': 'Von Karman (LD-Haack)',
    'lv_haack': 'LV-Haack (Sears-Haack)',
    'parabola': 'Parabola', 'blunt_cylinder': 'Blunt Cylinder',
}
_GRAIN_LABEL = {
    '': '', 'tubular': 'Tubular (progressive)',
    'rod_tube': 'Rod and tube (neutral)',
    'double_anchor': 'Double anchor (regressive)',
    'star': 'Star (neutral)', 'multi_fin': 'Multi-fin (two-phase)',
    'dual_composition': 'Dual composition (two-phase)',
}
_GUID_LABEL  = {'gravity_turn': 'Gravity Turn', 'loft': 'Loft'}

# Human label → internal key (reverse maps)
_NOSE_KEY  = {v: k for k, v in _NOSE_LABEL.items()}
_GRAIN_KEY = {v: k for k, v in _GRAIN_LABEL.items()}
_GUID_KEY  = {v: k for k, v in _GUID_LABEL.items()}

# Forden (2007) default Cd table
_FORDEN_MACH = [0.0, 0.85, 1.0, 1.2, 2.0, 4.5]
_FORDEN_CD   = [0.2, 0.20, 0.27, 0.27, 0.20, 0.20]


# ---------------------------------------------------------------------------
# Lazy openpyxl import
# ---------------------------------------------------------------------------
def _xl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise ImportError(
            'openpyxl is required for XLSX export/import.\n'
            'Install with:  pip install openpyxl'
        )


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def _fill(hex_color: str):
    from openpyxl.styles import PatternFill
    return PatternFill('solid', fgColor=hex_color)

def _font(bold=False, color='000000', size=10):
    from openpyxl.styles import Font
    return Font(bold=bold, color=color, size=size)

def _align(h='left', v='center', wrap=False):
    from openpyxl.styles import Alignment
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# Colour palette
_CH  = '1F497D'   # dark blue  — section headers
_CCH = '2E75B6'   # mid blue   — column headers
_CIN = 'FFFFC0'   # yellow     — user input cells
_CCO = 'D9D9D9'   # light grey — computed cells
_CLB = 'F2F2F2'   # off-white  — label / unit cells
_CCS = 'BDD7EE'   # light blue — computed-section headers


# ---------------------------------------------------------------------------
# Low-level cell writers
# ---------------------------------------------------------------------------
def _section(ws, row: int, title: str, computed: bool = False) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    c = ws.cell(row=row, column=1, value=title)
    c.fill      = _fill(_CCS if computed else _CH)
    c.font      = _font(bold=True, color='000000' if computed else 'FFFFFF', size=11)
    c.alignment = _align('left')
    ws.row_dimensions[row].height = 16

def _col_headers(ws, row: int, pairs: list) -> None:
    for col, text in pairs:
        c = ws.cell(row=row, column=col, value=text)
        c.fill      = _fill(_CCH)
        c.font      = _font(bold=True, color='FFFFFF')
        c.alignment = _align('center')

def _label(ws, row: int, text: str, unit: str = '', notes: str = '') -> None:
    for col, val in ((2, text), (3, unit or ''), (9, notes or '')):
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = _fill(_CLB)
        c.font      = _font(size=10 if col != 9 else 9,
                            color='000000' if col != 9 else '555555')
        c.alignment = _align('left', wrap=(col == 9))

def _inputs(ws, row: int, cols: list, values: list,
            fmt: str = 'General') -> None:
    for i, col in enumerate(cols):
        val = values[i] if i < len(values) else None
        c = ws.cell(row=row, column=col, value=val)
        c.fill          = _fill(_CIN)
        c.font          = _font()
        c.alignment     = _align('right')
        c.number_format = fmt

def _computed(ws, row: int, col: int, formula: str,
              fmt: str = '0.00') -> None:
    c = ws.cell(row=row, column=col, value=formula)
    c.fill          = _fill(_CCO)
    c.font          = _font(color='333333')
    c.alignment     = _align('right')
    c.number_format = fmt

def _dropdowns(ws, row: int, cols: list, options: list) -> None:
    """Add a single DataValidation object covering all cols at row."""
    from openpyxl.worksheet.datavalidation import DataValidation
    formula = '"' + ','.join(str(o) for o in options) + '"'
    dv = DataValidation(type='list', formula1=formula,
                        showDropDown=False, allow_blank=True)
    ws.add_data_validation(dv)
    for col in cols:
        ref = ws.cell(row=row, column=col).coordinate
        dv.add(ref)
        ws.cell(row=row, column=col).fill = _fill(_CIN)
        ws.cell(row=row, column=col).font = _font()

def _dropdown(ws, row: int, col: int, options: list) -> None:
    _dropdowns(ws, row, [col], options)


# ---------------------------------------------------------------------------
# Conversion helpers
# ---------------------------------------------------------------------------
def _nose_label(key: str) -> str:
    return _NOSE_LABEL.get(key or '', '')

def _grain_label(key: str) -> str:
    return _GRAIN_LABEL.get(key or '', '')

def _guid_label(key: str) -> str:
    return _GUID_LABEL.get(key or 'gravity_turn', 'Gravity Turn')

def _yn(val: bool) -> str:
    return 'YES' if val else 'NO'


# ---------------------------------------------------------------------------
# Stage-chain flattener
# ---------------------------------------------------------------------------
def _stage_dicts(params) -> list:
    """Return list of up to 4 missile_to_dict dicts, one per stage.
    thrust_N is injected directly from the params object because
    missile_to_dict derives (and discards) it from Isp."""
    from missile_models import missile_to_dict
    out, node = [], params
    while node is not None and len(out) < 4:
        d = missile_to_dict(node)
        d['thrust_N'] = node.thrust_N
        out.append(d)
        node = node.stage2
    return out


# ---------------------------------------------------------------------------
# Sheet builders  (stubs — filled section by section)
# ---------------------------------------------------------------------------
def _build_missile_sheet(ws, stages: list, top: dict) -> None:
    r = _R

    # Column widths
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    for ch in 'DEFG': ws.column_dimensions[ch].width = 16
    ws.column_dimensions['H'].width = 4
    ws.column_dimensions['I'].width = 34
    ws.freeze_panes = 'D8'

    # Title
    ws.merge_cells('A1:I1')
    c = ws.cell(row=1, column=1, value='Thrusty — Missile Parameter Template')
    c.fill = _fill('1A1A2E'); c.font = _font(bold=True, color='FFFFFF', size=14)
    c.alignment = _align('center'); ws.row_dimensions[1].height = 24

    # ── IDENTITY ─────────────────────────────────────────────────────────────
    _section(ws, r['name'] - 1, 'IDENTITY')
    _label(ws, r['name'], 'Missile Name', '', 'Short identifying name')
    _inputs(ws, r['name'], [4], [top.get('name', '')])
    ws.merge_cells(start_row=r['name'], start_column=4,
                   end_row=r['name'],   end_column=7)

    # ── STAGES ───────────────────────────────────────────────────────────────
    _section(ws, r['mass_init'] - 2, 'STAGES')
    _col_headers(ws, r['mass_init'] - 1, [
        (2, 'Parameter'), (3, 'Units'),
        (4, 'Stage 1'), (5, 'Stage 2'), (6, 'Stage 3'), (7, 'Stage 4'),
        (9, 'Notes'),
    ])

    def sv(key, cast=float):
        """Stage values: list of 4 (one per stage, None if stage absent)."""
        out = []
        for s in stages:
            v = s.get(key)
            out.append(cast(v) if v not in (None, '') else None)
        while len(out) < 4:
            out.append(None)
        return out

    def srow(rk, label, unit, key, cast=float, notes=''):
        _label(ws, r[rk], label, unit, notes)
        _inputs(ws, r[rk], _SCOLS, sv(key, cast))

    srow('mass_init',  'Initial (launch) mass',     'kg', 'mass_initial',
         notes='Structure + propellant + payload')
    srow('mass_prop',  'Propellant mass',            'kg', 'mass_propellant')
    srow('mass_final', 'Final (burnout) mass',       'kg', 'mass_final',
         notes='= Initial − Propellant')
    srow('diam',       'Diameter',                   'm',  'diameter_m')
    srow('length',     'Length',                     'm',  'length_m')
    srow('thrust',     'Vacuum thrust',              'N',  'thrust_N',
         notes='Reference only — model always derives thrust from Isp (see Derived thrust below)')
    srow('burn',       'Burn time',                  's',  'burn_time_s')
    srow('isp',        'Isp (specific impulse)',     's',  'isp_s',
         notes='Solid: 230–290  Storable liq: 280–310  Cryo: 420–450')
    srow('nozzle',     'Nozzle exit area',           'm²', 'nozzle_exit_area_m2',
         notes='0 = legacy 2% back-pressure approx')
    srow('peak_thr',   'Peak thrust (solid only)',   'N',  'thrust_peak_N',
         notes='0 = derive from thrust × fill factor')

    # Solid motor YES/NO
    _label(ws, r['solid'], 'Solid motor', '—', 'YES = cannot throttle or cut off')
    solid_vals = [_yn(s.get('solid_motor', False)) for s in stages]
    while len(solid_vals) < 4: solid_vals.append('NO')
    _inputs(ws, r['solid'], _SCOLS, solid_vals)
    _dropdowns(ws, r['solid'], _SCOLS, _YESNO_OPTS)

    # Grain type dropdown
    _label(ws, r['grain'], 'Grain type', '—',
           'Solid motors only; blank = constant thrust (liquid)')
    grain_vals = [_grain_label(s.get('grain_type', '')) for s in stages]
    while len(grain_vals) < 4: grain_vals.append('')
    _inputs(ws, r['grain'], _SCOLS, grain_vals)
    _dropdowns(ws, r['grain'], _SCOLS, _GRAIN_OPTS)

    # ── COMPUTED — STAGES ────────────────────────────────────────────────────
    _section(ws, _RC_STAGES, 'COMPUTED — STAGES  (do not edit)',
             computed=True)
    comp_labels = [
        (_RC_STAGES+1, 'Propellant mass check', 'kg',
         '= Initial − Final  (should equal Propellant mass above)'),
        (_RC_STAGES+2, 'Mass fraction',         '—',  'Propellant / Initial'),
        (_RC_STAGES+3, 'ΔV  (Tsiolkovsky)',     'm/s','Isp × g₀ × ln(m₀/m_f)'),
        (_RC_STAGES+4, 'T/W at ignition',       '—',  'Thrust / (Initial × g₀)'),
        (_RC_STAGES+5, 'Derived thrust',        'N',
         'Isp × 9.80665 × Propellant / Burn time  (cross-check for Vacuum thrust above)'),
    ]
    for row, lbl, unit, note in comp_labels:
        _label(ws, row, lbl, unit, note)

    for i, col in enumerate(_SLETRS):
        mi  = f'{col}{r["mass_init"]}'
        mf  = f'{col}{r["mass_final"]}'
        mp  = f'{col}{r["mass_prop"]}'
        isp = f'{col}{r["isp"]}'
        th  = f'{col}{r["thrust"]}'
        bn  = f'{col}{r["burn"]}'
        ci  = _SCOLS[i]
        _computed(ws, _RC_STAGES+1, ci,
                  f'=IF({mi}>0,{mi}-{mf},"—")', '#,##0')
        _computed(ws, _RC_STAGES+2, ci,
                  f'=IF({mi}>0,{mp}/{mi},"—")', '0.000')
        _computed(ws, _RC_STAGES+3, ci,
                  f'=IF(AND({mi}>0,{mf}>0),{isp}*9.80665*LN({mi}/{mf}),"—")',
                  '#,##0')
        _computed(ws, _RC_STAGES+4, ci,
                  f'=IF({mi}>0,{th}/({mi}*9.80665),"—")', '0.00')
        _computed(ws, _RC_STAGES+5, ci,
                  f'=IF(AND({isp}>0,{bn}>0),{isp}*9.80665*{mp}/{bn},"—")',
                  '#,##0')

    # ── BOOSTERS ─────────────────────────────────────────────────────────────
    _section(ws, r['b_n'] - 2, 'BOOSTERS  (strap-on, parallel to Stage 1)')
    _col_headers(ws, r['b_n'] - 1,
                 [(2, 'Parameter'), (3, 'Units'), (4, 'Value'), (9, 'Notes')])

    def brow(rk, label, unit, key, cast=float, notes=''):
        _label(ws, r[rk], label, unit, notes)
        v = top.get(key)
        _inputs(ws, r[rk], [4], [cast(v) if v not in (None, '') else None])

    brow('b_n',     'Number of boosters',          '—',  'n_boosters',
         cast=int,  notes='0 = no strap-ons')
    brow('b_thr',   'Vacuum thrust per booster',   'N',  'booster_thrust_n',
         notes='Per booster')
    brow('b_burn',  'Burn time',                   's',  'booster_burn_time_s')
    brow('b_inert', 'Inert mass per booster',      'kg', 'booster_inert_kg')
    brow('b_prop',  'Propellant mass per booster', 'kg', 'booster_prop_kg')
    brow('b_isp',   'Isp',                         's',  'booster_isp_s',
         notes='Solid: 230–290 s')
    brow('b_nozzle','Nozzle exit area',             'm²', 'booster_nozzle_area_m2',
         notes='0 = legacy approx')
    brow('b_diam',  'Diameter',                    'm',  'booster_diam_m')
    brow('b_len',   'Length  (0 = 2 × diameter)',  'm',  'booster_length_m')
    brow('b_cd',    'Zero-lift Cd',                '—',  'booster_cd',
         notes='Default 0.20 (tangent ogive)')
    brow('b_delay', 'Core ignition delay',         's',  'booster_core_delay_s',
         notes='0 = simultaneous (Soyuz); >0 = sequential (LVM3, Titan IIIC)')

    # ── COMPUTED — BOOSTERS ──────────────────────────────────────────────────
    _section(ws, _RC_BOOSTERS, 'COMPUTED — BOOSTERS  (do not edit)',
             computed=True)
    bn_  = f'D{r["b_n"]}';    bpr = f'D{r["b_prop"]}'
    bin_ = f'D{r["b_inert"]}'; bth = f'D{r["b_thr"]}'
    mi1  = f'D{r["mass_init"]}'; th1 = f'D{r["thrust"]}'

    _label(ws, _RC_BOOSTERS+1, 'Total booster propellant', 'kg')
    _computed(ws, _RC_BOOSTERS+1, 4,
              f'=IF({bn_}>0,{bn_}*{bpr},"—")', '#,##0')
    _label(ws, _RC_BOOSTERS+2, 'Total booster inert mass', 'kg')
    _computed(ws, _RC_BOOSTERS+2, 4,
              f'=IF({bn_}>0,{bn_}*{bin_},"—")', '#,##0')
    _label(ws, _RC_BOOSTERS+3,
           'Liftoff T/W  (all boosters + Stage 1)', '—',
           'Combined booster + Stage 1 thrust vs full launch mass')
    _computed(ws, _RC_BOOSTERS+3, 4,
              f'=IF({mi1}>0,({bn_}*{bth}+{th1})/({mi1}*9.80665),"—")', '0.00')

    # ── PAYLOAD & RV ─────────────────────────────────────────────────────────
    _section(ws, r['payload'] - 1, 'PAYLOAD & REENTRY VEHICLE')

    def prow(rk, label, unit, key, cast=float, notes=''):
        _label(ws, r[rk], label, unit, notes)
        v = top.get(key)
        _inputs(ws, r[rk], [4], [cast(v) if v not in (None, '') else None])

    prow('payload',  'Payload (total)',         'kg',    'payload_kg',
         notes='Bus + all RVs')
    prow('bus_mass', 'Bus (PBV) mass',          'kg',    'bus_mass_kg')
    prow('n_rvs',    'Number of RVs',           '—',     'num_rvs',     cast=int)
    prow('rv_mass',  'RV mass (each)',           'kg',    'rv_mass_kg')
    _label(ws, r['rv_sep'], 'RV separates at burnout', '—',
           'YES → empty stage follows a separate debris arc')
    _inputs(ws, r['rv_sep'], [4], [_yn(top.get('rv_separates', False))])
    _dropdown(ws, r['rv_sep'], 4, _YESNO_OPTS)
    prow('rv_beta',  'RV β (ballistic coeff.)', 'kg/m²', 'rv_beta_kg_m2',
         notes='Higher β → less drag deceleration')
    _label(ws, r['rv_shape'], 'RV nose shape', '—')
    _inputs(ws, r['rv_shape'], [4], [_nose_label(top.get('rv_shape', ''))])
    _dropdown(ws, r['rv_shape'], 4, _NOSE_OPTS)
    prow('rv_diam',  'RV diameter',             'm',     'rv_diameter_m')
    prow('rv_len',   'RV length',               'm',     'rv_length_m')
    prow('pbv_diam', 'PBV diameter',            'm',     'pbv_diameter_m')
    prow('pbv_len',  'PBV length',              'm',     'pbv_length_m')

    # ── SHROUD / FAIRING ─────────────────────────────────────────────────────
    _section(ws, r['shr_mass'] - 1, 'SHROUD / FAIRING')

    def shrow(rk, label, unit, key, cast=float, notes=''):
        _label(ws, r[rk], label, unit, notes)
        v = top.get(key)
        _inputs(ws, r[rk], [4], [cast(v) if v not in (None, '') else None])

    shrow('shr_mass',    'Shroud mass',       'kg', 'shroud_mass_kg',
          notes='0 = no shroud')
    shrow('shr_alt',     'Jettison altitude', 'km', 'shroud_jettison_alt_km',
          notes='Default 80 km')
    shrow('shr_len',     'Shroud length',     'm',  'shroud_length_m')
    shrow('shr_diam',    'Shroud diameter',   'm',  'shroud_diameter_m')
    _label(ws, r['shr_shape'], 'Shroud nose shape', '—')
    _inputs(ws, r['shr_shape'], [4],
            [_nose_label(top.get('shroud_nose_shape', ''))])
    _dropdown(ws, r['shr_shape'], 4, _NOSE_OPTS)
    shrow('shr_noselen', 'Shroud nose length', 'm', 'shroud_nose_length_m')

    # ── AERODYNAMICS ─────────────────────────────────────────────────────────
    _section(ws, r['nose_shape'] - 1, 'AERODYNAMICS')
    _label(ws, r['nose_shape'], 'Nose shape', '—')
    _inputs(ws, r['nose_shape'], [4], [_nose_label(top.get('nose_shape', ''))])
    _dropdown(ws, r['nose_shape'], 4, _NOSE_OPTS)
    _label(ws, r['nose_len'], 'Nose length', 'm',
           'Physical nose-cone length')
    _inputs(ws, r['nose_len'], [4], [top.get('nose_length_m', 0.0)])
    _label(ws, r['pay_diam'], 'Payload diameter', 'm',
           '0 = use body diameter')
    _inputs(ws, r['pay_diam'], [4], [top.get('payload_diameter_m', 0.0)])


def _build_cd_sheet(ws, mach: list, cd: list) -> None:
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 40

    ws.merge_cells('A1:C1')
    c = ws.cell(row=1, column=1, value='Cd vs Mach  —  Drag Coefficient Table')
    c.fill = _fill('1A1A2E'); c.font = _font(bold=True, color='FFFFFF', size=13)
    c.alignment = _align('center'); ws.row_dimensions[1].height = 22

    ws.merge_cells('A2:C2')
    note = ws.cell(row=2, column=1,
                   value='Forden (2007) defaults pre-filled.  '
                         'Edit or extend as needed.  '
                         'Leave blank to use the built-in nose-shape model instead.')
    note.fill = _fill(_CLB); note.font = _font(size=9, color='444444')
    note.alignment = _align('left', wrap=True); ws.row_dimensions[2].height = 28

    for col, hdr in ((1, 'Mach'), (2, 'Cd'), (3, 'Notes')):
        c = ws.cell(row=3, column=col, value=hdr)
        c.fill = _fill(_CCH); c.font = _font(bold=True, color='FFFFFF')
        c.alignment = _align('center')

    use_mach = mach if mach else _FORDEN_MACH
    use_cd   = cd   if cd   else _FORDEN_CD
    for i, (m, d) in enumerate(zip(use_mach, use_cd)):
        r = 4 + i
        mc = ws.cell(row=r, column=1, value=m)
        dc = ws.cell(row=r, column=2, value=d)
        for cell in (mc, dc):
            cell.fill = _fill(_CIN); cell.font = _font()
            cell.alignment = _align('right'); cell.number_format = '0.000'

    # Leave extra blank yellow rows for extension
    for i in range(len(use_mach), len(use_mach) + 6):
        for col in (1, 2):
            c = ws.cell(row=4 + i, column=col)
            c.fill = _fill(_CIN); c.alignment = _align('right')


def _build_reference_sheet(ws) -> None:
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 32

    ws.merge_cells('A1:E1')
    c = ws.cell(row=1, column=1, value='Reference Values  (read only)')
    c.fill = _fill('1A1A2E'); c.font = _font(bold=True, color='FFFFFF', size=13)
    c.alignment = _align('center'); ws.row_dimensions[1].height = 22

    _section(ws, 2, 'Propellant Types — Typical Isp and Mass Fraction')
    for col, hdr in ((1,'Propellant type'),(2,'Isp (s)'),(3,'Mass fraction'),
                     (4,'T/W at ignition'),(5,'Notes')):
        c = ws.cell(row=3, column=col, value=hdr)
        c.fill = _fill(_CCH); c.font = _font(bold=True, color='FFFFFF')
        c.alignment = _align('center')

    rows = [
        ('Solid composite (HTPB/AP)',      '230 – 290', '0.85 – 0.92',
         '1.5 – 3.0', 'Most modern SRBMs/MRBMs'),
        ('Solid double-base',              '200 – 240', '0.80 – 0.88',
         '1.5 – 2.5', 'Older/smaller motors'),
        ('Storable liquid (N₂O₄/UDMH)',   '280 – 310', '0.85 – 0.92',
         '1.2 – 2.0', 'Hypergolic; Scud, Nodong, DF-series'),
        ('Storable liquid (N₂O₄/MMH)',    '285 – 320', '0.85 – 0.92',
         '1.2 – 2.0', 'Hypergolic; many IRBMs'),
        ('Cryogenic (LOX / RP-1)',         '340 – 360', '0.88 – 0.94',
         '1.1 – 1.5', 'Requires fuelling at launch'),
        ('Cryogenic (LOX / LH₂)',          '420 – 460', '0.86 – 0.92',
         '1.0 – 1.4', 'Highest Isp; space launch vehicles'),
        ('Cryogenic (LOX / CH₄)',          '340 – 380', '0.87 – 0.93',
         '1.1 – 1.5', 'Methane; newer systems'),
    ]
    for i, row_data in enumerate(rows):
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=4 + i, column=col, value=val)
            c.fill = _fill(_CLB if i % 2 == 0 else 'FFFFFF')
            c.font = _font(size=10); c.alignment = _align('left')

    _section(ws, 12, 'Guidance Mode Quick Reference')
    guid_rows = [
        ('Gravity Turn', 'IRBM / ICBM',
         'Kick to near-horizontal early in flight; thrust locked to velocity vector.'),
        ('Loft',         'SRBM / MRBM',
         'Pitch to a fixed loft angle at a constant rate, then hold.'),
    ]
    for col, hdr in ((1,'Mode'),(2,'Typical use'),(4,'Description')):
        c = ws.cell(row=13, column=col, value=hdr)
        c.fill = _fill(_CCH); c.font = _font(bold=True, color='FFFFFF')
        c.alignment = _align('center')
    for i, (mode, use, desc) in enumerate(guid_rows):
        for col, val in ((1,mode),(2,use),(4,desc)):
            c = ws.cell(row=14+i, column=col, value=val)
            c.fill = _fill(_CLB if i % 2 == 0 else 'FFFFFF')
            c.font = _font(); c.alignment = _align('left')


# ---------------------------------------------------------------------------
# Read helpers (for import)
# ---------------------------------------------------------------------------
def _rnum(ws, row: int, col: int, default: float = 0.0) -> float:
    v = ws.cell(row=row, column=col).value
    if v is None or v == '' or v == '—':
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default

def _rint(ws, row: int, col: int, default: int = 0) -> int:
    return int(_rnum(ws, row, col, default))

def _rstr(ws, row: int, col: int, default: str = '') -> str:
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else default

def _rbool(ws, row: int, col: int, default: bool = False) -> bool:
    return _rstr(ws, row, col, 'NO').upper() == 'YES'


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
def export_missile_xlsx(path: str, params) -> None:
    """Write params to a fully filled-in XLSX template at path."""
    xl = _xl()
    stages = _stage_dicts(params)
    top    = stages[0] if stages else {}
    mach   = list(top.get('mach_table', []))
    cd     = list(top.get('cd_table',   []))

    wb = xl.Workbook()
    ws_m = wb.active
    ws_m.title = 'Missile'
    ws_c = wb.create_sheet('Cd Table')
    ws_r = wb.create_sheet('Reference')

    _build_missile_sheet(ws_m, stages, top)
    _build_cd_sheet(ws_c, mach, cd)
    _build_reference_sheet(ws_r)
    wb.save(path)


def import_missile_xlsx(path: str):
    """Read an XLSX template and return a MissileParams chain."""
    from missile_models import missile_from_dict
    xl = _xl()
    wb = xl.load_workbook(path, data_only=True)
    ws = wb['Missile']
    r  = _R

    name = _rstr(ws, r['name'], 4, 'Unnamed')

    # Read Cd table if present
    mach_t, cd_t = [], []
    if 'Cd Table' in wb.sheetnames:
        wc = wb['Cd Table']
        for row in range(4, 4 + 30):
            m = wc.cell(row=row, column=1).value
            d = wc.cell(row=row, column=2).value
            if m is None or d is None:
                break
            try:
                mach_t.append(float(m)); cd_t.append(float(d))
            except (TypeError, ValueError):
                break

    # Build per-stage dicts from columns D-G
    stage_dicts = []
    for i, col in enumerate(_SCOLS):
        mi = _rnum(ws, r['mass_init'],  col)
        if mi <= 0 and i > 0:
            break   # stage absent
        mp = _rnum(ws, r['mass_prop'],  col)
        mf = _rnum(ws, r['mass_final'], col)
        if mf <= 0 and mi > 0 and mp > 0:
            mf = mi - mp
        d = {
            'name':               name if i == 0 else f'{name} S{i+1}',
            'mass_initial':       mi,
            'mass_propellant':    mp,
            'mass_final':         mf,
            'diameter_m':         _rnum(ws, r['diam'],     col),
            'length_m':           _rnum(ws, r['length'],   col),
            'thrust_N':           _rnum(ws, r['thrust'],   col),
            'burn_time_s':        _rnum(ws, r['burn'],     col),
            'isp_s':              _rnum(ws, r['isp'],      col),
            'nozzle_exit_area_m2':_rnum(ws, r['nozzle'],  col),
            'solid_motor':        _rbool(ws, r['solid'],   col),
            'grain_type':   _GRAIN_KEY.get(_rstr(ws, r['grain'],   col), ''),
            'thrust_peak_N':      _rnum(ws, r['peak_thr'],col),
            'mach_table':         mach_t,
            'cd_table':           cd_t,
        }
        stage_dicts.append(d)

    # Chain stages (reverse order so we can set stage2 pointers)
    chain = None
    for sd in reversed(stage_dicts):
        sd['stage2'] = missile_from_dict(chain.__dict__ if chain else
                                         {k: None for k in []}) if False else None
        chain = missile_from_dict(sd)
        if len(stage_dicts) > 1:
            # Re-attach already-built later stage
            pass

    # Rebuild chain properly
    stages_built = [missile_from_dict(sd) for sd in stage_dicts]
    for i in range(len(stages_built) - 1):
        stages_built[i].stage2 = stages_built[i + 1]
    top_stage = stages_built[0] if stages_built else missile_from_dict(
        {'name': name, 'mass_initial': 0, 'mass_propellant': 0,
         'mass_final': 0, 'diameter_m': 0, 'length_m': 0,
         'thrust_N': 0, 'burn_time_s': 1, 'isp_s': 250})

    # Top-level fields (stored on stage 1 node only)
    top_stage.payload_kg          = _rnum(ws, r['payload'],  4)
    top_stage.bus_mass_kg         = _rnum(ws, r['bus_mass'], 4)
    top_stage.num_rvs             = _rint(ws, r['n_rvs'],    4, 1)
    top_stage.rv_mass_kg          = _rnum(ws, r['rv_mass'],  4)
    top_stage.rv_separates        = _rbool(ws, r['rv_sep'],  4)
    top_stage.rv_beta_kg_m2       = _rnum(ws, r['rv_beta'],  4)
    top_stage.rv_shape    = _NOSE_KEY.get(_rstr(ws, r['rv_shape'],  4), '')
    top_stage.rv_diameter_m       = _rnum(ws, r['rv_diam'],  4)
    top_stage.rv_length_m         = _rnum(ws, r['rv_len'],   4)
    top_stage.pbv_diameter_m      = _rnum(ws, r['pbv_diam'], 4)
    top_stage.pbv_length_m        = _rnum(ws, r['pbv_len'],  4)
    top_stage.shroud_mass_kg      = _rnum(ws, r['shr_mass'], 4)
    top_stage.shroud_jettison_alt_km = _rnum(ws, r['shr_alt'], 4, 80.0)
    top_stage.shroud_length_m     = _rnum(ws, r['shr_len'],  4)
    top_stage.shroud_diameter_m   = _rnum(ws, r['shr_diam'], 4)
    top_stage.shroud_nose_shape   = _NOSE_KEY.get(
                                      _rstr(ws, r['shr_shape'], 4), '')
    top_stage.shroud_nose_length_m= _rnum(ws, r['shr_noselen'], 4)
    top_stage.nose_shape  = _NOSE_KEY.get(_rstr(ws, r['nose_shape'], 4), '')
    top_stage.nose_length_m       = _rnum(ws, r['nose_len'],  4)
    top_stage.payload_diameter_m  = _rnum(ws, r['pay_diam'],  4)
    top_stage.n_boosters          = _rint(ws, r['b_n'],     4)
    top_stage.booster_thrust_n    = _rnum(ws, r['b_thr'],   4)
    top_stage.booster_burn_time_s = _rnum(ws, r['b_burn'],  4)
    top_stage.booster_inert_kg    = _rnum(ws, r['b_inert'], 4)
    top_stage.booster_prop_kg     = _rnum(ws, r['b_prop'],  4)
    top_stage.booster_isp_s       = _rnum(ws, r['b_isp'],   4)
    top_stage.booster_nozzle_area_m2 = _rnum(ws, r['b_nozzle'], 4)
    top_stage.booster_diam_m      = _rnum(ws, r['b_diam'],  4)
    top_stage.booster_length_m    = _rnum(ws, r['b_len'],   4)
    top_stage.booster_cd          = _rnum(ws, r['b_cd'],    4, 0.20)
    top_stage.booster_core_delay_s= _rnum(ws, r['b_delay'], 4)
    return top_stage


def make_blank_template(path: str) -> None:
    """Write a blank (unfilled) template the user fills in from scratch."""
    from missile_models import MissileParams
    blank = MissileParams(
        name='', mass_initial=0, mass_propellant=0, mass_final=0,
        diameter_m=0, length_m=0, thrust_N=0, burn_time_s=1, isp_s=250,
    )
    export_missile_xlsx(path, blank)
